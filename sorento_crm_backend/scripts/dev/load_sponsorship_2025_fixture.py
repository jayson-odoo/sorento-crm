#!/usr/bin/env python3
"""Load the client's JAN-DEC'25 sponsorship workbook into a LOCAL database, as a fixture.

PLAN-reporting-foundation.md slice S5 (AC-F1 to AC-F5). The CRM's own sponsorship history
starts in 2026; the 214 forms the client kept by hand in 2025 exist only in
`SPONSORSHIP REPORT JAN-Dec'25.xlsx`. Loading them locally is what lets the report be read
against a full year, and it is what AC-D5's export diff compares against.

**This is a developer fixture and it must never reach production.** Three things keep it
there: the loader refuses any `DATABASE_URL` that is not localhost (AC-F1), every row it
writes is stamped `source='fixture_2025'` so a stray one is identifiable and deletable in a
single statement (AC-F2), and nothing in `deploy.sh`, docker, alembic or CI references this
file. It is run by hand or not at all.

Run from `sorento_crm_backend/`:

    python scripts/dev/load_sponsorship_2025_fixture.py            # dry run (default)
    python scripts/dev/load_sponsorship_2025_fixture.py --apply    # write

    # Undo, if ever needed:
    DELETE FROM purchase_requests WHERE source = 'fixture_2025';   -- lines cascade

What the workbook holds and how it is read
------------------------------------------

12 monthly sheets (`JAN'25` .. `Dec'25`; the tab names vary in case and spelling - `Augt'25`,
`Sept'25` - so the month is taken from the first three letters plus the `'YY` suffix) with a
header on row 6, year sub-headers on row 7 and rows from row 8 down to a `GRAND TOTAL` row.
The SUMMARY pivot is not read: it is the shape the report itself produces.

**The month comes from the TAB NAME, not from the title cell A4.** A4 usually carries the
month as a datetime, but the client's `Dec'25` sheet carries 1 Nov 2025 there. Trusting it
would file 24 December forms under November: invisible in a year total, wrong in every cell
of the summary. The mismatch is reported rather than silently corrected.

`PS NO` reads `PSSF25- 001` in the workbook, with a space after the dash and three digits.
The CRM numbers every form `PSSF{yy}-####` (4 digits, shared with PR / SI / CMP), so the
loader **normalises to `PSSF25-0001`** - otherwise the fixture rows would sort and search
differently from every real form and the report's PS No column would show two formats.

`PROJECT VALUE` is a number, a `-`, or free text. A dash means "no value" and clears BOTH
columns; free text goes to `total_project_value_text` and contributes to no total, which is
the rule AC-B3 already fixed for the live rows. `SPONSHER/SPONSER PROJECT` is mapped onto
the `procurement_sponsor_subject` lookup using migration 243's own keyword lists, with the
raw text preserved in `sponsor_subject_other` when it lands on `others`.

`SAMPLE PRICE` becomes ONE `purchase_request_lines` row (`SAMPLE`, quantity 1), because that
is where the report reads the sample price from; a blank or zero price writes no line, so
the column stays empty rather than reading `0.00` (AC-B2).

`EXPECTED YEAR OF DELIVERY` (H..K = 2025..2028) is ticked on no row of the 2025 workbook, so
every fixture form has a NULL `expected_delivery_date`. The reader is written anyway, since
the client's 2026 sheets do use it.

Agent names
-----------

`SALES AGENT` is a first name in the workbook and a display name in `respond_contacts`. The
match is the EXACT, case-insensitive one `scripts/backfill_requested_by_contact.py` already
uses (name, or first + last), reused rather than re-stated, plus a small alias map for the
five short spellings the workbook uses for a contact whose record carries a surname. An
unresolved name is never guessed: the row keeps the typed name in `requested_by`, and the
name is listed with its row count at the end (AC-F4). `ACT`, `KH LIM` and `JAMYN` have no
contact at all (the workbook's own SUMMARY sheet spells the last one `JAYMYN`; neither
spelling exists), so those rows are attributed to nobody until someone creates the contact.

Idempotence
-----------

Keyed on `request_number` (AC-F3). A re-run updates the header in place and REPLACES the
lines, so editing a cell in the workbook and re-running converges; it never duplicates. A
form the fixture does not own - one with the same number and a different `source` - is left
untouched and reported, so a hand-raised PSSF25 form can never be overwritten by a fixture.
"""
from __future__ import annotations

import argparse
import os
import re
import sys
from collections import Counter
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.parse import urlsplit

import openpyxl

# Allow `from app.*` / `import scripts.*` when invoked as a file from the backend directory.
_BACKEND_ROOT = Path(__file__).resolve().parents[2]
if str(_BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(_BACKEND_ROOT))

from app.models.access import RespondContact  # noqa: E402
from app.models.procurement import PurchaseRequestHeader, PurchaseRequestLine  # noqa: E402

# The exact-match contact resolver, reused rather than restated: two copies of a matching
# rule are two rules the day one of them is corrected.
from scripts.backfill_requested_by_contact import (  # noqa: E402
    _build_contact_index,
    _resolve,
)

#: Stamped on every row this script writes. The kill switch, the audit trail and the
#: idempotence key's owner, all in one column.
SOURCE = "fixture_2025"

DEFAULT_WORKBOOK = _BACKEND_ROOT / "tests" / "fixtures" / "sponsorship_2025.xlsx"

#: Row 6 is the header, row 7 the year sub-headers, so the table starts here.
FIRST_DATA_ROW = 8

#: Column indexes on a monthly sheet (1-based, as openpyxl counts).
_COL_PS_NO = 1
_COL_AGENT = 2
_COL_CUSTOMER = 3
_COL_PROJECT_TITLE = 4
_COL_SUBJECT = 5
_COL_PROJECT_VALUE = 6
_COL_SAMPLE_PRICE = 7
_COL_YEAR_FIRST = 8
_COL_YEAR_LAST = 11
_YEAR_HEADER_ROW = 7
_TITLE_MONTH_CELL = "A4"

_MONTHS = (
    "JAN", "FEB", "MAR", "APR", "MAY", "JUN",
    "JUL", "AUG", "SEP", "OCT", "NOV", "DEC",
)
_MONTH_ABBR = (
    "Jan", "Feb", "Mar", "Apr", "May", "Jun",
    "Jul", "Aug", "Sep", "Oct", "Nov", "Dec",
)

#: Migration 243's own keyword lists, so a fixture row lands on the same lookup option the
#: CRM would have chosen for the same typed text.
_SHOWROOM = re.compile(r"showroom|show\s+room", re.IGNORECASE)
_MOCKUP = re.compile(r"mockup|mock[\s-]+up|sample|prototype", re.IGNORECASE)

#: Text that means "nothing was written here".
_BLANKS = {"", "-", "--", "n/a", "na"}

#: Workbook spelling (upper-cased) -> the display name the contact record actually carries.
#: Checked against `respond_contacts` on the local copy, 2026-08-26. Anything not here is
#: resolved on its own text or reported unmatched; nothing is fuzzy-matched.
AGENT_ALIASES: Dict[str, str] = {
    "CINDY": "Cindy Lee",
    "LEENA": "Leena Marzuki",
    "BRENDON": "Brendon Foo",
    "SHAHRUL": "Shahrul Sorento",
    "CK": "CK Lee",
}

#: Hosts that are this machine. Anything else is somebody's real database.
LOCAL_HOSTS = frozenset({"localhost", "127.0.0.1", "::1"})


class FixtureRefused(RuntimeError):
    """Raised instead of writing 214 fixture rows somewhere they do not belong."""


# --------------------------------------------------------------------------- the guard


def assert_local_database(url: Optional[str] = None) -> str:
    """Return the host when it is this machine, else raise FixtureRefused saying why.

    AC-F1. The check is on the HOST, not on the database name: a production database
    reached through a tunnel would still answer to a familiar name, but never to
    `localhost`, and this is the last thing standing between a developer fixture and the
    client's own register.

    `main()` passes the URL the ENGINE will use (`settings.database_url`). The
    `DATABASE_URL` fallback is for a caller that has one in hand and no app settings.
    """
    raw = url if url is not None else os.getenv("DATABASE_URL")
    if not raw:
        raise FixtureRefused(
            "Refusing to load the 2025 sponsorship fixture: DATABASE_URL is not set, so "
            "there is no way to tell which database this would write to. Point it at a "
            "localhost copy and try again."
        )

    host = (urlsplit(raw).hostname or "").strip().lower()
    if host in LOCAL_HOSTS:
        return host

    raise FixtureRefused(
        f"Refusing to load the 2025 sponsorship fixture: DATABASE_URL points at host "
        f"{host or '(none)'!r}, which is not localhost. This fixture writes 214 invented "
        f"sponsorship forms and belongs on a local copy only."
    )


# --------------------------------------------------------------------------- the parser
# Workbook -> plain dicts. No database, no ORM, so S4's export diff can reuse it and this
# half stays testable on the committed file alone.


def _clean(value: Any) -> Optional[str]:
    """Trimmed text, or None for a blank cell and for the workbook's placeholder dashes."""
    if value is None:
        return None
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return None if text.lower() in _BLANKS else text


def _decimal(value: Any) -> Optional[Decimal]:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    return None


def normalise_request_number(raw: Any) -> Optional[str]:
    """`PSSF25- 001` -> `PSSF25-0001`, the CRM's own `PSSF{yy}-####` rule.

    The workbook writes three digits and a stray space; every real form in the table is four
    digits with none. Left alone, the fixture rows would sort apart from the real ones and
    read as a second numbering scheme on the report's PS No column.
    """
    if raw is None:
        return None
    text = re.sub(r"\s+", "", str(raw)).upper()
    match = re.fullmatch(r"PSSF(\d{2})-(\d+)", text)
    if not match:
        return None
    return f"PSSF{match.group(1)}-{int(match.group(2)):04d}"


def map_sponsor_subject(raw: Any) -> tuple[Optional[str], Optional[str]]:
    """Free text -> (`showroom` | `mockup` | `others`, free text parked on `others`).

    Migration 243's keyword lists, so this agrees with what the CRM did to the live rows.
    The raw text is dropped when it adds nothing to the label it already chose ("OTHERS"
    under Others would render as "Others: OTHERS" on the report).
    """
    text = _clean(raw)
    if text is None:
        return "others", None
    if _SHOWROOM.search(text):
        return "showroom", None
    if _MOCKUP.search(text):
        return "mockup", None
    if text.strip().lower().rstrip("s") == "other":
        return "others", None
    return "others", text


def _month_from_sheet_name(title: str) -> Optional[date]:
    match = re.match(r"\s*([A-Za-z]{3})[A-Za-z]*\s*'?\s*(\d{2})", title)
    if not match:
        return None
    stem = match.group(1).upper()
    if stem not in _MONTHS:
        return None
    return date(2000 + int(match.group(2)), _MONTHS.index(stem) + 1, 1)


def _grand_total_row(worksheet) -> Optional[int]:
    for row in range(FIRST_DATA_ROW, worksheet.max_row + 1):
        for column in (_COL_PROJECT_TITLE, _COL_PS_NO, _COL_SUBJECT):
            value = worksheet.cell(row, column).value
            if isinstance(value, str) and "GRAND TOTAL" in value.upper():
                return row
    return None


def _delivery_year(worksheet, row: int) -> Optional[int]:
    """The earliest ticked year in H..K, read against the year sub-headers on row 7."""
    for column in range(_COL_YEAR_FIRST, _COL_YEAR_LAST + 1):
        if _clean(worksheet.cell(row, column).value) is None:
            continue
        header = worksheet.cell(_YEAR_HEADER_ROW, column).value
        if isinstance(header, (int, float)):
            return int(header)
        parsed = _clean(header)
        if parsed and parsed.isdigit():
            return int(parsed)
    return None


def _parse_sheet(worksheet) -> Optional[Dict[str, Any]]:
    month = _month_from_sheet_name(worksheet.title)
    if month is None:
        return None
    grand_total_row = _grand_total_row(worksheet)
    if grand_total_row is None:
        return None

    month_cell = worksheet[_TITLE_MONTH_CELL].value
    if isinstance(month_cell, datetime):
        month_cell = month_cell.date()
    if isinstance(month_cell, date):
        month_cell = date(month_cell.year, month_cell.month, 1)
    else:
        month_cell = None

    rows: List[Dict[str, Any]] = []
    for excel_row in range(FIRST_DATA_ROW, grand_total_row):
        request_number = normalise_request_number(worksheet.cell(excel_row, _COL_PS_NO).value)
        if request_number is None:
            continue  # a blank spacer line inside the table

        raw_value = worksheet.cell(excel_row, _COL_PROJECT_VALUE).value
        project_value = _decimal(raw_value)
        project_value_text = None if project_value is not None else _clean(raw_value)
        subject, subject_other = map_sponsor_subject(worksheet.cell(excel_row, _COL_SUBJECT).value)

        rows.append(
            {
                "sheet": worksheet.title,
                "excel_row": excel_row,
                "month": month,
                "request_number": request_number,
                "sales_agent": _clean(worksheet.cell(excel_row, _COL_AGENT).value),
                "customer_name": _clean(worksheet.cell(excel_row, _COL_CUSTOMER).value),
                "project_title": _clean(worksheet.cell(excel_row, _COL_PROJECT_TITLE).value),
                "sponsor_subject": subject,
                "sponsor_subject_other": subject_other,
                "project_value": project_value,
                "project_value_text": project_value_text,
                "sample_price": _decimal(worksheet.cell(excel_row, _COL_SAMPLE_PRICE).value),
                "expected_delivery_year": _delivery_year(worksheet, excel_row),
            }
        )

    return {
        "sheet": worksheet.title,
        "month": month,
        "month_cell": month_cell,
        "month_cell_disagrees": month_cell is not None and month_cell != month,
        "grand_total_row": grand_total_row,
        "sheet_total_project_value": _decimal(
            worksheet.cell(grand_total_row, _COL_PROJECT_VALUE).value
        ),
        "sheet_total_sample_price": _decimal(
            worksheet.cell(grand_total_row, _COL_SAMPLE_PRICE).value
        ),
        "rows": rows,
    }


def parse_workbook(path) -> List[Dict[str, Any]]:
    """Every monthly sheet as a plain dict, in calendar order. No database involved.

    `data_only=True` because the GRAND TOTAL cells are `=SUM(...)` formulas and what is
    wanted is the number Excel last computed - the client's own total, to check the parsed
    rows against.
    """
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    sheets = [
        parsed
        for parsed in (_parse_sheet(worksheet) for worksheet in workbook.worksheets)
        if parsed is not None
    ]
    workbook.close()
    return sorted(sheets, key=lambda sheet: sheet["month"])


def iter_rows(sheets: List[Dict[str, Any]]):
    for sheet in sheets:
        yield from sheet["rows"]


# --------------------------------------------------------------------------- the loader
# Dicts -> database. Everything above this line runs without one.


def _resolve_agent(name: Optional[str], by_name, by_fullname) -> Optional[str]:
    """The typed name, then its alias, resolved by the shared exact-match rule.

    Ambiguity resolves to nothing, exactly as the backfill treats it: two contacts called
    Cindy is a question for a human, not a coin toss.
    """
    if not name:
        return None
    contact_id, status = _resolve(name, by_name, by_fullname)
    if status == "matched":
        return contact_id
    alias = AGENT_ALIASES.get(name.strip().upper())
    if alias:
        contact_id, status = _resolve(alias, by_name, by_fullname)
        if status == "matched":
            return contact_id
    return None


def load(db, sheets: List[Dict[str, Any]], *, dry_run: bool = False) -> Dict[str, Any]:
    """Upsert every parsed row as an approved sponsorship form. Returns the counters.

    Idempotent on `request_number` (AC-F3): the header is updated in place and the lines
    are replaced, so the second run over an unchanged workbook leaves the database exactly
    as the first did.
    """
    by_name, by_fullname = _build_contact_index(db)

    result: Dict[str, Any] = {
        "sheets": len(sheets),
        "rows_seen": 0,
        "inserted": 0,
        "updated": 0,
        "lines_written": 0,
        "matched": 0,
        "unmatched": Counter(),
        "skipped_foreign": [],
        "contacts_indexed": len(by_name) + len(by_fullname),
    }

    for row in iter_rows(sheets):
        result["rows_seen"] += 1
        month: date = row["month"]
        stamp = datetime(month.year, month.month, 1)

        contact_id = _resolve_agent(row["sales_agent"], by_name, by_fullname)
        if contact_id:
            result["matched"] += 1
        elif row["sales_agent"]:
            result["unmatched"][row["sales_agent"]] += 1

        existing = (
            db.query(PurchaseRequestHeader)
            .filter(PurchaseRequestHeader.request_number == row["request_number"])
            .one_or_none()
        )
        if existing is not None and existing.source != SOURCE:
            # Not ours. A real form that happens to carry this number is left alone.
            result["skipped_foreign"].append(row["request_number"])
            continue

        if existing is None:
            result["inserted"] += 1
        else:
            result["updated"] += 1
        if row["sample_price"]:
            result["lines_written"] += 1

        # A dry run must not touch the ORM objects at all: mutating one dirties the
        # session, and the next autoflush would issue the UPDATE the run promised not to.
        if dry_run:
            continue

        # `Any`, because the models declare classic `Column(...)` rather than `Mapped[...]`:
        # to a type checker every attribute on a concretely-typed instance reads as
        # `Column[str]`, and every assignment below is an error. The rest of the app dodges
        # this by holding query results untyped; this says so out loud instead.
        header: Any = existing or PurchaseRequestHeader(request_number=row["request_number"])
        header.request_type = "sponsorship_form"
        header.source = SOURCE
        header.status = "approved"
        header.approval_status = "approved"
        header.request_date = month
        header.submitted_at = stamp
        header.approved_at = stamp
        header.customer_name = row["customer_name"]
        header.project_title = row["project_title"]
        header.sponsor_subject = row["sponsor_subject"]
        header.sponsor_subject_other = row["sponsor_subject_other"]
        header.total_project_value = row["project_value"]
        header.total_project_value_text = row["project_value_text"]
        header.expected_delivery_date = (
            date(row["expected_delivery_year"], 1, 1)
            if row["expected_delivery_year"]
            else None
        )
        header.requested_by = row["sales_agent"]
        header.requested_by_contact_id = contact_id

        if existing is None:
            db.add(header)

        db.flush()
        # Replace, never append: a second run must not grow the line list.
        db.query(PurchaseRequestLine).filter(
            PurchaseRequestLine.purchase_request_id == header.id
        ).delete(synchronize_session=False)
        if row["sample_price"]:
            db.add(
                PurchaseRequestLine(
                    purchase_request_id=header.id,
                    item_code="SAMPLE",
                    quantity=Decimal("1"),
                    unit_price=row["sample_price"],
                    total=row["sample_price"],
                    remark=SOURCE,
                    sort_order=0,
                )
            )
            db.flush()

    return result


# -------------------------------------------------------------------------- the summary


def _money(value: Optional[Decimal]) -> str:
    return "-" if value is None else f"{value:,.2f}"


def _month_label(month: date) -> str:
    return f"{_MONTH_ABBR[month.month - 1]}'{month.year % 100:02d}"


def format_summary(sheets: List[Dict[str, Any]], result: Dict[str, Any]) -> str:
    """What was read, what was written, and the per-month totals beside the sheet's own.

    The last part is the point: the client checks the report against these 12 GRAND TOTAL
    cells, so the loader prints both numbers side by side rather than asserting they agree.
    """
    lines: List[str] = []
    lines.append(f"Sheets read: {result['sheets']}")
    lines.append(
        f"Rows: {result['rows_seen']} seen, {result['inserted']} inserted, "
        f"{result['updated']} updated, {len(result['skipped_foreign'])} skipped (not ours)"
    )
    lines.append(f"Sample-price lines written: {result['lines_written']}")
    lines.append(
        f"Agents: {result['matched']} row(s) matched to a contact, "
        f"{sum(result['unmatched'].values())} row(s) left unattributed"
    )

    if result["skipped_foreign"]:
        lines.append("")
        lines.append("Skipped - a form with this number exists and is not the fixture's:")
        for number in result["skipped_foreign"]:
            lines.append(f"  {number}")

    lines.append("")
    lines.append("Unmatched sales agents (row keeps the typed name, no contact link):")
    if result["unmatched"]:
        for name, count in result["unmatched"].most_common():
            lines.append(f"  {name:<20} {count:>4} row(s)")
    else:
        lines.append("  (none)")

    lines.append("")
    lines.append(
        f"{'Month':<8} {'Rows':>5}  {'Project value':>18} {'sheet GRAND TOTAL':>18}  "
        f"{'Sample price':>14} {'sheet GRAND TOTAL':>18}"
    )
    for sheet in sheets:
        parsed_value = sum(
            (r["project_value"] for r in sheet["rows"] if r["project_value"] is not None),
            Decimal("0"),
        )
        parsed_sample = sum(
            (r["sample_price"] for r in sheet["rows"] if r["sample_price"] is not None),
            Decimal("0"),
        )
        agrees = _agrees(parsed_value, sheet["sheet_total_project_value"]) and _agrees(
            parsed_sample, sheet["sheet_total_sample_price"]
        )
        lines.append(
            f"{_month_label(sheet['month']):<8} {len(sheet['rows']):>5}  "
            f"{_money(parsed_value):>18} {_money(sheet['sheet_total_project_value']):>18}  "
            f"{_money(parsed_sample):>14} {_money(sheet['sheet_total_sample_price']):>18}"
            f"  {'match' if agrees else 'DIFFERS'}"
        )

    total_value = sum(
        (r["project_value"] for r in iter_rows(sheets) if r["project_value"] is not None),
        Decimal("0"),
    )
    total_sample = sum(
        (r["sample_price"] for r in iter_rows(sheets) if r["sample_price"] is not None),
        Decimal("0"),
    )
    lines.append(f"{'YEAR':<8} {result['rows_seen']:>5}  {_money(total_value):>18} "
                 f"{'':>18}  {_money(total_sample):>14}")

    disagreeing = [s["sheet"] for s in sheets if s["month_cell_disagrees"]]
    if disagreeing:
        lines.append("")
        lines.append(
            "Sheet(s) whose A4 title date disagrees with the tab name (tab name used): "
            + ", ".join(disagreeing)
        )

    return "\n".join(lines)


def _agrees(parsed: Decimal, sheet_total: Optional[Decimal]) -> bool:
    if sheet_total is None:
        return False
    return abs(parsed - sheet_total) < Decimal("0.01")


# ------------------------------------------------------------------------------- CLI


def main() -> int:
    parser = argparse.ArgumentParser(description="Load the 2025 sponsorship workbook (LOCAL ONLY).")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK), help="Path to the .xlsx")
    parser.add_argument("--dry-run", action="store_true", default=True, help="(default) report only")
    parser.add_argument("--apply", action="store_true", help="perform DB writes")
    args = parser.parse_args()
    if args.apply:
        args.dry_run = False

    # The URL the ENGINE will use, not the environment variable: `Settings` reads `.env`
    # through pydantic-settings, so an unset shell variable says nothing about where the
    # next connection goes, and a guard that checked the shell would wave that through.
    from app.config import settings

    try:
        host = assert_local_database(settings.database_url)
    except FixtureRefused as refusal:
        print(refusal)
        return 2

    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}")
        return 2

    sheets = parse_workbook(workbook_path)
    print(f"Database host: {host}")
    print(f"Workbook: {workbook_path}")

    from app.database import SessionLocal

    db = SessionLocal()
    try:
        result = load(db, sheets, dry_run=args.dry_run)
        if args.dry_run:
            db.rollback()
        else:
            db.commit()
        print()
        print(format_summary(sheets, result))
        print()
        print("[dry-run] no writes performed." if args.dry_run else "Writes committed.")
        return 0
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    raise SystemExit(main())
