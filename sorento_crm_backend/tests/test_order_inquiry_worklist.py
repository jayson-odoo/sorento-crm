"""Purchasing's cross-project order inquiry: the list, the summary and the workbook.

The per-project list next door answers "what did THIS project raise". It cannot answer
purchasing's own question, and not because it is inconvenient: an order ADOPTED from the
AutoCount book has `project_id` NULL by design, so its rows belong to no project and the
per-project query can never return them. Before this list they were reachable only from
the one sales order that raised them.

So the cases worth pinning here are the ones that would let that regress:

* an adopted row and an authored row appear in the SAME list;
* the PROJECT/CUSTOMER column falls back to the CORE order's customer when there is no
  project, rather than going blank (an inner join to Project is what made it blank);
* every filter narrows, including the two that are the screen's own controls;
* the sort set is CLOSED, nulls sort LAST in both directions, and the order is total;
* the workbook is one sheet per delivery MONTH with the client's own heading row.
"""
from __future__ import annotations

import io
import uuid
from datetime import date, datetime, timedelta
from decimal import Decimal

import openpyxl
import pytest
from sqlalchemy import text

from app.models.company import Company
from app.models.inventory import Warehouse
from app.models.order import Customer, SalesOrder, SalesOrderLine
from app.models.procurement import (
    InboundShipment,
    PurchaseOrder,
    PurchaseOrderLine,
    SPOAllocation,
    Supplier,
)
from app.models.product import Product, ProductCategory, UnitOfMeasure
from app.models.project_so import (
    INQUIRY_ACTIONED,
    INQUIRY_PLACED,
    INQUIRY_RAISED,
    IV_ALREADY_INBOUND,
    IV_ORDER,
    SO_STATUS_DRAFT,
    OrderInquiry,
    OrderInquiryRow,
    ProjectSalesOrder,
    ProjectSalesOrderLine,
)
from app.models.sales_agent import SalesAgent
from app.models.user import User
from app.services import project_seed_service

from ._pg_fixture import blank_session

MARKER = "zzt-oi-worklist"
BASE = "/api/v1/project-sales"
LIST = f"{BASE}/order-inquiries"

READ_ONLY = ["projects.projects.view"]
NO_GRANTS: list[str] = []


def _uid() -> str:
    return str(uuid.uuid4())


def _sorento(db) -> str:
    return db.execute(text("select id from companies where code = 'SRT'")).scalar()


def _user(db, name: str) -> str:
    user_id = _uid()
    db.add(User(id=user_id, email=f"{user_id}@zzt.test", name=name))
    db.flush()
    return user_id


def _product(db, code: str, name: str) -> Product:
    uom = UnitOfMeasure(id=_uid(), uom_code=f"ZZT{_uid()[:6]}", uom_name="Unit")
    category = ProductCategory(
        id=_uid(), category_code=f"ZZT-{_uid()[:8]}", category_name=f"{MARKER} cat"
    )
    db.add_all([uom, category])
    db.flush()
    row = Product(
        id=_uid(),
        product_code=code,
        product_name=name,
        category_id=category.id,
        base_uom_id=uom.id,
        list_price=Decimal("100.00"),
    )
    db.add(row)
    db.flush()
    return row


def _customer(db, company_id: str, name: str) -> Customer:
    row = Customer(
        id=_uid(),
        company_id=company_id,
        customer_code=f"ZZT-{_uid()[:8]}",
        customer_name=name,
    )
    db.add(row)
    db.flush()
    return row


def _inquiry_for(db, company_id: str, order: ProjectSalesOrder) -> OrderInquiry:
    inquiry = OrderInquiry(
        id=_uid(),
        company_id=company_id,
        project_sales_order_id=order.id,
        state=INQUIRY_RAISED,
    )
    db.add(inquiry)
    db.flush()
    return inquiry


def _row(db, company_id: str, inquiry: OrderInquiry, **fields) -> OrderInquiryRow:
    row = OrderInquiryRow(
        id=_uid(),
        company_id=company_id,
        order_inquiry_id=inquiry.id,
        verb=fields.pop("verb", IV_ORDER),
        state=fields.pop("state", INQUIRY_RAISED),
        qty=Decimal(str(fields.pop("qty", "10"))),
        **fields,
    )
    db.add(row)
    db.flush()
    return row


def _client(db, user_id: str, permissions):
    from fastapi.testclient import TestClient

    from app.database import get_db
    from app.dependencies import get_current_user, get_current_user_or_api_key
    from app.main import app
    from app.services.company_scope_resolver import apply_company_scope
    from app.services.user_service import UserPermissionService

    actor = {"id": user_id, "email": f"{user_id}@zzt.test", "role": "user"}
    app.dependency_overrides[get_db] = lambda: db
    app.dependency_overrides[get_current_user] = lambda: dict(actor)
    app.dependency_overrides[get_current_user_or_api_key] = lambda: dict(actor)
    app.dependency_overrides[apply_company_scope] = lambda: None

    originals = (
        UserPermissionService.check_user_has_permission,
        UserPermissionService.get_user_permission_slugs,
    )
    granted = list(permissions)
    UserPermissionService.check_user_has_permission = (
        lambda self, uid, slug: slug in granted
    )
    UserPermissionService.get_user_permission_slugs = lambda self, uid: list(granted)
    return TestClient(app), originals


def _restore(originals) -> None:
    from app.main import app
    from app.services.user_service import UserPermissionService

    UserPermissionService.check_user_has_permission = originals[0]
    UserPermissionService.get_user_permission_slugs = originals[1]
    app.dependency_overrides.clear()


def _seed(db, company_id: str, user_id: str) -> dict:
    """One authored project order and one ADOPTED AutoCount order, both with rows.

    Deliberately the two shapes together: everything this list exists for is that they
    are one list, and a fixture with only the authored shape would pass against the
    per-project query that cannot see the other one.
    """
    from app.services.project_service import register_project

    project = register_project(
        db,
        company_id=company_id,
        actor_user_id=user_id,
        developer_party_id=None,
        title=f"{MARKER} Tuju Residence",
    )

    # --- authored: a real project registration, delivering in March.
    authored = ProjectSalesOrder(
        id=_uid(),
        company_id=company_id,
        project_id=project.id,
        area_group="TOWER",
        provisional_ref=f"ZZT-PSO-{_uid()[:8]}",
        autocount_doc_no="SO363150",
        status=SO_STATUS_DRAFT,
        grouping_origin="area",
        published_at=datetime(2025, 8, 15, 9, 0),
    )
    db.add(authored)
    db.flush()
    authored_product = _product(db, f"ZZT-SRTWT107-{_uid()[:6]}", f"{MARKER} Wall hung WC")
    authored_line = ProjectSalesOrderLine(
        id=_uid(),
        company_id=company_id,
        project_sales_order_id=authored.id,
        line_no=1,
        product_id=authored_product.id,
        description=f"{MARKER} authored",
        qty=Decimal("91"),
        uom="UNIT",
        unit_price=Decimal("10.00"),
        amount=Decimal("910.00"),
        delivery_date=date(2026, 3, 2),
    )
    db.add(authored_line)
    db.flush()
    authored_inquiry = _inquiry_for(db, company_id, authored)
    authored_row = _row(
        db,
        company_id,
        authored_inquiry,
        so_line_id=authored_line.id,
        item_code=authored_product.product_code,
        qty="91",
        delivery_date=date(2026, 3, 2),
    )

    # --- adopted: the AutoCount book, no project registration at all.
    customer = _customer(db, company_id, f"{MARKER} Optad Sdn Bhd")
    core = SalesOrder(
        id=_uid(),
        company_id=company_id,
        so_number=f"ZZTSO{_uid()[:8]}",
        customer_id=customer.id,
        order_date=date(2026, 1, 8),
    )
    db.add(core)
    db.flush()
    adopted = ProjectSalesOrder(
        id=_uid(),
        company_id=company_id,
        project_id=None,
        so_id=core.id,
        provisional_ref=core.so_number,
        autocount_doc_no=core.so_number,
        status="adopted",
    )
    db.add(adopted)
    db.flush()
    adopted_product = _product(db, f"ZZT-SRTWC8605-{_uid()[:6]}", f"{MARKER} Close coupled WC")
    adopted_line = ProjectSalesOrderLine(
        id=_uid(),
        company_id=company_id,
        project_sales_order_id=adopted.id,
        line_no=1,
        product_id=adopted_product.id,
        description=f"{MARKER} adopted",
        qty=Decimal("85"),
        uom="UNIT",
        unit_price=Decimal("10.00"),
        amount=Decimal("850.00"),
        delivery_date=date(2026, 1, 19),
    )
    db.add(adopted_line)
    db.flush()
    adopted_inquiry = _inquiry_for(db, company_id, adopted)
    adopted_row = _row(
        db,
        company_id,
        adopted_inquiry,
        so_line_id=adopted_line.id,
        item_code=adopted_product.product_code,
        qty="85",
        delivery_date=date(2026, 1, 19),
        state=INQUIRY_ACTIONED,
    )
    # An undated row, so "nulls last in BOTH directions" has something to be about.
    undated_row = _row(
        db,
        company_id,
        adopted_inquiry,
        so_line_id=adopted_line.id,
        item_code=f"ZZT-UNDATED-{_uid()[:6]}",
        qty="6",
        delivery_date=None,
    )
    db.commit()
    return {
        "project": project,
        "authored": authored,
        "authored_row": authored_row,
        "adopted": adopted,
        "adopted_row": adopted_row,
        "undated_row": undated_row,
        "core": core,
        "customer": customer,
    }


def _placed_supply(
    db,
    company_id: str,
    row: OrderInquiryRow,
    *,
    supplier: Supplier | None = None,
    supplier_name: str = f"{MARKER} DAFUYUAN",
) -> dict:
    """A purchase order the row can actually be traced to, through its SPO reference.

    This is the ONLY link the schema holds today between an inquiry row and a placed
    order, so it is the only thing the SUPPLIER and PO NO columns are allowed to print.
    Everything else on the sheet is filled in by hand and stays blank here.
    """
    if supplier is None:
        supplier = Supplier(
            id=_uid(),
            company_id=company_id,
            supplier_code=f"ZZT-{_uid()[:8]}",
            supplier_name=supplier_name,
        )
        db.add(supplier)
    warehouse = Warehouse(
        id=_uid(),
        company_id=company_id,
        warehouse_code=f"ZZT{_uid()[:6]}",
        warehouse_name=f"{MARKER} WH",
    )
    db.add(warehouse)
    db.flush()
    order = PurchaseOrder(
        id=_uid(),
        company_id=company_id,
        po_number=f"ZZT-202601-{_uid()[:6]}",
        supplier_id=supplier.id,
    )
    db.add(order)
    db.flush()
    line = PurchaseOrderLine(
        id=_uid(),
        company_id=company_id,
        purchase_order_id=order.id,
        product_id=_product(db, f"ZZT-P-{_uid()[:6]}", f"{MARKER} po line").id,
        qty_ordered=Decimal("85"),
    )
    shipment = InboundShipment(
        id=_uid(),
        company_id=company_id,
        shipment_number=f"ZZT-{_uid()[:8]}",
        shipment_date=date(2026, 1, 2),
    )
    db.add_all([line, shipment])
    db.flush()
    spo_number = f"ZZT-SPO-{_uid()[:6]}"
    db.add(
        SPOAllocation(
            id=_uid(),
            company_id=company_id,
            spo_number=spo_number,
            inbound_shipment_id=shipment.id,
            warehouse_id=warehouse.id,
            allocated_quantity=85,
            product_id=line.product_id,
            po_line_id=line.id,
        )
    )
    row.spo_ref = spo_number
    row.verb = IV_ALREADY_INBOUND
    db.flush()
    db.commit()
    return {"supplier": supplier, "purchase_order": order}


def _purchase_order(db, company_id: str, **overrides) -> dict:
    """A standalone purchase order with one line - the "PO no" popup's own fixture,
    unlinked from any order inquiry row (unlike `_placed_supply`, which exists to make a
    ROW traceable through its SPO reference)."""
    supplier = Supplier(
        id=_uid(),
        company_id=company_id,
        supplier_code=f"ZZT-{_uid()[:8]}",
        supplier_name=f"{MARKER} PO SUPPLIER",
    )
    warehouse = Warehouse(
        id=_uid(),
        company_id=company_id,
        warehouse_code=f"ZZT{_uid()[:6]}",
        warehouse_name=f"{MARKER} PO WH",
    )
    db.add_all([supplier, warehouse])
    db.flush()
    order = PurchaseOrder(
        id=_uid(),
        company_id=company_id,
        po_number=f"ZZT-PO-{_uid()[:8]}",
        supplier_id=supplier.id,
        expected_date=overrides.get("expected_date", date(2026, 5, 1)),
        status=overrides.get("status", "active"),
    )
    db.add(order)
    db.flush()
    product = _product(db, f"ZZT-POLINE-{_uid()[:6]}", f"{MARKER} po line item")
    line = PurchaseOrderLine(
        id=_uid(),
        company_id=company_id,
        purchase_order_id=order.id,
        product_id=product.id,
        warehouse_id=warehouse.id,
        qty_ordered=Decimal("40"),
        qty_received=Decimal("15"),
    )
    db.add(line)
    db.commit()
    return {
        "order": order,
        "supplier": supplier,
        "warehouse": warehouse,
        "product": product,
        "line": line,
    }


def _api(permissions):
    from app.models.base import company_scope

    with blank_session() as db:
        company_id = _sorento(db)
        project_seed_service.run(db, company_id=company_id)
        user_id = _user(db, f"{MARKER} Eling")
        seeded = _seed(db, company_id, user_id)
        client, originals = _client(db, user_id, permissions)
        try:
            with company_scope(db, frozenset({company_id})):
                yield client, db, company_id, seeded
        finally:
            _restore(originals)


@pytest.fixture()
def api():
    yield from _api(READ_ONLY)


@pytest.fixture()
def stranger_api():
    yield from _api(NO_GRANTS)


# ------------------------------------------------------------------ the list


def test_the_list_carries_project_rows_and_adopted_rows_together(api):
    client, _db, _company_id, seeded = api

    response = client.get(LIST)

    assert response.status_code == 200, response.text
    body = response.json()
    ids = {row["id"] for row in body["data"]}
    assert seeded["authored_row"].id in ids
    assert seeded["adopted_row"].id in ids
    assert body["pagination"]["total"] == 3


def test_an_adopted_row_names_the_core_orders_customer_rather_than_nothing(api):
    client, _db, _company_id, seeded = api

    body = client.get(LIST).json()
    adopted = next(row for row in body["data"] if row["id"] == seeded["adopted_row"].id)

    # The inner join to Project is what used to blank this column for every adopted row.
    assert seeded["customer"].customer_name in (adopted["project_customer"] or "")
    assert adopted["so_number"] == seeded["core"].so_number
    assert adopted["so_date"] == "2026-01-08"
    assert adopted["core_sales_order_id"] == seeded["core"].id
    assert adopted["is_adopted"] is True


def test_a_project_row_names_the_project_and_reaches_its_own_document(api):
    client, _db, _company_id, seeded = api

    body = client.get(LIST).json()
    authored = next(row for row in body["data"] if row["id"] == seeded["authored_row"].id)

    assert f"{MARKER} Tuju Residence" in (authored["project_customer"] or "")
    assert authored["project_id"] == seeded["project"].id
    assert authored["project_sales_order_id"] == seeded["authored"].id
    assert authored["core_sales_order_id"] is None
    assert authored["product_name"].startswith(MARKER)


def test_the_per_project_list_agrees_with_the_worklist_on_the_customer_label(api):
    """One label, computed once. Two screens disagreeing is a support call."""
    client, _db, _company_id, seeded = api

    worklist = client.get(LIST).json()["data"]
    per_project = client.get(
        f"{BASE}/projects/{seeded['project'].id}/order-inquiry-rows"
    ).json()["data"]

    mine = next(row for row in worklist if row["id"] == seeded["authored_row"].id)
    theirs = next(row for row in per_project if row["id"] == seeded["authored_row"].id)
    assert mine["project_customer"] == theirs["project_customer"]


# ----------------------------------------------------------- the inquiry number


def _inquiry_no_of(db, row: OrderInquiryRow) -> str:
    return (
        db.query(OrderInquiry.inquiry_no)
        .filter(OrderInquiry.id == row.order_inquiry_id)
        .scalar()
    )


def test_every_row_names_the_inquiry_it_belongs_to(api):
    """`OI-000123`, off the row's own header - the number a person quotes.

    The S/O no beside it cannot answer "which instruction was I given": an amendment
    raises a SECOND inquiry on the same sales order, and both rows print the same
    sales-order number.
    """
    client, db, _company_id, seeded = api

    body = client.get(LIST).json()
    by_id = {row["id"]: row for row in body["data"]}

    # It survives `response_model`, which silently drops anything undeclared.
    assert "inquiry_no" in by_id[seeded["authored_row"].id], by_id[
        seeded["authored_row"].id
    ].keys()
    assert by_id[seeded["authored_row"].id]["inquiry_no"] == _inquiry_no_of(
        db, seeded["authored_row"]
    )
    assert by_id[seeded["adopted_row"].id]["inquiry_no"] == _inquiry_no_of(
        db, seeded["adopted_row"]
    )
    # The stamp is real, not a test fixture's invention.
    assert by_id[seeded["authored_row"].id]["inquiry_no"].startswith("OI-")


def test_the_query_box_matches_the_inquiry_number(api):
    """Purchasing is asked about "OI-000123" by name, so the one search box has to find
    it - the row is otherwise reachable only by knowing which sales order raised it."""
    client, db, _company_id, seeded = api

    number = _inquiry_no_of(db, seeded["authored_row"])
    body = client.get(LIST, params={"query": number}).json()

    assert [row["id"] for row in body["data"]] == [seeded["authored_row"].id]


def test_the_list_sorts_by_the_inquiry_number(api):
    """Nulls last in both directions, and the order is total, like every other column."""
    client, db, _company_id, seeded = api

    ascending = client.get(LIST, params={"sort": "inquiry_no", "dir": "asc"}).json()
    descending = client.get(LIST, params={"sort": "inquiry_no", "dir": "desc"}).json()

    numbers = [row["inquiry_no"] for row in ascending["data"]]
    assert numbers == sorted(numbers)
    assert [row["inquiry_no"] for row in descending["data"]] == sorted(
        numbers, reverse=True
    )
    assert _inquiry_no_of(db, seeded["adopted_row"]) in numbers


# ---------------------------------------------------------------- the filters


def test_the_delivery_month_filter_is_the_sheet_tab(api):
    client, _db, _company_id, seeded = api

    january = client.get(LIST, params={"delivery_month": "2026-01"}).json()
    assert [row["id"] for row in january["data"]] == [seeded["adopted_row"].id]

    march = client.get(LIST, params={"delivery_month": "2026-03"}).json()
    assert [row["id"] for row in march["data"]] == [seeded["authored_row"].id]


def test_the_state_filter_narrows_to_one_state(api):
    client, _db, _company_id, seeded = api

    body = client.get(LIST, params={"state": INQUIRY_ACTIONED}).json()

    assert [row["id"] for row in body["data"]] == [seeded["adopted_row"].id]


def test_the_project_filter_narrows_to_one_project(api):
    client, _db, _company_id, seeded = api

    body = client.get(LIST, params={"project_id": seeded["project"].id}).json()

    assert [row["id"] for row in body["data"]] == [seeded["authored_row"].id]


def test_the_raised_date_filter_is_the_per_day_tab(api):
    client, _db, _company_id, seeded = api
    # `created_at` is naive UTC; the tab is a Malaysian day.
    raised_on = (seeded["adopted_row"].created_at + timedelta(hours=8)).date().isoformat()

    body = client.get(LIST, params={"raised_date": raised_on}).json()

    assert body["pagination"]["total"] == 3
    assert client.get(LIST, params={"raised_date": "2001-01-01"}).json()[
        "pagination"
    ]["total"] == 0


def test_the_raised_date_tab_is_a_malaysian_day_not_a_utc_one(api):
    client, db, company_id, seeded = api
    inquiry = db.get(OrderInquiry, seeded["adopted_row"].order_inquiry_id)
    # 00:30 MYT on 20 Aug is 16:30 UTC on 19 Aug; 23:30 MYT on 19 Aug is 15:30 UTC.
    after_midnight = _row(
        db,
        company_id,
        inquiry,
        item_code=f"ZZT-MIDNIGHT-{_uid()[:6]}",
        qty="1",
        created_at=datetime(2026, 8, 19, 16, 30),
    )
    before_midnight = _row(
        db,
        company_id,
        inquiry,
        item_code=f"ZZT-EVENING-{_uid()[:6]}",
        qty="1",
        created_at=datetime(2026, 8, 19, 15, 30),
    )
    db.commit()

    on_the_20th = {
        row["id"]
        for row in client.get(LIST, params={"raised_date": "2026-08-20"}).json()["data"]
    }
    on_the_19th = {
        row["id"]
        for row in client.get(LIST, params={"raised_date": "2026-08-19"}).json()["data"]
    }

    assert after_midnight.id in on_the_20th
    assert after_midnight.id not in on_the_19th
    assert before_midnight.id in on_the_19th
    assert before_midnight.id not in on_the_20th


def test_the_query_box_matches_the_sales_order_the_item_the_product_and_the_customer(api):
    client, _db, _company_id, seeded = api

    by_so = client.get(LIST, params={"query": seeded["core"].so_number}).json()
    assert {row["id"] for row in by_so["data"]} == {
        seeded["adopted_row"].id,
        seeded["undated_row"].id,
    }

    by_item = client.get(LIST, params={"query": seeded["authored_row"].item_code}).json()
    assert [row["id"] for row in by_item["data"]] == [seeded["authored_row"].id]

    by_product = client.get(LIST, params={"query": "Wall hung WC"}).json()
    assert [row["id"] for row in by_product["data"]] == [seeded["authored_row"].id]

    by_customer = client.get(LIST, params={"query": "Optad"}).json()
    assert {row["id"] for row in by_customer["data"]} == {
        seeded["adopted_row"].id,
        seeded["undated_row"].id,
    }


def test_the_supplier_and_po_are_the_ones_the_row_actually_links_to(api):
    client, db, company_id, seeded = api
    placed = _placed_supply(db, company_id, seeded["adopted_row"])

    body = client.get(LIST).json()
    adopted = next(row for row in body["data"] if row["id"] == seeded["adopted_row"].id)
    authored = next(row for row in body["data"] if row["id"] == seeded["authored_row"].id)

    assert adopted["supplier"] == placed["supplier"].supplier_name
    assert adopted["po_number"] == placed["purchase_order"].po_number
    # Nothing is invented for a row nobody has placed yet: blank is what their sheet
    # writes and blank is what this says.
    assert authored["supplier"] is None
    assert authored["po_number"] is None

    filtered = client.get(
        LIST, params={"supplier_id": placed["supplier"].id}
    ).json()
    assert [row["id"] for row in filtered["data"]] == [seeded["adopted_row"].id]


# ------------------------------------------------------------------- location


def test_a_row_with_a_stamped_stock_location_prints_it_unchanged(api):
    """`stock_location` is stamped once at raise time - the donor an order-back row left
    oversold, or a confirmed allocation's warehouse - and LOCATION just reads it back."""
    client, db, _company_id, seeded = api
    seeded["authored_row"].stock_location = "BRW-BB"
    db.add(seeded["authored_row"])
    db.commit()

    body = client.get(LIST).json()
    row = next(r for r in body["data"] if r["id"] == seeded["authored_row"].id)

    assert row["location"] == "BRW-BB"


def test_a_row_with_no_stamped_location_falls_back_to_its_lines_own_warehouse(api):
    """A plain Buy row raised straight off the board has no confirmed allocation and no
    donor, so the column falls back to the fulfilment warehouse already on the line's own
    core sales order line - a real answer, just not one purchasing confirmed themselves."""
    client, db, company_id, seeded = api
    warehouse = Warehouse(
        id=_uid(),
        company_id=company_id,
        warehouse_code=f"ZZT{_uid()[:6]}",
        warehouse_name=f"{MARKER} fulfilment WH",
    )
    db.add(warehouse)
    db.flush()
    core_line = SalesOrderLine(
        id=_uid(),
        company_id=company_id,
        sales_order_id=seeded["core"].id,
        product_id=_product(db, f"ZZT-COREP-{_uid()[:6]}", f"{MARKER} core line").id,
        warehouse_id=warehouse.id,
        qty_ordered=Decimal("85"),
    )
    db.add(core_line)
    db.flush()
    adopted_line = (
        db.query(ProjectSalesOrderLine)
        .filter(ProjectSalesOrderLine.id == seeded["adopted_row"].so_line_id)
        .one()
    )
    adopted_line.core_sales_order_line_id = core_line.id
    seeded["adopted_row"].stock_location = None
    db.add_all([adopted_line, seeded["adopted_row"]])
    db.commit()

    body = client.get(LIST).json()
    row = next(r for r in body["data"] if r["id"] == seeded["adopted_row"].id)

    assert row["location"] == warehouse.warehouse_code


def test_a_row_with_neither_a_stamped_nor_a_line_location_prints_blank(api):
    client, _db, _company_id, seeded = api

    body = client.get(LIST).json()
    row = next(r for r in body["data"] if r["id"] == seeded["authored_row"].id)

    assert row["location"] is None


def test_sorting_by_location_is_accepted(api):
    client, _db, _company_id, _seeded = api

    response = client.get(LIST, params={"sort": "location"})

    assert response.status_code == 200, response.text


# ------------------------------------------------------------------- sorting


def test_an_unknown_sort_column_is_refused_rather_than_quietly_ignored(api):
    client, _db, _company_id, _seeded = api

    response = client.get(LIST, params={"sort": "whatever"})

    assert response.status_code == 422


def test_an_unknown_direction_is_refused(api):
    client, _db, _company_id, _seeded = api

    assert client.get(LIST, params={"dir": "sideways"}).status_code == 422


@pytest.mark.parametrize("direction", ["asc", "desc"])
def test_a_row_with_no_delivery_date_sorts_last_in_both_directions(api, direction):
    client, _db, _company_id, seeded = api

    body = client.get(
        LIST, params={"sort": "delivery_date", "dir": direction}
    ).json()

    assert body["data"][-1]["id"] == seeded["undated_row"].id


def test_every_advertised_sort_column_answers(api):
    client, _db, _company_id, _seeded = api
    from app.services.order_inquiry_worklist_service import SORTABLE_FIELDS

    for field in sorted(SORTABLE_FIELDS):
        response = client.get(LIST, params={"sort": field})
        assert response.status_code == 200, f"{field}: {response.text}"


def test_the_route_and_the_service_agree_on_the_sortable_set():
    """FastAPI cannot build a `Literal` from a runtime set, so this is what keeps them
    from drifting: a column the route accepts and the service refuses is a 500."""
    from typing import get_args

    from app.api.v1.projects.order_inquiries import WorklistSort
    from app.services.order_inquiry_worklist_service import SORTABLE_FIELDS

    assert set(get_args(WorklistSort)) == set(SORTABLE_FIELDS)


def test_the_order_is_total_so_paging_neither_repeats_nor_drops_a_row(api):
    client, _db, _company_id, _seeded = api

    first = client.get(LIST, params={"limit": 2, "page": 1, "sort": "state"}).json()
    second = client.get(LIST, params={"limit": 2, "page": 2, "sort": "state"}).json()

    seen = [row["id"] for row in first["data"]] + [row["id"] for row in second["data"]]
    assert len(seen) == len(set(seen)) == 3


# ------------------------------------------------------------------- summary


def test_the_summary_totals_the_visible_set_and_still_lists_every_month(api):
    client, _db, _company_id, _seeded = api

    everything = client.get(f"{LIST}/summary").json()
    assert everything["total_rows"] == 3
    assert everything["total_qty"] == "182"
    assert everything["by_state"] == {
        "raised": 2,
        "actioned": 1,
        "cancelled": 0,
        "total": 3,
    }
    assert [entry["label"] for entry in everything["by_month"]] == ["JAN 26", "MAR 26"]
    assert everything["by_month"][0]["rows"] == 1
    assert everything["by_month"][0]["qty"] == "85"

    january = client.get(f"{LIST}/summary", params={"delivery_month": "2026-01"}).json()
    assert january["total_rows"] == 1
    # The month strip is the control that changes month, so narrowing to one month must
    # not leave the user with one tab and no way back.
    assert [entry["month"] for entry in january["by_month"]] == ["2026-01", "2026-03"]


def test_the_summary_offers_the_suppliers_and_projects_actually_present(api):
    client, db, company_id, seeded = api
    placed = _placed_supply(db, company_id, seeded["adopted_row"])

    body = client.get(f"{LIST}/summary").json()

    assert [entry["id"] for entry in body["suppliers"]] == [placed["supplier"].id]
    assert [entry["id"] for entry in body["projects"]] == [seeded["project"].id]


# --------------------------------------------------------------------- agent


def test_a_row_names_the_core_orders_sales_agent_when_the_chain_resolves(api):
    """Read off the same core sales order the SO DATE / S/O NO columns already join to."""
    client, db, company_id, seeded = api
    agent = SalesAgent(
        id=_uid(),
        company_id=company_id,
        sales_agent=f"ZZT{_uid()[:6].upper()}",
        person_label=f"{MARKER} Sean",
    )
    db.add(agent)
    db.flush()
    seeded["core"].sales_agent_id = agent.id
    db.add(seeded["core"])
    db.commit()

    body = client.get(LIST).json()
    adopted = next(row for row in body["data"] if row["id"] == seeded["adopted_row"].id)

    assert adopted["agent_code"] == agent.sales_agent
    assert adopted["agent_label"] == agent.person_label


def test_a_row_has_no_agent_when_the_chain_does_not_resolve(api):
    client, _db, _company_id, seeded = api

    body = client.get(LIST).json()
    # The adopted row's own core order carries no agent, and the authored row reaches no
    # core order at all - both are stated absences, never a guess.
    adopted = next(row for row in body["data"] if row["id"] == seeded["adopted_row"].id)
    authored = next(row for row in body["data"] if row["id"] == seeded["authored_row"].id)

    assert adopted["agent_code"] is None
    assert authored["agent_code"] is None


def test_sorting_by_agent_is_accepted(api):
    client, _db, _company_id, _seeded = api

    response = client.get(LIST, params={"sort": "agent"})

    assert response.status_code == 200, response.text


# -------------------------------------------------------------------- export


def test_the_workbook_is_one_sheet_per_delivery_month_with_their_heading_row(api):
    client, _db, _company_id, _seeded = api

    response = client.get(f"{LIST}/export")

    assert response.status_code == 200, response.text
    assert "attachment;" in response.headers["content-disposition"]
    book = openpyxl.load_workbook(io.BytesIO(response.content))
    assert book.sheetnames == ["JAN 26", "MAR 26", "NO DATE"]
    sheet = book["JAN 26"]
    assert sheet.cell(row=1, column=1).value == "ORDER INQUIRY"
    assert [cell.value for cell in sheet[2][:9]] == [
        "SO DATE",
        "S/O NO",
        "ITEM CODE",
        "QTY",
        "TOTAL QTY",
        "DELIVERY DATE",
        "PROJECT/CUSTOMER",
        "SUPPLIER",
        "PO NO ",
    ]
    assert sheet.cell(row=2, column=10).value == "LOCATION"
    assert sheet.max_row == 3


def test_the_workbook_honours_the_filters_the_screen_is_showing(api):
    client, _db, _company_id, _seeded = api

    response = client.get(f"{LIST}/export", params={"delivery_month": "2026-03"})

    book = openpyxl.load_workbook(io.BytesIO(response.content))
    assert book.sheetnames == ["MAR 26"]


def test_total_qty_totals_the_supplier_and_item_run_not_the_item_across_suppliers(api):
    client, db, company_id, seeded = api
    inquiry = db.get(OrderInquiry, seeded["adopted_row"].order_inquiry_id)
    abc = f"ZZT-ABC-{_uid()[:6]}"
    dfe = f"ZZT-DEF-{_uid()[:6]}"

    def raised(code: str, qty: str, day: int) -> OrderInquiryRow:
        return _row(
            db,
            company_id,
            inquiry,
            item_code=code,
            qty=qty,
            delivery_date=date(2026, 5, day),
        )

    # Supplier A: ABC 10 + ABC 4, DEF 5. Supplier B: ABC 3.
    a_abc_first, a_abc_second, a_def, b_abc = (
        raised(abc, "10", 4),
        raised(abc, "4", 5),
        raised(dfe, "5", 4),
        raised(abc, "3", 4),
    )
    supplier_a = _placed_supply(
        db, company_id, a_abc_first, supplier_name=f"{MARKER} A FACTORY"
    )["supplier"]
    _placed_supply(db, company_id, a_abc_second, supplier=supplier_a)
    _placed_supply(db, company_id, a_def, supplier=supplier_a)
    _placed_supply(db, company_id, b_abc, supplier_name=f"{MARKER} B FACTORY")

    response = client.get(f"{LIST}/export", params={"delivery_month": "2026-05"})

    book = openpyxl.load_workbook(io.BytesIO(response.content))
    sheet = book["MAY 26"]
    body = [[cell.value for cell in row[:9]] for row in sheet.iter_rows(min_row=3)]
    assert [(row[7], row[2], row[3], row[4]) for row in body] == [
        (supplier_a.supplier_name, abc, 10.0, None),
        (supplier_a.supplier_name, abc, 4.0, 14.0),
        (supplier_a.supplier_name, dfe, 5.0, None),
        (f"{MARKER} B FACTORY", abc, 3.0, None),
    ]


def test_an_empty_export_is_still_a_workbook_a_person_can_open(api):
    client, _db, _company_id, _seeded = api

    response = client.get(f"{LIST}/export", params={"query": "zzt-nothing-matches"})

    book = openpyxl.load_workbook(io.BytesIO(response.content))
    assert book.sheetnames == ["ORDER INQUIRY"]
    assert book.worksheets[0].max_row == 2


# ---------------------------------------------------- taken from PO / remaining open


def _line_on_authored_order(db, company_id: str, seeded: dict, *, qty: str, day: int):
    """A fresh SO line on the seed's own authored order, so its rows are addressable
    through the existing project inquiry without disturbing the seed's own two rows."""
    line = ProjectSalesOrderLine(
        id=_uid(),
        company_id=company_id,
        project_sales_order_id=seeded["authored"].id,
        line_no=99,
        product_id=_product(db, f"ZZT-FLOW-{_uid()[:6]}", f"{MARKER} flow line").id,
        description=f"{MARKER} flow line",
        qty=Decimal(qty),
        uom="UNIT",
        unit_price=Decimal("10.00"),
        amount=Decimal("0"),
        delivery_date=date(2026, 4, day),
    )
    db.add(line)
    db.flush()
    return line


def test_taken_from_po_and_remaining_open_reflect_the_g2_cascade_split(api):
    """One line, split by the cascade into a placed allocation (5) and a raised
    remainder (14): both rows report the SAME pair, because both describe the same
    line - what was actually taken off a PO, and what still flows to reorder planning."""
    client, db, company_id, seeded = api
    inquiry = db.get(OrderInquiry, seeded["authored_row"].order_inquiry_id)
    line = _line_on_authored_order(db, company_id, seeded, qty="19", day=1)
    placed = _row(
        db,
        company_id,
        inquiry,
        so_line_id=line.id,
        item_code=f"{MARKER}-SPLIT",
        qty="5",
        state=INQUIRY_PLACED,
        delivery_date=date(2026, 4, 1),
    )
    raised = _row(
        db,
        company_id,
        inquiry,
        so_line_id=line.id,
        item_code=f"{MARKER}-SPLIT",
        qty="14",
        state=INQUIRY_RAISED,
        delivery_date=date(2026, 4, 1),
    )
    db.commit()

    body = client.get(LIST, params={"delivery_month": "2026-04"}).json()
    by_id = {row["id"]: row for row in body["data"]}

    assert by_id[placed.id]["taken_from_po"] == "5"
    assert by_id[placed.id]["remaining_open"] == "14"
    assert by_id[raised.id]["taken_from_po"] == "5"
    # A raised row IS the uncovered remainder, so its own quantity counts towards it.
    assert by_id[raised.id]["remaining_open"] == "14"


def test_a_redirected_placed_row_no_longer_counts_toward_taken_from_po(api):
    """The captain's ruling, 21 Aug 2026: a placed row a planning change REDIRECTED to
    replenish the shared pool is real placed quantity, just not this line's anymore - it
    must not count toward `taken_from_po` (or it would read as this line's need being
    covered by a PO that is actually bound for the pool). The redirected row's OWN cells
    still show what it now is: its `location` reads the pool, and its `note` says so."""
    client, db, company_id, seeded = api
    inquiry = db.get(OrderInquiry, seeded["authored_row"].order_inquiry_id)
    line = _line_on_authored_order(db, company_id, seeded, qty="20", day=3)
    redirected = _row(
        db,
        company_id,
        inquiry,
        so_line_id=line.id,
        item_code=f"{MARKER}-REDIRECT",
        qty="12",
        state=INQUIRY_PLACED,
        delivery_date=date(2026, 4, 3),
        redirected_to_pool=True,
        stock_location="ZZT-BRW",
        note="Redirected to replenish ZZT-BRW (was ZZT-OWN) - planning change batch abcd1234.",
    )
    raised = _row(
        db,
        company_id,
        inquiry,
        so_line_id=line.id,
        item_code=f"{MARKER}-REDIRECT",
        qty="8",
        state=INQUIRY_RAISED,
        delivery_date=date(2026, 4, 3),
    )
    db.commit()

    body = client.get(LIST, params={"delivery_month": "2026-04"}).json()
    by_id = {row["id"]: row for row in body["data"]}

    # Neither row counts the redirected 12 as taken - it does not serve this line anymore.
    assert by_id[redirected.id]["taken_from_po"] == "0"
    assert by_id[raised.id]["taken_from_po"] == "0"
    assert by_id[raised.id]["remaining_open"] == "8"
    # The redirected row's own cells name what actually happened to it.
    assert by_id[redirected.id]["location"] == "ZZT-BRW"
    assert "Redirected" in (by_id[redirected.id]["note"] or "")


def test_an_unplaced_lines_row_reports_zero_taken_and_its_full_qty_as_remaining(api):
    client, db, company_id, seeded = api
    inquiry = db.get(OrderInquiry, seeded["authored_row"].order_inquiry_id)
    line = _line_on_authored_order(db, company_id, seeded, qty="12", day=2)
    raised = _row(
        db,
        company_id,
        inquiry,
        so_line_id=line.id,
        item_code=f"{MARKER}-UNPLACED",
        qty="12",
        delivery_date=date(2026, 4, 2),
    )
    db.commit()

    body = client.get(LIST, params={"delivery_month": "2026-04"}).json()
    row = next(r for r in body["data"] if r["id"] == raised.id)

    assert row["taken_from_po"] == "0"
    assert row["remaining_open"] == "12"


# ---------------------------------------------------------------- the po popup


def test_the_po_detail_route_answers_the_header_and_every_line(api):
    client, db, company_id, _seeded = api
    placed = _purchase_order(db, company_id)

    response = client.get(f"{BASE}/order-inquiries/po/{placed['order'].id}")

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["po_number"] == placed["order"].po_number
    assert body["supplier_code"] == placed["supplier"].supplier_code
    assert body["supplier_name"] == placed["supplier"].supplier_name
    assert body["expected_date"] == "2026-05-01"
    assert body["status"] == "active"
    assert len(body["lines"]) == 1
    line = body["lines"][0]
    assert line["sku"] == placed["product"].product_code
    assert line["product_name"] == placed["product"].product_name
    assert line["qty_ordered"] == "40"
    assert line["qty_received"] == "15"
    assert line["remaining"] == "25"
    assert line["location"] == placed["warehouse"].warehouse_code


def test_the_po_detail_route_404s_on_a_po_from_another_company(api):
    client, db, _company_id, _seeded = api
    from app.models.base import company_scope
    from app.models.company import Company

    other_id = _uid()
    with company_scope(db, None):
        db.add(Company(id=other_id, name=f"{MARKER} Other PO Co", code=f"ZZ{_uid()[:6]}"))
        db.flush()
        placed = _purchase_order(db, other_id)

    response = client.get(f"{BASE}/order-inquiries/po/{placed['order'].id}")

    assert response.status_code == 404


def test_the_po_detail_route_404s_on_an_unknown_id(api):
    client, _db, _company_id, _seeded = api

    response = client.get(f"{BASE}/order-inquiries/po/{_uid()}")

    assert response.status_code == 404


def test_the_po_detail_route_denies_a_user_without_the_view_permission(stranger_api):
    client, db, company_id, _seeded = stranger_api
    placed = _purchase_order(db, company_id)

    response = client.get(f"{BASE}/order-inquiries/po/{placed['order'].id}")

    assert response.status_code == 403


# ---------------------------------------------------------- auth and company


def test_a_user_without_the_module_read_is_refused(stranger_api):
    client, _db, _company_id, _seeded = stranger_api

    assert client.get(LIST).status_code == 403
    assert client.get(f"{LIST}/summary").status_code == 403
    assert client.get(f"{LIST}/export").status_code == 403


def test_another_companys_rows_are_not_on_this_companys_list(api):
    client, db, _company_id, _seeded = api
    from app.models.base import company_scope

    other_id = _uid()
    # The seed itself has to run unscoped, and it has to hand the scope BACK: leaving the
    # session on the all-companies scope would make this test pass by making every other
    # company visible, which is the opposite of what it is checking.
    with company_scope(db, None):
        db.add(Company(id=other_id, name=f"{MARKER} Other", code=f"ZZ{_uid()[:6]}"))
        db.flush()
        order = ProjectSalesOrder(
            id=_uid(),
            company_id=other_id,
            project_id=None,
            provisional_ref=f"ZZT-OTHER-{_uid()[:8]}",
            autocount_doc_no="ZZT-OTHER-SO",
            status="adopted",
        )
        db.add(order)
        db.flush()
        inquiry = _inquiry_for(db, other_id, order)
        _row(db, other_id, inquiry, item_code="ZZT-OTHER-ITEM", qty="1")

    body = client.get(LIST).json()

    assert all(row["item_code"] != "ZZT-OTHER-ITEM" for row in body["data"])
    assert body["pagination"]["total"] == 3
