"""Tests for app.services.excel_macro_stripper."""
from __future__ import annotations

import io
import os
import zipfile

import pytest
from openpyxl import Workbook

from app.services.excel_macro_stripper import (
    XLSX_MIME,
    MacroWorkbookError,
    extract_macro_template_xlsx,
    is_xlsm_filename,
    maybe_strip,
    maybe_strip_upload,
    strip_macros_to_xlsx,
)


def _build_synthetic_xlsm() -> bytes:
    """Build a minimal .xlsm-like zip by injecting xl/vbaProject.bin into an xlsx."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "hello"
    ws["B1"] = 42
    buf = io.BytesIO()
    wb.save(buf)
    src_bytes = buf.getvalue()

    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(src_bytes)) as src, zipfile.ZipFile(
        out, "w", zipfile.ZIP_DEFLATED
    ) as dst:
        for name in src.namelist():
            data = src.read(name)
            if name == "[Content_Types].xml":
                data = data.replace(
                    b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
                    b"application/vnd.ms-excel.sheet.macroEnabled.main+xml",
                )
            dst.writestr(name, data)
        dst.writestr("xl/vbaProject.bin", b"FAKE_VBA_BYTES_FOR_TEST")
    return out.getvalue()


def _zip_names(blob: bytes) -> list[str]:
    return zipfile.ZipFile(io.BytesIO(blob)).namelist()


def test_is_xlsm_filename():
    assert is_xlsm_filename("foo.xlsm")
    assert is_xlsm_filename("FOO.XLSM")
    assert not is_xlsm_filename("foo.xlsx")
    assert not is_xlsm_filename("foo.xls")
    assert not is_xlsm_filename(None)  # type: ignore[arg-type]


def test_strip_macros_drops_vba_bytes():
    src = _build_synthetic_xlsm()
    assert "xl/vbaProject.bin" in _zip_names(src)

    cleaned, name = strip_macros_to_xlsx(src, "stock balance.xlsm")

    assert name == "stock balance.xlsx"
    assert "xl/vbaProject.bin" not in _zip_names(cleaned)
    # Sanity: still a valid spreadsheet zip with the workbook part present.
    assert "xl/workbook.xml" in _zip_names(cleaned)


def test_strip_macros_preserves_sheet_values():
    from openpyxl import load_workbook

    src = _build_synthetic_xlsm()
    cleaned, _ = strip_macros_to_xlsx(src, "stock balance.xlsm")

    wb = load_workbook(io.BytesIO(cleaned), data_only=False)
    ws = wb.active
    assert ws["A1"].value == "hello"
    assert ws["B1"].value == 42


def test_maybe_strip_passthrough_for_xlsx():
    payload = b"not-really-xlsx-but-should-passthrough"
    out_bytes, out_name = maybe_strip(payload, "foo.xlsx")
    assert out_bytes is payload
    assert out_name == "foo.xlsx"


def test_maybe_strip_upload_normalizes_mime():
    src = _build_synthetic_xlsm()
    cleaned, name, mime = maybe_strip_upload(
        src, "stock balance.xlsm", "application/vnd.ms-excel.sheet.macroEnabled.12"
    )
    assert name == "stock balance.xlsx"
    assert mime == XLSX_MIME
    assert "xl/vbaProject.bin" not in _zip_names(cleaned)


def test_maybe_strip_upload_passthrough_for_xlsx():
    payload = b"abc"
    cleaned, name, mime = maybe_strip_upload(payload, "foo.xlsx", "text/plain")
    assert cleaned is payload
    assert name == "foo.xlsx"
    assert mime == "text/plain"


# Real macro workbook committed as the FE e2e fixture (3 sheets: Active Loc /
# Master / Template). Falls back to skip when the monorepo layout differs.
_LOCAL_FIXTURE = os.path.join(
    os.path.dirname(__file__),
    "..",
    "..",
    "sorento_crm_frontend",
    "e2e",
    "fixtures",
    "stock balance - Macro Version.xlsm",
)


@pytest.mark.skipif(
    not os.path.exists(_LOCAL_FIXTURE),
    reason="Real .xlsm fixture only available on the developer machine",
)
def test_strip_macros_on_real_fixture():
    with open(_LOCAL_FIXTURE, "rb") as fh:
        src = fh.read()
    cleaned, name = strip_macros_to_xlsx(src, "stock balance - Macro Version.xlsm")
    assert name.endswith(".xlsx")
    names = _zip_names(cleaned)
    assert "xl/vbaProject.bin" not in names
    assert not any("vba" in n.lower() for n in names)


def test_keep_data_sheet_prefers_named_candidate():
    """Multi-sheet xlsm → only the named data sheet survives (not sheet 0)."""
    from openpyxl import Workbook, load_workbook

    wb = Workbook()
    wb.active.title = "Active Loc"
    wb.active["A1"] = "locations"
    wb.create_sheet("Master")["A1"] = "macro-ui"
    wb.create_sheet("Template")["A1"] = "stock-data"
    base = io.BytesIO()
    wb.save(base)

    out = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(base.getvalue())) as src, zipfile.ZipFile(
        out, "w", zipfile.ZIP_DEFLATED
    ) as dst:
        for name in src.namelist():
            dst.writestr(name, src.read(name))
        dst.writestr("xl/vbaProject.bin", b"FAKE_VBA")

    cleaned, name, mime = maybe_strip_upload(
        out.getvalue(),
        "stock balance.xlsm",
        "application/vnd.ms-excel.sheet.macroEnabled.12",
        keep_data_sheet=True,
        data_sheet_candidates=["Template", "Stock List"],
    )
    assert name == "stock balance.xlsx"
    assert mime == XLSX_MIME
    assert "xl/vbaProject.bin" not in _zip_names(cleaned)

    wb2 = load_workbook(io.BytesIO(cleaned))
    assert wb2.sheetnames == ["Template"]
    assert wb2["Template"]["A1"].value == "stock-data"


def test_keep_data_sheet_falls_back_to_first():
    """No candidate matches → falls back to sheet 0."""
    from openpyxl import Workbook, load_workbook

    wb = Workbook()
    wb.active.title = "Data"
    wb.active["A1"] = 1
    wb.create_sheet("Extra")["A1"] = 2
    buf = io.BytesIO()
    wb.save(buf)

    cleaned, name, mime = maybe_strip_upload(
        buf.getvalue(),
        "stock.xlsx",
        XLSX_MIME,
        keep_data_sheet=True,
        data_sheet_candidates=["Nonexistent"],
    )
    assert name == "stock.xlsx"
    assert mime == XLSX_MIME
    wb2 = load_workbook(io.BytesIO(cleaned))
    assert wb2.sheetnames == ["Data"]


def test_keep_data_sheet_case_insensitive_match():
    from openpyxl import Workbook, load_workbook

    wb = Workbook()
    wb.active.title = "Active Loc"
    wb.create_sheet("template")["A1"] = "x"
    buf = io.BytesIO()
    wb.save(buf)

    cleaned, _, _ = maybe_strip_upload(
        buf.getvalue(),
        "stock.xlsx",
        XLSX_MIME,
        keep_data_sheet=True,
        data_sheet_candidates=["Template"],
    )
    wb2 = load_workbook(io.BytesIO(cleaned))
    assert wb2.sheetnames == ["template"]


# ---------------------------------------------------------------------------
# extract_macro_template_xlsx - Stock List macro pipeline
# (docs/plans/PLAN-stock-list-xlsm-macro-upload.md)
# ---------------------------------------------------------------------------


def _multi_sheet_xlsm_bytes(*, with_template: bool) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "Active Loc"
    wb.active["A1"] = "locations"
    wb.create_sheet("Master")["A1"] = "macro-ui"
    if with_template:
        ws = wb.create_sheet("Template")
        ws["A1"] = "Item Code"
        ws["A2"] = "ABC-123"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_extract_template_passthrough_for_non_xlsm():
    payload = b"raw-xlsx-bytes"
    out_bytes, out_name, out_mime = extract_macro_template_xlsx(payload, "stock.xlsx", "m/t")
    assert out_bytes is payload
    assert out_name == "stock.xlsx"
    assert out_mime == "m/t"

    out_bytes, out_name, out_mime = extract_macro_template_xlsx(payload, "stock.xls", "m/t")
    assert out_bytes is payload
    assert out_name == "stock.xls"


def test_extract_template_keeps_only_template_sheet():
    from openpyxl import load_workbook

    cleaned, name, mime = extract_macro_template_xlsx(
        _multi_sheet_xlsm_bytes(with_template=True),
        "stock balance - Macro Version.xlsm",
        "application/vnd.ms-excel.sheet.macroEnabled.12",
    )
    assert name == "stock balance - Macro Version.xlsx"
    assert mime == XLSX_MIME
    assert "xl/vbaProject.bin" not in _zip_names(cleaned)
    wb = load_workbook(io.BytesIO(cleaned))
    assert wb.sheetnames == ["Template"]
    assert wb["Template"]["A2"].value == "ABC-123"


def test_extract_template_single_sheet_kept_without_template_name():
    from openpyxl import Workbook, load_workbook

    wb = Workbook()
    wb.active.title = "Sheet1"
    wb.active["A1"] = "x"
    buf = io.BytesIO()
    wb.save(buf)

    cleaned, name, mime = extract_macro_template_xlsx(buf.getvalue(), "single.xlsm", None)
    assert name == "single.xlsx"
    assert mime == XLSX_MIME
    assert load_workbook(io.BytesIO(cleaned)).sheetnames == ["Sheet1"]


def test_extract_template_multi_sheet_without_template_raises():
    with pytest.raises(MacroWorkbookError) as exc:
        extract_macro_template_xlsx(
            _multi_sheet_xlsm_bytes(with_template=False), "bad.xlsm", None
        )
    assert "Template" in str(exc.value)
    assert "Active Loc" in str(exc.value)


def test_extract_template_unreadable_xlsm_raises():
    with pytest.raises(MacroWorkbookError):
        extract_macro_template_xlsx(b"definitely-not-a-zip", "corrupt.xlsm", None)


def test_extract_template_require_data_rejects_empty_template():
    """DO-lines rule: macro Template present but unpopulated → actionable error."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "Master"
    wb.active["A1"] = "noise"
    t = wb.create_sheet("Template")
    t.append(["Doc No", "Item Code", "Location", "Qty"])  # headers only
    buf = io.BytesIO()
    wb.save(buf)

    with pytest.raises(MacroWorkbookError) as exc:
        extract_macro_template_xlsx(buf.getvalue(), "lines.xlsm", None, require_data=True)
    assert "no data rows" in str(exc.value)


def test_extract_template_require_data_accepts_populated_template():
    from openpyxl import Workbook, load_workbook

    wb = Workbook()
    wb.active.title = "Master"
    t = wb.create_sheet("Template")
    t.append(["Doc No", "Item Code", "Location", "Qty"])
    t.append(["PS202605-0001", "SRT-1", "BRW", 5])
    buf = io.BytesIO()
    wb.save(buf)

    cleaned, name, _ = extract_macro_template_xlsx(
        buf.getvalue(), "lines.xlsm", None, require_data=True
    )
    assert name == "lines.xlsx"
    assert load_workbook(io.BytesIO(cleaned)).sheetnames == ["Template"]


def test_extract_template_bakes_formulas_to_values():
    """data_only=True load means formula text never survives into the output."""
    from openpyxl import Workbook, load_workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    ws["A1"] = 2
    ws["B1"] = "='Template'!A1*2"
    wb.create_sheet("Master")["A1"] = 99
    buf = io.BytesIO()
    wb.save(buf)

    cleaned, _, _ = extract_macro_template_xlsx(buf.getvalue(), "f.xlsm", None)
    out = load_workbook(io.BytesIO(cleaned), data_only=False)
    val = out["Template"]["B1"].value
    assert not (isinstance(val, str) and val.startswith("="))


@pytest.mark.skipif(
    not os.path.exists(_LOCAL_FIXTURE),
    reason="Real .xlsm fixture only available on the developer machine",
)
def test_extract_template_real_fixture():
    from openpyxl import load_workbook

    with open(_LOCAL_FIXTURE, "rb") as fh:
        src = fh.read()
    cleaned, name, mime = extract_macro_template_xlsx(
        src, "stock balance - Macro Version.xlsm", "application/vnd.ms-excel.sheet.macroEnabled.12"
    )
    assert name == "stock balance - Macro Version.xlsx"
    assert mime == XLSX_MIME
    names = _zip_names(cleaned)
    assert not any("vba" in n.lower() for n in names)
    wb = load_workbook(io.BytesIO(cleaned), read_only=True)
    try:
        assert wb.sheetnames == ["Template"]
        ws = wb["Template"]
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        assert headers == ["Item Code", "Item Description", "Location", "On Hand Qty"]
    finally:
        wb.close()


@pytest.mark.skipif(
    not os.path.exists(_LOCAL_FIXTURE),
    reason="Real .xlsm fixture only available on the developer machine",
)
def test_keep_data_sheet_real_fixture():
    """User fixture has 3 sheets - only 'Template' (the data sheet) survives."""
    from openpyxl import load_workbook

    with open(_LOCAL_FIXTURE, "rb") as fh:
        src = fh.read()

    cleaned, name, mime = maybe_strip_upload(
        src,
        "stock balance - Macro Version.xlsm",
        "application/vnd.ms-excel.sheet.macroEnabled.12",
        keep_data_sheet=True,
        data_sheet_candidates=["Template", "Stock List", "Stock_List", "Sheet1"],
    )
    assert name == "stock balance - Macro Version.xlsx"
    assert mime == XLSX_MIME

    out_wb = load_workbook(io.BytesIO(cleaned), data_only=False, read_only=True)
    try:
        assert out_wb.sheetnames == ["Template"]
    finally:
        out_wb.close()
