"""R13 - with a retained stock list, the xlsx IS their file.

`PLAN-scm-fulfilment-feedback-p4.md` section 4, AC-D1 / D2 / D3 / D6. Ms Tee's ask, in her
words: "send them the same sheet back with the quantity to load filled in", and the captain's
round-2 ruling (Q1): the asked quantity is an APPENDED column K, their ten columns untouched.

So the golden test below is not a description of the July file, it is a comparison against it:
their own `2026-7-27  库存明细.xlsx` goes in (committed under `documentation/plans/scm/
fixtures/`), the export comes out, and every cell they wrote - value, font, fill, border,
number format, merge, width, row height - has to still be there. The five-column sheet of our
own is gone (AC-D6): a supplier must not receive a different document because of a file WE
failed to keep.

THE ROUND TRIP IS STILL THE CONTRACT (AC-C2 of part 3). The supplier's next stock list is very
often this file with the numbers changed, so whatever goes out has to come back in through
`supplier_inventory_reader.read_workbook` - asserted here on both layouts.
"""
from __future__ import annotations

import uuid
from io import BytesIO
from pathlib import Path

import openpyxl

from app.services.scm import container_request_xlsx as svc
from app.services.scm import supplier_document_model as model
from app.services.scm.supplier_inventory_reader import read_workbook
from tests._pg_fixture import pg_session
from tests.scm._outstanding_workbooks import require_aliases
from tests.scm.test_loading_plan import World

FIXTURE = (
    Path(__file__).resolve().parents[3]
    / "documentation"
    / "plans"
    / "scm"
    / "fixtures"
    / "jinbaichuan-stock-list-2026-07-27.xlsx"
)

SUPPLIER = {"supplier_code": "JBC", "supplier_name": "JINBAICHUAN"}

#: Where their sheet ends: header on row 2, data 3..119, `合计：` on 120, ten columns.
HEADER_ROW = 2
FIRST_DATA_ROW = 3
LAST_DATA_ROW = 119
TOTALS_ROW = 120
THEIR_COLS = 10
QTY_COL = 11


def _world(db) -> World:
    require_aliases(db, "supplier_inventory")
    return World(db)


def _sheet_model(
    db, lines: list[dict], *, monkeypatch, retained: bytes | None = ..., supplier_id=None
):
    data = FIXTURE.read_bytes() if retained is ... else retained
    monkeypatch.setattr(model, "_retained_stock_list", lambda _db, _sid: data)
    return model.build(
        db,
        supplier=SUPPLIER,
        supplier_id=str(supplier_id or uuid.uuid4()),
        lines=lines,
    )


def _line(item_code: str, qty: float, name: str = "Basin") -> dict:
    return {"item_code": item_code, "product_name": name, "qty": qty, "product_id": None}


def _open(data: bytes):
    return openpyxl.load_workbook(BytesIO(data)).active


def _style(cell) -> tuple:
    color = getattr(cell.font.color, "rgb", None) if cell.font.color else None
    fill = getattr(cell.fill.fgColor, "rgb", None) if cell.fill.fill_type else None
    return (
        cell.font.name,
        cell.font.sz,
        cell.font.b,
        color if isinstance(color, str) else None,
        cell.fill.fill_type,
        fill if isinstance(fill, str) else None,
        cell.border.left.style,
        cell.border.right.style,
        cell.border.top.style,
        cell.border.bottom.style,
        cell.alignment.horizontal,
        cell.alignment.vertical,
        cell.number_format,
    )


# --------------------------------------------------------------------------- #
# their workbook, cell for cell
# --------------------------------------------------------------------------- #


def test_every_cell_they_wrote_survives_untouched(monkeypatch):
    # AC-D1. The whole point of R13: this is THEIR file with one column written into it, not
    # a copy of their data in a workbook of ours. Value, font, fill, border and number format,
    # on every one of their 1,200 cells.
    with pg_session() as db:
        _world(db)
        sheet = _sheet_model(db, [_line("SRTWC8355-RL-250", 300)], monkeypatch=monkeypatch)

        out = _open(svc.render(sheet))
        theirs = _open(FIXTURE.read_bytes())

        for r in range(1, TOTALS_ROW + 1):
            for c in range(1, THEIR_COLS + 1):
                mine, yours = out.cell(row=r, column=c), theirs.cell(row=r, column=c)
                assert mine.value == yours.value, f"value at {mine.coordinate}"
                assert _style(mine) == _style(yours), f"style at {mine.coordinate}"


def test_their_merges_widths_and_row_heights_survive(monkeypatch):
    # AC-D1. A family is one 序号 and one volume across nine rows (`A3:A11`, `I3:I11`);
    # unmerging it would print the volume nine times, which reads as nine times the volume.
    with pg_session() as db:
        _world(db)
        sheet = _sheet_model(db, [_line("SRTWC8355-RL-250", 300)], monkeypatch=monkeypatch)

        out = _open(svc.render(sheet))
        theirs = _open(FIXTURE.read_bytes())

        assert sorted(str(m) for m in out.merged_cells.ranges) == sorted(
            str(m) for m in theirs.merged_cells.ranges
        )
        for letter in "ABCDEFGHIJ":
            assert out.column_dimensions[letter].width == (
                theirs.column_dimensions[letter].width
            ), letter
        for r in (1, 2, 3, 26, 119, 120):
            assert out.row_dimensions[r].height == theirs.row_dimensions[r].height, r


def test_column_k_carries_the_ask_and_is_styled_like_their_last_column(monkeypatch):
    # Q1 / AC-D1. Appended, never inserted; and it has to LOOK like a column of theirs or the
    # sheet reads as two documents stapled together.
    with pg_session() as db:
        _world(db)
        sheet = _sheet_model(
            db,
            [_line("SRTWC8355-RL-250", 300), _line("SRTSP131", 12)],
            monkeypatch=monkeypatch,
        )

        out = _open(svc.render(sheet))

        assert out.cell(row=HEADER_ROW, column=QTY_COL).value == svc.QTY_TO_LOAD_HEADER
        assert _style(out.cell(row=HEADER_ROW, column=QTY_COL)) == _style(
            out.cell(row=HEADER_ROW, column=THEIR_COLS)
        )
        asked = {
            out.cell(row=r, column=2).value: out.cell(row=r, column=QTY_COL).value
            for r in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1)
        }
        assert asked["SRTWC8355-RL-250"] == 300
        assert asked["SRTSP131"] == 12
        assert asked["SRTWC286-SH-150NEW"] is None
        assert out.column_dimensions["K"].width == out.column_dimensions["J"].width


def test_a_zero_ask_leaves_the_cell_empty(monkeypatch):
    # AC-D3 (AC-C3 of part 3 stands). A zero reads as "pack none of these".
    with pg_session() as db:
        _world(db)
        sheet = _sheet_model(db, [_line("SRTSP131", 0)], monkeypatch=monkeypatch)

        out = _open(svc.render(sheet))

        row = next(
            r
            for r in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1)
            if out.cell(row=r, column=2).value == "SRTSP131"
        )
        assert out.cell(row=row, column=QTY_COL).value is None


def test_the_totals_row_sums_our_column_too(monkeypatch):
    # AC-D1. Their three sums are their formulas, untouched; ours is one more of the same.
    with pg_session() as db:
        _world(db)
        sheet = _sheet_model(db, [_line("SRTSP131", 12)], monkeypatch=monkeypatch)

        out = _open(svc.render(sheet))

        assert out.cell(row=TOTALS_ROW, column=6).value == "=SUM(F3:F119)"
        assert out.cell(row=TOTALS_ROW, column=9).value == "=SUM(I3:I119)"
        assert out.cell(row=TOTALS_ROW, column=QTY_COL).value == "=SUM(K3:K119)"


def test_a_product_they_never_listed_is_appended_and_the_totals_row_moves_down(monkeypatch):
    # AC-D2. It is still part of the ask. Their 合计 row keeps its formulas and widens to
    # cover the rows we added - a total that stopped short of them would understate the ask.
    with pg_session() as db:
        _world(db)
        sheet = _sheet_model(
            db,
            [_line("SRTSP131", 12), _line("ZZT-NEW-1", 80, "New basin")],
            monkeypatch=monkeypatch,
        )

        out = _open(svc.render(sheet))

        assert out.cell(row=120, column=1).value == 39  # their last 序号 is 38
        assert out.cell(row=120, column=2).value == "ZZT-NEW-1"
        assert out.cell(row=120, column=10).value == model.NOT_ON_LIST_REMARK
        assert out.cell(row=120, column=QTY_COL).value == 80
        assert _style(out.cell(row=120, column=2)) == _style(
            out.cell(row=LAST_DATA_ROW, column=2)
        )
        assert out.row_dimensions[120].height == out.row_dimensions[LAST_DATA_ROW].height

        assert out.cell(row=121, column=1).value == "合计："
        assert out.cell(row=121, column=6).value == "=SUM(F3:F120)"
        assert out.cell(row=121, column=9).value == "=SUM(I3:I120)"
        assert out.cell(row=121, column=QTY_COL).value == "=SUM(K3:K120)"
        assert out.cell(row=120, column=1).value != "合计："
        merges = {str(m) for m in out.merged_cells.ranges}
        assert "A121:E121" in merges and "A120:E120" not in merges


def test_the_export_reads_back_through_the_stock_list_reader(monkeypatch):
    # AC-C2 of part 3, and the point of the whole slice: the supplier answers with the file we
    # sent them. An export the reader cannot parse breaks the loop silently.
    with pg_session() as db:
        _world(db)
        sheet = _sheet_model(db, [_line("SRTSP131", 12)], monkeypatch=monkeypatch)

        out = read_workbook(svc.render(sheet), db=db)

        assert out.ok, out.missing_columns
        assert [r.item_code for r in out.rows][:2] == [
            "SRTWC286-SH-150NEW",
            "SRTWC286-SH-180",
        ]
        assert out.rows[0].qty_unfinished == 0
        assert out.rows[1].qty_unfinished == 101


# --------------------------------------------------------------------------- #
# no retained file
# --------------------------------------------------------------------------- #


def test_without_a_retained_file_the_sheet_is_their_layout_in_our_hand(monkeypatch):
    # AC-D6. Same eleven columns, same title line, same yellow fields, same 合计 row - what
    # changes is only that we have no merges to draw, because we hold no family information.
    with pg_session() as db:
        w = _world(db)
        w.stock("A", packed=120, unfinished=340, cbm=0.21)
        sheet = _sheet_model(
            db,
            [_line(w.product("A").product_code, 500)],
            monkeypatch=monkeypatch,
            retained=None,
            supplier_id=w.supplier.id,
        )

        out = _open(svc.render(sheet))

        assert [out.cell(row=2, column=c).value for c in range(1, QTY_COL + 1)] == [
            "序号",
            "型号",
            "商标",
            "规格",
            "品名",
            "包装好库存",
            "空瓷",
            "体积(cbm)",
            "总体积(cbm)",
            "备注",
            svc.QTY_TO_LOAD_HEADER,
        ]
        assert out.cell(row=1, column=1).value == model.NO_FILE_TITLE
        assert out.cell(row=2, column=1).font.b is True
        assert out.cell(row=3, column=2).fill.fgColor.rgb == "FFFFFF00"
        assert out.cell(row=3, column=2).font.name == "宋体"
        assert out.row_dimensions[3].height == 18.75
        assert out.column_dimensions["B"].width == 28.7109375
        assert out.cell(row=4, column=1).value == "合计："
        assert out.cell(row=4, column=6).value == "=SUM(F3:F3)"
        assert out.cell(row=4, column=QTY_COL).value == "=SUM(K3:K3)"


def test_without_a_retained_file_the_export_still_reads_back(monkeypatch):
    # The round trip again, on the branch that has no file to copy: the no-file layout is not
    # allowed to be the one shape the reader cannot take back.
    with pg_session() as db:
        w = _world(db)
        w.stock("A", packed=120, unfinished=340)
        sheet = _sheet_model(
            db,
            [_line(w.product("A").product_code, 500)],
            monkeypatch=monkeypatch,
            retained=None,
            supplier_id=w.supplier.id,
        )

        out = read_workbook(svc.render(sheet), db=db)

        assert out.ok, out.missing_columns
        assert [r.item_code for r in out.rows] == [w.product("A").product_code]
        assert out.rows[0].qty_packed == 120
        assert out.rows[0].qty_unfinished == 340


def test_a_stored_file_that_will_not_open_still_produces_a_sheet(monkeypatch):
    # A send must never die because a stored file is corrupt: the ask is the point.
    with pg_session() as db:
        w = _world(db)
        sheet = _sheet_model(
            db,
            [_line(w.product("A").product_code, 500)],
            monkeypatch=monkeypatch,
            retained=b"not a workbook",
        )

        out = _open(svc.render(sheet))

        assert out.cell(row=2, column=QTY_COL).value == svc.QTY_TO_LOAD_HEADER


def test_build_goes_from_the_database_straight_to_bytes(monkeypatch):
    # `build` is what the send path and the download route call; `render` is what the model
    # test drives. They must not be able to disagree.
    with pg_session() as db:
        w = _world(db)
        monkeypatch.setattr(model, "_retained_stock_list", lambda _db, _sid: None)

        data = svc.build(
            db,
            supplier=SUPPLIER,
            supplier_id=str(w.supplier.id),
            lines=[_line(w.product("A").product_code, 500)],
        )

        assert _open(data).cell(row=2, column=QTY_COL).value == svc.QTY_TO_LOAD_HEADER


def test_the_filename_names_the_supplier_and_the_day():
    name = svc.filename({"supplier_code": "JBC"})

    assert name.startswith("container-request-JBC-")
    assert name.endswith(".xlsx")
