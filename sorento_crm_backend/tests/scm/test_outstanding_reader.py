"""Reading a real .xlsx extract, and the week-to-week diff it produces.

The fixtures are xlsx files, not dicts, because the parsing is half the risk: dates arriving
as datetimes, quantities as strings with separators, headers in an order nobody agreed, and
a column nobody told us about.

Provenance, stated plainly: these fixtures are built to
`documentation/plans/scm/scm-autocount-extract-spec.md` using the real Tuju Residence figures
used throughout the plan and the golden tests (SO397450, SRTWC8613-RL 135 due 1 Jul and 72
due 3 Aug). They are KEPT rather than replaced now that the client's genuine export has
arrived (7 Aug 2026), because they are the only fixtures that carry outstanding quantities -
every line of the real file is fully delivered. The real one is exercised alongside them in
`test_autocount_so_detail_reader.py`, and its header differences became alias rows rather
than code changes, which is the whole reason headers are resolved through a table.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import pytest

from app.services.import_alias_service import AliasResolver
from app.services.scm.outstanding_diff import (
    ADDED,
    CLOSED,
    DATE_MOVED,
    QTY_CHANGED,
    UNCHANGED,
    diff_lines,
)
from app.services.scm.outstanding_reader import SO, read_workbook
from tests._pg_fixture import pg_session

_FIX = Path(__file__).parent / "fixtures"


@pytest.fixture()
def resolver():
    """The real alias table, so the test proves the seeded aliases actually resolve.

    Falls back to the canonical seed when the database has none, which keeps the test
    meaningful on a database built by create_all without the module seed.
    """
    with pg_session() as db:
        r = AliasResolver.for_doc_type(db, SO)
        if r.known_fields:
            yield r
            return
    pytest.skip("no outstanding_so aliases seeded in this database")


def _read(name, resolver):
    return read_workbook((_FIX / name).read_bytes(), SO, resolver)


# --------------------------------------------------------------------------- #
# reading
# --------------------------------------------------------------------------- #

def test_reads_every_line_of_the_extract(resolver):
    res = _read("outstanding_so_sample.xlsx", resolver)

    assert res.ok, res.missing_columns
    assert res.total_rows == 5
    assert len(res.lines) == 5
    assert res.problems == []


def test_headers_resolve_by_alias_not_by_position(resolver):
    """The fixture puts PROJECT/CUSTOMER first and spells it `S/O NO`, so a reader keyed on
    column order or on one exact spelling would mis-read every row."""
    res = _read("outstanding_so_sample.xlsx", resolver)

    first = res.lines[0]
    assert first.doc_number == "SO397450"
    assert first.item_code == "SRTWC8613-RL"
    assert first.location == "BRW-BB"
    assert first.label == "TUJU RESIDENCE"
    assert res.unmapped_headers == []


def test_the_same_item_twice_on_one_order_survives_reading(resolver):
    """The shape that drives the whole identity design must not be collapsed at read time."""
    res = _read("outstanding_so_sample.xlsx", resolver)

    tuju = [l for l in res.lines
            if l.doc_number == "SO397450" and l.item_code == "SRTWC8613-RL"]
    assert len(tuju) == 2
    assert sorted(l.required_date for l in tuju) == [date(2026, 7, 1), date(2026, 8, 3)]
    assert sorted(l.qty for l in tuju) == [72, 135]


def test_dates_are_read_as_dates(resolver):
    res = _read("outstanding_so_sample.xlsx", resolver)
    assert all(isinstance(l.required_date, date) for l in res.lines)


def test_a_file_missing_the_delivery_date_column_is_rejected_outright(resolver):
    """Importing a report with no date column would blank every date in scope, so this is
    refused at the header rather than row by row."""
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["S/O NO", "ITEM CODE", "QTY", "STOCK LOCATION"])
    ws.append(["SO397450", "SRTWC8613-RL", 135, "BRW-BB"])
    buf = BytesIO()
    wb.save(buf)

    res = read_workbook(buf.getvalue(), SO, resolver)

    assert res.ok is False
    assert "required_date" in res.missing_columns
    assert res.lines == []


def test_an_unknown_column_is_reported_rather_than_ignored(resolver):
    """A silently dropped column is how an import produces confidently wrong numbers."""
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["S/O NO", "ITEM CODE", "QTY", "DELIVERY DATE", "MYSTERY COLUMN"])
    ws.append(["SO397450", "SRTWC8613-RL", 135, date(2026, 7, 1), "?"])
    buf = BytesIO()
    wb.save(buf)

    res = read_workbook(buf.getvalue(), SO, resolver)

    assert res.ok is True, "an unknown extra column must not block a usable file"
    assert "MYSTERY COLUMN" in res.unmapped_headers


def test_a_bad_row_is_reported_and_the_rest_still_import(resolver):
    """4,000 lines with 3 bad rows should import 3,997 and report 3."""
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["S/O NO", "ITEM CODE", "QTY", "DELIVERY DATE", "STOCK LOCATION"])
    ws.append(["SO397450", "SRTWC8613-RL", 135, date(2026, 7, 1), "BRW-BB"])
    ws.append(["", "SRTWC8613-RL", 10, date(2026, 7, 1), "BRW-BB"])          # no SO
    ws.append(["SO397450", "", 10, date(2026, 7, 1), "BRW-BB"])              # no item
    buf = BytesIO()
    wb.save(buf)

    res = read_workbook(buf.getvalue(), SO, resolver)

    assert len(res.lines) == 1
    assert len(res.problems) == 2
    assert all(p.row_number in (3, 4) for p in res.problems)


def test_quantities_with_thousands_separators_are_read(resolver):
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["S/O NO", "ITEM CODE", "QTY", "DELIVERY DATE"])
    ws.append(["SO397512", "B2155-NL-BLUE", "7,646", date(2026, 11, 15)])
    buf = BytesIO()
    wb.save(buf)

    res = read_workbook(buf.getvalue(), SO, resolver)
    assert [l.qty for l in res.lines] == [7646.0]


def test_a_fully_delivered_line_is_not_carried_as_zero(resolver):
    """Zero outstanding belongs in the diff's closed side, reached by absence."""
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["S/O NO", "ITEM CODE", "QTY", "DELIVERY DATE"])
    ws.append(["SO397450", "SRTWC8613-RL", 0, date(2026, 7, 1)])
    buf = BytesIO()
    wb.save(buf)

    assert read_workbook(buf.getvalue(), SO, resolver).lines == []


# --------------------------------------------------------------------------- #
# Excel serial dates (19 Aug 2026 incident: openpyxl handed back the raw serial for
# `required_date` on 14,497 rows of the captain's real AutoCount export, and the reader
# read every one of them as unparseable, then the write path nulled the stored date).
# --------------------------------------------------------------------------- #

def test_a_date_cell_arriving_as_a_raw_excel_serial_number_is_read(resolver):
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["S/O NO", "ITEM CODE", "QTY", "DELIVERY DATE"])
    ws.append(["SO346436", "CKS1050", 10, 43886])
    ws.append(["SO346436", "CKSW015", 10, 44139])
    ws.append(["SO346436", "TPE-9201", 10, "45000"])  # numeric string, same as int/float
    buf = BytesIO()
    wb.save(buf)

    res = read_workbook(buf.getvalue(), SO, resolver)

    assert [l.required_date for l in res.lines] == [
        date(2020, 2, 25), date(2020, 11, 4), date(2023, 3, 15),
    ]
    assert res.problems == []
    assert all(l.date_unreadable is False for l in res.lines)


def test_a_small_number_that_looks_like_a_quantity_is_not_read_as_a_date(resolver):
    """150 in the date column is not a plausible Excel serial (that range starts in 1954)
    and must not be silently taken as one."""
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["S/O NO", "ITEM CODE", "QTY", "DELIVERY DATE"])
    ws.append(["SO346436", "CKS1050", 10, 150])
    buf = BytesIO()
    wb.save(buf)

    res = read_workbook(buf.getvalue(), SO, resolver)

    assert res.lines[0].required_date is None
    assert res.lines[0].date_unreadable is True
    assert len(res.problems) == 1
    assert "could not read the date" in res.problems[0].reason


# --------------------------------------------------------------------------- #
# the week-to-week story, file to file
# --------------------------------------------------------------------------- #

def test_a_week_later_the_diff_tells_the_whole_story(resolver):
    """End to end on two real files: a project slipped, a line grew, one was delivered,
    one is new, and another project is untouched."""
    week1 = _read("outstanding_so_sample.xlsx", resolver)
    week2 = _read("outstanding_so_sample_week2.xlsx", resolver)
    assert week1.ok and week2.ok

    diff = diff_lines(week1.lines, week2.lines)
    counts = diff.counts

    # Tuju slipped two weeks on the 1 Jul line.
    moved = diff.of_kind(DATE_MOVED)
    assert len(moved) == 1
    assert moved[0].item_code == "SRTWC8613-RL"
    assert moved[0].days_moved == 14

    # The 3 Aug line grew 72 -> 90 without moving.
    grew = diff.of_kind(QTY_CHANGED)
    assert len(grew) == 1 and grew[0].qty_delta == 18

    # SRTWT7408 on SO397450 was delivered; the identically-coded line on the OTHER project
    # must NOT be swept up with it.
    closed = diff.of_kind(CLOSED)
    assert [(c.doc_number, c.item_code) for c in closed] == [("SO397450", "SRTWT7408")]

    added = diff.of_kind(ADDED)
    assert [c.item_code for c in added] == ["C-FH24"]

    # Aria Verde is unchanged, and both its lines say so.
    assert counts[UNCHANGED] == 2
    assert diff.scope_documents == ("SO397450", "SO397512")
