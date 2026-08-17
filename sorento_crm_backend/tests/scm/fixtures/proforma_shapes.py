"""The two real proforma-invoice shapes, reproduced cell-for-cell.

Both builders below were built by loading the ORIGINAL supplier files with
``openpyxl.load_workbook(..., data_only=True)`` and dumping
``ws.cell(row=r, column=c).value`` for every row/column in the used range, then
copying the result in literally. Column positions, header spellings (including the
newline in ``净重\\n(kg)``), blank labelled cells, totals rows and blank numbered rows
are exactly what the real files hold. The two originals are NOT committed to the repo
(``KAILU形式发票(Sorento)260717.xlsx`` carries real bank details); Kailu's bank-detail
rows below are placeholder text, not the real values.

``item_code_map`` lets a caller substitute item codes (e.g. a test's own seeded
product codes) without touching anything else about the shape: every cell whose value
equals a key is replaced with the mapped value, wherever in the sheet it sits.
"""
from __future__ import annotations

import datetime
from io import BytesIO
from typing import Optional

import openpyxl

_INVOICE_DATE = datetime.datetime(2026, 7, 31, 0, 0)

# --------------------------------------------------------------------------------- #
# Jinbaichuan multi-block pre-loading list -- "2026-7-31 SORENTO 预装清单.xlsx"
# --------------------------------------------------------------------------------- #

_TITLE = (
    "潮州市金百川卫浴科技有限公司 \n CHAOZHOU JINBAICHUAN SANITARY WARE TECHNOLOGY "
    "CO.,LTD\nTEL：13308786682  18144411999\nProforma Invoice\n\n"
)
_ADDRESS = (
    "No.5, Jalan Astana 2/KU2, Bandar Bukit Raja, 41050 Klang, Selangor, Malaysia."
)
_HEADER19 = [
    None, "产品型号", "认证编码", "图片", "品名", "规格", "孔距", "尺寸（mm）", "数量",
    "箱数", "净重\n(kg)", "总净重\n(kg)", "毛重", "总毛重", "体积(cbm)", "总体积(cbm)",
    "RMB", "金额（rmb）", "商标", "备注",
]

JINBAICHUAN_ROWS: list[list] = [
    [_TITLE, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None],
    ["Customer Name 客户名：", None, None, None, None, "SORENTO SDN BHD", None, None,
     None, None, "提单号：", None, None, None, "Date 日期：", None, None, _INVOICE_DATE,
     None, None],
    [None] * 20,
    ["Address 地址：", None, None, None, _ADDRESS, None, None, None, None, None,
     "Container No 货柜号：", None, None, None, None, None, None, None, None, None],
    [None] * 20,
    [None, None, None, None, None, None, None, None, None, None, "封条号：", None,
     None, None, None, None, None, None, None, None],
    [None] * 20,
    list(_HEADER19),
    [1, "SRTWC287A-RL-250", "", "", "连体马桶", "纸箱包装", "250", "730*370*690", 408,
     408, "40", 16320, "50", 20400, "0.17", 69.36, "250", 102000, "SORENTO", None],
    [None, None, None, None, None, None, None, None, None, 408, None, 16320, None,
     20400, None, 69.36, "总金额", 102000, None, None],
    [_TITLE, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None],
    ["Customer Name 客户名：", None, None, None, None, "SORENTO SDN BHD", None, None,
     None, None, "提单号：", None, None, None, "Date 日期：", None, None, _INVOICE_DATE,
     None, None],
    [None] * 20,
    ["Address 地址：", None, None, None, _ADDRESS, None, None, None, None, None,
     "Container No 货柜号：", None, None, None, None, None, None, None, None, None],
    [None] * 20,
    [None, None, None, None, None, None, None, None, None, None, "封条号：", None,
     None, None, None, None, None, None, None, None],
    [None] * 20,
    list(_HEADER19),
    [1, "SRTWC8152-SH-300-UF", "", "", "连体马桶", "纸箱包装", "300", "730*780*380",
     376, 376, "47", 17672, "52", 19552, 0.18, 67.68, "290", 109040, "SORENTO",
     "辉源UF盖板"],
    [None, None, None, None, None, None, None, None, None, 376, None, 17672, None,
     19552, None, 67.68, "总金额", 109040, None, None],
    [None] * 20,
    [_TITLE, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None],
    ["Customer Name 客户名：", None, None, None, None, "SORENTO SDN BHD", None, None,
     None, None, "提单号：", None, None, None, "Date 日期：", None, None, _INVOICE_DATE,
     None, None],
    [None] * 20,
    ["Address 地址：", None, None, None, _ADDRESS, None, None, None, None, None,
     "Container No 货柜号：", None, None, None, None, None, None, None, None, None],
    [None] * 20,
    [None, None, None, None, None, None, None, None, None, None, "封条号：", None,
     None, None, None, None, None, None, None, None],
    [None] * 20,
    list(_HEADER19),
    [1, "CWB242", "", "", "盆", "纸箱包装", "", "450*360*170", 200, 200, "7", 1400,
     "8", 1600, "0.03", 6, "34.5", 6900, "CABANA", None],
    [2, "CGB247", "247", "", "盆", "纸皮包装", "", "520*170*380", 100, 100, "8", 800,
     "9", 900, "0.034", 3.4, "45", 4500, "CABANA", None],
    [3, "SRTWB248", "248", "", "盆", "纸皮包装", "", "420*360*170", 200, 200, "8",
     1600, "9", 1800, "0.03", 6, "34.5", 6900, "SORENTO", None],
    [4, "SRTWB890", "890", "", "盆", "纸箱包装", "", "520*20*420", 250, 250, "14",
     3500, "15", 3750, "0.044", 11, "89.5", 22375, "SORENTO", None],
    [5, "SRTWC7405-RL-250", "", "", "分体马桶", "天地盖", "250", "680*360*400", 198,
     198, "23", 4554, "24", 4752, "0.09", 17.82, "128", 25344, "SORENTO", None],
    [6, "SRTWCY7405认证", "", "", "分体水箱", "天地盖", "", "390*390*180", 198, 198,
     "14", 2772, "15", 2970, "0.03", 5.94, "75", 14850, "SORENTO", None],
    [7, "7405盖板", "", "", "盖板", "纸箱包装", "", "470*470*380", 100, 10, "1.3", 130,
     "1.4", 140, "0.08", 0.8, "26", 2600, "", None],
    [8, "7604盖板", "", "", "盖板", "纸箱包装", "", "370*370*450", 98, 10, "2", 196,
     "2.2", 215.6, "0.08", 0.8, "22", 2156, "", None],
    [9, "CWC604-RL-250", "", "", "分体马桶", "纸箱包装", "250", "680*360*400", 34, 34,
     "25", 850, "26", 884, "0.09", 3.06, "128", 4352, "CABANA", None],
    [10, "CWCY604认证", "", "", "分体水箱", "天地盖", "", "390*390*180", 34, 34, "14",
     476, "15", 510, "0.03", 1.02, "75", 2550, "CABANA", None],
    [11, "604盖板", "", "", "盖板", "纸箱包装", "", "370*370*450", 34, 4, "1.3", 44.2,
     "1.4", 47.6, "0.08", 0.32, "22", 748, "", None],
    [12, "SRTSP124", "", "", "蹲便器", "纸箱包装", "", "550*440*210", 164, 164, "11",
     1804, "12", 1968, "0.06", 9.84, "65", 10660, "SORENTO", None],
    [13, "CWB242海关样品", "", "", "盆", "纸箱包装", "", "450*360*170", 5, 5, "7", 35,
     "8", 40, "0.03", 0.15, "39.5", 197.5, "CABANA", None],
    [14, "CGB247海关样品", "247", "", "盆", "纸皮包装", "", "520*170*380", 5, 5, "8",
     40, "9", 45, "0.02", 0.1, "55", 275, "CABANA", None],
    [15, "SRTWB248海关样品", "248", "", "盆", "纸箱包装", "", "420*360*170", 5, 5, "8",
     40, "9", 45, "0.03", 0.15, "37", 185, "SORENTO", None],
    [16, "SRTWB890海关样品", "890", "", "盆", "纸箱包装", "", "520*420*200", 5, 5,
     "14", 70, "15", 75, "0.044", 0.22, "94.5", 472.5, "SORENTO", None],
    [17, "SRTWC7405海关样品", "7405", "", "分体马桶", "纸箱包装", "250", "680*360*400",
     5, 5, "23", 115, "24", 120, "0.09", 0.45, "143", 715, "SORENTO", None],
    [18, "SRTWCY7405海关样品", "7405", "", "分体水箱", "纸箱包装", "", "390*390*180",
     5, 5, "14", 70, "15", 75, "0.03", 0.15, "80", 400, "SORENTO", None],
    [19, "SRTSP124海关样品", "", "", "蹲便器", "纸箱包装", "", "550*460*210", 10, 10,
     "11", 110, "12", 120, "0.06", 0.6, "70", 700, "SORENTO", None],
    [None, None, None, None, None, None, None, None, None, 1442, None, 18606.2, None,
     20057.2, None, 67.82, "总金额", 106880, None, None],
    [_TITLE, None, None, None, None, None, None, None, None, None, None, None, None,
     None, None, None, None, None, None, None],
    ["Customer Name 客户名：", None, None, None, None, "SORENTO SDN BHD", None, None,
     None, None, "提单号：", None, None, None, "Date 日期：", None, None, _INVOICE_DATE,
     None, None],
    [None] * 20,
    ["Address 地址：", None, None, None, _ADDRESS, None, None, None, None, None,
     "Container No 货柜号：", None, None, None, None, None, None, None, None, None],
    [None] * 20,
    [None, None, None, None, None, None, None, None, None, None, "封条号：", None,
     None, None, None, None, None, None, None, None],
    [None] * 20,
    list(_HEADER19),
    [1, "SRTWC283-SH-250", "", "", "连体马桶", "纸箱包装", "250", "700*370*840", 100,
     100, "50", 5000, "55", 5500, "0.2", 20, "308", 30800, "SORENTO", "辉源UF盖板"],
    [2, "SRTWC286-SH-200-NEW", "", "", "连体马桶", "纸箱包装", "200", "750*370*690",
     50, 50, "45", 2250, "49", 2450, 0.18, 9, "285", 14250, "SORENTO", None],
    [3, "SRTWC8152-RL-200", "", "", "连体马桶", "纸箱包装", "200", "720*360*760", 50,
     50, "47", 2350, "52", 2600, "0.18", 9, "315", 15750, "SORENTO", None],
    [4, "SRTWC8152-SH-300", "", "", "连体马桶", "纸箱包装", "300", "730*780*380", 100,
     100, "47", 4700, "52", 5200, 0.18, 18, "300", 30000, "SORENTO", None],
    [5, "SRTWC8354-SH-250", "", "", "连体马桶", "纸箱包装", 250, "720*360*760", 60, 60,
     "45", 2700, "51", 3060, "0.19", 11.4, "335", 20100, "SORENTO", None],
    [None, None, None, None, None, None, None, None, None, 360, None, 17000, None,
     18810, None, 67.4, "总金额", 110900, None, None],
    ["Customer Name 客户名：", None, None, None, None, "SORENTO SDN BHD", None, None,
     None, None, "提单号：", None, None, None, "Date 日期：", None, None, _INVOICE_DATE,
     None, None],
    [None] * 20,
    ["Address 地址：", None, None, None, _ADDRESS, None, None, None, None, None,
     "Container No 货柜号：", None, None, None, None, None, None, None, None, None],
    [None] * 20,
    [None, None, None, None, None, None, None, None, None, None, "封条号：", None,
     None, None, None, None, None, None, None, None],
    [None] * 20,
    list(_HEADER19),
    [1, "SRTWC8357-RL-180", "", "", "连体马桶", "纸箱包装", "180", "710*460*690", 30,
     30, "50", 1500, "57", 1710, "0.2", 6, "465", 13950, "SORENTO", None],
    [2, "SRTWC8357-RL-200", "", "", "连体马桶", "纸箱包装", "200", "710*460*690", 20,
     20, "50", 1000, "57", 1140, "0.2", 4, "465", 9300, "SORENTO", None],
    [3, "SRTWC8354-SH-250", "", "", "连体马桶", "纸箱包装", "250", "720*360*760", 40,
     40, "45", 1800, "51", 2040, "0.19", 7.6, "335", 13400, "SORENTO", None],
    [4, "SRTWC8354-SH-180", "", "", "连体马桶", "纸箱包装", "180", "720*360*760", 50,
     50, "45", 2250, "51", 2550, "0.19", 9.5, "335", 16750, "SORENTO", None],
    [5, None, "", "", "", "", "", "", None, None, "", "", "", "", "", "", "", "", "",
     None],
    [6, None, "", "", "", "", "", "", None, None, "", "", "", "", "", "", "", "", "",
     None],
    [7, None, "", "", "", "", "", "", None, None, "", "", "", "", "", "", "", "", "",
     None],
    [8, None, "", "", "", "", "", "", None, None, "", "", "", "", "", "", "", "", "",
     None],
    [None, None, None, None, None, None, None, None, None, 140, None, 6550, None,
     7440, None, 27.1, "总金额", 53400, None, None],
]

# --------------------------------------------------------------------------------- #
# Kailu single proforma -- "KAILU形式发票(Sorento)260717.xlsx"
# (bank rows 39-43 in the real file are placeholder text here, not the real values)
# --------------------------------------------------------------------------------- #

KAILU_ROWS: list[list] = [
    ["KAILU SANITARY WARE MANAGEMENT DEPARTMENT", None, None, None, None, None, None,
     None, None],
    ["Units 1260, 12/F, Emperor Group Centre, 288 Hennessy Road, Wan Chai, Hong Kong\n",
     None, None, None, None, None, None, None, None],
    [None] * 9,
    ["PROFORMA INVOICE - 形式发票", None, None, None, None, None, None, None, None],
    [None] * 9,
    ["SORENTO SDN. BHD.", None, None, None, None, "货单号：", "KL20260717", None, None],
    ["NO.5, JALAN ASTANA 2/KU2,", None, None, None, None, "日期：", "17.07.2026", None,
     None],
    ["BANDAR BUKIT RAJA,41050 KLANG,", None, None, None, None, None, None, None, None],
    ["SELANGOR DARUL EHSAN", None, None, None, None, None, None, None, None],
    ["Tel ：03-3393 1278 / 1678", None, None, None, None, None, None, None, None],
    [None] * 9,
    ["序号", "品名", "编号", "产品数量", "单价(元)", "总价（元）", "其他", None, None],
    [None, "BASIN COLD TAP", "SRTWT7443", 490, 65.5, 32095, None, None, None],
    [None, "BASIN COLD TAP", "SRTWT7443", 370, 66, 24420, "202605-S0060", None, None],
    [None, "BASIN COLD TAP", "SRTWT8203", 100, 90.5, 9050, None, None, None],
    [None, "BASIN COLD TAP", "SRTWT8203", 100, 91, 9100, "202605-S0084", None, None],
    [None, "CONCEALED SHOWER COLD TAP", "SRTWT8258\n-GM", 90, 63.5, 5715, None, None,
     None],
    [None, "WALL MOUNTED SHOWER UNION (SQUARE)", "SRTSC16", 150, 32, 4800, None, None,
     None],
    [None, "WALL MOUNTED SHOWER UNION (SQUARE)", "SRTSC16-BL", 300, 32.5, 9750, None,
     None, None],
    [None, "BASIN MIXER", "SRTWT5903", 10, 105, 1050, None, None, None],
    [None, "BASIN MIXER", "SRTWT5903", 210, 111, 23310, "202605-S0060", None, None],
    [None, "BASIN COLD TAP", "SRTWT8214-BL", 100, 67, 6700, None, None, None],
    [None, "BASIN COLD TAP", "SRTWT8214-GM", 100, 71, 7100, None, None, None],
    [None, "BASIN MIXER", "SRTWT8239-GM", 150, 122, 18300, None, None, None],
    [None, "BASIN MIXER", "SRTWT8267-GM", 50, 192, 9600, None, None, None],
    [None, "BASIN COLD TAP", "SRTSC07", 50, 52.5, 2625, None, None, None],
    [None, "BASIN COLD TAP", "SRTWT8264-GM", 200, 90, 18000, None, None, None],
    [None, "BASIN COLD TAP", "SRTWT7445-NEW", 500, 55, 27500, None, None, None],
    [None, "BASIN COLD TAP", "SRTWT7445-LV-WEPLS", 430, 57.5, 24725, None, None, None],
    [None, "BATHTUB FREESTANDING MIXER TAP", "SRTWT51030", 9, 620, 5580, None, None,
     None],
    [None, "PILLAR MOUNTED ELBOW ACTION TAP", "SRTEA803", 10, 55.5, 555, None, None,
     None],
    [None] * 9,
    [None] * 9,
    [None] * 9,
    ["合 计", None, None, None, None, 239975, None, None, None],
    [None] * 9,
    [None, None, None, None, None, None, "KAILU SANITARY WARE (凯露卫浴)", None, None],
    [None, None, None, None, None, None, "2026.07.17", None, None],
    ["Beneficiary Bank: PLACEHOLDER BANK, PLACEHOLDER BRANCH", None, None, None, None,
     None, None, None, None],
    ["Beneficiary Bank Address: PLACEHOLDER ADDRESS", None, None, None, None, None,
     None, None, None],
    ["Beneficiary account number: 0000000000", None, None, None, None, None, None,
     None, None],
    ["Swift Code: PLACEHOLDER", None, None, None, None, None, None, None, None],
    ["CNAPS Code：0000 0000 0000", None, None, None, None, None, None, None, None],
]


def _apply_item_code_map(rows: list[list], item_code_map: Optional[dict]) -> list[list]:
    if not item_code_map:
        return rows
    out = []
    for row in rows:
        out.append([item_code_map.get(v, v) if isinstance(v, str) else v for v in row])
    return out


def _to_bytes(rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def preloading_list_workbook(item_code_map: Optional[dict] = None) -> bytes:
    """The Jinbaichuan multi-block pre-loading list: 5 blocks, 30 priced lines."""
    return _to_bytes(_apply_item_code_map(JINBAICHUAN_ROWS, item_code_map))


def kailu_proforma_workbook(item_code_map: Optional[dict] = None) -> bytes:
    """The Kailu single proforma: 19 lines, 3 with a PO reference."""
    return _to_bytes(_apply_item_code_map(KAILU_ROWS, item_code_map))
