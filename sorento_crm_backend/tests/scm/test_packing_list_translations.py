"""AC-G3 - a shipment line's remarks carry `English (中文)` when they differ, and the
export prints it (R16, purchasing consolidation batch, lane C).

Postgres only, on a blank schema, seeding every FK target itself: CI's database is
empty. The AI provider is ALWAYS stubbed (`_stub_provider`) - `.env` carries a real
`OPENAI_API_KEY` for the app itself, and letting the miss reach the network here would
make this suite flaky and slow for no reason a stub cannot answer just as well.
"""
from __future__ import annotations

import json
import uuid
from io import BytesIO
from types import SimpleNamespace

import pytest

from app.models.ai_assistant import AIAssistantConfig
from app.models.import_alias import ImportFieldAlias
from app.models.procurement import InboundShipment, InboundShipmentLine, Supplier
from app.models.product import Product, ProductCategory, UnitOfMeasure
from app.services import translation_service as tsvc
from app.services.scm import consolidated_packing_list, packing_list_service
from tests._pg_fixture import blank_session

pytestmark = pytest.mark.usefixtures("no_live_llm")

MARKER = "ZZPLT"

# The headers migration 311 seeds for doc type `packing_list` - retyped here (same
# convention `test_packing_list_multi_supplier.py` uses) because CI's database is empty.
_ALIASES = [
    ("item_code", "产品型号"),
    ("product_name", "品名"),
    ("qty", "数量"),
    ("cartons", "箱数"),
    ("cbm_per_unit", "体积(cbm)"),
    ("container_no", "货柜号"),
    ("remark", "备注"),
]


@pytest.fixture
def db():
    with blank_session() as session:
        yield session


def _seed_aliases(db) -> None:
    for field, alias in _ALIASES:
        db.add(
            ImportFieldAlias(
                id=str(uuid.uuid4()), doc_type="packing_list", field=field, alias=alias, locale="zh"
            )
        )
    db.flush()


def _workbook(container: str, rows: list[tuple[str, float, float, float, str]]) -> bytes:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([f"货柜号：{container}"])
    ws.append(["产品型号", "品名", "数量", "箱数", "体积(cbm)", "备注"])
    for code, qty, cartons, cbm, remark in rows:
        ws.append([code, "座厕", qty, cartons, cbm, remark])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _configure_ai(db) -> AIAssistantConfig:
    cfg = AIAssistantConfig(
        id=str(uuid.uuid4()),
        provider="openai",
        model="gpt-4o-mini",
        temperature=0,
        system_prompt="",
        api_key_ciphertext="fake-key",
        enabled_tools=[],
        rag_enabled=True,
        is_enabled=True,
    )
    db.add(cfg)
    db.flush()
    return cfg


class _FakeProvider:
    def __init__(self, answers: dict[str, str]):
        self.answers = answers
        self.calls: list[list[dict]] = []

    def chat(self, messages, **_kwargs):
        self.calls.append(messages)
        user_content = messages[-1]["content"]
        lines = [
            ln.split(". ", 1)[1] if ". " in ln else ln
            for ln in user_content.rsplit("\n\n", 1)[-1].splitlines()
            if ln.strip()
        ]
        translations = [
            {"source": ln, "target": self.answers.get(ln, f"[{ln}]")} for ln in lines
        ]
        return SimpleNamespace(content=json.dumps({"translations": translations}))


def _stub_provider(monkeypatch, answers: dict[str, str]) -> _FakeProvider:
    fake = _FakeProvider(answers)
    monkeypatch.setattr(tsvc, "get_provider", lambda *a, **kw: fake)
    return fake


def _seed_world(db, tag: str):
    cat = ProductCategory(
        id=str(uuid.uuid4()), category_code=f"{MARKER}-CAT-{tag}", category_name="cat"
    )
    uom = UnitOfMeasure(id=str(uuid.uuid4()), uom_code=f"{MARKER}{tag}"[:20], uom_name="pcs")
    db.add_all([cat, uom])
    db.flush()
    supplier = Supplier(
        id=str(uuid.uuid4()), supplier_code=f"{MARKER}-S-{tag}",
        supplier_name=f"{MARKER} Factory", is_active=True,
    )
    db.add(supplier)
    product = Product(
        id=str(uuid.uuid4()), product_code=f"{MARKER}-{tag}", product_name="Toilet bowl",
        category_id=cat.id, base_uom_id=uom.id, list_price=0,
        is_active=True, is_discontinued=False,
    )
    db.add(product)
    db.flush()
    return supplier, product


def test_a_translated_remark_carries_english_and_chinese_and_the_export_prints_it(
    db, monkeypatch
):
    tag = uuid.uuid4().hex[:8].upper()
    supplier, product = _seed_world(db, tag)
    _seed_aliases(db)
    db.commit()

    _configure_ai(db)
    _stub_provider(monkeypatch, {"纸箱：2个": "Carton: 2 pieces"})

    container = f"{MARKER}U{tag}"
    out = packing_list_service.apply(
        db,
        _workbook(container, [(product.product_code, 10, 2, 0.21, "纸箱：2个")]),
        supplier_id=str(supplier.id),
    )
    db.commit()

    shipment_id = out["results"][0]["shipment_id"]
    line = (
        db.query(InboundShipmentLine)
        .filter(InboundShipmentLine.shipment_id == shipment_id)
        .one()
    )
    assert line.remarks == "Carton: 2 pieces (纸箱：2个)"

    payload = consolidated_packing_list.build(db, str(shipment_id))
    xlsx_bytes = consolidated_packing_list.to_xlsx(payload)

    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active
    # Column 19 is REMARKS, row 18 the first data line - `consolidated_packing_list.
    # _HEADER_ROW`/`_FIRST_LINE_ROW`. The export logic itself is untouched by this
    # lane; what is new is that `remarks` already carries the bilingual text by the
    # time this reads it.
    assert ws.cell(row=18, column=19).value == "Carton: 2 pieces (纸箱：2个)"


def test_an_english_remark_is_never_sent_to_the_model_and_stays_as_is(db, monkeypatch):
    tag = uuid.uuid4().hex[:8].upper()
    supplier, product = _seed_world(db, tag)
    _seed_aliases(db)
    db.commit()

    _configure_ai(db)
    fake = _stub_provider(monkeypatch, {})

    container = f"{MARKER}U{tag}"
    out = packing_list_service.apply(
        db,
        _workbook(container, [(product.product_code, 10, 2, 0.21, "loaded first")]),
        supplier_id=str(supplier.id),
    )
    db.commit()

    line = (
        db.query(InboundShipmentLine)
        .filter(InboundShipmentLine.shipment_id == out["results"][0]["shipment_id"])
        .one()
    )
    assert line.remarks == "loaded first"
    assert fake.calls == []


def test_a_line_with_no_remark_writes_no_remark(db, monkeypatch):
    tag = uuid.uuid4().hex[:8].upper()
    supplier, product = _seed_world(db, tag)
    _seed_aliases(db)
    db.commit()

    _configure_ai(db)
    fake = _stub_provider(monkeypatch, {})

    container = f"{MARKER}U{tag}"
    out = packing_list_service.apply(
        db,
        _workbook(container, [(product.product_code, 10, 2, 0.21, "")]),
        supplier_id=str(supplier.id),
    )
    db.commit()

    line = (
        db.query(InboundShipmentLine)
        .filter(InboundShipmentLine.shipment_id == out["results"][0]["shipment_id"])
        .one()
    )
    assert line.remarks is None
    assert fake.calls == []
