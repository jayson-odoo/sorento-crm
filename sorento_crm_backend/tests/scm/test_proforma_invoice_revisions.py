"""F5b - a supplier resending a container with new prices is a REVISION, not a second PI.

TEST-FIRST: `revision_candidate`, `apply(..., revision_of=...)`, `mark_as_revision_of` and
the diff on `serialize` do not exist at the time this file is written, so every test here is
expected to be red until they land.

Postgres via `pg_session` (rolled back at teardown) like the rest of this channel's suites,
because the reader resolves its header aliases from the alias table. Every row is seeded
under the marker `World` uses; nothing is borrowed from an existing table.

The shape under test is the real one: Jinbaichuan send the same pre-loading list again with
the same item codes and a different price, and the question the screen has to answer is
"what changed", not "which of these two documents is true".
"""
from __future__ import annotations

import uuid
from datetime import date

import pytest

from app.models.scm import ProformaInvoice
from app.services.error_handler import AppException
from app.services.scm import proforma_invoice_service as svc
from tests._pg_fixture import pg_session
from tests.scm.fixtures.proforma_shapes import kailu_proforma_workbook
from tests.scm.test_proforma_invoice_import import World, _invoices, _lines


#: Where 单价 sits on the Kailu shape (序号 / 品名 / 编号 / 产品数量 / 单价(元) / 总价（元）
#: / 其他), and the cell their 货单号 is written in.
_PRICE_COLUMN = 5
_PI_NUMBER_CELL = "G6"


def _kailu(w: World, *, price_factor: float = 1.0, pi_number=None) -> bytes:
    """The Kailu proforma, optionally re-priced and renumbered - the supplier's second send."""
    import openpyxl
    from io import BytesIO

    data = kailu_proforma_workbook({"SRTWT7443": w.code("A")})
    wb = openpyxl.load_workbook(BytesIO(data))
    ws = wb.active
    if price_factor != 1.0:
        # Only numeric cells are re-priced, so the letterhead and the header row are left
        # exactly as they are.
        for row in ws.iter_rows(min_col=_PRICE_COLUMN, max_col=_PRICE_COLUMN):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.value = round(cell.value * price_factor, 4)
    if pi_number is not None:
        ws[_PI_NUMBER_CELL] = pi_number
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _apply(db, w: World, data: bytes, **kwargs) -> dict:
    return svc.apply(db, data, supplier_id=str(w.supplier.id), actor="Ms Tee", **kwargs)


# --------------------------------------------------------------------------------- #
# AC-E6 - the dialog is OFFERED a revision, never forced into one
# --------------------------------------------------------------------------------- #


def test_a_resend_of_the_same_codes_offers_the_earlier_invoice_as_a_revision():
    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w))
        first = _invoices(db, w)[0]

        preview = svc.preview(
            db, _kailu(w, price_factor=1.1), supplier_id=str(w.supplier.id),
            source_ref="kailu-2.xlsx",
        )

        candidate = preview["documents"][0]["revision_candidate"]
        assert candidate["invoice_id"] == str(first.id)
        assert candidate["pi_number"] == first.pi_number
        assert candidate["overlap_pct"] == pytest.approx(100, abs=0.01)


def test_a_file_sharing_nothing_with_what_is_on_file_offers_no_revision():
    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w))

        # A different supplier's document entirely: same shape, codes we have never seen.
        other = World(db)
        preview = svc.preview(
            db, kailu_proforma_workbook({"SRTWT7443": other.code("Z")}),
            supplier_id=str(other.supplier.id), source_ref="other.xlsx",
        )

        assert preview["documents"][0]["revision_candidate"] is None


def test_an_already_converted_invoice_is_never_offered_as_a_revision_target():
    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w))
        svc.convert_to_draft_shipment(db, [str(_invoices(db, w)[0].id)])

        preview = svc.preview(
            db, _kailu(w, price_factor=1.1), supplier_id=str(w.supplier.id),
            source_ref="kailu-2.xlsx",
        )

        assert preview["documents"][0]["revision_candidate"] is None


# --------------------------------------------------------------------------------- #
# AC-E7 - a revision supersedes its predecessor, and keeps it
# --------------------------------------------------------------------------------- #


def test_uploading_as_a_revision_supersedes_the_prior_and_keeps_it_on_file():
    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w))
        first = _invoices(db, w)[0]

        out = _apply(
            db, w, _kailu(w, price_factor=1.1),
            source_ref="kailu-2.xlsx", revision_of={"1": str(first.id)},
        )

        db.refresh(first)
        assert first.status == "superseded"
        assert first.revision_no == 1
        new_id = out["results"][0]["invoice_id"]
        new = db.query(ProformaInvoice).filter(ProformaInvoice.id == new_id).one()
        assert new.status == "current"
        assert new.revision_no == 2
        assert str(new.revision_of_id) == str(first.id)
        # The predecessor is KEPT: it is what the supplier sent on the day.
        assert len(_invoices(db, w)) == 2


def test_a_revision_uploaded_from_the_same_file_gets_its_own_document_number():
    """Identity is (supplier, pi_number), and a re-upload of the same FILE derives the same
    positional number - so a revision taken from it has to be named apart or it would
    collide with the document it revises."""
    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w))
        first = _invoices(db, w)[0]

        out = _apply(db, w, _kailu(w, price_factor=1.1), revision_of={"1": str(first.id)})

        assert out["results"][0]["pi_number"] != first.pi_number
        assert len(_invoices(db, w)) == 2


def test_a_superseded_revision_cannot_be_revised_again():
    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w))
        first = _invoices(db, w)[0]
        _apply(db, w, _kailu(w, price_factor=1.1), revision_of={"1": str(first.id)})

        with pytest.raises(AppException) as exc:
            _apply(db, w, _kailu(w, price_factor=1.2), revision_of={"1": str(first.id)})
        assert exc.value.status_code == 409


def test_a_revision_of_another_suppliers_invoice_is_refused():
    with pg_session() as db:
        w = World(db)
        other = World(db)
        _apply(db, w, _kailu(w))
        theirs = _invoices(db, w)[0]

        with pytest.raises(AppException) as exc:
            svc.apply(
                db, kailu_proforma_workbook({"SRTWT7443": other.code("A")}),
                supplier_id=str(other.supplier.id), revision_of={"1": str(theirs.id)},
            )
        assert exc.value.status_code == 422


# --------------------------------------------------------------------------------- #
# AC-E8 - the detail says what changed
# --------------------------------------------------------------------------------- #


def test_the_detail_reads_revision_2_of_2_and_names_its_predecessor():
    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w))
        first = _invoices(db, w)[0]
        out = _apply(db, w, _kailu(w, price_factor=1.1), revision_of={"1": str(first.id)})

        detail = svc.serialize(db, svc.get_or_404(db, out["results"][0]["invoice_id"]))

        assert detail["revision_no"] == 2
        assert detail["revision_count"] == 2
        assert detail["revision_of_pi_number"] == first.pi_number
        assert [r["revision_no"] for r in detail["revisions"]] == [1, 2]
        # Read from the OLDER one too: the chain is the same chain from either end.
        older = svc.serialize(db, svc.get_or_404(db, str(first.id)))
        assert older["revision_count"] == 2
        assert older["status"] == "superseded"


def test_the_diff_names_every_line_whose_price_moved():
    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w))
        first = _invoices(db, w)[0]
        priced = [ln for ln in _lines(db, first.id) if ln.unit_price is not None]
        out = _apply(db, w, _kailu(w, price_factor=1.1), revision_of={"1": str(first.id)})

        detail = svc.serialize(db, svc.get_or_404(db, out["results"][0]["invoice_id"]))

        diff = detail["diff"]
        assert diff["price_changed_lines"] == len(priced)
        assert diff["compared_to_pi_number"] == first.pi_number
        # Kailu name the same model on two lines at two prices, so a change is identified
        # by the code AND which occurrence of it this is.
        sample = priced[0]
        entry = next(
            c for c in diff["changes"]
            if c["item_code"] == sample.item_code and c["occurrence"] == 1
        )
        assert entry["unit_price_changed"] is True
        assert entry["unit_price_was"] == pytest.approx(float(sample.unit_price))
        assert entry["unit_price_now"] == pytest.approx(float(sample.unit_price) * 1.1, rel=1e-3)
        assert entry["qty_changed"] is False


def test_the_first_revision_has_nothing_to_compare_itself_with():
    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w))

        detail = svc.serialize(db, _invoices(db, w)[0])

        assert detail["diff"] is None
        assert detail["revision_count"] == 1


# --------------------------------------------------------------------------------- #
# AC-E9 / AC-E10 - a superseded price is never a cost, and never converts
# --------------------------------------------------------------------------------- #


def test_a_superseded_revision_cannot_be_converted():
    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w))
        first = _invoices(db, w)[0]
        _apply(db, w, _kailu(w, price_factor=1.1), revision_of={"1": str(first.id)})

        with pytest.raises(AppException) as exc:
            svc.convert_to_draft_shipment(db, [str(first.id)])
        assert exc.value.status_code == 409
        assert exc.value.detail["code"] == "superseded"


def test_the_shipment_takes_the_current_revisions_price_not_the_superseded_one():
    from app.models.procurement import InboundShipmentLine

    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w))
        first = _invoices(db, w)[0]
        old_price = float(
            next(ln for ln in _lines(db, first.id) if ln.product_id).unit_price
        )
        out = _apply(db, w, _kailu(w, price_factor=1.1), revision_of={"1": str(first.id)})

        shipment = svc.convert_to_draft_shipment(db, [out["results"][0]["invoice_id"]])

        costs = [
            float(ln.unit_cost)
            for ln in db.query(InboundShipmentLine)
            .filter(InboundShipmentLine.shipment_id == shipment["shipment_id"])
            .all()
            if ln.unit_cost is not None
        ]
        assert costs
        assert all(abs(c - old_price) > 0.001 for c in costs)


def test_a_superseded_revision_cannot_be_adjusted():
    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w))
        first = _invoices(db, w)[0]
        _apply(db, w, _kailu(w, price_factor=1.1), revision_of={"1": str(first.id)})
        line = _lines(db, first.id)[0]

        with pytest.raises(AppException) as exc:
            svc.adjust_line(db, str(first.id), str(line.id), qty=1, actor="Ms Tee")
        assert exc.value.status_code == 409


# --------------------------------------------------------------------------------- #
# AC-E11 - a wrongly-created new PI can be linked to its predecessor afterwards
# --------------------------------------------------------------------------------- #


def test_marking_a_new_invoice_as_a_revision_of_its_predecessor():
    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w))
        first = _invoices(db, w)[0]
        # Uploaded as a NEW PI by mistake. Kailu STATE their document number, so a second
        # send under the same one would have updated the first invoice in place - the
        # mistake this action exists to undo is a genuinely new document nobody linked.
        out = _apply(
            db, w, _kailu(w, price_factor=1.1, pi_number="KL20260801"),
            source_ref="kailu-second.xlsx",
        )
        second_id = out["results"][0]["invoice_id"]

        detail = svc.mark_as_revision_of(db, second_id, str(first.id))

        db.refresh(first)
        assert first.status == "superseded"
        assert detail["revision_no"] == 2
        assert detail["revision_count"] == 2
        assert detail["diff"]["compared_to_pi_number"] == first.pi_number


def test_an_invoice_cannot_be_marked_as_a_revision_of_itself():
    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w))
        first = _invoices(db, w)[0]

        with pytest.raises(AppException) as exc:
            svc.mark_as_revision_of(db, str(first.id), str(first.id))
        assert exc.value.status_code == 422


# --------------------------------------------------------------------------------- #
# Review finding 4 - a chain cannot be made to eat its own tail
# --------------------------------------------------------------------------------- #


def test_marking_the_predecessor_as_a_revision_of_its_own_successor_is_refused():
    """A -> B, then B -> A closes the loop: the superseded original comes back as current,
    both documents answer "what is this container costing", and the chain walk relies on a
    guard to terminate rather than on the data being sane."""
    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w))
        first = _invoices(db, w)[0]
        out = _apply(db, w, _kailu(w, price_factor=1.1), revision_of={"1": str(first.id)})
        second_id = out["results"][0]["invoice_id"]

        with pytest.raises(AppException) as exc:
            svc.mark_as_revision_of(db, str(first.id), second_id)

        assert exc.value.status_code == 409
        assert exc.value.detail["code"] == "revision_cycle"
        db.refresh(first)
        assert first.status == "superseded"


def test_a_longer_loop_is_refused_too():
    """A -> B -> C, then C -> A. The cycle is two hops away, so a check that only looked at
    the immediate predecessor would let it through."""
    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w))
        first = _invoices(db, w)[0]
        second_id = _apply(
            db, w, _kailu(w, price_factor=1.1), revision_of={"1": str(first.id)}
        )["results"][0]["invoice_id"]
        third_id = _apply(
            db, w, _kailu(w, price_factor=1.2), revision_of={"1": second_id}
        )["results"][0]["invoice_id"]

        with pytest.raises(AppException) as exc:
            svc.mark_as_revision_of(db, str(first.id), third_id)

        assert exc.value.status_code == 409
        assert exc.value.detail["code"] == "revision_cycle"


# --------------------------------------------------------------------------------- #
# Browser pass 2, finding 1 - a SKIP row must not hide a revision candidate
# --------------------------------------------------------------------------------- #


def test_an_invoice_with_only_a_skip_row_is_still_offered_as_a_revision_target():
    """The dev JINBAICHUAN invoices carry a SKIP link - a line no product matched, recorded
    when a convert of ANOTHER invoice ran. The candidate query excluded any invoice with a
    link row at all, so the upload dialog offered no revision for exactly the documents the
    tester was re-uploading, and every second send landed as an independent PI."""
    from app.models.scm import ProformaInvoiceShipmentLink
    from app.models.procurement import InboundShipment

    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w))
        first = _invoices(db, w)[0]
        shipment = InboundShipment(
            id=str(uuid.uuid4()),
            shipment_number=f"ZZREV-{uuid.uuid4().hex[:8]}",
            shipment_date=date(2026, 8, 1),
            shipment_status="draft",
        )
        db.add(shipment)
        db.flush()
        db.add(ProformaInvoiceShipmentLink(
            id=str(uuid.uuid4()),
            proforma_invoice_id=first.id,
            proforma_invoice_line_id=_lines(db, first.id)[0].id,
            inbound_shipment_id=shipment.id,
            inbound_shipment_line_id=None,
            unmatched_reason="No catalogue product matches this line's item code.",
        ))
        db.flush()

        preview = svc.preview(
            db, _kailu(w, price_factor=1.1), supplier_id=str(w.supplier.id),
            source_ref="kailu-2.xlsx",
        )

        candidate = preview["documents"][0]["revision_candidate"]
        assert candidate is not None
        assert candidate["invoice_id"] == str(first.id)


def test_an_invoice_actually_on_a_container_is_still_not_offered():
    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w))
        svc.convert_to_draft_shipment(db, [str(_invoices(db, w)[0].id)])

        preview = svc.preview(
            db, _kailu(w, price_factor=1.1), supplier_id=str(w.supplier.id),
            source_ref="kailu-2.xlsx",
        )

        assert preview["documents"][0]["revision_candidate"] is None


# --------------------------------------------------------------------------------- #
# Browser pass 3, finding 1 - an explicit untick files a NEW invoice
# --------------------------------------------------------------------------------- #


def test_unticking_the_revision_offer_creates_a_second_invoice_from_the_same_file():
    """The identity a file derives is (supplier, pi_number), and the same file derives the
    same number - so an untick fell straight into the idempotent in-place replace and the
    upload reported "updated in place" with no new row. An explicit untick is an
    instruction: file this as a NEW document, and give it the next free number."""
    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w), source_ref="same.xlsx")
        first = _invoices(db, w)[0]

        out = _apply(db, w, _kailu(w), source_ref="same.xlsx", file_as_new=["1"])

        assert out["documents_created"] == 1
        assert out["documents_updated"] == 0
        created = out["results"][0]
        assert created["invoice_id"] != str(first.id)
        assert created["pi_number"] != first.pi_number
        assert created["pi_number"].startswith(first.pi_number)
        assert len(_invoices(db, w)) == 2


def test_the_new_invoice_is_not_a_revision_of_anything():
    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w), source_ref="same.xlsx")
        first = _invoices(db, w)[0]

        out = _apply(db, w, _kailu(w), source_ref="same.xlsx", file_as_new=["1"])

        new_id = out["results"][0]["invoice_id"]
        new = db.query(ProformaInvoice).filter(ProformaInvoice.id == new_id).one()
        assert new.revision_of_id is None
        assert new.revision_no == 1
        db.refresh(first)
        assert first.status == "current"


def test_without_the_untick_the_same_file_still_lands_in_place():
    """AC-P1.4 - the idempotency a nervous second Confirm relies on is not narrowed."""
    with pg_session() as db:
        w = World(db)
        _apply(db, w, _kailu(w), source_ref="same.xlsx")

        out = _apply(db, w, _kailu(w), source_ref="same.xlsx")

        assert out["documents_created"] == 0
        assert out["documents_updated"] == 1
        assert len(_invoices(db, w)) == 1
