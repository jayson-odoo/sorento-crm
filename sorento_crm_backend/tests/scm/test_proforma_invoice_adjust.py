"""F5 - the proforma invoice carries volume, and Sorento adjusts it to fit the container.

TEST-FIRST: the cbm columns, `adjust_line`, `remove_line`, `set_container_size`, the export
and the over-capacity guard do not exist at the time this file is written, so every test
here is expected to be red (AttributeError / a missing key) until they land.

Runs on the REAL Postgres database via `pg_session` (rolled back at teardown) like
`test_proforma_invoice_import.py`, because the reader resolves its header aliases from the
alias table - so this suite also proves migration 428's seed was actually applied rather
than merely written. Every row is seeded under the `ZZPIV` marker; nothing is borrowed.

The numbers are the real pre-loading list's (`2026-7-31 SORENTO 预装清单.xlsx`, reproduced
cell-for-cell in `fixtures/proforma_shapes.py`): five blocks at 69.36 / 67.68 / 67.82 /
67.4 / 27.1 cbm, four of them over a 65 cbm 40HQ (AC-D3).
"""
from __future__ import annotations

import uuid

import pytest

from app.models.scm import ContainerSize, ProformaInvoice, ProformaInvoiceLine
from app.services.error_handler import AppException
from app.services.scm import proforma_invoice_service as svc
from tests._pg_fixture import pg_session
from tests.scm.fixtures.proforma_shapes import (
    kailu_proforma_workbook,
    preloading_list_workbook,
)
from tests.scm.test_proforma_invoice_import import World, _invoices, _lines

MARKER = "ZZPIV"

#: What each of the five blocks of the real pre-loading list totals to (AC-D3).
BLOCK_CBM = [69.36, 67.68, 67.82, 67.4, 27.1]


def _seed_container_sizes(db) -> None:
    """Migration 336's own seed helper, called rather than assumed: CI's database is built
    with `create_all` and never runs a migration body, so no container size exists there."""
    import importlib.util
    from pathlib import Path

    versions = Path(__file__).resolve().parents[2] / "alembic" / "versions"
    spec = importlib.util.spec_from_file_location(
        "_m336_sizes", versions / "336_scm_supplier_inventory_loading_plan.py"
    )
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    m.seed_container_sizes(db.connection())
    db.flush()


def _apply_preloading(db, w: World) -> list[ProformaInvoice]:
    # One code per block we actually convert, so a convert test is refused for CAPACITY
    # rather than for having no catalogue match at all.
    data = preloading_list_workbook(
        {
            "SRTWC287A-RL-250": w.code("A"),
            "CWB242": w.code("B"),
            "SRTWC8357-RL-180": w.code("C"),
        }
    )
    svc.apply(db, data, supplier_id=str(w.supplier.id), actor="Ms Tee")
    return _invoices(db, w)


# --------------------------------------------------------------------------------- #
# AC-D1 / AC-D3 - the volume the document states is stored, per line
# --------------------------------------------------------------------------------- #


def test_the_preloading_list_stores_cartons_and_both_volumes_per_line():
    with pg_session() as db:
        w = World(db)
        invoices = _apply_preloading(db, w)

        first = _lines(db, invoices[0].id)[0]
        assert float(first.cartons) == 408
        assert float(first.cbm_per_unit) == pytest.approx(0.17)
        assert float(first.cbm_total) == pytest.approx(69.36)


def test_each_of_the_five_blocks_reads_its_own_container_volume():
    with pg_session() as db:
        w = World(db)
        invoices = _apply_preloading(db, w)

        totals = [
            sum(float(ln.cbm_total or 0) for ln in _lines(db, inv.id)) for inv in invoices
        ]
        assert totals == pytest.approx(BLOCK_CBM, abs=0.01)


def test_a_document_stating_no_volume_stores_null_never_zero():
    """Kailu's proforma has no volume column at all. NULL and 0 are different answers to
    "will this fit", and only one of them is honest (AC-D1)."""
    with pg_session() as db:
        w = World(db)
        svc.apply(db, kailu_proforma_workbook({"SRTWT7443": w.code("A")}),
                  supplier_id=str(w.supplier.id))
        invoice = _invoices(db, w)[0]

        for line in _lines(db, invoice.id):
            assert line.cbm_per_unit is None
            assert line.cbm_total is None
            assert line.cartons is None


def test_the_supplier_figures_are_frozen_at_import():
    with pg_session() as db:
        w = World(db)
        invoices = _apply_preloading(db, w)
        line = _lines(db, invoices[0].id)[0]

        assert float(line.supplier_qty) == float(line.qty) == 408
        assert float(line.supplier_unit_price) == float(line.unit_price) == 250


# --------------------------------------------------------------------------------- #
# AC-D2 / AC-D4 - the header states the fill against a named container
# --------------------------------------------------------------------------------- #


def test_the_header_states_the_volume_against_a_65_cbm_40hq():
    with pg_session() as db:
        _seed_container_sizes(db)
        w = World(db)
        invoices = _apply_preloading(db, w)

        out = svc.serialize(db, invoices[0])

        assert out["total_cbm"] == pytest.approx(69.36)
        assert out["container_cbm"] == pytest.approx(65)
        assert out["container_size_code"] == "40HQ"
        assert out["fill_pct"] == pytest.approx(106.7, abs=0.1)
        assert out["over_by_cbm"] == pytest.approx(4.36, abs=0.01)
        assert out["unmeasured_lines"] == 0


def test_an_unmeasured_document_counts_its_lines_rather_than_reading_zero():
    with pg_session() as db:
        _seed_container_sizes(db)
        w = World(db)
        svc.apply(db, kailu_proforma_workbook({"SRTWT7443": w.code("A")}),
                  supplier_id=str(w.supplier.id))

        out = svc.serialize(db, _invoices(db, w)[0])

        assert out["total_cbm"] is None
        assert out["unmeasured_lines"] == out["line_count"] > 0
        assert out["fill_pct"] is None


def test_the_container_size_is_changeable_on_the_invoice():
    with pg_session() as db:
        _seed_container_sizes(db)
        w = World(db)
        invoice = _apply_preloading(db, w)[0]
        small = (
            db.query(ContainerSize).filter(ContainerSize.code == "20GP").one()
        )

        out = svc.set_container_size(db, str(invoice.id), str(small.id))

        assert out["container_size_code"] == "20GP"
        assert out["container_cbm"] == pytest.approx(28)
        # And back to the tenant default, which is what a cleared select means.
        out = svc.set_container_size(db, str(invoice.id), None)
        assert out["container_size_code"] == "40HQ"


# --------------------------------------------------------------------------------- #
# AC-E1 / AC-E2 / AC-E3 - Sorento adjusts, the supplier's figure stays
# --------------------------------------------------------------------------------- #


def test_adjusting_a_line_keeps_the_suppliers_own_quantity():
    with pg_session() as db:
        _seed_container_sizes(db)
        w = World(db)
        invoice = _apply_preloading(db, w)[0]
        line = _lines(db, invoice.id)[0]

        out = svc.adjust_line(db, str(invoice.id), str(line.id), qty=380, actor="Ms Tee")

        db.refresh(line)
        assert float(line.qty) == 380
        assert float(line.supplier_qty) == 408
        assert out["lines"][0]["qty"] == 380
        assert out["lines"][0]["supplier_qty"] == 408


def test_adjusting_a_line_recomputes_its_volume_and_its_money():
    with pg_session() as db:
        _seed_container_sizes(db)
        w = World(db)
        invoice = _apply_preloading(db, w)[0]
        line = _lines(db, invoice.id)[0]

        out = svc.adjust_line(db, str(invoice.id), str(line.id), qty=380, actor="Ms Tee")

        db.refresh(line)
        assert float(line.cbm_total) == pytest.approx(0.17 * 380)
        assert float(line.amount) == pytest.approx(250 * 380)
        assert out["total_cbm"] == pytest.approx(0.17 * 380)
        assert out["over_by_cbm"] is None


def test_adjusting_stamps_who_and_when():
    with pg_session() as db:
        _seed_container_sizes(db)
        w = World(db)
        invoice = _apply_preloading(db, w)[0]
        line = _lines(db, invoice.id)[0]

        out = svc.adjust_line(db, str(invoice.id), str(line.id), qty=400, actor="Ms Tee")

        assert out["adjusted_by"] == "Ms Tee"
        assert out["adjusted_at"] is not None
        assert out["is_adjusted"] is True


def test_a_negative_quantity_is_refused():
    with pg_session() as db:
        w = World(db)
        invoice = _apply_preloading(db, w)[0]
        line = _lines(db, invoice.id)[0]

        with pytest.raises(AppException) as exc:
            svc.adjust_line(db, str(invoice.id), str(line.id), qty=-1, actor="Ms Tee")
        assert exc.value.status_code == 422


def test_a_line_of_another_invoice_is_a_404_not_a_silent_write():
    with pg_session() as db:
        w = World(db)
        invoices = _apply_preloading(db, w)
        stranger = _lines(db, invoices[1].id)[0]

        with pytest.raises(AppException) as exc:
            svc.adjust_line(db, str(invoices[0].id), str(stranger.id), qty=1, actor="x")
        assert exc.value.status_code == 404


def test_removing_a_line_takes_it_off_the_invoice_and_off_the_volume():
    with pg_session() as db:
        _seed_container_sizes(db)
        w = World(db)
        invoice = _apply_preloading(db, w)[2]
        before = _lines(db, invoice.id)
        victim = before[0]
        victim_cbm = float(victim.cbm_total)

        out = svc.remove_line(db, str(invoice.id), str(victim.id), actor="Ms Tee")

        assert out["line_count"] == len(before) - 1
        assert len(out["lines"]) == len(before) - 1
        assert out["total_cbm"] == pytest.approx(BLOCK_CBM[2] - victim_cbm, abs=0.01)
        assert out["is_adjusted"] is True
        assert (
            db.query(ProformaInvoiceLine)
            .filter(ProformaInvoiceLine.id == victim.id)
            .count()
            == 0
        )


# --------------------------------------------------------------------------------- #
# AC-E4 - the adjusted invoice goes back to the supplier as a workbook
# --------------------------------------------------------------------------------- #


def test_the_export_carries_the_adjusted_quantity_and_names_the_suppliers_own():
    import openpyxl
    from io import BytesIO

    with pg_session() as db:
        _seed_container_sizes(db)
        w = World(db)
        invoice = _apply_preloading(db, w)[0]
        line = _lines(db, invoice.id)[0]
        svc.adjust_line(db, str(invoice.id), str(line.id), qty=380, actor="Ms Tee")

        payload = svc.serialize(db, svc.get_or_404(db, str(invoice.id)))
        data = svc.to_xlsx(payload)

        ws = openpyxl.load_workbook(BytesIO(data)).active
        cells = [
            [c.value for c in row]
            for row in ws.iter_rows()
        ]
        flat = [str(v) for row in cells for v in row if v is not None]
        assert "380" in flat or 380 in [v for row in cells for v in row]
        # The supplier's own figure travels in the remark column, where their sheet keeps it.
        assert any("408" in s for s in flat)
        # Their header spelling, so the sheet is recognisable as their own document.
        assert any(s == "总体积(cbm)" for s in flat)
        assert svc.export_filename(payload).endswith(".xlsx")


def test_the_export_recomputes_the_amount_from_the_adjusted_quantity():
    import openpyxl
    from io import BytesIO

    with pg_session() as db:
        _seed_container_sizes(db)
        w = World(db)
        invoice = _apply_preloading(db, w)[0]
        line = _lines(db, invoice.id)[0]
        svc.adjust_line(db, str(invoice.id), str(line.id), qty=380, actor="Ms Tee")

        payload = svc.serialize(db, svc.get_or_404(db, str(invoice.id)))
        ws = openpyxl.load_workbook(BytesIO(svc.to_xlsx(payload))).active
        numbers = [
            c.value for row in ws.iter_rows() for c in row if isinstance(c.value, (int, float))
        ]
        assert 250 * 380 in numbers
        assert pytest.approx(0.17 * 380, abs=0.001) in numbers


# --------------------------------------------------------------------------------- #
# AC-E5 - an over-capacity conversion is refused, and overridable with a reason
# --------------------------------------------------------------------------------- #


def test_converting_an_over_capacity_invoice_is_refused_with_the_figures():
    with pg_session() as db:
        _seed_container_sizes(db)
        w = World(db)
        invoice = _apply_preloading(db, w)[0]

        with pytest.raises(AppException) as exc:
            svc.convert_to_draft_shipment(db, [str(invoice.id)])

        assert exc.value.status_code == 409
        body = exc.value.detail
        assert body["code"] == "over_capacity"
        assert "69.36" in body["message"] and "65" in body["message"]


def test_convert_anyway_with_a_reason_creates_the_shipment_and_records_why():
    from app.models.procurement import InboundShipment

    with pg_session() as db:
        _seed_container_sizes(db)
        w = World(db)
        invoice = _apply_preloading(db, w)[0]

        out = svc.convert_to_draft_shipment(
            db,
            [str(invoice.id)],
            override_capacity=True,
            override_reason="Second container booked",
        )

        shipment = (
            db.query(InboundShipment).filter(InboundShipment.id == out["shipment_id"]).one()
        )
        assert "Second container booked" in (shipment.notes or "")
        assert "over" in (shipment.notes or "").lower()


def test_convert_anyway_still_needs_a_reason():
    with pg_session() as db:
        _seed_container_sizes(db)
        w = World(db)
        invoice = _apply_preloading(db, w)[0]

        with pytest.raises(AppException) as exc:
            svc.convert_to_draft_shipment(db, [str(invoice.id)], override_capacity=True)
        assert exc.value.status_code == 422


def test_an_invoice_that_fits_converts_without_any_override():
    with pg_session() as db:
        _seed_container_sizes(db)
        w = World(db)
        # Block 5 is 27.1 cbm - comfortably inside a 65 cbm 40HQ.
        invoice = _apply_preloading(db, w)[4]

        out = svc.convert_to_draft_shipment(db, [str(invoice.id)])

        assert out["shipment_id"]


def test_an_unmeasured_invoice_converts_rather_than_being_refused_for_no_volume():
    """AC-H3: a PI without cbm converts as before. An unknown volume is not an over-capacity
    one, and refusing it would break the Kailu shape that has worked since G3b."""
    with pg_session() as db:
        _seed_container_sizes(db)
        w = World(db)
        svc.apply(db, kailu_proforma_workbook({"SRTWT7443": w.code("A")}),
                  supplier_id=str(w.supplier.id))
        invoice = _invoices(db, w)[0]

        out = svc.convert_to_draft_shipment(db, [str(invoice.id)])

        assert out["shipment_id"]


# --------------------------------------------------------------------------------- #
# The adjustment is closed once the goods are drafted onto a shipment
# --------------------------------------------------------------------------------- #


def test_a_converted_invoice_can_no_longer_be_adjusted():
    with pg_session() as db:
        _seed_container_sizes(db)
        w = World(db)
        invoice = _apply_preloading(db, w)[4]
        svc.convert_to_draft_shipment(db, [str(invoice.id)])
        line = _lines(db, invoice.id)[0]

        with pytest.raises(AppException) as exc:
            svc.adjust_line(db, str(invoice.id), str(line.id), qty=1, actor="Ms Tee")
        assert exc.value.status_code == 409
