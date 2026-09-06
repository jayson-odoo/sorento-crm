"""Supplier photos on a shipment line (R25/R26, section 12, purchasing consolidation
batch 6 Sep 2026, lane C, slice C3). UAC group L (AC-L1..L4).

Postgres only, `tests/_pg_fixture.py::blank_session`, own seeded chain (`ZZSLP`
marker). Storage is stubbed the way `test_packing_list_apply_files_attachment.py`
stubs it - upload/download are never a real network call in this suite.

No new migration: `EntityAttachmentLink` (`entity_attachment_links`) already carries
everything a new `inbound_shipment_line_photos` table would have - see the module
docstring on `app.services.scm.shipment_line_photos` and the plan's own
`## Deviations (lane C)`.
"""
from __future__ import annotations

import asyncio
import io
import uuid
from datetime import date

import openpyxl
import pytest
from fastapi import UploadFile
from starlette.datastructures import Headers

from app.models.base import set_company_scope
from app.models.entity_attachment import EntityAttachmentLink
from app.models.procurement import InboundShipment, InboundShipmentLine, Supplier
from app.models.product import Product, ProductCategory, UnitOfMeasure
from app.models.resources import Attachment, AttachmentDirectory, AttachmentType
from app.services.company_scope import DEFAULT_COMPANY_ID
from app.services.entity_attachment_service import EntityAttachmentService
from app.services.error_handler import AppException
from app.services.scm import consolidated_packing_list, shipment_line_photos
from tests._pg_fixture import blank_session, unique_code

MARKER = "ZZSLP"

# A hardcoded, valid 1x1 PNG (same fixture `test_ai_extract_route_off_the_loop.py`
# uses) - decodable by Pillow, so `store_thumbnail`/the export's own embed both work.
_TINY_PNG = (
    b"\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
    b"\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc\xf8\xcf\xc0"
    b"\x00\x00\x03\x01\x01\x00\xc9\xfe\x92\xef\x00\x00\x00\x00IEND\xaeB`\x82"
)


@pytest.fixture
def db():
    with blank_session() as session:
        set_company_scope(session, frozenset({DEFAULT_COMPANY_ID}))
        yield session


def _uom(db) -> str:
    uid = str(uuid.uuid4())
    db.add(UnitOfMeasure(id=uid, uom_code=unique_code("U")[:20], uom_name="pcs"))
    db.flush()
    return uid


def _category(db) -> str:
    cid = str(uuid.uuid4())
    db.add(
        ProductCategory(id=cid, category_code=unique_code("CAT"), category_name=f"{MARKER} category")
    )
    db.flush()
    return cid


def _product(db, code: str, *, category_id: str, uom_id: str) -> Product:
    p = Product(
        id=str(uuid.uuid4()),
        product_code=unique_code(code),
        product_name=code,
        category_id=category_id,
        base_uom_id=uom_id,
        list_price=0,
        is_active=True,
    )
    db.add(p)
    db.flush()
    return p


def _supplier(db) -> Supplier:
    s = Supplier(
        id=str(uuid.uuid4()), supplier_code=unique_code("SUP"), supplier_name=f"{MARKER} supplier", is_active=True
    )
    db.add(s)
    db.flush()
    return s


def _shipment(db, supplier_id) -> InboundShipment:
    s = InboundShipment(
        id=str(uuid.uuid4()),
        shipment_date=date(2026, 9, 1),
        supplier_id=supplier_id,
        shipping_container_number=unique_code("CONT"),
    )
    db.add(s)
    db.flush()
    return s


def _line(db, shipment_id, product_id, supplier_id, *, qty: int = 10) -> InboundShipmentLine:
    line = InboundShipmentLine(
        id=str(uuid.uuid4()),
        shipment_id=shipment_id,
        product_id=product_id,
        supplier_id=supplier_id,
        quantity_shipped=qty,
    )
    db.add(line)
    db.flush()
    return line


def _photo_type(
    db, *, code: str = "shipment_line_photo", allowed_extensions: str = "jpg,jpeg,png,webp,gif",
    default_directory_id: str | None = None,
) -> AttachmentType:
    t = AttachmentType(
        id=str(uuid.uuid4()),
        type_name="Shipment Line Photo",
        code=code,
        allowed_extensions=allowed_extensions,
        max_file_size_mb=10,
        default_directory_id=default_directory_id,
    )
    db.add(t)
    db.flush()
    return t


def _upload(name: str, *, data: bytes = _TINY_PNG, content_type: str = "image/png") -> UploadFile:
    return UploadFile(file=io.BytesIO(data), filename=name, headers=Headers({"content-type": content_type}))


def _link_photo(db, *, line_id: str, provider: str = "r2", filename: str = "photo.png") -> Attachment:
    """A photo attached WITHOUT going through `upload_photos` - for tests that need
    to control the attachment's own `storage_provider` (AC-L4) or seed many at once
    without paying the (stubbed) upload round trip each time."""
    att = Attachment(
        id=str(uuid.uuid4()),
        original_filename=filename,
        stored_filename=filename,
        file_path=f"https://cdn.test/{provider}/{filename}",
        mime_type="image/png",
        storage_provider=provider,
    )
    db.add(att)
    db.flush()
    EntityAttachmentService(db).link_existing_attachment(
        entity_type=shipment_line_photos.ENTITY_TYPE,
        entity_id=str(line_id),
        attachment_id=str(att.id),
    )
    return att


def _stub_backend(monkeypatch, *, provider: str = "r2") -> None:
    """Same shape `test_packing_list_apply_files_attachment.py`'s own `_stub_backend`
    uses, extended with `download_file` (the export and `_photo_bytes` read bytes
    back) and `get_signed_url` (every photo's `url`/`thumbnail_url` goes through it).
    """
    from app.services.storage_router import clear_signed_url_cache

    clear_signed_url_cache()
    monkeypatch.setattr("app.services.storage_router.default_provider", lambda: provider)
    monkeypatch.setattr(
        "app.services.storage_router.get_backend",
        lambda p: type(
            "StubBackend",
            (),
            {
                "upload_file": staticmethod(lambda **kw: (f"stub/{p}/key.png", "")),
                "download_file": staticmethod(lambda key: _TINY_PNG),
                "get_signed_url": staticmethod(
                    lambda key, expires_in=3600: f"https://signed.test/{p}/{key}"
                ),
                "get_cloudfront_base_url": staticmethod(lambda key: f"https://cdn.test/{key}"),
                "get_cdn_base_url": staticmethod(lambda key: f"https://cdn.test/{key}"),
            },
        )(),
    )


def _seed_line(db, *, code: str = "TAP", qty: int = 10):
    supplier = _supplier(db)
    category = _category(db)
    uom = _uom(db)
    product = _product(db, code, category_id=category, uom_id=uom)
    shipment = _shipment(db, supplier.id)
    line = _line(db, shipment.id, product.id, supplier.id, qty=qty)
    return shipment, line


# --------------------------------------------------------------------------- #
# AC-L1 - list / upload / delete
# --------------------------------------------------------------------------- #


def test_upload_then_list_returns_photos_in_order(db, monkeypatch):
    _stub_backend(monkeypatch)
    _photo_type(db)
    shipment, line = _seed_line(db)
    db.commit()

    out = asyncio.run(
        shipment_line_photos.upload_photos(
            db,
            shipment_id=str(shipment.id),
            line_id=str(line.id),
            files=[_upload("a.png"), _upload("b.png")],
            actor_id=None,
        )
    )

    assert [p["filename"] for p in out] == ["a.png", "b.png"]
    assert [p["sort_order"] for p in out] == [0, 1]
    assert all(p["url"] for p in out)
    assert all(p["thumbnail_url"] for p in out)
    assert all(p["attachment_id"] for p in out)

    by_line = shipment_line_photos.list_for_shipment(db, str(shipment.id))
    assert [p["filename"] for p in by_line[str(line.id)]] == ["a.png", "b.png"]


def test_upload_without_a_configured_type_is_a_named_400(db, monkeypatch):
    """No 'Shipment Line Photo' attachment type seeded - the upload must refuse with
    a clear error, never silently skip filing the photo (unlike the packing-list
    workbook's own best-effort filing, uploading IS the point of this endpoint)."""
    _stub_backend(monkeypatch)
    shipment, line = _seed_line(db)
    db.commit()

    with pytest.raises(AppException) as excinfo:
        asyncio.run(
            shipment_line_photos.upload_photos(
                db,
                shipment_id=str(shipment.id),
                line_id=str(line.id),
                files=[_upload("a.png")],
                actor_id=None,
            )
        )
    assert excinfo.value.status_code == 400


def test_upload_rejects_a_non_image_extension(db, monkeypatch):
    _stub_backend(monkeypatch)
    _photo_type(db)
    shipment, line = _seed_line(db)
    db.commit()

    with pytest.raises(AppException) as excinfo:
        asyncio.run(
            shipment_line_photos.upload_photos(
                db,
                shipment_id=str(shipment.id),
                line_id=str(line.id),
                files=[_upload("not-a-photo.pdf", content_type="application/pdf")],
                actor_id=None,
            )
        )
    assert excinfo.value.status_code == 400


def test_upload_rejects_a_non_image_extension_even_when_the_type_row_allows_it(db, monkeypatch):
    """Review round 1, item 2: the image guard is independent of the attachment
    type's own `allowed_extensions` - an admin later widening that row (e.g. to file
    a spec sheet under the same type) must not turn this into an arbitrary-file
    upload."""
    _stub_backend(monkeypatch)
    _photo_type(db, allowed_extensions="*")
    shipment, line = _seed_line(db)
    db.commit()

    with pytest.raises(AppException) as excinfo:
        asyncio.run(
            shipment_line_photos.upload_photos(
                db,
                shipment_id=str(shipment.id),
                line_id=str(line.id),
                files=[_upload("not-a-photo.pdf", content_type="application/pdf")],
                actor_id=None,
            )
        )
    assert excinfo.value.status_code == 400


def test_upload_derives_content_type_from_the_extension_when_none_is_sent(db, monkeypatch):
    """Review round 1, item 2: a bare multipart part with no `Content-Type` header
    must still store and thumbnail as the image it is."""
    _stub_backend(monkeypatch)
    _photo_type(db)
    shipment, line = _seed_line(db)
    db.commit()

    out = asyncio.run(
        shipment_line_photos.upload_photos(
            db,
            shipment_id=str(shipment.id),
            line_id=str(line.id),
            files=[_upload("a.png", content_type="")],
            actor_id=None,
        )
    )

    assert out[0]["thumbnail_url"], "no content-type must still thumbnail like an image"


def test_upload_files_into_the_type_own_default_directory(db, monkeypatch):
    """Review round 1, item 6: `directory_id` follows the attachment type's own
    `default_directory_id`, same convention `packing_list_service` reads it by."""
    _stub_backend(monkeypatch)
    directory_id = str(uuid.uuid4())
    db.add(AttachmentDirectory(id=directory_id, name=f"{MARKER} folder"))
    db.flush()
    _photo_type(db, default_directory_id=directory_id)
    shipment, line = _seed_line(db)
    db.commit()

    out = asyncio.run(
        shipment_line_photos.upload_photos(
            db,
            shipment_id=str(shipment.id),
            line_id=str(line.id),
            files=[_upload("a.png")],
            actor_id=None,
        )
    )

    attachment = db.query(Attachment).filter(Attachment.id == out[0]["attachment_id"]).first()
    assert str(attachment.directory_id) == directory_id


def test_upload_returns_the_lines_full_photo_list_not_just_this_batch(db, monkeypatch):
    """Review round 1, item 10: a second upload call returns EVERY photo on the line,
    the first batch included, not only what this call just added."""
    _stub_backend(monkeypatch)
    _photo_type(db)
    shipment, line = _seed_line(db)
    db.commit()

    asyncio.run(
        shipment_line_photos.upload_photos(
            db, shipment_id=str(shipment.id), line_id=str(line.id),
            files=[_upload("a.png")], actor_id=None,
        )
    )
    out = asyncio.run(
        shipment_line_photos.upload_photos(
            db, shipment_id=str(shipment.id), line_id=str(line.id),
            files=[_upload("b.png")], actor_id=None,
        )
    )

    assert [p["filename"] for p in out] == ["a.png", "b.png"]


def test_a_failure_after_the_put_leaves_no_orphan_object_and_keeps_earlier_files(db, monkeypatch):
    """Review round 1, item 4: the second file's own object (and thumbnail) purge on
    a forced `create_attachment` failure; the first file, already committed, stays."""
    _stub_backend(monkeypatch)
    _photo_type(db)
    shipment, line = _seed_line(db)
    db.commit()

    deleted: list[tuple[str, str]] = []
    monkeypatch.setattr(
        "app.services.scm.shipment_line_photos.delete_object_best_effort",
        lambda provider, key: deleted.append((provider, key)),
    )

    calls = {"n": 0}
    from app.services.resources_service import AttachmentService

    real_create = AttachmentService.create_attachment

    def _flaky_create(self, data, actor_id=None):
        calls["n"] += 1
        if calls["n"] == 2:
            raise RuntimeError("boom")
        return real_create(self, data, actor_id)

    monkeypatch.setattr(AttachmentService, "create_attachment", _flaky_create)

    with pytest.raises(AppException) as excinfo:
        asyncio.run(
            shipment_line_photos.upload_photos(
                db,
                shipment_id=str(shipment.id),
                line_id=str(line.id),
                files=[_upload("a.png"), _upload("b.png")],
                actor_id=None,
            )
        )

    assert excinfo.value.status_code == 400
    assert "a.png" in excinfo.value.detail["message"]
    assert "b.png" in excinfo.value.detail["message"]
    assert deleted, "the second file's own object was never purged"

    by_line = shipment_line_photos.list_for_shipment(db, str(shipment.id))
    assert [p["filename"] for p in by_line[str(line.id)]] == ["a.png"]


def test_delete_removes_the_link_and_the_attachment_row(db, monkeypatch):
    """AC-L2: the row (here, `EntityAttachmentLink` - the reused linkage mechanism)
    and the attachment it points at are both gone, via the FK's own CASCADE."""
    _stub_backend(monkeypatch)
    _photo_type(db)
    shipment, line = _seed_line(db)
    db.commit()

    out = asyncio.run(
        shipment_line_photos.upload_photos(
            db,
            shipment_id=str(shipment.id),
            line_id=str(line.id),
            files=[_upload("a.png")],
            actor_id=None,
        )
    )
    photo_id, attachment_id = out[0]["id"], out[0]["attachment_id"]

    shipment_line_photos.delete_photo(db, str(shipment.id), str(line.id), photo_id)

    assert db.query(EntityAttachmentLink).filter(EntityAttachmentLink.id == photo_id).first() is None
    assert db.query(Attachment).filter(Attachment.id == attachment_id).first() is None


def test_delete_an_unknown_photo_404s(db, monkeypatch):
    _stub_backend(monkeypatch)
    shipment, line = _seed_line(db)
    db.commit()

    with pytest.raises(AppException) as excinfo:
        shipment_line_photos.delete_photo(db, str(shipment.id), str(line.id), str(uuid.uuid4()))
    assert excinfo.value.status_code == 404


def test_delete_a_photo_under_another_shipment_404s_and_deletes_nothing(db, monkeypatch):
    """Review round 1, item 1: `EntityAttachmentLink` carries no company scope of its
    own, so matching on `photo_id` alone would let a caller who merely knows another
    shipment's photo id delete it - the delete has to be scoped to shipment_id/line_id
    too."""
    _stub_backend(monkeypatch)
    _photo_type(db)
    shipment, line = _seed_line(db)
    other_shipment, other_line = _seed_line(db, code="OTHER")
    db.commit()

    out = asyncio.run(
        shipment_line_photos.upload_photos(
            db,
            shipment_id=str(shipment.id),
            line_id=str(line.id),
            files=[_upload("a.png")],
            actor_id=None,
        )
    )
    photo_id, attachment_id = out[0]["id"], out[0]["attachment_id"]

    with pytest.raises(AppException) as excinfo:
        shipment_line_photos.delete_photo(
            db, str(other_shipment.id), str(other_line.id), photo_id
        )
    assert excinfo.value.status_code == 404
    assert db.query(EntityAttachmentLink).filter(EntityAttachmentLink.id == photo_id).first() is not None
    assert db.query(Attachment).filter(Attachment.id == attachment_id).first() is not None


# --------------------------------------------------------------------------- #
# AC-L4 - r2 and s3 both resolve through storage_router
# --------------------------------------------------------------------------- #


def test_list_resolves_both_an_r2_and_an_s3_photo(db, monkeypatch):
    _stub_backend(monkeypatch)
    shipment, line = _seed_line(db)
    _link_photo(db, line_id=line.id, provider="r2", filename="r2-photo.png")
    _link_photo(db, line_id=line.id, provider="s3", filename="s3-photo.png")
    db.commit()

    photos = shipment_line_photos.list_for_shipment(db, str(shipment.id))[str(line.id)]
    by_name = {p["filename"]: p for p in photos}

    assert "/r2/" in by_name["r2-photo.png"]["url"]
    assert "/s3/" in by_name["s3-photo.png"]["url"]


# --------------------------------------------------------------------------- #
# AC-L3 - the export
# --------------------------------------------------------------------------- #


def test_export_has_no_photo_columns_when_nothing_was_uploaded(db, monkeypatch):
    _stub_backend(monkeypatch)
    shipment, _line = _seed_line(db)
    db.commit()

    payload = consolidated_packing_list.build(db, str(shipment.id))
    wb = openpyxl.load_workbook(io.BytesIO(consolidated_packing_list.to_xlsx(payload)))
    ws = wb["RMB"]

    assert ws.cell(row=15, column=23).value is None  # W - one past TOTAL RM's own V


def test_export_prints_photo_1_through_n_after_the_busiest_line_and_anchors_each_image(
    db, monkeypatch
):
    """Two lines, 3 photos and 1 (Q5: no cap, `n` = the max) -> `PHOTO 1..3` headers,
    each image anchored on ITS OWN line's row."""
    _stub_backend(monkeypatch)
    supplier = _supplier(db)
    category = _category(db)
    uom = _uom(db)
    # Same supplier/factory block, so ordering is purely by product_code (AC-G3's own
    # rule `build()` already sorts by) - "TAP1" before "TAP2" regardless of the random
    # suffix `unique_code` appends, since that suffix never changes the given stem.
    product_a = _product(db, "TAP1", category_id=category, uom_id=uom)
    product_b = _product(db, "TAP2", category_id=category, uom_id=uom)
    shipment = _shipment(db, supplier.id)
    line_a = _line(db, shipment.id, product_a.id, supplier.id)
    line_b = _line(db, shipment.id, product_b.id, supplier.id)

    for i in range(3):
        _link_photo(db, line_id=line_a.id, filename=f"a-{i}.png")
    _link_photo(db, line_id=line_b.id, filename="b-0.png")
    db.commit()

    payload = consolidated_packing_list.build(db, str(shipment.id))
    xlsx = consolidated_packing_list.to_xlsx(payload)
    wb = openpyxl.load_workbook(io.BytesIO(xlsx))
    ws = wb["RMB"]

    # W, X, Y = columns 23, 24, 25; no PHOTO 4 (n stops at the busiest line's count).
    assert ws.cell(row=15, column=23).value == "PHOTO 1"
    assert ws.cell(row=15, column=24).value == "PHOTO 2"
    assert ws.cell(row=15, column=25).value == "PHOTO 3"
    assert ws.cell(row=15, column=26).value is None

    # RMB / TOTAL RM keep their existing letters (T / U) - the whole reason the photo
    # columns were appended after V rather than inserted before them (see
    # `_FIRST_PHOTO_COLUMN`'s own comment).
    assert ws.cell(row=15, column=20).value == "RMB"
    assert ws.cell(row=15, column=21).value == "TOTAL RM"

    assert len(ws._images) == 4  # 3 + 1
    placements = sorted(
        (img.anchor._from.row, img.anchor._from.col) for img in ws._images
    )
    # `_FIRST_LINE_ROW` is 18 (0-based 17): product_a's row, 3 images across columns
    # W/X/Y (0-based 22/23/24); product_b's row is the next one, one image at W.
    assert placements == [(17, 22), (17, 23), (17, 24), (18, 22)]


# --------------------------------------------------------------------------- #
# Routes - GET / POST / DELETE over the wire (review round 1, item 9)
# --------------------------------------------------------------------------- #

from fastapi.testclient import TestClient  # noqa: E402

from app.dependencies import get_current_user, get_current_user_or_api_key, get_db  # noqa: E402
from app.main import app  # noqa: E402
from app.services.company_scope_resolver import apply_company_scope  # noqa: E402
from app.services.user_service import UserPermissionService  # noqa: E402

#: The incumbent company every row this suite seeds is auto-stamped with, same
#: constant `test_consolidated_packing_list.py` reads by (`tests/conftest.py`).
_ROUTE_COMPANY_ID = "00000000-0000-0000-0000-000000000001"


def _route_caller(db, monkeypatch, *, permitted_slugs: set[str]) -> TestClient:
    """A TestClient reading the session the test seeded, holding exactly the given
    permission slugs - same shape `test_consolidated_packing_list.py`'s own
    `_caller` uses, generalised to more than one slug (GET needs
    `scm.dashboard.view`, POST/DELETE need `scm.reorder.run`)."""
    principal = {"id": str(uuid.uuid4()), "email": "zzt-slp@example.com"}
    app.dependency_overrides[get_db] = lambda: db
    app.dependency_overrides[get_current_user] = lambda: principal
    app.dependency_overrides[get_current_user_or_api_key] = lambda: principal
    monkeypatch.setattr(
        UserPermissionService,
        "check_user_has_permission",
        lambda self, uid, slug: slug in permitted_slugs,
    )

    scope = frozenset({_ROUTE_COMPANY_ID})
    set_company_scope(db, scope)

    async def _scope():
        set_company_scope(db, scope)
        return scope

    app.dependency_overrides[apply_company_scope] = _scope
    return TestClient(app)


@pytest.fixture
def route_client(db, monkeypatch):
    """Holds both the read and the write permission this router's routes use."""
    try:
        yield _route_caller(
            db, monkeypatch, permitted_slugs={"scm.dashboard.view", "scm.reorder.run"}
        )
    finally:
        app.dependency_overrides.clear()


@pytest.fixture
def reader_only_client(db, monkeypatch):
    """Holds only the read permission - every write route must refuse it."""
    try:
        yield _route_caller(db, monkeypatch, permitted_slugs={"scm.dashboard.view"})
    finally:
        app.dependency_overrides.clear()


def test_get_route_lists_photos_keyed_by_line_id(route_client, db, monkeypatch):
    _stub_backend(monkeypatch)
    shipment, line = _seed_line(db)
    _link_photo(db, line_id=line.id, filename="via-route.png")
    db.commit()

    r = route_client.get(f"/api/v1/scm/inbound-shipments/{shipment.id}/line-photos")

    assert r.status_code == 200, r.text
    body = r.json()
    assert [p["filename"] for p in body[str(line.id)]] == ["via-route.png"]
    assert {"id", "attachment_id", "sort_order", "thumbnail_url", "url", "filename"} <= set(
        body[str(line.id)][0]
    )


def test_get_route_requires_the_read_permission(db, monkeypatch):
    _stub_backend(monkeypatch)
    shipment, _line = _seed_line(db)
    db.commit()
    client = _route_caller(db, monkeypatch, permitted_slugs=set())
    try:
        r = client.get(f"/api/v1/scm/inbound-shipments/{shipment.id}/line-photos")
        assert r.status_code == 403, r.text
    finally:
        app.dependency_overrides.clear()


def test_post_route_uploads_a_photo_and_returns_the_lines_full_list(route_client, db, monkeypatch):
    _stub_backend(monkeypatch)
    _photo_type(db)
    shipment, line = _seed_line(db)
    db.commit()

    r = route_client.post(
        f"/api/v1/scm/inbound-shipments/{shipment.id}/lines/{line.id}/photos",
        files={"files": ("a.png", _TINY_PNG, "image/png")},
    )

    assert r.status_code == 200, r.text
    assert [p["filename"] for p in r.json()] == ["a.png"]


def test_post_route_requires_the_write_permission(reader_only_client, db, monkeypatch):
    _stub_backend(monkeypatch)
    _photo_type(db)
    shipment, line = _seed_line(db)
    db.commit()

    r = reader_only_client.post(
        f"/api/v1/scm/inbound-shipments/{shipment.id}/lines/{line.id}/photos",
        files={"files": ("a.png", _TINY_PNG, "image/png")},
    )

    assert r.status_code == 403, r.text


def test_delete_route_removes_the_photo(route_client, db, monkeypatch):
    _stub_backend(monkeypatch)
    _photo_type(db)
    shipment, line = _seed_line(db)
    db.commit()

    upload = route_client.post(
        f"/api/v1/scm/inbound-shipments/{shipment.id}/lines/{line.id}/photos",
        files={"files": ("a.png", _TINY_PNG, "image/png")},
    )
    photo_id = upload.json()[0]["id"]

    r = route_client.delete(
        f"/api/v1/scm/inbound-shipments/{shipment.id}/lines/{line.id}/photos/{photo_id}"
    )

    assert r.status_code == 200, r.text
    assert db.query(EntityAttachmentLink).filter(EntityAttachmentLink.id == photo_id).first() is None


def test_delete_route_404s_for_a_photo_under_another_shipment(route_client, db, monkeypatch):
    """Review round 1, item 1, proved over the wire: the cross-shipment guess 404s
    and deletes nothing."""
    _stub_backend(monkeypatch)
    _photo_type(db)
    shipment, line = _seed_line(db)
    other_shipment, other_line = _seed_line(db, code="OTHER")
    db.commit()

    upload = route_client.post(
        f"/api/v1/scm/inbound-shipments/{shipment.id}/lines/{line.id}/photos",
        files={"files": ("a.png", _TINY_PNG, "image/png")},
    )
    photo_id = upload.json()[0]["id"]

    r = route_client.delete(
        f"/api/v1/scm/inbound-shipments/{other_shipment.id}/lines/{other_line.id}/photos/{photo_id}"
    )

    assert r.status_code == 404, r.text
    assert db.query(EntityAttachmentLink).filter(EntityAttachmentLink.id == photo_id).first() is not None
