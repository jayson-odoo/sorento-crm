"""The order inquiry handshake: CS raises, purchasing acknowledges (`PLAN-scm-oi-handshake.md`).

Links are purchasing's word. Nothing is linked when a row is raised; the cascade runs at
ACKNOWLEDGE, at Link now, and at a purchase-order confirm, and every one of those three is
restricted to rows purchasing has taken on. What is pinned here, one test each:

* AC-H1/AC-H11 a board confirm raises `awaiting` rows and links NOTHING, even with an open
  purchase-order line sitting there waiting for the product;
* AC-H2 Acknowledge takes a batch on, stamps who and when, and runs the cascade for exactly
  those rows;
* AC-H3 a CS user (no `projects.order_inquiries.acknowledge`) is refused 403 by every
  one of the three write routes;
* AC-H4 the `ack` filter is a closed set, and the facet counts all four states with its own
  filter dropped;
* AC-H5 a reject needs a reason, and 422 says so;
* AC-H6 a rejected row leaves `scm.committed_v` and the plan's own committed SELECT, and the
  line goes back to the board undecided carrying the reason;
* AC-H7/AC-H8/AC-H9 an amend before acknowledgement is silent, one after it reads `changed`
  with the previous value and its links kept, and re-acknowledging returns it;
* AC-H10 the plan counts acknowledged and changed rows only, and awaiting rows are a count;
* AC-H13 Link now is scoped to the products named;
* AC-H14 every new column is on the wire from the list, the summary and the export.

Runs on the REAL database (rolled back), like `test_planning_change_apply_on_board`:
`scm.committed_v` is a view a migration installs and the blank scratch schema has none.
Every row is seeded here behind the marker - CI's database has no data.
"""
from __future__ import annotations

import io
from contextlib import contextmanager
from datetime import date, datetime
from decimal import Decimal

import pytest
from sqlalchemy import text
from sqlalchemy.orm import Session

from app.database import engine
from app.models.procurement import PurchaseOrder, PurchaseOrderLine, Supplier
from app.models.project_so import (
    ACK_ACKNOWLEDGED,
    ACK_AWAITING,
    ACK_CHANGED,
    ACK_REJECTED,
    INQUIRY_CANCELLED,
    IV_ORDER,
    OrderInquiryRow,
)
from app.services import project_seed_service
from app.services.project_order_inquiry_service import ProjectOrderInquiryService
from app.services.project_supply_service import ProjectSupplyService

from .test_planning_changes import (
    BASE,
    MARKER,
    _core_line,
    _core_so,
    _line_payload,
    _product,
    _project_line,
    _project_so,
    _sorento,
    _uid,
    _user,
    _warehouse,
)

VIEW = "projects.projects.view"
ACTION = "projects.order_inquiry.action"
EDIT = "projects.projects.edit"
ACKNOWLEDGE = "projects.order_inquiries.acknowledge"
PURCHASING = [VIEW, ACTION, ACKNOWLEDGE]
CS = [VIEW, EDIT, ACTION]

LIST = f"{BASE}/order-inquiries"
ACK_URL = f"{LIST}/acknowledge"
LINK_NOW = f"{LIST}/link-now"

WAS = date(2026, 8, 25)
NOW = date(2026, 8, 19)


# ---------------------------------------------------------------------------
# harness
# ---------------------------------------------------------------------------


@contextmanager
def _real_db_session():
    """The real database, every write discarded, and a route's own `rollback()` survived.

    Same reasoning as `test_planning_change_apply_on_board._real_db_session`: the view lives
    in the migrated schema, and `autoflush=False` is how the application's session is built,
    so a service that queries for its own pending write fails here as it would live.
    """
    connection = engine.connect()
    transaction = connection.begin()
    session = Session(
        bind=connection, join_transaction_mode="create_savepoint", autoflush=False
    )
    try:
        yield session
    finally:
        session.close()
        transaction.rollback()
        connection.close()


def _client(db, user_id: str, permissions):
    from fastapi.testclient import TestClient

    from app.database import get_db
    from app.dependencies import get_current_user, get_current_user_or_api_key
    from app.main import app
    from app.services.company_scope_resolver import apply_company_scope
    from app.services.user_service import UserPermissionService

    actor = {"id": user_id, "email": f"{user_id}@zzt.test", "role": "user"}
    app.dependency_overrides[get_db] = lambda: db
    app.dependency_overrides[get_current_user] = lambda: dict(actor)
    app.dependency_overrides[get_current_user_or_api_key] = lambda: dict(actor)
    app.dependency_overrides[apply_company_scope] = lambda: None

    originals = (
        UserPermissionService.check_user_has_permission,
        UserPermissionService.get_user_permission_slugs,
    )
    granted = list(permissions)
    UserPermissionService.check_user_has_permission = (
        lambda self, uid, slug: slug in granted
    )
    UserPermissionService.get_user_permission_slugs = lambda self, uid: list(granted)
    return TestClient(app), originals


def _restore(originals) -> None:
    from app.main import app
    from app.services.user_service import UserPermissionService

    UserPermissionService.check_user_has_permission = originals[0]
    UserPermissionService.get_user_permission_slugs = originals[1]
    app.dependency_overrides.clear()


class _World:
    def __init__(self, db, company_id, cs_user, buyer, project, product, warehouse,
                 plan_run):
        self.db = db
        self.company_id = company_id
        self.cs_user = cs_user
        self.buyer = buyer
        self.project = project
        self.product = product
        self.warehouse = warehouse
        #: The reorder run this suite's LINK HORIZON default is read off (S5, code review
        #: 27 Aug 2026). See `_pin_the_plan_horizon`.
        self.plan_run = plan_run


def _pin_the_plan_horizon(db, company_id) -> str:
    """The LATEST COMPLETED reorder run, seeded with NO "Plan until" (S5, code review
    27 Aug 2026).

    `plan_link_horizon` reads the latest completed run, and this suite runs on the SHARED
    local database - a copy of production, carrying real planning runs. Left alone, every
    acknowledge in this file and in `test_order_inquiry_handshake_edges.py` would link
    under whatever horizon somebody's last real run happened to name, and the same test
    would pass or fail depending on the copy. One seeded run, finishing now and naming no
    horizon, makes the default a fact of the fixture: NO horizon unless a test says
    otherwise, which the two AC-LH5 tests do by writing a date onto this run.
    """
    run_id = _uid()
    finished = datetime.utcnow()
    db.execute(
        text(
            "INSERT INTO scm.reorder_run (id, company_id, status, plan_horizon_date, "
            "started_at, finished_at, created_at) "
            "VALUES (:i, :c, 'completed', NULL, :f, :f, :f)"
        ),
        {"i": run_id, "c": company_id, "f": finished},
    )
    db.flush()
    return run_id


@pytest.fixture()
def world():
    from app.services.project_service import register_project

    with _real_db_session() as db:
        company_id = _sorento(db)
        project_seed_service.run(db, company_id=company_id)
        cs_user = _user(db, f"{MARKER} Eling")
        buyer = _user(db, f"{MARKER} Joey")
        project = register_project(
            db,
            company_id=company_id,
            actor_user_id=cs_user,
            developer_party_id=None,
            title=f"{MARKER} Yotu Builder {_uid()[:12]}",
        )
        product = _product(db)
        warehouse = _warehouse(db, f"ZZT-IB-{_uid()[:4]}", segment="project")
        plan_run = _pin_the_plan_horizon(db, company_id)
        db.flush()
        db.commit()
        yield _World(
            db, company_id, cs_user, buyer, project, product, warehouse, plan_run
        )


@pytest.fixture()
def api(world):
    """Two clients on one session: CS confirms on the board, purchasing acknowledges."""
    from app.models.base import company_scope

    cs_client, cs_originals = _client(world.db, world.cs_user, CS)
    try:
        with company_scope(world.db, frozenset({world.company_id})):
            yield cs_client, world
    finally:
        _restore(cs_originals)


@contextmanager
def _as_purchasing(world, permissions=PURCHASING):
    client, originals = _client(world.db, world.buyer, permissions)
    try:
        yield client
    finally:
        _restore(originals)
        # The CS client the `api` fixture installed is put back, so a test can go on
        # confirming on the board after purchasing has acted.
        _client(world.db, world.cs_user, CS)


# ---------------------------------------------------------------------------
# seeding
# ---------------------------------------------------------------------------


def _supplier(world) -> Supplier:
    supplier = Supplier(
        id=_uid(),
        company_id=world.company_id,
        supplier_code=f"ZZT-{_uid()[:8]}",
        supplier_name=f"{MARKER} DAFUYUAN",
    )
    world.db.add(supplier)
    world.db.flush()
    return supplier


def _open_po_line(world, *, qty, expected_date=date(2026, 8, 10), product=None):
    supplier = _supplier(world)
    po = PurchaseOrder(
        id=_uid(),
        company_id=world.company_id,
        po_number=f"ZZT-PO-{_uid()[:8]}",
        supplier_id=supplier.id,
        issue_date=date(2026, 6, 1),
        status="active",
    )
    world.db.add(po)
    world.db.flush()
    line = PurchaseOrderLine(
        id=_uid(),
        company_id=world.company_id,
        purchase_order_id=po.id,
        product_id=(product or world.product).id,
        warehouse_id=world.warehouse.id,
        qty_ordered=Decimal(str(qty)),
        qty_received=Decimal("0"),
        expected_date=expected_date,
        line_status="open",
    )
    world.db.add(line)
    world.db.flush()
    world.db.commit()
    return po, line


def _confirm(client, order_id, lines):
    return client.post(f"{BASE}/sales-orders/{order_id}/confirm", json={"lines": lines})


def _raise_one_row(api, *, qty="10", product=None):
    """One published order, one line, confirmed wholly as Buy: one raised inquiry row."""
    client, world = api
    db = world.db
    product = product or world.product
    core_so = _core_so(db, world.company_id)
    core_line = _core_line(
        db, core_so, product, world.warehouse, qty_ordered=qty, required_date=WAS
    )
    order = _project_so(
        db, world.project, so_id=core_so.id, autocount_doc_no=core_so.so_number
    )
    line = _project_line(db, order, line_no=1, product=product, core_line=core_line)
    db.commit()

    response = _confirm(client, order.id, [_line_payload(line.id, buy_qty=qty)])
    assert response.status_code == 200, response.text
    db.commit()
    return {
        "order": order,
        "line": line,
        "core_so": core_so,
        "core_line": core_line,
        "row": _order_row(world, line),
    }


def _order_row(world, line) -> OrderInquiryRow:
    return (
        world.db.query(OrderInquiryRow)
        .filter(
            OrderInquiryRow.so_line_id == line.id,
            OrderInquiryRow.verb == IV_ORDER,
            OrderInquiryRow.state != INQUIRY_CANCELLED,
        )
        .one()
    )


def _links_of(world, row) -> list:
    return ProjectOrderInquiryService(world.db)._links_of(str(row.id))


def _project_committed(world, *, planned: bool) -> Decimal:
    """What is committed for the product at its warehouse - as the VIEW says it, or as the
    PLAN's own SELECT does (`demand.horizon_committed_select_sql`, acknowledged only)."""
    from app.services.scm import demand

    if planned:
        sql = (
            f"SELECT COALESCE(SUM(project_committed), 0) FROM ("
            f"{demand.horizon_committed_select_sql()}) cv "
            "WHERE cv.product_id = :pid"
        )
        params = {"pid": str(world.product.id), "horizon": None}
    else:
        sql = (
            "SELECT COALESCE(SUM(project_committed), 0) FROM scm.committed_v "
            "WHERE product_id = :pid"
        )
        params = {"pid": str(world.product.id)}
    return Decimal(str(world.db.execute(text(sql), params).scalar() or 0))


# ---------------------------------------------------------------------------
# AC-H1 / AC-H11: raising links nothing
# ---------------------------------------------------------------------------


def test_a_board_confirm_raises_awaiting_rows_and_links_nothing(api):
    """The cascade left `supply.confirm` (plan section 3). An open purchase-order line for
    the very product is sitting there, and the row still comes out unlinked: linking is
    purchasing's word, and nobody has said it yet."""
    _client, world = api
    _open_po_line(world, qty=50)

    fixture = _raise_one_row(api)
    row = fixture["row"]

    assert row.ack_state == ACK_AWAITING
    assert row.acknowledged_by is None and row.acknowledged_at is None
    assert _links_of(world, row) == [], "nothing links until somebody acknowledges"


# ---------------------------------------------------------------------------
# AC-H2: acknowledge, one and many
# ---------------------------------------------------------------------------


def test_acknowledge_stamps_who_and_when_and_runs_the_cascade_for_those_rows(api):
    _client, world = api
    _open_po_line(world, qty=50)
    fixture = _raise_one_row(api)
    row = fixture["row"]

    with _as_purchasing(world) as buyer:
        response = buyer.post(ACK_URL, json={"row_ids": [str(row.id)]})
    assert response.status_code == 200, response.text
    world.db.commit()

    body = response.json()
    assert body["acknowledged"] == 1
    assert body["linked_rows"] == 1 and body["links"] >= 1

    world.db.refresh(row)
    assert row.ack_state == ACK_ACKNOWLEDGED
    assert str(row.acknowledged_by) == str(world.buyer)
    assert row.acknowledged_at is not None
    assert sum(Decimal(str(link.qty)) for link in _links_of(world, row)) == Decimal("10")


def test_acknowledge_takes_a_batch_on_in_one_press(api):
    _client, world = api
    first = _raise_one_row(api, qty="4")
    second = _raise_one_row(api, qty="6")

    with _as_purchasing(world) as buyer:
        response = buyer.post(
            ACK_URL,
            json={"row_ids": [str(first["row"].id), str(second["row"].id)]},
        )
    assert response.status_code == 200, response.text
    world.db.commit()
    assert response.json()["acknowledged"] == 2
    for fixture in (first, second):
        world.db.refresh(fixture["row"])
        assert fixture["row"].ack_state == ACK_ACKNOWLEDGED


def test_acknowledging_a_rejected_row_is_refused(api):
    """`awaiting` and `changed` are the two an acknowledgement may be taken on. A rejected
    row is purchasing's own refusal: taking it on again means CS re-deciding the line."""
    _client, world = api
    fixture = _raise_one_row(api)
    row = fixture["row"]

    with _as_purchasing(world) as buyer:
        reject = buyer.post(
            f"{LIST}/{row.id}/reject", json={"reason": "No supplier until November"}
        )
        assert reject.status_code == 200, reject.text
        world.db.commit()
        response = buyer.post(ACK_URL, json={"row_ids": [str(row.id)]})

    assert response.status_code == 422, response.text


# ---------------------------------------------------------------------------
# AC-H3: CS cannot acknowledge, reject or link
# ---------------------------------------------------------------------------


def test_a_cs_user_is_refused_by_every_write_route(api):
    """CS sees the column and the filter; the three actions are purchasing's grant."""
    cs_client, world = api
    fixture = _raise_one_row(api)
    row_id = str(fixture["row"].id)

    assert cs_client.post(ACK_URL, json={"row_ids": [row_id]}).status_code == 403
    assert (
        cs_client.post(f"{LIST}/{row_id}/reject", json={"reason": "no"}).status_code
        == 403
    )
    assert cs_client.post(LINK_NOW, json={}).status_code == 403
    # ... and the read they DO hold is unaffected.
    assert cs_client.get(LIST).status_code == 200


# ---------------------------------------------------------------------------
# AC-H4: the filter and the facet
# ---------------------------------------------------------------------------


def test_the_ack_filter_narrows_the_list_and_the_facet_counts_all_four_states(api):
    client, world = api
    awaiting = _raise_one_row(api, qty="4")
    acknowledged = _raise_one_row(api, qty="6")

    with _as_purchasing(world) as buyer:
        assert (
            buyer.post(
                ACK_URL, json={"row_ids": [str(acknowledged["row"].id)]}
            ).status_code
            == 200
        )
    world.db.commit()

    listed = client.get(LIST, params={"ack": "awaiting", "limit": 200}).json()
    ids = {row["id"] for row in listed["data"]}
    assert str(awaiting["row"].id) in ids
    assert str(acknowledged["row"].id) not in ids

    summary = client.get(
        f"{LIST}/summary", params={"ack": "awaiting"}
    ).json()
    # The facet drops its OWN filter, like every other control on this screen, so the
    # acknowledged row is still counted while the awaiting one is the only one listed.
    assert set(summary["ack"]) == {"awaiting", "acknowledged", "changed", "rejected"}
    assert summary["ack"]["acknowledged"] >= 1
    assert summary["ack"]["awaiting"] >= 1


def test_an_unknown_ack_value_is_refused(api):
    client, _world = api
    response = client.get(LIST, params={"ack": "maybe"})
    assert response.status_code == 422


# ---------------------------------------------------------------------------
# AC-H5 / AC-H6: reject
# ---------------------------------------------------------------------------


def test_a_reject_with_no_reason_is_refused(api):
    _client, world = api
    fixture = _raise_one_row(api)

    with _as_purchasing(world) as buyer:
        blank = buyer.post(f"{LIST}/{fixture['row'].id}/reject", json={"reason": "   "})
        missing = buyer.post(f"{LIST}/{fixture['row'].id}/reject", json={})

    assert blank.status_code == 422, blank.text
    assert missing.status_code == 422, missing.text
    world.db.refresh(fixture["row"])
    assert fixture["row"].ack_state == ACK_AWAITING


def test_a_rejected_row_records_who_when_and_why(api):
    _client, world = api
    fixture = _raise_one_row(api)
    row = fixture["row"]

    with _as_purchasing(world) as buyer:
        response = buyer.post(
            f"{LIST}/{row.id}/reject", json={"reason": "Factory closed until November"}
        )
    assert response.status_code == 200, response.text
    world.db.commit()

    body = response.json()
    assert body["ack_state"] == ACK_REJECTED
    assert body["rejected_reason"] == "Factory closed until November"
    assert body["rejected_by_name"] and f"{MARKER} Joey" in body["rejected_by_name"]
    world.db.refresh(row)
    assert str(row.rejected_by) == str(world.buyer) and row.rejected_at is not None


def test_a_rejected_row_leaves_committed_v_and_the_plans_own_demand(api):
    _client, world = api
    fixture = _raise_one_row(api)
    row = fixture["row"]

    before = _project_committed(world, planned=False)
    assert before >= Decimal("10"), "the raised row is owed, so the view counts it"

    with _as_purchasing(world) as buyer:
        assert (
            buyer.post(
                f"{LIST}/{row.id}/reject", json={"reason": "Nothing to buy this with"}
            ).status_code
            == 200
        )
    world.db.commit()

    assert _project_committed(world, planned=False) == before - Decimal("10")
    assert _project_committed(world, planned=True) == Decimal("0")


def test_a_rejected_rows_line_goes_back_to_the_board_undecided_with_the_reason(api):
    """AC-H6. The decision is uncovered through `confirm`'s own `uncover_line_ids` seam, so
    the line is planned again from scratch; the cell carries the refusal."""
    _client, world = api
    fixture = _raise_one_row(api)
    line = fixture["line"]
    order = fixture["order"]

    supply = ProjectSupplyService(world.db)
    assert supply.active_decision(str(order.id)) is not None

    with _as_purchasing(world) as buyer:
        assert (
            buyer.post(
                f"{LIST}/{fixture['row'].id}/reject",
                json={"reason": "Factory closed until November"},
            ).status_code
            == 200
        )
    world.db.commit()

    decision = supply.active_decision(str(order.id))
    covered = {
        str(snap.get("project_line_id"))
        for snap in ((decision.line_snapshots if decision else None) or [])
    }
    assert str(line.id) not in covered, "the line is undecided again"

    from app.services.project_fulfilment_board_service import FulfilmentBoardService

    cell = FulfilmentBoardService(world.db)._order_inquiries(
        [str(fixture["core_line"].id)]
    )[str(fixture["core_line"].id)]
    assert cell["ack_state"] == ACK_REJECTED
    assert cell["rejected_reason"] == "Factory closed until November"
    assert f"{MARKER} Joey" in (cell["rejected_by_name"] or "")


def test_the_boards_refusal_flag_hides_once_cs_has_decided_the_line_again(api):
    """The captain, review round: a refusal is a line coming BACK to CS, so it is news
    until they answer it. Once an active revision covers the line again, confirmed after
    the rejection, the cell is about that decision - a flag that outlived the answer reads
    as an open refusal on a line somebody has already dealt with."""
    _client, world = api
    fixture = _raise_one_row(api)
    core_id = str(fixture["core_line"].id)

    with _as_purchasing(world) as buyer:
        assert (
            buyer.post(
                f"{LIST}/{fixture['row'].id}/reject",
                json={"reason": "Factory closed until November"},
            ).status_code
            == 200
        )
    world.db.commit()

    from app.services.project_fulfilment_board_service import FulfilmentBoardService

    def cell_of():
        board = FulfilmentBoardService(world.db)
        frozen, _proposals = {}, None
        decision = ProjectSupplyService(world.db).active_decision(
            str(fixture["order"].id)
        )
        covering = {
            str(snap.get("core_line_id")): decision.confirmed_at
            for snap in ((decision.line_snapshots if decision else None) or [])
            if snap.get("core_line_id")
        }
        frozen = covering
        return board._order_inquiries([core_id], decided_at=frozen)[core_id]

    assert cell_of()["ack_state"] == ACK_REJECTED, "nobody has answered it yet"

    # CS decides the line again. The refused row stays readable; the cell is not about it.
    response = _confirm(
        _client, fixture["order"].id, [_line_payload(fixture["line"].id, buy_qty="10")]
    )
    assert response.status_code == 200, response.text
    world.db.commit()

    answered = cell_of()
    assert answered["ack_state"] != ACK_REJECTED
    assert answered["rejected_reason"] is None
    assert answered["rejected_by_name"] is None
    world.db.refresh(fixture["row"])
    assert fixture["row"].ack_state == ACK_REJECTED, "the refusal itself is still on record"


# ---------------------------------------------------------------------------
# AC-H7 / AC-H8 / AC-H9: change after acknowledgement
# ---------------------------------------------------------------------------


def _settle(world, fixture, *, qty, required_date=NOW):
    """What a planning change does to the line: settle its row in place (part 3)."""
    supply = ProjectSupplyService(world.db)
    from app.schemas.project_supply import ConfirmSupplyBody, ConfirmLine

    body = ConfirmSupplyBody(
        lines=[
            ConfirmLine(
                project_line_id=str(fixture["line"].id),
                timely_spo_qty="0",
                reserve=[],
                borrow=[],
                buy_qty=str(qty),
            )
        ]
    )
    fixture["core_line"].qty_ordered = Decimal(str(qty))
    fixture["core_line"].required_date = required_date
    fixture["line"].qty = Decimal(str(qty))
    fixture["line"].delivery_date = required_date
    world.db.flush()
    result = supply.confirm(
        fixture["order"],
        body,
        actor_user_id=world.cs_user,
        settle_in_place_line_ids=[str(fixture["line"].id)],
    )
    world.db.commit()
    return result


def test_an_amend_before_acknowledgement_leaves_the_row_awaiting(api):
    """AC-H7. CS is free to change what nobody has read; the row says nothing about it."""
    _client, world = api
    fixture = _raise_one_row(api)
    row = fixture["row"]

    _settle(world, fixture, qty="25")

    world.db.refresh(row)
    assert row.ack_state == ACK_AWAITING
    assert row.changed_at is None
    assert Decimal(str(row.qty)) == Decimal("25")


def test_an_amend_after_acknowledgement_reads_changed_and_keeps_its_links(api):
    """AC-H8. The row is updated in place with the previous value, and purchasing sees it
    as a change rather than as something it has already dealt with."""
    _client, world = api
    _open_po_line(world, qty=50)
    fixture = _raise_one_row(api)
    row = fixture["row"]

    with _as_purchasing(world) as buyer:
        assert buyer.post(ACK_URL, json={"row_ids": [str(row.id)]}).status_code == 200
    world.db.commit()
    assert _links_of(world, row), "acknowledging linked it"

    _settle(world, fixture, qty="25")

    world.db.refresh(row)
    assert row.ack_state == ACK_CHANGED
    assert row.changed_at is not None
    assert Decimal(str(row.qty)) == Decimal("25")
    assert _links_of(world, row), "a change keeps what the buyer already arranged"
    assert "Was 10" in (row.note or ""), "the previous value travels with the row"
    # As FIGURES, beside the sentence. The screen prints the Was / Now table off these;
    # it used to parse them back out of the note, where "Was 10, no previous delivery
    # date" gave up the quantity as `10,`.
    assert Decimal(str(row.previous_qty)) == Decimal("10")
    assert row.previous_delivery_date == WAS


def test_the_previous_value_reaches_the_wire_as_two_figures(api):
    """AC-H14 for the Was half: the list states `previous_qty` / `previous_delivery_date`,
    so nothing downstream has to read the note's prose to draw the change."""
    _client, world = api
    fixture = _raise_one_row(api)
    row = fixture["row"]

    with _as_purchasing(world) as buyer:
        assert buyer.post(ACK_URL, json={"row_ids": [str(row.id)]}).status_code == 200
    world.db.commit()
    _settle(world, fixture, qty="25")

    listed = _client.get(LIST, params={"query": str(fixture["core_so"].so_number)})
    assert listed.status_code == 200, listed.text
    wire = next(
        entry for entry in listed.json()["data"] if entry["id"] == str(row.id)
    )
    assert wire["ack_state"] == ACK_CHANGED
    assert wire["previous_qty"] == "10"
    assert wire["previous_delivery_date"] == WAS.isoformat()
    assert wire["qty"] == "25", "and the Now half is the row's own quantity"


def test_a_row_nobody_amended_states_no_previous_value(api):
    """No guess and no zero: a row that has never been settled says nothing was."""
    _client, world = api
    fixture = _raise_one_row(api)
    assert fixture["row"].previous_qty is None
    assert fixture["row"].previous_delivery_date is None


def test_re_acknowledging_a_changed_row_returns_it_and_links_the_remainder(api):
    _client, world = api
    _open_po_line(world, qty=30)
    fixture = _raise_one_row(api)
    row = fixture["row"]

    with _as_purchasing(world) as buyer:
        assert buyer.post(ACK_URL, json={"row_ids": [str(row.id)]}).status_code == 200
        world.db.commit()
        _settle(world, fixture, qty="25")
        # A second purchase order arrives for what the change added.
        _open_po_line(world, qty=40)
        response = buyer.post(ACK_URL, json={"row_ids": [str(row.id)]})
    assert response.status_code == 200, response.text
    world.db.commit()

    world.db.refresh(row)
    assert row.ack_state == ACK_ACKNOWLEDGED
    assert sum(Decimal(str(link.qty)) for link in _links_of(world, row)) == Decimal("25")


def test_a_supersede_of_an_acknowledged_row_raises_its_replacement_changed(api):
    """AC-H9. A reconfirm that cannot settle in place still owes purchasing the fact that
    this line is one they had already taken on."""
    _client, world = api
    fixture = _raise_one_row(api)
    row = fixture["row"]

    with _as_purchasing(world) as buyer:
        assert buyer.post(ACK_URL, json={"row_ids": [str(row.id)]}).status_code == 200
    world.db.commit()

    # A plain reconfirm of the same line: no settle-in-place seam, so the row is
    # superseded and a fresh one is raised in its place.
    response = _confirm(
        _client, fixture["order"].id, [_line_payload(fixture["line"].id, buy_qty="10")]
    )
    assert response.status_code == 200, response.text
    world.db.commit()

    world.db.refresh(row)
    assert row.state == INQUIRY_CANCELLED, "the old row was superseded"
    replacement = _order_row(world, fixture["line"])
    assert str(replacement.id) != str(row.id)
    assert replacement.ack_state == ACK_CHANGED


def _raise_two_rows(api, *, first_qty="10", second_qty="6"):
    """One order, TWO lines, both confirmed wholly as Buy: two raised inquiry rows.

    What the single-line fixture cannot express: a confirmation that names ONE line of an
    order carries the other, and a carried line's row is cancelled and re-raised under the
    new revision (13.4). That is the seam the handshake has to survive.
    """
    client, world = api
    db = world.db
    core_so = _core_so(db, world.company_id)
    first_core = _core_line(
        db, core_so, world.product, world.warehouse, qty_ordered=first_qty,
        required_date=WAS,
    )
    second_core = _core_line(
        db, core_so, world.product, world.warehouse, qty_ordered=second_qty,
        required_date=WAS,
    )
    order = _project_so(
        db, world.project, so_id=core_so.id, autocount_doc_no=core_so.so_number
    )
    first = _project_line(db, order, line_no=1, product=world.product, core_line=first_core)
    second = _project_line(db, order, line_no=2, product=world.product, core_line=second_core)
    db.commit()

    response = _confirm(
        client,
        order.id,
        [
            _line_payload(first.id, buy_qty=first_qty),
            _line_payload(second.id, buy_qty=second_qty),
        ],
    )
    assert response.status_code == 200, response.text
    db.commit()
    return {
        "order": order,
        "first": {"line": first, "core_line": first_core, "row": _order_row(world, first)},
        "second": {"line": second, "core_line": second_core, "row": _order_row(world, second)},
    }


def test_confirming_one_line_leaves_the_other_lines_acknowledgement_alone(api):
    """B1 (review). Both rows are acknowledged; CS then re-confirms line 2 alone.

    Line 2 is a change and reads as one. Line 1 was CARRIED - this confirmation said
    nothing about it - so its row must come out the far side still Acknowledged, with the
    same person and the same time on it. It did not: the carry cancels and re-raises the
    row, and the re-raise read the acknowledgement off rows it had just cancelled and
    promoted every one of them to `changed` - so every confirm of any other line of the
    order told the buyer that a row they had read had moved, with no Was and no Now,
    because nothing had.
    """
    _client, world = api
    fixture = _raise_two_rows(api)
    first_row, second_row = fixture["first"]["row"], fixture["second"]["row"]

    with _as_purchasing(world) as buyer:
        response = buyer.post(
            ACK_URL, json={"row_ids": [str(first_row.id), str(second_row.id)]}
        )
    assert response.status_code == 200, response.text
    world.db.commit()
    world.db.refresh(first_row)
    stamped_by, stamped_at = first_row.acknowledged_by, first_row.acknowledged_at
    assert stamped_by and stamped_at

    # CS confirms line 2 again, and names nothing else. (The same quantity: what makes
    # line 2 a change here is that this revision restates it, not the figure.)
    response = _confirm(
        _client, fixture["order"].id, [_line_payload(fixture["second"]["line"].id, buy_qty="6")]
    )
    assert response.status_code == 200, response.text
    world.db.commit()

    carried = _order_row(world, fixture["first"]["line"])
    assert carried.ack_state == ACK_ACKNOWLEDGED, (
        "the carried line was not changed by a confirmation that never named it"
    )
    assert str(carried.acknowledged_by) == str(stamped_by)
    assert carried.acknowledged_at == stamped_at
    assert carried.changed_at is None

    amended = _order_row(world, fixture["second"]["line"])
    assert amended.ack_state == ACK_CHANGED, "the line that DID change still says so"
    assert amended.changed_at is not None


def test_a_carried_line_that_nobody_acknowledged_stays_awaiting(api):
    """The other half of the same rule: a carry says nothing about a row nobody read."""
    _client, world = api
    fixture = _raise_two_rows(api)

    response = _confirm(
        _client, fixture["order"].id, [_line_payload(fixture["second"]["line"].id, buy_qty="6")]
    )
    assert response.status_code == 200, response.text
    world.db.commit()

    carried = _order_row(world, fixture["first"]["line"])
    assert carried.ack_state == ACK_AWAITING
    assert carried.acknowledged_by is None and carried.changed_at is None


def test_a_rejected_line_re_decided_raises_a_fresh_awaiting_row(api):
    """AC-H6's last clause, pinned against the inheritance rule above: what CS raises
    after a refusal is a NEW instruction nobody has read, never the refused one's state."""
    _client, world = api
    fixture = _raise_one_row(api)
    row = fixture["row"]

    with _as_purchasing(world) as buyer:
        assert buyer.post(ACK_URL, json={"row_ids": [str(row.id)]}).status_code == 200
        world.db.commit()
        reject = buyer.post(
            f"{LIST}/{row.id}/reject", json={"reason": "Factory closed until November"}
        )
        assert reject.status_code == 200, reject.text
    world.db.commit()

    response = _confirm(
        _client, fixture["order"].id, [_line_payload(fixture["line"].id, buy_qty="10")]
    )
    assert response.status_code == 200, response.text
    world.db.commit()

    fresh = _order_row(world, fixture["line"])
    assert str(fresh.id) != str(row.id)
    assert fresh.ack_state == ACK_AWAITING
    assert fresh.acknowledged_by is None and fresh.changed_at is None
    world.db.refresh(row)
    assert row.ack_state == ACK_REJECTED, "the refusal itself stays readable"


# ---------------------------------------------------------------------------
# AC-H10: what the plan counts
# ---------------------------------------------------------------------------


def test_the_plan_counts_acknowledged_and_changed_rows_only(api):
    _client, world = api
    fixture = _raise_one_row(api)
    row = fixture["row"]

    assert _project_committed(world, planned=True) == Decimal("0"), (
        "an awaiting row is not something to buy against"
    )
    assert _project_committed(world, planned=False) >= Decimal("10")

    with _as_purchasing(world) as buyer:
        assert buyer.post(ACK_URL, json={"row_ids": [str(row.id)]}).status_code == 200
    world.db.commit()

    assert _project_committed(world, planned=True) == Decimal("10")


def test_the_awaiting_count_is_reported_for_the_plan_page_chip(api):
    from app.services.scm.reorder_run_service import awaiting_acknowledgement_rows

    _client, world = api
    before = awaiting_acknowledgement_rows(world.db)
    fixture = _raise_one_row(api)
    assert awaiting_acknowledgement_rows(world.db) == before + 1

    with _as_purchasing(world) as buyer:
        assert (
            buyer.post(ACK_URL, json={"row_ids": [str(fixture["row"].id)]}).status_code
            == 200
        )
    world.db.commit()
    assert awaiting_acknowledgement_rows(world.db) == before


# ---------------------------------------------------------------------------
# AC-H13: link now
# ---------------------------------------------------------------------------


def test_link_now_links_acknowledged_rows_and_leaves_awaiting_ones_alone(api):
    _client, world = api
    acknowledged = _raise_one_row(api, qty="10")
    awaiting = _raise_one_row(api, qty="10")

    with _as_purchasing(world) as buyer:
        assert (
            buyer.post(
                ACK_URL, json={"row_ids": [str(acknowledged["row"].id)]}
            ).status_code
            == 200
        )
        world.db.commit()
        # Only now does the purchase order arrive, so the acknowledge itself linked
        # nothing and Link now is what has work to do.
        _open_po_line(world, qty=100)
        response = buyer.post(LINK_NOW, json={"product_ids": [str(world.product.id)]})
    assert response.status_code == 200, response.text
    world.db.commit()

    assert response.json()["placed_rows"] == 1
    assert _links_of(world, acknowledged["row"])
    assert _links_of(world, awaiting["row"]) == []


def test_link_now_is_scoped_to_the_products_it_is_given(api):
    _client, world = api
    other_product = _product(world.db)
    world.db.commit()
    mine = _raise_one_row(api, qty="10")
    theirs = _raise_one_row(api, qty="10", product=other_product)
    _open_po_line(world, qty=100)
    _open_po_line(world, qty=100, product=other_product)

    with _as_purchasing(world) as buyer:
        assert (
            buyer.post(
                ACK_URL,
                json={"row_ids": [str(mine["row"].id), str(theirs["row"].id)]},
            ).status_code
            == 200
        )
        world.db.commit()
        for row in (mine["row"], theirs["row"]):
            ProjectOrderInquiryService(world.db).unplace(
                str(row.id), actor_user_id=world.buyer
            )
        world.db.commit()
        response = buyer.post(LINK_NOW, json={"product_ids": [str(world.product.id)]})
    assert response.status_code == 200, response.text
    world.db.commit()

    assert _links_of(world, mine["row"])
    assert _links_of(world, theirs["row"]) == []


# ---------------------------------------------------------------------------
# AC-H14: on the wire
# ---------------------------------------------------------------------------


def test_every_new_field_reaches_the_list_the_row_and_the_export(api):
    client, world = api
    fixture = _raise_one_row(api)
    row = fixture["row"]

    with _as_purchasing(world) as buyer:
        assert (
            buyer.post(
                f"{LIST}/{row.id}/reject", json={"reason": "Factory closed"}
            ).status_code
            == 200
        )
    world.db.commit()

    listed = client.get(LIST, params={"ack": "rejected", "limit": 200}).json()
    on_wire = next(item for item in listed["data"] if item["id"] == str(row.id))
    for field in (
        "ack_state",
        "acknowledged_by_name",
        "acknowledged_at",
        "rejected_by_name",
        "rejected_at",
        "rejected_reason",
        "changed_at",
    ):
        assert field in on_wire, f"`response_model` dropped {field}"
    assert on_wire["ack_state"] == ACK_REJECTED
    assert on_wire["rejected_reason"] == "Factory closed"

    export = client.get(f"{LIST}/export", params={"ack": "rejected"})
    assert export.status_code == 200
    import openpyxl

    book = openpyxl.load_workbook(io.BytesIO(export.content))
    sheet = book[book.sheetnames[0]]
    headings = [cell.value for cell in sheet[2]]
    assert "ACKNOWLEDGED" in headings
    printed = [
        row_values[headings.index("ACKNOWLEDGED")]
        for row_values in sheet.iter_rows(min_row=3, values_only=True)
    ]
    assert any("Factory closed" in (value or "") for value in printed)


def test_the_sales_order_detail_carries_the_handshake(api):
    """The per-project / SO-detail serializer is a second reader of the same row, and it
    drops a field just as silently."""
    client, world = api
    fixture = _raise_one_row(api)

    body = client.get(f"{BASE}/sales-orders/{fixture['order'].id}/order-inquiry").json()
    row = next(item for item in body["rows"] if item["id"] == str(fixture["row"].id))
    assert row["ack_state"] == ACK_AWAITING
    assert "rejected_reason" in row and "changed_at" in row


# ---------------------------------------------------------------------------
# AC-LH1 / AC-LH2 / AC-LH4: the link horizon
#
# `PLAN-scm-oi-handshake.md` section 11. Every path that ties a document to a row takes a
# date to link up to; a row due after it is left Not linked and counted, so a 2030 order
# stops eating a purchase order a nearer one needed.
# ---------------------------------------------------------------------------

HORIZON = date(2026, 12, 31)
BEFORE_BOTH = date(2026, 9, 1)
NEAR = date(2026, 10, 1)
FAR = date(2030, 1, 1)

AUTO_PLACE = f"{LIST}/auto-place"
SUMMARY = f"{LIST}/summary"


def _due(world, row, when):
    """The row's delivery date, stated rather than inherited: the horizon is read off it
    and the fixture's own required date says nothing about 2030."""
    row.delivery_date = when
    world.db.flush()
    world.db.commit()
    return row


def _taken_off(world, line) -> Decimal:
    """What every link together claims off one purchase-order line."""
    by_po, _by_spo = ProjectOrderInquiryService(world.db)._linked_by_target()
    return by_po.get(str(line.id), Decimal("0"))


def test_acknowledge_leaves_a_row_due_after_the_horizon_not_linked(api):
    """AC-LH1. Two acknowledged rows of one product, one open purchase-order line of 100,
    and a horizon between them: the near row links, the far one stays Not linked and the
    line keeps the rest of its quantity for whoever needs it sooner."""
    _client, world = api
    _po, line = _open_po_line(world, qty=100)
    near = _due(world, _raise_one_row(api, qty="10")["row"], NEAR)
    far = _due(world, _raise_one_row(api, qty="10")["row"], FAR)

    with _as_purchasing(world) as buyer:
        response = buyer.post(
            ACK_URL,
            json={
                "row_ids": [str(near.id), str(far.id)],
                "link_up_to": HORIZON.isoformat(),
            },
        )
    assert response.status_code == 200, response.text
    world.db.commit()

    body = response.json()
    assert body["acknowledged"] == 2
    assert body["linked_rows"] == 1
    assert body["after_horizon"] == 1
    assert body["link_up_to"] == HORIZON.isoformat()
    assert sum(Decimal(str(link.qty)) for link in _links_of(world, near)) == Decimal("10")
    assert _links_of(world, far) == [], "the 2030 row took a purchase order it is not due on"
    assert _taken_off(world, line) == Decimal("10"), "the line kept its remainder"


def test_link_selected_says_how_many_it_left_after_the_horizon(api):
    """AC-LH2. The same two rows, acknowledged under a horizon that reaches neither, then
    Link selected under one that reaches the near one: the banner reads "1 linked, 1 after
    <date>", which is the pair of numbers this response carries."""
    _client, world = api
    _open_po_line(world, qty=100)
    near = _due(world, _raise_one_row(api, qty="10")["row"], NEAR)
    far = _due(world, _raise_one_row(api, qty="10")["row"], FAR)

    with _as_purchasing(world) as buyer:
        taken_on = buyer.post(
            ACK_URL,
            json={
                "row_ids": [str(near.id), str(far.id)],
                "link_up_to": BEFORE_BOTH.isoformat(),
            },
        )
        assert taken_on.status_code == 200, taken_on.text
        assert taken_on.json()["linked_rows"] == 0
        world.db.commit()

        response = buyer.post(
            AUTO_PLACE,
            json={
                "row_ids": [str(near.id), str(far.id)],
                "link_up_to": HORIZON.isoformat(),
            },
        )
    assert response.status_code == 200, response.text
    world.db.commit()

    body = response.json()
    assert body["placed_rows"] == 1
    assert body["after_horizon"] == 1
    assert body["link_up_to"] == HORIZON.isoformat()
    assert _links_of(world, far) == []


def test_a_row_with_no_delivery_date_is_inside_the_horizon(api):
    """AC-LH4. A blank date is not a far one: the row is still owed, nobody has said when,
    and refusing it a document would leave it unbought for a date nobody stated."""
    _client, world = api
    _open_po_line(world, qty=100)
    undated = _due(world, _raise_one_row(api, qty="10")["row"], None)

    with _as_purchasing(world) as buyer:
        response = buyer.post(
            ACK_URL,
            json={"row_ids": [str(undated.id)], "link_up_to": BEFORE_BOTH.isoformat()},
        )
    assert response.status_code == 200, response.text
    world.db.commit()

    assert response.json()["after_horizon"] == 0
    assert sum(
        Decimal(str(link.qty)) for link in _links_of(world, undated)
    ) == Decimal("10")


def _plan_until(world, when):
    """The "Plan until" of the latest completed reorder run - which IS the link horizon's
    default (S2, code review 27 Aug 2026).

    NOT `scm.priority_policy.reorder_coverage_until`, which this used to write. That field
    is the ladder's BUY-NOW line - a row needed after it is one the engine proposes buying
    - so using it as the link horizon meant the purchase order raised for those very rows
    could never be linked back to them. The run's own horizon is the date its netting
    stopped at, which is the honest answer to "how far out has anybody planned".
    """
    world.db.execute(
        text("UPDATE scm.reorder_run SET plan_horizon_date = :d WHERE id = :i"),
        {"d": when, "i": world.plan_run},
    )
    world.db.flush()


def test_the_summary_offers_the_plans_own_horizon_as_the_default(api):
    """AC-LH5's own half of the contract: the page does not invent a default date, it reads
    the reorder plan's own "Plan until". One setting, so the plan and the buyer cannot be
    working to two different horizons."""
    client, world = api
    _plan_until(world, HORIZON)
    world.db.commit()

    body = client.get(SUMMARY).json()

    assert body["link_up_to_default"] == HORIZON.isoformat(), (
        "`response_model` dropped link_up_to_default"
    )


def test_an_explicit_no_horizon_links_a_row_the_plan_does_not_reach(api):
    """S1 (code review, 27 Aug 2026). Clearing the date on the page means "no horizon", and
    that is a different instruction from naming none.

    Both used to travel as the same nothing: an empty box sent no `link_up_to`, the server
    read that as "the caller named none" and used the plan's own date - so once a plan run
    named a horizon, the page could not link a far-future row at ALL. The buyer could see
    the box empty and still be refused.
    """
    _client, world = api
    _plan_until(world, HORIZON)
    _open_po_line(world, qty=100)
    far = _due(world, _raise_one_row(api, qty="10")["row"], FAR)

    with _as_purchasing(world) as buyer:
        response = buyer.post(
            ACK_URL,
            json={"row_ids": [str(far.id)], "link_horizon": "none"},
        )
    assert response.status_code == 200, response.text
    world.db.commit()

    body = response.json()
    assert body["linked_rows"] == 1
    assert body["after_horizon"] == 0
    assert body["link_up_to"] is None
    assert body["link_horizon"] == "none", "`response_model` dropped link_horizon"
    assert sum(Decimal(str(link.qty)) for link in _links_of(world, far)) == Decimal("10")


def test_a_result_states_which_horizon_it_ran_under(api):
    """The other half of the same field: a pass that DID run to a date says so, so the FE
    never has to read a null `link_up_to` two ways."""
    _client, world = api
    _open_po_line(world, qty=100)
    near = _due(world, _raise_one_row(api, qty="10")["row"], NEAR)

    with _as_purchasing(world) as buyer:
        response = buyer.post(
            ACK_URL,
            json={"row_ids": [str(near.id)], "link_up_to": HORIZON.isoformat()},
        )
    assert response.status_code == 200, response.text

    body = response.json()
    assert body["link_horizon"] == "date"
    assert body["link_up_to"] == HORIZON.isoformat()


def test_link_selected_carries_the_explicit_no_horizon_too(api):
    """One rule, every door: the worklist's own press says the same thing about the same
    box as Acknowledge does."""
    _client, world = api
    _plan_until(world, HORIZON)
    _open_po_line(world, qty=100)
    far = _due(world, _raise_one_row(api, qty="10")["row"], FAR)

    with _as_purchasing(world) as buyer:
        taken_on = buyer.post(
            ACK_URL,
            json={"row_ids": [str(far.id)], "link_up_to": BEFORE_BOTH.isoformat()},
        )
        assert taken_on.json()["linked_rows"] == 0
        world.db.commit()

        response = buyer.post(
            AUTO_PLACE, json={"row_ids": [str(far.id)], "link_horizon": "none"}
        )
    assert response.status_code == 200, response.text
    world.db.commit()

    body = response.json()
    assert body["placed_rows"] == 1
    assert body["link_horizon"] == "none"
    assert body["link_up_to"] is None


def test_naming_a_date_horizon_without_a_date_is_refused(api):
    """`link_horizon: "date"` and no date is a caller contradicting itself. Refused rather
    than quietly read as one of the other two, which would link under a horizon nobody
    asked for."""
    _client, world = api
    row = _raise_one_row(api, qty="10")["row"]

    with _as_purchasing(world) as buyer:
        response = buyer.post(
            ACK_URL, json={"row_ids": [str(row.id)], "link_horizon": "date"}
        )

    assert response.status_code == 422, response.text


def test_a_caller_that_names_no_horizon_takes_the_plans_own(api):
    """The default is not "no horizon": a press that says nothing is a press under the
    plan's own horizon, which is what a purchase-order confirm uses too."""
    _client, world = api
    _plan_until(world, HORIZON)
    _open_po_line(world, qty=100)
    far = _due(world, _raise_one_row(api, qty="10")["row"], FAR)

    with _as_purchasing(world) as buyer:
        response = buyer.post(ACK_URL, json={"row_ids": [str(far.id)]})
    assert response.status_code == 200, response.text
    world.db.commit()

    body = response.json()
    assert body["linked_rows"] == 0
    assert body["after_horizon"] == 1
    assert body["link_up_to"] == HORIZON.isoformat()
    assert _links_of(world, far) == []
