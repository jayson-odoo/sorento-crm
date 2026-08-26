"""The export task itself (AC-A8), without a live worker.

RQ workers are SHARED across worktrees on this machine - a sibling checkout's worker will
happily steal a job enqueued here and run it against its own code. So the task function is
called directly, with the worker's own ``SessionLocal`` swapped for the test session and a
fake bucket in place of storage. What is proved here is what the drawer depends on: bytes
uploaded under a key namespaced by the download id, and a row that ends in 'ready' or in
'failed' with something readable - never one left spinning in 'processing'.

Run: pytest tests/test_report_export_task.py -q
"""
from __future__ import annotations

import uuid
from contextlib import contextmanager

import pytest

from tests import _report_fixture as fixture
from tests._pg_fixture import blank_session

KEY = "zzt_orders"
_USER_ID = str(uuid.uuid4())


class _FakeBackend:
    def __init__(self):
        self.uploads: list[tuple[str, bytes, str | None]] = []

    def upload_file(self, file_content, file_path, content_type=None, **_kwargs):
        self.uploads.append((file_path, file_content, content_type))
        return file_path, f"https://cdn.zzt.test/{file_path}"


@contextmanager
def _task_env(db, backend):
    """The task opens its OWN session (it runs in the worker); swap it and neuter close."""
    from unittest.mock import patch

    from app.tasks import report_export_tasks as tasks

    with patch.object(tasks, "SessionLocal", lambda: db), patch.object(
        tasks, "get_backend", lambda _provider: backend
    ), patch.object(tasks, "default_provider", lambda: "s3"), patch.object(
        db, "close", lambda: None
    ):
        yield tasks


@pytest.fixture
def db():
    from app.models.user import User

    with blank_session() as session:
        fixture.create_table(session)
        session.add(User(id=_USER_ID, email="exporter@zzt.test", name="Exporter", status="ACTIVE"))
        session.flush()
        yield session


@pytest.fixture
def registered():
    from app.services.reports import registry as reg

    definition = reg.register(fixture.definition())
    try:
        yield definition
    finally:
        reg._REGISTRY.pop(KEY, None)


@pytest.fixture
def download(db):
    from app.services.download_service import DownloadService

    return DownloadService(db).create(
        user_id=_USER_ID, kind="report_xlsx", filename="Scratch orders-2026.xlsx"
    )


def _params():
    return dict(fixture.definition().default_view["params"])


def test_the_task_renders_uploads_and_marks_the_download_ready(db, registered, download):
    from app.models.download import DownloadStatus, UserDownload

    backend = _FakeBackend()
    with _task_env(db, backend) as tasks:
        result = tasks.generate_report_xlsx(str(download.id), KEY, _params(), None, _USER_ID)

    assert result["status"] == "ready", result
    key, content, content_type = backend.uploads[0]
    assert key == f"exports/report-xlsx/{download.id}/Scratch orders-2026.xlsx"
    assert content[:4] == b"PK\x03\x04"  # a zip container, which an .xlsx is
    assert "spreadsheetml" in content_type

    row = db.query(UserDownload).filter(UserDownload.id == str(download.id)).one()
    assert row.status == DownloadStatus.READY.value
    assert row.storage_key == key
    assert row.storage_provider == "s3"
    assert row.filename == "Scratch orders-2026.xlsx"
    assert row.error is None


def test_the_task_honours_the_view_it_was_given(db, registered, download):
    from openpyxl import load_workbook
    from io import BytesIO

    view = {
        "params": _params(),
        "detail": {"columns": ["order_no", "amount"], "order": []},
        "pivot": {"rows": "region", "cols": "month", "measures": ["fee"]},
    }
    backend = _FakeBackend()
    with _task_env(db, backend) as tasks:
        assert tasks.generate_report_xlsx(
            str(download.id), KEY, _params(), view, _USER_ID
        )["status"] == "ready"

    workbook = load_workbook(BytesIO(backend.uploads[0][1]))
    assert workbook.sheetnames[0] == "SUMMARY"
    assert workbook.sheetnames[1:] == [
        "JAN'26", "FEB'26", "MAR'26", "APR'26", "MAY'26", "JUN'26",
        "JUL'26", "AUG'26", "SEP'26", "OCT'26", "NOV'26", "DEC'26",
    ]
    # Single-level headers merge up into the group row (AC-G4/G8).
    assert [c.value for c in workbook["JAN'26"][6]] == ["ORDER NO", "AMOUNT"]
    assert workbook["SUMMARY"]["A6"].value == "REGION"


def test_the_export_path_ignores_the_sync_caps(db, registered, download, monkeypatch):
    from app.services.reports import engine

    monkeypatch.setattr(engine, "DETAIL_ROW_CAP", 1)
    monkeypatch.setattr(engine, "PIVOT_CELL_CAP", 1)
    backend = _FakeBackend()
    with _task_env(db, backend) as tasks:
        result = tasks.generate_report_xlsx(str(download.id), KEY, _params(), None, _USER_ID)

    assert result["status"] == "ready", result


def test_an_unknown_report_fails_the_row_rather_than_the_queue(db, download):
    from app.models.download import DownloadStatus, UserDownload

    backend = _FakeBackend()
    with _task_env(db, backend) as tasks:
        result = tasks.generate_report_xlsx(str(download.id), "zzt_gone", _params(), None, _USER_ID)

    assert result["status"] == "failed"
    row = db.query(UserDownload).filter(UserDownload.id == str(download.id)).one()
    assert row.status == DownloadStatus.FAILED.value
    assert "zzt_gone" in (row.error or "")
    assert backend.uploads == []
