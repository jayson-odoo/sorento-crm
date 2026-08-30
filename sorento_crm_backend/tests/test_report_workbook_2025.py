"""AC-D5: the exported workbook against the client's own 2025 register, cell for cell.

This is the whole point of slice S5's fixture. The client keeps `SPONSORSHIP REPORT
JAN-Dec'25.xlsx` by hand; its twelve GRAND TOTAL cells and its SUMMARY C28 / C29 are numbers
they already trust. So: load those 214 forms into a blank schema through the loader itself,
run the report over 2025, render the export, and compare our GRAND TOTALs to theirs.

The rows are loaded HERE, into the test's own scratch schema, never read from the developer
copy that `scripts/dev/load_sponsorship_2025_fixture.py --apply` writes to. CI has no data
and the local database is a copy of production; a test that leaned on either would pass for
the wrong reason today and fail for no reason tomorrow.

Run: pytest tests/test_report_workbook_2025.py -q
"""
from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest

from tests._pg_fixture import blank_session

WORKBOOK = Path(__file__).resolve().parent / "fixtures" / "sponsorship_2025.xlsx"

SUMMARY = "SUMMARY"
MONTHS = [
    "JAN'25", "FEB'25", "MAR'25", "APR'25", "MAY'25", "JUN'25",
    "JUL'25", "AUG'25", "SEP'25", "OCT'25", "NOV'25", "DEC'25",
]

#: The client's own SUMMARY sheet, cells C28 and C29.
CLIENT_YEAR_PROJECT_VALUE = 257076027.91
CLIENT_YEAR_SAMPLE_PRICE = 518605.38

_GROUP_ROW = 6
_HEADER_ROW = 7
_FIRST_DATA_ROW = 8

_PARAMS = {
    "date_basis": "approved_at",
    "period": {"kind": "year", "year": 2025},
    "sales_agent": [],
    "status": ["approved", "processed_by_cs"],
}


@pytest.fixture(scope="module")
def sheets():
    """The client's workbook as plain dicts, read by the loader's own parser."""
    from scripts.dev.load_sponsorship_2025_fixture import parse_workbook

    return parse_workbook(WORKBOOK)


@pytest.fixture(scope="module")
def exported(sheets):
    """The 214 fixture forms, the report over 2025, and the workbook it exports."""
    from openpyxl import load_workbook

    from app.services.company_scope import set_company_scope
    from app.services.reports import engine
    from app.services.reports.definitions.sponsorship import REPORT
    from app.services.reports.xlsx_renderer import render_workbook
    from scripts.dev.load_sponsorship_2025_fixture import load

    with blank_session() as db:
        set_company_scope(db, None)
        result = load(db, sheets)
        assert result["inserted"] == 214, result

        data = engine.run_workbook(db, REPORT, _PARAMS)
        content = render_workbook(REPORT, data)

    return load_workbook(BytesIO(content))


def _headers(sheet) -> list:
    """The header of every column, wherever it sits: a single-level header is merged up
    into the group row (AC-G4), a tick column's year stays on the leaf row."""
    return [
        sheet.cell(row=_GROUP_ROW, column=index).value
        if sheet.cell(row=_HEADER_ROW, column=index).value is None
        else sheet.cell(row=_HEADER_ROW, column=index).value
        for index in range(1, sheet.max_column + 1)
    ]


def _column(sheet, header: str) -> int:
    return _headers(sheet).index(header) + 1


def _grand_total(sheet, label: str) -> float:
    """The GRAND TOTAL row's cell under a named column, as a float."""
    value = sheet.cell(row=sheet.max_row, column=_column(sheet, label)).value
    return _money(value)


def _money(value) -> float:
    return 0.0 if value in (None, "-") else round(float(value), 2)


# ------------------------------------------------------------------ shape (AC-D1/D2)


def test_the_export_is_summary_then_the_twelve_months_of_2025(exported):
    assert exported.sheetnames == [SUMMARY] + MONTHS


def test_every_sheet_opens_with_the_clients_title_block(exported):
    """AC-D2 / AC-G7 / AC-G11: the client's own four lines, in the client's own cells."""
    from datetime import date, datetime

    for name in [SUMMARY] + MONTHS:
        sheet = exported[name]
        assert sheet["A1"].value is None
        assert sheet["A2"].value == "SORENTO SDN BHD"
        assert sheet["A3"].value == "SPONSORSHIP"
        assert sheet["A5"].value == "DEPARTMENT:"
        assert sheet["C5"].value == "PROJECT SALES"

    assert exported[SUMMARY]["A4"].value == "Jan'25 to Dec'25"
    january = exported["JAN'25"]["A4"]
    assert isinstance(january.value, (date, datetime))
    assert (january.value.year, january.value.month) == (2025, 1)
    assert january.number_format == "mmm-yy"
    assert exported["DEC'25"]["A4"].value.month == 12


def test_a_monthly_sheet_carries_the_clients_own_header_row(exported):
    """AC-G8 / AC-G11. A..G are the client's seven columns, in their words and their
    order; H..K are the four delivery years, under one merged header, exactly as their
    own H6:K6 band prints them (empty on all 214 rows of 2025, and still four columns)."""
    sheet = exported["JAN'25"]

    assert _headers(sheet) == [
        "PS NO:",
        "SALES AGENT",
        "CUSTOMER NAME",
        "PROJECT TITLE",
        "SPONSHER PROJECT",
        "PROJECT VALUE",
        "SAMPLE PRICE",
        2025,
        2026,
        2027,
        2028,
    ]
    assert sheet["H6"].value == "EXPECTED YEAR OF DELIVERY"
    assert "H6:K6" in [str(r) for r in sheet.merged_cells.ranges]
    # Single-level headers are merged DOWN over both header rows, so the client's empty
    # band above every ungrouped column is not reproduced (AC-G4).
    assert "A6:A7" in [str(r) for r in sheet.merged_cells.ranges]
    assert sheet["A7"].value is None


def test_the_grand_total_money_sits_where_the_clients_own_sheet_puts_it(exported):
    """AC-G11. The client's JAN'25 GRAND TOTAL is F25 / G25 - the sixth and seventh
    columns, under PROJECT VALUE and SAMPLE PRICE, labelled to their left."""
    sheet = exported["JAN'25"]
    total_row = sheet.max_row

    assert _column(sheet, "PROJECT VALUE") == 6
    assert _column(sheet, "SAMPLE PRICE") == 7
    assert sheet.cell(row=total_row, column=5).value == "GRAND TOTAL"
    assert sheet.cell(row=total_row, column=5).font.bold
    assert sheet.cell(row=total_row, column=6).value is not None
    assert sheet.cell(row=total_row, column=7).value is not None


def test_each_sheet_holds_the_same_number_of_rows_the_client_typed(exported, sheets):
    for sheet, name in zip(sheets, MONTHS):
        rendered = exported[name]
        rows = rendered.max_row - _FIRST_DATA_ROW  # the GRAND TOTAL line is the last row
        assert rows == len(sheet["rows"]), name


# ------------------------------------------------------------------ numbers (AC-D5)


def test_each_monthly_grand_total_equals_the_clients_sheet(exported, sheets):
    """The twelve numbers the client checks the register against, to the cent."""
    for sheet, name in zip(sheets, MONTHS):
        rendered = exported[name]
        assert _grand_total(rendered, "PROJECT VALUE") == _money(
            sheet["sheet_total_project_value"]
        ), name
        assert _grand_total(rendered, "SAMPLE PRICE") == _money(
            sheet["sheet_total_sample_price"]
        ), name


def _labelled_row(sheet, label: str) -> int:
    for row in range(_FIRST_DATA_ROW, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == label:
            return row
    raise AssertionError(f"no row labelled {label!r}")


def test_the_summary_grand_total_equals_the_clients_c28_and_c29(exported):
    """AC-G10 / AC-G11. The client's own C28 and C29, under the same two labels, in the
    same column - the two numbers the register is kept for."""
    sheet = exported[SUMMARY]

    project_row = _labelled_row(sheet, "GRAND TOTAL PROJECT VALUE JAN-DEC'25")
    sample_row = _labelled_row(sheet, "GRAND TOTAL SAMPLE PRICE JAN-DEC'25")
    assert round(float(sheet.cell(row=project_row, column=3).value), 2) == CLIENT_YEAR_PROJECT_VALUE
    assert round(float(sheet.cell(row=sample_row, column=3).value), 2) == CLIENT_YEAR_SAMPLE_PRICE

    total_row = _labelled_row(sheet, "TOTAL SALES")
    assert round(float(sheet.cell(row=total_row, column=sheet.max_column - 1).value), 2) == (
        CLIENT_YEAR_PROJECT_VALUE
    )
    assert round(float(sheet.cell(row=total_row, column=sheet.max_column).value), 2) == (
        CLIENT_YEAR_SAMPLE_PRICE
    )


def test_the_summary_header_reads_like_the_clients(exported):
    """AC-G10. SALES AGENT, twelve uppercase month groups, then TOTAL VALUE (BY SALESMAN)."""
    sheet = exported[SUMMARY]

    assert sheet["A6"].value == "SALES AGENT"
    assert sheet["B6"].value == "JAN'25"
    assert sheet["B7"].value == "PROJECT VALUE"
    assert sheet["C7"].value == "SAMPLE PRICE"
    assert sheet.cell(row=_GROUP_ROW, column=sheet.max_column - 1).value == (
        "TOTAL VALUE (BY SALESMAN)"
    )


def test_the_monthly_totals_add_up_to_the_year(exported):
    project_value = sum(_grand_total(exported[name], "PROJECT VALUE") for name in MONTHS)
    sample_price = sum(_grand_total(exported[name], "SAMPLE PRICE") for name in MONTHS)
    assert round(project_value, 2) == CLIENT_YEAR_PROJECT_VALUE
    assert round(sample_price, 2) == CLIENT_YEAR_SAMPLE_PRICE


def test_the_summary_column_headers_are_the_twelve_months(exported):
    sheet = exported[SUMMARY]
    labels = [sheet.cell(row=_GROUP_ROW, column=col).value for col in range(2, sheet.max_column + 1)]
    named = [label for label in labels if label]
    assert named == [
        "JAN'25", "FEB'25", "MAR'25", "APR'25", "MAY'25", "JUN'25",
        "JUL'25", "AUG'25", "SEP'25", "OCT'25", "NOV'25", "DEC'25",
        "TOTAL VALUE (BY SALESMAN)",
    ]


def test_no_cell_anywhere_is_a_formula(exported):
    """AC-D4. The client's own sheet is SUM formulas; ours is what the engine computed."""
    for name in exported.sheetnames:
        for row in exported[name].iter_rows():
            for cell in row:
                assert not (isinstance(cell.value, str) and cell.value.startswith("="))
