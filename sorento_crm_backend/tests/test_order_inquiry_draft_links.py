"""A link is a DRAFT until purchasing confirms (`PLAN-scm-oi-draft-links.md`).

The handshake made linking purchasing's word and stopped the board from linking anything.
This reverses ONE half of that on purpose: the cascade runs again the moment CS confirms,
but what it writes is a DRAFT - a link on a row nobody has confirmed yet - and it stays a
draft until the row's own `ack_state` reads `acknowledged`. There is no state column on the
link (R1): the row IS the answer.

What is pinned here, one test each:

* AC-D1 a board confirm raises rows To confirm WITH their documents already found;
* AC-D2b an ORDER row (not only an ORDER BACK) drafts onto an SPO, SPO before PO, and only
  from a POOL location (R5, R11);
* AC-D3 a draft occupies the document's remaining quantity, so the next row is offered the
  rest and Confirm can never fail for want of quantity;
* AC-D4 the reorder plan still ignores a To confirm row's remainder;
* AC-D5 Confirm stamps the row and moves no link;
* AC-D6 Reject unplaces every link first, and the batch takes ONE reason;
* AC-D9 Auto link all re-deals a DRAFT onto a nearer document and never touches a confirmed
  row's link;
* AC-D11 a purchase-order confirm drafts To confirm rows;
* AC-D12 the `to_confirm` filter, its facet and the export;
* AC-D16/AC-D17 the SPO location and `late_days` on the wire;
* AC-D18/AC-D19 the two lightboxes.

Reuses `tests/test_order_inquiry_handshake.py`'s harness wholesale (`world` / `api`), for
the reason that file states: the seeding chain is the same one, `scm.committed_v` lives in
the migrated schema, and every row is seeded behind the `ZZT` marker because CI's database
has no data.
"""
from __future__ import annotations

import io
import uuid
from datetime import date, timedelta
from decimal import Decimal
from urllib.parse import quote

import pytest

from app.models.inventory import Warehouse
from app.models.procurement import (
    InboundShipment,
    PurchaseOrder,
    PurchaseOrderLine,
    SPOAllocation,
)
from app.models.project_so import (
    ACK_ACKNOWLEDGED,
    ACK_AWAITING,
    ACK_CHANGED,
    ACK_REJECTED,
    INQUIRY_CANCELLED,
    INQUIRY_PLACED,
    IV_ORDER,
    OrderInquiryLink,
    OrderInquiryRow,
)
from app.services.project_order_inquiry_service import ProjectOrderInquiryService

from .test_order_inquiry_handshake import (
    ACK_URL,
    BASE,
    FAR,
    HORIZON,
    LINK_NOW,
    LIST,
    MARKER,
    NOW,
    PURCHASING,
    WAS,
    _as_purchasing,
    _client,
    _confirm,
    _line_payload,
    _links_of,
    _open_po_line,
    _order_row,
    _product,
    _project_committed,
    _raise_one_row,
    _raise_two_rows,
    _restore,
    _settle,
    _supplier,
    _uid,
    _warehouse,
    api,
    world,
)

REJECT_BATCH = f"{LIST}/reject"
#: "Auto link all" - the worklist-wide re-deal (AC-D9), distinct from `LINK_NOW` (which the
#: page only offers after an upload and which reaches acknowledged rows via a narrower
#: seam). Named here rather than imported: the handshake test module defines its own
#: `AUTO_PLACE` mid-file, past the block this module already imports from.
AUTO_PLACE = f"{LIST}/auto-place"

#: A document that lands twelve days after the row needs it (AC-D17). `WAS` is the row's
#: own delivery date, so the number in the test is the arithmetic, not a magic constant.
LATE_BY = 12
LATE_ARRIVAL = WAS + timedelta(days=LATE_BY)


# ---------------------------------------------------------------------------
# seeding helpers this file owns
# ---------------------------------------------------------------------------


def _pooled(world) -> Warehouse:
    """Turn the world's own warehouse into a site that has a POOL above it.

    The world fixture builds one warehouse per test with no pool, and R11 is entirely
    about the pool set: an SPO line is linkable only at a pool location. Re-coding the
    fixture's own row (rather than seeding a second chain) keeps `_raise_one_row` and every
    location the rows already carry in step with it.
    """
    tag = uuid.uuid4().hex[:6].upper()
    pool = _warehouse(world.db, f"ZZP{tag}", segment="project")
    world.warehouse.warehouse_code = f"ZZP{tag}-IB"
    world.warehouse.pool_warehouse_id = pool.id
    world.db.flush()
    world.db.commit()
    return pool


def _spo_line(
    world,
    *,
    qty,
    warehouse,
    expected_date=date(2026, 8, 10),
    product=None,
    spo_number=None,
    line_no=1,
    shipment=None,
    location_code=None,
) -> SPOAllocation:
    """One OPEN shipping-order allocation - the shape the outstanding book now writes."""
    supplier = _supplier(world)
    allocation = SPOAllocation(
        id=_uid(),
        company_id=world.company_id,
        spo_number=spo_number or f"SPO-2026/08-{uuid.uuid4().hex[:4].upper()}",
        spo_line_number=line_no,
        product_id=(product or world.product).id,
        warehouse_id=warehouse.id if warehouse is not None else None,
        location_code=(
            location_code
            if location_code is not None
            else (warehouse.warehouse_code if warehouse is not None else None)
        ),
        allocated_quantity=Decimal(str(qty)),
        quantity_received=Decimal("0"),
        receipt_status="pending",
        line_status="open",
        source_system="scm_upload",
        issue_date=date(2026, 6, 1),
        expected_date=expected_date,
        supplier_id=supplier.id,
        inbound_shipment_id=shipment.id if shipment is not None else None,
    )
    world.db.add(allocation)
    world.db.flush()
    world.db.commit()
    return allocation


def _listed(client, row, **params):
    body = client.get(LIST, params={"limit": 200, **params}).json()
    return next(item for item in body["data"] if item["id"] == str(row.id))


def _link_documents(world, row) -> list:
    return [link.document for link in _links_of(world, row)]


# ---------------------------------------------------------------------------
# AC-D1: the board confirm finds the documents
# ---------------------------------------------------------------------------


def test_a_board_confirm_raises_a_to_confirm_row_that_already_holds_its_document(api):
    """The whole point of the plan: purchasing opens the page and the answer is on the row.

    The row is still `awaiting` - nobody has confirmed anything - and it carries a link,
    which is what a DRAFT is (R1).
    """
    _client, world = api
    po, _line = _open_po_line(world, qty=50)

    fixture = _raise_one_row(api)
    row = fixture["row"]

    assert row.ack_state == ACK_AWAITING
    assert _link_documents(world, row) == [po.po_number]


def test_a_row_nothing_can_cover_still_comes_out_unlinked(api):
    """AC-D2. No candidate is not an error, it is "Not found (new order)"."""
    _client, world = api

    row = _raise_one_row(api)["row"]

    assert row.ack_state == ACK_AWAITING
    assert _links_of(world, row) == []


def test_the_draft_is_the_rows_own_state_and_no_column_on_the_link(api):
    """R1. The link table gains nothing; a reader asks the row."""
    _client, world = api
    _open_po_line(world, qty=50)
    row = _raise_one_row(api)["row"]

    link = _links_of(world, row)[0]

    assert not hasattr(link, "status"), "a link state column is exactly what R1 refused"
    assert row.ack_state in (ACK_AWAITING, ACK_CHANGED), "which is what makes it a draft"


# ---------------------------------------------------------------------------
# AC-D2b / R5 / R11: the SPO side
# ---------------------------------------------------------------------------


def test_an_order_row_drafts_onto_a_shipping_order_before_any_purchase_order(api):
    """R5: "SPO link is always one, always SPO first then PO". Not only an ORDER BACK."""
    _client, world = api
    pool = _pooled(world)
    allocation = _spo_line(world, qty=50, warehouse=pool)
    _open_po_line(world, qty=50)

    row = _raise_one_row(api)["row"]

    assert row.verb == IV_ORDER
    assert _link_documents(world, row) == [allocation.spo_number]


def test_a_shipping_order_line_outside_the_pool_is_never_drafted(api):
    """R11: every location is SHOWN, only a pool location is TAKEN."""
    _client, world = api
    _pooled(world)
    off_pool = _warehouse(world.db, f"ZZO{uuid.uuid4().hex[:6].upper()}-BB")
    world.db.commit()
    allocation = _spo_line(world, qty=50, warehouse=off_pool)

    row = _raise_one_row(api)["row"]

    assert _links_of(world, row) == []
    # ... and it is still readable in the lightbox, which is the other half of R11.
    with _as_purchasing(world) as buyer:
        body = buyer.get(
            f"{LIST}/spo/{quote(allocation.spo_number, safe='')}"
        ).json()
    assert body["lines"][0]["location"] == off_pool.warehouse_code


# ---------------------------------------------------------------------------
# AC-D3: a draft occupies the quantity
# ---------------------------------------------------------------------------


def test_two_rows_are_never_drafted_onto_the_same_units(api):
    _client, world = api
    _open_po_line(world, qty=12)

    first = _raise_one_row(api, qty="10")["row"]
    second = _raise_one_row(api, qty="10")["row"]

    assert sum(Decimal(str(l.qty)) for l in _links_of(world, first)) == Decimal("10")
    assert sum(Decimal(str(l.qty)) for l in _links_of(world, second)) == Decimal("2")


# ---------------------------------------------------------------------------
# AC-D4: the plan is unmoved
# ---------------------------------------------------------------------------


def test_the_plan_still_ignores_a_to_confirm_rows_remainder(api):
    """A draft is not a decision, and the plan counts acknowledged and changed rows only."""
    _client, world = api
    _open_po_line(world, qty=50)

    _raise_one_row(api)

    assert _project_committed(world, planned=True) == Decimal("0")


# ---------------------------------------------------------------------------
# AC-D5: Confirm
# ---------------------------------------------------------------------------


def test_confirm_stamps_the_row_and_moves_no_link(api):
    _client, world = api
    po, _line = _open_po_line(world, qty=50)
    row = _raise_one_row(api)["row"]
    before = [str(link.id) for link in _links_of(world, row)]

    with _as_purchasing(world) as buyer:
        response = buyer.post(ACK_URL, json={"row_ids": [str(row.id)]})
    assert response.status_code == 200, response.text
    world.db.commit()

    world.db.refresh(row)
    assert row.ack_state == ACK_ACKNOWLEDGED
    assert [str(link.id) for link in _links_of(world, row)] == before
    assert _link_documents(world, row) == [po.po_number]


def test_confirm_fills_a_remainder_the_draft_could_not_cover(api):
    """The press still cascades: a row half covered when it was raised is finished here."""
    _client, world = api
    _open_po_line(world, qty=4)
    row = _raise_one_row(api, qty="10")["row"]
    assert sum(Decimal(str(l.qty)) for l in _links_of(world, row)) == Decimal("4")

    _open_po_line(world, qty=6)
    with _as_purchasing(world) as buyer:
        assert buyer.post(ACK_URL, json={"row_ids": [str(row.id)]}).status_code == 200
    world.db.commit()

    assert sum(Decimal(str(l.qty)) for l in _links_of(world, row)) == Decimal("10")


# ---------------------------------------------------------------------------
# AC-D6: Reject
# ---------------------------------------------------------------------------


def test_reject_unplaces_every_link_and_gives_the_quantity_back(api):
    """Today's reject refused a fully linked row outright, which with drafts is most of
    them. It takes the links down first instead, and the purchase order is free again."""
    _client, world = api
    _po, line = _open_po_line(world, qty=50)
    row = _raise_one_row(api, qty="10")["row"]
    assert _links_of(world, row), "the draft has to exist for the test to mean anything"
    assert row.state == INQUIRY_PLACED

    with _as_purchasing(world) as buyer:
        response = buyer.post(
            f"{LIST}/{row.id}/reject", json={"reason": "Factory closed"}
        )
    assert response.status_code == 200, response.text
    world.db.commit()

    world.db.refresh(row)
    assert row.ack_state == ACK_REJECTED
    assert _links_of(world, row) == []
    service = ProjectOrderInquiryService(world.db)
    by_po, _by_spo = service._linked_by_target()
    assert by_po.get(str(line.id), Decimal("0")) == Decimal("0")


def test_reject_on_a_placed_row_frees_the_pos_remaining_for_the_next_candidate(api):
    """Not just a number on a report: the freed quantity is real enough for a SECOND row's
    own draft to take it on the next re-deal. The first row takes the whole line, so a
    second row raised straight after it is left with nothing - the proof the reject really
    freed something rather than merely zeroing a count."""
    _client, world = api
    _open_po_line(world, qty=10)
    first = _raise_one_row(api, qty="10")["row"]
    assert sum(Decimal(str(l.qty)) for l in _links_of(world, first)) == Decimal("10")

    second = _raise_one_row(api, qty="5")["row"]
    assert _links_of(world, second) == [], "nothing was left for it while the first held it"

    with _as_purchasing(world) as buyer:
        assert (
            buyer.post(
                f"{LIST}/{first.id}/reject", json={"reason": "Cancelled by customer"}
            ).status_code
            == 200
        )
        world.db.commit()
        assert buyer.post(AUTO_PLACE, json={}).status_code == 200
    world.db.commit()

    world.db.refresh(second)
    assert sum(Decimal(str(l.qty)) for l in _links_of(world, second)) == Decimal("5")


def test_the_batch_reject_takes_one_reason_for_every_row(api):
    _client, world = api
    _open_po_line(world, qty=50)
    first = _raise_one_row(api, qty="4")["row"]
    second = _raise_one_row(api, qty="6")["row"]

    with _as_purchasing(world) as buyer:
        response = buyer.post(
            REJECT_BATCH,
            json={"row_ids": [str(first.id), str(second.id)], "reason": "Discontinued"},
        )
    assert response.status_code == 200, response.text
    world.db.commit()

    body = response.json()
    assert body["rejected"] == 2
    assert {entry["row_id"] for entry in body["results"]} == {
        str(first.id),
        str(second.id),
    }
    assert all(entry["ok"] for entry in body["results"])
    for row in (first, second):
        world.db.refresh(row)
        assert row.ack_state == ACK_REJECTED
        assert row.rejected_reason == "Discontinued"


def test_the_batch_reject_refuses_an_empty_reason(api):
    _client, world = api
    row = _raise_one_row(api)["row"]

    with _as_purchasing(world) as buyer:
        response = buyer.post(
            REJECT_BATCH, json={"row_ids": [str(row.id)], "reason": "   "}
        )

    assert response.status_code == 422, response.text


def test_the_batch_reject_refuses_the_whole_batch_when_one_row_cannot_be_refused(api):
    """A batch that half-happened is worse than one that did not: the buyer pressed once."""
    _client, world = api
    good = _raise_one_row(api, qty="4")["row"]
    already = _raise_one_row(api, qty="6")["row"]
    with _as_purchasing(world) as buyer:
        assert (
            buyer.post(
                f"{LIST}/{already.id}/reject", json={"reason": "First refusal"}
            ).status_code
            == 200
        )
        world.db.commit()
        response = buyer.post(
            REJECT_BATCH,
            json={"row_ids": [str(good.id), str(already.id)], "reason": "Second"},
        )

    assert response.status_code == 422, response.text
    world.db.rollback()
    world.db.refresh(good)
    assert good.ack_state == ACK_AWAITING, "nothing was written for the batch"


def test_the_batch_reject_refuses_the_whole_batch_when_one_row_is_cancelled(api):
    """A different branch of `_assert_rejectable` from the already-rejected case above -
    `order_inquiry_row_not_rejectable` rather than `order_inquiry_already_rejected` - and
    the same ALL-OR-NOTHING rule has to hold for it too."""
    _client, world = api
    good = _raise_one_row(api, qty="4")["row"]
    cancelled = _raise_one_row(api, qty="6")["row"]
    cancelled.state = INQUIRY_CANCELLED
    world.db.commit()

    with _as_purchasing(world) as buyer:
        response = buyer.post(
            REJECT_BATCH,
            json={"row_ids": [str(good.id), str(cancelled.id)], "reason": "Whole batch"},
        )

    assert response.status_code == 422, response.text
    world.db.rollback()
    world.db.refresh(good)
    assert good.ack_state == ACK_AWAITING, "nothing was written for the batch"


def test_a_cs_user_may_not_reject_a_batch(api):
    """AC-D8. Reject is purchasing's, on the acknowledge grant like every other press."""
    client, world = api
    row = _raise_one_row(api)["row"]

    response = client.post(
        REJECT_BATCH, json={"row_ids": [str(row.id)], "reason": "Not mine to say"}
    )

    assert response.status_code == 403


# ---------------------------------------------------------------------------
# AC-D9: the re-deal
# ---------------------------------------------------------------------------


def test_auto_link_all_moves_a_draft_onto_a_nearer_document(api):
    """R2. The whole reason a draft may be re-dealt: a better document arrived."""
    _client, world = api
    far, _far_line = _open_po_line(world, qty=50, expected_date=date(2026, 12, 1))
    row = _raise_one_row(api)["row"]
    assert _link_documents(world, row) == [far.po_number]

    near, _near_line = _open_po_line(world, qty=50, expected_date=date(2026, 7, 1))
    with _as_purchasing(world) as buyer:
        response = buyer.post(LINK_NOW, json={})
    assert response.status_code == 200, response.text
    world.db.commit()

    assert _link_documents(world, row) == [near.po_number]


def test_auto_link_all_never_moves_a_confirmed_rows_link(api):
    """The other half of R2, and the one that matters: a confirmed link is a promise."""
    _client, world = api
    far, _far_line = _open_po_line(world, qty=50, expected_date=date(2026, 12, 1))
    row = _raise_one_row(api)["row"]

    with _as_purchasing(world) as buyer:
        assert buyer.post(ACK_URL, json={"row_ids": [str(row.id)]}).status_code == 200
        world.db.commit()
        _open_po_line(world, qty=50, expected_date=date(2026, 7, 1))
        assert buyer.post(LINK_NOW, json={}).status_code == 200
    world.db.commit()

    assert _link_documents(world, row) == [far.po_number]


# ---------------------------------------------------------------------------
# AC-D11: the purchase-order confirm
# ---------------------------------------------------------------------------


def test_a_purchase_order_confirm_drafts_a_to_confirm_row(api):
    """A plan-generated purchase order is confirmed and the rows that sized it read it."""
    from app.services.scm.purchase_order_service import PurchaseOrderService

    _client, world = api
    row = _raise_one_row(api)["row"]
    assert _links_of(world, row) == []

    supplier = _supplier(world)
    po = PurchaseOrder(
        id=_uid(),
        company_id=world.company_id,
        po_number=f"ZZT-DRAFT-{_uid()[:8]}",
        supplier_id=supplier.id,
        status="draft_recommendation",
    )
    world.db.add(po)
    world.db.flush()
    world.db.add(
        PurchaseOrderLine(
            id=_uid(),
            company_id=world.company_id,
            purchase_order_id=po.id,
            product_id=world.product.id,
            warehouse_id=world.warehouse.id,
            qty_ordered=Decimal("50"),
            qty_received=Decimal("0"),
            expected_date=date(2026, 8, 10),
            line_status="open",
        )
    )
    world.db.commit()

    PurchaseOrderService(world.db).bulk_confirm([str(po.id)], actor=world.buyer)

    world.db.expire_all()
    row = world.db.query(OrderInquiryRow).filter(OrderInquiryRow.id == row.id).one()
    assert row.ack_state == ACK_AWAITING
    assert len(_links_of(world, row)) == 1


# ---------------------------------------------------------------------------
# AC-D12: the To confirm filter
# ---------------------------------------------------------------------------


def test_the_to_confirm_filter_is_awaiting_and_changed(api):
    client, world = api
    awaiting = _raise_one_row(api, qty="4")["row"]
    changed_fixture = _raise_one_row(api, qty="6")
    confirmed = _raise_one_row(api, qty="8")["row"]

    with _as_purchasing(world) as buyer:
        assert (
            buyer.post(
                ACK_URL,
                json={
                    "row_ids": [
                        str(changed_fixture["row"].id),
                        str(confirmed.id),
                    ]
                },
            ).status_code
            == 200
        )
        world.db.commit()
    _settle(world, changed_fixture, qty="9")
    world.db.commit()
    changed = _order_row(world, changed_fixture["line"])
    assert changed.ack_state == ACK_CHANGED

    body = client.get(LIST, params={"ack": "to_confirm", "limit": 200}).json()
    ids = {item["id"] for item in body["data"]}

    assert str(awaiting.id) in ids
    assert str(changed.id) in ids
    assert str(confirmed.id) not in ids


def test_the_summary_facet_counts_to_confirm(api):
    client, world = api
    _raise_one_row(api, qty="4")
    _raise_one_row(api, qty="6")

    facet = client.get(f"{LIST}/summary").json()["ack"]

    assert facet["to_confirm"] == facet["awaiting"] + facet["changed"]
    assert facet["to_confirm"] >= 2


def test_the_export_accepts_to_confirm(api):
    """Not a status check alone: `to_confirm` FILTERS the sheet the same way it filters the
    list - a row purchasing has confirmed is off it, one still owed is on it."""
    client, world = api
    to_confirm = _raise_one_row(api, qty="4")
    confirmed = _raise_one_row(api, qty="6")
    with _as_purchasing(world) as buyer:
        assert (
            buyer.post(ACK_URL, json={"row_ids": [str(confirmed["row"].id)]}).status_code
            == 200
        )
    world.db.commit()

    export = client.get(f"{LIST}/export", params={"ack": "to_confirm"})

    assert export.status_code == 200, export.text
    import openpyxl

    book = openpyxl.load_workbook(io.BytesIO(export.content))
    assert book.sheetnames
    numbers: set[str] = set()
    for name in book.sheetnames:
        for row in book[name].iter_rows(min_row=3, values_only=True):
            if row and row[1]:
                numbers.add(str(row[1]))
    assert to_confirm["core_so"].so_number in numbers
    assert confirmed["core_so"].so_number not in numbers


def test_an_unknown_acknowledgement_filter_is_still_refused(api):
    client, _world = api

    assert client.get(LIST, params={"ack": "maybe"}).status_code == 422


# ---------------------------------------------------------------------------
# AC-D16 / AC-D17: what the column reads
# ---------------------------------------------------------------------------


def test_a_late_document_says_how_many_days_late_it_is(api):
    client, world = api
    _open_po_line(world, qty=50, expected_date=LATE_ARRIVAL)
    row = _raise_one_row(api)["row"]

    listed = _listed(client, row)
    link = listed["links"][0]

    assert link["late"] is True
    assert link["late_days"] == LATE_BY


def test_a_document_that_lands_in_time_states_no_day_count(api):
    client, world = api
    _open_po_line(world, qty=50, expected_date=date(2026, 8, 1))
    row = _raise_one_row(api)["row"]

    link = _listed(client, row)["links"][0]

    assert link["late"] is False
    assert link["late_days"] is None


def test_an_spo_link_reads_its_pool_location_rather_than_a_line_number(api):
    client, world = api
    pool = _pooled(world)
    _spo_line(world, qty=50, warehouse=pool, line_no=14)
    row = _raise_one_row(api)["row"]

    link = _listed(client, row)["links"][0]

    assert link["kind"] == "spo"
    assert link["location"] == pool.warehouse_code
    assert link["line_label"] == "L14", "the line number moves into the title, it is kept"


def test_an_spo_link_with_no_warehouse_falls_back_to_the_books_own_code(api):
    """The banded history book states no warehouse; the code it printed is still a fact."""
    client, world = api
    _pooled(world)
    row = _raise_one_row(api)["row"]
    allocation = _spo_line(
        world, qty=50, warehouse=None, location_code="BRW-NOWHERE"
    )
    # Written straight, not through the cascade: this pins the SERIALIZER's fallback, and
    # a line at a code we hold no warehouse for is never a candidate (R11) by design.
    world.db.add(
        OrderInquiryLink(
            id=_uid(),
            company_id=world.company_id,
            row_id=row.id,
            spo_allocation_id=allocation.id,
            document=allocation.spo_number,
            qty=Decimal("10"),
        )
    )
    world.db.commit()

    link = _listed(client, row)["links"][0]

    assert link["location"] == "BRW-NOWHERE"


def test_the_sales_order_detail_carries_the_day_count_too(api):
    """The SO detail is a second reader of the same link, and `response_model` drops a
    field it has not been told about just as silently there."""
    client, world = api
    _open_po_line(world, qty=50, expected_date=LATE_ARRIVAL)
    fixture = _raise_one_row(api)

    body = client.get(
        f"{BASE}/sales-orders/{fixture['order'].id}/order-inquiry"
    ).json()
    row = next(item for item in body["rows"] if item["id"] == str(fixture["row"].id))

    assert row["links"][0]["late_days"] == LATE_BY


# ---------------------------------------------------------------------------
# AC-D18 / AC-D19: the lightboxes
# ---------------------------------------------------------------------------


def test_the_purchase_order_lightbox_names_who_is_holding_the_quantity(api):
    client, world = api
    po, _line = _open_po_line(world, qty=50)
    row = _raise_one_row(api, qty="10")["row"]

    body = client.get(f"{LIST}/po/{po.id}").json()

    assert body["allocations"], "the Allocated to panel reads off the links"
    allocation = body["allocations"][0]
    assert allocation["qty"] == "10"
    assert allocation["ack_state"] == ACK_AWAITING, "a draft, and the panel says Proposed"
    assert allocation["item_code"] == row.item_code
    assert allocation["inquiry_no"]


def test_the_purchase_order_lightbox_reads_a_confirmed_allocation_as_confirmed(api):
    client, world = api
    po, _line = _open_po_line(world, qty=50)
    row = _raise_one_row(api, qty="10")["row"]
    with _as_purchasing(world) as buyer:
        assert buyer.post(ACK_URL, json={"row_ids": [str(row.id)]}).status_code == 200
    world.db.commit()

    body = client.get(f"{LIST}/po/{po.id}").json()

    assert body["allocations"][0]["ack_state"] == ACK_ACKNOWLEDGED


def test_the_shipping_order_lightbox_answers_its_lines(api):
    client, world = api
    pool = _pooled(world)
    shipment = InboundShipment(
        id=_uid(),
        company_id=world.company_id,
        shipment_number=f"ZZT-SHIP-{_uid()[:8]}",
        shipment_date=date(2026, 8, 1),
        shipping_container_number=f"ZZTU{_uid()[:7].upper()}",
    )
    world.db.add(shipment)
    world.db.flush()
    allocation = _spo_line(world, qty=50, warehouse=pool, shipment=shipment)

    response = client.get(f"{LIST}/spo/{quote(allocation.spo_number, safe='')}")

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["spo_number"] == allocation.spo_number
    assert body["shipment_ref"] == shipment.shipment_number
    assert body["container_no"] == shipment.shipping_container_number
    line = body["lines"][0]
    assert line["sku"] == world.product.product_code
    assert line["allocated"] == "50"
    assert line["received"] == "0"
    assert line["remaining"] == "50"
    assert line["location"] == pool.warehouse_code


def test_the_shipping_order_lightbox_names_who_is_holding_it(api):
    client, world = api
    pool = _pooled(world)
    allocation = _spo_line(world, qty=50, warehouse=pool)
    row = _raise_one_row(api, qty="10")["row"]

    body = client.get(f"{LIST}/spo/{quote(allocation.spo_number, safe='')}").json()

    assert body["allocations"][0]["qty"] == "10"
    assert body["allocations"][0]["ack_state"] == ACK_AWAITING
    assert body["allocations"][0]["item_code"] == row.item_code


def test_the_shipping_order_lightbox_404s_on_a_number_nobody_holds(api):
    client, _world = api

    response = client.get(f"{LIST}/spo/{quote('SPO-2026/08-NOPE', safe='')}")

    assert response.status_code == 404


def test_the_shipping_order_lightbox_denies_a_user_without_the_view_grant(world):
    pool = _pooled(world)
    allocation = _spo_line(world, qty=50, warehouse=pool)
    client, originals = _client(world.db, world.buyer, [])
    try:
        response = client.get(f"{LIST}/spo/{quote(allocation.spo_number, safe='')}")
    finally:
        _restore(originals)

    assert response.status_code == 403


# ---------------------------------------------------------------------------
# Review round (28 Aug): the re-deal never costs a row the links it already has
# ---------------------------------------------------------------------------


def test_auto_link_all_keeps_the_draft_of_a_row_that_is_now_past_the_cut_off(api):
    """B1. The unplace used to run over the WHOLE scope before the per-row guards, so a
    drafted row the buyer had since moved past the cut off lost its documents and was
    reported as merely "held back". A press that leaves a row alone must leave its links
    alone: the cut off says "do not deal this one", not "take back what it holds"."""
    _client, world = api
    _po, _line = _open_po_line(world, qty=50)
    row = _raise_one_row(api, qty="10")["row"]
    before = [str(link.id) for link in _links_of(world, row)]
    assert before, "the draft has to exist for the test to mean anything"

    row.delivery_date = FAR
    world.db.flush()
    world.db.commit()

    with _as_purchasing(world) as buyer:
        response = buyer.post(AUTO_PLACE, json={"link_up_to": HORIZON.isoformat()})
    assert response.status_code == 200, response.text
    world.db.commit()

    # `>= 1`: this suite runs on the shared prod-copy database, whose own rows are past
    # this cut off too. What matters is that THIS row was counted and left alone.
    assert response.json()["after_horizon"] >= 1
    assert [str(link.id) for link in _links_of(world, row)] == before


def test_auto_link_all_keeps_the_draft_when_the_document_has_since_closed(api):
    """B1, the other half. The re-deal found no candidate at all - the purchase order was
    received and closed between the raise and the press - and the old answer is still the
    best one anybody has. Taking it down would leave the row reading "Not found (new
    order)" for a quantity that IS on its way."""
    _client, world = api
    po, line = _open_po_line(world, qty=50)
    row = _raise_one_row(api, qty="10")["row"]
    before = [str(link.id) for link in _links_of(world, row)]
    assert before

    line.line_status = "closed"
    world.db.flush()
    world.db.commit()

    with _as_purchasing(world) as buyer:
        assert buyer.post(AUTO_PLACE, json={}).status_code == 200
    world.db.commit()

    assert [str(link.id) for link in _links_of(world, row)] == before
    assert _link_documents(world, row) == [po.po_number]


def test_two_presses_of_auto_link_all_change_nothing_at_all(api):
    """S4. The re-deal deleted and rewrote identical links on every press, and wrote
    "Unlinked from X; Re-dealt by worklist" onto the row's note each time - so a buyer who
    pressed the button twice read a row that looked like it had moved twice. The take is
    computed first, and a take that matches what the row already holds is skipped."""
    _client, world = api
    _open_po_line(world, qty=50)
    row = _raise_one_row(api, qty="10")["row"]

    with _as_purchasing(world) as buyer:
        assert buyer.post(AUTO_PLACE, json={}).status_code == 200
        world.db.commit()
        world.db.refresh(row)
        first = [str(link.id) for link in _links_of(world, row)]
        note = row.note
        assert buyer.post(AUTO_PLACE, json={}).status_code == 200
    world.db.commit()

    world.db.refresh(row)
    assert [str(link.id) for link in _links_of(world, row)] == first
    assert row.note == note


# ---------------------------------------------------------------------------
# B2: a re-confirm of a DRAFTED row re-raises it rather than netting it as bought
# ---------------------------------------------------------------------------


def test_a_reconfirm_with_a_new_date_re_raises_the_drafted_row_on_that_date(api):
    """B2. A draft made the row `placed`, and the reconfirm netted it as supply already
    bought - so a board re-confirm carrying a new delivery date did nothing at all and the
    row went on saying the old one. A row nobody has confirmed is an instruction, not
    supply: it is unlinked, superseded, and re-raised on the new date (which the raise-time
    cascade then drafts again)."""
    _client, world = api
    po, _line = _open_po_line(world, qty=50)
    fixture = _raise_one_row(api)
    row = fixture["row"]
    assert _links_of(world, row), "the draft has to exist for the test to mean anything"

    fixture["core_line"].required_date = NOW
    fixture["line"].delivery_date = NOW
    world.db.flush()
    response = _confirm(
        _client, fixture["order"].id, [_line_payload(fixture["line"].id, buy_qty="10")]
    )
    assert response.status_code == 200, response.text
    world.db.commit()

    world.db.refresh(row)
    assert row.state == INQUIRY_CANCELLED, "the draft was netted as if it were bought"
    assert _links_of(world, row) == [], "and it kept the document with it"
    replacement = _order_row(world, fixture["line"])
    assert replacement.delivery_date == NOW
    assert _link_documents(world, replacement) == [po.po_number]


def test_a_reconfirm_that_lowers_a_drafted_rows_quantity_raises_no_exception(api):
    """B2. `placed > need` wrote a CANCEL_BALANCE exception - "purchasing bought 10, CS now
    wants 4" - about a purchase nobody had agreed to."""
    _client, world = api
    _open_po_line(world, qty=50)
    fixture = _raise_one_row(api, qty="10")
    row = fixture["row"]

    fixture["core_line"].qty_ordered = Decimal("4")
    fixture["line"].qty = Decimal("4")
    world.db.flush()
    response = _confirm(
        _client, fixture["order"].id, [_line_payload(fixture["line"].id, buy_qty="4")]
    )
    assert response.status_code == 200, response.text
    world.db.commit()

    assert response.json()["exceptions"] == []
    assert _cancel_balance_rows(world, fixture["line"]) == []
    live = _live_rows(world, fixture["line"])
    assert [str(item.qty) for item in live] == ["4.0000"]
    assert sum(
        Decimal(str(link.qty)) for link in _links_of(world, live[0])
    ) == Decimal("4")


def test_a_reconfirm_that_raises_a_drafted_rows_quantity_leaves_one_row(api):
    """B2. The netting split the line: 10 already "placed" plus a fresh 5, two rows in
    front of purchasing for one instruction. One row of 15, drafted."""
    _client, world = api
    _open_po_line(world, qty=50)
    fixture = _raise_one_row(api, qty="10")

    fixture["core_line"].qty_ordered = Decimal("15")
    fixture["line"].qty = Decimal("15")
    world.db.flush()
    response = _confirm(
        _client, fixture["order"].id, [_line_payload(fixture["line"].id, buy_qty="15")]
    )
    assert response.status_code == 200, response.text
    world.db.commit()

    live = _live_rows(world, fixture["line"])
    assert [str(item.qty) for item in live] == ["15.0000"]
    assert sum(
        Decimal(str(link.qty)) for link in _links_of(world, live[0])
    ) == Decimal("15")


def test_a_confirmed_rows_links_survive_a_reconfirm_untouched(api):
    """The other side of B2, and the rule it must not break: purchasing said yes, so the
    row IS supply and a reconfirm nets it exactly as it always did."""
    _client, world = api
    po, _line = _open_po_line(world, qty=50)
    fixture = _raise_one_row(api, qty="10")
    row = fixture["row"]
    with _as_purchasing(world) as buyer:
        assert buyer.post(ACK_URL, json={"row_ids": [str(row.id)]}).status_code == 200
    world.db.commit()
    before = [str(link.id) for link in _links_of(world, row)]

    response = _confirm(
        _client, fixture["order"].id, [_line_payload(fixture["line"].id, buy_qty="10")]
    )
    assert response.status_code == 200, response.text
    world.db.commit()

    world.db.refresh(row)
    assert row.ack_state == ACK_ACKNOWLEDGED
    assert row.state == INQUIRY_PLACED
    assert [str(link.id) for link in _links_of(world, row)] == before
    assert _link_documents(world, row) == [po.po_number]


# ---------------------------------------------------------------------------
# B3: a drafted row whose line leaves the revision is retired with it
# ---------------------------------------------------------------------------


def test_a_drafted_row_is_retired_when_its_line_leaves_the_revision(api):
    """B3. The retirement read `raised` only, and a drafted row is `placed` - so a line CS
    took back out of the decision left its row alive, holding purchase-order quantity for
    an instruction that no longer exists."""
    from app.services.project_supply_service import ProjectSupplyService

    _client, world = api
    _po, line = _open_po_line(world, qty=50)
    fixture = _raise_two_rows(api)
    dropped = fixture["first"]["row"]
    kept = fixture["second"]["row"]
    assert _links_of(world, dropped), "the draft has to exist for the test to mean anything"

    ProjectSupplyService(world.db).uncover_lines(
        fixture["order"],
        [str(fixture["first"]["line"].id)],
        actor_user_id=world.cs_user,
        reason="CS took the line back.",
    )
    world.db.commit()

    world.db.refresh(dropped)
    assert dropped.state == INQUIRY_CANCELLED
    assert _links_of(world, dropped) == [], "it held the quantity for ever"
    assert _live_rows(world, fixture["first"]["line"]) == []
    assert _links_of(world, _order_row(world, fixture["second"]["line"])), (
        "the line the revision kept keeps its own draft"
    )
    assert kept is not None


# ---------------------------------------------------------------------------
# B4: a batch reject over two lines of ONE order
# ---------------------------------------------------------------------------


def _raise_three_rows(api):
    """One order, THREE lines, all confirmed as Buy - the shape a batch reject needs.

    Two lines are refused and the third is not, so the un-decide has a surviving line to
    carry: that is what makes the press write a revision at all (an order whose every
    covered line is refused is superseded outright, and there is no "one revision per
    order" to count).
    """
    from .test_planning_changes import _core_line, _core_so, _project_line, _project_so

    client, world = api
    db = world.db
    core_so = _core_so(db, world.company_id)
    order = _project_so(
        db, world.project, so_id=core_so.id, autocount_doc_no=core_so.so_number
    )
    made = []
    for index, qty in enumerate(("10", "6", "4"), start=1):
        core_line = _core_line(
            db, core_so, world.product, world.warehouse, qty_ordered=qty,
            required_date=WAS,
        )
        line = _project_line(
            db, order, line_no=index, product=world.product, core_line=core_line
        )
        made.append({"line": line, "core_line": core_line, "qty": qty})
    db.commit()

    response = _confirm(
        client,
        order.id,
        [_line_payload(entry["line"].id, buy_qty=entry["qty"]) for entry in made],
    )
    assert response.status_code == 200, response.text
    db.commit()
    for entry in made:
        entry["row"] = _order_row(world, entry["line"])
    return {"order": order, "lines": made}


def test_the_batch_reject_of_two_lines_of_one_order_refuses_both(api):
    """B4. Each row was uncovered on its own, and uncovering ONE line writes a revision
    that cancels and re-raises the OTHER lines' rows - so the second refusal stamped a row
    that had just been superseded, and its live replacement went on sitting in front of
    purchasing as if nobody had refused it. The batch uncovers a whole order in one call:
    one revision, and every refused line named in it.
    """
    _client, world = api
    _open_po_line(world, qty=50)
    fixture = _raise_three_rows(api)
    first, second, kept = (entry["row"] for entry in fixture["lines"])
    assert _links_of(world, first), "the drafts have to exist for the test to mean anything"
    revisions_before = _revision_count(world, fixture["order"])

    with _as_purchasing(world) as buyer:
        response = buyer.post(
            REJECT_BATCH,
            json={"row_ids": [str(first.id), str(second.id)], "reason": "No supplier"},
        )
    assert response.status_code == 200, response.text
    world.db.commit()

    for row in (first, second):
        world.db.refresh(row)
        assert row.ack_state == ACK_REJECTED, (
            "a mid-batch revision moved the row this press was refusing"
        )
        assert _links_of(world, row) == []
    for entry in fixture["lines"][:2]:
        left = [
            item
            for item in _live_rows(world, entry["line"])
            if item.ack_state != ACK_REJECTED
        ]
        assert left == [], "a line the batch refused still has a row nobody refused"
    assert _revision_count(world, fixture["order"]) == revisions_before + 1, (
        "one press, one revision per order"
    )
    assert _live_rows(world, fixture["lines"][2]["line"]), (
        "the line nobody refused is still purchasing's work"
    )
    assert kept is not None


def _live_rows(world, line) -> list:
    """Every row of this line the page still shows - what `_order_row` asks for, without
    its "exactly one" demand, because half of these tests are about how many there are."""
    return (
        world.db.query(OrderInquiryRow)
        .filter(
            OrderInquiryRow.so_line_id == line.id,
            OrderInquiryRow.verb == IV_ORDER,
            OrderInquiryRow.state != INQUIRY_CANCELLED,
        )
        .order_by(OrderInquiryRow.created_at.asc())
        .all()
    )


def _cancel_balance_rows(world, line) -> list:
    """The "purchasing bought more than CS now wants" exception rows on this line."""
    from app.models.project_so import IV_CANCEL_BALANCE

    return (
        world.db.query(OrderInquiryRow)
        .filter(
            OrderInquiryRow.so_line_id == line.id,
            OrderInquiryRow.verb == IV_CANCEL_BALANCE,
            OrderInquiryRow.state != INQUIRY_CANCELLED,
        )
        .all()
    )


def _revision_count(world, order) -> int:
    from app.models.project_so import SOSupplyDecision

    return (
        world.db.query(SOSupplyDecision)
        .filter(SOSupplyDecision.project_sales_order_id == order.id)
        .count()
    )


# ---------------------------------------------------------------------------
# S1: the day count on the SCM sales order's own Linked to column
# ---------------------------------------------------------------------------


def test_the_scm_sales_order_detail_states_the_day_count_too(api):
    """S1. `SalesOrderLineLink` is a third `response_model` over the same link, and the SO
    detail's Lines tab already prints "arrives late" off it - so the number of days has to
    survive that schema as well, or the badge there says less than the same badge two
    screens away."""
    _client, world = api
    _open_po_line(world, qty=50, expected_date=LATE_ARRIVAL)
    fixture = _raise_one_row(api)

    with _as_purchasing(world, permissions=[*PURCHASING, "scm.dashboard.view"]) as buyer:
        response = buyer.get(f"/api/v1/scm/sales-orders/{fixture['core_so'].id}")
    assert response.status_code == 200, response.text

    lines = {line["id"]: line for line in response.json()["lines"]}
    link = lines[str(fixture["core_line"].id)]["linked_to"][0]
    assert link["late"] is True
    assert link["late_days"] == LATE_BY


# ---------------------------------------------------------------------------
# S5: a plan purchase order takes the draft it was bought for
# ---------------------------------------------------------------------------


def _po_at(world, *, qty, issue_date, expected_date, status="active"):
    """One purchase order and one line, with the dates the walk sorts on stated."""
    supplier = _supplier(world)
    po = PurchaseOrder(
        id=_uid(),
        company_id=world.company_id,
        po_number=f"ZZT-PO-{_uid()[:8]}",
        supplier_id=supplier.id,
        issue_date=issue_date,
        status=status,
    )
    world.db.add(po)
    world.db.flush()
    line = PurchaseOrderLine(
        id=_uid(),
        company_id=world.company_id,
        purchase_order_id=po.id,
        product_id=world.product.id,
        warehouse_id=world.warehouse.id,
        qty_ordered=Decimal(str(qty)),
        qty_received=Decimal("0"),
        expected_date=expected_date,
        line_status="open",
    )
    world.db.add(line)
    world.db.flush()
    world.db.commit()
    return po, line


def test_a_plan_purchase_order_confirm_moves_the_draft_it_was_bought_for(api):
    """S5. The purchase-order confirm is one of the four re-deal doors (section 5.4): the
    plan bought THIS order for these rows, so its own document beats the far one the raise
    could reach at the time. Only a draft moves."""
    from app.services.scm.purchase_order_service import PurchaseOrderService

    _client, world = api
    far, _far_line = _po_at(
        world, qty=50, issue_date=date(2026, 7, 1), expected_date=date(2027, 1, 1)
    )
    row = _raise_one_row(api, qty="10")["row"]
    assert _link_documents(world, row) == [far.po_number]

    plan_po, _line = _po_at(
        world,
        qty=50,
        issue_date=date(2026, 6, 1),
        expected_date=date(2026, 8, 10),
        status="draft_recommendation",
    )
    PurchaseOrderService(world.db).bulk_confirm([str(plan_po.id)], actor=world.buyer)
    world.db.commit()

    world.db.expire_all()
    row = world.db.query(OrderInquiryRow).filter(OrderInquiryRow.id == row.id).one()
    assert _link_documents(world, row) == [plan_po.po_number]


def test_a_plan_purchase_order_confirm_never_moves_a_confirmed_rows_link(api):
    """The same press, on a row purchasing has already confirmed: its link is a promise."""
    from app.services.scm.purchase_order_service import PurchaseOrderService

    _client, world = api
    far, _far_line = _po_at(
        world, qty=50, issue_date=date(2026, 7, 1), expected_date=date(2027, 1, 1)
    )
    row = _raise_one_row(api, qty="10")["row"]
    with _as_purchasing(world) as buyer:
        assert buyer.post(ACK_URL, json={"row_ids": [str(row.id)]}).status_code == 200
    world.db.commit()

    plan_po, _line = _po_at(
        world,
        qty=50,
        issue_date=date(2026, 6, 1),
        expected_date=date(2026, 8, 10),
        status="draft_recommendation",
    )
    PurchaseOrderService(world.db).bulk_confirm([str(plan_po.id)], actor=world.buyer)
    world.db.commit()

    world.db.expire_all()
    row = world.db.query(OrderInquiryRow).filter(OrderInquiryRow.id == row.id).one()
    assert _link_documents(world, row) == [far.po_number]


# ---------------------------------------------------------------------------
# S6: the plan page's "to confirm" chip counts drafted rows too
# ---------------------------------------------------------------------------


def test_the_to_confirm_count_still_sees_a_row_its_draft_made_placed(api):
    """S6. The count read `awaiting` rows in `raised` / `partly_linked` only, and a drafted
    row is `placed` - so the chip on the plan page emptied itself the moment the raise
    found a document, which is exactly when purchasing has something to confirm."""
    from app.services.scm import reorder_run_service

    _client, world = api
    _open_po_line(world, qty=50)
    before = reorder_run_service.awaiting_acknowledgement_rows(world.db)

    row = _raise_one_row(api, qty="10")["row"]
    assert row.state == INQUIRY_PLACED, "the row has to be drafted for this to mean anything"

    assert reorder_run_service.awaiting_acknowledgement_rows(world.db) == before + 1

    with _as_purchasing(world) as buyer:
        assert buyer.post(ACK_URL, json={"row_ids": [str(row.id)]}).status_code == 200
    world.db.commit()

    assert reorder_run_service.awaiting_acknowledgement_rows(world.db) == before
