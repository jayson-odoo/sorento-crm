"""Re-importing the Order Tracking Excel must not un-deliver an already-delivered DO.

Real production bug: the Master sheet carries the FULL order history on every
upload, but the Overall Tracking sheet is only a partial/rolling window (real
import_logs show master_rows routinely far exceeding tracking_rows). The Master
loop used to reset order_status_id to "New" unconditionally for every matched
order, even one that already had a real actual_delivery_date + Delivered status
recorded by an earlier run - wiping the status while leaving the date behind.
Traced from a CSV export where 5394/5394 delivery orders had actual_delivery_date
set but order_status stuck at "New Order".
"""
from __future__ import annotations

from datetime import date
from io import BytesIO

import pytest
from openpyxl import Workbook

from app.models.order import Order, OrderStatus
from app.services.order_service import OrderService
from tests._pg_fixture import blank_session


@pytest.fixture()
def db():
    with blank_session() as session:
        yield session


def _seed_statuses(db) -> dict[str, str]:
    ids = {}
    for code, name in [("NEW", "New Order"), ("DELIVERED", "Delivered"), ("COMPLETED", "Completed")]:
        row = OrderStatus(status_code=code, status_name=name)
        db.add(row)
        db.flush()
        ids[code] = row.id
    db.commit()
    return ids


def _build_workbook(*, master_rows: list[list], tracking_rows: list[list]) -> bytes:
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    master = wb.create_sheet("Master")
    master.append(["Doc. No.", "Date"])
    for r in master_rows:
        master.append(r)
    tracking = wb.create_sheet("Overall Tracking")
    tracking.append(["Doc Number", "Date"])
    for r in tracking_rows:
        tracking.append(r)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_master_only_reimport_does_not_undeliver_an_existing_order(db):
    status_ids = _seed_statuses(db)
    order = Order(
        order_number="DO-EXIST-1",
        order_date=date(2026, 7, 1),
        actual_delivery_date=date(2026, 7, 5),
        order_status_id=status_ids["DELIVERED"],
    )
    db.add(order)
    db.commit()

    svc = OrderService(db)
    # This run's Master sheet re-lists the order (as every run does); the
    # Overall Tracking sheet does NOT include it this time (the rolling-window
    # gap that triggers the bug).
    data = _build_workbook(
        master_rows=[["DO-EXIST-1", date(2026, 7, 1)]],
        tracking_rows=[],
    )

    result = svc.import_excel_tracking(data, "11111111-1111-1111-1111-111111111111")
    assert not result.get("errors"), result.get("errors")

    db.refresh(order)
    assert order.order_status_id == status_ids["DELIVERED"], (
        "an order already delivered must stay delivered when this run's Overall "
        "Tracking tab simply doesn't happen to re-list it"
    )
    assert order.actual_delivery_date == date(2026, 7, 5)


def test_master_reimport_still_defaults_a_never_delivered_order_to_new(db):
    """The fix must not defeat the original purpose: an order with no delivery
    date yet still gets (or stays) "New" on a Master reimport."""
    status_ids = _seed_statuses(db)
    order = Order(
        order_number="DO-PENDING-1",
        order_date=date(2026, 7, 1),
        actual_delivery_date=None,
        order_status_id=status_ids["NEW"],
    )
    db.add(order)
    db.commit()

    svc = OrderService(db)
    data = _build_workbook(
        master_rows=[["DO-PENDING-1", date(2026, 7, 1)]],
        tracking_rows=[],
    )

    result = svc.import_excel_tracking(data, "11111111-1111-1111-1111-111111111111")
    assert not result.get("errors"), result.get("errors")

    db.refresh(order)
    assert order.order_status_id == status_ids["NEW"]
    assert order.actual_delivery_date is None


def test_tracking_sheet_still_delivers_a_matched_order_in_the_same_run(db):
    """The Tracking-sheet path (the actual "mark delivered" mechanism) is untouched."""
    status_ids = _seed_statuses(db)
    order = Order(
        order_number="DO-NEW-1",
        order_date=date(2026, 7, 1),
        actual_delivery_date=None,
        order_status_id=status_ids["NEW"],
    )
    db.add(order)
    db.commit()

    svc = OrderService(db)
    data = _build_workbook(
        master_rows=[["DO-NEW-1", date(2026, 7, 1)]],
        tracking_rows=[["DO-NEW-1", date(2026, 7, 3)]],
    )

    result = svc.import_excel_tracking(data, "11111111-1111-1111-1111-111111111111")
    assert not result.get("errors"), result.get("errors")

    db.refresh(order)
    assert order.order_status_id == status_ids["DELIVERED"]
    assert order.actual_delivery_date is not None
