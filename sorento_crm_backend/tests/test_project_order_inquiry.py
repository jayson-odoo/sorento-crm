"""P10 derivation, the SCM handoff, the row state and the Excel (AC-I1, AC-I4 to AC-I7).

The netting arithmetic is proven without a database in
``test_project_order_inquiry_engine.py``. What is proven HERE is everything the engine
cannot see: that deriving writes the rows once and only once, that a pre-order on the
same project actually reaches the engine as a pool, that the ORDER row's stock location
is whatever the line was stamped with (never invented, never joined - AC-H5) and is left
empty when there is not one, that purchasing gets a task rather than an email, and that
the spreadsheet comes out with the client's own headings.

**Publish no longer derives, and neither does anything else on the SO path**
(PLAN-scm-front-planning.md section 4, AC-D01). `derive_for_sales_order` and its pool
netting are gone: the standard demand row is the confirmed Buy residual, written by
`refresh_for_decision` inside the atomic CS confirmation. The cases that pinned the SO
path's netting (pre-order pools, inbound SPO cover, the reserve window) went with the
behaviour, in the same commit; the ones about what happens AROUND the rows drive
`_confirmed_inquiry` below instead. Amendment derivation, and the netting the amendment
verbs still use, are untouched.

Postgres, blank scratch schema, rolled back at teardown. Every FK target is real: the
constraints this slice leans on are enforced, so an invented uuid aborts the transaction
rather than passing quietly.
"""
from __future__ import annotations

import io
import uuid
from datetime import date, datetime
from decimal import Decimal

import openpyxl
import pytest
from sqlalchemy import text

from app.models.inventory import Warehouse
from app.models.order import Customer, SalesOrder
from app.models.procurement import InboundShipment, SPOAllocation, Supplier
from app.models.product import Product, ProductCategory, UnitOfMeasure
from app.models.project_so import (
    INQUIRY_ACTIONED,
    INQUIRY_CANCELLED,
    INQUIRY_RAISED,
    IV_ALREADY_INBOUND,
    IV_CANCEL_BALANCE,
    IV_DELAY,
    IV_ORDER,
    IV_PRE_ORDERED,
    IV_RESERVE_AND_ORDER,
    SO_STATUS_DRAFT,
    SO_STATUS_PUBLISHED,
    OrderInquiry,
    OrderInquiryRow,
    ProjectSalesOrder,
    ProjectSalesOrderLine,
    SOAmendment,
)
from app.models.projects import TASK_LINK_ORDER_INQUIRY, ProjectTask
from app.models.user import User
from app.services import project_seed_service
from app.services.error_handler import AppException
from app.services.project_order_inquiry_service import (
    EXPORT_HEADINGS,
    EXPORT_SHEET,
    EXPORT_TITLE,
    ProjectOrderInquiryService,
)
from app.services.project_so_draft_service import ProjectSODraftService

from ._pg_fixture import blank_session

MARKER = "zzt-oi"


def _uid() -> str:
    return str(uuid.uuid4())


def _sorento(db) -> str:
    return db.execute(text("select id from companies where code = 'SRT'")).scalar()


def _user(db, name: str) -> str:
    user_id = _uid()
    db.add(User(id=user_id, email=f"{user_id}@zzt.test", name=name))
    db.flush()
    return user_id


def _product(db, code: str) -> Product:
    uom = UnitOfMeasure(id=_uid(), uom_code=f"ZZT{_uid()[:6]}", uom_name="Unit")
    category = ProductCategory(
        id=_uid(), category_code=f"ZZT-{_uid()[:8]}", category_name=f"{MARKER} cat"
    )
    db.add_all([uom, category])
    db.flush()
    row = Product(
        id=_uid(),
        product_code=code,
        product_name=f"{MARKER} {code}",
        category_id=category.id,
        base_uom_id=uom.id,
        list_price=Decimal("100.00"),
    )
    db.add(row)
    db.flush()
    return row


def _project(db, company_id: str, owner: str):
    from app.services.project_service import register_project

    return register_project(
        db,
        company_id=company_id,
        actor_user_id=owner,
        developer_party_id=None,
        title=f"{MARKER} Tuju Residences {_uid()[:6]}",
    )


def _sales_order(
    db,
    project,
    *,
    ref: str | None = None,
    is_pre_order: bool = False,
    is_sponsorship: bool = False,
    status: str = SO_STATUS_DRAFT,
    doc_no: str | None = None,
) -> ProjectSalesOrder:
    order = ProjectSalesOrder(
        id=_uid(),
        company_id=project.company_id,
        project_id=project.id,
        area_group="TOWER",
        provisional_ref=ref or f"ZZT-PSO-{_uid()[:8]}",
        autocount_doc_no=doc_no,
        status=status,
        is_pre_order=is_pre_order,
        is_sponsorship=is_sponsorship,
        grouping_origin="area",
    )
    db.add(order)
    db.flush()
    return order


def _line(db, order, product, qty, delivery_date, *, line_no=1) -> ProjectSalesOrderLine:
    row = ProjectSalesOrderLine(
        id=_uid(),
        company_id=order.company_id,
        project_sales_order_id=order.id,
        line_no=line_no,
        product_id=product.id,
        description=f"{MARKER} {product.product_code}",
        qty=Decimal(qty),
        uom="UNIT",
        unit_price=Decimal("10.00"),
        amount=Decimal(qty) * Decimal("10.00"),
        delivery_date=delivery_date,
    )
    db.add(row)
    db.flush()
    return row


def _warehouse(db, code: str) -> Warehouse:
    row = Warehouse(id=_uid(), warehouse_code=code, warehouse_name=f"{MARKER} {code}")
    db.add(row)
    db.flush()
    return row


def _inbound_spo(db, product, *, spo_number: str, qty: int, eta: date, landed: bool = False):
    supplier = Supplier(
        id=_uid(), supplier_code=f"ZZT-{_uid()[:8]}", supplier_name=f"{MARKER} supplier"
    )
    db.add(supplier)
    db.flush()
    shipment = InboundShipment(
        id=_uid(),
        shipment_number=f"ZZT-{_uid()[:8]}",
        supplier_id=supplier.id,
        shipment_date=date(2026, 5, 1),
        estimated_arrival_date=eta,
        actual_arrival_date=eta if landed else None,
        shipment_status="in_transit",
    )
    db.add(shipment)
    db.flush()
    warehouse = _warehouse(db, f"ZZT-{_uid()[:6]}")
    allocation = SPOAllocation(
        id=_uid(),
        spo_number=spo_number,
        inbound_shipment_id=shipment.id,
        warehouse_id=warehouse.id,
        product_id=product.id,
        allocated_quantity=qty,
        receipt_status="pending",
        quantity_received=0,
    )
    db.add(allocation)
    db.flush()
    return allocation


def _rows(db, inquiry_id):
    return (
        db.query(OrderInquiryRow)
        .filter(OrderInquiryRow.order_inquiry_id == inquiry_id)
        .order_by(OrderInquiryRow.created_at.asc())
        .all()
    )


@pytest.fixture()
def seeded():
    with blank_session() as db:
        company_id = _sorento(db)
        project_seed_service.run(db, company_id=company_id)
        owner = _user(db, f"{MARKER} Yana")
        yield db, company_id, owner


# ------------------------------------------------------------------ derivation

def _confirmed_inquiry(db, order, *, actor_user_id, buy=None):
    """The Buy-only handoff, written the way Stage 1C writes it.

    `derive_for_sales_order` is gone: the SO path no longer nets anything, and the only
    creator of a standard demand row is `refresh_for_decision`, inside the atomic
    confirmation (PLAN-scm-front-planning.md section 4). The cases below are about what
    happens AROUND those rows - the stock location on them, the purchasing task, the row
    states, the spreadsheet - so they need rows rather than a whole confirmation, and this
    writes them through the one writer that exists.

    ``buy`` defaults to each line's full quantity, which is what a line nothing covers
    resolves to and what these cases were originally written against.

    ``stock_location`` is read straight off ``line.stock_location`` - exactly what
    ``ProjectSupplyService._write_decision`` does after ``_restamp_stock_location`` has
    stamped it (AC-H5, single location, never a component join). Tests that want a
    location on the row set ``line.stock_location`` themselves rather than going through
    a real confirmation.
    """
    from app.models.project_so import SOSupplyDecision

    service = ProjectOrderInquiryService(db)
    lines = (
        db.query(ProjectSalesOrderLine)
        .filter(ProjectSalesOrderLine.project_sales_order_id == order.id)
        .order_by(ProjectSalesOrderLine.line_no.asc())
        .all()
    )
    revision = (
        db.query(SOSupplyDecision)
        .filter(SOSupplyDecision.project_sales_order_id == order.id)
        .count()
        + 1
    )
    decision = SOSupplyDecision(
        id=str(uuid.uuid4()),
        company_id=order.company_id,
        project_sales_order_id=order.id,
        revision_no=revision,
        # Only the first is active: two active revisions on one order is exactly what the
        # partial unique index refuses, and a fixture may not pretend otherwise.
        state="active" if revision == 1 else "superseded",
        line_snapshots=[{"line_no": line.line_no} for line in lines],
        confirmed_by=actor_user_id,
        confirmed_at=datetime.utcnow(),
    )
    db.add(decision)
    db.flush()
    result = service.refresh_for_decision(
        order,
        decision,
        [
            {
                "line": line,
                "line_no": line.line_no,
                "item_code": service._product_code(line.product_id),
                "buy_qty": Decimal(str(buy)) if buy is not None else Decimal(str(line.qty)),
                "required_date": line.delivery_date,
                "stock_location": line.stock_location,
            }
            for line in lines
        ],
        actor_user_id=actor_user_id,
    )
    return result["inquiry"]




def test_publishing_a_sales_order_raises_no_inquiry_row(seeded):
    """AC-D01 / AC-A04. Publish is not the handoff to purchasing any more.

    A published sales order has not yet been reconciled to AutoCount and CS has not yet
    composed its supply, so nothing about it is a purchase requirement: Reserve, Borrow and
    timely SPO cover may account for all of it. The inquiry is created inside atomic
    confirmation (PLAN-scm-front-planning.md section 4), and until then the whole SO reads
    **Needs CS review** and contributes zero demand.

    The derivation engine itself is unchanged and still pinned by every other test in this
    file; only its TRIGGER moved.
    """
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project)
    grating = _product(db, "CB6633")
    wc = _product(db, "SRTWC8613-RL")
    _line(db, order, grating, "600", date(2027, 1, 7), line_no=1)
    _line(db, order, wc, "135", date(2027, 2, 1), line_no=2)

    result = ProjectSODraftService(db).publish(order, actor_user_id=owner)

    assert result["status"] == SO_STATUS_PUBLISHED
    assert "order_inquiry_id" not in result
    assert (
        db.query(OrderInquiry)
        .filter(OrderInquiry.project_sales_order_id == order.id)
        .count()
        == 0
    )
    assert (
        db.query(OrderInquiryRow)
        .join(OrderInquiry, OrderInquiry.id == OrderInquiryRow.order_inquiry_id)
        .filter(OrderInquiry.project_sales_order_id == order.id)
        .count()
        == 0
    )


def test_publishing_a_sales_order_writes_no_core_sales_order(seeded):
    """Plan 5.2: publish is not a demand-class stamp point, because it stamps nothing.

    The two stamp points are the outstanding-order import and the Order Inquiry sheet
    import. Publish produces an AutoCount import FILE; the core `sales_orders` row arrives
    later, on the upload of what AutoCount made of it. A core row written here would be a
    third stamp point nobody classifies.
    """
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project)
    _line(db, order, _product(db, "CB6633"), "600", date(2027, 1, 7))
    before = db.query(SalesOrder).count()

    ProjectSODraftService(db).publish(order, actor_user_id=owner)

    db.refresh(order)
    assert order.so_id is None
    assert db.query(SalesOrder).count() == before


def test_deriving_twice_does_not_double_the_rows(seeded):
    """AC-I1 / AC-D05 idempotency: a second confirmation must not tell purchasing to buy
    it twice. The superseded row is CANCELLED rather than deleted, so the count of rows
    grows and the count of ACTIVE ones does not - the old instruction stays auditable.
    """
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project)
    _line(db, order, _product(db, "CB6633"), "600", date(2027, 1, 7))

    first = _confirmed_inquiry(db, order, actor_user_id=owner)
    second = _confirmed_inquiry(db, order, actor_user_id=owner)

    assert first.id == second.id
    active = [row for row in _rows(db, first.id) if row.state != INQUIRY_CANCELLED]
    assert len(active) == 1
    assert len(_rows(db, first.id)) == 2
    assert (
        db.query(OrderInquiry)
        .filter(OrderInquiry.project_sales_order_id == order.id)
        .count()
        == 1
    )


def test_a_stamped_line_becomes_the_stock_location(seeded):
    """AC-H5. The ORDER row's location is whatever the confirming service stamped on
    the line - ``ProjectSupplyService._restamp_stock_location``, exercised elsewhere.
    """
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project)
    line = _line(db, order, _product(db, "CB6633"), "600", date(2027, 1, 7))
    line.stock_location = "BRW-BB"
    db.flush()

    inquiry = _confirmed_inquiry(db, order, actor_user_id=owner)

    assert [row.stock_location for row in _rows(db, inquiry.id)] == ["BRW-BB"]


def test_an_unstamped_line_leaves_the_location_empty(seeded):
    """Never invented. Nobody has said where this comes from, so the column says nothing."""
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project)
    _line(db, order, _product(db, "CB6633"), "600", date(2027, 1, 7))

    inquiry = _confirmed_inquiry(db, order, actor_user_id=owner)

    assert [row.stock_location for row in _rows(db, inquiry.id)] == [None]


def test_no_allocation_at_all_leaves_the_location_empty(seeded):
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project)
    _line(db, order, _product(db, "CB6633"), "600", date(2027, 1, 7))

    inquiry = _confirmed_inquiry(db, order, actor_user_id=owner)

    assert [row.stock_location for row in _rows(db, inquiry.id)] == [None]


def test_a_multi_component_composition_still_names_one_location(seeded):
    """AC-H5: one row names ONE location, even when Reserve/Borrow drew on several
    warehouses to cover the rest of the line. That composition lives on the confirmed
    decision's snapshots, not on this column - a joined string here read as a real
    place purchasing could act on, and it never was one.
    """
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project)
    line = _line(db, order, _product(db, "CB6633"), "600", date(2027, 1, 7))
    # A multi-location composition still stamps the line's OWN fulfilment location,
    # not a join of every warehouse Reserve/Borrow happened to draw on.
    line.stock_location = "BRW-BB"
    db.flush()

    inquiry = _confirmed_inquiry(db, order, actor_user_id=owner)

    assert [row.stock_location for row in _rows(db, inquiry.id)] == ["BRW-BB"]


def test_stock_location_resolves_the_reconciled_core_lines_own_warehouse(seeded):
    """`_stock_location` (the amendment path, AC-H5) is the single fulfilment warehouse
    of the reconciled CORE line - never a join of Reserve/Borrow components.
    """
    from app.models.order import SalesOrder, SalesOrderLine

    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project)
    product = _product(db, "CB6633")
    warehouse = _warehouse(db, "BRW-IR")
    core_so = SalesOrder(
        id=_uid(), company_id=company_id, so_number=f"ZZT-CORE-{_uid()[:8]}"
    )
    db.add(core_so)
    db.flush()
    core_line = SalesOrderLine(
        id=_uid(),
        company_id=company_id,
        sales_order_id=core_so.id,
        product_id=product.id,
        warehouse_id=warehouse.id,
        qty_ordered=Decimal("10"),
        qty_delivered=Decimal("0"),
        required_date=date(2027, 1, 7),
    )
    db.add(core_line)
    db.flush()
    line = _line(db, order, product, "10", date(2027, 1, 7))
    line.core_sales_order_line_id = core_line.id
    db.flush()

    service = ProjectOrderInquiryService(db)
    assert service._stock_location(line.id) == "BRW-IR"


def test_stock_location_is_empty_with_no_reconciled_core_line(seeded):
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project)
    line = _line(db, order, _product(db, "CB6633"), "10", date(2027, 1, 7))

    service = ProjectOrderInquiryService(db)
    assert service._stock_location(line.id) is None


# ------------------------------------------------------------------ amendments


def test_publishing_an_amendment_derives_its_delta_in_purchasing_verbs(seeded):
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project, status=SO_STATUS_PUBLISHED, doc_no="SO397450")
    grating = _product(db, "CB6633")
    line = _line(db, order, grating, "570", date(2027, 1, 7))

    amendment = SOAmendment(
        id=_uid(),
        company_id=company_id,
        project_sales_order_id=order.id,
        delta_json={
            "rows": [
                {
                    "so_line_id": line.id,
                    "product_id": grating.id,
                    "product_code": "CB6633",
                    "verb": "DELAY",
                    "field": "delivery_date",
                    "from_value": "2026-07-01",
                    "to_value": "2027-01-07",
                    "qty": "570",
                },
                {
                    "so_line_id": line.id,
                    "product_id": grating.id,
                    "product_code": "CB6633",
                    "verb": "CANCEL BALANCE",
                    "field": "qty",
                    "from_value": "600",
                    "to_value": "570",
                    "qty": "30",
                },
            ]
        },
    )
    db.add(amendment)
    db.flush()

    inquiry = ProjectOrderInquiryService(db).derive_for_amendment(
        amendment, actor_user_id=owner
    )

    rows = _rows(db, inquiry.id)
    assert [(row.verb, row.qty) for row in rows] == [
        (IV_DELAY, Decimal("570.0000")),
        (IV_CANCEL_BALANCE, Decimal("30.0000")),
    ]
    # A verb on its own is not actionable: the row says what it moved from.
    assert rows[0].note == "Was 2026-07-01"
    assert rows[0].delivery_date == date(2027, 1, 7)
    assert rows[1].note == "Was 600, now 570"


def test_an_amendment_instruction_never_consumes_the_covering_pool(seeded):
    """A DELAY is about quantity purchasing already knows about."""
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    grating = _product(db, "CB6633")
    pre_order = _sales_order(
        db, project, is_pre_order=True, status=SO_STATUS_PUBLISHED, doc_no="SO383057"
    )
    _line(db, pre_order, grating, "600", date(2026, 1, 1))

    order = _sales_order(db, project, status=SO_STATUS_PUBLISHED, doc_no="SO397450")
    line = _line(db, order, grating, "600", date(2027, 1, 7))
    amendment = SOAmendment(
        id=_uid(),
        company_id=company_id,
        project_sales_order_id=order.id,
        delta_json={
            "rows": [
                {
                    "so_line_id": line.id,
                    "product_id": grating.id,
                    "product_code": "CB6633",
                    "verb": "DELAY",
                    "from_value": "2026-07-01",
                    "to_value": "2027-01-07",
                    "qty": "600",
                }
            ]
        },
    )
    db.add(amendment)
    db.flush()

    inquiry = ProjectOrderInquiryService(db).derive_for_amendment(
        amendment, actor_user_id=owner
    )

    rows = _rows(db, inquiry.id)
    assert [row.verb for row in rows] == [IV_DELAY]
    assert rows[0].covered_by is None


def test_an_amendment_derives_once_per_amendment(seeded):
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project, status=SO_STATUS_PUBLISHED)
    line = _line(db, order, _product(db, "CB6633"), "600", date(2027, 1, 7))
    amendment = SOAmendment(
        id=_uid(),
        company_id=company_id,
        project_sales_order_id=order.id,
        delta_json={
            "rows": [
                {
                    "so_line_id": line.id,
                    "product_code": "CB6633",
                    "verb": "DELAY",
                    "from_value": "2026-07-01",
                    "to_value": "2027-01-07",
                    "qty": "600",
                }
            ]
        },
    )
    db.add(amendment)
    db.flush()

    service = ProjectOrderInquiryService(db)
    first = service.derive_for_amendment(amendment, actor_user_id=owner)
    second = service.derive_for_amendment(amendment, actor_user_id=owner)

    assert first.id == second.id
    assert len(_rows(db, first.id)) == 1


# --------------------------------------------------------------- the SCM task


def test_purchasing_is_handed_a_task_with_the_rows_attached(seeded):
    """AC-I4. A task on the delivery phase pointing at the inquiry, not an email."""
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project, doc_no="SO397450")
    _line(db, order, _product(db, "CB6633"), "600", date(2027, 1, 7))

    service = ProjectOrderInquiryService(db)
    inquiry = _confirmed_inquiry(db, order, actor_user_id=owner)

    task = service.task_for(inquiry.id)
    assert task is not None
    assert task.linked_entity_type == TASK_LINK_ORDER_INQUIRY
    assert task.linked_entity_id == inquiry.id
    assert task.project_id == project.id
    assert task.task_phase == "delivery"
    assert task.category == "Purchasing"
    assert "SO397450" in task.name
    assert (
        db.query(ProjectTask)
        .filter(ProjectTask.linked_entity_id == inquiry.id)
        .count()
        == 1
    )


# ------------------------------------------------------------------ row state


def test_a_row_carries_who_actioned_it_and_when(seeded):
    """AC-I7. A state with nobody's name on it cannot answer "who do I ask"."""
    db, company_id, owner = seeded
    buyer = _user(db, f"{MARKER} purchasing")
    project = _project(db, company_id, owner)
    order = _sales_order(db, project)
    _line(db, order, _product(db, "CB6633"), "600", date(2027, 1, 7))

    service = ProjectOrderInquiryService(db)
    inquiry = _confirmed_inquiry(db, order, actor_user_id=owner)
    row_id = _rows(db, inquiry.id)[0].id

    body = service.mark_rows([row_id], state=INQUIRY_ACTIONED, actor_user_id=buyer)

    assert body[0]["state"] == INQUIRY_ACTIONED
    assert body[0]["actioned_at"] is not None
    assert body[0]["actioned_by_name"] == f"{MARKER} purchasing"
    db.refresh(inquiry)
    assert inquiry.state == INQUIRY_ACTIONED


def test_undoing_a_row_clears_the_claim_it_made(seeded):
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project)
    _line(db, order, _product(db, "CB6633"), "600", date(2027, 1, 7))

    service = ProjectOrderInquiryService(db)
    inquiry = _confirmed_inquiry(db, order, actor_user_id=owner)
    row_id = _rows(db, inquiry.id)[0].id
    service.mark_rows([row_id], state=INQUIRY_ACTIONED, actor_user_id=owner)

    body = service.mark_rows([row_id], state=INQUIRY_RAISED, actor_user_id=owner)

    assert body[0]["state"] == INQUIRY_RAISED
    assert body[0]["actioned_at"] is None
    assert body[0]["actioned_by_name"] is None
    db.refresh(inquiry)
    assert inquiry.state == INQUIRY_RAISED


def test_an_inquiry_stays_open_while_any_row_is_still_waiting(seeded):
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project)
    _line(db, order, _product(db, "CB6633"), "600", date(2027, 1, 7), line_no=1)
    _line(db, order, _product(db, "SRT382-6"), "135", date(2027, 2, 1), line_no=2)

    service = ProjectOrderInquiryService(db)
    inquiry = _confirmed_inquiry(db, order, actor_user_id=owner)
    rows = _rows(db, inquiry.id)
    service.mark_rows([rows[0].id], state=INQUIRY_ACTIONED, actor_user_id=owner)

    db.refresh(inquiry)
    assert inquiry.state == INQUIRY_RAISED

    service.mark_rows([rows[1].id], state=INQUIRY_CANCELLED, actor_user_id=owner)
    db.refresh(inquiry)
    assert inquiry.state == INQUIRY_ACTIONED


def test_an_unknown_state_is_refused(seeded):
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project)
    _line(db, order, _product(db, "CB6633"), "600", date(2027, 1, 7))
    service = ProjectOrderInquiryService(db)
    inquiry = _confirmed_inquiry(db, order, actor_user_id=owner)
    row_id = _rows(db, inquiry.id)[0].id

    with pytest.raises(AppException) as caught:
        service.mark_rows([row_id], state="done", actor_user_id=owner)

    assert caught.value.status_code == 422


# --------------------------------------------------------------------- reading


def test_the_list_serves_the_number_the_reader_knows_not_the_uuid(seeded):
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    customer = Customer(
        id=_uid(), customer_code=f"ZZT-{_uid()[:8]}", customer_name=f"{MARKER} BUIMACO"
    )
    db.add(customer)
    db.flush()
    order = _sales_order(db, project, doc_no="SO397450")
    _line(db, order, _product(db, "CB6633"), "600", date(2027, 1, 7))

    service = ProjectOrderInquiryService(db)
    _confirmed_inquiry(db, order, actor_user_id=owner)
    rows, total = service.list_rows(project.id)

    assert total == 1
    assert rows[0]["sales_order_ref"] == "SO397450"
    assert rows[0]["item_code"] == "CB6633"
    assert rows[0]["qty"] == "600"
    assert rows[0]["remark"] == "ORDER"
    assert project.title in (rows[0]["project_customer"] or "")


def test_the_list_filters_by_verb_and_by_state(seeded):
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    grating = _product(db, "CB6633")
    order = _sales_order(db, project, doc_no="SO397450")
    _line(db, order, grating, "100", date(2028, 7, 1), line_no=1)
    _line(db, order, _product(db, "SRT382-6"), "50", date(2028, 8, 3), line_no=2)

    service = ProjectOrderInquiryService(db)
    inquiry = _confirmed_inquiry(db, order, actor_user_id=owner)
    rows = _rows(db, inquiry.id)
    service.mark_rows([rows[0].id], state=INQUIRY_ACTIONED, actor_user_id=owner)

    ordered, ordered_total = service.list_rows(project.id, verb=[IV_ORDER])
    assert ordered_total == 2

    # A verb no row carries answers empty rather than answering everything.
    covered, covered_total = service.list_rows(project.id, verb=[IV_PRE_ORDERED])
    assert covered_total == 0 and covered == []

    open_rows, open_total = service.list_rows(project.id, state=[INQUIRY_RAISED])
    assert open_total == 1

    assert service.summary(project.id) == {
        "total": 2,
        "raised": 1,
        "actioned": 1,
        "cancelled": 0,
    }


def test_the_sales_order_view_carries_its_purchasing_task(seeded):
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project, doc_no="SO397450")
    _line(db, order, _product(db, "CB6633"), "600", date(2027, 1, 7))

    service = ProjectOrderInquiryService(db)
    _confirmed_inquiry(db, order, actor_user_id=owner)
    body = service.get_for_sales_order(order.id)

    assert body is not None
    assert body["task_id"]
    assert len(body["rows"]) == 1
    assert body["state"] == INQUIRY_RAISED


# ---------------------------------------------------------------------- export


def test_the_export_uses_the_clients_own_column_headings(seeded):
    """AC-I5, read off `expected-order-inquiry-2026-03-04.xlsx` in the golden set."""
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project, doc_no="SO397450")
    _line(db, order, _product(db, "CB6633"), "600", date(2027, 1, 7))

    service = ProjectOrderInquiryService(db)
    _confirmed_inquiry(db, order, actor_user_id=owner)
    filename, body = service.export_xlsx(project.id)

    assert filename.endswith(".xlsx")
    sheet = openpyxl.load_workbook(io.BytesIO(body)).worksheets[0]
    assert sheet.title == EXPORT_SHEET
    assert sheet.cell(row=1, column=1).value == EXPORT_TITLE
    headings = [cell.value for cell in sheet[2][: len(EXPORT_HEADINGS)]]
    assert headings == list(EXPORT_HEADINGS)
    assert headings == [
        "SO DATE",
        "S/O NO",
        "ITEM CODE",
        "QTY",
        "DELIVERY DATE",
        "PROJECT/CUSTOMER",
        "STOCK LOCATION",
        "REMARK",
    ]


def test_the_export_writes_the_remark_the_way_purchasing_reads_it(seeded):
    """An inbound row prints its SPO reference; every other row prints its verb.

    The inbound row is seeded rather than derived: the SO path raises Buy and nothing else
    now (section 4), and the coverage verbs it used to raise belong to an amendment. What
    is under test here is `_remark`, which is the same function either way.
    """
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    grating = _product(db, "CB6633")
    order = _sales_order(db, project, doc_no="SO397450")
    line = _line(db, order, grating, "100", date(2027, 1, 7), line_no=1)
    _line(db, order, _product(db, "SRT382-6"), "50", date(2027, 2, 1), line_no=2)

    service = ProjectOrderInquiryService(db)
    inquiry = _confirmed_inquiry(db, order, actor_user_id=owner)
    inbound = _rows(db, inquiry.id)[0]
    assert inbound.so_line_id == line.id
    inbound.verb = IV_ALREADY_INBOUND
    inbound.spo_ref = "202511-S0022"
    db.flush()
    _filename, body = service.export_xlsx(project.id)

    sheet = openpyxl.load_workbook(io.BytesIO(body)).worksheets[0]
    data = [
        (row[1], row[2], row[3], row[7])
        for row in sheet.iter_rows(min_row=3, values_only=True)
        if row[1]
    ]
    # An inbound row prints its SPO reference, exactly as the client's file does.
    assert ("SO397450", "CB6633", 100, "202511-S0022") in data
    assert ("SO397450", "SRT382-6", 50, "ORDER") in data


def test_the_export_leaves_an_unallocated_location_blank(seeded):
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project, doc_no="SO397450")
    _line(db, order, _product(db, "CB6633"), "600", date(2027, 1, 7))

    service = ProjectOrderInquiryService(db)
    _confirmed_inquiry(db, order, actor_user_id=owner)
    _filename, body = service.export_xlsx(project.id)

    sheet = openpyxl.load_workbook(io.BytesIO(body)).worksheets[0]
    first = next(sheet.iter_rows(min_row=3, values_only=True))
    assert not first[6]


def test_an_empty_project_still_exports_a_readable_sheet(seeded):
    """Nothing raised yet is a sheet with its headings, not a broken download."""
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)

    _filename, body = ProjectOrderInquiryService(db).export_xlsx(project.id)

    sheet = openpyxl.load_workbook(io.BytesIO(body)).worksheets[0]
    assert sheet.cell(row=1, column=1).value == EXPORT_TITLE
    assert [cell.value for cell in sheet[2][: len(EXPORT_HEADINGS)]] == list(EXPORT_HEADINGS)
    assert sheet.max_row == 2


# -------------------------------------------------------------------- AC-I6


def test_committed_demand_is_untouched_by_deriving_an_inquiry(seeded):
    """AC-I6. The inquiry says what to DO; it is never a second source of demand."""
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project)
    line = _line(db, order, _product(db, "CB6633"), "600", date(2027, 1, 7))
    before = (line.qty, line.delivery_date, line.product_id)

    _confirmed_inquiry(db, order, actor_user_id=owner)
    db.refresh(line)

    assert (line.qty, line.delivery_date, line.product_id) == before


# -------------------------------------------------------------------- AC-K4
#
# Pinned ahead of AC-K1, which does not build sponsorship drafts yet. The derivation reads
# no flag, so sponsorship works for FREE today -- which is exactly why it needs a test. A
# later change to the netting or the verb table could stop sponsorship stock from ever
# reaching purchasing and nothing would fail.


def test_a_sponsorship_sales_order_derives_rows_exactly_as_a_commercial_one(seeded):
    """AC-K4. Free stock still has to be bought, so purchasing still has to be told."""
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project, is_sponsorship=True)
    grating = _product(db, "CB6633")
    _line(db, order, grating, "600", date(2027, 1, 7))

    ProjectSODraftService(db).publish(order, actor_user_id=owner)
    inquiry = _confirmed_inquiry(db, order, actor_user_id=owner)

    rows = _rows(db, inquiry.id)
    assert [(row.item_code, row.qty, row.verb) for row in rows] == [
        ("CB6633", Decimal("600.0000"), IV_ORDER)
    ]
    assert {row.state for row in rows} == {INQUIRY_RAISED}


def test_a_price_zero_sponsorship_line_still_raises_an_instruction(seeded):
    """The instruction is about QUANTITY, never about money (AC-K1 prices at zero).

    A row suppressed because it was worth nothing is stock that never gets ordered.
    """
    db, company_id, owner = seeded
    project = _project(db, company_id, owner)
    order = _sales_order(db, project, is_sponsorship=True)
    product = _product(db, "SRT382-6")
    line = _line(db, order, product, "135", date(2027, 3, 1))
    line.unit_price = Decimal("0")
    line.amount = Decimal("0")
    db.flush()

    ProjectSODraftService(db).publish(order, actor_user_id=owner)
    inquiry = _confirmed_inquiry(db, order, actor_user_id=owner)

    rows = _rows(db, inquiry.id)
    assert [(row.item_code, row.qty) for row in rows] == [("SRT382-6", Decimal("135.0000"))]


