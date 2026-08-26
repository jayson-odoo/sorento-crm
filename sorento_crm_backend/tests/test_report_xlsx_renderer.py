"""The workbook the export writes (AC-D1 to AC-D4), on the synthetic dataset.

The client's file is SUMMARY plus twelve monthly tables, and so is ours: one tab per month
OF THE PERIOD, empty months included, because a register with eleven tabs reads as a lost
month rather than a quiet one. What is asserted here is the shape; the numbers are diffed
against the client's own workbook in tests/test_report_workbook_2025.py.

Three rules run through it:

- **Totals are VALUES, never formulas** (AC-D4). The workbook the client keeps by hand uses
  SUM formulas; ours writes what the engine computed, so Excel cannot recalculate a
  different answer from the screen the user exported.
- **Every sheet carries the same columns.** The tick-group members are derived once, over
  the whole period, so JAN and FEB cannot come out with different columns.
- **A blank cell is blank, never 0.00.**

Run: pytest tests/test_report_xlsx_renderer.py -q
"""
from __future__ import annotations

from io import BytesIO

import pytest

from tests import _report_fixture as fixture
from tests._pg_fixture import blank_session

SUMMARY = "SUMMARY"
MONTHS = [
    "JAN'26", "FEB'26", "MAR'26", "APR'26", "MAY'26", "JUN'26",
    "JUL'26", "AUG'26", "SEP'26", "OCT'26", "NOV'26", "DEC'26",
]

_GROUP_ROW = 6
_HEADER_ROW = 7
_FIRST_DATA_ROW = 8


@pytest.fixture
def workbook():
    from openpyxl import load_workbook

    from app.services.reports import engine
    from app.services.reports.xlsx_renderer import render_workbook

    with blank_session() as db:
        fixture.create_table(db)
        definition = fixture.definition()
        data = engine.run_workbook(db, definition, definition.default_view["params"])
        content = render_workbook(definition, data)

    assert isinstance(content, (bytes, bytearray))
    return load_workbook(BytesIO(content))


def _merged(sheet, anchor: str) -> str:
    """The merged range a top-left cell anchors, as a string, or "" if it is not merged."""
    for cell_range in sheet.merged_cells.ranges:
        if str(cell_range).startswith(f"{anchor}:"):
            return str(cell_range)
    return ""


def _headers(sheet) -> list:
    """The header of every column, wherever it sits: a single-level header is merged up
    into the group row (AC-G4), a tick column's year stays on the leaf row."""
    return [
        sheet.cell(row=_GROUP_ROW, column=index).value
        if sheet.cell(row=_HEADER_ROW, column=index).value is None
        else sheet.cell(row=_HEADER_ROW, column=index).value
        for index in range(1, sheet.max_column + 1)
    ]


def _column(sheet, header: str) -> int:
    return _headers(sheet).index(header) + 1


def _grand_total_label_column(sheet):
    """Where the GRAND TOTAL label landed on the last row of a detail sheet."""
    row = sheet.max_row
    for index in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=index).value == "GRAND TOTAL":
            return index
    return None


def _row_labels(sheet) -> list:
    """The PS-No column of every data row, excluding the GRAND TOTAL line."""
    return [
        sheet.cell(row=row, column=1).value
        for row in range(_FIRST_DATA_ROW, sheet.max_row)
    ]


# --------------------------------------------------------------------------- AC-D1


def test_the_summary_comes_first_then_one_sheet_per_month_of_the_period(workbook):
    """AC-D1. Twelve months in the period, twelve tabs, in calendar order."""
    assert workbook.sheetnames == [SUMMARY] + MONTHS


def test_a_month_with_no_rows_still_gets_its_sheet(workbook):
    """February booked nothing. The client's register still has a February."""
    sheet = workbook["FEB'26"]
    assert sheet["A2"].value == "ZZT SDN BHD"
    assert sheet["A6"].value == "ORDER NO"  # merged up into the group row (AC-G4/G8)
    assert _row_labels(sheet) == []
    assert _grand_total_label_column(sheet) is not None


def test_each_monthly_sheet_holds_only_that_months_rows(workbook):
    assert _row_labels(workbook["JAN'26"]) == ["Z-001", "Z-002"]
    assert _row_labels(workbook["MAR'26"]) == ["Z-003", "Z-004"]
    assert _row_labels(workbook["NOV'26"]) == ["Z-005"]


def test_a_row_outside_the_period_reaches_no_sheet(workbook):
    """Z-006 is a 2025 order; the period is 2026."""
    everywhere = [label for name in MONTHS for label in _row_labels(workbook[name])]
    assert "Z-006" not in everywhere


# --------------------------------------------------------------------------- AC-D2


def test_every_sheet_opens_with_the_title_block(workbook):
    """AC-D2 / AC-G7: the client's own four lines, in the client's own rows."""
    for name in [SUMMARY] + MONTHS:
        sheet = workbook[name]
        assert sheet["A1"].value is None
        assert sheet["A2"].value == "ZZT SDN BHD"
        assert sheet["A3"].value == "SCRATCH ORDERS"
        assert sheet["A5"].value == "DEPARTMENT:"
        assert sheet["C5"].value == "Scratch"
        assert sheet["A2"].font.bold and sheet["A3"].font.bold


def test_the_period_line_is_the_range_on_summary_and_a_real_date_on_a_monthly_sheet(workbook):
    """AC-G7. The client's monthly sheets carry a DATE in A4 formatted mmm-yy, which is
    what makes the month sortable and reformattable rather than a caption."""
    from datetime import date, datetime

    assert workbook[SUMMARY]["A4"].value == "Jan'26 to Dec'26"

    january = workbook["JAN'26"]["A4"]
    assert isinstance(january.value, (date, datetime))
    assert (january.value.year, january.value.month, january.value.day) == (2026, 1, 1)
    assert january.number_format == "mmm-yy"
    assert workbook["DEC'26"]["A4"].value.month == 12


# --------------------------------------------------------------------------- AC-D3


def test_the_summary_merges_each_column_value_over_its_measures(workbook):
    """AC-D3 / AC-G10: one merged month header spanning Amount and Fee, uppercase."""
    sheet = workbook[SUMMARY]
    assert sheet["A6"].value == "AGENT"
    assert _merged(sheet, "A6") == "A6:A7"
    assert sheet["B6"].value == "JAN'26"
    assert _merged(sheet, "B6") == "B6:C6"
    assert sheet["B7"].value == "AMOUNT"
    assert sheet["C7"].value == "FEE"


def test_a_monthly_sheet_merges_the_tick_group_header(workbook):
    """AC-D3: "DELIVERY YEAR" spans its 2026 and 2027 tick columns."""
    sheet = workbook["JAN'26"]
    group_row = [c.value for c in sheet[_GROUP_ROW]]
    assert "DELIVERY YEAR" in group_row
    anchor = sheet.cell(row=_GROUP_ROW, column=group_row.index("DELIVERY YEAR") + 1)
    assert _merged(sheet, anchor.coordinate)


def test_a_single_level_header_spans_both_header_rows(workbook):
    """AC-G4 / AC-G8. The client's sheet merges A6:A7 .. G6:G7; an empty band over every
    ungrouped column is the tell of a header that was drawn in two halves."""
    sheet = workbook["JAN'26"]
    assert _merged(sheet, "A6") == "A6:A7"
    assert sheet["A6"].value == "ORDER NO"
    assert sheet["A7"].value is None


def test_every_monthly_sheet_carries_the_same_columns(workbook):
    """The tick columns come from the whole period, not from one month's rows.

    January holds no 2027 delivery at all. Deriving the members per sheet would give
    January one tick column and March two, and the twelve tables would stop being the
    same table.
    """
    expected = _headers(workbook["JAN'26"])
    assert "2026" in [str(v) for v in expected] and "2027" in [str(v) for v in expected]  # noqa: E501
    for name in MONTHS:
        assert _headers(workbook[name]) == expected


def test_a_tick_reads_as_a_tick_not_as_true(workbook):
    """The client's sheet marks the delivery year with a tick, not the word TRUE."""
    sheet = workbook["JAN'26"]
    group_row = [c.value for c in sheet[_GROUP_ROW]]
    first_tick_col = group_row.index("DELIVERY YEAR") + 1
    ticks = {
        sheet.cell(row=row, column=first_tick_col).value
        for row in range(_FIRST_DATA_ROW, sheet.max_row)
    }
    assert ticks <= {"X", None}


# --------------------------------------------------------------------------- AC-D4


def test_the_summary_carries_a_row_per_row_value(workbook):
    sheet = workbook[SUMMARY]
    assert sheet["A8"].value == "Alice"
    assert sheet["A9"].value == "Bob"
    assert float(sheet["B8"].value) == 1250.25
    assert sheet["B8"].data_type == "n"


def test_a_missing_cell_reads_as_a_dash_never_as_zero(workbook):
    """AC-G9. Bob has nothing in January. The client's own sheet prints "-" there; a
    0.00 would claim he booked something and it came to nothing."""
    assert workbook[SUMMARY]["B9"].value == "-"


def _labelled_row(sheet, label: str) -> int:
    for row in range(_FIRST_DATA_ROW, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == label:
            return row
    raise AssertionError(f"no row labelled {label!r}")


def test_the_summary_column_totals_row_holds_values_not_formulas(workbook):
    """AC-D4 / AC-G10. A formula lets Excel recompute a different answer from the screen's."""
    sheet = workbook[SUMMARY]
    total_row = _labelled_row(sheet, "TOTAL")
    values = [
        sheet.cell(row=total_row, column=col).value for col in range(2, sheet.max_column + 1)
    ]
    assert float(values[-2]) == 1750.24  # the amount grand total, before the fee
    assert float(values[-1]) == 36.51


def test_the_summary_closes_with_one_labelled_total_row_per_measure(workbook):
    """AC-G10. The client's SUMMARY ends in two labelled lines carrying the year totals,
    which is the pair of numbers the whole register is kept for."""
    sheet = workbook[SUMMARY]
    amount_row = _labelled_row(sheet, "GRAND TOTAL AMOUNT JAN-DEC'26")
    fee_row = _labelled_row(sheet, "GRAND TOTAL FEE JAN-DEC'26")

    assert float(sheet.cell(row=amount_row, column=3).value) == 1750.24
    assert float(sheet.cell(row=fee_row, column=3).value) == 36.51
    assert sheet.cell(row=amount_row, column=1).font.bold
    assert _merged(sheet, f"A{amount_row}") == f"A{amount_row}:B{amount_row}"


def test_the_summary_row_total_group_closes_the_header(workbook):
    """AC-G10. The client names it TOTAL VALUE (BY SALESMAN); the kernel default is TOTAL,
    and the words are the definition's."""
    sheet = workbook[SUMMARY]
    assert sheet.cell(row=_GROUP_ROW, column=sheet.max_column - 1).value == "TOTAL"


def test_a_row_value_prints_as_it_is_stored(workbook):
    """AC-G10. Agent names are not uppercased: the register is read by the people in it."""
    assert workbook[SUMMARY]["A8"].value == "Alice"


def test_a_monthly_sheet_totals_its_own_measures(workbook):
    sheet = workbook["JAN'26"]
    total_row = sheet.max_row
    assert _grand_total_label_column(sheet) == _column(sheet, "AMOUNT") - 1
    amount_col = _column(sheet, "AMOUNT")
    fee_col = _column(sheet, "FEE")
    assert float(sheet.cell(row=total_row, column=amount_col).value) == 1250.25
    assert float(sheet.cell(row=total_row, column=fee_col).value) == 10.50


def test_the_monthly_totals_add_up_to_the_summary_grand_total(workbook):
    monthly = 0.0
    for name in MONTHS:
        sheet = workbook[name]
        amount_col = _column(sheet, "AMOUNT")
        value = sheet.cell(row=sheet.max_row, column=amount_col).value
        monthly += float(value) if value is not None else 0.0
    assert round(monthly, 2) == 1750.24


def test_no_cell_anywhere_in_the_workbook_is_a_formula(workbook):
    """AC-D4, across every sheet: nothing Excel could recalculate."""
    for name in workbook.sheetnames:
        for row in workbook[name].iter_rows():
            for cell in row:
                assert not (isinstance(cell.value, str) and cell.value.startswith("="))


def test_money_cells_carry_the_clients_accounting_format(workbook):
    """AC-G9. The RM sits left, the digits align, and a zero prints RM - like the client's."""
    from app.services.reports.xlsx_renderer import MONEY_FORMAT

    assert '"RM"' in MONEY_FORMAT and "#,##0.00" in MONEY_FORMAT
    assert workbook[SUMMARY]["B8"].number_format == MONEY_FORMAT
    sheet = workbook["JAN'26"]
    amount_col = _column(sheet, "AMOUNT")
    assert sheet.cell(row=_FIRST_DATA_ROW, column=amount_col).number_format == MONEY_FORMAT


# ------------------------------------------------------------------ AC-C5 / columns


def test_the_workbook_carries_exactly_the_views_columns_in_its_order():
    """AC-C5. What the user sees on screen is what the file holds, in that order."""
    from openpyxl import load_workbook

    from app.schemas.report import ReportViewConfig
    from app.services.reports import engine
    from app.services.reports.xlsx_renderer import render_workbook

    definition = fixture.definition()
    view = ReportViewConfig.model_validate(
        {
            "params": definition.default_view["params"],
            "detail": {"columns": ["amount", "order_no", "agent"], "order": []},
            "pivot": definition.default_view["pivot"],
        }
    )
    with blank_session() as db:
        fixture.create_table(db)
        data = engine.run_workbook(db, definition, definition.default_view["params"], view)
        content = render_workbook(definition, data)

    sheet = load_workbook(BytesIO(content))["JAN'26"]
    assert _headers(sheet) == ["AMOUNT", "ORDER NO", "AGENT"]


def _rendered(columns):
    """The JAN'26 sheet of a workbook rendered with exactly these detail columns."""
    from openpyxl import load_workbook

    from app.schemas.report import ReportViewConfig
    from app.services.reports import engine
    from app.services.reports.xlsx_renderer import render_workbook

    definition = fixture.definition()
    view = ReportViewConfig.model_validate(
        {
            "params": definition.default_view["params"],
            "detail": {"columns": list(columns), "order": []},
            "pivot": definition.default_view["pivot"],
        }
    )
    with blank_session() as db:
        fixture.create_table(db)
        data = engine.run_workbook(db, definition, definition.default_view["params"], view)
        content = render_workbook(definition, data)
    return load_workbook(BytesIO(content))["JAN'26"]


def test_the_grand_total_label_never_overwrites_a_money_total():
    """A view whose FIRST column is a measure has nowhere to the left to put the label, so
    it takes the first cell that carries no total rather than writing over the money."""
    sheet = _rendered(["amount", "order_no", "agent"])
    total_row = sheet.max_row

    values = [sheet.cell(row=total_row, column=c).value for c in range(1, sheet.max_column + 1)]
    assert "GRAND TOTAL" in values
    assert float(sheet.cell(row=total_row, column=1).value) == 1250.25


def test_the_grand_total_label_sits_in_the_column_before_the_first_measure():
    """AC-G9. The client types GRAND TOTAL beside the money, not at the far end of a row
    the reader has already scrolled past."""
    sheet = _rendered(["order_no", "agent", "amount"])
    assert _grand_total_label_column(sheet) == _column(sheet, "AMOUNT") - 1
    assert sheet.cell(row=sheet.max_row, column=_column(sheet, "AGENT")).font.bold


def test_the_grand_total_label_stays_inside_the_table_when_every_column_is_money():
    """AC-G9. A view of nothing but measures leaves no cell in the row free for the label,
    and it used to be written one column PAST the table: outside the border, outside the
    print area, and the first thing a reader loses. It takes a bordered row of its own
    directly above the totals instead, and the money stays where it was."""
    sheet = _rendered(["amount", "fee"])
    money_row = sheet.max_row
    label_row = money_row - 1

    assert sheet.cell(row=label_row, column=1).value == "GRAND TOTAL"
    assert sheet.cell(row=label_row, column=1).font.bold
    # Inside the table: every cell of the label row is bordered, like the rest.
    for column in range(1, sheet.max_column + 1):
        assert sheet.cell(row=label_row, column=column).border.left.style == "thin"
        assert sheet.cell(row=money_row, column=column).border.left.style == "thin"
    # Nothing was written over, and nothing landed past the last column.
    assert float(sheet.cell(row=money_row, column=1).value) == 1250.25
    assert float(sheet.cell(row=money_row, column=2).value) == 10.50
    assert sheet.cell(row=money_row, column=3).value is None
    assert sheet.cell(row=label_row, column=3).value is None


def test_a_date_column_is_a_real_date_not_a_string():
    """A text date cannot be sorted, filtered or formatted in Excel, which is most of what
    somebody opens the file to do."""
    from datetime import date, datetime

    sheet = _rendered(["order_no", "booked_on"])
    cell = sheet.cell(row=_FIRST_DATA_ROW, column=2)

    assert isinstance(cell.value, (date, datetime))
    assert cell.value.year == 2026 and cell.value.month == 1
    assert cell.number_format == "DD/MM/YYYY"


def test_an_empty_column_list_falls_back_to_the_definitions_default_columns():
    """A twenty-column sheet is unreadable, so a workbook never means the whole catalog.

    The SCREEN asks for the whole catalog and hides client-side, which is what makes
    ticking a column instant (AC-B7). A FILE has no Columns panel, so an unset view means
    the columns the report was designed around.
    """
    from openpyxl import load_workbook

    from app.schemas.report import ReportViewConfig
    from app.services.reports import engine
    from app.services.reports.xlsx_renderer import render_workbook

    definition = fixture.definition()
    definition.default_view["detail"]["columns"] = ["order_no", "amount"]
    view = ReportViewConfig.model_validate(
        {
            "params": definition.default_view["params"],
            "detail": {"columns": [], "order": []},
            "pivot": definition.default_view["pivot"],
        }
    )
    with blank_session() as db:
        fixture.create_table(db)
        data = engine.run_workbook(db, definition, definition.default_view["params"], view)
        content = render_workbook(definition, data)

    sheet = load_workbook(BytesIO(content))["JAN'26"]
    assert _headers(sheet) == ["ORDER NO", "AMOUNT"]


# --------------------------------------------------------------------------- AC-G7


def test_the_company_name_comes_from_system_settings():
    """AC-G7. A definition that names no company takes the letterhead from Settings.
    (A definition that does name one wins, because the live install's settings row still
    reads "Metronic"; see test_the_definition_names_the_company_when_settings_does_not.)"""
    from dataclasses import replace

    from openpyxl import load_workbook

    from app.models.user import SystemSetting
    from app.services.reports import engine
    from app.services.reports.xlsx_renderer import render_workbook

    base = fixture.definition()
    definition = replace(base, workbook=replace(base.workbook, company_name=""))
    with blank_session() as db:
        fixture.create_table(db)
        db.add(SystemSetting(id="zzt-settings", name="ZZT Holdings Berhad"))
        db.flush()
        data = engine.run_workbook(db, definition, definition.default_view["params"])
        content = render_workbook(definition, data)

    assert load_workbook(BytesIO(content))[SUMMARY]["A2"].value == "ZZT HOLDINGS BERHAD"


def test_the_definition_names_the_company_when_settings_does_not(workbook):
    """The definition's name is the letterhead whenever it has one; this is also what
    every test in this file renders under (the blank schema holds no settings row)."""
    assert workbook[SUMMARY]["A2"].value == "ZZT SDN BHD"


def test_the_definition_name_beats_a_settings_row_that_still_says_metronic():
    """The live install's system_settings.name is the template default; a definition that
    names the client must not be overridden by it."""
    from openpyxl import load_workbook

    from app.models.user import SystemSetting
    from app.services.reports import engine
    from app.services.reports.xlsx_renderer import render_workbook

    definition = fixture.definition()
    with blank_session() as db:
        fixture.create_table(db)
        db.add(SystemSetting(id="zzt-settings", name="Metronic"))
        db.flush()
        data = engine.run_workbook(db, definition, definition.default_view["params"])
        wb = load_workbook(BytesIO(render_workbook(definition, data)))
    assert wb[SUMMARY]["A2"].value == "ZZT SDN BHD"
