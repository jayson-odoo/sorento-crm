"""Route-level tests for the Stock List `.xlsm` macro upload pipeline.

Covers POST /api/v1/resource-management/attachments/replace-latest-stock-list
and the generic attachments POST with a Stock List type: macro uploads are
stripped of VBA and reduced to a values-only Template sheet before storage
(docs/plans/PLAN-stock-list-xlsm-macro-upload.md).

Auth bypass / sqlite-tables pattern copied from test_upload_activity_endpoint.py.
"""
from __future__ import annotations

import io
import zipfile

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

# MUST be first app import — resolves circular-import in app.modules.runtime.guards
from app.main import app  # noqa: E402


_USER_ID = "130c548f-048f-53b2-97a6-3a54676bea77"
_ROLE_ID = "7c50d6db-8dce-555a-85a2-86cf7756f33f"
# NB: must contain letters — sqlite's NUMERIC affinity on the UUID column
# coerces an all-digit uuid string to float and breaks row loading.
_TYPE_ID = "4b8bd2aa-c3a9-5b0d-b02a-4538c4bd2e04"
XLSM_MIME = "application/vnd.ms-excel.sheet.macroEnabled.12"


def _seed_user(db: Session) -> None:
    from app.models.user import User, UserRole, UserRoleAssignment

    db.add(
        UserRole(
            id=_ROLE_ID,
            slug="superadmin",
            name="Superadmin",
            description="",
            is_protected=True,
            is_default=False,
        )
    )
    db.flush()
    db.add(User(id=_USER_ID, email="u1@test.com", name="U1", status="ACTIVE"))
    db.flush()
    db.add(UserRoleAssignment(user_id=_USER_ID, role_id=_ROLE_ID))
    db.commit()


def _seed_stock_list_type(db: Session) -> None:
    from app.models.resources import AttachmentType

    db.add(
        AttachmentType(
            id=_TYPE_ID,
            type_name="Stock_List",
            allowed_extensions="xls,xlsx,xlsm",
            max_file_size_mb=10,
        )
    )
    db.commit()


class _FakeStorageBackend:
    """Captures the bytes the route would push to S3/R2."""

    def __init__(self) -> None:
        self.uploads: list[tuple[str, bytes, str | None]] = []

    def upload_file(self, *, file_content: bytes, file_path: str, content_type=None):
        self.uploads.append((file_path, file_content, content_type))
        return file_path, f"https://cdn.test/{file_path}"


@pytest.fixture
def client(monkeypatch):
    from app.dependencies import get_current_user, get_current_user_or_api_key, get_db
    from app.models.user import User, UserRole, UserRoleAssignment
    from app.models.resources import Attachment, AttachmentType
    from app.models.lookup import LookupBinding  # validator listener checks this table
    from app.models.embeddings import EmbeddingQueue  # change-listener inserts on Attachment create

    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )

    # Swap JSONB columns to generic JSON so SQLite create_all works.
    from sqlalchemy.dialects.postgresql import JSONB
    from sqlalchemy.types import JSON as GenericJSON

    for model in (Attachment, AttachmentType, EmbeddingQueue):
        for col in list(model.__table__.columns):
            if isinstance(col.type, JSONB):
                col.type = GenericJSON()

    from app.database import Base

    Base.metadata.create_all(
        engine,
        tables=[
            User.__table__,
            UserRole.__table__,
            UserRoleAssignment.__table__,
            AttachmentType.__table__,
            Attachment.__table__,
            LookupBinding.__table__,
            EmbeddingQueue.__table__,
        ],
    )

    SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    db = SessionLocal()
    _seed_user(db)
    _seed_stock_list_type(db)

    # Storage + side-channel stubs: capture uploads, skip webhook + access-level lookups.
    fake_backend = _FakeStorageBackend()
    import app.services.storage_router as storage_router
    import app.api.v1.resources.attachments as attachments_module
    from app.services.contact_access_type_service import ContactAccessTypeService

    monkeypatch.setattr(storage_router, "default_provider", lambda: "s3")
    monkeypatch.setattr(storage_router, "get_backend", lambda provider: fake_backend)
    monkeypatch.setattr(
        storage_router, "cdn_base_url", lambda provider, key: f"https://cdn.test/{key}"
    )
    monkeypatch.setattr(
        attachments_module, "_create_and_send_webhook", lambda *a, **k: None
    )
    monkeypatch.setattr(
        ContactAccessTypeService, "get_default_access_levels", lambda self: ["dealer"]
    )
    monkeypatch.setattr(
        ContactAccessTypeService,
        "validate_access_levels",
        lambda self, levels, field_name="access_levels": list(levels),
    )

    def _override_get_db():
        yield db

    def _override_current_user():
        return {"id": _USER_ID, "email": "u1@test.com"}

    app.dependency_overrides[get_db] = _override_get_db
    app.dependency_overrides[get_current_user] = _override_current_user
    app.dependency_overrides[get_current_user_or_api_key] = _override_current_user

    with TestClient(app) as c:
        yield c, db, fake_backend

    app.dependency_overrides.clear()
    db.close()


def _xlsm_bytes(*, with_template: bool = True) -> bytes:
    wb = Workbook()
    wb.active.title = "Active Loc"
    wb.active["A1"] = "locations"
    wb.create_sheet("Master")["A1"] = "macro-ui"
    if with_template:
        ws = wb.create_sheet("Template")
        ws.append(["Item Code", "Item Description", "Location", "On Hand Qty"])
        ws.append(["ABC-123", "Widget", "BRW", 10])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_bytes() -> bytes:
    wb = Workbook()
    wb.active.append(["Item Code", "On Hand Qty"])
    wb.active.append(["ABC-123", 5])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


REPLACE_URL = "/api/v1/resource-management/attachments/replace-latest-stock-list"
CREATE_URL = "/api/v1/resource-management/attachments/"


def test_replace_latest_xlsm_is_converted_to_template_only_xlsx(client):
    c, db, backend = client

    resp = c.post(
        REPLACE_URL,
        files={"file": ("stock balance - Macro Version.xlsm", _xlsm_bytes(), XLSM_MIME)},
    )
    assert resp.status_code == 201, resp.text
    body = resp.json()
    assert body["original_filename"] == "stock balance - Macro Version.xlsx"
    assert body["mime_type"].endswith("spreadsheetml.sheet")

    assert len(backend.uploads) == 1
    _, stored_bytes, stored_mime = backend.uploads[0]
    assert stored_mime is not None and stored_mime.endswith("spreadsheetml.sheet")
    assert not any(
        "vba" in n.lower() for n in zipfile.ZipFile(io.BytesIO(stored_bytes)).namelist()
    )
    wb = load_workbook(io.BytesIO(stored_bytes))
    assert wb.sheetnames == ["Template"]
    assert wb["Template"]["A2"].value == "ABC-123"


def test_replace_latest_xlsm_without_template_sheet_is_422(client):
    c, db, backend = client

    resp = c.post(
        REPLACE_URL,
        files={"file": ("bad.xlsm", _xlsm_bytes(with_template=False), XLSM_MIME)},
    )
    assert resp.status_code == 422, resp.text
    assert "Template" in resp.json()["detail"]
    assert backend.uploads == []


def test_replace_latest_xlsx_passes_through_unchanged(client):
    c, db, backend = client

    raw = _xlsx_bytes()
    resp = c.post(
        REPLACE_URL,
        files={
            "file": (
                "stock.xlsx",
                raw,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert resp.status_code == 201, resp.text
    assert resp.json()["original_filename"] == "stock.xlsx"
    assert backend.uploads[0][1] == raw  # byte-identical: no rewrite for non-xlsm


def test_replace_latest_requires_auth():
    # No dependency overrides: bare request must be rejected.
    with TestClient(app) as c:
        resp = c.post(
            REPLACE_URL, files={"file": ("stock.xlsm", b"x", XLSM_MIME)}
        )
    assert resp.status_code in (401, 403)


def test_generic_create_attachment_converts_stock_list_xlsm(client):
    c, db, backend = client

    resp = c.post(
        CREATE_URL,
        data={"attachment_type_id": _TYPE_ID},
        files={"file": ("stock balance - Macro Version.xlsm", _xlsm_bytes(), XLSM_MIME)},
    )
    assert resp.status_code in (200, 201), resp.text
    body = resp.json()
    assert body["original_filename"] == "stock balance - Macro Version.xlsx"

    _, stored_bytes, _ = backend.uploads[0]
    wb = load_workbook(io.BytesIO(stored_bytes))
    assert wb.sheetnames == ["Template"]


def test_generic_create_attachment_xlsm_without_template_is_422(client):
    c, db, backend = client

    resp = c.post(
        CREATE_URL,
        data={"attachment_type_id": _TYPE_ID},
        files={"file": ("bad.xlsm", _xlsm_bytes(with_template=False), XLSM_MIME)},
    )
    assert resp.status_code == 422, resp.text
    assert backend.uploads == []
