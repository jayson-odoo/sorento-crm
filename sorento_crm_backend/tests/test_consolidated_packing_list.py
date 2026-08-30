"""S10 - the Sorento packing list: one container, every factory that loaded it.

What Ms Tee used to build by hand: the lines of two or three suppliers under one container,
each factory subtotalled, a grand total, the SORENTO / MOCHA split, and a remarks column that
already says where the shipment differs from the loading plan that supplier was sent.

Everything asserted here is DERIVED. The factory is the supplier the line was uploaded as, the
company is the product's brand, the discrepancies are the latest supplier notice compared with
what actually arrived. None of it is typed in, so the test is about the derivation being right
rather than about a form being saved.

Postgres only, on a blank schema, seeding every FK target itself: CI's database is empty, so a
test that borrows an existing category, brand or supplier passes locally and fails there.
"""
from __future__ import annotations

import uuid
from datetime import date, datetime
from io import BytesIO
from urllib.parse import quote

import pytest
from fastapi.testclient import TestClient

from app.dependencies import get_current_user, get_current_user_or_api_key, get_db
from app.main import app
from app.models.procurement import InboundShipment, InboundShipmentLine, Supplier
from app.models.product import Brand, Product, ProductCategory, UnitOfMeasure
from app.models.scm import LoadingPlan
from app.models.supplier_notice import SupplierNotice, SupplierNoticeLine
from app.services.error_handler import AppException
from app.services.scm import consolidated_packing_list as svc
from tests._pg_fixture import blank_session

MARKER = "ZZCPL"
_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture
def db():
    with blank_session() as session:
        yield session


class World:
    """Two factories on one container, plus a line whose factory we were never told.

    Deliberately mixed: one MOCHA-brand product (the company split), one line the supplier
    stated a volume for and one that only the catalogue can measure (the cbm fallback), one
    line with no volume anywhere (so a partial cbm is visible as partial), and a supplier
    notice for the first factory that asked for more of one product, less of another, and one
    product that never turned up at all.
    """

    def __init__(self, db):
        self.db = db
        self.tag = uuid.uuid4().hex[:8].upper()

        self.category = ProductCategory(
            id=str(uuid.uuid4()),
            category_code=f"{MARKER}-CAT-{self.tag}",
            category_name=f"{MARKER} category",
        )
        self.uom = UnitOfMeasure(
            id=str(uuid.uuid4()), uom_code=f"{MARKER}{self.tag}"[:20], uom_name="pcs"
        )
        self.mocha = Brand(
            id=str(uuid.uuid4()), brand_code="MOCHA", brand_name="Mocha", is_active=True
        )
        self.sandel = Brand(
            id=str(uuid.uuid4()),
            brand_code=f"SANDEL-{self.tag}"[:50],
            brand_name="Sandel",
            is_active=True,
        )
        db.add_all([self.category, self.uom, self.mocha, self.sandel])
        db.flush()

        # Codes carry the shared tag then the sort key, so "ordered by product code" is a
        # deterministic assertion rather than whatever uuid4 happened to produce.
        self.tap = self.product("1TAP", brand=self.sandel)
        self.mocha_basin = self.product("2BASIN", brand=self.mocha, dims=(500, 400, 300))
        self.mat = self.product("3MAT")
        self.sink = self.product("4SINK")
        self.orphan = self.product("5ORPHAN")
        self.never_packed = self.product("6GONE")

        # Ordered by supplier NAME, so A sorts before B whatever their ids are.
        self.kailu = self.supplier("A-KAILU")
        self.caizhou = self.supplier("B-CAIZHOU")

        self.shipment = InboundShipment(
            id=str(uuid.uuid4()),
            shipment_number=f"{MARKER}-SH-{self.tag}",
            shipment_date=date(2026, 8, 1),
            shipping_container_number=f"{MARKER}U{self.tag}",
            bill_of_lading_number=f"BL-{self.tag}",
            shipment_status="in_transit",
        )
        db.add(self.shipment)
        db.flush()

        # The tap is the fully MEASURED line: material, pack size, carton and both weights,
        # which is what the container workbook derives every one of its figures from. The
        # rest are deliberately bare, because a container read off a PDF states none of it
        # and the same sheet has to print that container too.
        self.line(
            self.kailu,
            self.tap,
            qty=490,
            cartons=86,
            cbm="2.1053",
            remarks="fragile",
            material="不锈钢",
            pcs_per_carton=10,
            carton_length_cm=34,
            carton_width_cm=24,
            carton_height_cm=30,
            net_weight_per_carton="7.000",
            gross_weight_per_carton="8.300",
            unit_cost="65.50",
            currency="CNY",
        )
        self.line(self.kailu, self.mocha_basin, qty=100, cartons=10, unit_cost="100.00")
        self.line(self.kailu, self.mat, qty=120, cartons=12)
        self.line(self.caizhou, self.sink, qty=50, cartons=5, cbm="1.5000")
        self.line(None, self.orphan, qty=7, cartons=1)

    # -- seeding ---------------------------------------------------------- #

    def costed(self) -> None:
        """The container as somebody who has typed its paperwork up would see it.

        Separate from `__init__` so the bare container stays the default: an untyped
        header and blank costs are the state every container starts in, and the sheet has
        to be printable then too.
        """
        self.shipment.loading_date = date(2026, 7, 17)
        self.shipment.etd_date = date(2026, 7, 23)
        self.shipment.estimated_arrival_date = date(2026, 7, 25)
        self.shipment.eta_delay_date = date(2026, 7, 27)
        self.shipment.seal_number = "J0713349"
        self.shipment.forwarder_order_ref = "CNH1098313"
        self.shipment.consignee = "SORENTO SDN BHD"
        self.shipment.shipper = "SHENZHEN XINDESHENG TRADING CO.,LTD"
        self.shipment.china_forwarder = "ONE TOUCH"
        self.shipment.free_days_available = 14
        self.shipment.delivery_warehouse = "BRW"
        self.shipment.clearance_cost = 2700
        self.shipment.china_freight_cost = 13950
        self.shipment.insurance_rate = 1
        self.db.flush()

    def product(self, key: str, *, brand: Brand | None = None, dims=None) -> Product:
        p = Product(
            id=str(uuid.uuid4()),
            product_code=f"{MARKER}-{self.tag}-{key}",
            product_name=f"{key} product",
            category_id=self.category.id,
            base_uom_id=self.uom.id,
            brand_id=brand.id if brand else None,
            list_price=0,
            is_active=True,
        )
        if dims:
            p.dimensions_length, p.dimensions_width, p.dimensions_height = dims
        self.db.add(p)
        self.db.flush()
        return p

    def supplier(self, name: str) -> Supplier:
        s = Supplier(
            id=str(uuid.uuid4()),
            supplier_code=f"{MARKER}-{name}-{self.tag}"[:50],
            supplier_name=f"{MARKER} {name}",
            is_active=True,
        )
        self.db.add(s)
        self.db.flush()
        return s

    def line(self, supplier, product, *, qty, cartons, cbm=None, remarks=None, **measured):
        """One line on the container. `measured` is whatever the supplier's file stated.

        Left out by default so the "nobody measured this" path stays the common case in
        these tests: a container read off a PDF has none of it, and the workbook has to
        print that container too.
        """
        row = InboundShipmentLine(
            id=str(uuid.uuid4()),
            shipment_id=self.shipment.id,
            supplier_id=supplier.id if supplier else None,
            product_id=product.id,
            quantity_shipped=qty,
            cartons_count=cartons,
            cbm=cbm,
            remarks=remarks,
            **measured,
        )
        self.db.add(row)
        self.db.flush()
        return row

    def notice(
        self,
        supplier,
        lines,
        *,
        created_at: datetime,
        plan: bool = True,
        kind: str = "pack",
        sent_at: datetime | None = None,
    ):
        """One notice with its lines, as the supplier received it.

        `kind` is `pack` (goods the factory holds and was asked to load) unless a test wants
        `produce` (goods it still has to make), which is not a pack plan at all.
        """
        loading_plan = None
        if plan:
            loading_plan = LoadingPlan(
                id=str(uuid.uuid4()),
                supplier_id=supplier.id,
                container_count=1,
                container_cbm=68,
                capacity_cbm=68,
            )
            self.db.add(loading_plan)
            self.db.flush()
        row = SupplierNotice(
            id=str(uuid.uuid4()),
            supplier_id=supplier.id,
            loading_plan_id=loading_plan.id if loading_plan else None,
            channel="email",
            status="sent",
            created_at=created_at,
            sent_at=sent_at,
        )
        self.db.add(row)
        self.db.flush()
        for i, (product, qty) in enumerate(lines):
            self.db.add(
                SupplierNoticeLine(
                    id=str(uuid.uuid4()),
                    notice_id=row.id,
                    product_id=product.id,
                    item_code=product.product_code,
                    product_name=product.product_name,
                    qty=qty,
                    kind=kind,
                    sort_order=i,
                )
            )
        self.db.flush()
        return row

    def notice_for_kailu(self) -> SupplierNotice:
        """The plan Kailu was sent: more taps, fewer mats, and a product that never came.

        An older notice is seeded first so "the latest one" is a claim under test rather than
        an accident of there being only one.
        """
        self.notice(
            self.kailu,
            [(self.tap, 1000)],
            created_at=datetime(2026, 7, 1, 9, 0, 0),
        )
        return self.notice(
            self.kailu,
            [(self.tap, 500), (self.mat, 100), (self.never_packed, 100)],
            created_at=datetime(2026, 8, 1, 9, 0, 0),
        )


def _factory(payload: dict, name_fragment: str) -> dict:
    return next(f for f in payload["factories"] if name_fragment in (f["supplier_name"] or ""))


def _line(factory: dict, code_suffix: str) -> dict:
    return next(l for l in factory["lines"] if l["product_code"].endswith(code_suffix))


# --------------------------------------------------------------------------- #
# grouping and arithmetic
# --------------------------------------------------------------------------- #


def test_the_container_is_grouped_by_factory_with_the_unknown_one_last(db):
    # The header of a mixed container names nobody, so the factory has to come off the line.
    w = World(db)

    out = svc.build(db, str(w.shipment.id))

    assert out["shipment_id"] == str(w.shipment.id)
    assert out["container_no"] == w.shipment.shipping_container_number
    assert out["bl_no"] == w.shipment.bill_of_lading_number
    assert out["status"] == "in_transit"
    assert [f["supplier_name"] for f in out["factories"]] == [
        w.kailu.supplier_name,
        w.caizhou.supplier_name,
        "Unassigned",
    ]
    assert out["factories"][-1]["supplier_id"] is None
    # Within a factory, the order of the sheet is the order of the model numbers.
    assert [l["product_code"] for l in out["factories"][0]["lines"]] == [
        w.tap.product_code,
        w.mocha_basin.product_code,
        w.mat.product_code,
    ]


def test_each_factory_subtotals_its_own_lines_and_says_how_much_volume_it_knows(db):
    w = World(db)

    out = svc.build(db, str(w.shipment.id))

    kailu = _factory(out, "A-KAILU")
    assert kailu["subtotal"]["lines"] == 3
    assert kailu["subtotal"]["qty"] == 710
    assert kailu["subtotal"]["cartons"] == 108
    # 2.1053 stated by the supplier + 6.0 measured from the catalogue. The mat has no
    # volume anywhere, so two of three lines are known and the figure says so.
    assert kailu["subtotal"]["cbm"] == pytest.approx(8.1053)
    assert kailu["subtotal"]["cbm_known_lines"] == 2

    unassigned = _factory(out, "Unassigned")
    assert unassigned["subtotal"] == {
        "lines": 1,
        "qty": 7,
        "cartons": 1,
        "cbm": 0.0,
        "cbm_known_lines": 0,
    }


def test_the_grand_total_is_the_whole_container(db):
    w = World(db)

    out = svc.build(db, str(w.shipment.id))

    assert out["total"]["lines"] == 5
    assert out["total"]["qty"] == 767
    assert out["total"]["cartons"] == 114
    assert out["total"]["cbm"] == pytest.approx(9.6053)
    assert out["total"]["cbm_known_lines"] == 3


def test_a_missing_volume_falls_back_to_the_catalogue_and_then_to_nothing(db):
    # A cbm invented for a product nobody measured would be worse than an empty cell.
    w = World(db)

    out = svc.build(db, str(w.shipment.id))
    kailu = _factory(out, "A-KAILU")

    assert _line(kailu, "1TAP")["cbm"] == pytest.approx(2.1053)
    assert _line(kailu, "2BASIN")["cbm"] == pytest.approx(6.0)  # 500*400*300mm x 100
    assert _line(kailu, "3MAT")["cbm"] is None


def test_the_company_split_is_read_off_the_brand_and_always_prints_both_rows(db):
    w = World(db)

    out = svc.build(db, str(w.shipment.id))

    kailu = _factory(out, "A-KAILU")
    assert _line(kailu, "2BASIN")["company"] == "MOCHA"
    assert _line(kailu, "2BASIN")["brand"] == "MOCHA"
    assert _line(kailu, "1TAP")["company"] == "SORENTO"
    # No brand at all is Sorento's, the same as SANDEL / CABANA are.
    assert _line(_factory(out, "B-CAIZHOU"), "4SINK")["company"] == "SORENTO"

    split = {row["company"]: row for row in out["split"]}
    assert [row["company"] for row in out["split"]] == ["SORENTO", "MOCHA"]
    assert split["SORENTO"]["lines"] == 4
    assert split["SORENTO"]["qty"] == 667
    assert split["SORENTO"]["cartons"] == 104
    assert split["SORENTO"]["cbm"] == pytest.approx(3.6053)
    assert split["SORENTO"]["cbm_known_lines"] == 2
    assert split["MOCHA"]["lines"] == 1
    assert split["MOCHA"]["qty"] == 100
    assert split["MOCHA"]["cartons"] == 10
    assert split["MOCHA"]["cbm"] == pytest.approx(6.0)
    assert split["MOCHA"]["cbm_known_lines"] == 1


def test_both_split_rows_are_present_even_when_one_company_shipped_nothing(db):
    w = World(db)
    db.query(InboundShipmentLine).filter(
        InboundShipmentLine.product_id == w.mocha_basin.id
    ).delete(synchronize_session=False)

    out = svc.build(db, str(w.shipment.id))

    mocha = next(row for row in out["split"] if row["company"] == "MOCHA")
    assert mocha == {
        "company": "MOCHA",
        "lines": 0,
        "qty": 0,
        "cartons": 0,
        "cbm": 0.0,
        "cbm_known_lines": 0,
    }


# --------------------------------------------------------------------------- #
# R20 - the document prints the shipment, and no comparison against the plan
# --------------------------------------------------------------------------- #


def test_the_payload_compares_the_container_against_no_loading_plan(db):
    """It used to. The remarks column carried "Loading plan asked 500, packed 490", the
    factory carried the notice it was measured against, and a model the plan asked for and
    nobody loaded was listed under the block.

    All of it is gone (R20). This is the document the forwarder, the factories and the
    clearance agent read, and a line on it for goods that never went into the container is
    read as goods that shipped. The supplier notice itself is untouched - what was removed is
    printing OUR reading of it on THEIR packing list.
    """
    w = World(db)
    w.notice_for_kailu()

    out = svc.build(db, str(w.shipment.id))

    for factory in out["factories"]:
        assert "not_packed" not in factory
        assert "notice_id" not in factory
        assert "loading_plan_id" not in factory
        assert "has_pack_plan" not in factory
        assert "notice_created_at" not in factory
        assert "notice_sent_at" not in factory
        for line in factory["lines"]:
            assert "discrepancies" not in line


def test_the_remarks_are_the_suppliers_own_words_and_only_those(db):
    w = World(db)
    w.notice_for_kailu()

    out = svc.build(db, str(w.shipment.id))
    kailu = _factory(out, "A-KAILU")

    assert _line(kailu, "1TAP")["remarks"] == "fragile"
    assert _line(kailu, "3MAT")["remarks"] is None


# --------------------------------------------------------------------------- #
# the edges
# --------------------------------------------------------------------------- #


def test_a_container_with_no_lines_is_an_empty_list_not_a_404(db):
    # "Read but not yet unpacked" is a state the screen has to be able to draw.
    w = World(db)
    db.query(InboundShipmentLine).filter(
        InboundShipmentLine.shipment_id == w.shipment.id
    ).delete(synchronize_session=False)

    out = svc.build(db, str(w.shipment.id))

    assert out["factories"] == []
    assert out["total"] == {"lines": 0, "qty": 0, "cartons": 0, "cbm": 0.0, "cbm_known_lines": 0}
    assert [row["company"] for row in out["split"]] == ["SORENTO", "MOCHA"]
    assert all(row["qty"] == 0 for row in out["split"])


def test_an_unknown_container_is_a_404(db):
    with pytest.raises(AppException) as err:
        svc.build(db, str(uuid.uuid4()))

    assert err.value.status_code == 404


def test_an_id_that_is_not_an_id_is_a_404_not_a_broken_session(db):
    with pytest.raises(AppException) as err:
        svc.build(db, "not-a-uuid")

    assert err.value.status_code == 404
    # The session is still usable: the refusal came before the UUID column was ever compared.
    assert db.query(InboundShipment).count() == 0


# --------------------------------------------------------------------------- #
# the workbook
# --------------------------------------------------------------------------- #


def _sheet(payload: dict):
    """The workbook as openpyxl sees it: formulas as formulas, values as values."""
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(svc.to_xlsx(payload)))
    assert wb.sheetnames == ["RMB"]
    return wb["RMB"]


def _cells(payload: dict) -> list[list]:
    return [list(row) for row in _sheet(payload).iter_rows(values_only=True)]


def _column(ws, letter: str) -> list:
    return [ws[f"{letter}{r}"].value for r in range(1, ws.max_row + 1)]


def _row_of(ws, letter: str, value) -> int:
    """The row whose cell in `letter` equals `value`. The tests read by content, not index."""
    for r in range(1, ws.max_row + 1):
        if ws[f"{letter}{r}"].value == value:
            return r
    raise AssertionError(f"no row with {letter} = {value!r}")


def test_the_header_block_prints_the_container_and_its_paperwork(db):
    # Twelve labelled lines above the goods. Every one of them is a column on the
    # container, so a workbook missing one is a column that never reached the sheet.
    w = World(db)
    w.costed()
    payload = svc.build(db, str(w.shipment.id))

    ws = _sheet(payload)
    labels = {ws[f"A{r}"].value: ws[f"B{r}"].value for r in range(1, 13)}

    assert labels["CONTAINER :"] == w.shipment.shipping_container_number
    assert labels["SEAL NO : "] == "J0713349"
    assert labels["SO :"] == "CNH1098313"
    assert labels["CONSIGNEE :"] == "SORENTO SDN BHD"
    assert labels["SHIPPER :"] == "SHENZHEN XINDESHENG TRADING CO.,LTD"
    assert labels["CHINA AGENT : "] == "ONE TOUCH"
    assert labels["FREE DAYS : "] == "14 FREEDAYS"
    assert labels["DELIVERY WAREHOUSE : "] == "BRW"
    # Dates are written AS dates in the reference's own dd/mm/yyyy format (R20 fidelity),
    # not as the ISO strings the payload carries - openpyxl reads one back as a datetime.
    assert labels["ETD :"].date() == date(2026, 7, 23)
    # The REVISED eta, not the first-published one: `eta_delay_date` is the accurate figure.
    assert labels["ETA : "].date() == date(2026, 7, 27)
    # Every factory on board, in the order the blocks below run.
    assert labels["FACTORY :"] == f"{w.kailu.supplier_name}, {w.caizhou.supplier_name}"


def test_the_header_of_a_container_nobody_has_typed_up_is_blank_not_invented(db):
    w = World(db)
    payload = svc.build(db, str(w.shipment.id))

    ws = _sheet(payload)

    assert ws["B5"].value is None  # seal
    assert ws["B8"].value is None  # shipper
    assert ws["B11"].value is None  # free days


def test_the_two_row_column_header_is_the_fscu_one(db):
    w = World(db)
    payload = svc.build(db, str(w.shipment.id))

    ws = _sheet(payload)

    assert [ws[f"{c}15"].value for c in "ABCDEFGH"] == [
        "FACTORY", "NO", "MODEL", "DESCRIPTION", "MATERIAL", "QTY", "PCS / CTN", "CTN QTY",
    ]
    assert ws["I15"].value == "SIZE (CM)"
    assert [ws["I16"].value, ws["J16"].value, ws["K16"].value] == ["L", "W", "H"]
    assert [ws[f"{c}15"].value for c in "LMNOPQRSTU"] == [
        "CBM\n/ CTN", "TOTAL CBM", "NW", "GW", "TOTAL NW", "TOTAL GW",
        # `TOTAL RM` is the reference workbook's own heading, kept as it is written there.
        "LOGO", "REMARKS", "RMB", "TOTAL RM",
    ]
    # The lines scroll under the header rather than past it.
    assert ws.freeze_panes == "A17"


def test_one_block_per_factory_with_its_mocha_goods_under_their_own_heading(db):
    # MOCHA is invoiced separately, so the same factory's MOCHA lines are their own block -
    # the footer's per-company apportionment has to be readable off the blocks above it.
    w = World(db)
    payload = svc.build(db, str(w.shipment.id))

    ws = _sheet(payload)
    headings = [v for v in _column(ws, "A")[16:] if v and v != "-"]

    assert headings.count(w.kailu.supplier_name) == 2  # tap + mat, both SORENTO
    assert f"{w.kailu.supplier_name} (MOCHA)" in headings
    # A factory with no MOCHA line gets one block, not an empty second one.
    assert f"{w.caizhou.supplier_name} (MOCHA)" not in headings


def test_the_lines_are_numbered_once_across_the_whole_container(db):
    w = World(db)
    payload = svc.build(db, str(w.shipment.id))

    ws = _sheet(payload)
    numbers = [v for v in _column(ws, "B")[16:] if isinstance(v, int)]

    assert numbers == list(range(1, 6))


def test_a_measured_line_derives_its_cartons_volume_and_weights_as_formulas(db):
    # Formulas, not computed numbers: the recipient corrects a quantity in Excel - that is
    # what the sheet is for - and a workbook of frozen totals would keep printing the old
    # ones underneath the corrected line.
    w = World(db)
    payload = svc.build(db, str(w.shipment.id))

    ws = _sheet(payload)
    r = _row_of(ws, "C", w.tap.product_code)

    assert ws[f"E{r}"].value == "不锈钢"
    assert ws[f"F{r}"].value == 490
    assert ws[f"G{r}"].value == 10
    assert ws[f"H{r}"].value == f"=F{r}/G{r}"
    assert [ws[f"I{r}"].value, ws[f"J{r}"].value, ws[f"K{r}"].value] == [34.0, 24.0, 30.0]
    assert ws[f"L{r}"].value == f"=I{r}*J{r}*K{r}/10^6"
    assert ws[f"M{r}"].value == f"=H{r}*L{r}"
    assert ws[f"P{r}"].value == f"=N{r}*H{r}"
    assert ws[f"Q{r}"].value == f"=O{r}*H{r}"
    assert ws[f"U{r}"].value == f"=T{r}*F{r}"


def test_an_unmeasured_line_states_what_it_knows_and_derives_nothing(db):
    # A formula with no inputs prints a zero, and a container planned against a zero volume
    # arrives too full to close.
    w = World(db)
    payload = svc.build(db, str(w.shipment.id))

    ws = _sheet(payload)
    r = _row_of(ws, "C", w.sink.product_code)

    assert ws[f"G{r}"].value is None
    assert ws[f"H{r}"].value == 5  # the carton count the packing list stated
    assert ws[f"L{r}"].value is None
    assert ws[f"M{r}"].value == 1.5  # the volume it stated, not a formula
    assert ws[f"P{r}"].value is None
    assert ws[f"Q{r}"].value is None


def test_one_stated_weight_is_read_as_the_gross_one(db):
    # `weight_per_carton` is the only weight the line has ever had and every packing list
    # that states one states the gross. Read as a fallback rather than migrated.
    w = World(db)
    w.line(w.caizhou, w.never_packed, qty=10, cartons=2, weight_per_carton="18.70")
    payload = svc.build(db, str(w.shipment.id))

    ws = _sheet(payload)
    r = _row_of(ws, "C", w.never_packed.product_code)

    assert ws[f"N{r}"].value is None
    assert ws[f"O{r}"].value == 18.7


def test_every_block_subtotals_its_own_rows(db):
    w = World(db)
    payload = svc.build(db, str(w.shipment.id))

    ws = _sheet(payload)
    first = _row_of(ws, "C", w.tap.product_code)
    subtotal = first + 2  # tap, mat, then the subtotal: both are SORENTO Kailu lines

    assert ws[f"F{subtotal}"].value == f"=SUM(F{first}:F{first + 1})"
    assert ws[f"M{subtotal}"].value == f"=SUM(M{first}:M{first + 1})"
    assert ws[f"U{subtotal}"].value == f"=SUM(U{first}:U{first + 1})"
    # The factory's own amount sits beside its lines, merged down them.
    assert ws[f"V{first}"].value == f"=SUM(U{first}:U{first + 1})"
    assert ws[f"V{subtotal}"].value == f"=SUM(V{first})"


def test_the_grand_total_sums_the_subtotals_and_never_the_lines_twice(db):
    w = World(db)
    payload = svc.build(db, str(w.shipment.id))

    ws = _sheet(payload)
    dash = _row_of(ws, "A", "-")
    total = dash + 1
    subtotals = [
        r
        for r in range(17, dash)
        if isinstance(ws[f"F{r}"].value, str) and ws[f"F{r}"].value.startswith("=SUM(F")
    ]

    assert ws[f"F{total}"].value == "=SUM(" + ",".join(f"F{r}" for r in subtotals) + ")"
    # Kailu SORENTO, Kailu MOCHA, Caizhou, and the lines whose factory we were never told.
    assert len(subtotals) == 4


def test_the_footer_splits_the_container_between_the_two_companies(db):
    w = World(db)
    w.costed()
    payload = svc.build(db, str(w.shipment.id))

    ws = _sheet(payload)
    sorento = _row_of(ws, "L", "SORENTO")
    mocha = _row_of(ws, "L", "MOCHA")
    total = _row_of(ws, "A", "-") + 1

    # Clearance and China freight follow the VOLUME, insurance follows the AMOUNT: that is
    # how the forwarder bills them.
    assert ws[f"N{sorento}"].value == f"=M{sorento}/M{total}*2700.0"
    assert ws[f"O{sorento}"].value == f"=U{sorento}/U{total}*1.0"
    assert ws[f"P{sorento}"].value == f"=M{sorento}/M{total}*13950.0"
    assert ws[f"U{mocha}"].value.startswith("=SUM(U")

    grand = mocha + 1
    assert ws[f"M{grand}"].value == f"=M{sorento}+M{mocha}"
    assert ws[f"U{grand}"].value == f"=U{sorento}+U{mocha}"

    labels = grand + 1
    assert [ws[f"{c}{labels}"].value for c in "MNOP"] == [
        "CBM", "CLEARANCE", "INSURANCE", "CHINA FREIGHT",
    ]
    assert ws[f"U{labels}"].value == "TOTAL AMOUNT"
    assert ws[f"C{labels}"].value == "订单号:CNH1098313"
    assert ws[f"C{labels + 1}"].value == f"柜号:{w.shipment.shipping_container_number}"
    assert ws[f"C{labels + 2}"].value == "封号:J0713349"


def test_a_container_with_no_costs_typed_prints_the_split_and_no_apportionment(db):
    # The split is what the sheet is for; the costs are typed later, and a zero in their
    # place would read as a container that cost nothing to clear.
    w = World(db)
    payload = svc.build(db, str(w.shipment.id))

    ws = _sheet(payload)
    sorento = _row_of(ws, "L", "SORENTO")

    assert ws[f"M{sorento}"].value.startswith("=SUM(M")
    assert ws[f"N{sorento}"].value is None
    assert ws[f"O{sorento}"].value is None
    assert ws[f"P{sorento}"].value is None
    # And the total row does not add two blanks up into a 0 underneath them.
    grand = _row_of(ws, "L", "MOCHA") + 1
    assert ws[f"M{grand}"].value is not None
    assert ws[f"N{grand}"].value is None
    assert ws[f"P{grand}"].value is None


def test_the_workbook_writes_quantities_as_numbers_not_text(db):
    # A quantity the recipient cannot sum in Excel is a picture of a packing list.
    w = World(db)
    payload = svc.build(db, str(w.shipment.id))

    ws = _sheet(payload)
    r = _row_of(ws, "C", w.tap.product_code)

    assert ws[f"F{r}"].value == 490
    assert isinstance(ws[f"T{r}"].value, float)


def test_the_download_is_named_after_the_container_with_nothing_a_filesystem_argues_with(db):
    w = World(db)
    w.shipment.shipping_container_number = "FSCU 810/3365"
    db.flush()
    payload = svc.build(db, str(w.shipment.id))

    assert svc.export_filename(payload) == "FSCU8103365-packing-list.xlsx"


def test_a_container_with_no_number_falls_back_to_the_shipment_number(db):
    w = World(db)
    w.shipment.shipping_container_number = None
    db.flush()
    payload = svc.build(db, str(w.shipment.id))

    assert svc.export_filename(payload) == f"{w.shipment.shipment_number}-packing-list.xlsx"


def test_the_workbook_prints_only_what_is_on_the_container(db):
    """R20 - a model the loading plan asked for and nobody loaded gets no row.

    It used to get one, with an empty quantity and "Not packed - loading plan asked 100" in
    REMARKS. On the document the forwarder and the clearance agent read, a row with a model
    number on it is goods in the box.
    """
    w = World(db)
    w.notice_for_kailu()
    payload = svc.build(db, str(w.shipment.id))

    ws = _sheet(payload)
    printed = [
        str(cell.value)
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    ]
    assert w.never_packed.product_code not in printed
    assert not any("Not packed" in value for value in printed)
    assert not any("loading plan" in value.lower() for value in printed)


def test_a_container_with_no_lines_still_produces_a_readable_sheet(db):
    # A container read but not unpacked is a real state, and the screen offering the export
    # cannot know which ones are empty.
    w = World(db)
    db.query(InboundShipmentLine).filter(
        InboundShipmentLine.shipment_id == w.shipment.id
    ).delete(synchronize_session=False)
    db.flush()
    payload = svc.build(db, str(w.shipment.id))

    ws = _sheet(payload)
    total = _row_of(ws, "A", "-") + 1

    assert ws[f"F{total}"].value == 0
    assert ws[f"M{_row_of(ws, 'L', 'MOCHA')}"].value == 0


# --------------------------------------------------------------------------- #
# over the wire
# --------------------------------------------------------------------------- #


#: The incumbent company every row this suite seeds is auto-stamped with (tests/conftest.py).
SORENTO_COMPANY_ID = "00000000-0000-0000-0000-000000000001"


def _caller(db, monkeypatch, *, email: str, permitted: bool) -> TestClient:
    """A TestClient reading the session the test seeded, with a company to read it under.

    The company half is not optional: the api router depends on `apply_company_scope`, which
    in a real request resolves the caller's companies from their token. Overridden away, the
    session falls to a fail-closed empty scope and every owned row - shipment included -
    disappears, which reads as a 404 from the route rather than as a missing test principal.
    """
    from app.models.base import set_company_scope
    from app.services.company_scope_resolver import apply_company_scope
    from app.services.user_service import UserPermissionService

    principal = {"id": str(uuid.uuid4()), "email": email}
    app.dependency_overrides[get_db] = lambda: db
    app.dependency_overrides[get_current_user] = lambda: principal
    # The scm router is wrapped in the module guard, which authenticates separately.
    app.dependency_overrides[get_current_user_or_api_key] = lambda: principal
    monkeypatch.setattr(
        UserPermissionService,
        "check_user_has_permission",
        lambda self, uid, slug: permitted and slug == "scm.dashboard.view",
    )

    scope = frozenset({SORENTO_COMPANY_ID})
    set_company_scope(db, scope)

    async def _scope():
        set_company_scope(db, scope)
        return scope

    app.dependency_overrides[apply_company_scope] = _scope
    return TestClient(app)


@pytest.fixture
def client(db, monkeypatch):
    """A caller holding `scm.dashboard.view`."""
    try:
        yield _caller(db, monkeypatch, email="zzt-scm@example.com", permitted=True)
    finally:
        app.dependency_overrides.clear()


@pytest.fixture
def stranger(db, monkeypatch):
    """The same caller without the read permission."""
    try:
        yield _caller(db, monkeypatch, email="zzt-nobody@example.com", permitted=False)
    finally:
        app.dependency_overrides.clear()


def test_the_packing_list_route_returns_the_consolidated_list(db, client):
    w = World(db)
    w.notice_for_kailu()

    r = client.get(f"/api/v1/scm/inbound-shipments/{w.shipment.id}/packing-list")

    assert r.status_code == 200, r.text
    body = r.json()
    assert len(body["factories"]) == 3
    assert body["total"]["qty"] == 767
    assert [row["company"] for row in body["split"]] == ["SORENTO", "MOCHA"]


def test_the_export_route_returns_a_workbook_named_after_the_container(db, client):
    import openpyxl

    w = World(db)

    r = client.get(f"/api/v1/scm/inbound-shipments/{w.shipment.id}/packing-list/export")

    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == _XLSX
    expected_name = f"{w.shipment.shipping_container_number}-packing-list.xlsx"
    assert r.headers["content-disposition"] == (
        f'attachment; filename="{expected_name}"; filename*=UTF-8\'\'{quote(expected_name)}'
    )
    ws = openpyxl.load_workbook(BytesIO(r.content))["RMB"]
    factory_column = [ws[f"A{i}"].value for i in range(1, ws.max_row + 1)]
    company_column = [ws[f"L{i}"].value for i in range(1, ws.max_row + 1)]
    assert w.kailu.supplier_name in factory_column
    assert f"{w.kailu.supplier_name} (MOCHA)" in factory_column
    assert "SORENTO" in company_column
    assert "MOCHA" in company_column


def test_an_unknown_container_is_a_404_over_the_wire(db, client):
    r = client.get(f"/api/v1/scm/inbound-shipments/{uuid.uuid4()}/packing-list")

    assert r.status_code == 404, r.text


def test_a_container_id_that_is_not_an_id_is_a_404_over_the_wire(db, client):
    r = client.get("/api/v1/scm/inbound-shipments/not-a-uuid/packing-list")

    assert r.status_code == 404, r.text
    r = client.get("/api/v1/scm/inbound-shipments/not-a-uuid/packing-list/export")

    assert r.status_code == 404, r.text


def test_both_routes_require_the_scm_read_permission(db, stranger):
    w = World(db)

    assert stranger.get(
        f"/api/v1/scm/inbound-shipments/{w.shipment.id}/packing-list"
    ).status_code == 403
    assert stranger.get(
        f"/api/v1/scm/inbound-shipments/{w.shipment.id}/packing-list/export"
    ).status_code == 403
