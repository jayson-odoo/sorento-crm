"""Order tracking import must read 'Overall Tracking', never 'Daily Tracking'."""
from __future__ import annotations

from io import BytesIO
from unittest.mock import MagicMock

import openpyxl
import pytest
from openpyxl import Workbook

from app.services.order_service import OrderService


def _build_workbook(sheets: list[str], rows: dict[str, list[list]] | None = None) -> bytes:
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    for name in sheets:
        ws = wb.create_sheet(name)
        for r in (rows or {}).get(name, []):
            ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _service():
    db = MagicMock()
    db.query.return_value.filter.return_value.all.return_value = []
    return OrderService(db)


def test_import_rejects_workbook_with_only_daily_tracking_sheet():
    data = _build_workbook(["Master", "Daily Tracking"])
    with pytest.raises(ValueError, match="Overall Tracking"):
        _service().import_excel_tracking(data, "user-1", validate_only=True)


def test_import_rejects_workbook_missing_master_sheet():
    data = _build_workbook(["Overall Tracking"])
    with pytest.raises(ValueError, match="Master"):
        _service().import_excel_tracking(data, "user-1", validate_only=True)


def test_workbook_with_overall_tracking_picks_overall_rows_not_daily(monkeypatch):
    """When both sheets are present, the Overall Tracking sheet is the source.

    We patch ``openpyxl.load_workbook`` only to assert that the service then
    reads the ``Overall Tracking`` sheet - without depending on the rest of
    the import pipeline.
    """
    data = _build_workbook(
        ["Master", "Daily Tracking", "Overall Tracking"],
        rows={
            "Master": [["Doc. No."], ["DOC-MASTER-1"]],
            "Daily Tracking": [["Doc Number"], ["DOC-DAILY-ONLY"]],
            "Overall Tracking": [["Doc Number"], ["DOC-OVERALL-ONLY"]],
        },
    )

    accessed_sheets: list[str] = []

    real_workbook = openpyxl.load_workbook(BytesIO(data), data_only=True)

    class TrackingWorkbook:
        sheetnames = real_workbook.sheetnames

        def __getitem__(self, name):
            accessed_sheets.append(name)
            return real_workbook[name]

    monkeypatch.setattr(openpyxl, "load_workbook", lambda *a, **kw: TrackingWorkbook())

    service = _service()
    # Allow the import to bail anywhere after sheet selection; we only care
    # which sheets were accessed.
    try:
        service.import_excel_tracking(data, "user-1", validate_only=True)
    except Exception:
        pass

    assert "Overall Tracking" in accessed_sheets
    assert "Daily Tracking" not in accessed_sheets
