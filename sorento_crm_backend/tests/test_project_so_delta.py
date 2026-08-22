"""P11 revision delta, amendment and order change notice.

The revision under test is the client's own: `delivery-schedule-buimaco-r1.pdf` against
`delivery-schedule-slg-r2.pdf`, both committed in the golden set. Every date below is
transcribed from those two documents.

R1 runs the TOWER monthly from 1 July 2026 to 1 June 2027. R2 runs it fortnightly from
7 January 2027 to 10 June 2027, and the three COMMON AREA rows move from
1 Jul 2026 / 1 Oct 2026 / 1 Jun 2027 to 24 Jun 2027 / 8 Jul 2027 / 22 Jul 2027. Quantities
are identical in both. R2 was issued by SLG Construction while the PO came from Buimaco,
so it is a different `delivery_schedules` row for the same commitment -- which the delta
has to accept.

The column totals are the customer's own printed TOTAL QTY row: 927 for `SRTWC8613-RL`
(135 + 72 x 11) and 894 for `SRTWC8608-RL` (124 + 66 x 11 tower, plus 44 in the common
area). The fixtures reproduce them exactly, so a spread error would fail the build before
the delta was ever computed.

Fixture builders are shared with ``test_project_so_draft`` rather than copied: two
divergent copies of "what a PO version looks like" is how a suite starts testing itself.
"""
from __future__ import annotations

import io
from datetime import date, timedelta
from decimal import Decimal

import openpyxl
import pytest

from app.models.project_so import (
    DeliveryScheduleCell,
    OrderChangeNotice,
    ProjectDeliveryPhase,
    ProjectSalesOrderLine,
    SOAmendment,
)
from app.services.error_handler import AppException
from app.services.project_so_delta_service import (
    RESERVE_WINDOW_DAYS,
    ProjectSODeltaService,
)
from app.services.project_so_draft_service import ProjectSODraftService

from ._pg_fixture import blank_session
from .test_project_so_draft import (
    _cell,
    _lines,
    _party,
    _phase,
    _po,
    _po_version,
    _product,
    _project,
    _quotation,
    _schedule,
    _sorento,
    _user,
    project_seed_service,
)

MARKER = "zzt-delta"

# TOWER, R1 then R2, straight off the two documents: (sequence, label, R1, R2, qty P1, qty P2)
TOWER = [
    (1, "Level 2 & 7", date(2026, 7, 1), date(2027, 1, 7), "135", "124"),
    (2, "Level 8 & 10", date(2026, 8, 3), date(2027, 1, 21), "72", "66"),
    (3, "Level 11 to 13", date(2026, 9, 1), date(2027, 2, 4), "72", "66"),
    (4, "Level 14 to 16", date(2026, 10, 1), date(2027, 2, 18), "72", "66"),
    (5, "Level 17 to 19", date(2026, 11, 2), date(2027, 3, 4), "72", "66"),
    (6, "Level 20 to 22", date(2026, 12, 1), date(2027, 3, 18), "72", "66"),
    (7, "Level 23 to 25", date(2027, 1, 4), date(2027, 4, 1), "72", "66"),
    (8, "Level 26 to 28", date(2027, 2, 1), date(2027, 4, 15), "72", "66"),
    (9, "Level 29 to 31", date(2027, 3, 1), date(2027, 4, 29), "72", "66"),
    (10, "Level 32 to 34", date(2027, 4, 5), date(2027, 5, 13), "72", "66"),
    (11, "Level 35 to 37", date(2027, 5, 3), date(2027, 5, 27), "72", "66"),
    (12, "Level 38 to 40", date(2027, 6, 1), date(2027, 6, 10), "72", "66"),
]

# COMMON AREA carries NO label at all, which is finding G6: identity is (area, sequence).
COMMON = [
    (13, date(2026, 7, 1), date(2027, 6, 24)),
    (14, date(2026, 10, 1), date(2027, 7, 8)),
    (15, date(2027, 6, 1), date(2027, 7, 22)),
]


@pytest.fixture()
def seeded():
    with blank_session() as db:
        company_id = _sorento(db)
        project_seed_service.run(db, company_id=company_id)
        owner = _user(db, f"{MARKER} Maryam")
        yield db, company_id, owner


def _phases_payload(phases, *, use_r2: bool):
    """The `phases` block of a version's extracted_json.

    Dates come from here rather than from `project_delivery_phases`, because that table is
    keyed on (project, area_group, sequence) and a confirmed revision UPDATES the row: the
    version's own extraction is the only immutable record of what THAT document said.
    """
    payload = []
    for sequence, label, r1, r2, _q1, _q2 in phases[0]:
        payload.append(
            {
                "area_group": "TOWER",
                "sequence": sequence,
                "label": label,
                "delivery_date": (r2 if use_r2 else r1).isoformat(),
            }
        )
    for sequence, r1, r2 in phases[1]:
        payload.append(
            {
                "area_group": "COMMON AREA",
                "sequence": sequence,
                "label": None,
                "delivery_date": (r2 if use_r2 else r1).isoformat(),
            }
        )
    return {"phases": payload}


def _tuju(db, company_id, owner):
    """The whole R1 build: PO, quotation, schedule R1, phases, cells, drafted orders."""
    project = _project(db, company_id, owner)
    p1 = _product(db, "SRTWC8613-RL")
    p2 = _product(db, "SRTWC8608-RL")
    p3 = _product(db, "SRTFV1001")
    quotation = _quotation(
        db,
        project,
        lines=[(p1, "927", "392.85"), (p2, "894", "352.10"), (p3, "16", "295.85")],
    )
    party = _party(db, company_id)
    po = _po(db, project, party=party, quotation_version=quotation, po_number="HQ/26/01/121")
    po_version = _po_version(
        db,
        po,
        lines=[
            (1, p1, "SRTWC8613-RL", "927", "SETS", "392.85", "364171.95", False),
            (2, p2, "SRTWC8608-RL", "894", "SETS", "352.10", "314777.40", False),
            (3, p3, "SRTFV1001", "16", "NOS", "295.85", "4733.60", False),
        ],
    )
    r1 = _schedule(db, project, po, po_version=po_version, revision_label="ORIGINAL")
    r1.extracted_json = _phases_payload((TOWER, COMMON), use_r2=False)
    db.flush()

    phases = {}
    for sequence, label, r1_date, _r2, q1, q2 in TOWER:
        phase = _phase(
            db,
            project,
            area_group="TOWER",
            sequence=sequence,
            label=label,
            delivery_date=r1_date,
            version=r1,
        )
        phases[("TOWER", sequence)] = phase
        _cell(db, r1, phase, p1, q1)
        _cell(db, r1, phase, p2, q2)
    for sequence, r1_date, _r2 in COMMON:
        phase = _phase(
            db,
            project,
            area_group="COMMON AREA",
            sequence=sequence,
            label=None,
            delivery_date=r1_date,
            version=r1,
        )
        phases[("COMMON AREA", sequence)] = phase
    # 44 of P2 and 16 of P3 in the common area, so the printed column totals close:
    # 850 tower + 44 = 894, and 16 for the flush valve.
    _cell(db, r1, phases[("COMMON AREA", 15)], p2, "44")
    _cell(db, r1, phases[("COMMON AREA", 13)], p3, "16")

    service = ProjectSODraftService(db)
    built = service.build(po.id, r1.id)
    orders = {row["area_group"]: service.get_order(row["id"]) for row in built["data"]}
    for order in orders.values():
        service.publish(order, actor_user_id=owner)
    return project, po, po_version, r1, phases, orders, (p1, p2, p3)


def _revision(db, project, po, po_version, phases, products, *, use_r2=True, tweaks=None):
    """A second schedule version -- a different `delivery_schedules` row, as R2 really was."""
    p1, p2, p3 = products
    r2 = _schedule(
        db,
        project,
        po,
        po_version=po_version,
        revision_label="REVISED 1 - 23/7/2026",
        version_no=1,
    )
    r2.extracted_json = _phases_payload((TOWER, COMMON), use_r2=use_r2)
    db.flush()
    for sequence, _label, _r1, _r2, q1, q2 in TOWER:
        phase = phases[("TOWER", sequence)]
        _cell(db, r2, phase, p1, (tweaks or {}).get(("TOWER", sequence, "p1"), q1))
        _cell(db, r2, phase, p2, (tweaks or {}).get(("TOWER", sequence, "p2"), q2))
    _cell(db, r2, phases[("COMMON AREA", 15)], p2, "44")
    _cell(db, r2, phases[("COMMON AREA", 13)], p3, "16")
    return r2


# --------------------------------------------------------------- the golden delta


def test_the_r1_to_r2_revision_delays_every_phase_and_touches_no_quantity(seeded):
    """The client's own revision, in their own verbs.

    Twelve TOWER phases and three COMMON AREA phases move; not one quantity does. The
    phase-level count is what purchasing reads (`12 tower DELAY rows`); the row-level count
    is 12 phases x 2 products = 24, because an amendment is applied line by line.
    """
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    r2 = _revision(db, project, po, po_version, phases, products)

    service = ProjectSODeltaService(db)
    tower = service.preview(orders["TOWER"].id, schedule_version_id=r2.id)

    assert len(tower["phase_summary"]) == 12
    assert {entry["verb"] for entry in tower["phase_summary"]} == {"DELAY"}
    assert [entry["sequence"] for entry in tower["phase_summary"]] == list(range(1, 13))
    assert tower["phase_summary"][0]["delivery_date_from"] == "2026-07-01"
    assert tower["phase_summary"][0]["delivery_date_to"] == "2027-01-07"
    assert tower["phase_summary"][-1]["delivery_date_to"] == "2027-06-10"

    assert tower["verb_summary"]["DELAY"] == 24
    assert tower["verb_summary"]["ADVANCE"] == 0
    assert tower["verb_summary"]["ORDER"] == 0
    assert tower["verb_summary"]["CANCEL BALANCE"] == 0
    assert tower["quantities_changed"] is False
    assert all(row["field"] == "delivery_date" for row in tower["rows"])
    # Every row names the product and the phase, so no UUID is the only identifier.
    assert all(row["product_code"] and row["phase_label_from"] for row in tower["rows"])
    assert tower["from"]["label"] == "ORIGINAL"
    assert tower["to"]["label"] == "REVISED 1 - 23/7/2026"
    assert tower["unmatched"] == []

    common = service.preview(orders["COMMON AREA"].id, schedule_version_id=r2.id)
    assert len(common["phase_summary"]) == 3
    assert {entry["verb"] for entry in common["phase_summary"]} == {"DELAY"}
    # Sequence 14 carries no quantity for this order's products, so it has no line and no
    # row -- but the phase still moved, and saying so is the point of the summary.
    by_sequence = {entry["sequence"]: entry for entry in common["phase_summary"]}
    assert by_sequence[14]["line_count"] == 0
    assert by_sequence[13]["delivery_date_to"] == "2027-06-24"
    assert by_sequence[15]["delivery_date_to"] == "2027-07-22"
    assert common["quantities_changed"] is False
    assert common["verb_summary"]["DELAY"] == 2


def test_a_quantity_drop_reads_as_cancel_balance(seeded):
    """`CANCEL BALANCE 30 NOS` is how they write it, so that is what the row says."""
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    r2 = _revision(
        db,
        project,
        po,
        po_version,
        phases,
        products,
        use_r2=False,
        tweaks={("TOWER", 1, "p1"): "105"},
    )

    tower = ProjectSODeltaService(db).preview(orders["TOWER"].id, schedule_version_id=r2.id)
    cancels = [row for row in tower["rows"] if row["verb"] == "CANCEL BALANCE"]
    assert len(cancels) == 1
    assert cancels[0]["field"] == "qty"
    assert cancels[0]["from_value"] == "135"
    assert cancels[0]["to_value"] == "105"
    assert cancels[0]["qty"] == "30"
    assert cancels[0]["product_code"] == "SRTWC8613-RL"
    assert tower["quantities_changed"] is True
    # Dates did not move in this revision, so nothing else is proposed.
    assert tower["verb_summary"]["DELAY"] == 0


def test_an_increase_landing_inside_the_reserve_window_reads_as_reserve_and_order(seeded):
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    p1, p2, p3 = products
    soon = date.today() + timedelta(days=RESERVE_WINDOW_DAYS - 5)
    r2 = _schedule(db, project, po, po_version=po_version, revision_label="URGENT")
    payload = _phases_payload((TOWER, COMMON), use_r2=False)
    for entry in payload["phases"]:
        if entry["area_group"] == "TOWER" and entry["sequence"] == 1:
            entry["delivery_date"] = soon.isoformat()
    r2.extracted_json = payload
    db.flush()
    for sequence, _label, _r1d, _r2d, q1, q2 in TOWER:
        phase = phases[("TOWER", sequence)]
        _cell(db, r2, phase, p1, "200" if sequence == 1 else q1)
        _cell(db, r2, phase, p2, q2)
    _cell(db, r2, phases[("COMMON AREA", 15)], p2, "44")
    _cell(db, r2, phases[("COMMON AREA", 13)], p3, "16")

    tower = ProjectSODeltaService(db).preview(orders["TOWER"].id, schedule_version_id=r2.id)
    verbs = {row["verb"] for row in tower["rows"] if row["field"] == "qty"}
    assert verbs == {"RESERVE & ORDER"}
    assert tower["verb_summary"]["RESERVE & ORDER"] == 1


def test_a_phase_that_cannot_be_matched_is_reported_rather_than_dropped(seeded):
    """Silently dropping it is how a delta lies."""
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    p1, p2, p3 = products
    r2 = _schedule(db, project, po, po_version=po_version, revision_label="REVISED 1")
    payload = _phases_payload((TOWER, COMMON), use_r2=True)
    # The customer's revision simply does not carry Level 38 to 40 any more.
    payload["phases"] = [
        entry
        for entry in payload["phases"]
        if not (entry["area_group"] == "TOWER" and entry["sequence"] == 12)
    ]
    r2.extracted_json = payload
    db.flush()
    for sequence, _label, _r1d, _r2d, q1, q2 in TOWER:
        if sequence == 12:
            continue
        phase = phases[("TOWER", sequence)]
        _cell(db, r2, phase, p1, q1)
        _cell(db, r2, phase, p2, q2)
    _cell(db, r2, phases[("COMMON AREA", 15)], p2, "44")
    _cell(db, r2, phases[("COMMON AREA", 13)], p3, "16")

    tower = ProjectSODeltaService(db).preview(orders["TOWER"].id, schedule_version_id=r2.id)
    assert len(tower["phase_summary"]) == 11
    assert len(tower["unmatched"]) == 1
    entry = tower["unmatched"][0]
    assert entry["reason"] == "phase not found in the new version"
    assert entry["sequence"] == 12
    assert "Level 38 to 40" in entry["detail"]
    # The quantity at risk is named, so nobody has to work it out: 72 + 66.
    assert "138" in entry["detail"]


def test_a_delivery_moved_to_another_area_reads_as_change_so_no(seeded):
    """AC-G3's corroborated re-match, which is the client's `CHANGE SO NO`.

    Phase identity is `(area_group, sequence)`, so a delivery that moves to another area
    looks like one phase disappearing and another appearing. Reporting both as unmatched
    would be two half-truths; matching them on the sequence, corroborated by an identical
    quantity vector, is the one row that says what happened.
    """
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    p1, p2, p3 = products
    # The COMMON AREA delivery at sequence 13 becomes a PODIUM delivery at sequence 13,
    # same date, same 16 units.
    podium = _phase(
        db,
        project,
        area_group="PODIUM",
        sequence=13,
        label=None,
        delivery_date=date(2026, 7, 1),
        version=None,
    )
    r2 = _schedule(db, project, po, po_version=po_version, revision_label="REVISED 1")
    payload = _phases_payload((TOWER, COMMON), use_r2=False)
    for entry in payload["phases"]:
        if entry["area_group"] == "COMMON AREA" and entry["sequence"] == 13:
            entry["area_group"] = "PODIUM"
    r2.extracted_json = payload
    db.flush()
    for sequence, _label, _r1d, _r2d, q1, q2 in TOWER:
        phase = phases[("TOWER", sequence)]
        _cell(db, r2, phase, p1, q1)
        _cell(db, r2, phase, p2, q2)
    _cell(db, r2, phases[("COMMON AREA", 15)], p2, "44")
    _cell(db, r2, podium, p3, "16")

    common = ProjectSODeltaService(db).preview(
        orders["COMMON AREA"].id, schedule_version_id=r2.id
    )
    moves = [row for row in common["rows"] if row["verb"] == "CHANGE SO NO"]
    assert len(moves) == 1
    assert moves[0]["field"] == "area_group"
    assert moves[0]["from_value"] == "COMMON AREA"
    assert moves[0]["to_value"] == "PODIUM"
    assert moves[0]["qty"] == "16"
    assert moves[0]["product_code"] == "SRTFV1001"
    # Reported as ONE row, so neither side is left in `unmatched`.
    assert common["unmatched"] == []


# ------------------------------------------------------------ amendment and OCN


def test_creating_an_amendment_auto_drafts_its_change_notice_from_the_same_delta(seeded):
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    r2 = _revision(db, project, po, po_version, phases, products)

    service = ProjectSODeltaService(db)
    created = service.create(
        orders["TOWER"].id,
        schedule_version_id=r2.id,
        reason="Main contractor pushed the whole programme by six months.",
        actor_user_id=owner,
    )

    amendment = db.query(SOAmendment).filter(SOAmendment.id == created["amendment_id"]).one()
    ocn = db.query(OrderChangeNotice).filter(OrderChangeNotice.id == created["ocn_id"]).one()
    assert amendment.status == "proposed"
    assert amendment.ocn_id == ocn.id
    assert amendment.from_version_kind == "schedule"
    assert amendment.to_version_id == r2.id
    assert amendment.verb_summary["DELAY"] == 24
    assert created["ocn_number"].startswith("OCN-")
    assert ocn.approver_id is None
    assert ocn.reason.startswith("Main contractor")
    assert ocn.source_document_kind == "revised_schedule"
    assert ocn.source_document_id == r2.id
    # The change table IS the delta, not a retyped copy of it.
    assert ocn.change_table_json["verb_summary"] == amendment.verb_summary
    assert len(ocn.change_table_json["rows"]) == len(amendment.delta_json["rows"])
    assert len(ocn.change_table_json["phase_summary"]) == 12


def test_publishing_an_amendment_is_refused_without_an_approver(seeded):
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    r2 = _revision(db, project, po, po_version, phases, products)
    service = ProjectSODeltaService(db)
    created = service.create(
        orders["TOWER"].id,
        schedule_version_id=r2.id,
        reason="Programme pushed six months.",
        actor_user_id=owner,
    )

    with pytest.raises(AppException) as excinfo:
        service.publish_amendment(created["amendment_id"], actor_user_id=owner)
    assert excinfo.value.status_code == 409
    assert "approver" in str(excinfo.value.detail).lower()


def test_publishing_an_amendment_applies_the_dates_and_amends_the_order(seeded):
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    r2 = _revision(db, project, po, po_version, phases, products)
    tower = orders["TOWER"]
    before = {line.id: (line.qty, line.delivery_date) for line in _lines(db, tower.id)}

    service = ProjectSODeltaService(db)
    created = service.create(
        tower.id,
        schedule_version_id=r2.id,
        reason="Programme pushed six months.",
        actor_user_id=owner,
    )
    ocn = db.query(OrderChangeNotice).filter(OrderChangeNotice.id == created["ocn_id"]).one()
    ocn.approver_id = owner
    db.flush()

    body = service.publish_amendment(created["amendment_id"], actor_user_id=owner)
    assert body["status"] == "published"
    assert body["sales_order_status"] == "amended"
    assert body["applied_rows"] == 24
    db.refresh(ocn)
    assert ocn.approved_at is not None

    after = {line.id: (line.qty, line.delivery_date) for line in _lines(db, tower.id)}
    assert {qty for qty, _d in after.values()} == {qty for qty, _d in before.values()}
    assert sorted({d for _q, d in after.values()}) == sorted(
        {r2_date for _s, _l, _r1d, r2_date, _q1, _q2 in TOWER}
    )


def test_an_unpublished_draft_is_edited_not_amended(seeded):
    """Finding G9: nothing is committed, so there is nothing to raise an OCN about."""
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    r2 = _revision(db, project, po, po_version, phases, products)
    tower = orders["TOWER"]
    tower.status = "draft"
    db.flush()

    with pytest.raises(AppException) as excinfo:
        ProjectSODeltaService(db).create(
            tower.id,
            schedule_version_id=r2.id,
            reason="Programme pushed six months.",
            actor_user_id=owner,
        )
    assert excinfo.value.status_code == 409
    assert "draft" in str(excinfo.value.detail).lower()


def test_an_identical_revision_is_refused_rather_than_stored_empty(seeded):
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    same = _revision(db, project, po, po_version, phases, products, use_r2=False)

    with pytest.raises(AppException) as excinfo:
        ProjectSODeltaService(db).create(
            orders["TOWER"].id,
            schedule_version_id=same.id,
            reason="Checking whether anything moved.",
            actor_user_id=owner,
        )
    assert excinfo.value.status_code == 422
    assert "identical" in str(excinfo.value.detail).lower()


def test_naming_two_versions_at_once_is_refused(seeded):
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    r2 = _revision(db, project, po, po_version, phases, products)

    with pytest.raises(AppException) as excinfo:
        ProjectSODeltaService(db).preview(
            orders["TOWER"].id, schedule_version_id=r2.id, po_version_id=po_version.id
        )
    assert excinfo.value.status_code == 422


def test_a_revised_po_reports_a_price_move_without_inventing_a_verb(seeded):
    """Purchasing does not act on a price, so it goes to `unmatched` with both figures."""
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    p1, p2, p3 = products
    revised = _po_version(
        db,
        po,
        version_no=2,
        lines=[
            (1, p1, "SRTWC8613-RL", "927", "SETS", "400.00", "370800.00", False),
            (2, p2, "SRTWC8608-RL", "894", "SETS", "352.10", "314777.40", False),
            (3, p3, "SRTFV1001", "16", "NOS", "295.85", "4733.60", False),
        ],
    )

    delta = ProjectSODeltaService(db).preview(
        orders["TOWER"].id, po_version_id=revised.id
    )
    assert delta["rows"] == []
    reasons = [entry["reason"] for entry in delta["unmatched"]]
    assert "unit price changed and no purchasing verb applies" in reasons
    detail = next(
        entry["detail"]
        for entry in delta["unmatched"]
        if entry["reason"].startswith("unit price")
    )
    assert "392.85" in detail and "400.00" in detail


def test_a_revised_po_that_drops_a_quantity_cancels_the_balance(seeded):
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    p1, p2, p3 = products
    revised = _po_version(
        db,
        po,
        version_no=2,
        lines=[
            (1, p1, "SRTWC8613-RL", "900", "SETS", "392.85", "353565.00", False),
            (2, p2, "SRTWC8608-RL", "894", "SETS", "352.10", "314777.40", False),
            (3, p3, "SRTFV1001", "16", "NOS", "295.85", "4733.60", False),
        ],
    )

    delta = ProjectSODeltaService(db).preview(orders["TOWER"].id, po_version_id=revised.id)
    cancels = [row for row in delta["rows"] if row["verb"] == "CANCEL BALANCE"]
    assert len(cancels) == 1
    assert cancels[0]["qty"] == "27"
    assert cancels[0]["product_code"] == "SRTWC8613-RL"
    assert delta["from"]["kind"] == "po_version"
    assert delta["to"]["version_no"] == 2


def test_the_amendment_detail_reads_as_a_chain(seeded):
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    r2 = _revision(db, project, po, po_version, phases, products)
    service = ProjectSODeltaService(db)
    created = service.create(
        orders["TOWER"].id,
        schedule_version_id=r2.id,
        reason="Programme pushed six months.",
        actor_user_id=owner,
    )

    detail = service.get_amendment(created["amendment_id"])
    assert detail["provisional_ref"] == orders["TOWER"].provisional_ref
    assert detail["from_version_label"] == "ORIGINAL"
    assert detail["to_version_label"] == "REVISED 1 - 23/7/2026"
    assert detail["ocn_number"] == created["ocn_number"]
    assert detail["ocn_approver_name"] is None
    assert detail["status"] == "proposed"
    assert detail["verb_summary"]["DELAY"] == 24
    assert len(detail["delta"]["phase_summary"]) == 12
    # Every row is accepted by default -- that is today's all-or-nothing behaviour.
    assert detail["accepted_count"] == 24
    assert detail["declined_count"] == 0
    assert all(row["decision"] == "accepted" for row in detail["delta"]["rows"])
    assert all(row["declined_reason"] is None for row in detail["delta"]["rows"])
    assert all(row["row_key"] == str(i) for i, row in enumerate(detail["delta"]["rows"]))


# --------------------------------------------------- section 9.3: accept / decline


def _created_amendment(db, orders, r2, owner):
    service = ProjectSODeltaService(db)
    created = service.create(
        orders["TOWER"].id,
        schedule_version_id=r2.id,
        reason="Programme pushed six months.",
        actor_user_id=owner,
    )
    amendment = db.query(SOAmendment).filter(SOAmendment.id == created["amendment_id"]).one()
    ocn = db.query(OrderChangeNotice).filter(OrderChangeNotice.id == created["ocn_id"]).one()
    ocn.approver_id = owner
    db.flush()
    return service, amendment, ocn


def test_declining_a_row_needs_a_reason(seeded):
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    r2 = _revision(db, project, po, po_version, phases, products)
    service, amendment, _ocn = _created_amendment(db, orders, r2, owner)

    with pytest.raises(AppException) as excinfo:
        service.update_row_decisions(amendment.id, {"0": {"decision": "declined"}})
    assert excinfo.value.status_code == 422
    assert excinfo.value.detail["code"] == "amendment_row_reason_required"


def test_an_unknown_row_key_is_refused(seeded):
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    r2 = _revision(db, project, po, po_version, phases, products)
    service, amendment, _ocn = _created_amendment(db, orders, r2, owner)

    with pytest.raises(AppException) as excinfo:
        service.update_row_decisions(
            amendment.id, {"99": {"decision": "accepted"}}
        )
    assert excinfo.value.status_code == 422
    assert excinfo.value.detail["code"] == "amendment_row_unknown"


def test_declining_one_row_leaves_that_line_untouched_and_the_rest_amended(seeded):
    """AC-R05/R06-shaped: publish applies only the accepted rows; the declined row
    stays in `delta_json` with its decision, and its own line is untouched."""
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    r2 = _revision(db, project, po, po_version, phases, products)
    service, amendment, ocn = _created_amendment(db, orders, r2, owner)

    declined_row = amendment.delta_json["rows"][0]
    declined_line_id = declined_row["so_line_id"]
    before_date = declined_row["from_value"]
    assert declined_row["field"] == "delivery_date"

    detail = service.update_row_decisions(
        amendment.id,
        {"0": {"decision": "declined", "reason": "Client wants this phase to stay."}},
    )
    assert detail["declined_count"] == 1
    assert detail["accepted_count"] == 23
    assert detail["delta"]["rows"][0]["decision"] == "declined"
    assert detail["delta"]["rows"][0]["declined_reason"] == "Client wants this phase to stay."
    assert detail["delta"]["rows"][1]["decision"] == "accepted"
    assert detail["delta"]["rows"][1]["declined_reason"] is None

    body = service.publish_amendment(amendment.id, actor_user_id=owner)
    assert body["applied_rows"] == 23

    declined_line = (
        db.query(ProjectSalesOrderLine)
        .filter(ProjectSalesOrderLine.id == declined_line_id)
        .one()
    )
    assert declined_line.delivery_date.isoformat() == before_date

    db.refresh(amendment)
    assert amendment.verb_summary["DELAY"] == 23
    db.refresh(ocn)
    assert len(ocn.change_table_json["rows"]) == 23
    assert ocn.change_table_json["verb_summary"]["DELAY"] == 23
    # The declined row's own decision survives publish, unrewritten.
    published_detail = service.get_amendment(amendment.id)
    assert published_detail["delta"]["rows"][0]["decision"] == "declined"
    assert published_detail["declined_count"] == 1
    assert published_detail["accepted_count"] == 23


def test_row_decisions_are_refused_once_the_amendment_is_no_longer_proposed(seeded):
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    r2 = _revision(db, project, po, po_version, phases, products)
    service, amendment, _ocn = _created_amendment(db, orders, r2, owner)
    service.publish_amendment(amendment.id, actor_user_id=owner)

    with pytest.raises(AppException) as excinfo:
        service.update_row_decisions(
            amendment.id, {"1": {"decision": "declined", "reason": "too late now"}}
        )
    assert excinfo.value.status_code == 409
    assert excinfo.value.detail["code"] == "amendment_not_proposed"


# ------------------------------------------------- section 9.4: AutoCount change list


def test_the_autocount_change_list_lists_only_accepted_rows(seeded):
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    r2 = _revision(db, project, po, po_version, phases, products)
    service, amendment, _ocn = _created_amendment(db, orders, r2, owner)
    service.update_row_decisions(
        amendment.id,
        {"0": {"decision": "declined", "reason": "Client wants this phase to stay."}},
    )

    payload = service.build_autocount_change_list(amendment.id)
    assert payload["declined_count"] == 1
    assert len(payload["rows"]) == 23
    row = payload["rows"][0]
    assert row["verb"] == "DELAY"
    assert row["so_number"] == orders["TOWER"].provisional_ref
    assert row["old_date"] and row["new_date"]
    assert row["old_qty"] is None and row["new_qty"] is None


def test_the_autocount_change_list_workbook_headings_rows_and_declined_footer(seeded):
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    r2 = _revision(db, project, po, po_version, phases, products)
    service, amendment, _ocn = _created_amendment(db, orders, r2, owner)
    service.update_row_decisions(
        amendment.id,
        {"0": {"decision": "declined", "reason": "Client wants this phase to stay."}},
    )

    filename, body = service.autocount_change_list_workbook(amendment.id)
    assert filename.endswith(".xlsx")

    workbook = openpyxl.load_workbook(io.BytesIO(body))
    sheet = workbook.active
    assert sheet["A1"].value.startswith("AUTOCOUNT CHANGE LIST")
    headings = [cell.value for cell in sheet[2]]
    assert headings == [
        "S/O NO", "LINE", "ITEM CODE", "PRODUCT", "VERB",
        "OLD QTY", "NEW QTY", "OLD DATE", "NEW DATE", "NEW S/O NO",
    ]
    data_rows = [row for row in sheet.iter_rows(min_row=3, max_row=25, values_only=True)]
    assert len(data_rows) == 23
    assert all(row[4] == "DELAY" for row in data_rows)

    old_date_cell = sheet.cell(row=3, column=8)
    new_date_cell = sheet.cell(row=3, column=9)
    assert isinstance(old_date_cell.value, date)
    assert isinstance(new_date_cell.value, date)
    assert old_date_cell.number_format == "DD/MM/YYYY"
    assert new_date_cell.number_format == "DD/MM/YYYY"

    footer_values = [
        row[0] for row in sheet.iter_rows(min_row=26, values_only=True) if row and row[0]
    ]
    assert "Declined: 1 rows" in footer_values


# ------------------------------------------- section 9.5: AUTHORED_LIVE_STATUSES


def test_amending_an_adopted_order_is_refused_the_same_way_a_draft_is(seeded):
    """9.5: `create()` reads the shared `AUTHORED_LIVE_STATUSES` constant rather than
    an inline literal pair -- an adopted order (never drafted, never published here)
    is refused the same way a draft is."""
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    r2 = _revision(db, project, po, po_version, phases, products)
    tower = orders["TOWER"]
    tower.status = "adopted"
    db.flush()

    with pytest.raises(AppException) as excinfo:
        ProjectSODeltaService(db).create(
            tower.id,
            schedule_version_id=r2.id,
            reason="Programme pushed six months.",
            actor_user_id=owner,
        )
    assert excinfo.value.status_code == 409


# ----------------------------------------- section 9.7(d): accepted proposal override


def test_an_accepted_proposals_override_reads_per_line_against_the_untouched_product(
    seeded,
):
    """An accepted revision proposal (section 9.7b/c) writes
    ``delivery_date_override`` on ONE product's cells; the delta reads it ahead of
    the phase's own date for THAT product's lines and proposes ADVANCE/DELAY per
    line, while the other product under the very same phases still reads the
    phase's plain, unmatched date -- nothing about it changed."""
    db, company_id, owner = seeded
    project, po, po_version, _r1, phases, orders, products = _tuju(db, company_id, owner)
    p1, p2, _p3 = products
    r2 = _revision(db, project, po, po_version, phases, products)

    # Simulate the accept step: p1's twelve TOWER cells on r2 get a cadence that
    # starts well before every one of R1's own dates, so every p1 line reads
    # ADVANCE regardless of the plain R2 delay sitting on the phase itself.
    phase_by_id = {
        row.id: row
        for row in db.query(ProjectDeliveryPhase)
        .filter(ProjectDeliveryPhase.project_id == project.id)
        .all()
    }
    p1_cells = (
        db.query(DeliveryScheduleCell)
        .filter(
            DeliveryScheduleCell.version_id == r2.id,
            DeliveryScheduleCell.product_id == p1.id,
        )
        .all()
    )
    tower_p1_cells = [
        cell for cell in p1_cells if phase_by_id[cell.phase_id].area_group == "TOWER"
    ]
    assert len(tower_p1_cells) == 12
    ordered = sorted(tower_p1_cells, key=lambda cell: phase_by_id[cell.phase_id].sequence)
    override_start = date(2025, 1, 1)
    for index, cell in enumerate(ordered):
        cell.delivery_date_override = override_start + timedelta(days=14 * index)
    db.flush()

    tower = ProjectSODeltaService(db).preview(orders["TOWER"].id, schedule_version_id=r2.id)
    p1_rows = [row for row in tower["rows"] if row["product_code"] == "SRTWC8613-RL"]
    p2_rows = [row for row in tower["rows"] if row["product_code"] == "SRTWC8608-RL"]
    assert len(p1_rows) == 12
    assert all(row["verb"] == "ADVANCE" for row in p1_rows)
    assert {row["to_value"] for row in p1_rows} == {
        (override_start + timedelta(days=14 * i)).isoformat() for i in range(12)
    }
    # p2 carries no override: same phases, still the plain R2 delay, untouched.
    assert len(p2_rows) == 12
    assert all(row["verb"] == "DELAY" for row in p2_rows)
    assert {row["to_value"] for row in p2_rows} == {
        r2_date.isoformat() for _s, _l, _r1d, r2_date, _q1, _q2 in TOWER
    }
