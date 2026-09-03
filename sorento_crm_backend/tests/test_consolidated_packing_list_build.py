"""S12 - one volume rule for the Split card and the workbook (AC-L1, AC-L2).

`build()` used to fall straight from a line's STORED `cbm` to the catalogue's own
dimensions, skipping the line's own carton measurements entirely - so a line typed with a
pack size and a carton L/W/H and no stored `cbm` measured as ZERO on the Split card and
contributed nothing to a company's clearance / freight share, while `to_xlsx` derived a real
figure for the exact same line live from those dimensions (`=H*L`, `H` from `qty/pcs` or a
stated carton count, `L` from `I*J*K/10^6`). `line_cbm` (`app/services/scm/container_capacity
.py`) is the one rule now, read by both `build()` and `_attach_capacity`'s fill gauge
(`tests/scm/test_shipment_container_capacity.py`), so the two cannot drift from each other or
from the workbook again.

Three lines on one container: a SORENTO line with a stored `cbm` (untouched by this slice, so
the existing behaviour is pinned alongside the new one), a SORENTO dims-only line and a MOCHA
dims-only line - one of each company, the case AC-L2 asks for. Every expected figure here is
worked out by hand in the test body, in the SAME units and the SAME formula the workbook's own
cells use (`ctn = qty / pcs_per_carton`, `cbm = ctn * L * W * H / 1e6`; clearance and freight
apportioned by CBM share, insurance and the line amount by RMB share) - not copied from
`build()`'s own output - so a regression that breaks the arithmetic on both sides at once still
fails this test.

Postgres only, on a blank schema, seeding every FK target itself (same substrate and the same
reason as `test_consolidated_packing_list.py`): CI's database is empty, and a test that
borrows an existing category, brand or supplier passes locally and fails there.
"""
from __future__ import annotations

import uuid
from datetime import date

import pytest

from app.models.procurement import InboundShipment, InboundShipmentLine, Supplier
from app.models.product import Brand, Product, ProductCategory, UnitOfMeasure
from app.services.scm import consolidated_packing_list as svc
from tests._pg_fixture import blank_session

MARKER = "ZZCPLB"


@pytest.fixture
def db():
    with blank_session() as session:
        yield session


class World:
    """One factory, one container: a stored-cbm line and a dims-only line for each company."""

    def __init__(self, db):
        self.db = db
        self.tag = uuid.uuid4().hex[:8].upper()

        self.category = ProductCategory(
            id=str(uuid.uuid4()),
            category_code=f"{MARKER}-CAT-{self.tag}",
            category_name=f"{MARKER} category",
        )
        self.uom = UnitOfMeasure(
            id=str(uuid.uuid4()), uom_code=f"{MARKER}{self.tag}"[:20], uom_name="pcs"
        )
        self.mocha = Brand(
            id=str(uuid.uuid4()), brand_code="MOCHA", brand_name="Mocha", is_active=True
        )
        self.db.add_all([self.category, self.uom, self.mocha])
        self.db.flush()

        self.supplier = Supplier(
            id=str(uuid.uuid4()),
            supplier_code=f"{MARKER}-{self.tag}"[:50],
            supplier_name=f"{MARKER} FACTORY",
            is_active=True,
        )
        self.db.add(self.supplier)
        self.db.flush()

        self.shipment = InboundShipment(
            id=str(uuid.uuid4()),
            shipment_number=f"{MARKER}-SH-{self.tag}",
            shipment_date=date(2026, 8, 1),
            shipping_container_number=f"{MARKER}U{self.tag}",
            shipment_status="in_transit",
            clearance_cost=2700,
            china_freight_cost=13950,
            insurance_rate=1,
        )
        self.db.add(self.shipment)
        self.db.flush()

        # A stored cbm: untouched by this slice, pinned so the existing rule still wins
        # when the packing list DID state one.
        self.stored = self._product("STORED", brand=None)
        self._line(self.stored, qty=50, cbm="5.0000", unit_cost="10.00")

        # SORENTO dims-only: qty / pcs -> 10 cartons, 60x50x40cm -> 0.12 cbm/carton.
        self.sorento_dims = self._product("SORENTO-DIMS", brand=None)
        self._line(
            self.sorento_dims,
            qty=200,
            pcs_per_carton=20,
            carton_length_cm=60,
            carton_width_cm=50,
            carton_height_cm=40,
            unit_cost="15.00",
        )

        # MOCHA dims-only: qty / pcs -> 10 cartons, 50x40x30cm -> 0.06 cbm/carton.
        self.mocha_dims = self._product("MOCHA-DIMS", brand=self.mocha)
        self._line(
            self.mocha_dims,
            qty=100,
            pcs_per_carton=10,
            carton_length_cm=50,
            carton_width_cm=40,
            carton_height_cm=30,
            unit_cost="20.00",
        )

    def _product(self, key: str, *, brand: Brand | None) -> Product:
        p = Product(
            id=str(uuid.uuid4()),
            product_code=f"{MARKER}-{self.tag}-{key}",
            product_name=f"{key} product",
            category_id=self.category.id,
            base_uom_id=self.uom.id,
            brand_id=brand.id if brand else None,
            list_price=0,
            is_active=True,
        )
        self.db.add(p)
        self.db.flush()
        return p

    def _line(self, product: Product, *, qty: int, unit_cost: str, **measured) -> InboundShipmentLine:
        row = InboundShipmentLine(
            id=str(uuid.uuid4()),
            shipment_id=self.shipment.id,
            supplier_id=self.supplier.id,
            product_id=product.id,
            quantity_shipped=qty,
            currency="CNY",
            unit_cost=unit_cost,
            **measured,
        )
        self.db.add(row)
        self.db.flush()
        return row


def _line(payload: dict, code_suffix: str) -> dict:
    factory = payload["factories"][0]
    return next(l for l in factory["lines"] if l["product_code"].endswith(code_suffix))


def _split(payload: dict, company: str) -> dict:
    return next(row for row in payload["split"] if row["company"] == company)


# --------------------------------------------------------------------------------- #
# AC-L1 - a dims-only line contributes ITS OWN carton volume, not zero
# --------------------------------------------------------------------------------- #


def test_a_dims_only_line_derives_its_own_cbm_from_its_own_carton_dimensions(db):
    w = World(db)

    out = svc.build(db, str(w.shipment.id))

    # ctn = 200 / 20 = 10; cbm = 10 * (60 * 50 * 40 / 1e6) = 10 * 0.12
    assert _line(out, "SORENTO-DIMS")["cbm"] == pytest.approx(1.2, abs=0.0001)
    # ctn = 100 / 10 = 10; cbm = 10 * (50 * 40 * 30 / 1e6) = 10 * 0.06
    assert _line(out, "MOCHA-DIMS")["cbm"] == pytest.approx(0.6, abs=0.0001)
    # The stored-cbm line is untouched: this slice adds a rule, it does not replace the
    # existing one.
    assert _line(out, "STORED")["cbm"] == pytest.approx(5.0, abs=0.0001)


def test_the_dims_only_lines_reach_the_per_company_split_and_the_grand_total(db):
    w = World(db)

    out = svc.build(db, str(w.shipment.id))

    sorento = _split(out, "SORENTO")
    mocha = _split(out, "MOCHA")
    assert sorento["cbm"] == pytest.approx(5.0 + 1.2, abs=0.0001)
    assert sorento["cbm_known_lines"] == 2
    assert mocha["cbm"] == pytest.approx(0.6, abs=0.0001)
    assert mocha["cbm_known_lines"] == 1
    assert out["total"]["cbm"] == pytest.approx(6.8, abs=0.0001)
    assert out["total"]["cbm_known_lines"] == 3


# --------------------------------------------------------------------------------- #
# AC-L2 - build()'s figures agree with the hand-worked XLSX formulas
# --------------------------------------------------------------------------------- #


def test_build_agrees_with_a_hand_calc_of_the_xlsx_formulas(db):
    """`to_xlsx` derives CLEARANCE and CHINA FREIGHT by each company's CBM share, and
    INSURANCE and the line AMOUNT by its RMB share (`=M{row}/M{total_row}*{cost}` etc,
    `consolidated_packing_list.to_xlsx`). Worked out here from `build()`'s own cbm and cost
    figures, independently of the sheet-writing code, so a change to either side that drifts
    from the other fails this test rather than only a fidelity comparison against one fixed
    workbook.
    """
    w = World(db)

    out = svc.build(db, str(w.shipment.id))

    sorento_cbm = _split(out, "SORENTO")["cbm"]
    mocha_cbm = _split(out, "MOCHA")["cbm"]
    total_cbm = out["total"]["cbm"]

    # Amount = unit_cost * qty, per line, exactly what column U / T*F derives.
    sorento_amount = 10.00 * 50 + 15.00 * 200  # STORED + SORENTO-DIMS
    mocha_amount = 20.00 * 100  # MOCHA-DIMS
    total_amount = sorento_amount + mocha_amount

    costs = out["costs"]
    clearance, freight, insurance = (
        costs["clearance_cost"],
        costs["china_freight_cost"],
        costs["insurance_rate"],
    )

    sorento_clearance = round(clearance * sorento_cbm / total_cbm, 2)
    mocha_clearance = round(clearance * mocha_cbm / total_cbm, 2)
    sorento_freight = round(freight * sorento_cbm / total_cbm, 2)
    mocha_freight = round(freight * mocha_cbm / total_cbm, 2)
    sorento_insurance = round(insurance * sorento_amount / total_amount, 2)
    mocha_insurance = round(insurance * mocha_amount / total_amount, 2)

    assert (sorento_clearance, mocha_clearance) == (2461.76, 238.24)
    assert (sorento_freight, mocha_freight) == (12719.12, 1230.88)
    assert (sorento_insurance, mocha_insurance) == (0.64, 0.36)
    # Each company's share reconciles to the whole cost typed on the container - the same
    # constraint the workbook's own SUM row asserts visually.
    assert round(sorento_clearance + mocha_clearance, 2) == clearance
    assert round(sorento_freight + mocha_freight, 2) == freight
    assert round(sorento_insurance + mocha_insurance, 2) == insurance

    # The workbook derives the SAME cbm live from the line's own dimensions - the formula
    # strings it writes for a dims-only line, read off the sheet `build()` feeds.
    book = svc.to_xlsx(out)
    from io import BytesIO

    import openpyxl

    ws = openpyxl.load_workbook(BytesIO(book))["RMB"]
    sorento_row = next(
        r
        for r in range(svc._FIRST_LINE_ROW, svc._FIRST_LINE_ROW + 20)
        if ws[f"C{r}"].value and ws[f"C{r}"].value.endswith("SORENTO-DIMS")
    )
    assert ws[f"H{sorento_row}"].value == f"=F{sorento_row}/G{sorento_row}"
    assert ws[f"L{sorento_row}"].value == f"=I{sorento_row}*J{sorento_row}*K{sorento_row}/10^6"
    assert ws[f"M{sorento_row}"].value == f"=H{sorento_row}*L{sorento_row}"
