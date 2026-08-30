"""AC-F7 - the container workbook IS `FSCU8103365.xlsx`, cell by cell.

The reference file (committed at `documentation/plans/scm/fixtures/FSCU8103365.xlsx`) is the
one Ms Tee builds by hand and the one the forwarder, the factories and our clearance agent all
read. A workbook that is nearly the familiar one is worse than none, because every reader has
to work out which column moved - so the expectations here are READ OFF the reference rather
than restated, and a change to the builder that drifts from it fails.

A shipment is seeded from the reference's own rows (three factories, one of them with MOCHA
goods under their own heading, one line nobody stated a pack size for), exported through
`build()` + `to_xlsx()`, and compared against the reference for: the header block, the two-row
column header with its merges, column widths, row heights, fonts, number formats, the per-line
formulas, the block amount merged down its rows, the subtotals, the rule row, the grand total,
the SORENTO / MOCHA footer and the 订单号 / 柜号 / 封号 lines.

WHERE THE REFERENCE DISAGREES WITH ITSELF, one form is picked and named here:

  * `CBM / CTN` is `=I*J*K/10^6` on some rows and `=I*J*K/1000000` on others. `10^6` is used.
  * `TOTAL RMB` is `=PRODUCT(T,F)` on most rows and `=T*F` on others. `T*F` is used.
  * A block's amount cell repeats as `=SUM(V18)` on most subtotal rows and as a range on two.
    `=SUM(V<first row of the block>)` is used.
  * Subtotal rows are 15.95 high in the first four blocks and 20.1 in the last four. 15.95.
  * The rule row of dashes skips H and S in the reference. All 22 columns get one here: a
    rule with two gaps in it reads as a missing figure rather than as a rule.
  * The grand total is a sum of every LINE cell in the reference. It sums the SUBTOTALS here
    (AC-F3.5 of part 3), which cannot count a quantity twice.
  * The reference freezes no panes. This freezes below the column header (AC-F3.4).
"""
from __future__ import annotations

import re
import uuid
from datetime import date
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

from app.models.procurement import InboundShipment, InboundShipmentLine, Supplier
from app.models.product import Brand, Product, ProductCategory, UnitOfMeasure
from app.services.scm import consolidated_packing_list as svc
from tests._pg_fixture import blank_session

MARKER = "ZZFID"

REFERENCE = (
    Path(__file__).resolve().parents[2]
    / "documentation"
    / "plans"
    / "scm"
    / "fixtures"
    / "FSCU8103365.xlsx"
)

#: Blocks of the reference this test reproduces: (A-column factory name, row range, brand).
#: Chosen for coverage rather than for size - AFFANNI carries both companies, HONGJIE's row
#: states a carton count and no pack size, and TECA's logo is a third brand that still counts
#: as SORENTO.
_BLOCKS = [
    ("AFFANNI 阿凡尼", range(18, 25), "SORENTO"),
    ("AFFANNI 阿凡尼", range(26, 29), "MOCHA"),
    ("CAIZHOU 彩洲 ", range(30, 35), "SORENTO"),
    ("HONGJIE 鸿洁", range(69, 70), "SORENTO"),
    ("TECA 缔家", range(62, 63), "SANDEL"),
]


@pytest.fixture(scope="module")
def reference():
    assert REFERENCE.exists(), f"reference workbook missing: {REFERENCE}"
    return openpyxl.load_workbook(REFERENCE)["RMB"]


@pytest.fixture
def db():
    with blank_session() as session:
        yield session


def _num(cell):
    return None if cell.value in (None, "") else float(cell.value)


class World:
    """The reference's own lines, on a container in our database."""

    def __init__(self, db, reference):
        self.db = db
        self.ref = reference
        self.tag = uuid.uuid4().hex[:8].upper()
        self.category = ProductCategory(
            id=str(uuid.uuid4()),
            category_code=f"{MARKER}-CAT-{self.tag}",
            category_name=f"{MARKER} category",
        )
        self.uom = UnitOfMeasure(
            id=str(uuid.uuid4()), uom_code=f"{MARKER}{self.tag}"[:20], uom_name="pcs"
        )
        db.add_all([self.category, self.uom])
        db.flush()
        self.brands: dict[str, Brand] = {}
        self.suppliers: dict[str, Supplier] = {}

        self.shipment = InboundShipment(
            id=str(uuid.uuid4()),
            shipment_number=f"{MARKER}-SH-{self.tag}",
            shipment_date=date(2026, 7, 20),
            shipping_container_number="FSCU8103365",
            shipment_status="in_transit",
            # The twelve header lines, from the reference's own B column.
            loading_date=date(2026, 7, 17),
            etd_date=date(2026, 7, 23),
            estimated_arrival_date=date(2026, 7, 27),
            seal_number="J0713349",
            forwarder_order_ref="CNH1098313",
            consignee="SORENTO SDN BHD",
            shipper="SHENZHEN XINDESHENG TRADING CO.,LTD",
            china_forwarder="ONE TOUCH",
            free_days_available=14,
            delivery_warehouse="BRW",
            clearance_cost=2700,
            china_freight_cost=13950,
            insurance_rate=1,
        )
        db.add(self.shipment)
        db.flush()

        self.rows: list[dict] = []
        for name, rng, brand_code in _BLOCKS:
            for r in rng:
                self.rows.append(self._line(name, r, brand_code))
        db.flush()

    def brand(self, code: str) -> Brand:
        if code not in self.brands:
            # MOCHA is the one the split is keyed off, so it keeps its real code; the rest
            # are tagged, because a brand code is unique and CI's schema is shared.
            brand_code = "MOCHA" if code == "MOCHA" else f"{code}-{self.tag}"[:50]
            row = Brand(
                id=str(uuid.uuid4()), brand_code=brand_code, brand_name=code, is_active=True
            )
            self.db.add(row)
            self.db.flush()
            self.brands[code] = row
        return self.brands[code]

    def supplier(self, name: str) -> Supplier:
        if name not in self.suppliers:
            row = Supplier(
                id=str(uuid.uuid4()),
                supplier_code=f"{MARKER}-{len(self.suppliers)}-{self.tag}"[:50],
                # The NAME is the reference's own, so the block heading can be compared
                # against it directly.
                supplier_name=name,
                is_active=True,
            )
            self.db.add(row)
            self.db.flush()
            self.suppliers[name] = row
        return self.suppliers[name]

    def _line(self, factory: str, r: int, brand_code: str) -> dict:
        ref = self.ref
        product = Product(
            id=str(uuid.uuid4()),
            # The reference names one model twice (CAIZHOU rows 30 and 33), and a product
            # code is unique per company, so the reference ROW is part of the code here.
            product_code=f"{ref[f'C{r}'].value}-{r}-{self.tag}",
            product_name=ref[f"D{r}"].value,
            category_id=self.category.id,
            base_uom_id=self.uom.id,
            brand_id=self.brand(brand_code).id,
            list_price=0,
            is_active=True,
        )
        self.db.add(product)
        self.db.flush()
        stated_cartons = ref[f"H{r}"].value
        line = InboundShipmentLine(
            id=str(uuid.uuid4()),
            shipment_id=self.shipment.id,
            supplier_id=self.supplier(factory).id,
            product_id=product.id,
            quantity_shipped=int(_num(ref[f"F{r}"])),
            cartons_count=int(stated_cartons) if isinstance(stated_cartons, (int, float)) else 1,
            material=ref[f"E{r}"].value,
            pcs_per_carton=_num(ref[f"G{r}"]),
            carton_length_cm=_num(ref[f"I{r}"]),
            carton_width_cm=_num(ref[f"J{r}"]),
            carton_height_cm=_num(ref[f"K{r}"]),
            net_weight_per_carton=_num(ref[f"N{r}"]),
            gross_weight_per_carton=_num(ref[f"O{r}"]),
            unit_cost=_num(ref[f"T{r}"]),
            currency="CNY",
            remarks=ref[f"S{r}"].value,
        )
        self.db.add(line)
        return {"row": r, "factory": factory, "brand": brand_code, "line": line}


@pytest.fixture
def sheet(db, reference):
    """The workbook this builder produces for the reference's lines."""
    world = World(db, reference)
    payload = svc.build(db, str(world.shipment.id))
    book = openpyxl.load_workbook(BytesIO(svc.to_xlsx(payload)))
    return book["RMB"], world


# --------------------------------------------------------------------------------- #
# the sheet, the header block and the column header
# --------------------------------------------------------------------------------- #


def test_the_sheet_is_the_reference_tab(sheet, reference):
    out, _world = sheet
    assert out.title == reference.title == "RMB"


def test_the_twelve_header_labels_are_the_reference_ones(sheet, reference):
    out, _world = sheet
    for row in range(1, 13):
        assert out[f"A{row}"].value == reference[f"A{row}"].value


def test_the_three_header_dates_are_dates_formatted_the_reference_way(sheet, reference):
    """A date, not the ISO string the payload carries: Excel left-aligns that as text and
    prints 2026-07-17 on a document everybody else reads dd/mm/yyyy.

    openpyxl reads a date cell back as a `datetime`, in this workbook and in the reference
    alike, so the day is what is compared. The reference's own ETA is TEXT ('27/07/2026') in
    a date-formatted cell - one of the hand-typing faults this sheet exists to stop.
    """
    out, _world = sheet
    assert out["B1"].value.date() == reference["B1"].value.date() == date(2026, 7, 17)
    assert out["B2"].value.date() == reference["B2"].value.date() == date(2026, 7, 23)
    assert out["B3"].value.date() == date(2026, 7, 27)
    for coord in ("B1", "B2", "B3"):
        assert out[coord].number_format == reference[coord].number_format
        assert "dd/mm/yyyy" in out[coord].number_format


def test_the_rest_of_the_header_block_is_what_was_typed_on_the_container(sheet, reference):
    out, _world = sheet
    for coord in ("B4", "B5", "B6", "B7", "B8", "B9", "B11", "B12"):
        assert out[coord].value == reference[coord].value, coord


def test_the_two_row_column_header_matches_label_for_label(sheet, reference):
    out, _world = sheet
    for column in "ABCDEFGHIJKLMNOPQRSTUV":
        for row in (svc._HEADER_ROW, svc._SUBHEADER_ROW):
            coord = f"{column}{row}"
            assert out[coord].value == reference[coord].value, coord


def test_the_column_header_merges_are_the_reference_ones(sheet, reference):
    out, _world = sheet
    header_rows = {svc._HEADER_ROW, svc._SUBHEADER_ROW}
    expected = {
        str(m)
        for m in reference.merged_cells.ranges
        if m.min_row in header_rows and m.max_row in header_rows
    }
    actual = {
        str(m)
        for m in out.merged_cells.ranges
        if m.min_row in header_rows and m.max_row in header_rows
    }
    assert actual == expected


def test_the_column_header_is_bold_calibri_12_and_centred(sheet, reference):
    out, _world = sheet
    for column in "ABCDEFGHILMNOPQRSTU":
        cell = out[f"{column}{svc._HEADER_ROW}"]
        if cell.value is None:
            continue
        assert (cell.font.name, cell.font.sz, cell.font.b) == ("Calibri", 12.0, True), column
        assert cell.alignment.horizontal == "center"
        assert cell.alignment.vertical == "center"


def test_the_column_widths_are_the_reference_ones(sheet, reference):
    out, _world = sheet
    for column in "ABCDEFGHIJKLMNOPQRSTUV":
        expected = reference.column_dimensions.get(column)
        actual = out.column_dimensions.get(column)
        expected_width = getattr(expected, "width", None) if expected else None
        actual_width = getattr(actual, "width", None) if actual else None
        if expected_width is None:
            assert actual_width is None, f"{column} is sized here and not in the reference"
        else:
            assert actual_width == pytest.approx(expected_width, abs=0.05), column


def test_the_header_block_row_that_wraps_keeps_its_height(sheet, reference):
    out, _world = sheet
    assert out.row_dimensions[12].height == pytest.approx(reference.row_dimensions[12].height)


# --------------------------------------------------------------------------------- #
# the lines
# --------------------------------------------------------------------------------- #


def _line_rows(out) -> list[int]:
    """Every row that carries a numbered line."""
    return [
        r
        for r in range(svc._FIRST_LINE_ROW, out.max_row + 1)
        if isinstance(out.cell(row=r, column=2).value, int)
    ]


def _subtotal_rows(out) -> list[int]:
    """A block subtotal sums a RANGE; the grand total sums the subtotal cells by name."""
    return [
        r
        for r in range(svc._FIRST_LINE_ROW, out.max_row + 1)
        if re.fullmatch(r"=SUM\(F\d+:F\d+\)", str(out.cell(row=r, column=6).value or ""))
    ]


def _grand_total_row(out) -> int:
    return next(
        r
        for r in range(out.max_row, svc._FIRST_LINE_ROW, -1)
        if str(out.cell(row=r, column=6).value or "").startswith("=SUM(F")
        and ":" not in str(out.cell(row=r, column=6).value)
    )


def test_every_line_is_the_reference_font_centred_and_35_1_high(sheet, reference):
    out, _world = sheet
    for r in _line_rows(out):
        assert out.row_dimensions[r].height == pytest.approx(
            reference.row_dimensions[18].height
        )
        for column in "ABCDEFGHIJKLMNOPQRSTU":
            cell = out[f"{column}{r}"]
            assert (cell.font.name, cell.font.sz) == ("Calibri", 12.0), f"{column}{r}"
            assert cell.alignment.horizontal == "center", f"{column}{r}"
            assert cell.alignment.vertical == "center", f"{column}{r}"


def test_every_line_carries_the_reference_number_formats(sheet, reference):
    out, _world = sheet
    for r in _line_rows(out):
        for column in "IJKT":
            assert out[f"{column}{r}"].number_format == "0.00", f"{column}{r}"
        for column in "LMNOPQ":
            assert out[f"{column}{r}"].number_format == "0.00;[Red]0.00", f"{column}{r}"
        assert out[f"U{r}"].number_format == reference["U18"].number_format == "#,##0.00"


def test_every_line_derives_the_same_six_figures_as_the_reference(sheet):
    out, _world = sheet
    for r in _line_rows(out):
        assert out[f"H{r}"].value in (f"=F{r}/G{r}", out[f"H{r}"].value)
        if isinstance(out[f"H{r}"].value, str):
            assert out[f"H{r}"].value == f"=F{r}/G{r}"
        assert out[f"L{r}"].value == f"=I{r}*J{r}*K{r}/10^6"
        assert out[f"M{r}"].value == f"=H{r}*L{r}"
        assert out[f"P{r}"].value == f"=N{r}*H{r}"
        assert out[f"Q{r}"].value == f"=O{r}*H{r}"
        assert out[f"U{r}"].value == f"=T{r}*F{r}"


def test_a_line_with_no_pack_size_states_its_carton_count_instead(sheet, reference):
    """Reference row 69 (HONGJIE) states 4 cartons and no pcs per carton."""
    out, world = sheet
    code = f"{reference['C69'].value}-69-{world.tag}"
    row = next(r for r in _line_rows(out) if out[f"C{r}"].value == code)
    assert out[f"G{row}"].value is None
    assert out[f"H{row}"].value == int(reference["H69"].value)


def test_the_line_values_are_the_reference_values(sheet, reference):
    out, world = sheet
    by_code = {out[f"C{r}"].value: r for r in _line_rows(out)}
    for seeded in world.rows:
        ref_row = seeded["row"]
        code = f"{reference[f'C{ref_row}'].value}-{ref_row}-{world.tag}"
        r = by_code[code]
        assert out[f"D{r}"].value == reference[f"D{ref_row}"].value
        assert out[f"E{r}"].value == reference[f"E{ref_row}"].value
        assert float(out[f"F{r}"].value) == float(reference[f"F{ref_row}"].value)
        assert _num(out[f"I{r}"]) == _num(reference[f"I{ref_row}"])
        assert _num(out[f"J{r}"]) == _num(reference[f"J{ref_row}"])
        assert _num(out[f"K{r}"]) == _num(reference[f"K{ref_row}"])
        assert _num(out[f"N{r}"]) == _num(reference[f"N{ref_row}"])
        assert _num(out[f"O{r}"]) == _num(reference[f"O{ref_row}"])
        assert _num(out[f"T{r}"]) == _num(reference[f"T{ref_row}"])
        # REMARKS is the supplier's own note and nothing else (R20).
        assert (out[f"S{r}"].value or "") == (reference[f"S{ref_row}"].value or "")


def test_the_lines_are_numbered_once_across_the_container(sheet):
    out, world = sheet
    numbers = [out.cell(row=r, column=2).value for r in _line_rows(out)]
    assert numbers == list(range(1, len(world.rows) + 1))


def test_a_mocha_block_is_headed_the_reference_way(sheet):
    out, _world = sheet
    headings = {out[f"A{r}"].value for r in _line_rows(out)}
    assert "AFFANNI 阿凡尼" in headings
    assert "AFFANNI 阿凡尼 (MOCHA)" in headings


# --------------------------------------------------------------------------------- #
# the block amount, the subtotals, the rule and the grand total
# --------------------------------------------------------------------------------- #


def test_each_blocks_amount_is_summed_once_and_merged_down_its_rows(sheet):
    out, _world = sheet
    merges = {str(m) for m in out.merged_cells.ranges}
    for start, end in _block_ranges(out):
        assert out[f"V{start}"].value == f"=SUM(U{start}:U{end})"
        if end > start:
            assert f"V{start}:V{end}" in merges
        assert out[f"V{start}"].number_format == "#,##0.00"


def _block_ranges(out) -> list[tuple[int, int]]:
    ranges: list[tuple[int, int]] = []
    start = None
    for r in range(svc._FIRST_LINE_ROW, out.max_row + 1):
        numbered = isinstance(out.cell(row=r, column=2).value, int)
        if numbered and start is None:
            start = r
        elif not numbered and start is not None:
            ranges.append((start, r - 1))
            start = None
    return ranges


def test_every_block_subtotals_its_own_rows_in_bold_red(sheet, reference):
    out, _world = sheet
    blocks = _block_ranges(out)
    subtotals = _subtotal_rows(out)
    assert len(subtotals) == len(blocks)
    for (start, end), r in zip(blocks, subtotals):
        assert r == end + 1
        for column in ("F", "G", "H", "M", "P", "Q", "U"):
            assert out[f"{column}{r}"].value == f"=SUM({column}{start}:{column}{end})"
            cell = out[f"{column}{r}"]
            assert cell.font.b is True, f"{column}{r}"
            assert cell.font.color is not None and cell.font.color.rgb == "FFFF0000"
        assert out[f"V{r}"].value == f"=SUM(V{start})"
        assert out.row_dimensions[r].height == pytest.approx(
            reference.row_dimensions[25].height
        )
        for column in ("F", "G", "H"):
            assert out[f"{column}{r}"].number_format == reference[f"{column}25"].number_format
        for column in ("M", "P", "Q"):
            assert out[f"{column}{r}"].number_format == reference[f"{column}25"].number_format
        assert out[f"U{r}"].number_format == reference["U25"].number_format


def test_a_rule_of_dashes_closes_the_blocks(sheet):
    out, _world = sheet
    rule = _grand_total_row(out) - 1
    assert [out.cell(row=rule, column=c).value for c in range(1, 23)] == ["-"] * 22


def test_the_grand_total_sums_the_subtotals_and_prints_them_the_reference_way(
    sheet, reference
):
    out, _world = sheet
    total_row = _grand_total_row(out)
    refs = _subtotal_rows(out)
    for column in ("F", "G", "H", "M", "P", "Q", "U"):
        expected = "=SUM(" + ",".join(f"{column}{r}" for r in refs) + ")"
        assert out[f"{column}{total_row}"].value == expected
        assert out[f"{column}{total_row}"].font.b is True
        assert (
            out[f"{column}{total_row}"].number_format
            == reference[f"{column}81"].number_format
        ), column


# --------------------------------------------------------------------------------- #
# the footer
# --------------------------------------------------------------------------------- #


def _footer_rows(out) -> tuple[int, int, int, int]:
    """(sorento, mocha, totals, labels), by the CBM label the reference puts on the last."""
    labels = next(
        r
        for r in range(out.max_row, svc._FIRST_LINE_ROW, -1)
        if out.cell(row=r, column=13).value == "CBM"
    )
    return labels - 3, labels - 2, labels - 1, labels


def test_the_footer_apportions_the_container_the_reference_way(sheet, reference):
    out, _world = sheet
    sorento, mocha, totals, labels = _footer_rows(out)
    total_row = _grand_total_row(out)

    assert out[f"L{sorento}"].value == "SORENTO"
    assert out[f"T{sorento}"].value == "SORENTO"
    assert out[f"L{mocha}"].value == "MOCHA"
    assert out[f"T{mocha}"].value == "MOCHA"
    for r in (sorento, mocha):
        # Clearance and China freight follow the VOLUME, insurance follows the AMOUNT.
        assert out[f"N{r}"].value == f"=M{r}/M{total_row}*2700.0"
        assert out[f"O{r}"].value == f"=U{r}/U{total_row}*1.0"
        assert out[f"P{r}"].value == f"=M{r}/M{total_row}*13950.0"
        for column in ("M", "N", "O", "P", "U"):
            assert out[f"{column}{r}"].font.b is True, f"{column}{r}"
            assert (
                out[f"{column}{r}"].number_format == reference[f"{column}83"].number_format
            ), f"{column}{r}"

    for column in ("M", "N", "O", "P", "U"):
        assert out[f"{column}{totals}"].value == f"={column}{sorento}+{column}{mocha}"
        assert out[f"{column}{totals}"].font.b is True

    for column, label in (
        ("M", "CBM"),
        ("N", "CLEARANCE"),
        ("O", "INSURANCE"),
        ("P", "CHINA FREIGHT"),
        ("U", "TOTAL AMOUNT"),
    ):
        assert out[f"{column}{labels}"].value == reference[f"{column}86"].value == label


def test_the_three_forwarder_lines_are_written_the_reference_way(sheet, reference):
    out, _world = sheet
    _s, _m, _t, labels = _footer_rows(out)
    assert out[f"C{labels}"].value == reference["C86"].value == "订单号:CNH1098313"
    assert out[f"C{labels + 1}"].value == reference["C87"].value == "柜号:FSCU8103365"
    assert out[f"C{labels + 2}"].value == reference["C88"].value == "封号:J0713349"


# --------------------------------------------------------------------------------- #
# R20 - what the workbook no longer prints
# --------------------------------------------------------------------------------- #


def test_the_workbook_prints_the_shipment_and_nothing_else(sheet):
    out, _world = sheet
    printed = [
        str(cell.value)
        for row in out.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    ]
    assert not any("Not packed" in v for v in printed)
    assert not any("loading plan" in v.lower() for v in printed)


def test_the_payload_carries_no_plan_comparison(db, reference):
    world = World(db, reference)
    payload = svc.build(db, str(world.shipment.id))
    for factory in payload["factories"]:
        assert "not_packed" not in factory
        assert "has_pack_plan" not in factory
        assert "notice_id" not in factory
        for line in factory["lines"]:
            assert "discrepancies" not in line
