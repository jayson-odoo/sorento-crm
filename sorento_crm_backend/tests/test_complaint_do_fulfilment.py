"""Service-level tests for ComplaintFulfilmentService — the centralized
complaint <-> Delivery Order auto-fulfilment helper.

Covers the UAC lines of docs/plans/PLAN-complaint-do-auto-fulfilment.md:

  Linking : L1 L2 L3 L4 L5 L6 L7 L8
  Fulfil  : F1 F2 F3 F4 F5 F6 F7 F8 F9
  Freeze  : Z4 (pending DO editable + unlink on token removal)
  Notify  : N4 (notify-once via delivery_notified_at), N7 (dry-run no send/no link)
  Perf    : P1 (delta-only / batched — no per-row complaint queries)

Runs against the live Postgres test DB (same convention as the other complaint
service tests): seed rows with unique prefixes, assert, clean up. The new
``complaint_fulfilment_orders`` table already exists at alembic head 251.

External notification dispatch (Respond/email/RQ) is NOT exercised here — these
tests drive ``reconcile_links_and_status`` directly and inspect the returned
``(warnings, notify_payloads)`` plus DB state. Notification side effects have
their own file (test_complaint_do_notify.py).
"""
from __future__ import annotations

import uuid
from datetime import datetime
from typing import Iterator, Optional

import pytest
from sqlalchemy import event, text
from sqlalchemy.orm import Session

from app.database import SessionLocal, engine
from app.models.complaints import Complaint, ComplaintFulfilmentOrder
from app.models.order import Order
from app.services.complaint_fulfilment_service import ComplaintFulfilmentService

C_PREFIX = "CDF-"
O_PREFIX = "CDFDO-"


def _cleanup(conn) -> None:
    for sql in (
        "DELETE FROM complaint_fulfilment_orders WHERE order_id IN "
        "(SELECT id FROM orders WHERE order_number LIKE 'CDFDO-%')",
        "DELETE FROM complaint_fulfilment_orders WHERE complaint_id IN "
        "(SELECT id FROM complaints WHERE complaint_number LIKE 'CDF-%')",
        "DELETE FROM orders WHERE order_number LIKE 'CDFDO-%'",
        "DELETE FROM complaints WHERE complaint_number LIKE 'CDF-%'",
    ):
        try:
            conn.execute(text(sql))
        except Exception:
            conn.rollback()
    conn.commit()


@pytest.fixture(autouse=True)
def _clean_state():
    with engine.connect() as conn:
        _cleanup(conn)
    yield
    with engine.connect() as conn:
        _cleanup(conn)


@pytest.fixture
def db() -> Iterator[Session]:
    s = SessionLocal()
    try:
        yield s
    finally:
        s.rollback()
        s.close()


# --------------------------------------------------------------------------- #
# seed helpers
# --------------------------------------------------------------------------- #
def _mk_complaint(
    db: Session,
    number: str,
    *,
    status: str = "processed_by_cs",
    contact_id: Optional[str] = None,
    respond_inbox_url: Optional[str] = None,
) -> Complaint:
    c = Complaint(
        id=str(uuid.uuid4()),
        complaint_number=number,
        status=status,
        contact_id=contact_id,
        respond_inbox_url=respond_inbox_url,
    )
    db.add(c)
    db.commit()
    db.refresh(c)
    return c


def _delivered_status_id(db: Session) -> str:
    """The DELIVERED order-status id (a DO is 'delivered' only with date AND this status)."""
    row = db.execute(
        text("SELECT id FROM order_statuses WHERE lower(status_code) = 'delivered' LIMIT 1")
    ).first()
    assert row is not None, "seed order_statuses must contain a 'delivered' code"
    return str(row[0])


def _mk_order(
    db: Session,
    number: str,
    *,
    remarks_cs: Optional[str] = None,
    delivered: bool = False,
    cancelled: bool = False,
) -> Order:
    o = Order(
        id=str(uuid.uuid4()),
        order_number=number,
        remarks_cs=remarks_cs,
        # Delivered = delivery date AND a delivered status (both required).
        actual_delivery_date=datetime(2026, 6, 1, 10, 0) if delivered else None,
        order_status_id=_delivered_status_id(db) if delivered else None,
        is_cancelled=cancelled,
    )
    db.add(o)
    db.commit()
    db.refresh(o)
    return o


def _mark_delivered(db: Session, order: Order, when: Optional[datetime] = None) -> None:
    """Flip an existing order to delivered: sets BOTH the date and the delivered status."""
    order.actual_delivery_date = when or datetime(2026, 6, 2, 9, 0)
    order.order_status_id = _delivered_status_id(db)
    db.commit()


def _reconcile(db: Session, *changes, dry_run: bool = False):
    """Run the helper for the given (order, old_remarks) changes."""
    svc = ComplaintFulfilmentService(db)
    payload = [{"order": o, "old_remarks": old} for (o, old) in changes]
    warnings, notify = svc.reconcile_links_and_status(payload, dry_run=dry_run)
    if not dry_run:
        db.commit()
    return warnings, notify


def _links(db: Session, complaint_id: str) -> list[ComplaintFulfilmentOrder]:
    return (
        db.query(ComplaintFulfilmentOrder)
        .filter(ComplaintFulfilmentOrder.complaint_id == str(complaint_id))
        .all()
    )


def _status(db: Session, complaint_id: str) -> str:
    db.expire_all()
    c = db.query(Complaint).filter(Complaint.id == complaint_id).first()
    return c.status if c else ""


# --------------------------------------------------------------------------- #
# Linking
# --------------------------------------------------------------------------- #
def test_L1_single_token_links(db: Session) -> None:
    """UAC-L1: a DO whose Remarks CS names one processed_by_cs complaint links."""
    c = _mk_complaint(db, f"{C_PREFIX}0042")
    o = _mk_order(db, f"{O_PREFIX}1", remarks_cs=f"{C_PREFIX}0042")
    warnings, _ = _reconcile(db, (o, None))
    assert warnings == []
    links = _links(db, c.id)
    assert len(links) == 1 and str(links[0].order_id) == str(o.id)


def test_L2_multi_token_ampersand(db: Session) -> None:
    """UAC-L2: "CMP-1 & CMP-2" links the DO to both complaints."""
    c1 = _mk_complaint(db, f"{C_PREFIX}0001")
    c2 = _mk_complaint(db, f"{C_PREFIX}0002")
    o = _mk_order(db, f"{O_PREFIX}2", remarks_cs=f"{C_PREFIX}0001 & {C_PREFIX}0002")
    _reconcile(db, (o, None))
    assert len(_links(db, c1.id)) == 1
    assert len(_links(db, c2.id)) == 1


def test_L3_non_matching_token_ignored(db: Session) -> None:
    """UAC-L3: a token matching no complaint is ignored — no error, no warning."""
    o = _mk_order(db, f"{O_PREFIX}3", remarks_cs=f"{C_PREFIX}NOSUCH")
    warnings, notify = _reconcile(db, (o, None))
    assert warnings == []
    assert notify == []


def test_L4_not_processed_by_cs_not_linked_warns(db: Session) -> None:
    """UAC-L4: token -> complaint in an earlier open state is not linked + warns."""
    c = _mk_complaint(db, f"{C_PREFIX}OPEN", status="responded")
    o = _mk_order(db, f"{O_PREFIX}4", remarks_cs=f"{C_PREFIX}OPEN")
    warnings, _ = _reconcile(db, (o, None))
    assert _links(db, c.id) == []
    assert any("not yet processed by CS" in w for w in warnings), warnings


def test_L5_closed_or_rejected_not_linked_warns(db: Session) -> None:
    """UAC-L5: token -> closed/rejected complaint is not linked + warns."""
    closed = _mk_complaint(db, f"{C_PREFIX}CLOSED", status="closed")
    rejected = _mk_complaint(db, f"{C_PREFIX}REJ", status="rejected")
    o = _mk_order(
        db, f"{O_PREFIX}5", remarks_cs=f"{C_PREFIX}CLOSED {C_PREFIX}REJ"
    )
    warnings, _ = _reconcile(db, (o, None))
    assert _links(db, closed.id) == []
    assert _links(db, rejected.id) == []
    assert len([w for w in warnings if "not linked" in w]) == 2, warnings


def test_L6_identical_reimport_no_new_link_no_event(db: Session) -> None:
    """UAC-L6: re-running with identical Remarks CS adds no link and no notify."""
    c = _mk_complaint(db, f"{C_PREFIX}IDEM")
    o = _mk_order(db, f"{O_PREFIX}6", remarks_cs=f"{C_PREFIX}IDEM")
    _reconcile(db, (o, None))
    assert len(_links(db, c.id)) == 1
    # second pass, same remarks
    warnings, notify = _reconcile(db, (o, f"{C_PREFIX}IDEM"))
    assert len(_links(db, c.id)) == 1
    assert notify == []


def test_L7_case_insensitive_and_separators(db: Session) -> None:
    """UAC-L7: match is case-insensitive; '&', ',', whitespace all split."""
    c1 = _mk_complaint(db, f"{C_PREFIX}AA")
    c2 = _mk_complaint(db, f"{C_PREFIX}BB")
    c3 = _mk_complaint(db, f"{C_PREFIX}CC")
    # mixed case + comma + ampersand + extra whitespace
    o = _mk_order(
        db,
        f"{O_PREFIX}7",
        remarks_cs=f"  {C_PREFIX.lower()}aa ,  {C_PREFIX.lower()}bb &{C_PREFIX.lower()}cc ",
    )
    warnings, _ = _reconcile(db, (o, None))
    assert warnings == []
    assert len(_links(db, c1.id)) == 1
    assert len(_links(db, c2.id)) == 1
    assert len(_links(db, c3.id)) == 1


def test_L8_many_to_many(db: Session) -> None:
    """UAC-L8: one DO links many complaints; one complaint accumulates many DOs."""
    c1 = _mk_complaint(db, f"{C_PREFIX}M1")
    c2 = _mk_complaint(db, f"{C_PREFIX}M2")
    # one DO -> two complaints
    o1 = _mk_order(db, f"{O_PREFIX}81", remarks_cs=f"{C_PREFIX}M1 {C_PREFIX}M2")
    _reconcile(db, (o1, None))
    # second DO -> complaint M1 as well (partial shipments)
    o2 = _mk_order(db, f"{O_PREFIX}82", remarks_cs=f"{C_PREFIX}M1")
    _reconcile(db, (o2, None))
    assert {str(l.order_id) for l in _links(db, c1.id)} == {str(o1.id), str(o2.id)}
    assert {str(l.order_id) for l in _links(db, c2.id)} == {str(o1.id)}


# --------------------------------------------------------------------------- #
# Fulfil / status
# --------------------------------------------------------------------------- #
def test_F1_single_delivered_becomes_fulfilled(db: Session) -> None:
    """UAC-F1: lone linked DO delivers -> complaint fulfilled."""
    c = _mk_complaint(db, f"{C_PREFIX}F1")
    o = _mk_order(db, f"{O_PREFIX}F1", remarks_cs=f"{C_PREFIX}F1")
    _reconcile(db, (o, None))
    assert _status(db, c.id) == "processed_by_cs"
    _mark_delivered(db, o)
    _reconcile(db, (o, f"{C_PREFIX}F1"))
    assert _status(db, c.id) == "fulfilled"


def test_F2_one_delivered_one_pending_stays_open(db: Session) -> None:
    """UAC-F2: two linked DOs, one delivered one pending -> stays processed_by_cs."""
    c = _mk_complaint(db, f"{C_PREFIX}F2")
    o1 = _mk_order(db, f"{O_PREFIX}F2a", remarks_cs=f"{C_PREFIX}F2", delivered=True)
    o2 = _mk_order(db, f"{O_PREFIX}F2b", remarks_cs=f"{C_PREFIX}F2", delivered=False)
    _reconcile(db, (o1, None), (o2, None))
    assert _status(db, c.id) == "processed_by_cs"


def test_F3_both_delivered_fulfilled(db: Session) -> None:
    """UAC-F3: both linked DOs delivered -> fulfilled."""
    c = _mk_complaint(db, f"{C_PREFIX}F3")
    o1 = _mk_order(db, f"{O_PREFIX}F3a", remarks_cs=f"{C_PREFIX}F3", delivered=True)
    o2 = _mk_order(db, f"{O_PREFIX}F3b", remarks_cs=f"{C_PREFIX}F3", delivered=True)
    _reconcile(db, (o1, None), (o2, None))
    assert _status(db, c.id) == "fulfilled"


def test_F4_cancelled_excluded(db: Session) -> None:
    """UAC-F4: one delivered + one cancelled -> fulfilled (cancelled excluded)."""
    c = _mk_complaint(db, f"{C_PREFIX}F4")
    o1 = _mk_order(db, f"{O_PREFIX}F4a", remarks_cs=f"{C_PREFIX}F4", delivered=True)
    o2 = _mk_order(
        db, f"{O_PREFIX}F4b", remarks_cs=f"{C_PREFIX}F4", delivered=False, cancelled=True
    )
    _reconcile(db, (o1, None), (o2, None))
    assert _status(db, c.id) == "fulfilled"


def test_F5_reopen_on_new_pending_link(db: Session) -> None:
    """UAC-F5: a fulfilled complaint reopens to processed_by_cs when a new
    non-delivered DO links."""
    c = _mk_complaint(db, f"{C_PREFIX}F5")
    o1 = _mk_order(db, f"{O_PREFIX}F5a", remarks_cs=f"{C_PREFIX}F5", delivered=True)
    _reconcile(db, (o1, None))
    assert _status(db, c.id) == "fulfilled"
    o2 = _mk_order(db, f"{O_PREFIX}F5b", remarks_cs=f"{C_PREFIX}F5", delivered=False)
    _reconcile(db, (o2, None))
    assert _status(db, c.id) == "processed_by_cs"
    # reopen must clear the stale resolution timestamp set during fulfilment
    db.expire_all()
    reopened = db.query(Complaint).filter(Complaint.id == c.id).first()
    assert reopened is not None and reopened.resolved_at is None


def test_F6_reopen_then_deliver_refulfils(db: Session) -> None:
    """UAC-F6: after reopen, the new DO delivers -> fulfilled again."""
    c = _mk_complaint(db, f"{C_PREFIX}F6")
    o1 = _mk_order(db, f"{O_PREFIX}F6a", remarks_cs=f"{C_PREFIX}F6", delivered=True)
    _reconcile(db, (o1, None))
    o2 = _mk_order(db, f"{O_PREFIX}F6b", remarks_cs=f"{C_PREFIX}F6", delivered=False)
    _reconcile(db, (o2, None))
    assert _status(db, c.id) == "processed_by_cs"
    _mark_delivered(db, o2, datetime(2026, 6, 3, 8, 0))
    _reconcile(db, (o2, f"{C_PREFIX}F6"))
    assert _status(db, c.id) == "fulfilled"


def test_F7_cancel_lone_pending_fulfils(db: Session) -> None:
    """UAC-F7: cancelling the lone pending DO -> all non-cancelled delivered -> fulfilled."""
    c = _mk_complaint(db, f"{C_PREFIX}F7")
    o1 = _mk_order(db, f"{O_PREFIX}F7a", remarks_cs=f"{C_PREFIX}F7", delivered=True)
    o2 = _mk_order(db, f"{O_PREFIX}F7b", remarks_cs=f"{C_PREFIX}F7", delivered=False)
    _reconcile(db, (o1, None), (o2, None))
    assert _status(db, c.id) == "processed_by_cs"
    o2.is_cancelled = True
    db.commit()
    _reconcile(db, (o2, f"{C_PREFIX}F7"))
    assert _status(db, c.id) == "fulfilled"


def test_F8_sticky_terminal_unchanged(db: Session) -> None:
    """UAC-F8: a closed/rejected complaint is never linked, so a delivery never
    flips its status (sticky)."""
    closed = _mk_complaint(db, f"{C_PREFIX}F8", status="closed")
    o = _mk_order(
        db, f"{O_PREFIX}F8", remarks_cs=f"{C_PREFIX}F8", delivered=True
    )
    _reconcile(db, (o, None))
    assert _links(db, closed.id) == []
    assert _status(db, closed.id) == "closed"


def test_F9_sla_resolved_emit_on_fulfil_only(db: Session, monkeypatch) -> None:
    """UAC-F9: fulfil emits SLA 'resolved'; reopen does not re-emit."""
    import app.services.form_sla_service as sla_mod

    calls: list[tuple] = []
    monkeypatch.setattr(
        sla_mod,
        "emit_form_event",
        lambda db, etype, eid, event, **kw: calls.append((etype, eid, event)),
    )

    c = _mk_complaint(db, f"{C_PREFIX}F9")
    o1 = _mk_order(db, f"{O_PREFIX}F9a", remarks_cs=f"{C_PREFIX}F9", delivered=True)
    _reconcile(db, (o1, None))
    assert _status(db, c.id) == "fulfilled"
    resolved = [x for x in calls if x[2] == "resolved" and x[1] == str(c.id)]
    assert len(resolved) == 1, calls

    # Reopen: a new pending DO links -> processed_by_cs, NO new resolved emit.
    calls.clear()
    o2 = _mk_order(db, f"{O_PREFIX}F9b", remarks_cs=f"{C_PREFIX}F9", delivered=False)
    _reconcile(db, (o2, None))
    assert _status(db, c.id) == "processed_by_cs"
    assert [x for x in calls if x[2] == "resolved"] == []


# --------------------------------------------------------------------------- #
# Freeze — pending DO editable + unlink (UAC-Z4)
# --------------------------------------------------------------------------- #
def test_Z4_pending_unlink_on_token_removal(db: Session) -> None:
    """UAC-Z4: a pending DO's Remarks CS is editable — removing a token unlinks the
    complaint and the complaint stays open (processed_by_cs)."""
    c = _mk_complaint(db, f"{C_PREFIX}Z4")
    o = _mk_order(db, f"{O_PREFIX}Z4", remarks_cs=f"{C_PREFIX}Z4")
    _reconcile(db, (o, None))
    assert len(_links(db, c.id)) == 1
    # not delivered -> not frozen -> remove the token
    o.remarks_cs = ""
    db.commit()
    _reconcile(db, (o, f"{C_PREFIX}Z4"))
    assert _links(db, c.id) == []
    assert _status(db, c.id) == "processed_by_cs"


def test_Z4_pending_do_not_locked(db: Session) -> None:
    """UAC-Z4 / freeze signal: a not-yet-delivered linked DO is NOT locked."""
    c = _mk_complaint(db, f"{C_PREFIX}Z4B")
    o = _mk_order(db, f"{O_PREFIX}Z4B", remarks_cs=f"{C_PREFIX}Z4B")
    _reconcile(db, (o, None))
    assert ComplaintFulfilmentService(db).is_remarks_cs_locked(o) is False


def test_locked_when_delivered_and_linked(db: Session) -> None:
    """Freeze signal: a delivered + linked DO reports remarks_cs_locked True."""
    c = _mk_complaint(db, f"{C_PREFIX}LOCK")
    o = _mk_order(db, f"{O_PREFIX}LOCK", remarks_cs=f"{C_PREFIX}LOCK", delivered=True)
    _reconcile(db, (o, None))
    assert ComplaintFulfilmentService(db).is_remarks_cs_locked(o) is True


def _status_id(db: Session, code: str) -> str:
    row = db.execute(
        text("SELECT id FROM order_statuses WHERE lower(status_code) = :c LIMIT 1"),
        {"c": code.lower()},
    ).first()
    assert row is not None, f"seed order_statuses must contain '{code}'"
    return str(row[0])


def test_AND_date_without_delivered_status_not_fulfilled(db: Session) -> None:
    """Delivered = date AND delivered/completed status. A date under a non-delivered
    status (e.g. NEW) does NOT fulfil; flipping the status to DELIVERED then fulfils."""
    c = _mk_complaint(db, f"{C_PREFIX}AND")
    o = Order(
        id=str(uuid.uuid4()),
        order_number=f"{O_PREFIX}AND",
        remarks_cs=f"{C_PREFIX}AND",
        actual_delivery_date=datetime(2026, 6, 1, 10, 0),  # date set...
        order_status_id=_status_id(db, "NEW"),  # ...but status is NOT delivered
        is_cancelled=False,
    )
    db.add(o)
    db.commit()
    db.refresh(o)
    _reconcile(db, (o, None))
    assert _status(db, c.id) == "processed_by_cs"  # date alone is not enough

    # Now flip the status to DELIVERED (date already set) -> delivered -> fulfilled.
    o.order_status_id = _status_id(db, "DELIVERED")
    db.commit()
    _reconcile(db, (o, f"{C_PREFIX}AND"))
    assert _status(db, c.id) == "fulfilled"


def test_unfreeze_via_status(db: Session) -> None:
    """Unfreeze-via-status: a delivered+linked DO is locked, but evaluating the lock
    against an incoming non-delivered status returns False (so the same PUT that moves
    the DO off DELIVERED may also edit Remarks CS)."""
    c = _mk_complaint(db, f"{C_PREFIX}UNFRZ")
    o = _mk_order(db, f"{O_PREFIX}UNFRZ", remarks_cs=f"{C_PREFIX}UNFRZ", delivered=True)
    _reconcile(db, (o, None))
    svc = ComplaintFulfilmentService(db)
    assert svc.is_remarks_cs_locked(o) is True
    # Hypothetical: status moved to NEW (date unchanged) -> no longer delivered -> unlocked.
    assert (
        svc.is_locked_for_state(o, _status_id(db, "NEW"), o.actual_delivery_date)
        is False
    )
    # Still locked if it stays DELIVERED.
    assert (
        svc.is_locked_for_state(o, _status_id(db, "DELIVERED"), o.actual_delivery_date)
        is True
    )


def test_Z2_put_rejects_locked_remarks_cs(db: Session) -> None:
    """UAC-Z2: PUT /orders/{id} changing a frozen Remarks CS -> 422 REMARKS_CS_LOCKED."""
    from app.schemas.order import OrderUpdate
    from app.services.order_service import OrderService
    from app.services.error_handler import AppException

    c = _mk_complaint(db, f"{C_PREFIX}Z2")
    o = _mk_order(db, f"{O_PREFIX}Z2", remarks_cs=f"{C_PREFIX}Z2", delivered=True)
    _reconcile(db, (o, None))  # delivered + linked => frozen

    svc = OrderService(db)
    with pytest.raises(AppException) as exc:
        svc.update_order(
            str(o.id), OrderUpdate(remarks_cs="CDF-CHANGED"), updated_by="40e836b9-b307-5b9d-b7b4-1dea8f2110b8"
        )
    assert exc.value.status_code == 422
    detail = exc.value.detail
    code = detail.get("code") if isinstance(detail, dict) else None
    assert code == "REMARKS_CS_LOCKED", detail


def test_Z2_put_allows_other_fields_when_locked(db: Session) -> None:
    """A frozen DO still accepts non-Remarks-CS edits (only Remarks CS is locked)."""
    from app.schemas.order import OrderUpdate
    from app.services.order_service import OrderService

    c = _mk_complaint(db, f"{C_PREFIX}Z2B")
    o = _mk_order(db, f"{O_PREFIX}Z2B", remarks_cs=f"{C_PREFIX}Z2B", delivered=True)
    _reconcile(db, (o, None))

    svc = OrderService(db)
    # Same Remarks CS value (no change) + an unrelated field -> not rejected.
    updated = svc.update_order(
        str(o.id),
        OrderUpdate(remarks_cs=f"{C_PREFIX}Z2B", delivery_remarks_cs="ok"),
        updated_by=str(uuid.uuid4()),
    )
    assert updated.delivery_remarks_cs == "ok"


# --------------------------------------------------------------------------- #
# Freeze on the REAL import path (UAC-Z1)
# --------------------------------------------------------------------------- #
def _build_tracking_workbook(master_rows: list[list]) -> bytes:
    """Minimal valid Master + Overall Tracking workbook for import_excel_tracking."""
    from io import BytesIO
    from openpyxl import Workbook

    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)
    ms = wb.create_sheet("Master")
    ms.append(["Doc. No.", "Remarks CS"])
    for r in master_rows:
        ms.append(r)
    ts = wb.create_sheet("Overall Tracking")
    ts.append(["Doc Number"])  # header only — delivery already set in DB
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_Z1_import_freeze_warns_and_skips_remarks_cs(db: Session) -> None:
    """UAC-Z1: a REAL import (validate_only=False) carrying a changed Remarks CS for a
    delivered + linked DO keeps the DB value, skips that field, and surfaces a warning
    (the same warning that lands in the ImportLog)."""
    from app.services.order_service import OrderService

    c = _mk_complaint(db, f"{C_PREFIX}Z1")
    o = _mk_order(db, f"{O_PREFIX}Z1", remarks_cs=f"{C_PREFIX}Z1", delivered=True)
    _reconcile(db, (o, None))  # delivered + linked => Remarks CS frozen
    assert len(_links(db, c.id)) == 1

    data = _build_tracking_workbook([[f"{O_PREFIX}Z1", "CDF-CHANGED-IN-IMPORT"]])
    result = OrderService(db).import_excel_tracking(
        data, str(uuid.uuid4()), validate_only=False
    )

    warns = [
        w.get("warning", "") if isinstance(w, dict) else str(w)
        for w in result.get("warnings", [])
    ]
    assert any("change ignored" in w for w in warns), warns
    assert any(c.complaint_number in w for w in warns), warns

    # DB Remarks CS kept (field skipped), not overwritten by the import value.
    db.expire_all()
    fresh = db.query(Order).filter(Order.id == o.id).first()
    assert fresh is not None and fresh.remarks_cs == f"{C_PREFIX}Z1"


def test_Z1_import_freeze_warning_in_validate_only_response(db: Session) -> None:
    """UAC-Z1 (test-import half): the same freeze warning surfaces in the dry-run /
    validate-only import response."""
    from app.services.order_service import OrderService

    c = _mk_complaint(db, f"{C_PREFIX}Z1V")
    o = _mk_order(db, f"{O_PREFIX}Z1V", remarks_cs=f"{C_PREFIX}Z1V", delivered=True)
    _reconcile(db, (o, None))

    data = _build_tracking_workbook([[f"{O_PREFIX}Z1V", "CDF-CHANGED-DRYRUN"]])
    result = OrderService(db).import_excel_tracking(
        data, str(uuid.uuid4()), validate_only=True
    )
    assert any("change ignored" in w for w in result.get("warnings", [])), result


def test_Z1_import_not_frozen_when_not_delivered(db: Session) -> None:
    """UAC-Z1 negative / UAC-Z4: a linked but NOT-delivered DO is not frozen — an
    import may still change its Remarks CS (no warning, value updated)."""
    from app.services.order_service import OrderService

    c = _mk_complaint(db, f"{C_PREFIX}Z1N")
    o = _mk_order(db, f"{O_PREFIX}Z1N", remarks_cs=f"{C_PREFIX}Z1N", delivered=False)
    _reconcile(db, (o, None))

    data = _build_tracking_workbook([[f"{O_PREFIX}Z1N", f"{C_PREFIX}Z1N CDF-EXTRA"]])
    result = OrderService(db).import_excel_tracking(
        data, str(uuid.uuid4()), validate_only=False
    )
    warns = [
        w.get("warning", "") if isinstance(w, dict) else str(w)
        for w in result.get("warnings", [])
    ]
    assert not any("change ignored" in w for w in warns), warns
    db.expire_all()
    fresh = db.query(Order).filter(Order.id == o.id).first()
    assert fresh is not None and fresh.remarks_cs == f"{C_PREFIX}Z1N CDF-EXTRA"


# --------------------------------------------------------------------------- #
# Backfill — stamp-without-send + idempotent (UAC-B1 / UAC-B2)
# --------------------------------------------------------------------------- #
def test_B1_backfill_stamps_without_send(db: Session) -> None:
    """UAC-B1: the backfill drives the centralized helper with old_remarks=None,
    creates links, sets `fulfilled` for already-delivered DOs, and STAMPS
    delivery_notified_at — but DISCARDS the notify payloads so no historical message
    is sent. Mirrors scripts/backfill_complaint_fulfilment_orders.py."""
    c = _mk_complaint(db, f"{C_PREFIX}B1")
    o = _mk_order(db, f"{O_PREFIX}B1", remarks_cs=f"{C_PREFIX}B1", delivered=True)
    svc = ComplaintFulfilmentService(db)

    # Exactly what the backfill script does: dry_run=False, then DROP the payloads.
    warnings, notify_payloads = svc.reconcile_links_and_status(
        [{"order": o, "old_remarks": None}], dry_run=False
    )
    db.commit()
    # The script suppresses these — they are never dispatched.
    suppressed = len(notify_payloads)
    assert suppressed == 1

    links = _links(db, c.id)
    assert len(links) == 1
    assert links[0].delivery_notified_at is not None  # stamped, despite no send
    assert _status(db, c.id) == "fulfilled"


def test_B2_backfill_idempotent_rerun(db: Session) -> None:
    """UAC-B2: a second backfill pass creates no duplicate links, re-stamps nothing,
    and produces zero new notify payloads (so a re-run never spams)."""
    c = _mk_complaint(db, f"{C_PREFIX}B2")
    o = _mk_order(db, f"{O_PREFIX}B2", remarks_cs=f"{C_PREFIX}B2", delivered=True)
    svc = ComplaintFulfilmentService(db)

    svc.reconcile_links_and_status([{"order": o, "old_remarks": None}], dry_run=False)
    db.commit()
    first_links = _links(db, c.id)
    assert len(first_links) == 1
    first_stamp = first_links[0].delivery_notified_at
    assert first_stamp is not None

    # Re-run the backfill batch identically.
    _, notify2 = svc.reconcile_links_and_status(
        [{"order": o, "old_remarks": None}], dry_run=False
    )
    db.commit()
    second_links = _links(db, c.id)
    assert len(second_links) == 1  # no duplicate link row
    assert second_links[0].delivery_notified_at == first_stamp  # not re-stamped
    assert notify2 == []  # no re-notify
    assert _status(db, c.id) == "fulfilled"


def test_backfill_script_dry_run_executes(monkeypatch) -> None:
    """UAC-B1/B2 (script entrypoint): the backfill CLI runs in --dry-run mode,
    returns 0, and writes nothing (rolled back)."""
    import importlib
    import sys

    backfill = importlib.import_module("scripts.backfill_complaint_fulfilment_orders")
    monkeypatch.setattr(sys, "argv", ["backfill", "--dry-run"])
    assert backfill.main() == 0


# --------------------------------------------------------------------------- #
# Endpoint auth (UAC — the two new GET routes reject an unauthenticated caller)
# --------------------------------------------------------------------------- #
def test_complaint_fulfilment_orders_endpoint_requires_auth() -> None:
    from fastapi.testclient import TestClient
    from app.main import app

    with TestClient(app) as client:
        res = client.get(
            f"/api/v1/complaints-management/complaints/{uuid.uuid4()}/fulfilment-orders"
        )
    assert res.status_code in (401, 403), res.text


def test_order_fulfilled_complaints_endpoint_requires_auth() -> None:
    from fastapi.testclient import TestClient
    from app.main import app

    with TestClient(app) as client:
        res = client.get(
            f"/api/v1/order-management/orders/{uuid.uuid4()}/fulfilled-complaints"
        )
    assert res.status_code in (401, 403), res.text


# --------------------------------------------------------------------------- #
# Notify-once idempotency (UAC-N4) + dry-run (UAC-N7) at the reconcile boundary
# --------------------------------------------------------------------------- #
def test_N4_notify_once_via_delivery_notified_at(db: Session) -> None:
    """UAC-N4: a delivered DO yields a notify payload once; a re-run with the same
    delivered state yields none (delivery_notified_at gate)."""
    c = _mk_complaint(db, f"{C_PREFIX}N4")
    o = _mk_order(db, f"{O_PREFIX}N4", remarks_cs=f"{C_PREFIX}N4", delivered=True)
    _, notify1 = _reconcile(db, (o, None))
    assert len(notify1) == 1
    links = _links(db, c.id)
    assert links[0].delivery_notified_at is not None
    # Re-run: delivery already notified -> no payload.
    _, notify2 = _reconcile(db, (o, f"{C_PREFIX}N4"))
    assert notify2 == []


def test_N7_dry_run_no_links_no_payloads(db: Session) -> None:
    """UAC-N7: dry-run (validate-only) creates no links and emits no notify payloads."""
    c = _mk_complaint(db, f"{C_PREFIX}N7")
    o = _mk_order(db, f"{O_PREFIX}N7", remarks_cs=f"{C_PREFIX}N7", delivered=True)
    svc = ComplaintFulfilmentService(db)
    warnings, notify = svc.reconcile_links_and_status(
        [{"order": o, "old_remarks": None}], dry_run=True
    )
    db.rollback()
    assert notify == []
    assert _links(db, c.id) == []
    assert _status(db, c.id) == "processed_by_cs"


# --------------------------------------------------------------------------- #
# Performance — delta-only / batched (UAC-P1)
# --------------------------------------------------------------------------- #
def _count_complaint_selects(db: Session, fn) -> int:
    """Count SELECTs against the complaints table issued while running ``fn``."""
    counter = {"n": 0}

    def _before(conn, cursor, statement, params, context, executemany):
        s = " ".join(statement.lower().split())
        if s.startswith("select") and "from complaints " in s:
            counter["n"] += 1

    event.listen(engine, "before_cursor_execute", _before)
    try:
        fn()
    finally:
        event.remove(engine, "before_cursor_execute", _before)
    return counter["n"]


def test_P1_no_change_import_issues_no_complaint_query(db: Session) -> None:
    """UAC-P1: a delta-only import with no changed rows runs zero complaint queries."""
    svc = ComplaintFulfilmentService(db)
    n = _count_complaint_selects(
        db, lambda: svc.reconcile_links_and_status([], dry_run=False)
    )
    assert n == 0


def test_P1_complaint_lookup_is_batched_not_per_row(db: Session) -> None:
    """UAC-P1: the complaint lookup is one batched IN-query regardless of the number
    of changed orders (no per-row complaint query)."""
    c1 = _mk_complaint(db, f"{C_PREFIX}P1A")
    c2 = _mk_complaint(db, f"{C_PREFIX}P1B")
    c3 = _mk_complaint(db, f"{C_PREFIX}P1C")
    o1 = _mk_order(db, f"{O_PREFIX}P1a", remarks_cs=f"{C_PREFIX}P1A")
    o2 = _mk_order(db, f"{O_PREFIX}P1b", remarks_cs=f"{C_PREFIX}P1B")
    o3 = _mk_order(db, f"{O_PREFIX}P1c", remarks_cs=f"{C_PREFIX}P1C")
    svc = ComplaintFulfilmentService(db)

    def _run():
        svc.reconcile_links_and_status(
            [
                {"order": o1, "old_remarks": None},
                {"order": o2, "old_remarks": None},
                {"order": o3, "old_remarks": None},
            ],
            dry_run=True,
        )

    # Only the token->complaint lookup hits the complaints table on dry-run; it is a
    # single batched IN query, NOT one per order (would be 3 if per-row).
    n = _count_complaint_selects(db, _run)
    db.rollback()
    assert n == 1, f"expected one batched complaint lookup, got {n}"
