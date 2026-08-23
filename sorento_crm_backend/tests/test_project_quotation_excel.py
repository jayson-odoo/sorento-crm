"""S8 of the quotation DOCUMENT layer: exporting an ISSUE to Excel, one sheet per scope.

Written BEFORE the implementation. The Excel export exists for one reader: the customer's QS,
who otherwise re-types the whole quotation to re-price it (AC-F2). Every rule asserted here is a
rule that reader depends on, and each one fails silently rather than loudly:

- **One sheet per scope, in the document's own order.** A QS re-pricing the guard house should
  not have to scroll past the townhouse, which is exactly why this is deliberately NOT the
  sample's single banded sheet.
- **Every sheet states its own total, and the first sheet states the grand total.** A printed
  sheet with no total is a price list nobody can check, and a workbook with no grand total makes
  the reader add the tabs up by hand - which is the moment the numbers stop matching ours.
- **A rate-only line prints the words `rate only` and is in no total** (AC-C2 / AC-D3). Writing
  0.00 in that cell tells a QS the item is free, and a QS re-pricing in Excel will sum the
  column; both errors are invisible on the page.
- **Money is a NUMBER with a format, never a pre-formatted string.** The whole point of the
  Excel artifact is arithmetic. A right-aligned `"1,000.00"` looks identical on screen and
  produces zero in every SUM the reader writes.
- **The PRODUCT IMAGE column collapses per SHEET when no line on that sheet has an image**
  (AC-F4), judged per sheet rather than per workbook: a scope of blank cells is worse than no
  column, and the sheet next to it may legitimately need one.
- **A sheet title is a hostile string.** Scope labels are typed off the customer's own bill of
  quantities, and Excel caps a title at 31 characters, forbids ``[]:*?/\\`` and refuses
  duplicates. Any of those makes the whole workbook unopenable rather than merely ugly.

Postgres only, via ``blank_session``. Every row carries the ``zzt-qxlsx`` marker, because the
dev database this runs against holds a copy of production data.
"""
from __future__ import annotations

import uuid
from decimal import Decimal
from io import BytesIO

from sqlalchemy import text

from app.models.numbering import DocumentNumberingRule
from app.models.product import Product, ProductCategory, UnitOfMeasure
from app.models.projects import ProjectParty
from app.models.resources import Attachment
from app.models.user import User
from app.services import project_seed_service

from ._pg_fixture import blank_session

MARKER = "zzt-qxlsx"

# Off the real workbook shape: one priced line and one rate-only alternate, at money values
# that cannot be confused for each other.
PRICED_RATE = "250.00"
PRICED_QTY = 4
# Lines in the "many photographs" size test. Three distinct photographs prove the per-line ratio;
# 52 proved the same thing at 160 s a run.
LINE_COUNT = 3
PRICED_TOTAL = Decimal("1000.00")
RATE_ONLY_RATE = "180.00"
RATE_ONLY_QTY = 7
RATE_ONLY_TOTAL = Decimal("1260.00")  # written nowhere: the cell says "rate only" instead

SIGNATURE_DATA_URI = "data:image/png;base64,zzt"

# Excel's own rules, restated here rather than imported from the implementation: a test that
# read the production constant would still pass if that constant were wrong.
FORBIDDEN_TITLE_CHARS = set("[]:*?/\\")
MAX_TITLE_LEN = 31


def _uid() -> str:
    return str(uuid.uuid4())


def _sorento(db) -> str:
    """Sorento's company id as a STRING, the shape the request path carries."""
    return str(db.execute(text("select id from companies where code = 'SRT'")).scalar())


def _user(db, name: str) -> str:
    user_id = _uid()
    db.add(User(id=user_id, email=f"{user_id}@zzt.test", name=name))
    db.flush()
    return user_id


def _uom(db, code="NO") -> str:
    row = UnitOfMeasure(id=_uid(), uom_code=f"ZZT{_uid()[:4]}", uom_name=code)
    db.add(row)
    db.flush()
    return row.id


def _category(db, name: str) -> ProductCategory:
    row = ProductCategory(
        id=_uid(), category_code=f"ZZT-{_uid()[:8]}", category_name=f"{MARKER} {name}"
    )
    db.add(row)
    db.flush()
    return row


def _product(db, category_id: str, uom_id: str, *, description: str) -> Product:
    row = Product(
        id=_uid(),
        product_code=f"ZZT-{_uid()[:8]}",
        product_name=f"{MARKER} WC Suite",
        description=description,
        category_id=category_id,
        base_uom_id=uom_id,
        list_price=Decimal("300.00"),
    )
    db.add(row)
    db.flush()
    return row


def _party(db, company_id: str, *, address: str, phone: str) -> ProjectParty:
    row = ProjectParty(
        id=_uid(),
        company_id=company_id,
        party_type="developer",
        name=f"{MARKER} Nadi Cergas {_uid()[:6]}",
        address=address,
        phone=phone,
    )
    db.add(row)
    db.flush()
    return row


def _numbering_rule(db, company_id: str) -> DocumentNumberingRule:
    """Seeded, not borrowed: CI's database is empty, so a test that assumed an existing
    `project_quotation` rule would pass only on a developer's machine."""
    scoped = hasattr(DocumentNumberingRule, "company_id")
    query = db.query(DocumentNumberingRule).filter(
        DocumentNumberingRule.doc_type == "project_quotation"
    )
    if scoped:
        query = query.filter(DocumentNumberingRule.company_id == company_id)
    rule = query.first()
    if rule is None:
        rule = DocumentNumberingRule(id=_uid(), doc_type="project_quotation")
        if scoped:
            rule.company_id = company_id
        db.add(rule)
    rule.enabled = True
    rule.prefix_template = f"{MARKER}/Q/"
    rule.number_digits = 4
    rule.next_value = 141
    rule.start_value = 141
    rule.reset_policy = "none"
    rule.last_reset_key = None
    db.flush()
    return rule


def _attachment(db, company_id: str, name: str = "wc.png") -> Attachment:
    row = Attachment(
        id=_uid(),
        company_id=company_id,
        original_filename=f"{MARKER}-{name}",
        stored_filename=f"{MARKER}-{name.replace('.', '-renamed.')}",
        file_path=f"https://cdn.zzt.test/products/{MARKER}/{name}",
        mime_type="image/png",
    )
    db.add(row)
    db.flush()
    return row


def _photo_bytes(width: int = 1200, height: int = 900, seed: int = 0) -> bytes:
    """A real photograph-shaped JPEG, so the downscale is genuinely exercised.

    Noise rather than a flat colour: a flat image compresses to almost nothing, which would make
    every size assertion in this file pass for the wrong reason. ``seed`` makes two photographs
    genuinely different, so a size measured over 52 of them is not measuring one of them.
    """
    from PIL import Image

    image = Image.frombytes(
        "RGB",
        (width, height),
        bytes(
            (x * 7 + y * 13 + c * 61 + seed * 97) % 256
            for y in range(height)
            for x in range(width)
            for c in range(3)
        ),
    )
    buffer = BytesIO()
    image.save(buffer, format="JPEG", quality=95)
    return buffer.getvalue()


class _FakeBackend:
    """Stands in for S3/R2 so the picture path is exercised without network or credentials."""

    def __init__(self, payload: bytes):
        self.payload = payload
        self.keys: list = []

    def download_file(self, key):
        self.keys.append(key)
        return self.payload


def _save(workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _issue(db, document, owner):
    """Sign, then issue. An unsigned document cannot be issued (AC-H1), so every export in this
    file goes through the same two acts a salesperson performs."""
    from app.services import project_quotation_document_service as qdocs

    qdocs.sign_as_sorento(
        db,
        document=document,
        actor_user_id=owner,
        payload={
            "mode": "draw",
            "signer_name": document.signatory_name or f"{MARKER} Baser Ramli",
            "image_data_uri": SIGNATURE_DATA_URI,
        },
    )
    return qdocs.issue(db, document=document, actor_user_id=owner)


def _setup(db, *, developer_address=None, developer_phone=None):
    """The common chain: company, seed, numbering rule, owner, catalogue, project."""
    from app.services.project_service import register_project

    company_id = _sorento(db)
    project_seed_service.run(db, company_id=company_id)
    _numbering_rule(db, company_id)
    owner = _user(db, f"{MARKER} Baser Ramli")
    uom = _uom(db)
    category = _category(db, "Sanitary Ware")
    party = None
    if developer_address is not None:
        party = _party(db, company_id, address=developer_address, phone=developer_phone or "")
    project = register_project(
        db,
        company_id=company_id,
        actor_user_id=owner,
        developer_party_id=party.id if party is not None else None,
        title=f"{MARKER} Cadangan Membina Pangsapuri {_uid()[:12]}",
    )
    return {
        "company_id": company_id,
        "owner": owner,
        "uom": uom,
        "category": category,
        "party": party,
        "project": project,
    }


# --------------------------------------------------------------- reading a workbook


def _values(sheet) -> list:
    """Every non-empty cell value on a sheet, flattened. For presence assertions only."""
    return [
        cell.value
        for row in sheet.iter_rows()
        for cell in row
        if cell.value not in (None, "")
    ]


def _row_of(sheet, label):
    """The 1-based row index whose FIRST matching cell equals ``label`` exactly.

    Exact equality on purpose: `TOTAL` must not match `TOTAL AMOUNT`, or a test asserting the
    scope total would happily read the grand total instead and pass on a workbook that never
    printed a scope total at all.
    """
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == label:
                return cell.row
    return None


def _header_row(sheet):
    """The row index of the column-header row, found by its first column."""
    return _row_of(sheet, "ITEM")


def _headers(sheet) -> list:
    row_index = _header_row(sheet)
    assert row_index is not None, f"no header row on {sheet.title}"
    return [
        cell.value
        for cell in next(sheet.iter_rows(min_row=row_index, max_row=row_index))
        if cell.value not in (None, "")
    ]


def _column_index(sheet, header: str) -> int:
    row_index = _header_row(sheet)
    for cell in next(sheet.iter_rows(min_row=row_index, max_row=row_index)):
        if cell.value == header:
            return cell.column
    raise AssertionError(f"column {header!r} not on sheet {sheet.title!r}")


def _amount_cell(sheet, row_index: int):
    """The money cell on a given row, read out of the GRAND TOTAL column."""
    return sheet.cell(row=row_index, column=_column_index(sheet, "GRAND TOTAL"))


def _labelled_amount(sheet, label: str):
    """The value in the money column of the row carrying ``label``."""
    row_index = _row_of(sheet, label)
    assert row_index is not None, f"{label!r} is not on sheet {sheet.title!r}"
    return _amount_cell(sheet, row_index).value


def _data_rows(sheet) -> list:
    """The line rows: everything below the header row and above the TOTAL row."""
    start = _header_row(sheet) + 1
    stop = _row_of(sheet, "TOTAL")
    assert stop is not None, f"no TOTAL row on sheet {sheet.title!r}"
    return list(range(start, stop))


def _money(value) -> Decimal:
    """Compare money exactly, however openpyxl chose to hand the number back."""
    return Decimal(str(value))


# ----------------------------------------------------------------------- AC-F2


def test_the_workbook_carries_one_sheet_per_scope_in_the_documents_own_order():
    """The client's word for a scope is a tab, and this is the whole reason the Excel export is
    not the sample's single banded sheet. Order matters as much as count: the workbook has to
    read in the same sequence as the paper the customer is holding, or the two cannot be
    checked against each other."""
    from app.services import project_quotation_document_service as qdocs
    from app.services import project_quotation_excel_service as qxlsx
    from app.services import project_quotation_service as quotes

    with blank_session() as db:
        env = _setup(db)
        owner = env["owner"]
        product = _product(
            db, env["category"].id, env["uom"], description=f"{MARKER} close-coupled WC, white"
        )

        document = qdocs.create_document(db, project=env["project"], actor_user_id=owner)
        for label, price, qty in (
            (f"{MARKER} Townhouse", "250.00", 4),
            (f"{MARKER} Guard House", "400.00", 3),
            (f"{MARKER} Surau", "150.00", 2),
        ):
            scope = qdocs.add_scope(
                db, document=document, scope_label=label, actor_user_id=owner
            )
            quotes.upsert_line(
                db,
                version=quotes.current_version(db, scope.id),
                actor_user_id=owner,
                payload={"product_id": product.id, "unit_price": price, "quantity": qty},
            )

        issued = _issue(db, document, owner)
        workbook = qxlsx.build_issue_workbook(db, issued)

        assert workbook.sheetnames == [
            f"{MARKER} Townhouse",
            f"{MARKER} Guard House",
            f"{MARKER} Surau",
        ]


def test_the_first_sheet_carries_the_letterhead_and_the_rest_carry_the_reference():
    """A workbook is printed a sheet at a time. Sheet 1 has to identify the paper the way the
    PDF's first page does (sender, Our Ref including the revision, date, To, Attn, subject), and
    every later sheet has to carry enough - its scope and that same reference - that a single
    printed page found on a desk can still be matched back to the quotation it belongs to."""
    from app.services import project_quotation_document_service as qdocs
    from app.services import project_quotation_excel_service as qxlsx
    from app.services import project_quotation_service as quotes

    with blank_session() as db:
        env = _setup(
            db,
            developer_address=f"{MARKER} Level 8, Menara Lama, Kuala Lumpur",
            developer_phone="03-1111 1111",
        )
        owner = env["owner"]
        product = _product(
            db, env["category"].id, env["uom"], description=f"{MARKER} close-coupled WC, white"
        )

        document = qdocs.create_document(db, project=env["project"], actor_user_id=owner)
        document.your_ref = f"{MARKER}/NC/2026/007"
        document.attn_name = f"{MARKER} Kelly"
        document.subject_title = f"{MARKER} cadangan membina pangsapuri"
        db.flush()

        for label in (f"{MARKER} Townhouse", f"{MARKER} Guard House"):
            scope = qdocs.add_scope(
                db, document=document, scope_label=label, actor_user_id=owner
            )
            quotes.upsert_line(
                db,
                version=quotes.current_version(db, scope.id),
                actor_user_id=owner,
                payload={"product_id": product.id, "unit_price": PRICED_RATE, "quantity": 2},
            )

        issued = _issue(db, document, owner)
        workbook = qxlsx.build_issue_workbook(db, issued)

        first = _values(workbook.worksheets[0])
        assert issued.our_ref_text in first
        assert "R1" in str(issued.our_ref_text)  # the revision is part of the reference
        assert "Our Ref" in first
        assert f"{MARKER}/NC/2026/007" in first
        assert f"{MARKER} Level 8, Menara Lama, Kuala Lumpur" in first
        assert any(f"{MARKER} Kelly" in str(value) for value in first)
        assert any(f"{MARKER} cadangan membina pangsapuri".upper() in str(value) for value in first)
        assert env["party"].name in first

        # The later sheet is light on purpose - the letterhead is not repeated - but it is still
        # identifiable on its own.
        second = _values(workbook.worksheets[1])
        assert f"{MARKER} Guard House" in second
        assert any(str(issued.our_ref_text) in str(value) for value in second)
        # And it does NOT re-print the whole letterhead.
        assert f"{MARKER} Level 8, Menara Lama, Kuala Lumpur" not in second


def test_every_sheet_states_its_own_total_and_the_first_sheet_states_the_grand_total():
    """A sheet handed to somebody on its own has to answer "what does this scope cost", and the
    workbook as a whole has to answer "what does the quotation cost" without the reader adding
    the tabs up by hand. Both numbers come off the issue snapshot rather than being re-summed
    here, so the workbook cannot disagree with the record or with the PDF."""
    from app.services import project_quotation_document_service as qdocs
    from app.services import project_quotation_excel_service as qxlsx
    from app.services import project_quotation_service as quotes

    with blank_session() as db:
        env = _setup(db)
        owner = env["owner"]
        product = _product(
            db, env["category"].id, env["uom"], description=f"{MARKER} close-coupled WC, white"
        )

        document = qdocs.create_document(db, project=env["project"], actor_user_id=owner)
        for label, price, qty in (
            (f"{MARKER} Townhouse", "250.00", 4),
            (f"{MARKER} Guard House", "400.00", 3),
        ):
            scope = qdocs.add_scope(
                db, document=document, scope_label=label, actor_user_id=owner
            )
            quotes.upsert_line(
                db,
                version=quotes.current_version(db, scope.id),
                actor_user_id=owner,
                payload={"product_id": product.id, "unit_price": price, "quantity": qty},
            )

        issued = _issue(db, document, owner)
        workbook = qxlsx.build_issue_workbook(db, issued)

        assert issued.grand_total == Decimal("2200.00")
        assert _money(_labelled_amount(workbook.worksheets[0], "TOTAL")) == Decimal("1000.00")
        assert _money(_labelled_amount(workbook.worksheets[1], "TOTAL")) == Decimal("1200.00")
        assert _money(_labelled_amount(workbook.worksheets[0], "TOTAL AMOUNT")) == Decimal(
            "2200.00"
        )
        # Stated once, on the first sheet. Repeated on every tab it reads as a per-sheet figure.
        assert _row_of(workbook.worksheets[1], "TOTAL AMOUNT") is None


# ------------------------------------------------------------- AC-C2 / AC-D3


def test_a_rate_only_line_reads_rate_only_and_is_absent_from_every_total():
    """The sample carries five rate-only alternates. In Excel the failure mode is worse than on
    paper: a QS selects the money column and sums it, so a 0.00 standing in for the alternate is
    read as free AND its rate is invited into any total the reader builds. The words, the
    absence of a number in that cell, and the arithmetic are asserted together because each one
    alone still lets money be wrong."""
    from app.services import project_quotation_document_service as qdocs
    from app.services import project_quotation_excel_service as qxlsx
    from app.services import project_quotation_service as quotes

    with blank_session() as db:
        env = _setup(db)
        owner = env["owner"]
        product = _product(
            db, env["category"].id, env["uom"], description=f"{MARKER} close-coupled WC, white"
        )

        document = qdocs.create_document(db, project=env["project"], actor_user_id=owner)
        scope = qdocs.add_scope(
            db, document=document, scope_label=f"{MARKER} Townhouse", actor_user_id=owner
        )
        version = quotes.current_version(db, scope.id)
        quotes.upsert_line(
            db,
            version=version,
            actor_user_id=owner,
            payload={
                "product_id": product.id,
                "unit_price": PRICED_RATE,
                "quantity": PRICED_QTY,
                "sort_order": 0,
                "description_snapshot": f"{MARKER} WC suite",
            },
        )
        quotes.upsert_line(
            db,
            version=version,
            actor_user_id=owner,
            payload={
                "product_id": product.id,
                "unit_price": RATE_ONLY_RATE,
                "quantity": RATE_ONLY_QTY,
                "sort_order": 1,
                "is_rate_only": True,
                "description_snapshot": f"{MARKER} OKU grab bar alternate",
            },
        )

        issued = _issue(db, document, owner)
        sheet = qxlsx.build_issue_workbook(db, issued).worksheets[0]

        # The alternate is on the sheet and its rate is quoted, so the customer can price it.
        values = _values(sheet)
        assert f"{MARKER} OKU grab bar alternate" in values
        assert Decimal("180.00") in [
            _money(v) for v in values if isinstance(v, (int, float, Decimal))
        ]

        amounts = {row: _amount_cell(sheet, row).value for row in _data_rows(sheet)}
        assert "rate only" in [str(value) for value in amounts.values()]
        # Nothing numeric in that cell at all: not the extended rate, and not a zero.
        rate_only_cells = [v for v in amounts.values() if str(v) == "rate only"]
        assert len(rate_only_cells) == 1
        numeric_amounts = [
            _money(v) for v in amounts.values() if isinstance(v, (int, float, Decimal))
        ]
        assert numeric_amounts == [PRICED_TOTAL]
        assert RATE_ONLY_TOTAL not in numeric_amounts
        assert Decimal("0.00") not in numeric_amounts

        # Both totals are the priced line alone.
        assert issued.grand_total == PRICED_TOTAL
        assert _money(_labelled_amount(sheet, "TOTAL")) == PRICED_TOTAL
        assert _money(_labelled_amount(sheet, "TOTAL AMOUNT")) == PRICED_TOTAL
        assert _money(_labelled_amount(sheet, "TOTAL")) != PRICED_TOTAL + RATE_ONLY_TOTAL


# ----------------------------------------------------- money is arithmetic, not text


def test_money_lands_in_a_numeric_cell_with_a_format_not_as_a_pre_formatted_string():
    """This is why the export exists. A cell holding the string "1,000.00" is indistinguishable
    from a number on screen and contributes nothing to any SUM, so a QS re-pricing the scope
    silently works from a column of zeroes. Asserted on the rate, the line amount and both
    totals, because a renderer can easily get one of the four right by accident."""
    from app.services import project_quotation_document_service as qdocs
    from app.services import project_quotation_excel_service as qxlsx
    from app.services import project_quotation_service as quotes

    with blank_session() as db:
        env = _setup(db)
        owner = env["owner"]
        product = _product(
            db, env["category"].id, env["uom"], description=f"{MARKER} close-coupled WC, white"
        )

        document = qdocs.create_document(db, project=env["project"], actor_user_id=owner)
        scope = qdocs.add_scope(
            db, document=document, scope_label=f"{MARKER} Townhouse", actor_user_id=owner
        )
        quotes.upsert_line(
            db,
            version=quotes.current_version(db, scope.id),
            actor_user_id=owner,
            payload={
                "product_id": product.id,
                "unit_price": PRICED_RATE,
                "quantity": PRICED_QTY,
                "uom": "NO",
            },
        )

        issued = _issue(db, document, owner)
        sheet = qxlsx.build_issue_workbook(db, issued).worksheets[0]

        line_row = _data_rows(sheet)[0]
        checked = [
            _amount_cell(sheet, line_row),
            sheet.cell(row=line_row, column=_column_index(sheet, "UNIT RATE (RM)")),
            _amount_cell(sheet, _row_of(sheet, "TOTAL")),
            _amount_cell(sheet, _row_of(sheet, "TOTAL AMOUNT")),
        ]
        for cell in checked:
            assert not isinstance(cell.value, str), f"{cell.coordinate} is text: {cell.value!r}"
            assert isinstance(cell.value, (int, float, Decimal)), cell.coordinate
            assert "0.00" in cell.number_format, f"{cell.coordinate}: {cell.number_format}"

        assert _money(checked[0].value) == PRICED_TOTAL
        assert _money(checked[1].value) == Decimal("250.00")

        # Quantity too: it is the other half of every re-price the reader does.
        qty = sheet.cell(row=line_row, column=_column_index(sheet, "QTY"))
        assert not isinstance(qty.value, str)
        assert _money(qty.value) == Decimal("4")


# ----------------------------------------------------------------------- AC-C3


def test_a_band_label_heads_its_section_once_as_a_row_of_its_own():
    """The band is the customer's own BQ heading. Repeated down every line it is noise a QS has
    to read past, and as a value in a cell of the item row it cannot be told apart from a
    description. It heads the section once, on its own row, so the sheet lines up against the
    bill of quantities the customer sent us."""
    from app.services import project_quotation_document_service as qdocs
    from app.services import project_quotation_excel_service as qxlsx
    from app.services import project_quotation_service as quotes

    with blank_session() as db:
        env = _setup(db)
        owner = env["owner"]
        product = _product(
            db, env["category"].id, env["uom"], description=f"{MARKER} close-coupled WC, white"
        )

        document = qdocs.create_document(db, project=env["project"], actor_user_id=owner)
        scope = qdocs.add_scope(
            db, document=document, scope_label=f"{MARKER} Townhouse", actor_user_id=owner
        )
        version = quotes.current_version(db, scope.id)

        first_band = f"{MARKER} BILL NO 3 PAGE 15 4"
        second_band = f"{MARKER} OPTIONAL ITEMS FOR OKU TOILET"
        rows = [
            (first_band, f"{MARKER} WC suite"),
            (None, f"{MARKER} angle valve"),
            (second_band, f"{MARKER} grab bar"),
        ]
        for index, (band, description) in enumerate(rows):
            quotes.upsert_line(
                db,
                version=version,
                actor_user_id=owner,
                payload={
                    "product_id": product.id,
                    "unit_price": PRICED_RATE,
                    "quantity": 1,
                    "sort_order": index,
                    "band_label": band,
                    "description_snapshot": description,
                },
            )

        issued = _issue(db, document, owner)
        sheet = qxlsx.build_issue_workbook(db, issued).worksheets[0]
        values = [str(v) for v in _values(sheet)]

        assert values.count(first_band) == 1
        assert values.count(second_band) == 1
        # A row of its own, above the lines it heads.
        assert _row_of(sheet, first_band) < _row_of(sheet, f"{MARKER} WC suite")
        assert _row_of(sheet, f"{MARKER} angle valve") < _row_of(sheet, second_band)
        assert _row_of(sheet, second_band) < _row_of(sheet, f"{MARKER} grab bar")
        # And it is a heading, not a line: no money on the band row.
        assert _amount_cell(sheet, _row_of(sheet, first_band)).value in (None, "")


# ---------------------------------------------------------- hostile sheet titles


def test_a_long_punctuated_or_duplicated_scope_label_still_makes_a_valid_unique_title():
    """Scope labels are free text typed off the customer's BQ, so `Blok A/B [Phase 1]` and two
    scopes with the same name are both ordinary. Excel does not treat them as ordinary: over 31
    characters or carrying any of ``[]:*?/\\`` and the workbook will not open, and a duplicate
    title raises before a single byte is written. The whole export is lost, not one cell, which
    is why this is asserted as invariants rather than on one hand-picked string - and why the
    de-duplication has to be deterministic, so the same issue exported twice is the same file."""
    from app.services import project_quotation_document_service as qdocs
    from app.services import project_quotation_excel_service as qxlsx
    from app.services import project_quotation_service as quotes

    long_label = (
        f"{MARKER}: Blok A/B [Phase 1] * Cadangan Membina Pangsapuri Rumah Idam Selangor"
    )

    with blank_session() as db:
        env = _setup(db)
        owner = env["owner"]
        product = _product(
            db, env["category"].id, env["uom"], description=f"{MARKER} close-coupled WC, white"
        )

        document = qdocs.create_document(db, project=env["project"], actor_user_id=owner)
        # The same hostile label three times: long, punctuated AND duplicated at once, which is
        # what a copied-and-tweaked scope actually looks like.
        for _ in range(3):
            scope = qdocs.add_scope(
                db, document=document, scope_label=long_label, actor_user_id=owner
            )
            quotes.upsert_line(
                db,
                version=quotes.current_version(db, scope.id),
                actor_user_id=owner,
                payload={"product_id": product.id, "unit_price": PRICED_RATE, "quantity": 1},
            )

        issued = _issue(db, document, owner)
        names = qxlsx.build_issue_workbook(db, issued).sheetnames

        assert len(names) == 3
        assert len(set(names)) == 3, names
        for name in names:
            assert name, "an empty sheet title is refused by Excel"
            assert len(name) <= MAX_TITLE_LEN, name
            assert not (set(name) & FORBIDDEN_TITLE_CHARS), name
            assert name == name.strip()
            # Still recognisable as the scope it came from, or the reader cannot find their tab.
            assert name.startswith(f"{MARKER} Blok A")

        # Deterministic: a re-export of the same issue is the same workbook, so the suffix
        # cannot be a random or time-based token.
        assert qxlsx.build_issue_workbook(db, issued).sheetnames == names

        # And the totals still land per sheet, which is what a naive "just truncate" fix breaks
        # by collapsing two scopes onto one tab.
        workbook = qxlsx.build_issue_workbook(db, issued)
        for sheet in workbook.worksheets:
            assert _money(_labelled_amount(sheet, "TOTAL")) == Decimal("250.00")


# ----------------------------------------------------------------------- AC-F4


def test_the_product_image_column_is_judged_per_sheet_not_per_workbook():
    """A column of blank cells squeezes the columns that do carry information into an
    unreadable width, so it collapses when no line needs it. Per SHEET, because a workbook
    decision would put an empty column on the guard house for the sake of the townhouse - and,
    the other way round, would drop the column from a scope that genuinely has pictures the
    moment its neighbour has none."""
    from app.services import project_quotation_document_service as qdocs
    from app.services import project_quotation_excel_service as qxlsx
    from app.services import project_quotation_service as quotes

    with blank_session() as db:
        env = _setup(db)
        owner = env["owner"]
        product = _product(
            db, env["category"].id, env["uom"], description=f"{MARKER} close-coupled WC, white"
        )
        image = _attachment(db, env["company_id"])

        document = qdocs.create_document(db, project=env["project"], actor_user_id=owner)
        with_image = qdocs.add_scope(
            db, document=document, scope_label=f"{MARKER} Townhouse", actor_user_id=owner
        )
        quotes.upsert_line(
            db,
            version=quotes.current_version(db, with_image.id),
            actor_user_id=owner,
            payload={
                "product_id": product.id,
                "unit_price": PRICED_RATE,
                "quantity": PRICED_QTY,
                "image_attachment_id": image.id,
            },
        )
        without = qdocs.add_scope(
            db, document=document, scope_label=f"{MARKER} Guard House", actor_user_id=owner
        )
        quotes.upsert_line(
            db,
            version=quotes.current_version(db, without.id),
            actor_user_id=owner,
            payload={"product_id": product.id, "unit_price": "400.00", "quantity": 3},
        )

        issued = _issue(db, document, owner)
        workbook = qxlsx.build_issue_workbook(db, issued)

        assert "PRODUCT IMAGE" in _headers(workbook.worksheets[0])
        assert "PRODUCT IMAGE" not in _headers(workbook.worksheets[1])

        # The rest of the sample's column set is on both sheets either way.
        for sheet in workbook.worksheets:
            for header in (
                "ITEM",
                "TECHNICAL SPEC",
                "DESCRIPTION",
                "BRAND",
                "PRODUCT CODE",
                "QTY",
                "UNIT RATE (RM)",
                "COMPLETE SET",
                "GRAND TOTAL",
            ):
                assert header in _headers(sheet), f"{header} missing from {sheet.title}"


# ----------------------------------------------------------------------- XLS-1/2


def test_the_image_column_carries_the_picture_itself_anchored_over_an_empty_cell(monkeypatch):
    """S21, reversing this module's earlier "filename, not picture" decision.

    The reason for that decision was real and still is: an openpyxl drawing is not a cell value,
    so it does not travel when the QS sorts or filters the sheet. It is outweighed by the client
    opening their own issued quotation - `Cabana Elmina- nadi cergas R2.xlsx`, 24 photographs,
    every one anchored in column B beside its line - and asking for that. The PRODUCT CODE column
    still identifies a row after a sort, so a drifting picture is untidy rather than ambiguous.

    The cell VALUE stays empty, exactly as in their workbook: a filename underneath would show
    through around the picture."""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    from app.services import product_image_service as images
    from app.services import project_quotation_document_service as qdocs
    from app.services import project_quotation_excel_service as qxlsx
    from app.services import project_quotation_service as quotes

    with blank_session() as db:
        env = _setup(db)
        owner = env["owner"]
        product = _product(
            db, env["category"].id, env["uom"], description=f"{MARKER} close-coupled WC, white"
        )
        image = _attachment(db, env["company_id"])
        monkeypatch.setattr(
            images, "get_backend", lambda provider: _FakeBackend(_photo_bytes(1200, 900))
        )

        document = qdocs.create_document(db, project=env["project"], actor_user_id=owner)
        scope = qdocs.add_scope(
            db, document=document, scope_label=f"{MARKER} Townhouse", actor_user_id=owner
        )
        quotes.upsert_line(
            db,
            version=quotes.current_version(db, scope.id),
            actor_user_id=owner,
            payload={
                "product_id": product.id,
                "unit_price": PRICED_RATE,
                "quantity": PRICED_QTY,
                "image_attachment_id": image.id,
            },
        )

        workbook = qxlsx.build_issue_workbook(db, _issue(db, document, owner))
        sheet = workbook.worksheets[0]
        row = _data_rows(sheet)[0]
        column = _column_index(sheet, "PRODUCT IMAGE")

        assert len(sheet._images) == 1
        # Anchored to the LINE's own cell in the picture column, which is what makes the
        # photograph read as belonging to that item rather than floating over the sheet.
        assert sheet._images[0].anchor == f"{get_column_letter(column)}{row}"
        assert sheet.cell(row=row, column=column).value is None
        # XLS-4: tall enough that the picture is not lying over the line below it.
        assert (sheet.row_dimensions[row].height or 0) >= sheet._images[0].height * 0.75

        # And it survives the save, at the same cell. openpyxl only turns a string anchor into a
        # real drawing anchor when the workbook is written, so the in-memory object proves nothing
        # about the file the customer opens.
        reopened = load_workbook(BytesIO(_save(workbook)))
        placed = reopened.worksheets[0]._images[0].anchor._from
        assert (placed.row + 1, placed.col + 1) == (row, column)


def test_a_line_with_no_picture_leaves_its_cell_empty_rather_than_explaining_itself(monkeypatch):
    """The workbook goes to the customer next to the PDF. Our internal "nobody has chosen this
    product's photo yet" belongs on the screen where somebody can act on it, not in a cell they
    read. The column is here because a NEIGHBOURING line has a picture."""
    from app.services import product_image_service as images
    from app.services import project_quotation_document_service as qdocs
    from app.services import project_quotation_excel_service as qxlsx
    from app.services import project_quotation_service as quotes

    with blank_session() as db:
        env = _setup(db)
        owner = env["owner"]
        monkeypatch.setattr(
            images, "get_backend", lambda provider: _FakeBackend(_photo_bytes(600, 600))
        )
        document = qdocs.create_document(db, project=env["project"], actor_user_id=owner)
        scope = qdocs.add_scope(
            db, document=document, scope_label=f"{MARKER} Townhouse", actor_user_id=owner
        )
        version = quotes.current_version(db, scope.id)
        quotes.upsert_line(
            db,
            version=version,
            actor_user_id=owner,
            payload={
                "product_id": _product(
                    db, env["category"].id, env["uom"], description=f"{MARKER} WC"
                ).id,
                "unit_price": PRICED_RATE,
                "quantity": PRICED_QTY,
                "image_attachment_id": _attachment(db, env["company_id"]).id,
            },
        )
        quotes.upsert_line(
            db,
            version=version,
            actor_user_id=owner,
            payload={
                "product_id": _product(
                    db, env["category"].id, env["uom"], description=f"{MARKER} bottle trap"
                ).id,
                "unit_price": "6.50",
                "quantity": 2,
            },
        )

        workbook = qxlsx.build_issue_workbook(db, _issue(db, document, owner))
        sheet = workbook.worksheets[0]
        column = _column_index(sheet, "PRODUCT IMAGE")
        rows = _data_rows(sheet)

        assert len(sheet._images) == 1
        assert sheet.cell(row=rows[1], column=column).value is None
        assert not any(
            "photo" in str(value).lower() for value in _values(sheet) if value is not None
        )


def test_a_picture_that_cannot_be_fetched_leaves_the_cell_empty_rather_than_failing_the_export(
    monkeypatch,
):
    """XLS-6. The reader is waiting for a price, not a photograph."""
    from app.services import product_image_service as images
    from app.services import project_quotation_document_service as qdocs
    from app.services import project_quotation_excel_service as qxlsx
    from app.services import project_quotation_service as quotes

    with blank_session() as db:
        env = _setup(db)
        owner = env["owner"]

        def _boom(provider):
            raise RuntimeError(f"{MARKER} storage down")

        monkeypatch.setattr(images, "get_backend", _boom)
        document = qdocs.create_document(db, project=env["project"], actor_user_id=owner)
        scope = qdocs.add_scope(
            db, document=document, scope_label=f"{MARKER} Townhouse", actor_user_id=owner
        )
        quotes.upsert_line(
            db,
            version=quotes.current_version(db, scope.id),
            actor_user_id=owner,
            payload={
                "product_id": _product(
                    db, env["category"].id, env["uom"], description=f"{MARKER} WC"
                ).id,
                "unit_price": PRICED_RATE,
                "quantity": PRICED_QTY,
                "image_attachment_id": _attachment(db, env["company_id"]).id,
            },
        )

        workbook = qxlsx.build_issue_workbook(db, _issue(db, document, owner))
        sheet = workbook.worksheets[0]
        # The column still stands (the data says the line has a picture) and the money is intact.
        assert "PRODUCT IMAGE" in _headers(sheet)
        assert not sheet._images
        assert PRICED_TOTAL in _values(sheet)


def test_a_multi_line_workbook_of_photographs_stays_a_file_somebody_can_open(monkeypatch):
    """XLS-5, measured rather than asserted in the abstract. The client's real quotation runs to
    52+ lines and the live catalogue's chosen photographs average 1.1 MB. Size scales linearly
    per line, so a handful of distinct photographs measures the same ratio as 52 did, in seconds
    rather than minutes of CI."""
    from openpyxl import load_workbook

    from app.services import product_image_service as images
    from app.services import project_quotation_document_service as qdocs
    from app.services import project_quotation_excel_service as qxlsx
    from app.services import project_quotation_service as quotes

    with blank_session() as db:
        env = _setup(db)
        owner = env["owner"]
        # A different photograph per line, and different BYTES: 52 copies of one picture is one
        # picture once the zip container has deduplicated it, which would measure nothing.
        original = _photo_bytes(1600, 1600)
        photos: dict = {}

        class _PerKey:
            def download_file(self, key):
                if key not in photos:
                    photos[key] = _photo_bytes(1600, 1600, seed=len(photos) + 1)
                return photos[key]

        monkeypatch.setattr(images, "get_backend", lambda provider: _PerKey())

        document = qdocs.create_document(db, project=env["project"], actor_user_id=owner)
        scope = qdocs.add_scope(
            db, document=document, scope_label=f"{MARKER} Townhouse", actor_user_id=owner
        )
        version = quotes.current_version(db, scope.id)
        for index in range(LINE_COUNT):
            quotes.upsert_line(
                db,
                version=version,
                actor_user_id=owner,
                payload={
                    "product_id": _product(
                        db, env["category"].id, env["uom"], description=f"{MARKER} item {index}"
                    ).id,
                    "unit_price": PRICED_RATE,
                    "quantity": PRICED_QTY,
                    "image_attachment_id": _attachment(
                        db, env["company_id"], f"item-{index}.jpg"
                    ).id,
                },
            )

        xlsx_bytes, _ = qxlsx.render_issue_xlsx(db, _issue(db, document, owner))
        naive = LINE_COUNT * len(original)
        print(
            f"\n{LINE_COUNT}-line workbook: {len(xlsx_bytes) / 1024:.0f} KB "
            f"({LINE_COUNT} originals would be {naive / 1024 / 1024:.1f} MB)"
        )
        assert len(xlsx_bytes) < 4 * 1024 * 1024, f"{len(xlsx_bytes)} bytes"
        assert len(xlsx_bytes) < naive / 10

        reopened = load_workbook(BytesIO(xlsx_bytes))
        assert len(reopened.worksheets[0]._images) == LINE_COUNT


# --------------------------------------------------------------- the actual bytes


def test_the_export_returns_bytes_that_open_as_a_real_workbook_named_after_the_reference():
    """Everything above asserts an in-memory Workbook; this asserts the file. A workbook that
    only ever gets inspected as an object can be unsaveable (an illegal title, a bad number
    format), and the customer is the one who finds out. The filename mirrors the PDF's so a
    saved file is still identifiable by the reference the customer quotes back."""
    from openpyxl import load_workbook

    from app.services import project_quotation_document_service as qdocs
    from app.services import project_quotation_excel_service as qxlsx
    from app.services import project_quotation_pdf_service as qpdf
    from app.services import project_quotation_service as quotes

    with blank_session() as db:
        env = _setup(db)
        owner = env["owner"]
        product = _product(
            db, env["category"].id, env["uom"], description=f"{MARKER} close-coupled WC, white"
        )

        document = qdocs.create_document(db, project=env["project"], actor_user_id=owner)
        scope = qdocs.add_scope(
            db, document=document, scope_label=f"{MARKER} Townhouse", actor_user_id=owner
        )
        quotes.upsert_line(
            db,
            version=quotes.current_version(db, scope.id),
            actor_user_id=owner,
            payload={
                "product_id": product.id,
                "unit_price": PRICED_RATE,
                "quantity": PRICED_QTY,
            },
        )

        issued = _issue(db, document, owner)
        payload, filename = qxlsx.render_issue_xlsx(db, issued)

        assert isinstance(payload, bytes)
        assert payload[:2] == b"PK"  # a zip container, which is what an xlsx is
        assert filename.endswith(".xlsx")
        assert "R1" in filename
        # Same stem as the PDF: the two artifacts of one issue sit together in a folder.
        assert filename[: -len(".xlsx")] == qpdf.build_filename(issued)[: -len(".pdf")]

        reopened = load_workbook(BytesIO(payload))
        assert reopened.sheetnames == [f"{MARKER} Townhouse"]
        reread = reopened.worksheets[0]
        assert _money(_labelled_amount(reread, "TOTAL")) == PRICED_TOTAL
        assert _money(_labelled_amount(reread, "TOTAL AMOUNT")) == PRICED_TOTAL


# --------------------------------------------------------------------- AC-F3


def test_the_workbook_exports_the_issue_snapshot_and_not_the_live_rows():
    """Same rule as the PDF, and the same silent failure: reading the scope's CURRENT version
    would make a re-export next year show a revision priced months later under the reference the
    customer holds. Asserted here too rather than trusted from the sibling, because the Excel
    path could easily grow its own loader."""
    from app.services import project_quotation_document_service as qdocs
    from app.services import project_quotation_excel_service as qxlsx
    from app.services import project_quotation_service as quotes

    with blank_session() as db:
        env = _setup(db)
        owner = env["owner"]
        product = _product(
            db, env["category"].id, env["uom"], description=f"{MARKER} close-coupled WC, white"
        )

        document = qdocs.create_document(db, project=env["project"], actor_user_id=owner)
        scope = qdocs.add_scope(
            db, document=document, scope_label=f"{MARKER} Townhouse", actor_user_id=owner
        )
        quotes.upsert_line(
            db,
            version=quotes.current_version(db, scope.id),
            actor_user_id=owner,
            payload={
                "product_id": product.id,
                "unit_price": PRICED_RATE,
                "quantity": PRICED_QTY,
            },
        )

        r1 = _issue(db, document, owner)

        v2 = quotes.revise(db, quotation=scope, actor_user_id=owner)
        quotes.upsert_line(
            db,
            version=v2,
            actor_user_id=owner,
            payload={"product_id": product.id, "unit_price": "999.99", "quantity": PRICED_QTY},
        )
        db.flush()
        db.expire_all()

        sheet = qxlsx.build_issue_workbook(db, r1).worksheets[0]
        amounts = [
            _money(v)
            for v in _values(sheet)
            if isinstance(v, (int, float, Decimal))
        ]

        assert _money(_labelled_amount(sheet, "TOTAL")) == PRICED_TOTAL
        assert Decimal("3999.96") not in amounts
        assert Decimal("999.99") not in amounts


# ------------------------------------------------------- component lines (real data)


def test_a_component_of_a_set_leaves_its_money_cells_empty_not_zero():
    """The Excel sheet is the one a QS re-prices in, so a zero here is worse than on paper: it
    is a number that sums, sorts and charts. Real quotations list the parts of a complete set on
    their own rows (the Tuju Residences document has a pedestal at 305.55 followed by 4 rows at
    nothing, because the money is on the parent), and none of them is flagged rate-only, since
    that flag means a quoted alternate.

    Empty, not the string "0.00" and not the number 0: a text cell would break the QS's own
    formulas just as badly. Asserted against the PDF's arithmetic so the two exports cannot
    drift into disagreeing about the same issue."""
    from app.services import project_quotation_document_service as qdocs
    from app.services import project_quotation_excel_service as qxlsx
    from app.services import project_quotation_service as quotes

    with blank_session() as db:
        env = _setup(db)
        owner = env["owner"]
        parent = _product(
            db, env["category"].id, env["uom"], description=f"{MARKER} close-coupled pedestal"
        )
        component = _product(
            db, env["category"].id, env["uom"], description=f"{MARKER} cistern only, no charge"
        )

        document = qdocs.create_document(db, project=env["project"], actor_user_id=owner)
        scope = qdocs.add_scope(
            db, document=document, scope_label=f"{MARKER} Townhouse", actor_user_id=owner
        )
        version = quotes.current_version(db, scope.id)
        quotes.upsert_line(
            db,
            version=version,
            actor_user_id=owner,
            payload={"product_id": parent.id, "unit_price": "305.55", "quantity": "894"},
        )
        quotes.upsert_line(
            db,
            version=version,
            actor_user_id=owner,
            payload={"product_id": component.id, "unit_price": "0.00", "quantity": "894"},
        )

        issued = _issue(db, document, owner)
        sheet = qxlsx.build_issue_workbook(db, issued).worksheets[0]

        rows = list(sheet.iter_rows())
        component_row = next(
            row
            for row in rows
            if any(str(cell.value or "").startswith(f"{MARKER} cistern only") for cell in row)
        )
        parent_row = next(
            row
            for row in rows
            if any(str(cell.value or "").startswith(f"{MARKER} close-coupled pedestal") for cell in row)
        )

        # The component is still listed, and the customer still receives 894 of them.
        assert any(cell.value == 894 for cell in component_row)
        # Nothing anywhere on its row reads as money, in either the rate or the amount column.
        assert not any(
            cell.value in (0, Decimal("0.00"), "0.00", "0", 0.0) for cell in component_row
        )

        # The priced parent is untouched, and the money is still a NUMBER, not text.
        assert any(cell.value == Decimal("305.55") for cell in parent_row)
        assert any(cell.value == Decimal("273161.70") for cell in parent_row)
        assert issued.grand_total == Decimal("273161.70")
