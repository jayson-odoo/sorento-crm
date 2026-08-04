"""Container status import: matching and writing (slice 3, part 2).

The rules that carry risk:

* **A3 - match across EVERY status.** The packing-list matcher deliberately only
  looks at not-fully-received shipments. Reusing it here would create a duplicate
  row for every container that already completed, and 318 of the 407 rows in the
  file are on the archived `Arrived` tab.
* **A5 - a blank cell never clears.** A re-upload of an older sheet must not wipe
  dates somebody has since filled in.
* **A4 - idempotent.** Importing the same file twice must change nothing the
  second time, including `updated_at`.
* **B4 - remarks are activity feed entries**, not columns and not internal notes
  (which are private to their author and would hide them).

Counts are asserted against seeded rows and the committed workbook, never against
whatever the local database happens to contain - the local DB is a copy of
production and CI's is empty.
"""
from __future__ import annotations

import uuid
from datetime import date
from pathlib import Path

import pytest

from app.models.activities import ActivityEvent
from app.models.procurement import InboundShipment, Supplier
from app.services.container_status_import import parse_container_status_workbook
from app.services.container_status_service import (
    ContainerStatusImportService,
)
from tests._pg_fixture import blank_session, unique_code


FIXTURE = (
    Path(__file__).resolve().parents[2]
    / "sorento_crm_frontend"
    / "e2e"
    / "fixtures"
    / "container-status-2026.xlsx"
)

requires_fixture = pytest.mark.skipif(
    not FIXTURE.exists(), reason="real workbook fixture not present in this checkout"
)


@pytest.fixture
def db():
    with blank_session() as s:
        yield s


@pytest.fixture
def supplier_id(db):
    supplier = Supplier(
        id=str(uuid.uuid4()),
        supplier_code=unique_code("SUP"),
        supplier_name="ZZT Container Status Supplier",
    )
    db.add(supplier)
    db.flush()
    return supplier.id


def _shipment(db, supplier_id, container, *, status="in_transit", **kwargs):
    shipment = InboundShipment(
        id=str(uuid.uuid4()),
        shipment_number=unique_code("SHP"),
        supplier_id=supplier_id,
        shipment_date=date(2026, 7, 1),
        shipping_container_number=container,
        shipment_status=status,
        **kwargs,
    )
    db.add(shipment)
    db.flush()
    return shipment


def _one_row_workbook(container: str, **cells) -> bytes:
    """A single data row, headers derived from the cells asked for."""
    import openpyxl
    from io import BytesIO

    header_for = {
        "eta_delay_date": "ETA DELAY",
        "eta_date": "ETA",
        "gatepass_date": "GATEPASS",
        "inspection_date": "INSPECTION",
        "approval_date": "APPROVAL",
        "liner_code": "LINER",
        "loc": "LOC",
        "consignee": "CONSIGNEE",
        "free_days_available": "FREE DAYS AVAILABLE",
        "remarks": "REMARKS 1",
    }
    keys = list(cells)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fitting"
    ws.append(["FITTING CONTAINER 2026"])
    ws.append(["NO", "CONTAINER"] + [header_for[k] for k in keys])
    ws.append([1, container] + [cells[k] for k in keys])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _service(db):
    return ContainerStatusImportService(db)


# ------------------------------------------------------------------ matching


def test_matches_a_completed_shipment_instead_of_creating_a_duplicate(db, supplier_id):
    """A3. 318 of 407 rows are archived containers; the packing-list matcher's
    not-fully-received rule would duplicate every one of them."""
    existing = _shipment(db, supplier_id, "GXYU5106903", status="fully_received")

    parsed = parse_container_status_workbook(
        _one_row_workbook("GXYU5106903", eta_delay_date=date(2026, 7, 8))
    )
    result = _service(db).apply(parsed, user_id=None)

    assert result["updated"] == 1
    assert result["created"] == 0
    db.expire(existing)
    assert existing.eta_delay_date == date(2026, 7, 8)


def test_matches_through_separator_differences(db, supplier_id):
    existing = _shipment(db, supplier_id, "GXYU-5106903")

    parsed = parse_container_status_workbook(
        _one_row_workbook("gxyu 5106903", gatepass_date=date(2026, 7, 17))
    )
    result = _service(db).apply(parsed, user_id=None)

    assert (result["updated"], result["created"]) == (1, 0)
    db.expire(existing)
    assert existing.gatepass_date == date(2026, 7, 17)


def test_creates_a_shipment_for_a_container_the_system_has_never_seen(db):
    parsed = parse_container_status_workbook(
        _one_row_workbook("TGBU9807730", eta_date=date(2026, 7, 19), liner_code="MSC")
    )
    result = _service(db).apply(parsed, user_id=None)

    assert (result["updated"], result["created"]) == (0, 1)
    created = (
        db.query(InboundShipment)
        .filter(InboundShipment.shipping_container_number == "TGBU9807730")
        .one()
    )
    assert created.eta_date == date(2026, 7, 19)
    assert created.liner_code == "MSC"
    assert created.source_sheet == "Fitting"
    # The sheet carries no supplier, and supplier_id is nullable.
    assert created.supplier_id is None


def test_a_created_shipment_gets_a_shipment_date_it_can_actually_store(db):
    """`shipment_date` is NOT NULL. The sheet has no such column, so the importer
    must supply something rather than blowing up on the insert."""
    parsed = parse_container_status_workbook(
        _one_row_workbook("TGBU9807730", eta_date=date(2026, 7, 19))
    )
    _service(db).apply(parsed, user_id=None)

    created = (
        db.query(InboundShipment)
        .filter(InboundShipment.shipping_container_number == "TGBU9807730")
        .one()
    )
    assert created.shipment_date is not None


# --------------------------------------------------------- blank never clears


def test_a_blank_cell_does_not_clear_a_value_already_recorded(db, supplier_id):
    """A5. Re-uploading last week's sheet must not wipe this week's work."""
    existing = _shipment(
        db,
        supplier_id,
        "GXYU5106903",
        eta_delay_date=date(2026, 7, 8),
        gatepass_date=date(2026, 7, 17),
    )

    # A sheet that mentions only the ETA delay. Gatepass is absent, not blank.
    parsed = parse_container_status_workbook(
        _one_row_workbook("GXYU5106903", eta_delay_date=date(2026, 7, 12))
    )
    _service(db).apply(parsed, user_id=None)

    db.expire(existing)
    assert existing.eta_delay_date == date(2026, 7, 12), "a value present wins"
    assert existing.gatepass_date == date(2026, 7, 17), "an absent cell changes nothing"


def test_an_empty_string_cell_is_treated_as_absent_not_as_a_clear(db, supplier_id):
    existing = _shipment(db, supplier_id, "GXYU5106903", consignee="Sorento (Mocha)")

    parsed = parse_container_status_workbook(
        _one_row_workbook("GXYU5106903", consignee="")
    )
    _service(db).apply(parsed, user_id=None)

    db.expire(existing)
    assert existing.consignee == "Sorento (Mocha)"


# ------------------------------------------------------------- idempotency


def test_importing_the_same_row_twice_changes_nothing_the_second_time(db, supplier_id):
    """A4. Includes `updated_at`: a no-op write would still bump it, so the
    importer must not touch a row whose values already agree."""
    _shipment(db, supplier_id, "GXYU5106903")
    workbook = _one_row_workbook(
        "GXYU5106903", eta_delay_date=date(2026, 7, 12), liner_code="CMA"
    )

    first = _service(db).apply(parse_container_status_workbook(workbook), user_id=None)
    db.flush()
    shipment = (
        db.query(InboundShipment)
        .filter(InboundShipment.shipping_container_number == "GXYU5106903")
        .one()
    )
    stamp = shipment.updated_at

    second = _service(db).apply(parse_container_status_workbook(workbook), user_id=None)
    db.flush()
    db.expire(shipment)

    assert first["updated"] == 1
    assert second["updated"] == 0
    assert second["unchanged"] == 1
    assert shipment.updated_at == stamp


@requires_fixture
def test_the_real_workbook_imports_once_and_is_then_a_no_op(db, supplier_id):
    """The golden case, seeded rather than borrowed.

    Three real containers are seeded so the update path is exercised; the rest of
    the file creates. Counts come from the committed FILE, so this holds on an
    empty CI database as well as locally.
    """
    for container in ("GXYU5106903", "OOCU8630645", "CICU1013499"):
        _shipment(db, supplier_id, container)

    parsed = parse_container_status_workbook(FIXTURE.read_bytes())
    assert len(parsed.rows) == 407

    first = _service(db).apply(parsed, user_id=None)
    assert first["updated"] == 3
    assert first["created"] == 404
    assert first["updated"] + first["created"] == 407

    total = db.query(InboundShipment).count()
    assert total == 407, "3 seeded rows were matched, not duplicated"

    second = _service(db).apply(
        parse_container_status_workbook(FIXTURE.read_bytes()), user_id=None
    )
    assert (second["created"], second["updated"]) == (0, 0)
    assert second["unchanged"] == 407
    assert db.query(InboundShipment).count() == 407


# ----------------------------------------------------------------- remarks


def test_remarks_land_in_the_activity_feed_not_on_the_shipment(db, supplier_id):
    """B4. `internal_notes` would be wrong - they are private to their author."""
    shipment = _shipment(db, supplier_id, "GXYU5106903")

    parsed = parse_container_status_workbook(
        _one_row_workbook("GXYU5106903", remarks="held at port pending CIDB")
    )
    _service(db).apply(parsed, user_id=None)
    db.flush()

    events = (
        db.query(ActivityEvent)
        .filter(
            ActivityEvent.entity_type == "inbound_shipment",
            ActivityEvent.entity_id == shipment.id,
        )
        .all()
    )
    assert len(events) == 1
    assert "held at port pending CIDB" in (events[0].body_text or "")
    assert events[0].kind == "user_update"


def test_the_same_remark_is_not_appended_again_on_re_import(db, supplier_id):
    """Otherwise a daily re-upload grows the feed without adding information."""
    shipment = _shipment(db, supplier_id, "GXYU5106903")
    workbook = _one_row_workbook("GXYU5106903", remarks="held at port")

    service = _service(db)
    service.apply(parse_container_status_workbook(workbook), user_id=None)
    db.flush()
    service.apply(parse_container_status_workbook(workbook), user_id=None)
    db.flush()

    count = (
        db.query(ActivityEvent)
        .filter(
            ActivityEvent.entity_type == "inbound_shipment",
            ActivityEvent.entity_id == shipment.id,
        )
        .count()
    )
    assert count == 1


# ------------------------------------------------------------------ dry run


@requires_fixture
def test_the_dry_run_reports_what_would_happen_and_writes_nothing(db, supplier_id):
    _shipment(db, supplier_id, "GXYU5106903")
    before = db.query(InboundShipment).count()

    report = _service(db).validate(FIXTURE.read_bytes())

    assert report["valid"] is True
    assert report["summary"] == {
        "total_rows": 407,
        "would_update": 1,
        "would_create": 406,
        "error_count": 0,
    }
    assert db.query(InboundShipment).count() == before, "a dry run must not write"


@requires_fixture
def test_the_dry_run_warns_about_the_alias_and_the_blank_rows_but_not_blocks(db):
    """The operator acts on aliases and skipped rows. Block detection is how the
    parser works, so it is deliberately not a warning."""
    report = _service(db).validate(FIXTURE.read_bytes())
    joined = " ".join(report["warnings"])

    assert "RL" in joined
    assert "475" in joined
    assert "header block" not in joined.lower()


def test_a_bad_container_becomes_an_error_the_operator_can_locate(db):
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fitting"
    ws.append(["TITLE"])
    ws.append(["NO", "CONTAINER", "ETA DELAY"])
    ws.append([1, "NOTACONTAINER", date(2026, 7, 8)])
    buffer = BytesIO()
    wb.save(buffer)

    report = _service(db).validate(buffer.getvalue())

    assert report["valid"] is False
    assert report["summary"]["error_count"] == 1
    assert any("NOTACONTAINER" in e and "Fitting" in e for e in report["errors"])


def test_a_file_that_is_not_a_container_status_sheet_fails_the_dry_run(db):
    import openpyxl
    from io import BytesIO

    wb = openpyxl.Workbook()
    wb.active.append(["NO", "SHIPMENT"])
    buffer = BytesIO()
    wb.save(buffer)

    report = _service(db).validate(buffer.getvalue())

    assert report["valid"] is False
    assert any("CONTAINER" in e for e in report["errors"])
    assert report["summary"]["total_rows"] == 0


def test_creating_without_a_company_scope_fails_with_a_sentence_not_a_constraint(db):
    """`inbound_shipments.company_id` is NOT NULL on a migrated database and is
    filled by the insert auto-stamp, which only fires when a company scope is set.
    A job with no company snapshot runs system-scoped, so the stamp has nothing to
    write and every insert dies.

    The blank scratch schema is built by `create_all` from the MODEL, where the
    column is deliberately nullable, so this test cannot reproduce the Postgres
    violation - it pins the GUARD instead, which is what turns the failure into a
    sentence the operator can act on.
    """
    from unittest.mock import patch

    from app.services.container_status_service import ContainerStatusCompanyScopeError

    parsed = parse_container_status_workbook(
        _one_row_workbook("TGBU9807730", eta_date=date(2026, 7, 19))
    )
    with patch(
        "app.services.job_service.active_company_id_from_scope", return_value=None
    ):
        with pytest.raises(ContainerStatusCompanyScopeError) as excinfo:
            _service(db).apply(parsed, user_id=None)

    assert "company" in str(excinfo.value).lower()


def test_an_update_only_import_does_not_need_a_company_scope(db, supplier_id):
    """The guard fires only when the run would CREATE. Updating existing rows
    stamps nothing, so a system-scoped re-upload must still work."""
    from unittest.mock import patch

    _shipment(db, supplier_id, "GXYU5106903")
    parsed = parse_container_status_workbook(
        _one_row_workbook("GXYU5106903", eta_delay_date=date(2026, 7, 12))
    )
    with patch(
        "app.services.job_service.active_company_id_from_scope", return_value=None
    ):
        result = _service(db).apply(parsed, user_id=None)

    assert result["updated"] == 1
