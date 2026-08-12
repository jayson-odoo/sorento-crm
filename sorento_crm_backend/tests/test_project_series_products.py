"""S18 - a series nominates PRODUCTS as well as categories.

The client's definition of "standard" is a list of product codes, not a set of groups:
"any product that is not in the sheet that I provided you are flagged as non standard".
Measured against their template, expressing it with categories would have called 15,048
products standard in order to capture 167, and would never have flagged the sibling SKU
the alert exists for. So membership is answered from BOTH sides, and the four cases below
are the whole truth table: nominated directly, inherited from a category, both, neither.

The importer is specified here too, and its most important case is the one that produces
NO membership: a code the catalogue does not carry has to come back to the user verbatim.
An unmatched code is the client's data telling us something (their sheet quotes base codes
that the catalogue only stocks in suffixed variants), and a loader that silently dropped
them would turn that signal into a smaller number nobody could interrogate.
"""
from __future__ import annotations

import uuid
from decimal import Decimal

import pytest

from sqlalchemy import text

from app.models.base import company_scope
from app.models.product import Product, ProductCategory, UnitOfMeasure
from app.services.error_handler import AppException

from ._pg_fixture import blank_session

MARKER = "zzt-series-products"


def _uid() -> str:
    return str(uuid.uuid4())


def _sorento(db) -> str:
    return db.execute(text("select id from companies where code = 'SRT'")).scalar()


def _uom(db) -> str:
    row = UnitOfMeasure(id=_uid(), uom_code=f"ZZT{_uid()[:4]}", uom_name="Piece")
    db.add(row)
    db.flush()
    return row.id


def _category(db, name: str, parent_id: str | None = None) -> ProductCategory:
    row = ProductCategory(
        id=_uid(),
        category_code=f"ZZT-{_uid()[:8]}",
        category_name=f"{MARKER} {name}",
        parent_category_id=parent_id,
    )
    db.add(row)
    db.flush()
    return row


def _product(db, code: str, category_id: str, uom_id: str) -> Product:
    row = Product(
        id=_uid(),
        product_code=code,
        product_name=f"{MARKER} {code}",
        category_id=category_id,
        base_uom_id=uom_id,
        list_price=Decimal("1000.00"),
    )
    db.add(row)
    db.flush()
    return row


def _series(db, company_id: str, name: str = "Template"):
    from app.models.projects import ProjectSeries

    series = ProjectSeries(id=_uid(), company_id=company_id, name=f"{MARKER} {name}")
    db.add(series)
    db.flush()
    return series


# ------------------------------------------------------- membership truth table


def test_a_product_nominated_by_name_is_in_the_series():
    """Direct membership, with NO category nominated at all. This is the shape the
    client's own template takes, and the shape the old model could not express."""
    from app.services import project_pricing_service as pricing

    with blank_session() as db:
        company_id = _sorento(db)
        uom = _uom(db)
        category = _category(db, "Basins")
        listed = _product(db, "ZZT-CWB-242", category.id, uom)
        sibling = _product(db, "ZZT-CWB-243", category.id, uom)
        series = _series(db, company_id)

        pricing.set_series_products(db, series=series, product_ids=[listed.id])

        assert pricing.is_in_series(db, series_id=series.id, product=listed) is True
        # The sibling shares the category and is still NOT standard, which is the whole
        # reason product-level membership had to exist.
        assert pricing.is_in_series(db, series_id=series.id, product=sibling) is False


def test_a_product_under_a_nominated_category_is_still_in_the_series():
    """Categories keep working, unchanged. This is an addition, not a replacement."""
    from app.services import project_pricing_service as pricing

    with blank_session() as db:
        company_id = _sorento(db)
        uom = _uom(db)
        top = _category(db, "Sanitary Ware")
        leaf = _category(db, "Wall-hung Basins", parent_id=top.id)
        inherited = _product(db, "ZZT-INH-1", leaf.id, uom)
        series = _series(db, company_id)

        pricing.set_series_categories(db, series=series, category_ids=[top.id])

        assert pricing.is_in_series(db, series_id=series.id, product=inherited) is True


def test_both_sources_combine_rather_than_one_overriding_the_other():
    from app.services import project_pricing_service as pricing

    with blank_session() as db:
        company_id = _sorento(db)
        uom = _uom(db)
        covered = _category(db, "Basins")
        elsewhere = _category(db, "Kitchen Sinks")
        by_category = _product(db, "ZZT-BOTH-CAT", covered.id, uom)
        by_name = _product(db, "ZZT-BOTH-NAME", elsewhere.id, uom)
        series = _series(db, company_id)

        pricing.set_series_categories(db, series=series, category_ids=[covered.id])
        pricing.set_series_products(db, series=series, product_ids=[by_name.id])

        assert pricing.is_in_series(db, series_id=series.id, product=by_category) is True
        assert pricing.is_in_series(db, series_id=series.id, product=by_name) is True


def test_neither_nominated_nor_inherited_is_out_of_the_series():
    from app.services import project_pricing_service as pricing

    with blank_session() as db:
        company_id = _sorento(db)
        uom = _uom(db)
        covered = _category(db, "Basins")
        elsewhere = _category(db, "Kitchen Sinks")
        listed = _product(db, "ZZT-NEITHER-IN", covered.id, uom)
        outside = _product(db, "ZZT-NEITHER-OUT", elsewhere.id, uom)
        series = _series(db, company_id)

        pricing.set_series_categories(db, series=series, category_ids=[covered.id])
        pricing.set_series_products(db, series=series, product_ids=[listed.id])

        assert pricing.is_in_series(db, series_id=series.id, product=outside) is False


def test_the_membership_is_expanded_once_and_reused_across_a_whole_version():
    """The pre-expansion the guardrail pass depends on: one hierarchy walk and one
    product read for a 52-line version, not 52 of each."""
    from app.services import project_pricing_service as pricing

    with blank_session() as db:
        company_id = _sorento(db)
        uom = _uom(db)
        covered = _category(db, "Basins")
        inherited = _product(db, "ZZT-EXP-CAT", covered.id, uom)
        named = _product(db, "ZZT-EXP-NAME", _category(db, "Loose").id, uom)
        outside = _product(db, "ZZT-EXP-OUT", _category(db, "Sinks").id, uom)
        series = _series(db, company_id)
        pricing.set_series_categories(db, series=series, category_ids=[covered.id])
        pricing.set_series_products(db, series=series, product_ids=[named.id])

        membership = pricing.series_membership(db, series.id)

        assert (
            pricing.is_in_series(
                db, series_id=series.id, product=inherited, membership=membership
            )
            is True
        )
        assert (
            pricing.is_in_series(
                db, series_id=series.id, product=named, membership=membership
            )
            is True
        )
        assert (
            pricing.is_in_series(
                db, series_id=series.id, product=outside, membership=membership
            )
            is False
        )


def test_replacing_the_nominated_products_reconciles_rather_than_rebuilding():
    from app.models.projects import ProjectSeriesProduct
    from app.services import project_pricing_service as pricing

    with blank_session() as db:
        company_id = _sorento(db)
        uom = _uom(db)
        keep = _product(db, "ZZT-KEEP", _category(db, "Loose").id, uom)
        drop = _product(db, "ZZT-DROP", _category(db, "Loose").id, uom)
        add = _product(db, "ZZT-ADD", _category(db, "Loose").id, uom)
        series = _series(db, company_id)
        pricing.set_series_products(db, series=series, product_ids=[keep.id, drop.id])

        pricing.set_series_products(db, series=series, product_ids=[keep.id, add.id])

        rows = {
            row.product_id
            for row in db.query(ProjectSeriesProduct)
            .filter(ProjectSeriesProduct.series_id == series.id)
            .all()
        }
        assert rows == {keep.id, add.id}


# ---------------------------------------------------------------- the importer


def test_codes_are_matched_through_the_repos_shared_normaliser():
    """The client's sheet carries trailing spaces, internal double spaces and dashes
    where the catalogue has none. ``variant_link_service.normalize_code`` is already the
    answer to that question everywhere else, so the importer asks it rather than
    inventing a tenth spelling of the same rule."""
    from app.services import project_pricing_service as pricing

    with blank_session() as db:
        company_id = _sorento(db)
        uom = _uom(db)
        product = _product(db, "ZZTCWB242", _category(db, "Loose").id, uom)
        series = _series(db, company_id)

        with company_scope(db, frozenset({company_id})):
            report = pricing.apply_series_product_codes(
                db,
                series=series,
                codes=["  zzt-cwb  242 ", "ZZT CWB-242"],
                mode="append",
            )

        assert report["submitted"] == 2
        # Both cells are the same code once normalised, so they buy one membership row.
        assert report["unique_codes"] == 1
        assert report["added"] == 1
        assert report["unmatched_codes"] == []
        assert pricing.is_in_series(db, series_id=series.id, product=product) is True


def test_a_code_the_catalogue_does_not_carry_comes_back_verbatim():
    """The requirement in one assertion: an unmatched code is REPORTED, never dropped."""
    from app.services import project_pricing_service as pricing

    with blank_session() as db:
        company_id = _sorento(db)
        uom = _uom(db)
        known = _product(db, "ZZT-KNOWN-1", _category(db, "Loose").id, uom)
        series = _series(db, company_id)

        with company_scope(db, frozenset({company_id})):
            report = pricing.apply_series_product_codes(
                db,
                series=series,
                codes=["ZZT-KNOWN-1", "ZZT-NOT-IN-CATALOGUE", "MWC7605-RL"],
                mode="append",
            )

        assert report["added"] == 1
        assert report["matched_codes"] == 1
        # Verbatim as submitted, so the admin can paste it back into their own sheet.
        assert report["unmatched_codes"] == ["ZZT-NOT-IN-CATALOGUE", "MWC7605-RL"]
        assert pricing.is_in_series(db, series_id=series.id, product=known) is True


def test_importing_the_same_list_twice_adds_nothing_and_says_so():
    from app.services import project_pricing_service as pricing

    with blank_session() as db:
        company_id = _sorento(db)
        uom = _uom(db)
        _product(db, "ZZT-IDEM-1", _category(db, "Loose").id, uom)
        series = _series(db, company_id)

        with company_scope(db, frozenset({company_id})):
            pricing.apply_series_product_codes(
                db, series=series, codes=["ZZT-IDEM-1"], mode="append"
            )
            second = pricing.apply_series_product_codes(
                db, series=series, codes=["ZZT-IDEM-1"], mode="append"
            )

        assert second["added"] == 0
        assert second["already_present"] == 1
        assert second["product_count"] == 1


def test_replace_mode_removes_what_the_new_list_leaves_out_and_counts_it():
    from app.services import project_pricing_service as pricing

    with blank_session() as db:
        company_id = _sorento(db)
        uom = _uom(db)
        _product(db, "ZZT-REP-KEEP", _category(db, "Loose").id, uom)
        _product(db, "ZZT-REP-DROP", _category(db, "Loose").id, uom)
        _product(db, "ZZT-REP-ADD", _category(db, "Loose").id, uom)
        series = _series(db, company_id)

        with company_scope(db, frozenset({company_id})):
            pricing.apply_series_product_codes(
                db,
                series=series,
                codes=["ZZT-REP-KEEP", "ZZT-REP-DROP"],
                mode="append",
            )
            report = pricing.apply_series_product_codes(
                db,
                series=series,
                codes=["ZZT-REP-KEEP", "ZZT-REP-ADD"],
                mode="replace",
            )

        assert report["added"] == 1
        assert report["already_present"] == 1
        assert report["removed"] == 1
        assert report["product_count"] == 2


def test_a_list_with_no_usable_code_is_refused_rather_than_written():
    """An empty paste would otherwise WIPE the series in replace mode while reporting a
    cheerful zero."""
    from app.services import project_pricing_service as pricing

    with blank_session() as db:
        company_id = _sorento(db)
        series = _series(db, company_id)

        with company_scope(db, frozenset({company_id})):
            with pytest.raises(AppException) as exc:
                pricing.apply_series_product_codes(
                    db, series=series, codes=["", "   "], mode="replace"
                )

        assert exc.value.status_code == 422


def test_another_companys_product_is_never_matched_by_code():
    """Two companies stock the same codes. A series belongs to one of them, and matching
    across the boundary would nominate a row the admin cannot even see."""
    from app.services import project_pricing_service as pricing

    with blank_session() as db:
        company_id = _sorento(db)
        other_id = _uid()
        db.execute(
            text(
                "insert into companies (id, code, name) values (:id, :code, :name)"
            ),
            {"id": other_id, "code": f"ZZ{_uid()[:4]}", "name": f"{MARKER} Other"},
        )
        uom = _uom(db)
        with company_scope(db, frozenset({other_id})):
            theirs = _product(db, "ZZT-SHARED-CODE", _category(db, "Loose").id, uom)
            assert theirs.company_id == other_id

        series = _series(db, company_id)
        with company_scope(db, frozenset({company_id})):
            report = pricing.apply_series_product_codes(
                db, series=series, codes=["ZZT-SHARED-CODE"], mode="append"
            )

        assert report["added"] == 0
        assert report["unmatched_codes"] == ["ZZT-SHARED-CODE"]


# ------------------------------------------------------------- sheet extraction


def test_codes_are_read_out_of_a_workbook_by_its_PRODUCT_CODE_heading():
    """The client's template puts the header on row 2 and the codes in column F, across
    three sheets. Read by HEADING, never by position: a column that moves must not turn
    every description into a product code."""
    import io

    import openpyxl

    from app.services import project_series_import_service as sheets

    book = openpyxl.Workbook()
    first = book.active
    first.title = "wares"
    first.append(["PROPOSED CABANA ITEMS"])
    first.append(["ITEM ", "PRODUCT IMAGE ", " DESCRIPTION ", "BRAND ", "PRODUCT CODE", "DEVELOPERS"])
    first.append([1, None, "One piece WC", "CABANA", "CWC7601-S-RL", 220])
    first.append([2, None, "Closed couple WC", "CABANA", " CWB 242 ", 200])
    second = book.create_sheet("shower")
    second.append(["PROPOSED BASIN TAP"])
    second.append(["ITEM ", "PRODUCT IMAGE ", " DESCRIPTION ", "BRAND ", "PRODUCT CODE"])
    second.append([1, None, "Rain shower head", "SORENTO", "SRTSH1040-T"])
    buffer = io.BytesIO()
    book.save(buffer)

    codes = sheets.extract_product_codes(buffer.getvalue(), filename="template.xlsx")

    assert codes == ["CWC7601-S-RL", "CWB 242", "SRTSH1040-T"]


def test_a_sheet_with_no_product_code_column_is_refused_with_the_reason():
    import io

    import openpyxl

    from app.services import project_series_import_service as sheets

    book = openpyxl.Workbook()
    book.active.append(["ITEM", "DESCRIPTION", "BRAND"])
    book.active.append([1, "One piece WC", "CABANA"])
    buffer = io.BytesIO()
    book.save(buffer)

    with pytest.raises(AppException) as exc:
        sheets.extract_product_codes(buffer.getvalue(), filename="template.xlsx")

    assert exc.value.status_code == 422
    assert "product code" in exc.value.detail["message"].lower()


def test_a_csv_of_bare_codes_is_read_as_a_single_column():
    """Not every list arrives as the client's template. A one-column CSV with no heading
    at all is the other thing an admin will hand us."""
    from app.services import project_series_import_service as sheets

    codes = sheets.extract_product_codes(
        b"CWC7601-S-RL\nCWB 242\n\nSRTSH1040-T\n", filename="codes.csv"
    )

    assert codes == ["CWC7601-S-RL", "CWB 242", "SRTSH1040-T"]


def test_a_cell_holding_two_codes_on_two_lines_is_read_as_two():
    """Straight off the client's own sheet: one PRODUCT CODE cell carries `MAB7050-WH`
    above `SRTWHBWP`. Read whole it matches nothing and reports as one unreadable string;
    split, both are looked up and the miss report names the code that is really missing."""
    import io

    import openpyxl

    from app.services import project_series_import_service as sheets

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["ITEM ", " DESCRIPTION ", "PRODUCT CODE"])
    sheet.append([1, "Basin", "MAB7050-WH\nSRTWHBWP"])
    buffer = io.BytesIO()
    book.save(buffer)

    assert sheets.extract_product_codes(buffer.getvalue(), filename="t.xlsx") == [
        "MAB7050-WH",
        "SRTWHBWP",
    ]


# ---------------------------------------------------------------- T2: price and discount
#
# The sheet carries two numbers beside each code and S18 deliberately did not read them
# ("price floors already have their own model"). The client has since asked for exactly
# those numbers on the series, so the narrow reader widens. Measured across their book of
# 151 codes: 95 carry a price, 56 carry a discount, and the DISTRIBUTORS column is written
# two different ways in two tabs.


def test_a_percentage_is_understood_however_the_sheet_spells_it():
    """`6 % MAX` in `wares` and `0.06` in `fittings` are the same six percent.

    The rule has to be stated because it is a guess either way round: a value BELOW one is a
    fraction and is multiplied out, a value of one or more is already a percentage. Anything
    carrying a literal `%` is a percentage whatever its magnitude, so `6 % MAX` can never be
    read as six hundred percent. Their whole book is 6, 8 and 10, and every spelling of those
    lands on the same three numbers.
    """
    from app.services import project_series_import_service as sheets

    assert sheets.normalise_percent("6 % MAX") == Decimal("6")
    assert sheets.normalise_percent("8 % MAX") == Decimal("8")
    assert sheets.normalise_percent("0.06") == Decimal("6")
    assert sheets.normalise_percent(0.1) == Decimal("10")
    assert sheets.normalise_percent(6) == Decimal("6")
    assert sheets.normalise_percent("6%") == Decimal("6")

    # Silence stays silence. A blank discount is NOT zero - reading it as "no discount
    # permitted" would put a hard floor under the 56 codes that carry a price and no
    # percentage.
    assert sheets.normalise_percent(None) is None
    assert sheets.normalise_percent("") is None
    assert sheets.normalise_percent("   ") is None
    assert sheets.normalise_percent("n/a") is None


def test_a_price_is_read_as_money_and_rubbish_is_not_invented():
    from app.services import project_series_import_service as sheets

    assert sheets.normalise_price(220) == Decimal("220")
    assert sheets.normalise_price("200") == Decimal("200")
    assert sheets.normalise_price("RM 1,250.50") == Decimal("1250.50")
    assert sheets.normalise_price(None) is None
    assert sheets.normalise_price("") is None
    assert sheets.normalise_price("TBC") is None
    # Negative money is not a price anybody meant.
    assert sheets.normalise_price(-5) is None


def test_the_sheet_is_read_as_code_price_and_discount_together():
    """One pass returns the row, not three parallel lists that can fall out of step.

    Mirrors the real workbook: DEVELOPERS holds the price and DISTRIBUTORS the percentage,
    the singular spellings appear in another tab, and a tab with neither column is still
    read for its codes.
    """
    import io

    import openpyxl

    from app.services import project_series_import_service as sheets

    book = openpyxl.Workbook()
    wares = book.active
    wares.title = "wares"
    wares.append(["PROPOSED CABANA ITEMS"])
    wares.append(
        ["ITEM ", " DESCRIPTION ", "BRAND ", "PRODUCT CODE", "DEVELOPERS", "DISTRIBUTORS"]
    )
    wares.append([1, "One piece WC", "CABANA", "CWC7601-S-RL", 220, "6 % MAX"])
    wares.append([2, "Closed couple WC", "CABANA", "CWC1009-RL", 200, "8 % MAX"])
    # A priced row with no percentage at all - 56 of the client's codes look like this.
    wares.append([3, "Cistern", "CABANA", "CWC605-RL", 170, None])

    fittings = book.create_sheet("fittings")
    fittings.append(["PROPOSED CABANA ITEMS"])
    # The SINGULAR spellings, and the fraction form of the percentage.
    fittings.append(
        ["ITEM ", " DESCRIPTION ", "BRAND ", "PRODUCT CODE", "DEVELOPER", "DISTRIBUTOR"]
    )
    fittings.append([1, "Angle valve", "CABANA", "B2155-BLUE", 8, 0.06])

    shower = book.create_sheet("shower")
    shower.append(["PROPOSED BASIN TAP"])
    shower.append(["ITEM ", " DESCRIPTION ", "BRAND ", "PRODUCT CODE"])
    shower.append([1, "Rain shower head", "SORENTO", "SRTSH1040-T"])

    buffer = io.BytesIO()
    book.save(buffer)

    rows = sheets.extract_series_rows(buffer.getvalue(), filename="template.xlsx")

    assert [row.code for row in rows] == [
        "CWC7601-S-RL",
        "CWC1009-RL",
        "CWC605-RL",
        "B2155-BLUE",
        "SRTSH1040-T",
    ]
    assert [row.selling_price for row in rows] == [
        Decimal("220"),
        Decimal("200"),
        Decimal("170"),
        Decimal("8"),
        None,
    ]
    assert [row.max_discount_pct for row in rows] == [
        Decimal("6"),
        Decimal("8"),
        None,
        Decimal("6"),
        None,
    ]


def test_reading_codes_alone_still_works_for_callers_that_only_want_them():
    """`extract_product_codes` is the S18 entry point and several callers still use it."""
    import io

    import openpyxl

    from app.services import project_series_import_service as sheets

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.append(["ITEM ", "PRODUCT CODE", "DEVELOPERS"])
    sheet.append([1, "CWC7601-S-RL", 220])
    buffer = io.BytesIO()
    book.save(buffer)

    assert sheets.extract_product_codes(buffer.getvalue(), filename="t.xlsx") == [
        "CWC7601-S-RL"
    ]


def test_importing_a_priced_sheet_stores_the_price_and_the_percentage():
    from app.models.projects import ProjectSeriesProduct
    from app.services import project_pricing_service as pricing
    from app.services.project_series_import_service import SeriesSheetRow

    with blank_session() as db:
        company_id = _sorento(db)
        uom = _uom(db)
        category = _category(db, "Sanitary").id
        wc = _product(db, "CWC7601-S-RL", category, uom)
        valve = _product(db, "B2155-BLUE", category, uom)
        bare = _product(db, "SRTSH1040-T", category, uom)
        series = _series(db, company_id)

        report = pricing.apply_series_product_codes(
            db,
            series=series,
            codes=[
                SeriesSheetRow("CWC7601-S-RL", Decimal("220"), Decimal("6")),
                SeriesSheetRow("B2155-BLUE", Decimal("8"), Decimal("6")),
                # Priced, no percentage - 56 of the client's 151 codes look like this.
                SeriesSheetRow("SRTSH1040-T", None, None),
            ],
            mode="replace",
        )
        assert report["matched_codes"] == 3

        rows = {
            row.product_id: row
            for row in db.query(ProjectSeriesProduct)
            .filter(ProjectSeriesProduct.series_id == series.id)
            .all()
        }
        assert rows[wc.id].selling_price == Decimal("220.00")
        assert rows[wc.id].max_discount_pct == Decimal("6.00")
        assert rows[valve.id].selling_price == Decimal("8.00")
        # Stated nothing, so it holds nothing. NOT zero.
        assert rows[bare.id].selling_price is None
        assert rows[bare.id].max_discount_pct is None


def test_re_importing_a_corrected_sheet_updates_a_price_it_does_not_skip_it():
    """The row is already on the series, so it counts as "already present" - and the whole
    point of re-importing is that the NUMBER changed. Treating an existing nomination as
    nothing-to-do would silently keep last year's price."""
    from app.models.projects import ProjectSeriesProduct
    from app.services import project_pricing_service as pricing
    from app.services.project_series_import_service import SeriesSheetRow

    with blank_session() as db:
        company_id = _sorento(db)
        uom = _uom(db)
        wc = _product(db, "CWC7601-S-RL", _category(db, "Sanitary").id, uom)
        series = _series(db, company_id)

        pricing.apply_series_product_codes(
            db,
            series=series,
            codes=[SeriesSheetRow("CWC7601-S-RL", Decimal("220"), Decimal("6"))],
            mode="append",
        )
        report = pricing.apply_series_product_codes(
            db,
            series=series,
            codes=[SeriesSheetRow("CWC7601-S-RL", Decimal("245"), Decimal("8"))],
            mode="append",
        )

        assert report["added"] == 0
        assert report["already_present"] == 1
        row = (
            db.query(ProjectSeriesProduct)
            .filter(ProjectSeriesProduct.product_id == wc.id)
            .one()
        )
        assert row.selling_price == Decimal("245.00")
        assert row.max_discount_pct == Decimal("8.00")


def test_a_code_only_paste_never_erases_a_price_somebody_typed():
    """The paste box sends bare codes. Re-nominating a product that already carries a price
    must not blank it: the admin said "these are in the series", not "these cost nothing"."""
    from app.models.projects import ProjectSeriesProduct
    from app.services import project_pricing_service as pricing
    from app.services.project_series_import_service import SeriesSheetRow

    with blank_session() as db:
        company_id = _sorento(db)
        uom = _uom(db)
        wc = _product(db, "CWC7601-S-RL", _category(db, "Sanitary").id, uom)
        series = _series(db, company_id)
        pricing.apply_series_product_codes(
            db,
            series=series,
            codes=[SeriesSheetRow("CWC7601-S-RL", Decimal("220"), Decimal("6"))],
            mode="append",
        )

        # Plain strings, the S18 shape, still accepted.
        pricing.apply_series_product_codes(
            db, series=series, codes=["CWC7601-S-RL"], mode="append"
        )

        row = (
            db.query(ProjectSeriesProduct)
            .filter(ProjectSeriesProduct.product_id == wc.id)
            .one()
        )
        assert row.selling_price == Decimal("220.00")
        assert row.max_discount_pct == Decimal("6.00")


def test_the_series_floor_is_the_price_less_the_percentage():
    """220 at 6% is 206.80. The number a refusal is argued from, so it is computed once here
    and used by both the screen and the engine."""
    from app.services.project_pricing_service import series_floor

    assert series_floor(Decimal("220"), Decimal("6")) == Decimal("206.80")
    assert series_floor(Decimal("200"), Decimal("8")) == Decimal("184.00")
    assert series_floor(Decimal("8"), Decimal("6")) == Decimal("7.52")
    # Rounded to the cent, half up, so a floor is never a fraction of one.
    assert series_floor(Decimal("99.99"), Decimal("10")) == Decimal("89.99")
    # Zero percent IS a real instruction when somebody types it: sell at exactly this price.
    assert series_floor(Decimal("220"), Decimal("0")) == Decimal("220.00")


def test_a_series_with_only_half_the_answer_sets_no_floor_at_all():
    """AC-C4, and the reason the columns are nullable.

    A price with no percentage must fall through to `price_floor_rules` rather than becoming
    a hard floor at the selling price - 56 of the client's 151 codes are exactly that shape,
    and the alternative puts every one of them in breach on the first discount.
    """
    from app.services.project_pricing_service import series_floor

    assert series_floor(Decimal("220"), None) is None
    assert series_floor(None, Decimal("6")) is None
    assert series_floor(None, None) is None
    # Nonsense is refused rather than propagated as a negative floor.
    assert series_floor(Decimal("220"), Decimal("140")) is None
