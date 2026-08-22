"""P10 order inquiry ROUTES (AC-I4, AC-I5, AC-I7).

Route-level because the wiring is its own seam: the list, the export and the mark are
each reachable only if the router mounts BEFORE `/projects/{project_id}`, and a captured
path 404s while looking exactly like a project that has no inquiry.

The grants are the point of the auth cases. Reading is the project's view grant; acting
on a row is `projects.order_inquiry.action`, which is purchasing's rather than the
project owner's, so a reader must be refused the mark.
"""
from __future__ import annotations

import io
import uuid
from datetime import date, datetime
from decimal import Decimal

import openpyxl
import pytest
from sqlalchemy import text

from app.models.product import Product, ProductCategory, UnitOfMeasure
from app.models.project_so import (
    INQUIRY_ACTIONED,
    INQUIRY_RAISED,
    SO_STATUS_DRAFT,
    ProjectSalesOrder,
    ProjectSalesOrderLine,
)
from app.models.user import User
from app.services import project_seed_service

from ._pg_fixture import blank_session

MARKER = "zzt-oi-route"
BASE = "/api/v1/project-sales"

READ_ONLY = ["projects.projects.view"]
PURCHASING = ["projects.projects.view", "projects.order_inquiry.action"]


def _uid() -> str:
    return str(uuid.uuid4())


def _sorento(db) -> str:
    return db.execute(text("select id from companies where code = 'SRT'")).scalar()


def _user(db, name: str) -> str:
    user_id = _uid()
    db.add(User(id=user_id, email=f"{user_id}@zzt.test", name=name))
    db.flush()
    return user_id


def _product(db, code: str) -> Product:
    uom = UnitOfMeasure(id=_uid(), uom_code=f"ZZT{_uid()[:6]}", uom_name="Unit")
    category = ProductCategory(
        id=_uid(), category_code=f"ZZT-{_uid()[:8]}", category_name=f"{MARKER} cat"
    )
    db.add_all([uom, category])
    db.flush()
    row = Product(
        id=_uid(),
        product_code=code,
        product_name=f"{MARKER} {code}",
        category_id=category.id,
        base_uom_id=uom.id,
        list_price=Decimal("100.00"),
    )
    db.add(row)
    db.flush()
    return row


def _client(db, user_id: str, permissions):
    from fastapi.testclient import TestClient

    from app.database import get_db
    from app.dependencies import get_current_user, get_current_user_or_api_key
    from app.main import app
    from app.services.company_scope_resolver import apply_company_scope
    from app.services.user_service import UserPermissionService

    actor = {"id": user_id, "email": f"{user_id}@zzt.test", "role": "user"}
    app.dependency_overrides[get_db] = lambda: db
    app.dependency_overrides[get_current_user] = lambda: dict(actor)
    app.dependency_overrides[get_current_user_or_api_key] = lambda: dict(actor)
    app.dependency_overrides[apply_company_scope] = lambda: None

    originals = (
        UserPermissionService.check_user_has_permission,
        UserPermissionService.get_user_permission_slugs,
    )
    granted = list(permissions)
    UserPermissionService.check_user_has_permission = (
        lambda self, uid, slug: slug in granted
    )
    UserPermissionService.get_user_permission_slugs = lambda self, uid: list(granted)
    return TestClient(app), originals


def _restore(originals) -> None:
    from app.main import app
    from app.services.user_service import UserPermissionService

    UserPermissionService.check_user_has_permission = originals[0]
    UserPermissionService.get_user_permission_slugs = originals[1]
    app.dependency_overrides.clear()


def _seed_inquiry(db, company_id: str, user_id: str):
    from app.services.project_order_inquiry_service import ProjectOrderInquiryService
    from app.services.project_service import register_project

    project = register_project(
        db,
        company_id=company_id,
        actor_user_id=user_id,
        developer_party_id=None,
        title=f"{MARKER} Tuju Residences {_uid()[:6]}",
    )
    order = ProjectSalesOrder(
        id=_uid(),
        company_id=company_id,
        project_id=project.id,
        area_group="TOWER",
        provisional_ref=f"ZZT-PSO-{_uid()[:8]}",
        autocount_doc_no="SO397450",
        status=SO_STATUS_DRAFT,
        grouping_origin="area",
    )
    db.add(order)
    db.flush()
    for index, (code, qty) in enumerate((("CB6633", "600"), ("SRT382-6", "135")), start=1):
        db.add(
            ProjectSalesOrderLine(
                id=_uid(),
                company_id=company_id,
                project_sales_order_id=order.id,
                line_no=index,
                product_id=_product(db, code).id,
                description=f"{MARKER} {code}",
                qty=Decimal(qty),
                uom="UNIT",
                unit_price=Decimal("10.00"),
                amount=Decimal(qty) * Decimal("10.00"),
                delivery_date=date(2028, 7, 1),
            )
        )
    db.flush()
    # Written the way Stage 1C writes it: one Buy row per line, inside what the atomic
    # confirmation does (`refresh_for_decision`). The route tests are about the ROUTES,
    # so the rows are made by the one writer rather than by a whole confirmation.
    from app.models.project_so import SOSupplyDecision

    lines = (
        db.query(ProjectSalesOrderLine)
        .filter(ProjectSalesOrderLine.project_sales_order_id == order.id)
        .order_by(ProjectSalesOrderLine.line_no.asc())
        .all()
    )
    decision = SOSupplyDecision(
        id=_uid(),
        company_id=company_id,
        project_sales_order_id=order.id,
        revision_no=1,
        state="active",
        line_snapshots=[{"line_no": line.line_no} for line in lines],
        confirmed_by=user_id,
        confirmed_at=datetime.utcnow(),
    )
    db.add(decision)
    db.flush()
    service = ProjectOrderInquiryService(db)
    inquiry = service.refresh_for_decision(
        order,
        decision,
        [
            {
                "line": line,
                "line_no": line.line_no,
                "item_code": service._product_code(line.product_id),
                "buy_qty": Decimal(str(line.qty)),
                "required_date": line.delivery_date,
                "stock_location": None,
            }
            for line in lines
        ],
        actor_user_id=user_id,
    )["inquiry"]
    db.commit()
    return project, order, inquiry


def _api(permissions):
    from app.models.base import company_scope

    with blank_session() as db:
        company_id = _sorento(db)
        project_seed_service.run(db, company_id=company_id)
        user_id = _user(db, f"{MARKER} Eling")
        project, order, inquiry = _seed_inquiry(db, company_id, user_id)
        client, originals = _client(db, user_id, permissions)
        try:
            with company_scope(db, frozenset({company_id})):
                yield client, db, project, order, inquiry, user_id
        finally:
            _restore(originals)


@pytest.fixture()
def api():
    yield from _api(PURCHASING)


@pytest.fixture()
def reader_api():
    yield from _api(READ_ONLY)


def test_the_rows_list_is_reachable_and_not_captured_by_the_project_route(api):
    client, _db, project, _order, _inquiry, _user_id = api

    response = client.get(f"{BASE}/projects/{project.id}/order-inquiry-rows")

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["pagination"]["total"] == 2
    assert {row["item_code"] for row in body["data"]} == {"CB6633", "SRT382-6"}
    assert all(row["sales_order_ref"] == "SO397450" for row in body["data"])
    assert all(row["state"] == INQUIRY_RAISED for row in body["data"])


def test_the_rows_list_filters_by_verb(api):
    client, _db, project, _order, _inquiry, _user_id = api

    response = client.get(
        f"{BASE}/projects/{project.id}/order-inquiry-rows", params={"verb": "ORDER"}
    )

    assert response.status_code == 200, response.text
    assert response.json()["pagination"]["total"] == 2

    none_left = client.get(
        f"{BASE}/projects/{project.id}/order-inquiry-rows", params={"verb": "DELAY"}
    )
    assert none_left.json()["pagination"]["total"] == 0
    assert none_left.json()["empty"] is True


def test_the_summary_counts_what_is_still_open(api):
    client, _db, project, _order, _inquiry, _user_id = api

    body = client.get(f"{BASE}/projects/{project.id}/order-inquiry-summary").json()

    assert body == {"total": 2, "raised": 2, "actioned": 0, "cancelled": 0}


def test_the_sales_order_view_returns_the_inquiry_with_its_task(api):
    client, _db, _project, order, inquiry, _user_id = api

    response = client.get(f"{BASE}/sales-orders/{order.id}/order-inquiry")

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["id"] == inquiry.id
    assert body["task_id"]
    assert len(body["rows"]) == 2


def test_a_sales_order_with_nothing_raised_says_so_rather_than_returning_nothing(api):
    client, db, project, _order, _inquiry, _user_id = api
    other = ProjectSalesOrder(
        id=_uid(),
        company_id=project.company_id,
        project_id=project.id,
        area_group="COMMON AREA",
        provisional_ref=f"ZZT-PSO-{_uid()[:8]}",
        status=SO_STATUS_DRAFT,
        grouping_origin="area",
    )
    db.add(other)
    db.commit()

    response = client.get(f"{BASE}/sales-orders/{other.id}/order-inquiry")

    assert response.status_code == 404
    assert "publishes" in response.json()["message"]


def test_marking_a_row_records_who_and_when(api):
    client, _db, project, _order, _inquiry, _user_id = api
    rows = client.get(f"{BASE}/projects/{project.id}/order-inquiry-rows").json()["data"]

    response = client.post(
        f"{BASE}/order-inquiry-rows/mark",
        json={"row_ids": [rows[0]["id"]], "state": INQUIRY_ACTIONED},
    )

    assert response.status_code == 200, response.text
    body = response.json()
    assert body[0]["state"] == INQUIRY_ACTIONED
    assert body[0]["actioned_at"]
    assert body[0]["actioned_by_name"] == f"{MARKER} Eling"


def test_marking_refuses_a_state_that_is_not_one_of_the_three(api):
    client, _db, project, _order, _inquiry, _user_id = api
    rows = client.get(f"{BASE}/projects/{project.id}/order-inquiry-rows").json()["data"]

    response = client.post(
        f"{BASE}/order-inquiry-rows/mark",
        json={"row_ids": [rows[0]["id"]], "state": "done"},
    )

    assert response.status_code == 422


def test_a_reader_may_read_the_rows_but_not_act_on_them(reader_api):
    client, _db, project, _order, _inquiry, _user_id = reader_api

    rows = client.get(f"{BASE}/projects/{project.id}/order-inquiry-rows")
    assert rows.status_code == 200

    refused = client.post(
        f"{BASE}/order-inquiry-rows/mark",
        json={"row_ids": [rows.json()["data"][0]["id"]], "state": INQUIRY_ACTIONED},
    )
    assert refused.status_code == 403


def test_the_export_downloads_a_workbook_with_the_clients_headings(api):
    client, _db, project, _order, _inquiry, _user_id = api

    response = client.get(f"{BASE}/projects/{project.id}/order-inquiry-export")

    assert response.status_code == 200, response.text
    assert "attachment;" in response.headers["content-disposition"]
    assert response.headers["content-disposition"].endswith('.xlsx"')
    sheet = openpyxl.load_workbook(io.BytesIO(response.content)).worksheets[0]
    assert sheet.cell(row=1, column=1).value == "ORDER INQUIRY"
    assert [cell.value for cell in sheet[2][:8]] == [
        "SO DATE",
        "S/O NO",
        "ITEM CODE",
        "QTY",
        "DELIVERY DATE",
        "PROJECT/CUSTOMER",
        "STOCK LOCATION",
        "REMARK",
    ]
    assert sheet.max_row == 4


def test_the_export_honours_the_same_filters_as_the_list(api):
    client, _db, project, _order, _inquiry, _user_id = api

    response = client.get(
        f"{BASE}/projects/{project.id}/order-inquiry-export",
        params={"query": "CB6633"},
    )

    sheet = openpyxl.load_workbook(io.BytesIO(response.content)).worksheets[0]
    codes = [row[2] for row in sheet.iter_rows(min_row=3, values_only=True)]
    assert codes == ["CB6633"]


def test_an_unknown_project_is_a_404_not_an_empty_list(api):
    client, _db, _project, _order, _inquiry, _user_id = api

    response = client.get(f"{BASE}/projects/{_uid()}/order-inquiry-rows")

    assert response.status_code == 404
