"""S2 quotation DOCUMENT routes (UAC-project-quotation-document, Groups A, B, D, G).

Written against the UAC rather than against the handlers: nothing in this file was derived
from reading ``app/api/v1/projects/quotation_documents.py``, so it states what the routes owe
a client rather than restating what they happen to do.

Route-level rather than service-level because the HTTP seam carries decisions the service
cannot: whether a refusal reaches the client as a 422 it can render, whether a document
belonging to somebody else's project leaks through a URL that names THIS project, and
whether a permission a salesperson does not hold actually stops the write. The sibling
suite ``test_project_quotation_document.py`` already proves the arithmetic; nothing here
re-proves it, it proves the arithmetic reaches the wire.

The three cases worth the most:

- **The grand total on the wire excludes rate-only lines** (AC-C2 / AC-D3). The service
  test pins the sum; this pins what a client actually reads, because a serializer that
  recomputed the total itself would pass every service test and still send the customer a
  number Sorento never priced.
- **An issued document refuses deletion** (AC-A6) through the route, with a message that
  names withdrawal. A silent hide leaves a reference in a customer inbox this system
  cannot explain.
- **An issued version refuses line edits** (AC-B4) through the LINE route, which is a
  different router from the one being built. The freeze is derived from the issue rows, so
  it has to hold on an endpoint whose author never thought about documents.

Postgres only, via ``blank_session``. Every row created here carries the ``zzt-qdocroute``
marker, including the customer-facing document number, so nothing in this file can touch
the real data the dev database holds.
"""
from __future__ import annotations

import uuid
from contextlib import contextmanager
from datetime import date
from decimal import Decimal
from urllib.parse import quote

import pytest
from sqlalchemy import text

from app.models.numbering import DocumentNumberingRule
from app.models.product import Product, ProductCategory, UnitOfMeasure
from app.models.projects import ProjectParty
from app.models.user import User
from app.services import project_seed_service

from ._pg_fixture import blank_session

MARKER = "zzt-qdocroute"

# The quotation router mounts under the project-sales module prefix, so the paths the
# plan writes as ``/api/v1/projects/{id}/quotation-documents`` are reached here.
BASE = "/api/v1/project-sales"

VIEW = "projects.projects.view"
EDIT = "projects.projects.edit"
DELETE = "projects.projects.delete"

ALL_SLUGS = [
    VIEW,
    "projects.projects.create",
    EDIT,
    DELETE,
    "projects.projects.manage",
    "projects.types.view",
    "projects.types.edit",
]

# Off the real workbook (Cabana Elmina - nadi cergas R2): one priced line and one rate-only
# alternate at the SAME quantity, so a total that wrongly counted the alternate lands on a
# number that is obviously wrong rather than plausibly wrong.
SAMPLE_QTY = "1046"
PRICED_RATE = "250.00"
RATE_ONLY_RATE = "180.00"
TOWNHOUSE_TOTAL = Decimal("261500.00")  # 1046 x 250
GUARD_HOUSE_TOTAL = Decimal("1600.00")  # 4 x 400
CORRECT_GRAND_TOTAL = Decimal("263100.00")
WRONG_GRAND_TOTAL_IF_RATE_ONLY_COUNTED = Decimal("451380.00")  # + 1046 x 180



def _sign(client, root: str, document_id: str) -> None:
    """Sign a draft over HTTP so it can be issued.

    AC-H1 is a real contract change from the client's both-sides-e-sign decision: an unsigned
    quotation cannot be issued. These tests predate the gate, so they sign first rather than having
    the assertion loosened.
    """
    response = client.post(
        f"{root}/{document_id}/sign",
        json={
            "signer_name": f"{MARKER} Baser",
            "mode": "draw",
            "image_data_uri": "data:image/png;base64,zzt",
        },
    )
    assert response.status_code == 201, response.text

def _uid() -> str:
    return str(uuid.uuid4())


def _sorento(db) -> str:
    """Sorento's company id as a STRING, the shape the request path carries."""
    return str(db.execute(text("select id from companies where code = 'SRT'")).scalar())


def _user(db, name: str) -> str:
    user_id = _uid()
    db.add(User(id=user_id, email=f"{user_id}@zzt.test", name=name))
    db.flush()
    return user_id


def _uom(db) -> str:
    row = UnitOfMeasure(id=_uid(), uom_code=f"ZZT{_uid()[:4]}", uom_name="Piece")
    db.add(row)
    db.flush()
    return row.id


def _category(db, name: str) -> ProductCategory:
    row = ProductCategory(
        id=_uid(),
        category_code=f"ZZT-{_uid()[:8]}",
        category_name=f"{MARKER} {name}",
    )
    db.add(row)
    db.flush()
    return row


def _product(db, category_id: str, uom_id: str, list_price: str) -> Product:
    row = Product(
        id=_uid(),
        product_code=f"ZZT-{_uid()[:8]}",
        product_name=f"{MARKER} WC Suite",
        description="Close-coupled WC suite, white",
        category_id=category_id,
        base_uom_id=uom_id,
        list_price=Decimal(list_price),
    )
    db.add(row)
    db.flush()
    return row


def _party(db, company_id: str, *, address: str, phone: str) -> ProjectParty:
    row = ProjectParty(
        id=_uid(),
        company_id=company_id,
        party_type="developer",
        name=f"{MARKER} Nadi Cergas {_uid()[:6]}",
        address=address,
        phone=phone,
    )
    db.add(row)
    db.flush()
    return row


def _numbering_rule(db, company_id: str) -> DocumentNumberingRule:
    """Seed the `project_quotation` rule AC-A0 requires rather than borrowing one.

    CI's database is empty, so a test that assumed the module seed had already created this
    rule would pass only on a developer's machine. The prefix carries the marker so even the
    customer-facing document number is recognisable as test data.
    """
    scoped = hasattr(DocumentNumberingRule, "company_id")

    query = db.query(DocumentNumberingRule).filter(
        DocumentNumberingRule.doc_type == "project_quotation"
    )
    if scoped:
        query = query.filter(DocumentNumberingRule.company_id == company_id)
    rule = query.first()

    if rule is None:
        rule = DocumentNumberingRule(id=_uid(), doc_type="project_quotation")
        if scoped:
            rule.company_id = company_id
        db.add(rule)

    rule.enabled = True
    rule.prefix_template = f"{MARKER}/Q/"
    rule.number_digits = 4
    rule.next_value = 1
    rule.start_value = 1
    rule.reset_policy = "none"
    rule.last_reset_key = None
    db.flush()
    return rule


def _client(db, user_id: str):
    from fastapi.testclient import TestClient

    from app.database import get_db
    from app.dependencies import get_current_user, get_current_user_or_api_key
    from app.main import app
    from app.services.company_scope_resolver import apply_company_scope
    from app.services.user_service import UserPermissionService

    actor = {"id": user_id, "email": f"{user_id}@zzt.test", "role": "superadmin"}
    app.dependency_overrides[get_db] = lambda: db
    app.dependency_overrides[get_current_user] = lambda: dict(actor)
    app.dependency_overrides[get_current_user_or_api_key] = lambda: dict(actor)
    # See tests/test_project_quotation_routes.py: the router-level resolver would otherwise
    # re-stamp the scope from a request that carries no active company.
    app.dependency_overrides[apply_company_scope] = lambda: None

    originals = (
        UserPermissionService.check_user_has_permission,
        UserPermissionService.get_user_permission_slugs,
    )
    UserPermissionService.check_user_has_permission = lambda self, uid, slug: True
    UserPermissionService.get_user_permission_slugs = lambda self, uid: list(ALL_SLUGS)
    return TestClient(app), originals


def _restore(originals) -> None:
    from app.main import app
    from app.services.user_service import UserPermissionService

    UserPermissionService.check_user_has_permission = originals[0]
    UserPermissionService.get_user_permission_slugs = originals[1]
    app.dependency_overrides.clear()


@contextmanager
def _without_permission(slug: str):
    """Run the block as a user who holds every project permission EXCEPT ``slug``.

    Patched on the class, the same way the fixture grants them, so the route's
    ``require_permission`` dependency and the service's in-body ownership check both see
    the same reduced grant. Restored on exit whatever the block does.
    """
    from app.services.user_service import UserPermissionService

    granted = [s for s in ALL_SLUGS if s != slug]
    original_check = UserPermissionService.check_user_has_permission
    original_slugs = UserPermissionService.get_user_permission_slugs
    UserPermissionService.check_user_has_permission = (
        lambda self, uid, wanted, _denied=slug: wanted != _denied
    )
    UserPermissionService.get_user_permission_slugs = lambda self, uid: list(granted)
    try:
        yield
    finally:
        UserPermissionService.check_user_has_permission = original_check
        UserPermissionService.get_user_permission_slugs = original_slugs


@pytest.fixture()
def api():
    """A client, a seeded project with a developer party, and a numbering rule.

    The party exists because AC-A3's snapshot is the whole point of the create route: a
    project with no developer would let a document be created with an empty recipient block
    and prove nothing.
    """
    from app.models.base import company_scope
    from app.services.project_service import register_project

    with blank_session() as db:
        company_id = _sorento(db)
        project_seed_service.run(db, company_id=company_id)
        _numbering_rule(db, company_id)
        user_id = _user(db, f"{MARKER} Baser")
        party = _party(
            db,
            company_id,
            address=f"{MARKER} Level 8, Menara Lama, Kuala Lumpur",
            phone="03-1111 1111",
        )
        project = register_project(
            db,
            company_id=company_id,
            actor_user_id=user_id,
            developer_party_id=party.id,
            title=f"{MARKER} Cabana Elmina {_uid()[:12]}",
        )
        db.commit()
        client, originals = _client(db, user_id)
        try:
            with company_scope(db, frozenset({company_id})):
                yield client, db, company_id, user_id, project, party
        finally:
            _restore(originals)


# --------------------------------------------------------------------- helpers


def _create_document(client, project_id: str, **body) -> dict:
    response = client.post(
        f"{BASE}/projects/{project_id}/quotation-documents", json=body or {}
    )
    assert response.status_code == 201, response.text
    return response.json()


def _add_scope(client, project_id: str, document_id: str, label: str) -> dict:
    response = client.post(
        f"{BASE}/projects/{project_id}/quotation-documents/{document_id}/scopes",
        json={"scope_label": label},
    )
    assert response.status_code == 201, response.text
    return response.json()


def _current_version_id(db, quotation_id: str) -> str:
    from app.services import project_quotation_service as quotes

    return quotes.current_version(db, quotation_id).id


def _document_row(db, document_id: str):
    """The ORM row behind a document the client created over HTTP.

    Needed where the CUSTOMER's own act is being simulated: they reach the issue through a public
    token route that has its own suite, and re-walking that here would be testing the token
    rather than what the salesperson's screens are handed back.
    """
    from app.models.projects import ProjectQuotationDocument

    return (
        db.query(ProjectQuotationDocument)
        .filter(ProjectQuotationDocument.id == document_id)
        .first()
    )


def _add_priced_line(client, version_id: str, product, *, price: str, qty: str) -> dict:
    response = client.post(
        f"{BASE}/quotation-versions/{version_id}/lines",
        json={"product_id": product.id, "unit_price": price, "quantity": qty},
    )
    assert response.status_code == 201, response.text
    return response.json()


def _add_rate_only_line(db, quotation_id: str, product, actor_user_id: str, *, price: str):
    """A quoted alternate with a rate and no money, added through the service.

    Deliberately not through the line route: ``is_rate_only`` is a Group C field on the LINE
    schema and this slice builds the DOCUMENT routes, so requiring it on the wire here would
    make a document test fail for a reason that belongs to another slice. What is under test
    is the total the document route reports, and this puts the line in front of it.
    """
    from app.services import project_quotation_service as quotes

    line = quotes.upsert_line(
        db,
        version=quotes.current_version(db, quotation_id),
        actor_user_id=actor_user_id,
        payload={
            "product_id": product.id,
            "unit_price": price,
            "quantity": SAMPLE_QTY,
            "is_rate_only": True,
        },
    )
    db.commit()
    return line


def _money(value) -> Decimal:
    """Compare money exactly, however the serializer chose to put it on the wire."""
    return Decimal(str(value))


# ---------------------------------------------------------- AC-A1 / AC-A4 / AC-D2


def test_a_document_reads_back_with_its_number_recipient_and_a_total_that_ignores_rate_only(
    api,
):
    """This is the screen a salesperson checks before pressing Issue, so every fact on it is
    a fact the customer is about to be sent. The number identifies the paper, the recipient
    block is what gets posted to, and the grand total is the figure that gets paid - and the
    figure is the one that can be silently wrong, because a rate-only alternate looks exactly
    like a priced line until you add them up."""
    client, db, _company_id, user_id, project, party = api
    uom = _uom(db)
    category = _category(db, "Sanitary Ware")
    product = _product(db, category.id, uom, "300.00")
    db.commit()

    document = _create_document(
        client,
        project.id,
        your_ref=f"{MARKER}/NC/2026/018",
        attn_name="Kelly",
        subject_title=f"{MARKER} CADANGAN MEMBINA PANGSAPURI RUMAH IDAM",
        doc_date="2026-02-26",
    )

    assert document["document_no"] == f"{MARKER}/Q/0001"
    assert document["your_ref"] == f"{MARKER}/NC/2026/018"
    assert document["attn_name"] == "Kelly"
    assert document["subject_title"] == f"{MARKER} CADANGAN MEMBINA PANGSAPURI RUMAH IDAM"
    assert document["doc_date"] == "2026-02-26"
    # AC-A3: snapshotted onto the document, not read live off the party.
    assert document["recipient_name_snapshot"] == party.name
    assert document["recipient_address_snapshot"] == (
        f"{MARKER} Level 8, Menara Lama, Kuala Lumpur"
    )
    assert document["recipient_phone_snapshot"] == "03-1111 1111"

    townhouse = _add_scope(client, project.id, document["id"], f"{MARKER} Townhouse")
    guard_house = _add_scope(client, project.id, document["id"], f"{MARKER} Guard House")

    priced = _add_priced_line(
        client,
        _current_version_id(db, townhouse["id"]),
        product,
        price=PRICED_RATE,
        qty=SAMPLE_QTY,
    )
    assert _money(priced["line_total"]) == TOWNHOUSE_TOTAL
    _add_rate_only_line(db, townhouse["id"], product, user_id, price=RATE_ONLY_RATE)

    _add_priced_line(
        client,
        _current_version_id(db, guard_house["id"]),
        product,
        price="400.00",
        qty="4",
    )

    detail = client.get(
        f"{BASE}/projects/{project.id}/quotation-documents/{document['id']}"
    )
    assert detail.status_code == 200, detail.text
    body = detail.json()

    # AC-A4: the scopes come back in tab order, which is the order the PDF prints its bands.
    assert [scope["scope_label"] for scope in body["scopes"]] == [
        f"{MARKER} Townhouse",
        f"{MARKER} Guard House",
    ]
    assert _money(body["grand_total"]) == CORRECT_GRAND_TOTAL
    assert _money(body["grand_total"]) != WRONG_GRAND_TOTAL_IF_RATE_ONLY_COUNTED

    listed = client.get(f"{BASE}/projects/{project.id}/quotation-documents")
    assert listed.status_code == 200, listed.text
    envelope = listed.json()
    assert [row["id"] for row in envelope["data"]] == [document["id"]]
    assert envelope["empty"] is False
    assert envelope["pagination"]["total"] == 1
    assert _money(envelope["data"][0]["grand_total"]) == CORRECT_GRAND_TOTAL


# --------------------------------------------------------------------- AC-A2


def test_a_new_document_arrives_already_filled_in_from_what_the_system_knows(api):
    """Journey step 2: the only thing asked for is what is not already known. A blank
    recipient block and a blank subject would make every quotation a re-typing exercise
    against data the project already holds, and re-typed addresses are how a quotation gets
    posted to the wrong office."""
    client, _db, _company_id, _user_id, project, party = api

    document = _create_document(client, project.id)

    assert document["subject_title"] == project.title
    assert document["doc_date"] == date.today().isoformat()
    assert document["recipient_party_id"] == party.id
    assert document["recipient_name_snapshot"] == party.name
    assert document["recipient_address_snapshot"] == (
        f"{MARKER} Level 8, Menara Lama, Kuala Lumpur"
    )
    # Nothing is invented: what the system cannot derive stays empty rather than guessed.
    assert document["your_ref"] is None
    assert document["attn_name"] is None


def test_the_header_can_be_corrected_without_starting_the_quotation_again(api):
    """`Attn:` is wrong, or the customer sends their own reference after the document was
    opened. Without an edit route the salesperson deletes and rebuilds, which burns a
    customer-facing running number and loses the scopes already priced."""
    client, _db, _company_id, _user_id, project, _party = api

    document = _create_document(client, project.id)

    updated = client.patch(
        f"{BASE}/projects/{project.id}/quotation-documents/{document['id']}",
        json={
            "attn_name": "Kelly Tan",
            "your_ref": f"{MARKER}/NC/2026/021",
            "subject_title": f"{MARKER} CADANGAN MEMBINA PANGSAPURI",
        },
    )
    assert updated.status_code == 200, updated.text
    body = updated.json()
    assert body["attn_name"] == "Kelly Tan"
    assert body["your_ref"] == f"{MARKER}/NC/2026/021"
    assert body["subject_title"] == f"{MARKER} CADANGAN MEMBINA PANGSAPURI"
    # The running number is claimed once and never re-drawn (AC-A0).
    assert body["document_no"] == document["document_no"]

    reread = client.get(
        f"{BASE}/projects/{project.id}/quotation-documents/{document['id']}"
    ).json()
    assert reread["attn_name"] == "Kelly Tan"


def test_a_scope_can_be_renamed_and_reordered_so_the_bands_print_in_walkthrough_order(api):
    """The salesperson walks the development in an order the customer recognises, and the
    printed bands have to match it or the QS cannot line the quotation up against their bill.
    Renaming matters for the same reason: "Block A" turns out to be "Guard House" only after
    someone has already priced it."""
    client, _db, _company_id, _user_id, project, _party = api

    document = _create_document(client, project.id)
    townhouse = _add_scope(client, project.id, document["id"], f"{MARKER} Townhouse")
    guard_house = _add_scope(client, project.id, document["id"], f"{MARKER} Block A")

    renamed = client.patch(
        f"{BASE}/projects/{project.id}/quotation-documents/{document['id']}"
        f"/scopes/{guard_house['id']}",
        json={"scope_label": f"{MARKER} Guard House"},
    )
    assert renamed.status_code == 200, renamed.text
    assert renamed.json()["scope_label"] == f"{MARKER} Guard House"

    reordered = client.patch(
        f"{BASE}/projects/{project.id}/quotation-documents/{document['id']}"
        f"/scopes/{townhouse['id']}",
        json={"sort_order": 5},
    )
    assert reordered.status_code == 200, reordered.text

    body = client.get(
        f"{BASE}/projects/{project.id}/quotation-documents/{document['id']}"
    ).json()
    assert [scope["scope_label"] for scope in body["scopes"]] == [
        f"{MARKER} Guard House",
        f"{MARKER} Townhouse",
    ]


# ----------------------------------------------------------------------- RBAC


def test_someone_who_may_not_edit_the_project_cannot_change_its_quotation(api):
    """A quotation is a price commitment, and the read grant is handed out widely - support,
    coordination, anyone who needs to see what was quoted. If the write routes only checked
    that a user was logged in, every reader would be able to re-price a document that is
    about to be sent. Checked on EVERY mutating route, because a single unguarded one is the
    whole hole."""
    client, _db, _company_id, _user_id, project, _party = api

    document = _create_document(client, project.id)
    scope = _add_scope(client, project.id, document["id"], f"{MARKER} Townhouse")
    root = f"{BASE}/projects/{project.id}/quotation-documents"

    with _without_permission(EDIT):
        denials = {
            "create document": client.post(root, json={}),
            "edit document": client.patch(
                f"{root}/{document['id']}", json={"attn_name": "Nobody"}
            ),
            "add scope": client.post(
                f"{root}/{document['id']}/scopes",
                json={"scope_label": f"{MARKER} Sneaked In"},
            ),
            "edit scope": client.patch(
                f"{root}/{document['id']}/scopes/{scope['id']}",
                json={"scope_label": f"{MARKER} Renamed By Nobody"},
            ),
            # No signing here on purpose: this caller has no edit permission, so the 403 lands
            # before the signature gate is ever consulted.
            "issue": client.post(f"{root}/{document['id']}/issue"),
        }
    for what, response in denials.items():
        assert response.status_code == 403, f"{what} was allowed: {response.text}"

    with _without_permission(DELETE):
        refused = client.delete(f"{root}/{document['id']}")
    assert refused.status_code == 403, refused.text

    # The refusals refused: nothing was written on the way to the 403.
    survivor = client.get(f"{root}/{document['id']}").json()
    assert survivor["attn_name"] is None
    assert [s["scope_label"] for s in survivor["scopes"]] == [f"{MARKER} Townhouse"]


# ----------------------------------------------------------------- validation


def test_a_scope_with_no_real_label_is_refused_before_it_reaches_a_tab(api):
    """A scope IS its label - it is the tab on screen and the band heading in the printed
    document. A blank or space-only one produces an unnamed section in a document a customer
    receives, and nothing downstream can repair it because there was never anything to
    repair."""
    client, _db, _company_id, _user_id, project, _party = api

    document = _create_document(client, project.id)
    root = f"{BASE}/projects/{project.id}/quotation-documents/{document['id']}/scopes"

    for label in ("", "   ", "\t\n"):
        refused = client.post(root, json={"scope_label": label})
        assert refused.status_code == 422, f"{label!r} was accepted: {refused.text}"

    missing = client.post(root, json={})
    assert missing.status_code == 422, missing.text


def test_a_malformed_document_id_is_answered_not_crashed(api):
    """A truncated or hand-typed id arrives from a stale bookmark and from every crawler.
    Reaching the database with it raises a driver error the route turns into a 500, which
    pages someone at night and tells the client nothing it can act on. The honest answer to
    "this id cannot exist" is the same one given for "this id does not exist"."""
    client, _db, _company_id, _user_id, project, _party = api

    root = f"{BASE}/projects/{project.id}/quotation-documents"
    for bad in ("not-a-uuid", "123", "00000000-0000-0000-0000"):
        for response in (
            client.get(f"{root}/{bad}"),
            client.patch(f"{root}/{bad}", json={"attn_name": "x"}),
            client.delete(f"{root}/{bad}"),
            client.post(f"{root}/{bad}/scopes", json={"scope_label": f"{MARKER} X"}),
        ):
            assert 400 <= response.status_code < 500, (
                f"{bad!r} produced {response.status_code}: {response.text}"
            )


# ---------------------------------------------------------------------- 404s


def test_a_document_belonging_to_another_project_is_not_reachable_through_this_one(api):
    """Nesting the document under a project is a claim, not decoration: the URL says this
    document is part of this pursuit. If the handler looked the document up by id alone, the
    nesting would be a lie and anybody holding one project's id could read another project's
    prices through it - a leak that no screen ever shows."""
    client, db, company_id, user_id, project, _party = api
    from app.services.project_service import register_project

    other_project = register_project(
        db,
        company_id=company_id,
        actor_user_id=user_id,
        developer_party_id=None,
        title=f"{MARKER} Seri Tanjung {_uid()[:12]}",
    )
    db.commit()

    document = _create_document(client, project.id)

    stolen = client.get(
        f"{BASE}/projects/{other_project.id}/quotation-documents/{document['id']}"
    )
    assert stolen.status_code == 404, stolen.text

    unknown = client.get(
        f"{BASE}/projects/{project.id}/quotation-documents/{_uid()}"
    )
    assert unknown.status_code == 404, unknown.text

    # The other project's own list stays empty rather than borrowing this one's document.
    listed = client.get(f"{BASE}/projects/{other_project.id}/quotation-documents").json()
    assert listed["data"] == []
    assert listed["empty"] is True


# --------------------------------------------------------------------- AC-A6


def test_an_issued_document_refuses_deletion_and_a_draft_takes_its_scopes_with_it(api):
    """Deleting an issued document deletes the record of what a customer is holding in their
    inbox, and the refusal has to be visible: a salesperson who believes it is gone will
    re-quote the same scope under a new number. A draft nobody has seen is the opposite case
  - it deletes cleanly, and its scopes go with it, because a scope has no meaning without
    the document that carries it."""
    from app.models.projects import ProjectQuotation

    client, db, _company_id, _user_id, project, _party = api
    root = f"{BASE}/projects/{project.id}/quotation-documents"

    issued = _create_document(client, project.id)
    _add_scope(client, project.id, issued["id"], f"{MARKER} Townhouse")
    _sign(client, root, issued["id"])
    assert client.post(f"{root}/{issued['id']}/issue").status_code == 201

    refused = client.delete(f"{root}/{issued['id']}")
    assert refused.status_code == 422, refused.text
    assert "withdraw" in refused.json()["message"].lower()
    assert client.get(f"{root}/{issued['id']}").status_code == 200, "the refusal deleted it"

    draft = _create_document(client, project.id)
    scope = _add_scope(client, project.id, draft["id"], f"{MARKER} Reception")

    deleted = client.delete(f"{root}/{draft['id']}")
    assert deleted.status_code == 204, deleted.text
    assert client.get(f"{root}/{draft['id']}").status_code == 404

    db.expire_all()
    assert (
        db.query(ProjectQuotation).filter(ProjectQuotation.id == scope["id"]).first() is None
    ), "the document went but its scope was orphaned"


# --------------------------------------------------------------------- AC-B2


def test_reissuing_stamps_r1_then_r2_which_is_what_the_customer_quotes_back(api):
    """`Our Ref ... (R2)` is the string a customer reads down the phone. It has to advance on
    every issue and identify exactly one thing: if a second issue could reuse R1, the
    reference stops naming the paper they are holding and no conversation about it can be
    settled. The grand total is frozen onto each issue for the same reason - what was sent
    stays readable after the live document moves on."""
    client, db, _company_id, user_id, project, _party = api
    uom = _uom(db)
    category = _category(db, "Sanitary Ware")
    product = _product(db, category.id, uom, "300.00")
    db.commit()

    root = f"{BASE}/projects/{project.id}/quotation-documents"
    document = _create_document(client, project.id)
    scope = _add_scope(client, project.id, document["id"], f"{MARKER} Townhouse")
    _add_priced_line(
        client,
        _current_version_id(db, scope["id"]),
        product,
        price=PRICED_RATE,
        qty=SAMPLE_QTY,
    )

    _sign(client, root, document["id"])
    first = client.post(f"{root}/{document['id']}/issue")
    assert first.status_code == 201, first.text
    r1 = first.json()
    assert r1["issue_no"] == 1
    assert r1["our_ref_text"] == f"{document['document_no']} (R1)"
    assert _money(r1["grand_total"]) == TOWNHOUSE_TOTAL

    # A revision re-prices the scope, then goes out as R2.
    revised = client.post(f"{BASE}/quotations/{scope['id']}/revise")
    assert revised.status_code == 201, revised.text
    carried = client.get(
        f"{BASE}/quotation-versions/{revised.json()['id']}/lines"
    ).json()["data"][0]
    client.put(
        f"{BASE}/quotation-versions/{revised.json()['id']}/lines/{carried['id']}",
        json={"unit_price": "230.00"},
    )

    _sign(client, root, document["id"])
    second = client.post(f"{root}/{document['id']}/issue")
    assert second.status_code == 201, second.text
    r2 = second.json()
    assert r2["issue_no"] == 2
    assert r2["our_ref_text"] == f"{document['document_no']} (R2)"
    assert _money(r2["grand_total"]) == Decimal("240580.00")  # 1046 x 230

    history = client.get(f"{root}/{document['id']}/issues")
    assert history.status_code == 200, history.text
    envelope = history.json()
    assert [row["issue_no"] for row in envelope["data"]] == [2, 1]
    assert envelope["pagination"]["total"] == 2
    # R1 still reads as it was sent, at the price it was sent at.
    assert _money(envelope["data"][1]["grand_total"]) == TOWNHOUSE_TOTAL


def test_the_customers_answer_reaches_the_wire_on_the_document_and_in_the_list(api):
    """S17. The salesperson's two screens read this off the DOCUMENT, not the issue history.

    The client asked "when i request changes, how can i see it from the system?", and the answer
    is a banner on the quotation plus a badge on the project's quotation list - neither of which
    fetches issues. So the decision has to survive the response model, which is exactly where
    this codebase has lost a new field before: a manual dict computing it perfectly and a schema
    that never declared it drops it silently on the way out.

    Three things at the wire: the words arrive, acceptance outranks a request, and re-issuing
    clears it (the revision the customer holds is the one being reported).
    """
    from app.services import project_quotation_document_service as qdocs

    client, db, _company_id, _user_id, project, _party = api
    uom = _uom(db)
    category = _category(db, "Sanitary Ware")
    product = _product(db, category.id, uom, "300.00")
    db.commit()

    root = f"{BASE}/projects/{project.id}/quotation-documents"
    document = _create_document(client, project.id)
    scope = _add_scope(client, project.id, document["id"], f"{MARKER} Townhouse")
    _add_priced_line(
        client,
        _current_version_id(db, scope["id"]),
        product,
        price=PRICED_RATE,
        qty=SAMPLE_QTY,
    )
    _sign(client, root, document["id"])
    issued = client.post(f"{root}/{document['id']}/issue")
    assert issued.status_code == 201, issued.text

    # Nothing asked yet: the keys are present and empty, never absent.
    quiet = client.get(f"{root}/{document['id']}").json()
    assert quiet["customer_decision"] is None
    assert quiet["changes_requested_note"] is None
    assert quiet["accepted_at"] is None

    record = qdocs.current_issue(db, _document_row(db, document["id"]))
    note = f"{MARKER} can you provide me more discount"
    qdocs.request_changes(db, record=record, note=note, requester_name=f"{MARKER} Kelly")
    db.commit()

    asked = client.get(f"{root}/{document['id']}").json()
    assert asked["customer_decision"] == "changes_requested"
    assert asked["changes_requested_note"] == note
    assert asked["changes_requested_by_name"] == f"{MARKER} Kelly"
    assert asked["changes_requested_at"] is not None
    # And on the list the project tab renders, which is the other screen that has to say it.
    listed = client.get(root).json()["data"][0]
    assert listed["customer_decision"] == "changes_requested"

    qdocs.accept_issue(
        db,
        record=record,
        signer_name=f"{MARKER} Kelly",
        mode="draw",
        image_data_uri="data:image/png;base64,zzt",
    )
    db.commit()

    signed = client.get(f"{root}/{document['id']}").json()
    assert signed["customer_decision"] == "accepted"
    assert signed["accepted_at"] is not None
    # The words stay on record - they happened - but they are no longer the standing answer.
    assert signed["changes_requested_note"] == note

    # A fresh revision resets it: R2 is what the customer now holds, and nobody has answered it.
    _sign(client, root, document["id"])
    reissued = client.post(f"{root}/{document['id']}/issue")
    assert reissued.status_code == 201, reissued.text
    fresh = client.get(f"{root}/{document['id']}").json()
    assert fresh["current_issue_no"] == 2
    assert fresh["customer_decision"] is None
    assert fresh["accepted_at"] is None
    assert fresh["changes_requested_note"] is None


# --------------------------------------------------------------------- AC-B4


def test_the_line_endpoint_refuses_to_edit_a_version_the_customer_already_holds(api):
    """The freeze is DERIVED from the issue rows, and the line routes belong to a different
    router whose author never thought about documents. So the rule has to hold where the edit
    actually arrives: without this, issuing R1 and then editing a price rewrites the very rows
    R1 claims to contain, and the PDF on file quietly stops matching the record behind it. The
    code is asserted, not just the status, because the client renders "open a revision" off
    it and a generic 422 gives the user nothing to do."""
    client, db, _company_id, _user_id, project, _party = api
    uom = _uom(db)
    category = _category(db, "Sanitary Ware")
    product = _product(db, category.id, uom, "300.00")
    db.commit()

    root = f"{BASE}/projects/{project.id}/quotation-documents"
    document = _create_document(client, project.id)
    scope = _add_scope(client, project.id, document["id"], f"{MARKER} Townhouse")
    version_id = _current_version_id(db, scope["id"])
    line = _add_priced_line(client, version_id, product, price=PRICED_RATE, qty="4")

    _sign(client, root, document["id"])
    assert client.post(f"{root}/{document['id']}/issue").status_code == 201

    edited = client.put(
        f"{BASE}/quotation-versions/{version_id}/lines/{line['id']}",
        json={"unit_price": "230.00"},
    )
    assert edited.status_code == 422, edited.text
    assert edited.json()["code"] == "quotation_version_issued"

    added = client.post(
        f"{BASE}/quotation-versions/{version_id}/lines",
        json={"product_id": product.id, "unit_price": "180.00", "quantity": "1"},
    )
    assert added.status_code == 422, added.text
    assert added.json()["code"] == "quotation_version_issued"

    # Nothing half-landed: the issue still reports the money it was stamped with.
    still = client.get(f"{root}/{document['id']}/issues").json()["data"][0]
    assert _money(still["grand_total"]) == Decimal("1000.00")  # 4 x 250
    lines = client.get(f"{BASE}/quotation-versions/{version_id}/lines").json()["data"]
    assert len(lines) == 1
    assert _money(lines[0]["unit_price"]) == Decimal("250.00")


# --------------------------------------------------------------------- AC-G4 (PDF)


def _issue_one(client, db, project, *, price: str = PRICED_RATE, qty: str = "4") -> tuple:
    """A signed, issued document with one priced line. Returns ``(root, document, issue)``."""
    uom = _uom(db)
    category = _category(db, "Sanitary Ware")
    product = _product(db, category.id, uom, "300.00")
    db.commit()

    root = f"{BASE}/projects/{project.id}/quotation-documents"
    document = _create_document(client, project.id)
    scope = _add_scope(client, project.id, document["id"], f"{MARKER} Townhouse")
    _add_priced_line(
        client, _current_version_id(db, scope["id"]), product, price=price, qty=qty
    )
    _sign(client, root, document["id"])
    issued = client.post(f"{root}/{document['id']}/issue")
    assert issued.status_code == 201, issued.text
    return root, document, issued.json()


def test_the_pdf_download_returns_a_real_pdf_named_after_the_reference(api):
    """This is the file that goes in the customer's inbox, so two things have to hold on the
    wire: the bytes are a PDF a browser will open inline rather than a JSON error rendered as
    a download, and the filename carries `Our Ref (R1)`. A saved attachment named
    `download.pdf` cannot be matched back to the paper it is, which is the whole reason the
    reference exists."""
    client, db, _company_id, _user_id, project, _party = api
    root, document, issue = _issue_one(client, db, project)

    response = client.get(f"{root}/{document['id']}/issues/{issue['id']}/pdf")
    if response.status_code == 503:
        pytest.skip(f"WeasyPrint unavailable on this host: {response.text}")

    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith("application/pdf")
    # A PDF, not a 200 carrying an error page. The magic bytes are the only honest check.
    assert response.content[:5] == b"%PDF-"
    expected = f"quotation-{MARKER}-Q-0001-R1.pdf"
    assert response.headers["content-disposition"] == (
        f'inline; filename="{expected}"; filename*=UTF-8\'\'{quote(expected)}'
    )


def test_a_missing_rendering_library_is_a_503_that_names_itself_not_a_500(api, monkeypatch):
    """Prod hosts need cairo/pango installed and CI hosts may not have them. Left as a bare
    500 this looks like a bug in the quotation and sends somebody reading application code;
    as a 503 with `pdf_rendering_unavailable` it sends them to the host. The FE also renders
    off the code, so it is asserted rather than just the status."""
    from app.services import project_quotation_pdf_service as pdf_service
    from app.services.complaint_pdf_service import PDFRenderingUnavailable

    client, db, _company_id, _user_id, project, _party = api
    root, document, issue = _issue_one(client, db, project)

    def _explode(_db, _issue):
        raise PDFRenderingUnavailable("libpango not found")

    monkeypatch.setattr(pdf_service, "render_issue_pdf", _explode)

    response = client.get(f"{root}/{document['id']}/issues/{issue['id']}/pdf")
    assert response.status_code == 503, response.text
    body = response.json()
    assert body["code"] == "pdf_rendering_unavailable"
    assert "libpango not found" in body["message"]


def test_a_revision_id_from_another_document_cannot_be_downloaded_through_this_one(api):
    """The URL names a project, a document AND an issue, so the issue has to be checked
    against the document rather than merely fetched by id. Without it, a known issue id
    downloads through any document the caller may view, and the price list of a quotation
    they were never shown lands as a PDF."""
    client, db, _company_id, _user_id, project, _party = api
    root, document, _issue = _issue_one(client, db, project)

    # A second document on the same project: same permissions, different paper.
    other = _create_document(client, project.id)

    stranger = client.get(f"{root}/{other['id']}/issues/{_issue['id']}/pdf")
    assert stranger.status_code == 404, stranger.text
    assert stranger.json()["code"] == "quotation_issue_not_found"

    unknown = client.get(f"{root}/{document['id']}/issues/{_uid()}/pdf")
    assert unknown.status_code == 404, unknown.text

    malformed = client.get(f"{root}/{document['id']}/issues/not-a-uuid/pdf")
    assert malformed.status_code in (400, 404, 422), malformed.text


def test_someone_who_may_not_view_the_project_cannot_download_its_quotation(api):
    """A quotation PDF is the full price list. Read permission on the project is the gate, and
    it has to be enforced at the route: the renderer takes an issue row and would happily
    render one for anybody who reached it."""
    client, db, _company_id, _user_id, project, _party = api
    root, document, issue = _issue_one(client, db, project)

    with _without_permission(VIEW):
        denied = client.get(f"{root}/{document['id']}/issues/{issue['id']}/pdf")
    assert denied.status_code in (401, 403), denied.text


# --------------------------------------------------------------------- AC-F2 (Excel)


def test_the_excel_download_returns_a_workbook_that_downloads_rather_than_previews(api):
    """The Excel export exists because the customer's QS re-prices in Excel (AC-F2), so what
    lands on the wire has to be a workbook their spreadsheet opens, not a JSON error served with
    a spreadsheet content type. Sent as an ATTACHMENT rather than inline: nobody previews a
    workbook in a browser tab, and an inline disposition on a binary a browser cannot render is
    how a download turns into a blank page. The filename carries `Our Ref (R1)` for the same
    reason the PDF's does - a saved `download.xlsx` cannot be matched back to the paper it is."""
    from io import BytesIO

    from openpyxl import load_workbook

    client, db, _company_id, _user_id, project, _party = api
    root, document, issue = _issue_one(client, db, project)

    response = client.get(f"{root}/{document['id']}/issues/{issue['id']}/xlsx")
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    expected = f"quotation-{MARKER}-Q-0001-R1.xlsx"
    assert response.headers["content-disposition"] == (
        f'attachment; filename="{expected}"; filename*=UTF-8\'\'{quote(expected)}'
    )

    # A real workbook, not a 200 carrying an error page. One sheet per scope, with the money on
    # it, so the route is proven to be wired to the exporter rather than to an empty stub.
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == [f"{MARKER} Townhouse"]
    assert Decimal("1000.00") in [  # 4 x 250
        _money(cell.value)
        for row in workbook.worksheets[0].iter_rows()
        for cell in row
        if isinstance(cell.value, (int, float, Decimal))
    ]


def test_a_revision_id_from_another_document_cannot_be_exported_through_this_one(api):
    """Same leak as the PDF route, and it has to be closed separately because it is a separate
    handler: the URL names a project, a document AND an issue, so the issue is checked against
    the document rather than merely fetched by id. Without it a known issue id exports through
    any document the caller may view, and a price list they were never shown lands as a
    spreadsheet - which is the more useful artifact to leak, since it is already machine
    readable."""
    client, db, _company_id, _user_id, project, _party = api
    root, document, issue = _issue_one(client, db, project)

    other = _create_document(client, project.id)

    stranger = client.get(f"{root}/{other['id']}/issues/{issue['id']}/xlsx")
    assert stranger.status_code == 404, stranger.text
    assert stranger.json()["code"] == "quotation_issue_not_found"

    unknown = client.get(f"{root}/{document['id']}/issues/{_uid()}/xlsx")
    assert unknown.status_code == 404, unknown.text

    malformed = client.get(f"{root}/{document['id']}/issues/not-a-uuid/xlsx")
    assert malformed.status_code in (400, 404, 422), malformed.text


def test_someone_who_may_not_view_the_project_cannot_export_its_quotation(api):
    """The workbook is the full price list in a form that can be re-sent, re-priced and pasted
    into somebody else's tender. Read permission on the project is the gate and it is enforced at
    the route, because the exporter takes an issue row and would build a workbook for anybody who
    reached it."""
    client, db, _company_id, _user_id, project, _party = api
    root, document, issue = _issue_one(client, db, project)

    with _without_permission(VIEW):
        denied = client.get(f"{root}/{document['id']}/issues/{issue['id']}/xlsx")
    assert denied.status_code in (401, 403), denied.text


# ------------------------------------------------------------------- AC-H1 (wire)


def test_the_document_tells_the_client_whether_it_has_been_signed(api):
    """The Issue CTA is gated on the signature, so the client has to be able to SEE the
    signature after a page refresh, not only in the response to its own sign call.

    Without it the client has to guess, and the only available guess - "it is issued, so it
    must have been signed" - is wrong twice. It is wrong on a document issued before the
    signature gate existed, where it enables an Issue button the server then refuses with
    `quotation_document_unsigned`. And it is wrong the other way on a signed DRAFT, where a
    refresh loses the signature and the client disables issuing on a document that is ready to
    go. Same family as the get_user rule: a column that never reaches the serializer never
    reaches the screen."""
    client, db, _company_id, _user_id, project, _party = api
    root = f"{BASE}/projects/{project.id}/quotation-documents"
    document = _create_document(client, project.id)

    before = client.get(f"{root}/{document['id']}").json()
    assert before["is_signed"] is False
    assert before["signatory_signature"] is None

    _sign(client, root, document["id"])

    after = client.get(f"{root}/{document['id']}").json()
    assert after["is_signed"] is True
    signature = after["signatory_signature"]
    assert signature is not None
    # Enough to RENDER, not just a boolean: the screen shows who signed and when beside the ink.
    assert signature["signer_name"] == f"{MARKER} Baser"
    assert signature["image_data_uri"] == "data:image/png;base64,zzt"
    assert signature["mode"] == "draw"
    assert signature["signed_at"]

    # It survives into the list, which is where a salesperson scans for what is ready to issue.
    listed = next(
        row
        for row in client.get(root).json()["data"]
        if row["id"] == document["id"]
    )
    assert listed["is_signed"] is True
