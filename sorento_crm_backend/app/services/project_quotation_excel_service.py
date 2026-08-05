"""Export ONE quotation issue to Excel, one sheet per scope (AC-F2).

The PDF sibling renders the document Sorento SENDS. This renders the artifact the customer's QS
WORKS IN: they re-price in Excel, and without this file they re-type the whole quotation to do it.
Two consequences run through the module, and both are the reason it is not simply the PDF's single
banded sheet exported:

1. **One sheet per scope**, in the document's own order. A QS re-pricing the guard house should
   not have to scroll past the townhouse, and each sheet ends with its own total so a single
   printed tab is a complete statement.
2. **Money is a NUMBER with a number format, never a pre-formatted string.** A cell holding
   ``"1,000.00"`` looks identical on screen and contributes zero to every SUM the reader writes,
   so the arithmetic they do silently disagrees with ours. Amounts are read as ``Decimal`` off the
   snapshot and written as ``Decimal``; nothing here accumulates a total in ``float``.

Everything is read through the PDF service's traversal (``_document_of`` / ``_load_scopes`` /
``_sender``) rather than re-derived from the models. That is deliberate: the two artifacts describe
the same issue, so a second loader would be a second chance to disagree about which version's lines
an issue contains or what a scope totalled. Reading the ISSUE SNAPSHOT is what makes a re-export
next year the thing that was sent (AC-F3).

A rate-only line writes the words ``rate only`` and no number at all (AC-C2 / AC-D3). In Excel that
matters more than on paper: 0.00 in that cell tells a QS the item is free AND invites its rate into
any column they sum.

**Deviation from the PDF, deliberate.** The ``PRODUCT IMAGE`` column carries the image's FILENAME,
not the picture. The column still collapses per sheet when no line on that sheet has an image
(AC-F4), which is the rule that matters for readability. Embedding would mean pulling bytes from
object storage and anchoring floating pictures over cells (an openpyxl drawing is not a cell value,
so it does not survive the sort/filter a QS immediately applies), and it would add a Pillow
dependency this backend does not declare. The workbook is a pricing artifact; the PDF remains the
one with the photographs.

Also deliberately absent: the cover letter and the terms. They are prose belonging to the document
of record, and pasting them into a spreadsheet the reader is about to re-sort adds nothing they can
act on. The PDF carries both.
"""
from __future__ import annotations

import io
import logging
import os
import re
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.models.projects import (
    ProjectQuotationDocument,
    ProjectQuotationIssue,
    ProjectQuotationLine,
)
from app.models.resources import Attachment
from app.services import project_quotation_pdf_service as pdf

logger = logging.getLogger(__name__)

# Number formats, not formatted strings. The header already says "(RM)", so only the grand total
# repeats the currency - it is the figure somebody reads out loud.
MONEY_FORMAT = "#,##0.00"
GRAND_TOTAL_FORMAT = '"RM "#,##0.00'
# `1046`, not `1046.00`: the column is stored Numeric(12,2) but quantities are counts.
QTY_FORMAT = "#,##0.##"

SCOPE_TOTAL_LABEL = "TOTAL"
GRAND_TOTAL_LABEL = "TOTAL AMOUNT"

# Excel refuses to open a workbook whose sheet title is over 31 characters or contains any of
# these, and openpyxl raises on the characters before a byte is written. Scope labels are free text
# typed off the customer's bill of quantities, so both are ordinary input here.
_TITLE_LIMIT = 31
_ILLEGAL_TITLE_CHARS = re.compile(r"[\[\]:*?/\\]")
_WHITESPACE = re.compile(r"\s+")
# Excel reserves this one name. Seeding it as "taken" sends it through the same de-duplication
# path as any other collision instead of needing a rule of its own.
_RESERVED_TITLES = frozenset({"history"})

# Kept out of the number format unless it is safe: a stray quote in the unit would produce a format
# string Excel rejects, which loses the whole workbook rather than one cell.
_UOM_SAFE = re.compile(r"[^0-9A-Za-z .%/-]")

_HEADER_FILL = PatternFill("solid", fgColor="EEEEEE")
_BAND_FILL = PatternFill("solid", fgColor="F4F4F4")
_BOLD = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=12)
_ITALIC = Font(italic=True)
_WRAP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="top")
_CENTER = Alignment(horizontal="center", vertical="top")

# Wide enough that the reader does not have to widen every column before they can read it.
_COLUMN_WIDTHS: Dict[str, int] = {
    "ITEM": 6,
    "PRODUCT IMAGE": 26,
    "TECHNICAL SPEC": 34,
    "DESCRIPTION": 42,
    "BRAND": 14,
    "PRODUCT CODE": 16,
    "QTY": 12,
    "UNIT RATE (RM)": 14,
    "COMPLETE SET": 14,
    "GRAND TOTAL": 16,
}


# --------------------------------------------------------------------- values


def _text(value: Any) -> Optional[str]:
    """Plain text, or nothing at all.

    Not the PDF's ``_cell``: that HTML-escapes, and an ampersand reaching a spreadsheet as
    ``&amp;`` is a visible defect in a document the customer reads.
    """
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _date_text(value: Any) -> Optional[str]:
    if isinstance(value, (datetime, date)):
        return value.strftime("%d/%m/%Y")
    return _text(value)


def _amount(value: Any) -> Decimal:
    """Money as a Decimal, straight off the snapshot. Never via float."""
    return Decimal(value if value is not None else 0)


def _line_amount(value: Any) -> Optional[Decimal]:
    """Money on a LINE, where zero means "no separate charge" and the cell is left empty.

    Set components are quoted on their own rows at nothing, because the price sits on the parent
    item. In a workbook a printed `0.00` is worse than on paper: the QS re-pricing this sheet
    gets a number that sums, sorts and charts as a real free product. ``None`` leaves the cell
    genuinely empty rather than writing a string, which would break their formulas instead.

    Totals keep ``_amount``: a scope that really does add up to nothing should say so.
    """
    amount = _amount(value)
    return None if amount == Decimal("0") else amount


def _qty_format(uom: Any) -> str:
    """The quantity's own format, carrying the unit as a literal suffix.

    The unit belongs beside the number the customer reads, but the sample's column set has no UOM
    column and the cell has to stay numeric - a QS multiplies it. A format suffix gives both:
    ``1046 NO`` on screen, ``1046`` to every formula.
    """
    unit = _UOM_SAFE.sub("", str(uom or "")).strip()[:12]
    return f'{QTY_FORMAT}" {unit}"' if unit else QTY_FORMAT


# ---------------------------------------------------------------- sheet titles


def _trim(value: str, limit: int) -> str:
    """Cut to ``limit`` and clean the edges. Excel also rejects a leading/trailing apostrophe."""
    return value[: max(limit, 0)].strip().strip("'").strip()


def _sheet_title(label: Any, position: int, used: Set[str]) -> str:
    """A title Excel will accept, unique in this workbook, and derived only from its inputs.

    Deterministic on purpose: a random or time-based suffix would make two exports of the same
    issue two different files, and the whole point of an issue is that it does not move.
    """
    text = _WHITESPACE.sub(" ", _ILLEGAL_TITLE_CHARS.sub(" ", str(label or ""))).strip()
    # An unlabelled scope is still a scope, and it still has to be findable.
    fallback = f"Scope {position}"
    candidate = _trim(text, _TITLE_LIMIT) or fallback

    attempt = 2
    while candidate.casefold() in used:
        suffix = f" ({attempt})"
        base = _trim(text, _TITLE_LIMIT - len(suffix)) or fallback
        candidate = f"{base}{suffix}"
        attempt += 1

    used.add(candidate.casefold())
    return candidate


# --------------------------------------------------------------------- images


def _image_names(db: Session, lines: Sequence[ProjectQuotationLine]) -> Dict[str, str]:
    """Filename per attachment id, fetched in one query for the whole scope.

    Best-effort by design: a missing attachment row leaves the cell empty rather than failing the
    export. The reader is waiting for a price, not a filename.
    """
    ids = {str(line.image_attachment_id) for line in lines if line.image_attachment_id}
    if not ids:
        return {}
    rows = db.query(Attachment).filter(Attachment.id.in_(ids)).all()
    names: Dict[str, str] = {}
    for row in rows:
        name = _text(getattr(row, "original_filename", None)) or _text(
            getattr(row, "stored_filename", None)
        )
        if name:
            names[str(row.id)] = name
    return names


# -------------------------------------------------------------------- writing


def _put(
    sheet: Worksheet,
    row: int,
    column: int,
    value: Any,
    *,
    font: Optional[Font] = None,
    number_format: Optional[str] = None,
    alignment: Optional[Alignment] = None,
    fill: Optional[PatternFill] = None,
):
    cell = sheet.cell(row=row, column=column, value=value)
    if font is not None:
        cell.font = font
    if number_format is not None:
        cell.number_format = number_format
    if alignment is not None:
        cell.alignment = alignment
    if fill is not None:
        cell.fill = fill
    return cell


def _write_letterhead(
    sheet: Worksheet,
    document: ProjectQuotationDocument,
    issue: ProjectQuotationIssue,
    sender: Dict[str, Any],
    row: int,
) -> int:
    """The first sheet identifies the paper the way the PDF's first page does.

    Only the first sheet carries it: repeated on every tab it is noise, and the later sheets get a
    light header instead so a single printed page is still identifiable.
    """
    name = _text(sender.get("name"))
    if name:
        _put(sheet, row, 1, name, font=_TITLE_FONT)
        row += 1
    for line in sender.get("address_lines") or []:
        _put(sheet, row, 1, _text(line))
        row += 1
    phone = _text(sender.get("phone"))
    if phone:
        _put(sheet, row, 1, f"Tel: {phone}")
        row += 1

    row += 1
    # `our_ref_text` already carries the revision suffix, e.g. `SRT/Q/0141 (R1)`. That IS the
    # reference the customer quotes back, so it is written whole rather than rebuilt here.
    for label, value in (
        ("Our Ref", _text(issue.our_ref_text)),
        ("Your Ref", _text(document.your_ref)),
        ("Date", _date_text(document.doc_date or issue.issued_at)),
    ):
        if value is None:
            continue
        _put(sheet, row, 1, label, font=_BOLD)
        _put(sheet, row, 2, value)
        row += 1

    row += 1
    _put(sheet, row, 1, "To:", font=_BOLD)
    row += 1
    recipient: List[Optional[str]] = [_text(document.recipient_name_snapshot)]
    recipient += [_text(line) for line in pdf._lines_of(document.recipient_address_snapshot)]
    recipient += [_text(line) for line in pdf._lines_of(document.recipient_phone_snapshot)]
    for line in recipient:
        if line is None:
            continue
        _put(sheet, row, 1, line)
        row += 1
    attn = _text(document.attn_name)
    if attn:
        _put(sheet, row, 1, f"Attn: {attn}", font=_BOLD)
        row += 1

    subject = _text(document.subject_title)
    if subject:
        row += 1
        _put(sheet, row, 1, subject.upper(), font=_BOLD)
        row += 1
    return row + 1


def _write_light_header(sheet: Worksheet, issue: ProjectQuotationIssue, row: int) -> int:
    """Enough for a sheet printed on its own: the reference it belongs to.

    Its scope label follows immediately below, written by the scope block itself, so a page found
    on a desk answers both "which quotation" and "which scope".
    """
    ref = _text(issue.our_ref_text)
    if ref:
        _put(sheet, row, 1, f"Our Ref: {ref}")
        row += 1
    return row + 1


def _write_scope(
    db: Session,
    sheet: Worksheet,
    scope: Dict[str, Any],
    row: int,
    *,
    grand_total: Optional[Decimal],
) -> None:
    lines: Sequence[ProjectQuotationLine] = scope["lines"]

    # Judged per SHEET (AC-F4): a workbook-wide decision would either put an empty column on a
    # scope that has no pictures, or drop it from one that does because its neighbour has none.
    show_image = any(line.image_attachment_id for line in lines)
    columns = [c for c in pdf._COLUMNS if c != pdf._IMAGE_COLUMN or show_image]
    money_column = len(columns)
    images = _image_names(db, lines) if show_image else {}
    positions = {name: index for index, name in enumerate(columns, start=1)}

    # The label verbatim, not upper-cased: it is the customer's own wording and it is also the tab
    # name, so the two have to read as the same thing.
    label = _text(scope.get("label"))
    if label:
        _put(sheet, row, 1, label, font=_BOLD)
        row += 1

    header_row = row
    for name, column in positions.items():
        _put(
            sheet,
            header_row,
            column,
            name,
            font=_BOLD,
            fill=_HEADER_FILL,
            alignment=_WRAP_LEFT,
        )
        sheet.column_dimensions[get_column_letter(column)].width = _COLUMN_WIDTHS.get(name, 16)
    # A scope runs to hundreds of lines; the headings have to stay visible while it is read.
    sheet.freeze_panes = f"A{header_row + 1}"
    row = header_row + 1

    current_band: Optional[str] = None
    for position, line in enumerate(lines, start=1):
        band = _text(line.band_label)
        # Once, above the line that opens the band. Repeated down every row it is noise a QS reads
        # past; in the item row's own cells it cannot be told apart from a description.
        if band and band != current_band:
            current_band = band
            _put(sheet, row, 1, band, font=_BOLD, fill=_BAND_FILL, alignment=_WRAP_LEFT)
            sheet.merge_cells(
                start_row=row, start_column=1, end_row=row, end_column=money_column
            )
            row += 1

        # Through pdf, deliberately: the workbook and the document are the same quotation, so
        # one definition of the item number keeps them from disagreeing about it.
        _put(
            sheet,
            row,
            positions["ITEM"],
            pdf.item_number(line, position),
            alignment=_CENTER,
        )
        if show_image:
            _put(
                sheet,
                row,
                positions[pdf._IMAGE_COLUMN],
                images.get(str(line.image_attachment_id or "")),
                alignment=_WRAP_LEFT,
            )
        _put(
            sheet,
            row,
            positions["TECHNICAL SPEC"],
            _text(line.technical_spec),
            alignment=_WRAP_LEFT,
        )
        _put(
            sheet,
            row,
            positions["DESCRIPTION"],
            _text(line.description_snapshot),
            alignment=_WRAP_LEFT,
        )
        _put(sheet, row, positions["BRAND"], _text(line.brand_snapshot), alignment=_WRAP_LEFT)
        _put(
            sheet,
            row,
            positions["PRODUCT CODE"],
            _text(line.product_code_snapshot),
            alignment=_WRAP_LEFT,
        )
        _put(
            sheet,
            row,
            positions["QTY"],
            _amount(line.quantity),
            number_format=_qty_format(line.uom),
            alignment=_RIGHT,
        )
        _put(
            sheet,
            row,
            positions["UNIT RATE (RM)"],
            _line_amount(line.unit_price),
            number_format=MONEY_FORMAT,
            alignment=_RIGHT,
        )
        _put(
            sheet, row, positions["COMPLETE SET"], _text(line.complete_set), alignment=_WRAP_LEFT
        )
        if line.is_rate_only:
            # The words, and NO number. A 0.00 here reads as free, and its rate would be pulled
            # into any total the reader sums (AC-C2 / AC-D3).
            _put(
                sheet,
                row,
                money_column,
                pdf.RATE_ONLY_TEXT,
                font=_ITALIC,
                alignment=_RIGHT,
            )
        else:
            _put(
                sheet,
                row,
                money_column,
                _line_amount(line.line_total),
                number_format=MONEY_FORMAT,
                alignment=_RIGHT,
            )
        row += 1

    if not lines:
        _put(sheet, row, 1, "No items priced in this scope.", font=_ITALIC)
        row += 1

    # The FROZEN scope total off the issue row, not a re-sum: the workbook must not be able to
    # disagree with the record behind it, or with the PDF of the same issue.
    label_column = max(money_column - 1, 1)
    _put(sheet, row, label_column, SCOPE_TOTAL_LABEL, font=_BOLD, alignment=_RIGHT)
    _put(
        sheet,
        row,
        money_column,
        _amount(scope.get("total")),
        font=_BOLD,
        number_format=MONEY_FORMAT,
        alignment=_RIGHT,
    )
    row += 1

    if grand_total is not None:
        # Stated once, on the first sheet. On every tab it would read as a per-sheet figure.
        row += 1
        _put(sheet, row, label_column, GRAND_TOTAL_LABEL, font=_BOLD, alignment=_RIGHT)
        _put(
            sheet,
            row,
            money_column,
            _amount(grand_total),
            font=_BOLD,
            number_format=GRAND_TOTAL_FORMAT,
            alignment=_RIGHT,
        )


def _prepare_for_print(sheet: Worksheet) -> None:
    """Landscape and fitted to one page wide, matching the PDF.

    These sheets get printed and put in front of a QS. Left at Excel's default a ten-column
    quotation breaks across two sheets of paper, splitting the money column off the descriptions.
    """
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    if sheet.sheet_properties.pageSetUpPr is not None:
        sheet.sheet_properties.pageSetUpPr.fitToPage = True


# --------------------------------------------------------------------- public


def build_issue_workbook(db: Session, issue: ProjectQuotationIssue) -> Workbook:
    """One workbook for one issue: a sheet per scope, in the document's own order (AC-F2).

    Exposed separately from ``render_issue_xlsx`` so the layout can be asserted cell by cell
    without going through a save.
    """
    document = pdf._document_of(db, issue)
    sender = pdf._sender(db, document)
    # The PDF's traversal, not a second one: same issue, same snapshot, same numbers.
    scopes = pdf._load_scopes(db, issue)

    workbook = Workbook()
    # Every sheet here is a scope, so the default one is removed rather than reused - reusing it
    # would leave an untitled empty sheet whenever the scope list is empty.
    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)

    used: Set[str] = set(_RESERVED_TITLES)
    for position, scope in enumerate(scopes, start=1):
        sheet = workbook.create_sheet(title=_sheet_title(scope.get("label"), position, used))
        first = position == 1
        row = (
            _write_letterhead(sheet, document, issue, sender, 1)
            if first
            else _write_light_header(sheet, issue, 1)
        )
        _write_scope(
            db, sheet, scope, row, grand_total=issue.grand_total if first else None
        )
        _prepare_for_print(sheet)

    if not workbook.worksheets:
        # A workbook with no sheets cannot be saved at all, so an issue carrying no scope gets the
        # letterhead and its (zero) total rather than a file the reader cannot open.
        sheet = workbook.create_sheet(title="Quotation")
        row = _write_letterhead(sheet, document, issue, sender, 1)
        _write_scope(
            db,
            sheet,
            {"label": None, "lines": [], "total": _amount(issue.grand_total)},
            row,
            grand_total=issue.grand_total,
        )
        _prepare_for_print(sheet)

    return workbook


def build_filename(issue: ProjectQuotationIssue) -> str:
    """The PDF's name with an ``.xlsx`` suffix.

    Derived from ``pdf.build_filename`` rather than re-implemented so the two artifacts of one
    issue can never drift apart: they sit next to each other in whatever folder the customer saves
    them to, under the reference they quote back to us.
    """
    return f"{os.path.splitext(pdf.build_filename(issue))[0]}.xlsx"


def render_issue_xlsx(db: Session, issue: ProjectQuotationIssue) -> Tuple[bytes, str]:
    """Return ``(xlsx_bytes, filename)`` for one issue."""
    buffer = io.BytesIO()
    build_issue_workbook(db, issue).save(buffer)
    return buffer.getvalue(), build_filename(issue)
