"""ReportResult -> an .xlsx workbook, laid out like the register the client keeps by hand.

SUMMARY first (the pivot), then ONE SHEET PER MONTH of the period, named the way the
client's own file names them (JAN'25 .. DEC'25). A month with no rows still gets its sheet:
a register with eleven tabs reads as a lost month rather than a quiet one.

**The layout is the client's, cell for cell** (AC-G7 to AC-G10): the title block occupies
rows 2-5 (company, report, the month as a real date, DEPARTMENT), the header is two rows,
6 and 7, with every single-level label merged vertically and every group merged across its
members, and the table starts on row 8. What differs per report is DATA, not code: the
words, the widths and the two summary labels come from the definition's `WorkbookSpec`, and
a report that declares none of them gets its own column labels uppercased.

**Totals are VALUES, never formulas** (AC-D4). The workbook the client keeps uses SUM
formulas; ours writes exactly what the engine computed, so an Excel recalculation cannot
produce a different answer from the screen the user exported.

The renderer knows nothing about months, dates or filters. It is handed a
``engine.WorkbookData`` - a summary and a list of named sheets - and draws it.
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import Dict, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.report import (
    ReportColumn,
    ReportColumnGroup,
    ReportDetailLayout,
    ReportPivotLayout,
)
from app.services.reports.engine import WorkbookData
from app.services.reports.registry import ReportDefinition, WorkbookSpec

#: Accounting, the way the client's own cells are formatted: the RM is pinned left, the
#: digits align on the decimal, and a zero prints "RM -" rather than "RM 0.00" (AC-G9).
MONEY_FORMAT = '_-"RM"* #,##0.00_-;-"RM"* #,##0.00_-;_-"RM"* "-"??_-;_-@_-'
#: What a money cell with NO value prints. The client types it by hand; a blank cell reads
#: as an oversight and a 0.00 claims a number nobody has.
NO_VALUE = "-"
# The CRM writes dates dd/mm/yyyy everywhere else, and a text date cannot be sorted,
# filtered or reformatted in Excel - which is most of what a file is opened to do.
DATE_FORMAT = "DD/MM/YYYY"
MONTH_FORMAT = "mmm-yy"
TICK = "X"

_COMPANY_FONT = Font(bold=True, size=26)
_REPORT_FONT = Font(bold=True, size=22)
_PERIOD_FONT = Font(bold=True, size=22)
_DEPARTMENT_FONT = Font(bold=True, size=12)
_HEADER_FONT = Font(bold=True, size=12)
_TOTAL_FONT = Font(bold=True)

_CENTRE = Alignment(horizontal="center", vertical="center")
_CENTRE_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center")

_THIN = Side(style="thin")
_BOX = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Row 1 is left empty, as the client's own sheets leave it. 2-5 are the title block, the
# header is 6-7, and the table starts at 8.
_COMPANY_ROW = 2
_REPORT_ROW = 3
_PERIOD_ROW = 4
_DEPARTMENT_ROW = 5
_GROUP_ROW = 6
_HEADER_ROW = 7
_FIRST_DATA_ROW = 8

_TITLE_HEIGHTS = {_COMPANY_ROW: 35.25, _REPORT_ROW: 30.75, _PERIOD_ROW: 29.25, _DEPARTMENT_ROW: 19.5}


def _header_text(spec: WorkbookSpec, key: str, label: str) -> str:
    """The word the client's own sheet prints over this column.

    The definition names the ones where their wording is not ours ("SPONSHER PROJECT"); the
    rest is the column's own label, uppercased, which is the whole layout for report #2.
    """
    return spec.headers.get(key) or (label or "").upper()


def _width(spec: WorkbookSpec, key: str) -> float:
    return spec.column_widths.get(key, spec.default_width)


def _decimal(value: Optional[str]) -> Optional[Decimal]:
    return Decimal(value) if value not in (None, "") else None


def _write_money(sheet: Worksheet, row: int, column: int, value: Optional[str], *, bold=False):
    """A money cell: the number when there is one, the client's own "-" when there is not."""
    cell = sheet.cell(row=row, column=column)
    amount = _decimal(value)
    cell.value = NO_VALUE if amount is None else amount
    cell.number_format = MONEY_FORMAT
    if bold:
        cell.font = _TOTAL_FONT
    return cell


def _write_date(sheet: Worksheet, row: int, column: int, value: Optional[str]) -> None:
    """An ISO date string as a REAL date cell. Anything unparseable stays the text it was."""
    cell = sheet.cell(row=row, column=column)
    try:
        cell.value = date.fromisoformat(str(value)[:10])
    except ValueError:
        cell.value = value
        return
    cell.number_format = DATE_FORMAT


def _title_block(
    sheet: Worksheet,
    definition: ReportDefinition,
    company: str,
    width: int,
    period_text: Optional[str],
    period_date: Optional[date],
) -> None:
    """The four lines every sheet opens with, merged across the table (AC-G7)."""
    spec = definition.workbook
    lines = (
        (_COMPANY_ROW, (company or spec.company_name).upper(), _COMPANY_FONT),
        (_REPORT_ROW, (spec.report_title or definition.title).upper(), _REPORT_FONT),
        (_PERIOD_ROW, period_date or period_text, _PERIOD_FONT),
    )
    for row, value, font in lines:
        cell = sheet.cell(row=row, column=1, value=value)
        cell.font = font
        cell.alignment = _CENTRE
        if row == _PERIOD_ROW and period_date is not None:
            cell.number_format = MONTH_FORMAT
        for column in range(1, width + 1):
            sheet.cell(row=row, column=column).border = _BOX
        if width > 1:
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)

    label = sheet.cell(row=_DEPARTMENT_ROW, column=1, value="DEPARTMENT:")
    label.font = _DEPARTMENT_FONT
    label.alignment = _CENTRE
    if width > 1:
        sheet.merge_cells(
            start_row=_DEPARTMENT_ROW,
            start_column=1,
            end_row=_DEPARTMENT_ROW,
            end_column=min(2, width),
        )
    if width > 2:
        sheet.cell(row=_DEPARTMENT_ROW, column=3, value=spec.department or "").alignment = _LEFT

    for row, height in _TITLE_HEIGHTS.items():
        sheet.row_dimensions[row].height = height


def _style_header_cells(sheet: Worksheet, width: int) -> None:
    for row in (_GROUP_ROW, _HEADER_ROW):
        for column in range(1, width + 1):
            cell = sheet.cell(row=row, column=column)
            cell.font = _HEADER_FONT
            cell.alignment = _CENTRE_WRAP
            cell.border = _BOX
    sheet.row_dimensions[_GROUP_ROW].height = 32.25


def _write_header(
    sheet: Worksheet,
    spec: WorkbookSpec,
    columns: Sequence[ReportColumn],
    groups: Sequence[ReportColumnGroup],
    positions: Dict[str, int],
) -> None:
    """The client's two-row header: single levels merged DOWN, groups merged ACROSS.

    A single-level label written on the leaf row alone leaves an empty band above every
    ungrouped column, which is the tell of a header drawn in two halves (AC-G4, AC-G8).
    """
    grouped = {key for group in groups for key in group.keys}

    for column in columns:
        index = positions[column.key]
        if column.key in grouped:
            # A tick column's header is the year itself, written as a number the way the
            # client writes it, and it stays on the leaf row under its group.
            label = column.label
            value = int(label) if str(label).isdigit() else label
            sheet.cell(row=_HEADER_ROW, column=index, value=value)
            continue
        sheet.cell(
            row=_GROUP_ROW, column=index, value=_header_text(spec, column.key, column.label)
        )
        sheet.merge_cells(
            start_row=_GROUP_ROW, start_column=index, end_row=_HEADER_ROW, end_column=index
        )

    for group in groups:
        members = [positions[key] for key in group.keys if key in positions]
        if not members:
            continue
        first, last = min(members), max(members)
        sheet.cell(
            row=_GROUP_ROW, column=first, value=_header_text(spec, group.source, group.label)
        )
        if last > first:
            sheet.merge_cells(
                start_row=_GROUP_ROW, start_column=first, end_row=_GROUP_ROW, end_column=last
            )

    _style_header_cells(sheet, len(columns))


def _render_detail(
    sheet: Worksheet,
    definition: ReportDefinition,
    detail: ReportDetailLayout,
    company: str,
    period_text: Optional[str],
    period_date: Optional[date],
) -> None:
    spec = definition.workbook
    columns = list(detail.columns)
    positions = {column.key: index for index, column in enumerate(columns, start=1)}
    width = max(len(columns), 1)

    _title_block(sheet, definition, company, width, period_text, period_date)
    _write_header(sheet, spec, columns, detail.column_groups, positions)

    row = _FIRST_DATA_ROW
    for record in detail.rows:
        for column in columns:
            index = positions[column.key]
            value = record.get(column.key)
            if column.type == "money":
                _write_money(sheet, row, index, value)
            elif column.type == "date":
                if value is not None:
                    _write_date(sheet, row, index, value)
            elif column.type == "bool":
                if value:
                    sheet.cell(row=row, column=index, value=TICK).alignment = _CENTRE
            elif value is not None:
                sheet.cell(row=row, column=index, value=value).alignment = _LEFT
            sheet.cell(row=row, column=index).border = _BOX
        row += 1

    _write_detail_total(sheet, detail, columns, positions, row)

    for column in columns:
        key = next(
            (g.source for g in detail.column_groups if column.key in g.keys), column.key
        )
        sheet.column_dimensions[get_column_letter(positions[column.key])].width = _width(spec, key)


def _write_detail_total(
    sheet: Worksheet,
    detail: ReportDetailLayout,
    columns: List[ReportColumn],
    positions: Dict[str, int],
    row: int,
) -> None:
    """GRAND TOTAL: the label beside the money, the money as the engine computed it.

    ONE rule, in three steps, because the columns are whatever the user left on screen:

    1. The cell immediately BEFORE the first measure, which is where the client types it -
       a column or two left of the first amount, where the eye lands on its way to it.
    2. When that cell carries a total of its own (a view whose first column IS a measure,
       so there is nothing to its left), the first cell in the row that does not. The label
       then sits to the RIGHT of the money rather than over it: writing it over a total
       would lose the number silently.
    3. When EVERY visible column is a totalled measure, there is no free cell at all, so
       the label takes column 1 of a bordered row directly ABOVE the totals. It used to be
       written one column past the table - outside the border and outside the print area.

    The label never leaves the table, and it never replaces an amount.
    """
    totalled = {c.key for c in columns if c.type == "money" and c.key in detail.totals}
    measure_positions = [positions[c.key] for c in columns if c.type == "money"]
    preferred = min(measure_positions) - 1 if measure_positions else 1
    free = [positions[c.key] for c in columns if c.key not in totalled]
    label_column = preferred if preferred in free else (min(free) if free else 1)
    label_row = row
    if not free:
        # Step 3: the money moves DOWN a row, so the label can have one of its own.
        row += 1

    cell = sheet.cell(row=label_row, column=label_column, value="GRAND TOTAL")
    cell.font = _TOTAL_FONT
    cell.alignment = _LEFT
    for column in columns:
        if column.key in totalled:
            _write_money(sheet, row, positions[column.key], detail.totals[column.key], bold=True)
    for index in range(1, len(columns) + 1):
        sheet.cell(row=label_row, column=index).border = _BOX
        sheet.cell(row=row, column=index).border = _BOX


def _render_summary(
    sheet: Worksheet,
    definition: ReportDefinition,
    pivot: ReportPivotLayout,
    company: str,
    period_label: str,
    period_compact: str,
) -> None:
    """The client's SUMMARY: salesman by month, a TOTAL SALES line, then the year totals."""
    spec = definition.workbook
    measures = pivot.measures
    span = max(len(measures), 1)
    width = 1 + span * (len(pivot.col_dim.values) + 1)

    _title_block(sheet, definition, company, width, period_label, None)

    sheet.cell(
        row=_GROUP_ROW, column=1, value=_header_text(spec, pivot.row_dim.key, pivot.row_dim.label)
    )
    sheet.merge_cells(start_row=_GROUP_ROW, start_column=1, end_row=_HEADER_ROW, end_column=1)

    def _group(column: int, label: str) -> None:
        sheet.cell(row=_GROUP_ROW, column=column, value=label)
        if span > 1:
            sheet.merge_cells(
                start_row=_GROUP_ROW,
                start_column=column,
                end_row=_GROUP_ROW,
                end_column=column + span - 1,
            )
        for offset, measure in enumerate(measures):
            sheet.cell(
                row=_HEADER_ROW,
                column=column + offset,
                value=_header_text(spec, measure.key, measure.label),
            )
            sheet.column_dimensions[get_column_letter(column + offset)].width = _width(
                spec, measure.key
            )

    column = 2
    for value in pivot.col_dim.values:
        _group(column, ((pivot.col_dim.value_labels or {}).get(value, value) or "").upper())
        column += span
    total_column = column
    _group(total_column, spec.summary_row_total_label)
    _style_header_cells(sheet, width)
    sheet.column_dimensions["A"].width = _width(spec, pivot.row_dim.key)

    row = _FIRST_DATA_ROW
    for row_value in pivot.row_values:
        # As stored: the register is read by the people named in it.
        sheet.cell(row=row, column=1, value=row_value).alignment = _LEFT
        column = 2
        for col_value in pivot.col_dim.values:
            cell_measures = pivot.cells.get(row_value, {}).get(col_value, {})
            for offset, measure in enumerate(measures):
                _write_money(sheet, row, column + offset, cell_measures.get(measure.key))
            column += span
        for offset, measure in enumerate(measures):
            _write_money(
                sheet,
                row,
                total_column + offset,
                pivot.row_totals.get(row_value, {}).get(measure.key),
            )
        for index in range(1, width + 1):
            sheet.cell(row=row, column=index).border = _BOX
        row += 1

    total_row = row
    label = sheet.cell(row=total_row, column=1, value=spec.summary_total_row_label)
    label.font = _TOTAL_FONT
    label.alignment = _CENTRE
    column = 2
    for col_value in pivot.col_dim.values:
        for offset, measure in enumerate(measures):
            _write_money(
                sheet, total_row, column + offset, pivot.col_totals.get(col_value, {}).get(measure.key)
            )
        column += span
    for offset, measure in enumerate(measures):
        _write_money(sheet, total_row, total_column + offset, pivot.grand_total.get(measure.key))
    for index in range(1, width + 1):
        sheet.cell(row=total_row, column=index).border = _BOX

    # One labelled line per measure, the way the client closes their own sheet: the two
    # numbers the whole register is kept for, spelled out rather than read off a corner.
    row = total_row + 2
    for measure in measures:
        title = " ".join(
            part
            for part in ("GRAND TOTAL", _header_text(spec, measure.key, measure.label), period_compact)
            if part
        )
        cell = sheet.cell(row=row, column=1, value=title)
        cell.font = _TOTAL_FONT
        cell.alignment = _CENTRE
        if width > 1:
            sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        _write_money(sheet, row, 3, pivot.grand_total.get(measure.key), bold=True)
        row += 1


def render_workbook(definition: ReportDefinition, data: WorkbookData) -> bytes:
    """The whole workbook as bytes: SUMMARY, then one sheet per month of the period."""
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "SUMMARY"
    company = data.company_name or definition.workbook.company_name
    _render_summary(
        summary_sheet,
        definition,
        data.summary,
        company,
        data.period_label,
        data.period_compact_label,
    )

    for sheet in data.sheets:
        # Excel refuses a tab name over 31 characters; a month never is, but a report whose
        # period is a custom range still names its sheets through here.
        _render_detail(
            workbook.create_sheet(title=sheet.name[:31]),
            definition,
            sheet.detail,
            company,
            sheet.label,
            sheet.month_start,
        )

    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
