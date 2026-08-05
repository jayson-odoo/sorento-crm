"""Read `Container Status 2026.xlsx` into shipment field values.

The workbook is not the flat sheet-per-tab file it looks like, and every rule
below exists because reading it the obvious way produces wrong answers silently.

**Blocks, not sheets.** Several tabs stack more than one titled section, each
with its own header row. The real file has 9 blocks across 5 tabs (`Fitting` at
rows 2 and 31, `Ceramic` at 2, 69 and 75, `Arrived` at 2,
`Arrived - Joint Mocha Container` at 2 and 22, `Arrived (Mocha) Joint BL` at 2).
So parsing is anchored on rows whose cell text is exactly ``CONTAINER``: each one
OPENS a block and supplies that block's column positions. A repeated header is a
section boundary, never a data row - treat it as data and you report four bogus
ISO 6346 rejects, which is what an earlier draft did.

**Names, not positions.** Header names drift between tabs. `Ceramic` heads its
liner column ``RL`` while every other tab heads it ``LINER``; `Arrived` uses
``W/H ARRIVALS`` for ``WAREHOUSE ARRIVALS``. Reading column 4 positionally
mislabels 55 liners and nothing complains, so every header resolves through
:data:`HEADER_ALIASES` and an unrecognised one is reported rather than guessed at.

**Blank rows are not errors.** 475 numbered rows carry no container number. They
are scaffolding the maintainer types ahead of time.

Nothing here touches the database. The parser returns data; matching and writing
live in :mod:`app.services.container_status_service`.
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any, Iterable, Optional

logger = logging.getLogger(__name__)


class ContainerStatusParseError(Exception):
    """The file is not a container status workbook at all."""


# ISO 6346: four letters (owner code + category) then seven digits.
_ISO_6346 = re.compile(r"^[A-Z]{4}\d{7}$")

# Same set the packing-list matcher strips, so a candidate found in SQL survives
# the Python-side comparison. Kept in sync with
# procurement_service._CONTAINER_STRIP_CHARS.
_CONTAINER_STRIP_CHARS = (" ", "-", "/", ".", "_")

# Openpyxl stops reporting a real last row on these sheets (max_row runs to
# 1,046,173 on Ceramic), so scanning is bounded by a run of empty rows instead.
_EMPTY_RUN_LIMIT = 50
_MAX_COLUMNS = 70

#: Canonical header -> the shipment column it fills.
FIELD_MAP: dict[str, str] = {
    "LOC": "loc",
    "LINER": "liner_code",
    "CHINA FORWARDER": "china_forwarder",
    "MALAYSIA FORWARDER": "malaysia_forwarder",
    "CONSIGNEE": "consignee",
    "FREE DAYS AVAILABLE": "free_days_available",
    "STACKED": "stacked",
    "LOADING": "loading_date",
    "ETC": "etc_date",
    "ETD": "etd_date",
    "ETA": "eta_date",
    "ETA DELAY": "eta_delay_date",
    "INSPECTION": "inspection_date",
    "APPROVAL": "approval_date",
    "GATEPASS": "gatepass_date",
    "DELIVERY WAREHOUSE": "delivery_warehouse",
    "WAREHOUSE ARRIVALS": "warehouse_arrival_date",
    "INFORMED COLLECTION": "informed_collection_date",
    "COLLECTION": "collection_date",
    "COA PERMIT NO": "coa_permit_no",
}

#: Columns the parser recognises and deliberately does not import.
#: ``FACTORY`` and ``PRODUCT`` are outside the signed-off frontend contract; the
#: cost block is deferred (D9) and survives in the retained original file. Naming
#: them here is what keeps them out of the "unrecognised header" warning, so a
#: genuinely new column still stands out.
IGNORED_HEADERS: frozenset[str] = frozenset(
    {
        "NO",
        "FACTORY",
        "PRODUCT",
        # Present on every tab, filled 6 / 4 / 4 / 4 times across 407 containers.
        # Read by nothing, so not imported at all (D34); the retained source file
        # keeps them.
        "ATA",
        "ORI DOC RECEIVED DATE",
        "K1 SUBMISSION",
        "YARD ARRIVALS",
        "REMARKS 1",
        "REMARKS 2",
        "REMARKS 3",
        "2ND LABOUR CHARGE",
        "K1 RATE",
        "CHINA FORWARDING COST (RMB)",
        "PRODUCT VALUE (RMB)",
        "DATE RECEIVE K1",
        "SST AMOUNT",
        "LOCAL CLEARANCE CHARGES",
        "KOPI",
        "DEMURRAGE",
        "PORT CHARGES",
        "DETENTION",
        "TOTAL LOCAL CLEARANCE",
        "TOTAL FREIGHT",
        "CARGO INSURANCE",
        "SRT? MBS? BOTH?",
        "USD",
        "SST",
        "RATE",
        "SYSTEM",
        "ACTUAL",
        "DIFF",
        # A rate cell used as its own header on three tabs.
        "0.014",
    }
)

#: Canonical header -> the spellings actually found in the file. Each entry is a
#: measured difference, not a defensive guess.
HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    # Ceramic heads the liner column RL. 55 rows depend on this one line.
    "LINER": ("RL",),
    "WAREHOUSE ARRIVALS": ("W/H ARRIVALS",),
    "CHINA FORWARDING COST (RMB)": ("CHINA FREIGHT (RMB)",),
    "SST": ("10% SST",),
    "DEMURRAGE": ("DEMURRANGE",),
}

#: Reverse lookup, built once.
_ALIAS_TO_CANONICAL: dict[str, str] = {
    alias.upper(): canonical
    for canonical, aliases in HEADER_ALIASES.items()
    for alias in aliases
}

_REMARKS_HEADERS = ("REMARKS 1", "REMARKS 2", "REMARKS 3")

_DATE_FIELDS = frozenset(
    name for name in FIELD_MAP.values() if name.endswith("_date")
)
_INT_FIELDS = frozenset({"free_days_available"})


def normalize_container(value: Optional[str]) -> str:
    """Uppercase and strip separators. Twin of ``_container_match_key``."""
    if value is None:
        return ""
    key = str(value).strip()
    if not key:
        return ""
    for ch in _CONTAINER_STRIP_CHARS:
        key = key.replace(ch, "")
    return key.upper()


@dataclass
class ParsedBlock:
    """One titled section of one tab."""

    sheet: str
    #: 1-based Excel row of the header. Quoted in warnings so a maintainer can
    #: find it.
    header_row: int
    title: Optional[str]
    #: Canonical header -> 0-based column index, for THIS block only.
    columns: dict[str, int]
    row_count: int = 0


@dataclass
class ParsedRow:
    sheet: str
    excel_row: int
    container: str
    container_key: str
    #: Shipment column -> coerced value. Only non-empty cells appear, which is
    #: what makes blank-never-clears fall out naturally downstream.
    values: dict[str, Any] = field(default_factory=dict)
    #: REMARKS 1/2/3, in order, non-empty only. These become activity feed
    #: entries, never columns (B4).
    remarks: list[str] = field(default_factory=list)


@dataclass
class RejectedRow:
    sheet: str
    excel_row: int
    container: str
    reason: str


@dataclass
class Occurrence:
    sheet: str
    excel_row: int


@dataclass
class Collision:
    container_key: str
    occurrences: list[Occurrence]


@dataclass
class ParsedWorkbook:
    blocks: list[ParsedBlock] = field(default_factory=list)
    rows: list[ParsedRow] = field(default_factory=list)
    rejected: list[RejectedRow] = field(default_factory=list)
    collisions: list[Collision] = field(default_factory=list)
    blank_row_count: int = 0
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        # A header cell holding 0.014 arrives as a float; a numeric "header" like
        # 21 must not become "21.0" and miss its alias.
        return str(int(value)).strip()
    return str(value).strip()


def _header_key(value: Any) -> str:
    return _text(value).upper()


def _canonical_header(raw: str) -> Optional[str]:
    """Canonical name for a header cell, or None when it means nothing to us."""
    key = raw.upper()
    if not key:
        return None
    if key in FIELD_MAP or key in IGNORED_HEADERS or key in _REMARKS_HEADERS:
        return key
    return _ALIAS_TO_CANONICAL.get(key)


def _coerce_date(value: Any) -> Optional[date]:
    """Excel hands back datetimes, serials and the occasional typed string."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        # Excel serial. 1899-12-30 is the epoch openpyxl uses for the 1900 system.
        try:
            from openpyxl.utils.datetime import from_excel

            converted = from_excel(float(value))
        except Exception:  # noqa: BLE001 - a junk serial is not worth a failure
            return None
        if isinstance(converted, datetime):
            return converted.date()
        return converted if isinstance(converted, date) else None
    text = str(value).strip()
    if not text:
        return None
    for pattern in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    return None


def _coerce_int(value: Any) -> Optional[int]:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _sheet_rows(worksheet) -> list[tuple]:
    """Rows up to the real end of content.

    ``worksheet.max_row`` is unusable here - it reports over a million rows on
    three of the five tabs - so read until a long run of wholly empty rows.
    """
    rows: list[tuple] = []
    empty_run = 0
    for row in worksheet.iter_rows(min_row=1, max_col=_MAX_COLUMNS, values_only=True):
        if not any(cell not in (None, "") for cell in row):
            empty_run += 1
            if empty_run > _EMPTY_RUN_LIMIT and rows:
                break
        else:
            empty_run = 0
        rows.append(row)
    while rows and not any(cell not in (None, "") for cell in rows[-1]):
        rows.pop()
    return rows


def _block_title(rows: list[tuple], header_index: int) -> Optional[str]:
    """The titled line above the header, e.g. "FITTING CONTAINER 2026 (JOINT
    MOCHA CONTAINER)". Recorded for traceability only."""
    if header_index == 0:
        return None
    for cell in rows[header_index - 1]:
        text = _text(cell)
        if text:
            return text
    return None


def _find_header_rows(rows: list[tuple]) -> list[tuple[int, int]]:
    """(row index, container column index) for every header row in this tab."""
    found: list[tuple[int, int]] = []
    for index, row in enumerate(rows):
        for column, cell in enumerate(row):
            if _header_key(cell) == "CONTAINER":
                found.append((index, column))
                break
    return found


def parse_container_status_workbook(file_data: bytes) -> ParsedWorkbook:
    """Parse the workbook. Never touches the database.

    Raises :class:`ContainerStatusParseError` only when the file cannot be a
    container status workbook at all - no tab has a ``CONTAINER`` header. Every
    other problem is reported on the result so the operator can decide.
    """
    import openpyxl

    try:
        workbook = openpyxl.load_workbook(
            BytesIO(file_data), data_only=True, read_only=True
        )
    except Exception as exc:  # noqa: BLE001 - surfaced to the operator as-is
        raise ContainerStatusParseError(f"Could not read the workbook: {exc}") from exc

    result = ParsedWorkbook()
    unknown_headers: dict[str, set[str]] = {}
    aliased: set[tuple[str, str, str]] = set()
    seen: dict[str, list[Occurrence]] = {}

    for worksheet in workbook.worksheets:
        rows = _sheet_rows(worksheet)
        header_positions = _find_header_rows(rows)
        if not header_positions:
            continue

        for position, (header_index, container_column) in enumerate(header_positions):
            end_index = (
                header_positions[position + 1][0] - 1
                if position + 1 < len(header_positions)
                else len(rows) - 1
            )
            header_row = rows[header_index]

            columns: dict[str, int] = {}
            for column, cell in enumerate(header_row):
                raw = _header_key(cell)
                if not raw or raw == "CONTAINER":
                    continue
                canonical = _canonical_header(raw)
                if canonical is None:
                    unknown_headers.setdefault(raw, set()).add(worksheet.title)
                    continue
                if canonical != raw:
                    aliased.add((canonical, raw, worksheet.title))
                # First occurrence wins: a tab repeating LOCAL CLEARANCE CHARGES
                # twice must not have its earlier column silently replaced.
                columns.setdefault(canonical, column)

            block = ParsedBlock(
                sheet=worksheet.title,
                header_row=header_index + 1,
                title=_block_title(rows, header_index),
                columns=columns,
            )
            result.blocks.append(block)

            for row_index in range(header_index + 1, end_index + 1):
                row = rows[row_index]
                excel_row = row_index + 1
                raw_container = (
                    _text(row[container_column])
                    if container_column < len(row)
                    else ""
                )
                if not raw_container:
                    # A numbered row with no container is scaffolding, not an error.
                    result.blank_row_count += 1
                    continue

                key = normalize_container(raw_container)
                if not _ISO_6346.match(key):
                    result.rejected.append(
                        RejectedRow(
                            sheet=worksheet.title,
                            excel_row=excel_row,
                            container=raw_container,
                            reason=(
                                f"'{raw_container}' is not a valid ISO 6346 container "
                                "number (four letters then seven digits)"
                            ),
                        )
                    )
                    continue

                parsed_row = ParsedRow(
                    sheet=worksheet.title,
                    excel_row=excel_row,
                    container=raw_container,
                    container_key=key,
                    values={"source_sheet": worksheet.title},
                )

                for canonical, column in columns.items():
                    if column >= len(row):
                        continue
                    cell = row[column]
                    if canonical in _REMARKS_HEADERS:
                        text = _text(cell)
                        if text:
                            parsed_row.remarks.append(text)
                        continue
                    field_name = FIELD_MAP.get(canonical)
                    if field_name is None:
                        continue
                    if field_name in _DATE_FIELDS:
                        coerced = _coerce_date(cell)
                    elif field_name in _INT_FIELDS:
                        coerced = _coerce_int(cell)
                    else:
                        coerced = _text(cell) or None
                    # Only non-empty values are carried, which is what makes
                    # blank-never-clears (A5) fall out of the apply step for free.
                    if coerced is not None:
                        parsed_row.values[field_name] = coerced

                # Remarks are read in column order; the sheet's own order is 1,2,3.
                parsed_row.remarks = [r for r in parsed_row.remarks if r]

                result.rows.append(parsed_row)
                block.row_count += 1
                seen.setdefault(key, []).append(
                    Occurrence(sheet=worksheet.title, excel_row=excel_row)
                )

    workbook.close()

    if not result.blocks:
        raise ContainerStatusParseError(
            "No tab in this workbook has a CONTAINER column, so it is not a "
            "container status sheet."
        )

    result.collisions = [
        Collision(container_key=key, occurrences=occurrences)
        for key, occurrences in seen.items()
        if len(occurrences) > 1
    ]

    result.warnings.extend(_alias_warnings(aliased))
    result.warnings.extend(_unknown_header_warnings(unknown_headers))
    result.warnings.extend(_blank_row_warning(result))
    result.warnings.extend(_empty_block_warning(result))

    return result


def _alias_warnings(aliased: Iterable[tuple[str, str, str]]) -> list[str]:
    """Only aliases for columns we actually IMPORT are worth telling anyone about.

    Three of the five live aliases are on the cost block (`CHINA FREIGHT (RMB)`,
    `DEMURRANGE`, `10% SST`), which this importer deliberately skips. Warning about
    those buries the one that matters - `LINER <- RL`, which decides whether 55
    Ceramic rows get a liner at all.
    """
    grouped: dict[str, list[str]] = {}
    for canonical, raw, sheet in sorted(aliased):
        if canonical not in FIELD_MAP:
            continue
        grouped.setdefault(canonical, []).append(f'"{raw}" on {sheet}')
    return [
        f'Header matched by alias: {canonical} <- {", ".join(sources)}.'
        for canonical, sources in grouped.items()
    ]


def _unknown_header_warnings(unknown: dict[str, set[str]]) -> list[str]:
    if not unknown:
        return []
    parts = [
        f'"{header}" on {", ".join(sorted(sheets))}'
        for header, sheets in sorted(unknown.items())
    ]
    return [
        "Unrecognised column(s), not imported: "
        + "; ".join(parts)
        + ". Add them to the field map or the ignore list."
    ]


def _blank_row_warning(result: ParsedWorkbook) -> list[str]:
    if not result.blank_row_count:
        return []
    return [
        f"{result.blank_row_count} numbered rows carry no container number and "
        "were skipped without an error."
    ]


def _empty_block_warning(result: ParsedWorkbook) -> list[str]:
    empty = [b for b in result.blocks if b.row_count == 0]
    if not empty:
        return []
    named = ", ".join(f"{b.sheet} row {b.header_row}" for b in empty)
    return [
        f"{len(empty)} of {len(result.blocks)} sections hold no containers "
        f"({named}). Empty scaffolding, nothing to import."
    ]
