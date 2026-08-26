"""Purchasing's own order inquiry: every raised row, across every subject.

**Why this exists next to `project_order_inquiry_service`.** That one reads the rows of
ONE project. Purchasing does not work a project at a time, and more to the point it
cannot: an order adopted from the AutoCount book has `project_id` NULL by design
(`PLAN-fulfilment-planning-from-autocount-so.md` section 2), so its rows belong to no
project and no per-project query will ever return them. Confirming supply on an adopted
order raised rows that were reachable only from the one sales order that raised them.

**The shape is not invented here.** It is the workbook purchasing already works from -
`JAN - DEC 2026 ORDER.xlsx`, sheet title `ORDER INQUIRY`, columns SO DATE / S/O NO / ITEM
CODE / QTY / TOTAL QTY / DELIVERY DATE / PROJECT/CUSTOMER / SUPPLIER / PO NO, one sheet
per delivery MONTH plus a tab per day showing what was raised that day. The columns are
read off that file rather than retyped, the month is the primary axis rather than one
filter among many, and the export writes the same workbook back out.

**SUPPLIER and PO NO are never guessed.** On their sheet a blank in those two columns
means "not placed yet", and that is exactly what the system knows: the only link the
schema holds between an inquiry row and a placed purchase order is the row's SPO
reference, through `spo_allocations.po_line_id`. A row with no such link prints blank
rather than the product's usual supplier, because purchasing reads that column as a
statement that an order exists.

**Each of the screen's own controls is computed ignoring its own filter** (the month
strip, the supplier list, the project list, the people who raised them). A control that
empties itself the moment it is used cannot be used a second time. The visible TOTALS
honour every filter, month included, because that strip is a statement about what is on
screen.
"""
from __future__ import annotations

import io
import logging
from datetime import date, datetime
from decimal import Decimal
from itertools import groupby
from typing import Any, Dict, List, Optional, Sequence, Tuple

from sqlalchemy import Date, String, cast, func, or_, select
from sqlalchemy.orm import Session, aliased

from app.models.inventory import Warehouse
from app.models.order import Customer, SalesOrder, SalesOrderLine
from app.models.procurement import PurchaseOrder, PurchaseOrderLine, SPOAllocation, Supplier
from app.models.product import Product
from app.models.project_so import (
    ACK_ACKNOWLEDGED,
    ACK_AWAITING,
    ACK_CHANGED,
    ACK_REJECTED,
    ACK_STATES,
    INQUIRY_ACTIONED,
    INQUIRY_CANCELLED,
    INQUIRY_PLACED,
    INQUIRY_RAISED,
    INQUIRY_PARTLY_LINKED,
    IV_ORDER,
    IV_ORDER_BACK,
    OrderInquiry,
    OrderInquiryLink,
    OrderInquiryRow,
    ProjectSalesOrder,
    ProjectSalesOrderLine,
    SOSupplyDecision,
)
from app.models.projects import Project, ProjectParty, ProjectPurchaseOrder
from app.models.sales_agent import SalesAgent
from app.models.user import User
from app.services.error_handler import AppException
from app.services.project_order_inquiry_service import (
    ProjectOrderInquiryService,
    project_customer_label,
)

logger = logging.getLogger(__name__)

_ZERO = Decimal("0")

#: The states pre-seeded at zero on the summary strip: every state a row can be in,
#: `partly_linked` and `placed` included (section 3.I). Declared on the schema too
#: (`OrderInquiryStateCounts`), because `response_model` silently drops a key it has not
#: been told about - the strip reported four states while the rows carried six.
INQUIRY_STATES = (
    INQUIRY_RAISED,
    INQUIRY_PARTLY_LINKED,
    INQUIRY_ACTIONED,
    INQUIRY_CANCELLED,
    INQUIRY_PLACED,
)

#: The four acknowledgement states, pre-seeded at zero on the facet so a state nothing is
#: in still offers itself in the filter (`PLAN-scm-oi-handshake.md` section 4). Read off
#: the model's own tuple rather than retyped, so a fifth one cannot arrive here unnoticed.
ACK_FILTER_STATES = ACK_STATES

#: Every column the list renders is sortable, and this is the CLOSED SET. An unknown
#: value is a 422 rather than a silent fall back to the default, for the same reason it
#: is on the fulfilment worklist: a grid drawing a sort arrow on a column the server
#: ignored is a screen lying about what it is showing.
#:
#: The route declares the same names as a `Literal` (FastAPI cannot build one from a
#: runtime set) and a test asserts the two agree.
SORTABLE_FIELDS = frozenset(
    {
        "inquiry_no",
        "so_date",
        "so_number",
        "item_code",
        "product_name",
        "qty",
        "delivery_date",
        "project_customer",
        "supplier",
        "po_number",
        "state",
        "raised_at",
        "raised_by_name",
        "location",
        "agent",
    }
)

#: Delivery date ascending, because that is the order the work is due in and the order
#: their own sheets are in. Sorting is an override of this, never a replacement for
#: having one.
DEFAULT_SORT_FIELD = "delivery_date"

#: Their tab names, not the browser's: `JAN 26`, `JUNE 26`, `JULY 26`, `SEPT 26`. Read
#: off the workbook, because a renamed tab is a tab their own habits stop finding.
MONTH_WORD = (
    "JAN",
    "FEB",
    "MAR",
    "APR",
    "MAY",
    "JUNE",
    "JULY",
    "AUG",
    "SEPT",
    "OCT",
    "NOV",
    "DEC",
)

EXPORT_TITLE = "ORDER INQUIRY"
#: Where the rows that have no delivery date go. They are not dropped: an instruction
#: with no date is still an instruction, and one that vanishes from the file is one
#: nobody places.
EXPORT_UNDATED_SHEET = "NO DATE"
#: Read off `JAN - DEC 2026 ORDER.xlsx` row 2, trailing space and all. Retyping it
#: tidier is how a column their filters no longer find gets introduced.
EXPORT_HEADINGS = (
    "SO DATE",
    "S/O NO",
    "ITEM CODE",
    "QTY",
    "TOTAL QTY",
    "DELIVERY DATE",
    "PROJECT/CUSTOMER",
    "SUPPLIER",
    "PO NO ",
    "LOCATION",
    # APPENDED, never inserted: their own filters and habits are keyed on the columns
    # above being where they have always been (`PLAN-scm-oi-handshake.md` section 4).
    "ACKNOWLEDGED",
)

# The two routes a row can be attributed by, joined ONCE through a coalesce rather than
# twice through two aliases of `customers`. The precedence is the same either way - the
# billed project party first, the core sales order's own customer when there is none -
# and one join keeps the company-scope predicate correct: `with_loader_criteria` emits an
# UNALIASED `customers.company_id` into each aliased ON clause, which Postgres rejects
# outright ("invalid reference to FROM-clause entry").
_CUSTOMER_ID = func.coalesce(ProjectParty.customer_id, SalesOrder.customer_id)

# The purchase order this row's FIRST link sits on - what the Supplier column reads and
# what the row's `po_id` addresses. Off the LINKS since section 3.I: a row holds many now,
# and `order_inquiry_rows.po_line_id` is the derived display of the first one rather than
# the record. Ordered by when the link was made, so "first" is an order in time.
#
# Correlated EXPLICITLY - an auto-correlated `exists`/scalar subquery loses its FROM
# clauses the moment this query is reshaped (the `shipment_supplier_predicate` lesson).
_LINKED_PO_ID = (
    select(PurchaseOrderLine.purchase_order_id)
    .select_from(OrderInquiryLink)
    .join(PurchaseOrderLine, PurchaseOrderLine.id == OrderInquiryLink.po_line_id)
    .where(OrderInquiryLink.row_id == OrderInquiryRow.id)
    .order_by(OrderInquiryLink.linked_at.asc(), OrderInquiryLink.id.asc())
    .limit(1)
    .correlate(OrderInquiryRow)
    .scalar_subquery()
)
# The row's OWN link on an SPO, and through it the purchase order that shipping order
# draws down (when the book ever names one - `po_line_id` is NULL on every migrated SPO
# allocation, so this is the path a system-raised SPO takes and no other).
_SPO_LINKED_PO_ID = (
    select(PurchaseOrderLine.purchase_order_id)
    .select_from(OrderInquiryLink)
    .join(SPOAllocation, SPOAllocation.id == OrderInquiryLink.spo_allocation_id)
    .join(PurchaseOrderLine, PurchaseOrderLine.id == SPOAllocation.po_line_id)
    .where(OrderInquiryLink.row_id == OrderInquiryRow.id)
    .order_by(OrderInquiryLink.linked_at.asc(), OrderInquiryLink.id.asc())
    .limit(1)
    .correlate(OrderInquiryRow)
    .scalar_subquery()
)
# The row's OWN `spo_ref` - the coverage reference the netting engine writes on an
# ALREADY INBOUND / PRE-ORDERED row (`ProjectOrderInquiryService._write`), which is not a
# placement at all and never became a link. Kept as the last leg: those rows are traceable
# to a purchase order through the shipping order that covers them, and dropping the leg
# when placements moved to the links table would have blanked their Supplier and PO no
# columns for a reason that has nothing to do with linking.
_SPO_REF_PLACED_PO_ID = (
    select(PurchaseOrderLine.purchase_order_id)
    .select_from(SPOAllocation)
    .join(PurchaseOrderLine, PurchaseOrderLine.id == SPOAllocation.po_line_id)
    .where(SPOAllocation.spo_number == OrderInquiryRow.spo_ref)
    .limit(1)
    .correlate(OrderInquiryRow)
    .scalar_subquery()
)
_PLACED_PO_ID = func.coalesce(
    _LINKED_PO_ID, _SPO_LINKED_PO_ID, _SPO_REF_PLACED_PO_ID
)

#: Does this row hold a link of each kind? The "Linked" filter's own predicates (AC-I5),
#: stated once so the filter and the column cannot disagree about what "linked to a PO"
#: means.
_HAS_PO_LINK = (
    select(OrderInquiryLink.id)
    .where(
        OrderInquiryLink.row_id == OrderInquiryRow.id,
        OrderInquiryLink.po_line_id.isnot(None),
    )
    .correlate(OrderInquiryRow)
    .exists()
)
_HAS_SPO_LINK = (
    select(OrderInquiryLink.id)
    .where(
        OrderInquiryLink.row_id == OrderInquiryRow.id,
        OrderInquiryLink.spo_allocation_id.isnot(None),
    )
    .correlate(OrderInquiryRow)
    .exists()
)
_HAS_ANY_LINK = (
    select(OrderInquiryLink.id)
    .where(OrderInquiryLink.row_id == OrderInquiryRow.id)
    .correlate(OrderInquiryRow)
    .exists()
)


def _linked_qty(*where) -> Any:
    """How much of this row sits on documents, correlated to the row itself.

    One builder for the three sums the cards need (SPO, PO, everything), so the three
    cannot come apart from each other or from `_HAS_PO_LINK` / `_HAS_SPO_LINK` above.
    """
    return (
        select(func.coalesce(func.sum(OrderInquiryLink.qty), 0))
        .where(OrderInquiryLink.row_id == OrderInquiryRow.id, *where)
        .correlate(OrderInquiryRow)
        .scalar_subquery()
    )


#: The three quantities the strip's cards are (section 3.I2, AC-I11): what sits on SPO
#: allocations, what sits on purchase order lines, and the remainder nobody has put
#: anywhere - which is what still flows to reorder planning.
#:
#: NEVER NEGATIVE. A row linked beyond its own quantity is a data question, and a negative
#: Buy would quietly cancel out a real one somewhere else in the same total.
#:
#: `_UNLINKED_QTY` is a per-row remainder and says nothing about whether that remainder is
#: still OWED - `_NOT_OWED_STATES` below is what answers that, and every reader of this
#: column applies it.
_SPO_LINKED_QTY = _linked_qty(OrderInquiryLink.spo_allocation_id.isnot(None))
_PO_LINKED_QTY = _linked_qty(OrderInquiryLink.po_line_id.isnot(None))
_UNLINKED_QTY = func.greatest(OrderInquiryRow.qty - _linked_qty(), 0)

#: The two states whose quantity is NOT OWED any more, so neither the three cards nor the
#: `kind` filter counts them: `cancelled` was called off, and `actioned` has already been
#: answered somewhere else. `_quantity_flow_by_so_line` below and `scm.committed_v` both
#: drop `actioned` for exactly this reason, so counting an actioned row's remainder here
#: would put a quantity on the Buy card that reorder planning itself does not carry - and
#: purchasing would buy it twice.
_NOT_OWED_STATES = (INQUIRY_CANCELLED, INQUIRY_ACTIONED)

# `SO DATE` is the date on the DOCUMENT. For an adopted order that is the core sales
# order's own order date; an authored one that has never been to AutoCount falls back to
# when it published, and then to when it was written, so the column is never empty for a
# row that plainly has a date somewhere.
_SO_DATE = func.coalesce(
    SalesOrder.order_date,
    cast(ProjectSalesOrder.published_at, Date),
    cast(ProjectSalesOrder.created_at, Date),
)
_SO_NUMBER = func.coalesce(
    ProjectSalesOrder.autocount_doc_no, ProjectSalesOrder.provisional_ref
)
_CUSTOMER_NAME = Customer.customer_name
# What `PROJECT/CUSTOMER` sorts and filters on. The printed label starts with the
# customer when there is one and with the project when there is not, so this is the same
# ordering the eye reads down the column.
_PROJECT_CUSTOMER = func.coalesce(_CUSTOMER_NAME, Project.title)
_RAISED_AT = OrderInquiryRow.created_at
# The per-day tab is a MALAYSIAN day. `created_at` is stored naive UTC (the session runs
# with `timezone=utc`), so a row raised at 00:30 MYT belongs to the tab of the day that
# is still, in UTC, the evening before.
_RAISED_DAY = cast(
    func.timezone("Asia/Kuala_Lumpur", func.timezone("UTC", _RAISED_AT)), Date
)
# WHO told purchasing to buy THIS ROW. Per ROW, not per header, and the difference is the
# whole point: the header's `raised_by` is RE-STAMPED on every reconfirm (AC-H4), so
# reading it here would print the latest reconfirmer's name beside an older row's own
# clock - a row A raised on 12 Aug would read "B, 12/08 10:25" the moment B reconfirmed
# one other line. The row's own answer is the revision that raised it:
# `supply_decision_id` -> `so_supply_decisions.confirmed_by`, the same person the header
# stamps at that moment (PLAN section 3.H).
#
# An amendment-born row (`ProjectOrderInquiryService._write`) carries NO decision at all -
# it is raised off the amendment, not off a supply revision - so it falls back to its
# header's `raised_by`, which for that inquiry is the person who published the amendment
# and is never re-stamped (an amendment raises its OWN inquiry).
#
# The id never leaves the service: a screen printing a UUID at a buyer is a screen they
# cannot use, so the filter takes an id and every read gives a name.
_RAISED_BY_ID = func.coalesce(SOSupplyDecision.confirmed_by, OrderInquiry.raised_by)
_RAISED_BY_NAME = User.name
# Where the PO gets placed for, not where the item is bought TO. `stock_location` on the
# row is stamped once, at raise time: the DONOR the take left oversold for an order-back
# row, or the confirmed allocation's warehouse for a plan/confirmed row
# (`project_order_inquiry_service._stock_location`). A plain Buy row raised straight off
# the board has neither, so it falls back to the fulfilment warehouse already on the
# line's own core sales order line - still a real answer, just not one purchasing
# confirmed themselves. Blank only when the row traces to no line at all.
_LOCATION = func.coalesce(OrderInquiryRow.stock_location, Warehouse.warehouse_code)

#: WHO acknowledged the row and who rejected it, by name. Two aliases of `users`, because
#: the same query already joins that table once for the person who RAISED the row and the
#: three answers are different people. Both OUTER and both on a primary key, so a row
#: nobody has acknowledged - or one whose acknowledger has since been removed - still
#: reaches the list.
_ACK_USER = aliased(User)
_REJECT_USER = aliased(User)

_SORT_EXPRESSIONS = {
    "inquiry_no": OrderInquiry.inquiry_no,
    "so_date": _SO_DATE,
    "so_number": _SO_NUMBER,
    "item_code": OrderInquiryRow.item_code,
    "product_name": Product.product_name,
    "qty": OrderInquiryRow.qty,
    "delivery_date": OrderInquiryRow.delivery_date,
    "project_customer": _PROJECT_CUSTOMER,
    "supplier": Supplier.supplier_name,
    "po_number": PurchaseOrder.po_number,
    "state": OrderInquiryRow.state,
    "raised_at": _RAISED_AT,
    "raised_by_name": _RAISED_BY_NAME,
    "location": _LOCATION,
    "agent": SalesAgent.sales_agent,
}

_COLUMNS = (
    OrderInquiryRow.id.label("id"),
    # `OI-000123`, off the header this row already joins to - no second query, and no
    # second opinion about which inquiry a row belongs to. The S/O no cannot stand in for
    # it: an amendment raises a SECOND inquiry on the same sales order.
    OrderInquiry.inquiry_no.label("inquiry_no"),
    OrderInquiryRow.so_line_id.label("so_line_id"),
    OrderInquiryRow.item_code.label("item_code"),
    OrderInquiryRow.qty.label("qty"),
    OrderInquiryRow.delivery_date.label("delivery_date"),
    OrderInquiryRow.state.label("state"),
    OrderInquiryRow.verb.label("verb"),
    OrderInquiryRow.note.label("note"),
    OrderInquiryRow.cited_document.label("cited_document"),
    _RAISED_AT.label("raised_at"),
    _RAISED_BY_NAME.label("raised_by_name"),
    _SO_DATE.label("so_date"),
    _SO_NUMBER.label("so_number"),
    Product.id.label("product_id"),
    Product.product_name.label("product_name"),
    _CUSTOMER_NAME.label("customer_name"),
    Project.id.label("project_id"),
    Project.title.label("project_title"),
    ProjectSalesOrder.id.label("project_sales_order_id"),
    ProjectSalesOrder.is_pre_order.label("is_pre_order"),
    SalesOrder.id.label("core_sales_order_id"),
    Supplier.id.label("supplier_id"),
    Supplier.supplier_name.label("supplier"),
    PurchaseOrder.id.label("po_id"),
    PurchaseOrder.po_number.label("po_number"),
    _LOCATION.label("location"),
    SalesAgent.sales_agent.label("agent_code"),
    SalesAgent.person_label.label("agent_label"),
    # The handshake (`PLAN-scm-oi-handshake.md`). Every one of them on the wire, because
    # the column prints a different sentence per state and the filter counts all four.
    OrderInquiryRow.ack_state.label("ack_state"),
    OrderInquiryRow.acknowledged_at.label("acknowledged_at"),
    OrderInquiryRow.rejected_at.label("rejected_at"),
    OrderInquiryRow.rejected_reason.label("rejected_reason"),
    OrderInquiryRow.changed_at.label("changed_at"),
    _ACK_USER.name.label("acknowledged_by_name"),
    _REJECT_USER.name.label("rejected_by_name"),
)


def _dec(value: Any) -> Decimal:
    if value is None:
        return _ZERO
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value))
    except Exception:  # noqa: BLE001 - a malformed stored number is data, not a crash
        return _ZERO


def _qty_str(value: Decimal) -> str:
    """`600`, not `600.0000`. ``normalize()`` alone turns 100 into `1E+2`."""
    return format(_dec(value).normalize(), "f")


def ack_label(row: Dict[str, Any]) -> str:
    """The handshake as one printed phrase, for the export (AC-H14).

    The same four readings the column on screen carries, written out rather than coloured:
    a spreadsheet has no badge, and "Awaiting" beside a blank name is what tells a buyer
    reading the file that nobody has taken this instruction on yet.
    """
    state = row.get("ack_state") or ACK_AWAITING
    if state == ACK_ACKNOWLEDGED:
        who = row.get("acknowledged_by_name")
        when = row.get("acknowledged_at")
        stamp = when.strftime("%Y-%m-%d %H:%M") if when else ""
        return " ".join(part for part in ("Acknowledged", who, stamp) if part)
    if state == ACK_CHANGED:
        when = row.get("changed_at")
        return f"Changed {when.strftime('%Y-%m-%d')}" if when else "Changed"
    if state == ACK_REJECTED:
        reason = (row.get("rejected_reason") or "").strip()
        who = row.get("rejected_by_name")
        head = f"Rejected by {who}" if who else "Rejected"
        return f"{head}: {reason}" if reason else head
    return "Awaiting"


def month_label(month: str) -> str:
    """`2026-01` to `JAN 26`, spelled the way their tab is."""
    year, _, number = month.partition("-")
    return f"{MONTH_WORD[int(number) - 1]} {year[2:]}"


def _month_bounds(month: str) -> Tuple[date, date]:
    """The half-open range a `YYYY-MM` covers. Refuses anything that is not a month."""
    try:
        first = datetime.strptime(month.strip(), "%Y-%m").date()
    except (ValueError, AttributeError):
        raise AppException(
            422,
            f"'{month}' is not a delivery month. Use YYYY-MM.",
            code="invalid_delivery_month",
        )
    following = (
        date(first.year + 1, 1, 1) if first.month == 12 else date(first.year, first.month + 1, 1)
    )
    return first, following


def _as_day(value: str) -> date:
    try:
        return datetime.strptime(value.strip(), "%Y-%m-%d").date()
    except (ValueError, AttributeError):
        raise AppException(
            422,
            f"'{value}' is not a date. Use YYYY-MM-DD.",
            code="invalid_raised_date",
        )


class OrderInquiryWorklistService:
    """Reads, totals and exports every raised instruction, whoever it belongs to."""

    def __init__(self, db: Session):
        self.db = db

    # ------------------------------------------------------------------ query

    def _base(
        self,
        *,
        query: Optional[str] = None,
        delivery_month: Optional[str] = None,
        raised_date: Optional[str] = None,
        state: Optional[str] = None,
        project_id: Optional[str] = None,
        supplier_id: Optional[str] = None,
        raised_by: Optional[str] = None,
        linked: Optional[str] = None,
        kind: Optional[str] = None,
        ack: Optional[str] = None,
    ):
        """Every inquiry row in the company, with everything a column needs beside it.

        Every join is OUTER except the two that define a row's existence (its inquiry and
        the sales order the inquiry was raised on). An inner join to `Project` is what
        made the PROJECT/CUSTOMER column blank for adopted rows - worse, it dropped them
        from the list entirely, which is why nobody noticed the blank.
        """
        base = (
            self.db.query(OrderInquiryRow.id)
            .select_from(OrderInquiryRow)
            .join(OrderInquiry, OrderInquiry.id == OrderInquiryRow.order_inquiry_id)
            .join(
                ProjectSalesOrder,
                ProjectSalesOrder.id == OrderInquiry.project_sales_order_id,
            )
            # The revision that raised this row, and the person who confirmed it. Both
            # OUTER and both on a primary key, so a row with no decision (the amendment
            # path) or a decision whose user has since been removed still reaches the
            # list rather than dropping out of it.
            .outerjoin(
                SOSupplyDecision,
                SOSupplyDecision.id == OrderInquiryRow.supply_decision_id,
            )
            .outerjoin(User, User.id == _RAISED_BY_ID)
            .outerjoin(_ACK_USER, _ACK_USER.id == OrderInquiryRow.acknowledged_by)
            .outerjoin(_REJECT_USER, _REJECT_USER.id == OrderInquiryRow.rejected_by)
            .outerjoin(
                ProjectSalesOrderLine,
                ProjectSalesOrderLine.id == OrderInquiryRow.so_line_id,
            )
            .outerjoin(Product, Product.id == ProjectSalesOrderLine.product_id)
            # The LOCATION fallback: the line's own core sales order line, then the
            # fulfilment warehouse on that line. Joined on primary keys throughout, so
            # neither join can multiply a row.
            .outerjoin(
                SalesOrderLine,
                SalesOrderLine.id == ProjectSalesOrderLine.core_sales_order_line_id,
            )
            .outerjoin(Warehouse, Warehouse.id == SalesOrderLine.warehouse_id)
            .outerjoin(SalesOrder, SalesOrder.id == ProjectSalesOrder.so_id)
            # Who sold it. The same core sales order the SO DATE / S/O NO columns already
            # read off - `core_sales_order_line_id` is only ever set to a line of THIS
            # order (`project_so_reconciliation_service`), so this is the one join, not a
            # per-row lookup.
            .outerjoin(SalesAgent, SalesAgent.id == SalesOrder.sales_agent_id)
            .outerjoin(Project, Project.id == ProjectSalesOrder.project_id)
            .outerjoin(
                ProjectPurchaseOrder,
                ProjectPurchaseOrder.id == ProjectSalesOrder.purchase_order_id,
            )
            .outerjoin(
                ProjectParty, ProjectParty.id == ProjectPurchaseOrder.issuing_party_id
            )
            .outerjoin(Customer, Customer.id == _CUSTOMER_ID)
            .outerjoin(PurchaseOrder, PurchaseOrder.id == _PLACED_PO_ID)
            .outerjoin(Supplier, Supplier.id == PurchaseOrder.supplier_id)
        )
        if delivery_month:
            first, following = _month_bounds(delivery_month)
            base = base.filter(
                OrderInquiryRow.delivery_date >= first,
                OrderInquiryRow.delivery_date < following,
            )
        if raised_date:
            base = base.filter(_RAISED_DAY == _as_day(raised_date))
        if state:
            base = base.filter(OrderInquiryRow.state == state)
        if project_id:
            base = base.filter(ProjectSalesOrder.project_id == project_id)
        if supplier_id:
            base = base.filter(Supplier.id == supplier_id)
        if raised_by:
            # By id, off the ROW's own raiser (its revision's confirmer, the header only
            # when there is no revision). The screen picks the person from the summary's
            # own list, so this never has to guess which "Cindy" was meant.
            base = base.filter(_RAISED_BY_ID == raised_by)
        if linked:
            # WHERE the row is linked (AC-I5), which is a different question from what
            # STATE it is in: a buyer asking "what have I still not put on anything"
            # wants `none`, and one chasing shipping orders wants `spo`. `po` and `spo`
            # mean "holds at least one link of that kind", so a row split across both
            # answers to either - it genuinely is on both.
            # `any` is internal - "Unlink all" asks for every row that holds a link, and
            # the route's own whitelist offers po / spo / none to a caller.
            if linked == "any":
                base = base.filter(_HAS_ANY_LINK)
            elif linked == "po":
                base = base.filter(_HAS_PO_LINK)
            elif linked == "spo":
                base = base.filter(_HAS_SPO_LINK)
            elif linked == "none":
                base = base.filter(~_HAS_ANY_LINK)
            else:
                raise AppException(
                    422,
                    f"'{linked}' is not a link filter. Use po, spo or none.",
                    code="invalid_linked_filter",
                )
        if kind:
            # WHAT the row still needs, which is what the three cards above both views
            # ask (section 3.I2) and a different question from `linked`: a row linked 5
            # of 8 to a purchase order carries `po` AND `buy`, so it answers to either
            # card, while `linked=po` only asks where its links point. A row in one of
            # `_NOT_OWED_STATES` carries no kind at all - its quantity is not owed any
            # more, and its links are history (`links_for_rows` already hides them), so
            # counting it would tell purchasing to buy something somebody has called off
            # or already answered.
            if kind not in ("spo", "po", "buy"):
                raise AppException(
                    422,
                    f"'{kind}' is not a supply kind. Use spo, po or buy.",
                    code="invalid_kind_filter",
                )
            base = base.filter(OrderInquiryRow.state.notin_(_NOT_OWED_STATES))
            if kind == "spo":
                base = base.filter(_HAS_SPO_LINK)
            elif kind == "po":
                base = base.filter(_HAS_PO_LINK)
            else:
                base = base.filter(_UNLINKED_QTY > 0)
        if ack:
            # WHERE THE HANDSHAKE STANDS (`PLAN-scm-oi-handshake.md` section 4), which is
            # a third question beside `state` and `linked`: purchasing's own worklist is
            # "what have I not acknowledged yet", and CS's reading of the same list is
            # "what has purchasing refused". A closed set, refused rather than ignored,
            # for the same reason `state` and `kind` are - a filter nothing can equal
            # reads on screen as "no work to do".
            if ack not in ACK_FILTER_STATES:
                raise AppException(
                    422,
                    f"'{ack}' is not an acknowledgement state. Use "
                    f"{', '.join(ACK_FILTER_STATES)}.",
                    code="invalid_ack_filter",
                )
            base = base.filter(OrderInquiryRow.ack_state == ack)
        if query:
            like = f"%{query.strip()}%"
            base = base.filter(
                or_(
                    OrderInquiryRow.item_code.ilike(like),
                    OrderInquiryRow.spo_ref.ilike(like),
                    # Purchasing is asked about "OI-000123" by name; without this the row is
                    # reachable only by knowing which sales order raised it.
                    OrderInquiry.inquiry_no.ilike(like),
                    cast(_SO_NUMBER, String).ilike(like),
                    Product.product_name.ilike(like),
                    Product.product_code.ilike(like),
                    Customer.customer_name.ilike(like),
                    Project.title.ilike(like),
                    Project.project_code.ilike(like),
                    # The CS who raised it. By name, and by the FRONT of the email
                    # address rather than anywhere inside it: a buyer types "cindy",
                    # and matching `%cindy%` across a whole address would also return
                    # every row whose raiser happens to work at cindy.com.
                    User.name.ilike(like),
                    User.email.ilike(f"{query.strip()}%"),
                )
            )
        return base

    def list_rows(
        self,
        *,
        page: int = 1,
        limit: int = 50,
        sort: Optional[str] = None,
        direction: Optional[str] = "asc",
        **filters,
    ) -> Dict[str, Any]:
        """One page of the worklist, in a TOTAL and STABLE order.

        Three rules hold for every sortable column, the same three the fulfilment
        worklist holds to:

        * nulls sort LAST in BOTH directions. Postgres would put them last ascending and
          first descending, so reversing "delivery date" would answer with the rows that
          have no date at all. A missing value is not an extreme value;
        * every sort ends on the row id, so a tie breaks the same way on page 2 as on
          page 1 and a paged read neither repeats a row nor drops one;
        * an unknown column is refused here as well as at the route, so a service caller
          (a test, a script, the MCP) gets the same answer an HTTP caller gets.
        """
        field = sort or DEFAULT_SORT_FIELD
        if field not in SORTABLE_FIELDS:
            raise AppException(
                422,
                f"'{field}' is not a column this list can be sorted by.",
                code="unsortable_field",
            )
        if direction not in (None, "asc", "desc"):
            raise AppException(
                422,
                f"'{direction}' is not a sort direction. Use asc or desc.",
                code="invalid_sort_direction",
            )
        descending = direction == "desc"
        page = max(page, 1)

        base = self._base(**filters)
        total = int(
            base.with_entities(func.count(OrderInquiryRow.id)).order_by(None).scalar() or 0
        )
        column = _SORT_EXPRESSIONS[field]
        ordering = column.desc().nulls_last() if descending else column.asc().nulls_last()
        rows = (
            base.with_entities(*_COLUMNS)
            .order_by(ordering, OrderInquiryRow.id.asc())
            .offset((page - 1) * limit)
            .limit(limit)
            .all()
        )
        product_by_row, link_candidates = self._link_candidate_context(rows)
        flow = self._quantity_flow_by_so_line(rows)
        links = ProjectOrderInquiryService(self.db).links_for_rows(
            [row.id for row in rows]
        )
        return {
            "data": [
                self._serialize(row, product_by_row, link_candidates, flow, links)
                for row in rows
            ],
            "pagination": {"total": total, "page": page, "limit": limit},
            "empty": total == 0,
        }

    def _link_candidate_context(
        self, rows
    ) -> Tuple[Dict[str, Optional[str]], Dict[str, set]]:
        """`has_link_candidate` for a whole PAGE, bulk-answered once. Row id -> product id
        (the line's own reconciled product first, the item code second - the same
        precedence `ProjectOrderInquiryService._resolve_product_id` uses per row), and the
        and the products that still have something to link to, by kind."""
        by_code = {row.item_code for row in rows if not row.product_id and row.item_code}
        code_products = (
            dict(
                self.db.query(Product.product_code, Product.id)
                .filter(Product.product_code.in_(list(by_code)))
                .all()
            )
            if by_code
            else {}
        )
        product_by_row: Dict[str, Optional[str]] = {
            row.id: row.product_id or code_products.get(row.item_code) for row in rows
        }
        candidates = ProjectOrderInquiryService(self.db).link_candidate_products(
            set(product_by_row.values())
        )
        return product_by_row, candidates

    def _quantity_flow_by_so_line(self, rows) -> Dict[str, Dict[str, Decimal]]:
        """"Taken from PO" and "Remaining" for a whole PAGE, bulk-answered once (the
        captain, 20 Aug: "show the quantity, quantity taken from PO, and the remaining
        quantity, cause this is what flows to reorder planning").

        Read off the LINKS since section 3.I, which is what makes the two figures true
        again. Before it, a cascade SPLIT a line's rows and the pair was "sum the placed
        siblings" against "sum the raised siblings" - correct only because the split had
        moved the arithmetic into the row count. A row now keeps its full quantity and
        carries links, so per `so_line_id`:

        * `taken` - the sum of every LINK on the line's rows, whatever document it names;
        * `remaining` - the sum of `qty - linked` across them, which is exactly
          `scm.committed_v`'s own confirmed leg (migration 422) and therefore exactly what
          still flows to reorder planning.

        Scoped to the verbs `committed_v` counts: `ORDER` and, since part 2 section 4b,
        `ORDER_BACK`. Rows in `actioned` / `cancelled` are out, as they always were.

        A row a planning change REDIRECTED to replenish the shared pool
        (`planning_change_service._apply_placed_redirect`, the captain's ruling 21 Aug
        2026) is excluded from both: it is still real linked quantity, just not this
        line's anymore, and counting it here would read as this line's need being covered
        by a purchase order that is actually bound for the pool.

        One grouped query plus one aggregate over the links, not one query per row.
        """
        so_line_ids = {row.so_line_id for row in rows if row.so_line_id}
        if not so_line_ids:
            return {}
        linked = (
            func.coalesce(
                select(func.sum(OrderInquiryLink.qty))
                .where(OrderInquiryLink.row_id == OrderInquiryRow.id)
                .correlate(OrderInquiryRow)
                .scalar_subquery(),
                0,
            )
        ).label("linked")
        agg = (
            self.db.query(
                OrderInquiryRow.so_line_id,
                func.coalesce(func.sum(linked), 0),
                func.coalesce(
                    func.sum(func.greatest(OrderInquiryRow.qty - linked, 0)), 0
                ),
            )
            .filter(
                OrderInquiryRow.so_line_id.in_(so_line_ids),
                OrderInquiryRow.verb.in_((IV_ORDER, IV_ORDER_BACK)),
                OrderInquiryRow.state.in_(
                    (INQUIRY_PLACED, INQUIRY_PARTLY_LINKED, INQUIRY_RAISED)
                ),
                OrderInquiryRow.redirected_to_pool.is_(False),
            )
            .group_by(OrderInquiryRow.so_line_id)
            .all()
        )
        return {
            so_line_id: {"taken": _dec(taken), "remaining": _dec(remaining)}
            for so_line_id, taken, remaining in agg
        }

    def _serialize(
        self,
        row,
        product_by_row: Optional[Dict[str, Optional[str]]] = None,
        link_candidates: Optional[Dict[str, set]] = None,
        flow: Optional[Dict[str, Dict[str, Decimal]]] = None,
        links: Optional[Dict[str, List[Dict[str, Any]]]] = None,
    ) -> Dict[str, Any]:
        line_flow = (flow or {}).get(row.so_line_id, {})
        row_links = (links or {}).get(row.id, [])
        linked_qty = sum((_dec(link["qty"]) for link in row_links), _ZERO)
        return {
            "id": row.id,
            "inquiry_no": row.inquiry_no,
            "so_date": row.so_date,
            "so_number": row.so_number,
            "item_code": row.item_code,
            "product_name": row.product_name,
            "qty": _qty_str(_dec(row.qty)),
            "delivery_date": row.delivery_date,
            "project_customer": project_customer_label(
                row.customer_name, row.project_title, row.is_pre_order
            ),
            "supplier": row.supplier,
            "supplier_id": row.supplier_id,
            "po_number": row.po_number,
            "po_id": row.po_id,
            "location": row.location,
            "taken_from_po": _qty_str(line_flow.get("taken", _ZERO)),
            "remaining_open": _qty_str(line_flow.get("remaining", _ZERO)),
            # WHERE this row's quantity sits (AC-I5), off the ONE reader the per-project
            # list and the SCM sales-order detail also use.
            "links": row_links,
            "linked_qty": _qty_str(linked_qty),
            "cited_document": row.cited_document,
            "has_link_candidate": (
                ProjectOrderInquiryService.has_link_candidate(
                    row.verb, product_by_row.get(row.id), link_candidates
                )
                if product_by_row and link_candidates
                else False
            ),
            "agent_code": row.agent_code,
            "agent_label": row.agent_label,
            "state": row.state,
            "ack_state": row.ack_state,
            "acknowledged_by_name": row.acknowledged_by_name,
            "acknowledged_at": row.acknowledged_at,
            "rejected_by_name": row.rejected_by_name,
            "rejected_at": row.rejected_at,
            "rejected_reason": row.rejected_reason,
            "changed_at": row.changed_at,
            "raised_at": row.raised_at,
            "raised_by_name": row.raised_by_name,
            "verb": row.verb,
            "note": row.note,
            "project_id": row.project_id,
            "project_sales_order_id": row.project_sales_order_id,
            "core_sales_order_id": row.core_sales_order_id,
            # An adopted record is a mirror of a core sales order and has no project
            # registration; that pair is the whole distinction and the screen links on it.
            "is_adopted": bool(row.core_sales_order_id) and row.project_id is None,
        }

    # -------------------------------------------------------------- po detail

    def get_po_detail(self, po_id: str) -> Dict[str, Any]:
        """The "PO no" cell's popup: this purchase order's own header and every one of
        its lines (the captain, 20 Aug).

        Purchasing on this worklist holds `projects.projects.view`, not
        `scm.dashboard.view` - the same permission gotcha "Place on PO" already worked
        around - so this reads `purchase_orders` / `purchase_order_lines` off the
        PROJECTS router by plain ORM query rather than calling the SCM purchase-orders
        route. Both tables are `CompanyScopedMixin`, so the session-level company scope
        listener (`app.services.company_scope`) already restricts every query here to the
        caller's own company - a foreign company's PO id resolves to nothing, same as an
        unknown one.
        """
        po = self.db.query(PurchaseOrder).filter(PurchaseOrder.id == po_id).first()
        if po is None:
            raise AppException(
                status_code=404,
                message="That purchase order no longer exists.",
                code="order_inquiry_po_not_found",
            )
        supplier = (
            self.db.query(Supplier).filter(Supplier.id == po.supplier_id).first()
            if po.supplier_id
            else None
        )
        lines = (
            self.db.query(
                Product.product_code,
                Product.product_name,
                PurchaseOrderLine.qty_ordered,
                PurchaseOrderLine.qty_received,
                Warehouse.warehouse_code,
            )
            .select_from(PurchaseOrderLine)
            .outerjoin(Product, Product.id == PurchaseOrderLine.product_id)
            .outerjoin(Warehouse, Warehouse.id == PurchaseOrderLine.warehouse_id)
            .filter(PurchaseOrderLine.purchase_order_id == po.id)
            .order_by(Product.product_code.asc().nulls_last())
            .all()
        )
        return {
            "id": po.id,
            "po_number": po.po_number,
            "supplier_code": supplier.supplier_code if supplier else None,
            "supplier_name": supplier.supplier_name if supplier else None,
            "expected_date": po.expected_date,
            "status": po.status,
            "lines": [
                {
                    "sku": sku,
                    "product_name": product_name,
                    "qty_ordered": _qty_str(_dec(qty_ordered)),
                    "qty_received": _qty_str(_dec(qty_received)),
                    "remaining": _qty_str(_dec(qty_ordered) - _dec(qty_received)),
                    "location": warehouse_code,
                }
                for sku, product_name, qty_ordered, qty_received, warehouse_code in lines
            ],
        }

    # ---------------------------------------------------------------- summary

    def summary(self, **filters) -> Dict[str, Any]:
        """The strip above the list, and the controls beside it.

        Totals honour EVERY filter, month included, because they describe what is on
        screen. Each axis drops its OWN filter, because a control that empties itself the
        moment you use it cannot be used a second time.
        """
        visible = self._base(**filters)
        state_rows = (
            visible.with_entities(
                OrderInquiryRow.state,
                func.count(OrderInquiryRow.id),
                func.coalesce(func.sum(OrderInquiryRow.qty), 0),
            )
            .group_by(OrderInquiryRow.state)
            .all()
        )
        by_state = {state: 0 for state in INQUIRY_STATES}
        total_rows = 0
        total_qty = _ZERO
        for state, count, qty in state_rows:
            by_state[state] = int(count)
            total_rows += int(count)
            total_qty += _dec(qty)
        by_state["total"] = total_rows

        return {
            "total_rows": total_rows,
            "total_qty": _qty_str(total_qty),
            "by_state": by_state,
            "by_month": self._by_month({**filters, "delivery_month": None}),
            "suppliers": self._suppliers({**filters, "supplier_id": None}),
            "projects": self._projects({**filters, "project_id": None}),
            "raised_by": self._raised_by({**filters, "raised_by": None}),
            # The three cards, computed with the CARD FILTER ITSELF DROPPED, for the
            # reason every other axis here drops its own: a card that empties the two
            # beside it the moment it is pressed cannot be pressed a second time.
            "kinds": self._kinds({**filters, "kind": None}),
            # The four acknowledgement counts, computed with the ACK FILTER ITSELF
            # DROPPED, for the reason every other axis here drops its own.
            "ack": self._acks({**filters, "ack": None}),
        }

    def _acks(self, filters: Dict[str, Any]) -> Dict[str, int]:
        """How many rows sit at each acknowledgement state (AC-H4).

        ROWS, not quantity: acknowledging is a decision per instruction, and "three rows
        awaiting" is what the press says. Pre-seeded at zero so a state nothing is in is
        still offered - a filter value that disappears the moment it empties cannot be
        used to check that it emptied.
        """
        counts = {state: 0 for state in ACK_FILTER_STATES}
        rows = (
            self._base(**filters)
            .with_entities(OrderInquiryRow.ack_state, func.count(OrderInquiryRow.id))
            .group_by(OrderInquiryRow.ack_state)
            .order_by(None)
            .all()
        )
        for state, count in rows:
            if state in counts:
                counts[state] = int(count)
        return counts

    def _kinds(self, filters: Dict[str, Any]) -> Dict[str, str]:
        """Quantity per kind over every matching row (AC-I11): on SPO allocations, on
        purchase order lines, and the unlinked remainder that still has to be bought.

        SUMMED SERVER-SIDE OVER THE WHOLE MATCHING SET, never over a page, so a card
        cannot claim less than pressing it reveals. Cancelled and actioned rows are
        dropped by the same rule the `kind` filter drops them (`_NOT_OWED_STATES`), so
        the cards and the rows agree.
        """
        spo, po, buy = (
            self._base(**filters)
            .with_entities(
                func.coalesce(func.sum(_SPO_LINKED_QTY), 0),
                func.coalesce(func.sum(_PO_LINKED_QTY), 0),
                func.coalesce(func.sum(_UNLINKED_QTY), 0),
            )
            .filter(OrderInquiryRow.state.notin_(_NOT_OWED_STATES))
            .order_by(None)
            .one()
        )
        return {
            "spo": _qty_str(_dec(spo)),
            "po": _qty_str(_dec(po)),
            "buy": _qty_str(_dec(buy)),
        }

    def _by_month(self, filters: Dict[str, Any]) -> List[Dict[str, Any]]:
        month = func.to_char(OrderInquiryRow.delivery_date, "YYYY-MM")
        rows = (
            self._base(**filters)
            .with_entities(
                month.label("month"),
                func.count(OrderInquiryRow.id),
                func.coalesce(func.sum(OrderInquiryRow.qty), 0),
            )
            .filter(OrderInquiryRow.delivery_date.isnot(None))
            .group_by(month)
            .order_by(month.asc())
            .all()
        )
        return [
            {
                "month": value,
                "label": month_label(value),
                "rows": int(count),
                "qty": _qty_str(_dec(qty)),
            }
            for value, count, qty in rows
        ]

    def _suppliers(self, filters: Dict[str, Any]) -> List[Dict[str, Any]]:
        rows = (
            self._base(**filters)
            .with_entities(
                Supplier.id, Supplier.supplier_name, func.count(OrderInquiryRow.id)
            )
            .filter(Supplier.id.isnot(None))
            .group_by(Supplier.id, Supplier.supplier_name)
            .order_by(Supplier.supplier_name.asc())
            .all()
        )
        return [
            {"id": supplier_id, "label": name, "rows": int(count)}
            for supplier_id, name, count in rows
        ]

    def _projects(self, filters: Dict[str, Any]) -> List[Dict[str, Any]]:
        rows = (
            self._base(**filters)
            .with_entities(Project.id, Project.title, func.count(OrderInquiryRow.id))
            .filter(Project.id.isnot(None))
            .group_by(Project.id, Project.title)
            .order_by(Project.title.asc())
            .all()
        )
        return [
            {"id": project_id, "label": title, "rows": int(count)}
            for project_id, title, count in rows
        ]

    def _raised_by(self, filters: Dict[str, Any]) -> List[Dict[str, Any]]:
        """The people who have actually raised an inquiry, and how many rows each.

        Scoped to the rows in view rather than to `users`, deliberately: a picker built
        from the user table would list hundreds of names, almost all of which return
        nothing, and the one question this filter answers is "show me what CS raised".
        Bounded by the same fact - a company has a handful of people who confirm supply -
        so it ships whole with the summary rather than as a searched endpoint.
        """
        rows = (
            self._base(**filters)
            .with_entities(
                _RAISED_BY_ID, User.name, func.count(OrderInquiryRow.id)
            )
            .filter(_RAISED_BY_ID.isnot(None))
            .group_by(_RAISED_BY_ID, User.name)
            .order_by(User.name.asc().nulls_last())
            .all()
        )
        return [
            {"id": user_id, "label": name or "Unnamed user", "rows": int(count)}
            for user_id, name, count in rows
        ]

    # ------------------------------------------------------------- unplace all

    def unplace_all_preview(self, **filters) -> Dict[str, Any]:
        """The confirm dialog's own numbers (the captain, 21 Aug: "why i cannot unplace
        all" - the answer was the count and the scope were wrong, not that the action
        should be blocked), resolved server-side against the SAME filters `list_rows`
        reads - `state` is never one of them, because this is always about LINKED rows,
        whatever else is filtered. Since section 3.I that is a link test rather than a
        state test: a partly linked row holds links too, and leaving it out would have made
        "Unlink all" quietly refuse to give back the half a cascade had covered.

        `product_code`/`product_name` are best-effort labelling only, not a second scope:
        when every matching row resolves to the SAME product, the dialog can say which
        one; when they do not (or none resolves any), it says nothing rather than picking
        one arbitrarily. `LIMIT 2` is enough to tell "one" from "more than one" without
        pulling the whole matching set.
        """
        # `linked` is forced, so whatever the caller sent for it is dropped rather than
        # passed twice - the same treatment `state` gets at the route, and for the same
        # reason: this action is about LINKED rows however the list happens to be
        # filtered. Sending both is a TypeError, which the route turns into a 500 and the
        # dialog into "could not check linked rows".
        filters.pop("linked", None)
        visible = self._base(**filters, linked="any")
        count = int(
            visible.with_entities(func.count(OrderInquiryRow.id)).order_by(None).scalar()
            or 0
        )
        product_code: Optional[str] = None
        product_name: Optional[str] = None
        if count:
            distinct_products = (
                visible.with_entities(Product.product_code, Product.product_name)
                .distinct()
                .limit(2)
                .all()
            )
            if len(distinct_products) == 1 and distinct_products[0][0]:
                product_code, product_name = distinct_products[0]
        return {
            "count": count,
            "product_code": product_code,
            "product_name": product_name,
        }

    def unplace_all(self, *, actor_user_id: str, **filters) -> int:
        """"Unplace all" for the CURRENT worklist scope - the SAME filters `list_rows`
        reads, forced to placed. Resolved as a fresh id list against the full matching
        set, never against whatever page happened to be loaded: the worklist paginates
        server-side (`list_rows`'s own `offset`/`limit`), so a client-derived scope would
        silently miss every row behind page 1. No filters at all means every linked row
        in the company - "unlink all" with nothing narrowing it is exactly that.

        The actual write is `ProjectOrderInquiryService.unplace_rows` - this method's own
        job stops at resolving WHICH rows are in scope; the two can never disagree about
        what a linked row is because both read off the same "holds a link" predicate.
        """
        filters.pop("linked", None)
        visible = self._base(**filters, linked="any")
        row_ids = [row_id for (row_id,) in visible.all()]
        return ProjectOrderInquiryService(self.db).unplace_rows(row_ids)

    # ----------------------------------------------------------------- export

    def export_xlsx(self, **filters) -> Tuple[str, bytes]:
        """The filtered set as their own workbook: one sheet per delivery month.

        Within a sheet the rows go SUPPLIER then ITEM CODE, which is the order their own
        sheets are in and the order a buyer places in - one purchase order per factory.
        TOTAL QTY prints on the last row of each supplier-and-item-code run and totals
        that run, which is the one thing on their sheet that is arithmetic rather than
        typing. The run is the pair, not the item code alone: the same item bought from
        two factories is two purchase orders, and a total across both is a quantity
        nobody places.

        Generated per request rather than stored, exactly as the per-project export is: a
        stored file goes stale the moment supply is reconfirmed, and a stale instruction
        is the thing this replaces.
        """
        import openpyxl

        rows = self._all_rows(**filters)
        grouped: Dict[str, List[Dict[str, Any]]] = {}
        for row in rows:
            key = (
                row["delivery_date"].strftime("%Y-%m")
                if row["delivery_date"]
                else ""
            )
            grouped.setdefault(key, []).append(row)

        workbook = openpyxl.Workbook()
        # A new workbook opens with one blank sheet; every sheet here is named for a
        # month, so the blank one goes rather than sitting in front of them called
        # "Sheet".
        default_sheet = workbook.active
        if default_sheet is not None:
            workbook.remove(default_sheet)
        # Dated months in order, undated last: a row with no date is still an instruction
        # and is never dropped from the file.
        for key in sorted(month for month in grouped if month):
            self._write_sheet(workbook, month_label(key), grouped[key])
        if "" in grouped:
            self._write_sheet(workbook, EXPORT_UNDATED_SHEET, grouped[""])
        if not workbook.sheetnames:
            # An empty result is still a workbook a person can open and see the headings
            # of, rather than a file their spreadsheet refuses.
            self._write_sheet(workbook, EXPORT_TITLE, [])

        buffer = io.BytesIO()
        workbook.save(buffer)
        filename = f"order-inquiry-{date.today().isoformat()}.xlsx"
        return filename, buffer.getvalue()

    def _all_rows(self, **filters) -> List[Dict[str, Any]]:
        """The same set the list serves, unpaged, in the workbook's own order."""
        rows = (
            self._base(**filters)
            .with_entities(*_COLUMNS)
            .order_by(
                OrderInquiryRow.delivery_date.asc().nulls_last(),
                Supplier.supplier_name.asc().nulls_last(),
                OrderInquiryRow.item_code.asc().nulls_last(),
                OrderInquiryRow.id.asc(),
            )
            .all()
        )
        return [self._serialize(row) for row in rows]

    def _write_sheet(
        self, workbook, title: str, rows: Sequence[Dict[str, Any]]
    ) -> None:
        sheet = workbook.create_sheet(title=title[:31])
        sheet.append([EXPORT_TITLE])
        sheet.append(list(EXPORT_HEADINGS))
        # `￿` sorts after every real string, so a missing supplier or item code
        # lands at the end rather than at the top where a buyer would read it first.
        ordered = sorted(
            rows,
            key=lambda row: (
                (row.get("supplier") or "￿"),
                (row.get("item_code") or "￿"),
                row.get("delivery_date") or date.max,
            ),
        )
        runs = groupby(
            ordered, key=lambda row: (row.get("supplier"), row.get("item_code"))
        )
        for (_supplier, code), members in runs:
            run = list(members)
            total = sum((_dec(row.get("qty")) for row in run), _ZERO)
            for index, row in enumerate(run):
                last_of_run = index == len(run) - 1
                # Only where it says something the QTY column does not: a single-row
                # run has its own quantity as its total, and printing it twice is noise.
                run_total = float(total) if last_of_run and len(run) > 1 else None
                sheet.append(
                    [
                        row.get("so_date"),
                        row.get("so_number") or "",
                        code or "",
                        float(_dec(row.get("qty"))),
                        run_total,
                        row.get("delivery_date"),
                        row.get("project_customer") or "",
                        # Blank means nobody has placed it, exactly as it does on their
                        # sheet.
                        row.get("supplier") or "",
                        row.get("po_number") or "",
                        row.get("location") or "",
                        ack_label(row),
                    ]
                )
