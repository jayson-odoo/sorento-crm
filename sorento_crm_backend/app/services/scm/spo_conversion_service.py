""""Create SPO" - a shipment's uncovered lines become a real CRM purchase order.

`PLAN-scm-proforma-to-spo.md`'s Amendment, second decision: "Separate button after
packing-list apply. The shipment page gets a Create SPO action she presses when ready.
Suggestion logic as originally planned (match open PO lines by product / stated po_ref /
delivery date, net on hand + incoming SPO), but the BASE quantity is the PACKED qty, not the
invoice qty."

Two moves, two functions, the same shape as `container_request_service` (`build` / `send`)
and `proforma_invoice_service` (`convert_to_draft_shipment`):

  * `suggest` - a pure read. Per shipment line: the PACKED quantity (`quantity_shipped`,
    never the PI's invoiced figure - the Amendment's own correction), what an OPEN purchase
    order to the SAME supplier already covers (matched by product, PINNED to one PO when the
    line's originating PI line(s) stated a `po_ref`), and on hand + incoming SPO company-wide
  - same three figures `container_request_service._stock_context` reads, same reason: this
    is one company deciding what it still needs, not one warehouse's question. The suggested
    qty is `packed - po_covered - on_hand - incoming_spo`, floored at zero, editable. A line
    with nothing left to ask for reads as COVERED (unticked by default, one line saying why);
    a line with no supplier at all cannot convert (a container with an unattributed factory
    line, the n8n PDF path) and says so.
  * `create` - writes ONE `purchase_orders` header per SUPPLIER represented on the shipment
    (a container is routinely several factories' goods, and AutoCount POs are per supplier
    too), lines carrying the confirmed quantities, source-marked `SOURCE_SYSTEM` so it can
    never be mistaken for an AutoCount import. A container can carry MANY SPOs over its life
    (R1, `PLAN-scm-spo-planner-feedback-3sep.md`, `create`'s own "Seventh amendment"): each
    run is judged per LINE, against that line's own remainder, and refused (422) only once
    nothing is left to convert at all - never a blanket refusal the moment any SPO exists.

**Where the new SPO counts as supply.** `scm.po_ordered_v` reads `purchase_order_lines`
by STATUS and LINE_STATUS only - no `source_system` predicate - so an `active` / `open` CRM
SPO line is counted as "ordered" the moment it is created, exactly like an AutoCount import
would be (migration 337's `PO_ORDERED_V`). `scm.on_order_v` reads `spo_allocations`
exclusively and is UNCHANGED by this module on purpose: an allocation is a WAREHOUSE decision
(which location this shipment's goods land in), a separate, later, explicit step
(`allocation_suggestion_service.approve` / `SPOAllocationService.create_allocation`) that
this action does not take for her - `po_history_service`'s own SPO-history rows draw the
identical line (deliberately NOT written to `spo_allocations` either). So a freshly created
CRM SPO shows as ORDERED on every book/reorder screen immediately; it becomes INCOMING SPO
supply only once somebody allocates it to a warehouse, same as any other SPO.

**Reconciliation** (the next AutoCount book import matching this SPO by number) is explicitly
OUT OF SCOPE here, per the plan's design notes - the worksheet is the handoff, and the office
keys it into AutoCount by hand.

**Second amendment (captain, 21 Aug 00:40): the planner table.** `suggest` now answers two
further questions per line, both requested as a loading-plan-style table rather than the
first cut's flat checkbox list:

  * **Which PO covers this, earliest first.** `po_covered_qty` used to be one summed figure;
    it is now also broken down into `po_takes` - the per-PO quantities an EARLIEST-FIRST
    cascade would take, the identical discipline `project_order_inquiry_service._cascade_take`
    embodies for Place-on-PO (soonest `PurchaseOrderLine.expected_date`, then the PO's own
    number, then the line id) - COPIED here rather than imported, since that module is a large
    stateful class built around order-inquiry rows and importing it only for this one pure
    algorithm would drag its whole dependency graph into a module that has none of it (the
    same reasoning `_stock_context` above already gives for copying rather than importing
    `container_request_service`'s stock figures). The summed total is unchanged - a `need`
    (the packed qty) bigger than every candidate line combined still simply returns less than
    `need`, so `po_covered_qty` and `suggested_qty` read byte-identical to before this
    breakdown existed; only the per-PO detail is new.
  * **Which warehouse the SPO should land at.** `location_options` - one entry per candidate
    location carrying any live figure for this product (`location_stock_service.
    location_stock_for_product`, the SAME per-location reader the reorder Buy-row popup and
    the fulfilment board's `stock_detail` already use, so this screen's numbers can never
    disagree with theirs), each with its outstanding SO, on hand, and incoming SPO - the "after
    figure" (what `available` becomes once THIS SPO lands there) is left to the caller to add
    to whatever quantity it is proposing, since that quantity is edited live on screen and a
    server round-trip per keystroke would be the wrong shape for it. Ranked by the SAME
    Fulfilment Priority policy every other draw-down in this module uses
    (`priority.factors_for_demand_rows` - project earlier delivery first, then retail, never a
    second sort); `suggested_warehouse_id` is simply the top of that order. A product with no
    open demand or stock anywhere in the company (a first-time landing) falls back to the one
    "counts as available" warehouse (`allocation_suggestion_service._default_warehouse` -
    reused, not re-invented, for the identical reason that module already gives).

**One confirm, two writes - not a second approve step.** `create`'s `lines` payload gains an
optional `warehouse_id` per line. When present, the SAME action that mints the SPO also writes
the `spo_allocations` row for it (`SPOAllocationService.create_allocation`, the identical write
`allocation_suggestion_service.approve` uses) - so `scm.on_order_v` starts counting the new
supply at that warehouse the moment she presses Confirm, rather than a second screen. When
ABSENT (every caller before this amendment, and any line the buyer leaves unallocated), no
allocation is written for that line - `on_order_v` stays silent about it exactly as before,
which is the invariant `test_created_spo_lines_are_absent_from_on_order_v_until_allocated`
already locks in and this amendment does not disturb: allocation was always a decision, not a
default, and it still is - it is simply now a decision made on the SAME screen instead of a
second one, because the captain's plan for this table has her decide both there.

**Third amendment (captain live case, 21 Aug): delete the SPO, and self-heal a link that
outlives it.** He created SPOs from a draft shipment, then deleted their `spo_allocations` on
the SPO Allocations screen, and the planner tab stayed on "SPO already created" with no way
back - the SPO header and its `shipment_line_spo_link` row were both still there, only the
allocation was gone. Two moves:

  * `unwind` - the mirror of `create`. Deletes the `purchase_order_lines` + `purchase_orders`
    headers a shipment's `create` minted, the `shipment_line_spo_link` rows for the whole
    shipment, and any `spo_allocations` still hanging off those PO lines. Guarded twice: only
    `source_system == crm_spo` headers (an AutoCount import is refused, 409), and only headers
    linked to THIS shipment. Exposed on the planner as a Delete action on the already-converted
    state.
  * `_heal_stale_links` - `suggest`'s own defence against a CRM SPO removed by some path OTHER
    than `unwind` (a generic PO delete, a bad migration): a `shipment_line_spo_link` naming a
    PO/PO line that no longer exists is deleted on read, with a warning log, rather than
    trusted. A shipment left fully stale returns to `suggest`'s normal (non-converted) state; a
    shipment left partially stale (some SPOs deleted, some not) keeps `already_converted: true`
    and names only the SPOs still alive - `self_heal_note` says how many were cleared either
    way, non-null only when this run actually cleaned something up.

**Fourth amendment (captain, DB evidence, live case): the po_ref pin is supplier-scoped and
that defeats it.** PI line SRTWT7443 states po_ref `202605-S0060`, an open PO with 1,880 of
the product - but that PO is booked under the importer's name-squashed KAILU identity while
the shipment line itself carries the PI's own `400-J006` supplier code. `_po_cascade_lines`'s
pinned call ANDed `PurchaseOrder.supplier_id == supplier_id` onto the pin, so it returned
nothing and the planner read the line as covering 0 - exactly backwards from the module's own
rule, "a stated po_ref outranks inference". `_pinned_po_candidates` now resolves the STATED
number on its own (`po_number` is unique on `purchase_orders` in practice, so this is almost
always exactly one purchase order) and trusts it regardless of supplier spelling; the supplier
filter only comes back as a tiebreak on the rare/defensive case of more than one match. The
UN-pinned (product-match) path is unchanged - it still filters by supplier, because inference
with no stated document has nothing else to anchor it to.

**Fifth amendment (captain doctrine correction, 21 Aug): the arithmetic was inverted.** His own
words: "when there is PO, then we only can do SPO... it is when we got PO, then we only can
pull from the PO to form SPO." Every amendment above treated an open PO as competing supply
that REDUCED the suggested SPO qty (`packed - po_covered - on_hand - incoming_spo`). That is
backwards: an SPO is the SHIPMENT LEG of an existing PO, not a request that shrinks once a PO
happens to exist. Forming an SPO PULLS quantity FROM open PO lines; the pull IS the SPO's
composition, not a deduction from it.

  * **New arithmetic.** `suggested_qty` is now `po_covered_qty` itself - the SAME
    earliest-first `po_takes` cascade as before (soonest `expected_date`, then PO number),
    just no longer subtracted from anything. `_cascade_take`'s `need` argument is the packed
    qty, so the cascade still naturally caps at `min(packed, total pullable)` - the ARITHMETIC
    is unchanged, only what the caller DOES with the number is: it becomes the ask, not the
    remainder.
  * **On hand / incoming SPO drop out of the formula.** They stay on the response as CONTEXT
    (cheap - the same `_stock_context` query as before) but no longer net anything; a shipment
    with an open PO behind it is offered in full even while stock sits in the warehouse,
    because forming the SPO is about accounting for the PO, not about whether the goods are
    needed this minute.
  * **A packed quantity with no PO backing cannot become an SPO line.** `no_po_qty =
    max(packed - po_covered_qty, 0)` names the portion nothing open can back. When
    `po_covered_qty` is zero - nothing at all is pullable - the WHOLE line is `cannot_convert`
    (same shape as the no-supplier case, unselectable), reason `_REASON_NO_PO`. A
    PARTIALLY-backed line stays selectable at `po_covered_qty`; `no_po_qty` and a short note
    travel on `reason` so the shortfall is visible, never silently dropped.
    **SUPERSEDED by the sixth amendment below (captain's ruling, 3 Sep): a no-PO or
    partially-backed line is no longer `cannot_convert` at all - it converts at the buyer's
    typed quantity, PO-backed or not.**
  * **`covered` is gone.** It used to mean "nothing left to ask for, because a PO or stock
    already covers it" - a concept that only made sense when a PO was a deduction. There is no
    replacement flag: a line either has something pullable (`cannot_convert: false`,
    `suggested_qty = po_covered_qty`) or it does not (`cannot_convert: true`).

**Recording the pull, and the honesty question the plan asked to settle.** `create` decides,
per confirmed line, exactly which open PO line(s) the confirmed quantity draws down (the same
earliest-first cascade `suggest` already showed, re-run against LIVE data rather than trusted
from the earlier read - the standard "recompute at write time" rule every write path in this
module already follows). Two ways to record that pull were on the table:

  1. **Advance the source PO line's accounting** - `po_line.qty_received += take_qty` on every
     PO line a take draws from, the IDENTICAL write `allocation_suggestion_service.approve`
     already makes when a shipment draws down a PO line (see that module's own docstring:
     "the allocation raises `scm.on_order_v`; this makes `scm.po_ordered_v` fall"). Chosen.
  2. **Link only** - record which PO lines were touched and by how much, but leave their
     `qty_received` untouched, deferring the netting to the AutoCount book import that later
     reconciles this SPO by number.

  (1) is the honest one, and by a wide margin. Under (2), for as long as reconciliation has not
  happened - which is every day between "Create SPO" and the next AutoCount import, an interval
  this plan's own design notes already flag as open-ended - the SAME physical goods would count
  TWICE in `scm.po_ordered_v`: once as the original PO line's still-open balance, and again as
  the brand-new CRM SPO line's `qty_ordered`. Every screen that reads "ordered" (the PO book,
  the container request's own PO column, the reorder engine's outstanding-PO context) would
  overstate supply by exactly what "Create SPO" just did, silently, for however long
  reconciliation takes. (1) keeps the total byte-identical before and after: the source line's
  open balance falls by the take, the new SPO line's `qty_ordered` rises by the same take, nets
  to zero movement in the company-wide "ordered" total - the conversion re-attributes an
  existing order to its shipment, it does not conjure a second one.

  **How the pull is recorded, without a new table.** `shipment_line_spo_link` already carries a
  UNIQUE constraint on `inbound_shipment_line_id` (migration 406) - one row per shipment line,
  ever - so a shipment line that cascades across MORE than one PO line cannot get a second link
  row for its second source. Rather than a migration to lift that (out of scope for this pass,
  flagged below), the new CRM SPO's own `purchase_order_lines.source_ref` - a free `String`
  column, unused by any `crm_spo` line before this - now carries the take breakdown as JSON:
  `[{"po_line_id": ..., "qty": ...}, ...]`, one entry per source PO line the confirmed quantity
  drew from. This is what makes (1) SAFELY reversible: `unwind` parses it back and un-advances
  exactly what `create` advanced, per source line, before deleting the CRM SPO rows - without
  it, deleting a CRM SPO would have left the source PO permanently short by whatever it lent,
  a real loss with no SPO left to explain it. A dedicated per-take link table (queryable in SQL
  without parsing JSON) is the natural follow-up if this trail needs to be machine-readable
  from outside this module - noted, not built, here.

**Sixth amendment (captain's ruling, 3 Sep): the PO cap is removed.** The fifth amendment's own
"a packed quantity with no PO backing cannot become an SPO line" is corrected again - what she
asked for was never conditional on a PO existing to back it; a PO only ever decided how much of
the ask this module could ALREADY reconcile against AutoCount's own book. `create`'s `need =
min(requested, packed)` is now the SPO line's quantity outright, not capped a second time at
`po_covered_qty`. The PO cascade still runs exactly as the fifth amendment describes and still
PULLS from - and advances - only the open PO line(s) it actually reaches; whatever `need` is left
over once the cascade stops (`no_po_qty = need - po_covered_qty`, floored at zero) is written onto
the SAME new SPO line with no pull behind it - `source_ref.pulls` still names only the covered
part, `source_ref.no_po_qty` names the rest. `cannot_convert` no longer means "nothing open to
pull from"; it is true ONLY for a line with no supplier at all (the n8n PDF path, unchanged) - a
line with a supplier and zero open PO is now a perfectly convertible line, simply one the office
will have to raise a fresh PO for after the fact. `suggested_qty` keeps reading `po_covered_qty` -
not `need` - because it is the DEFAULT the screen's input starts at, not a cap; the buyer can type
past it. `_REASON_NO_PO` stays as informational text on a partially- or un-backed line (never a
skip reason in `create` anymore), reworded so it no longer implies the line is blocked.
"""
from __future__ import annotations

import json
import logging
import uuid
from datetime import date as _date
from decimal import Decimal
from io import BytesIO
from typing import Any, Optional, Sequence

from sqlalchemy import func, text
from sqlalchemy.orm import Session

from app.models.order import Customer, SalesOrder, SalesOrderLine
from app.models.procurement import (
    InboundShipment,
    InboundShipmentLine,
    PurchaseOrder,
    PurchaseOrderLine,
    SPOAllocation,
    Supplier,
)
from app.models.inventory import Warehouse
from app.models.projects import Project
from app.models.sales_agent import SalesAgent
from app.models.scm import ProformaInvoiceLine, ProformaInvoiceShipmentLink, ShipmentLineSpoLink
from app.services.company_scope_sql import company_sql_predicate
from app.services.error_handler import AppException
from app.services.numbering_service import NumberingService
from app.services.scm.demand_class import PROJECT as _DEMAND_CLASS_PROJECT
from app.services.scm.pool_predicate import ACTIVE_SITE_POOL_SQL
from app.services.scm.supplier_scope import is_uuid as _is_uuid

logger = logging.getLogger(__name__)

#: The CRM-originated marker on `purchase_orders.source_system` / `purchase_order_lines.
#: source_system`. Distinct from every AutoCount import stamp (`scm_po_history`,
#: `scm_spo_history`, `scm_upload`) and from the reorder engine's own draft marker
#: (`scm_recommendation`) - see the module docstring for every consumer this was checked
#: against (`po_ordered_v`, `on_order_v`, `purchase_order_service._source_label`,
#: `outstanding_import_service`'s history guard).
SOURCE_SYSTEM = "crm_spo"

#: A CRM SPO counts as "ordered" the moment it exists (the module docstring) - so it is
#: created ACTIVE + OPEN, never a draft a later Confirm step would have to promote. Mirrors
#: `purchase_order_service._ON_ORDER_STATUSES` / `_OPEN_LINE_STATUS`.
_STATUS = "active"
_LINE_STATUS = "open"

#: `NumberingService` doc_type for the CRM SPO's own series, kept distinct from every
#: AutoCount pattern (`######-S####`, `SPO-####/##-####`) and from the CRM's own canonical
#: PO series (`PO-{year}/{month}-####`, `decision_service`) so an AutoCount import can never
#: collide with a number this module minted. Falls back to a random suffix when no numbering
#: rule is configured, same shape as `proforma_invoice_service._draft_shipment_number`.
_NUMBER_DOC_TYPE = "purchase_order_crm_spo"
_NUMBER_PREFIX = "CRM-SPO"

_REASON_NO_SUPPLIER = "No supplier recorded on this shipment line, so it cannot be added to an SPO."
_REASON_NOT_SELECTED = "Not selected."
#: The doctrine correction's own line, verbatim from the captain - a packed quantity nothing
#: open can back cannot become an SPO line at all, same shape as the no-supplier case.
_REASON_NO_PO = (
    "No open PO to pull from; the SPO line is written without PO backing."
)


def _uuid() -> str:
    return str(uuid.uuid4())


def _f(value: Any) -> Optional[float]:
    return None if value is None else float(value)


def _g(value: float) -> str:
    """Trim a quantity for a sentence - `12.0` reads `12`, `12.5` reads `12.5`."""
    return f"{value:g}"


def _qty_str(value: Any) -> str:
    """`600`, not `600.0000` - and never `1.23457e+06`, which is where `_g`'s `%g` drifts
    past six figures (F6, review round). Two `linked_to` entries in one sales-order line's
    cell (S7's SPO coverage beside a project's own order-inquiry link) must format the SAME
    quantity alike, so this is a copy of `project_order_inquiry_service._qty_str` - the
    reasoning the module docstring already gives for `_cascade_take` applies here too: that
    module is a large stateful class, and importing it for one three-line helper would drag
    its whole dependency graph into a module that has none of it.
    """
    try:
        dec = value if isinstance(value, Decimal) else Decimal(str(value))
    except Exception:  # noqa: BLE001 - a malformed stored number is data, not a crash
        dec = Decimal(0)
    return format(dec.normalize(), "f")


def _shipment_or_404(db: Session, shipment_id: str) -> InboundShipment:
    if not _is_uuid(shipment_id):
        raise AppException(404, "Inbound shipment not found")
    row = db.query(InboundShipment).filter(InboundShipment.id == shipment_id).one_or_none()
    if row is None:
        raise AppException(404, "Inbound shipment not found")
    return row


def _existing_links(db: Session, shipment_id: str) -> list[ShipmentLineSpoLink]:
    return (
        db.query(ShipmentLineSpoLink)
        .filter(ShipmentLineSpoLink.inbound_shipment_id == shipment_id)
        .all()
    )


def _existing_spos(db: Session, shipment_id: str) -> list[dict]:
    """Every CRM SPO ever created off this shipment (R1: "many SPOs per container" - one row
    per `create` RUN that produced something, not one per shipment). Oldest first, so the
    planner's grid reads in the order the SPOs were made. `line_count` / `total_qty` are
    this PO's OWN lines - one create run always writes exactly one header per supplier, so
    they are that header's whole content, not a re-derivation of the shipment link rows."""
    rows = (
        db.query(PurchaseOrder)
        .join(ShipmentLineSpoLink, ShipmentLineSpoLink.purchase_order_id == PurchaseOrder.id)
        .filter(ShipmentLineSpoLink.inbound_shipment_id == shipment_id)
        .distinct()
        .order_by(PurchaseOrder.created_at.asc())
        .all()
    )
    return [
        {
            "purchase_order_id": str(po.id),
            "po_number": po.po_number,
            "supplier_id": str(po.supplier_id) if po.supplier_id else None,
            "supplier_name": po.supplier.supplier_name if po.supplier else None,
            "line_count": len(po.lines),
            "total_qty": sum(float(ln.qty_ordered or 0) for ln in po.lines),
            "created_at": po.created_at.isoformat() if po.created_at else None,
            "status": po.status,
        }
        for po in rows
    ]


def _already_spo_qty_by_line(
    db: Session, shipment_line_ids: Sequence[str]
) -> dict[str, dict[str, Any]]:
    """Per shipment line, across EVERY `create` run against this shipment so far (R1): how
    much of it is already an SPO line's own `qty_ordered`, and which SPO number(s) - oldest
    first - are responsible. `ShipmentLineSpoLink` can carry SEVERAL matched rows per line
    now (migration 469 dropped the one-row-per-line-ever unique index a partial conversion
    would otherwise violate), so this sums every one of them rather than trusting there is
    only one - the same "recompute, never trust a single row" discipline the rest of this
    module already follows for a PO's own open balance."""
    ids = [str(sid) for sid in shipment_line_ids if sid]
    if not ids:
        return {}
    rows = (
        db.query(
            ShipmentLineSpoLink.inbound_shipment_line_id,
            PurchaseOrderLine.qty_ordered,
            PurchaseOrder.po_number,
        )
        .join(
            PurchaseOrderLine,
            PurchaseOrderLine.id == ShipmentLineSpoLink.purchase_order_line_id,
        )
        .join(PurchaseOrder, PurchaseOrder.id == PurchaseOrderLine.purchase_order_id)
        .filter(
            ShipmentLineSpoLink.inbound_shipment_line_id.in_(ids),
            ShipmentLineSpoLink.purchase_order_line_id.isnot(None),
        )
        .all()
    )
    out: dict[str, dict[str, Any]] = {}
    for line_id, qty_ordered, po_number in rows:
        entry = out.setdefault(str(line_id), {"qty": 0.0, "spo_numbers": []})
        entry["qty"] += float(qty_ordered or 0)
        if po_number and po_number not in entry["spo_numbers"]:
            entry["spo_numbers"].append(po_number)
    for entry in out.values():
        entry["spo_numbers"].sort()
    return out


def _heal_stale_links(
    db: Session, shipment_id: str, links: list[ShipmentLineSpoLink]
) -> tuple[list[ShipmentLineSpoLink], int]:
    """Verify every MATCHED link still names a live PO/PO line, and self-heal the ones that do
    not (the captain's second ask, 21 Aug: allocations deleted elsewhere left the planner stuck
    on "SPO already created" with nothing behind it). A CRM SPO removed by any path other than
    `unwind` above - a generic PO delete, a bad migration - leaves `shipment_line_spo_link`
    behind: both FKs here are `ON DELETE SET NULL` (migration 406), so Postgres itself already
    clears `purchase_order_id` / `purchase_order_line_id` the moment the row they name is
    deleted, however that deletion happened. The tell is the COMBINATION this can only ever
    reach by that route: `unmatched_reason IS NULL` (so it was written as a MATCH, never a
    skip - every skip `create` writes always carries a reason) with `purchase_order_id IS NULL`
    (so whatever it matched is now gone). Trusting the row at face value is exactly what left
    the captain's planner stuck; this re-derives the truth from what Postgres already did to it
    rather than repeating that mistake. A `purchase_order_id` that is still present is
    re-queried for real (belt and braces against a schema where the delete path did not go
    through the FK at all), same for a lone orphaned `purchase_order_line_id` under a header
    that is otherwise still alive.

    Stale rows are DELETED here, on read, so the very next `suggest` answers honestly - back to
    its normal (non-`already_converted`) state when every link was stale, or naming only the
    SPOs still alive when some were (the partially-alive case: some SPOs deleted, some not).

    Returns the surviving links and how many were cleaned up, so the caller can say so.
    """
    candidates = [l for l in links if l.unmatched_reason is None]
    if not candidates:
        return links, 0

    po_ids = {str(l.purchase_order_id) for l in candidates if l.purchase_order_id}
    line_ids = {str(l.purchase_order_line_id) for l in candidates if l.purchase_order_line_id}
    live_po_ids = (
        {str(r[0]) for r in db.query(PurchaseOrder.id).filter(PurchaseOrder.id.in_(po_ids)).all()}
        if po_ids
        else set()
    )
    live_line_ids = (
        {
            str(r[0])
            for r in db.query(PurchaseOrderLine.id).filter(PurchaseOrderLine.id.in_(line_ids)).all()
        }
        if line_ids
        else set()
    )

    alive: list[ShipmentLineSpoLink] = []
    stale: list[ShipmentLineSpoLink] = []
    for link in links:
        if link.unmatched_reason is not None:
            alive.append(link)  # a genuine skip - nothing named, nothing to verify
            continue
        if link.purchase_order_id is None and link.purchase_order_line_id is None:
            # The SET-NULL signature: this was a MATCH (no reason) with nothing left to
            # name - its PO was deleted by some path other than `unwind`.
            stale.append(link)
            continue
        po_ok = link.purchase_order_id is None or str(link.purchase_order_id) in live_po_ids
        line_ok = link.purchase_order_line_id is None or str(link.purchase_order_line_id) in live_line_ids
        (alive if (po_ok and line_ok) else stale).append(link)

    if stale:
        logger.warning(
            "shipment_line_spo_link: %d stale row(s) for shipment %s point at a deleted "
            "purchase order/line - cleaning up on read",
            len(stale),
            shipment_id,
        )
        for link in stale:
            db.delete(link)
        db.flush()

    return alive, len(stale)


def _po_refs_for_line(db: Session, shipment_line_id: str) -> set[str]:
    """The `po_ref`(s) named on the PI line(s) this shipment line was drafted from, if any.

    A shipment line reached through the draft-from-PI convert can aggregate several PI
    lines (same product, same supplier, across invoices - `proforma_invoice_service`'s own
    grouping); a shipment reached from a real packing-list upload has none at all, and this
    returns empty for it.
    """
    rows = (
        db.query(ProformaInvoiceLine.po_ref)
        .join(
            ProformaInvoiceShipmentLink,
            ProformaInvoiceShipmentLink.proforma_invoice_line_id == ProformaInvoiceLine.id,
        )
        .filter(
            ProformaInvoiceShipmentLink.inbound_shipment_line_id == shipment_line_id,
            ProformaInvoiceLine.po_ref.isnot(None),
        )
        .all()
    )
    return {r[0].strip() for r in rows if r[0] and r[0].strip()}


def _cascade_take(
    lines: list[tuple[PurchaseOrderLine, PurchaseOrder, float]], need: float
) -> list[tuple[PurchaseOrderLine, PurchaseOrder, float]]:
    """Earliest-first cascade take - the SAME discipline `project_order_inquiry_service.
    _cascade_take` embodies for Place-on-PO: `min(open balance, still needed)` off each
    candidate in the order it is given, stopping once `need` is covered. `lines` is already
    sorted earliest-first by the caller (`_po_cascade_lines`); each item carries its own open
    quantity. A `need` bigger than every candidate combined simply returns less than `need` -
    the caller decides what an unfinished walk means (here, the SPO still asks for the rest)."""
    still = need
    takes: list[tuple[PurchaseOrderLine, PurchaseOrder, float]] = []
    for line, po, available in lines:
        if still <= 0:
            break
        take = min(available, still)
        if take > 0:
            takes.append((line, po, take))
            still -= take
    return takes


def _open_line_rows(q) -> list[tuple[PurchaseOrderLine, PurchaseOrder, float]]:
    """Shared tail of both cascade lookups below: oldest DOCUMENT first, `available` computed
    for every matching line - INCLUDING one with nothing left (`available == 0`, S5).

    A zero-available line was dropped here before S5, so a line another SPO had already
    pulled down to zero simply vanished from the candidate set - `_cascade_take` never took
    from it either way (`min(0, still)` is always 0), but `suggest`'s taken-only rows need to
    SEE it to say "this line is fully pulled by SPO-1" instead of nothing. Keeping it here
    rather than a second, unfiltered query means the two readers can never see a different
    candidate set for the same shipment line (`_candidate_lines_for_line`).

    Ordered by the purchase order's own `issue_date` since the captain's Q8 ruling (26 Aug):
    the buyer draws down the order they raised first, and a line's expected date is when it
    is due, not when it was ordered. The two disagree routinely - a PO raised in January can
    be due after one raised in March - and the old ordering answered the wrong question,
    which is why "which PO covers this" read differently here than in the PO book.
    `order_date` is NOT the column: it does not exist on `purchase_orders`.
    """
    rows = q.order_by(
        PurchaseOrder.issue_date.asc().nulls_last(),
        PurchaseOrderLine.expected_date.asc().nulls_last(),
        PurchaseOrder.po_number.asc(),
        PurchaseOrderLine.id.asc(),
    ).all()
    out: list[tuple[PurchaseOrderLine, PurchaseOrder, float]] = []
    for line, po in rows:
        available = max(float(line.qty_ordered or 0) - float(line.qty_received or 0), 0.0)
        out.append((line, po, available))
    return out


def _po_cascade_lines(
    db: Session, supplier_id: str, product_id: str
) -> list[tuple[PurchaseOrderLine, PurchaseOrder, float]]:
    """Open PO lines to this supplier for this product, EARLIEST `expected_date` FIRST, then
    document sequence - `project_order_inquiry_service._candidates_for_row`'s own
    ordering, so a "which PO covers this" answer here can never disagree with what the
    Place-on-PO cascade would say about the same purchase order. The PRODUCT-match (inference)
    path only - a STATED po_ref is resolved by `_pinned_po_candidates` below, which does NOT
    filter by supplier (see that function's docstring for why the two paths diverge)."""
    q = (
        db.query(PurchaseOrderLine, PurchaseOrder)
        .join(PurchaseOrder, PurchaseOrder.id == PurchaseOrderLine.purchase_order_id)
        .filter(
            PurchaseOrder.supplier_id == supplier_id,
            PurchaseOrder.status.notin_(("draft", "draft_recommendation")),
            PurchaseOrderLine.product_id == product_id,
            PurchaseOrderLine.line_status == "open",
        )
    )
    return _open_line_rows(q)


def _pinned_po_candidates(
    db: Session, po_number: str, supplier_id: str, product_id: str
) -> list[tuple[PurchaseOrderLine, PurchaseOrder, float]]:
    """Resolve a STATED po_ref - the module's own rule, "a stated po_ref outranks inference" -
    WITHOUT filtering by supplier first (captain, DB evidence, live case): a PI's supplier name
    and the CRM book's supplier for the same factory can be spelled differently (an importer's
    name-squashed identity vs the PI's own code) - PI line SRTWT7443 states po_ref
    202605-S0060, an open PO with 1,880 of the product, booked under KAILU, while the shipment
    line itself carries supplier 400-J006. Filtering the PIN by supplier as well as by number
    defeated the pin in exactly the case it exists for, and the planner read the line as
    covering 0.

    `po_number` is UNIQUE on `purchase_orders` in practice, so this almost always resolves to
    exactly one purchase order - trust the stated document then, regardless of supplier
    spelling. If it somehow resolves to MORE than one (the schema does not forbid it even
    though the live data does), the supplier is the tiebreak: narrow to the match(es) that
    also agree with it, the same filter the inference path always applies.
    """
    matches = db.query(PurchaseOrder).filter(PurchaseOrder.po_number == po_number).all()
    if len(matches) > 1:
        matches = [po for po in matches if str(po.supplier_id) == str(supplier_id)]
    po_ids = [po.id for po in matches]
    if not po_ids:
        return []
    q = (
        db.query(PurchaseOrderLine, PurchaseOrder)
        .join(PurchaseOrder, PurchaseOrder.id == PurchaseOrderLine.purchase_order_id)
        .filter(
            PurchaseOrderLine.purchase_order_id.in_(po_ids),
            PurchaseOrder.status.notin_(("draft", "draft_recommendation")),
            PurchaseOrderLine.product_id == product_id,
            PurchaseOrderLine.line_status == "open",
        )
    )
    return _open_line_rows(q)


def _candidate_lines_for_line(
    db: Session, ln: InboundShipmentLine
) -> tuple[Optional[str], list[tuple[PurchaseOrderLine, PurchaseOrder, float]]]:
    """The PO lines a Create SPO for this shipment line could draw from - pin first (a
    stated `po_ref` outranks inference), else the plain product/supplier match - shared by
    `_match_takes_for_line`'s cascade AND `suggest`'s taken-only rows (S5), so the SAME
    resolution decides both what the cascade may walk and which lines the plan even shows.
    Every matching line comes back, INCLUDING one with `available == 0` (`_open_line_rows`,
    S5) - `_cascade_take` never takes from one anyway, but a taken-only row needs to see it.

    Returns `(resolved_by, candidates)`: `resolved_by` names WHICH resolution produced the
    list ('po_ref' when a stated ref resolved to at least one line, 'product' when the plain
    match did, `None` when neither named a line at all) - a statement of PROVENANCE, not a
    promise that the cascade will actually take anything from it (a pin can resolve to lines
    with nothing open); `_match_takes_for_line` still decides `matched_by` on its own terms.
    """
    supplier_id = str(ln.supplier_id) if ln.supplier_id else None
    if supplier_id is None:
        return None, []
    po_refs = _po_refs_for_line(db, str(ln.id))
    if len(po_refs) == 1:
        pinned = _pinned_po_candidates(db, next(iter(po_refs)), supplier_id, str(ln.product_id))
        if pinned:
            return "po_ref", pinned
    product_lines = _po_cascade_lines(db, supplier_id, str(ln.product_id))
    return ("product" if product_lines else None), product_lines


def _match_takes_for_line(
    db: Session,
    ln: InboundShipmentLine,
    need: float,
    only_po_lines: Optional[set[str]] = None,
) -> tuple[Optional[str], list[tuple[PurchaseOrderLine, PurchaseOrder, float]]]:
    """The one PO-matching decision this module makes, shared by `suggest` (a read against
    `need = packed`) and `create` (a FRESH read, at write time, against `need = the confirmed
    qty` - never the earlier `suggest` numbers, so a PO that moved between the two calls is
    never double-spent). Pin first (a stated `po_ref` outranks inference), fall back to a
    plain product/supplier match. Returns `(matched_by, takes)` - `takes` empty and
    `matched_by` `None` when nothing is pullable at all."""
    supplier_id = str(ln.supplier_id) if ln.supplier_id else None
    if supplier_id is None or need <= 0:
        return None, []

    def _keep(candidates):
        """The candidate lines the buyer left ticked, BEFORE the cascade walks them.

        Filtering the cascade's OUTPUT instead was the bug: with 80 packed against PO A
        (50 open) and PO B (100 open), the walk takes 50 and 30, and dropping A's take left
        30 - when B alone can cover all 80. The tick changes which lines are available, so
        the walk has to run again over what is left, not have a slice cut out of it.
        """
        if only_po_lines is None:
            return candidates
        return [c for c in candidates if str(c[0].id) in only_po_lines]

    resolved_by, candidates = _candidate_lines_for_line(db, ln)
    matched_by: Optional[str] = None
    takes: list[tuple[PurchaseOrderLine, PurchaseOrder, float]] = []
    if resolved_by == "po_ref":
        pinned_lines = _keep(candidates)
        if pinned_lines:
            takes = _cascade_take(pinned_lines, need)
            # F1 (review round): a pin can resolve to a line with NOTHING open (S5 stopped
            # dropping `available == 0` lines from the candidate set), so `matched_by` is
            # `po_ref` only when the cascade actually took something from it - an exhausted
            # pin otherwise blocked the product-match fallback below from ever running, and
            # the line read `cannot_convert` with an open PO to the same supplier sitting
            # right there.
            if takes:
                matched_by = "po_ref"
    if matched_by is None:
        # Either the resolution already landed on the product match, or a po_ref pin
        # existed but the buyer ticked out every one of its lines - the ORIGINAL fallback:
        # try the plain product/supplier match fresh, never the pin's own candidates twice.
        product_lines = candidates if resolved_by in (None, "product") else _po_cascade_lines(
            db, supplier_id, str(ln.product_id)
        )
        product_lines = _keep(product_lines)
        takes = _cascade_take(product_lines, need)
        if takes:
            matched_by = "product"
    return matched_by, takes


def _spo_pulls_by_po_line(
    db: Session, po_line_ids: Optional[list[str]] = None
) -> dict[str, list[dict]]:
    """Every CRM SPO line that has pulled from one of these purchase-order lines, oldest SPO
    number first - `{po_line_id: [{"spo_number", "qty"}]}` (S5, `po_takes[].taken_qty` /
    `taken_by`).

    Read off the source line's own record: `create` advances `qty_received` on the PO line
    AND writes `source_ref.pulls` on the SPO line it created (`parse_source_ref`), so this is
    a read of a fact already on file, not a second table to keep in step. Kept separate from
    `PurchaseOrderService._spo_takes_of` (the PO detail's own occupancy panel), which needs
    the LANDING too (packing list, warehouses, arrival date) this helper does not fetch -
    the same one-query-per-purpose reasoning `_stock_context` above gives for not sharing a
    query across two different-shaped answers.

    `po_line_ids=None` (F7, review round) answers for EVERY line any CRM SPO has ever pulled
    from, not a chosen set - `suggest` calls this ONCE per shipment, for every candidate
    across every shipment line, rather than once PER LINE the way it used to: each call scans
    the whole `crm_spo` book regardless of how many ids it is asked about, so calling it once
    per shipment line multiplied that scan by the shipment's own line count for no reason.
    """
    if po_line_ids is not None and not po_line_ids:
        return {}
    wanted = set(po_line_ids) if po_line_ids is not None else None
    rows = (
        db.query(PurchaseOrderLine.source_ref, PurchaseOrder.po_number)
        .join(PurchaseOrder, PurchaseOrder.id == PurchaseOrderLine.purchase_order_id)
        .filter(
            PurchaseOrderLine.source_system == SOURCE_SYSTEM,
            PurchaseOrderLine.source_ref.isnot(None),
        )
        .all()
    )
    out: dict[str, list[dict]] = {}
    for source_ref, po_number in rows:
        for po_line_id, qty in parse_source_ref(source_ref)["pulls"]:
            if wanted is None or po_line_id in wanted:
                out.setdefault(po_line_id, []).append({"spo_number": po_number, "qty": qty})
    for entries in out.values():
        entries.sort(key=lambda e: e["spo_number"] or "")
    return out


def _spo_so_coverage_rows(
    db: Session, *, product_id: Optional[str] = None
) -> list[tuple[str, str, float, str]]:
    """Every `(so_line_id, spo_number, qty, po_line_id)` a CRM SPO line's own
    `source_ref.so_coverage` names, oldest SPO number first - the shared row scan behind
    `_spo_cover_by_so_line` (per-PRODUCT, S5's planner `taken_by`) and `coverage_for_so_lines`
    (per-SO-LINE-ID set, S7's retail sales-order `linked_to`), so the planner's `taken_by`
    and the sales order's "Linked to" column can never name a different SPO for the same
    line. `po_line_id` is the SPO's OWN `purchase_order_lines.id` - the same id
    `spo_allocations.po_line_id` carries - so a caller can join to the allocation for a
    location / arrival date (`coverage_for_so_lines`).
    """
    q = (
        db.query(PurchaseOrderLine.id, PurchaseOrderLine.source_ref, PurchaseOrder.po_number)
        .join(PurchaseOrder, PurchaseOrder.id == PurchaseOrderLine.purchase_order_id)
        .filter(
            PurchaseOrderLine.source_system == SOURCE_SYSTEM,
            PurchaseOrderLine.source_ref.isnot(None),
        )
    )
    if product_id is not None:
        q = q.filter(PurchaseOrderLine.product_id == product_id)
    out: list[tuple[str, str, float, str]] = []
    for po_line_id, source_ref, po_number in q.all():
        for so_line_id, qty in parse_source_ref(source_ref)["so_coverage"]:
            out.append((so_line_id, po_number, qty, str(po_line_id)))
    out.sort(key=lambda row: row[1] or "")
    return out


def _spo_cover_by_so_line(db: Session, product_id: str) -> dict[str, list[dict]]:
    """Every CRM SPO line pointed at a retail sales-order line for THIS product, oldest SPO
    number first - `{so_line_id: [{"spo_number", "qty"}]}` (S5, `so_coverage[].taken_qty` /
    `taken_by`).

    The retail half writes no link row - the links table hangs off an order-inquiry row and
    a retail line has none - so `source_ref.so_coverage` is the only record that the
    quantity has already been promised. Without it the same 30 pieces were offered, and
    default-ticked, on every container of the month.

    A thin grouping over `_spo_so_coverage_rows`, which does the actual read. An unwound SPO
    takes its lines with it, so its record disappears with it, which is correct: the promise
    was undone.
    """
    out: dict[str, list[dict]] = {}
    for so_line_id, spo_number, qty, _po_line_id in _spo_so_coverage_rows(
        db, product_id=product_id
    ):
        out.setdefault(so_line_id, []).append({"spo_number": spo_number, "qty": qty})
    return out


def coverage_for_so_lines(db: Session, so_line_ids: Sequence[str]) -> dict[str, list[dict]]:
    """`SalesOrderLineLink`-shaped entries for every retail sales-order line a CRM SPO
    already covers (S7, AC-G2..G6) - `{so_line_id: [SalesOrderLineLink dict]}`.

    `sales_order_service`'s "Linked to" column has something to read for a PROJECT line (an
    `order_inquiry_links` row) but nothing for a RETAIL one: the retail tick is recorded only
    on the covering SPO line's own `source_ref.so_coverage` - the links table hangs off an
    inquiry row and a retail line has none - so without this a retail line an SPO already
    promised reads "-" forever.

    Shares its row scan with `_spo_cover_by_so_line` (`_spo_so_coverage_rows`), so the
    planner's `taken_by` and this column can never name a different SPO for the same line.

    `location` is the covering SPO line's OWN `spo_allocations` row - the warehouse the
    container is actually going to, first by allocation id when the SPO line was split
    across several - `None` while nothing has been allocated yet. `expected_date` is that
    allocation's shipment ETA (`estimated_arrival_date`), `None` when the allocation carries
    no booked shipment, or there is no allocation at all. `line_label` and `late`/`late_days`
    are the fields `links_for_rows` states for an order-inquiry link that a retail line has
    no equivalent of: no per-line numbering to print, and no `required_date` promise on the
    retail line itself to compare an SPO's arrival against.
    """
    wanted = {sid for sid in so_line_ids if sid}
    if not wanted:
        return {}
    rows = [row for row in _spo_so_coverage_rows(db) if row[0] in wanted]
    if not rows:
        return {}

    po_line_ids = {po_line_id for _so, _spo, _qty, po_line_id in rows}
    alloc_by_line: dict[str, tuple[Optional[str], Optional[_date]]] = {}
    for po_line_id, warehouse_code, eta in (
        db.query(
            SPOAllocation.po_line_id, Warehouse.warehouse_code,
            InboundShipment.estimated_arrival_date,
        )
        .outerjoin(Warehouse, Warehouse.id == SPOAllocation.warehouse_id)
        .outerjoin(InboundShipment, InboundShipment.id == SPOAllocation.inbound_shipment_id)
        .filter(SPOAllocation.po_line_id.in_(po_line_ids))
        .order_by(SPOAllocation.id.asc())
        .all()
    ):
        # First allocation for this SPO line wins (a split line has several) - the docstring's
        # own "first if several".
        alloc_by_line.setdefault(str(po_line_id), (warehouse_code, eta))

    out: dict[str, list[dict]] = {}
    for so_line_id, spo_number, qty, po_line_id in rows:
        warehouse_code, eta = alloc_by_line.get(po_line_id, (None, None))
        out.setdefault(so_line_id, []).append({
            "kind": "spo",
            "document": spo_number,
            "line_label": None,
            "qty": _qty_str(qty),
            "location": warehouse_code,
            "expected_date": eta.isoformat() if eta else None,
            "late": False,
            "late_days": None,
        })
    return out


#: The site-pool test, from the one module that spells it (`pool_predicate`). It was copied
#: here, and into two other files, on the reasoning `_stock_context` below gives for copying
#: its whole query - but this one line is not that: both cells this module prints open a
#: dialog counting ACTIVE POOL rows only (`location_stock_service.location_stock_for_product`
#: for On hand, `container_request_drill` for Incoming SPO), so the cells count the same
#: locations or they cannot foot, and a rule that must be identical is one rule.
_ACTIVE_POOL = ACTIVE_SITE_POOL_SQL


def _stock_context(db: Session, product_ids: list[str]) -> dict[str, dict]:
    """On hand + incoming SPO, per product, at ACTIVE SITE POOLS - COPIED from
    `container_request_service._stock_context` rather than imported (same reasoning that
    module gives for copying `loading_plan_service._catalogue_cbm`: two lanes touching the
    same file is a worse cost than a few duplicated lines). Same figures, same views, so this
    screen and the container request can never disagree about what is on hand or incoming.

    CONTEXT ONLY since the doctrine correction (module docstring, fifth amendment) - neither
    figure feeds `suggested_qty` any more. Kept because it is cheap (one query, already paid
    for by every earlier version of this module) and still useful to see beside the ask.

    Context still has to foot to the dialog it opens (AC-G3: "sum = cell"). Both figures used
    to sum every warehouse the net-position view names, while the On hand lightbox lists
    active pool locations only and the Incoming SPO dialog filters `w.is_active AND _POOL` -
    so the planner printed one number and the reader who clicked it landed on another. Closed
    locations and project bins leave the cells here for the same reason they left the
    container request's (`_ACTIVE_POOL` above)."""
    if not product_ids:
        return {}
    prod_scope, prod_params = company_sql_predicate(db, "p.company_id", param_prefix="scp")
    wh_scope, wh_params = company_sql_predicate(db, "w.company_id", param_prefix="scw")
    where = ["np.product_id::text = ANY(:pids)", _ACTIVE_POOL]
    if prod_scope:
        where.append(prod_scope)
    if wh_scope:
        where.append(wh_scope)
    sql = f"""
        SELECT np.product_id::text AS product_id,
               SUM(COALESCE(np.quantity_on_hand, 0)) AS on_hand,
               SUM(COALESCE(np.on_order, 0)) AS incoming_spo
        FROM scm.net_position_v np
        JOIN products p ON p.id = np.product_id
        JOIN warehouses w ON w.id = np.warehouse_id
        WHERE {' AND '.join(where)}
        GROUP BY np.product_id
    """
    rows = db.execute(
        text(sql), {"pids": product_ids, **prod_params, **wh_params}
    ).mappings().all()
    return {
        r["product_id"]: {
            "on_hand": float(r["on_hand"] or 0),
            "incoming_spo": float(r["incoming_spo"] or 0),
        }
        for r in rows
    }


def _demand_by_warehouse(db: Session, product_id: str) -> dict[str, dict]:
    """Per-warehouse open-SO demand shape for ONE product - the same predicate
    `location_stock_service.location_stock_for_product`'s own `so_qty` aggregate uses
    (`SalesOrder.status == 'open'`, `demand.is_open_demand()`), grouped further by earliest
    need-by date, oldest document date and demand class so `_location_options` can rank
    candidate locations through the shared Fulfilment Priority policy instead of a private
    sort. Class follows the repo's own rule (not a literal match): a warehouse with ANY
    project-class line ranks as project (project need is what the plan cares about seeing
    first), else retail when any line names a class, else unclassified.

    `payment_terms_days` is sourced the SAME way the auto-place ranking already does
    (`priority.payment_terms_by_customer`, `project_supply_service`/`project_fulfilment_board
    _service`'s own reads) rather than left hard-coded absent: every customer with open demand
    behind this warehouse is looked up, and the SHORTEST of their terms is kept - the same
    "shorter is higher" reading `priority.factors_for_demand_rows` gives the `customer_credit`
    factor generally, applied here at the location's most credit-sensitive customer. None
    (ABSENT, never a false best/worst) when nobody behind this location has been assessed."""
    from app.services.scm import priority
    from app.services.scm.demand import is_open_demand

    rows = (
        db.query(
            SalesOrderLine.warehouse_id,
            func.min(SalesOrderLine.required_date).label("required_date"),
            func.min(SalesOrder.order_date).label("order_date"),
            func.bool_or(SalesOrder.demand_class == "project").label("has_project"),
            func.bool_or(
                SalesOrder.demand_class.isnot(None) & (SalesOrder.demand_class != "project")
            ).label("has_retail"),
        )
        .join(SalesOrder, SalesOrder.id == SalesOrderLine.sales_order_id)
        .filter(
            SalesOrderLine.product_id == product_id,
            SalesOrder.status == "open",
            is_open_demand(),
            SalesOrderLine.warehouse_id.isnot(None),
        )
        .group_by(SalesOrderLine.warehouse_id)
        .all()
    )
    if not rows:
        return {}

    customer_pairs = (
        db.query(SalesOrderLine.warehouse_id, SalesOrder.customer_id)
        .join(SalesOrder, SalesOrder.id == SalesOrderLine.sales_order_id)
        .filter(
            SalesOrderLine.product_id == product_id,
            SalesOrder.status == "open",
            is_open_demand(),
            SalesOrderLine.warehouse_id.isnot(None),
            SalesOrder.customer_id.isnot(None),
        )
        .distinct()
        .all()
    )
    customers_by_warehouse: dict[str, set[str]] = {}
    for wh_id, cust_id in customer_pairs:
        customers_by_warehouse.setdefault(str(wh_id), set()).add(str(cust_id))
    all_customer_ids = {cid for ids in customers_by_warehouse.values() for cid in ids}
    terms_by_customer = priority.payment_terms_by_customer(db, list(all_customer_ids))

    out: dict[str, dict] = {}
    for r in rows:
        klass = "project" if r.has_project else "retail" if r.has_retail else None
        wh_key = str(r.warehouse_id)
        terms_here = [
            terms_by_customer[cid]
            for cid in customers_by_warehouse.get(wh_key, ())
            if terms_by_customer.get(cid) is not None
        ]
        out[wh_key] = {
            "required_date": r.required_date,
            "order_date": r.order_date,
            "demand_class": klass,
            "payment_terms_days": min(terms_here) if terms_here else None,
        }
    return out


def _demand_lines_by_warehouse(db: Session, product_id: str) -> dict[str, list[dict]]:
    """The INDIVIDUAL open SO lines behind `_demand_by_warehouse`'s aggregates - the doctrine
    correction's second drill: "what SO am I covering", per candidate location, earliest need
    date first. Same predicate `location_stock_for_product`'s own `so_qty` aggregate uses
    (`SalesOrder.status == 'open'`, `demand.is_open_demand()`); read once per product rather
    than once per candidate location. `agent_name` is a cheap outer join (`sales_agents`
    already carries a plain name column) - absent, never fabricated, on a line with no agent
    on file.

    Deliberately NOT cascaded/capped against the SPO qty here - a location's demand list is
    the same regardless of how much of it this particular SPO ends up satisfying, and the
    qty being proposed is edited live on screen (the same reason `_location_options.available`
    leaves its own "after figure" to the caller). The FE runs the identical earliest-first
    cascade against these rows and the live qty to answer "which of these does THIS SPO cover"
    and to bucket the SO-coverage schedule matrix - mirroring `po_takes`, computed the same way
    just on the other side of the shipment.
    """
    from app.services.scm.demand import demand_qty, is_open_demand

    owed = demand_qty()
    rows = (
        db.query(
            SalesOrderLine.warehouse_id,
            SalesOrder.so_number,
            Customer.customer_name,
            SalesAgent.sales_agent,
            SalesOrderLine.required_date,
            SalesOrder.order_date,
            owed.label("qty"),
        )
        .join(SalesOrder, SalesOrder.id == SalesOrderLine.sales_order_id)
        .outerjoin(Customer, Customer.id == SalesOrder.customer_id)
        .outerjoin(SalesAgent, SalesAgent.id == SalesOrder.sales_agent_id)
        .filter(
            SalesOrderLine.product_id == product_id,
            SalesOrder.status == "open",
            is_open_demand(),
            SalesOrderLine.warehouse_id.isnot(None),
        )
        .order_by(
            SalesOrderLine.required_date.asc().nulls_last(),
            SalesOrder.order_date.asc().nulls_last(),
            SalesOrder.so_number.asc(),
        )
        .all()
    )
    out: dict[str, list[dict]] = {}
    for r in rows:
        out.setdefault(str(r.warehouse_id), []).append({
            "so_number": r.so_number,
            "customer_name": r.customer_name,
            "agent_name": r.sales_agent,
            "required_date": r.required_date.isoformat() if r.required_date else None,
            "order_date": r.order_date.isoformat() if r.order_date else None,
            "qty": float(r.qty or 0),
        })
    return out


#: How a ticked piece of demand is named on the wire, and what the create sends back.
#: `project:<order inquiry row id>` / `retail:<sales order line id>` - the family is part of
#: the key because the two are different records, and only one of them can carry a link.
_COVERAGE_PROJECT = "project"
_COVERAGE_RETAIL = "retail"


def _project_coverage(db: Session, product_id: str) -> list[dict]:
    """Unlinked project demand for this product - the order-inquiry rows (R1, part 2 P3).

    Only ORDER BACK rows: an ORDER is a new purchase and goes on a purchase order, and
    `project_order_inquiry_service` refuses an SPO link on any other verb - so offering one
    here would offer a tick that cannot be honoured.

    `qty` is what the row still needs, net of every link it already carries. A row half
    placed is offered for the other half; a FULLY placed row is still RETURNED, at `qty: 0`
    (S5), carrying `taken_qty` / `taken_by` so the plan can grey it out rather than have it
    vanish with no word about where it went.
    """
    from app.models.project_so import (
        INQUIRY_CANCELLED,
        IV_ORDER_BACK,
        OrderInquiry,
        OrderInquiryLink,
        OrderInquiryRow,
        ProjectSalesOrder,
        ProjectSalesOrderLine,
    )

    rows = (
        db.query(
            OrderInquiryRow,
            ProjectSalesOrder.autocount_doc_no,
            ProjectSalesOrder.provisional_ref,
            Project.title,
        )
        .join(ProjectSalesOrderLine, ProjectSalesOrderLine.id == OrderInquiryRow.so_line_id)
        .join(OrderInquiry, OrderInquiry.id == OrderInquiryRow.order_inquiry_id)
        .join(ProjectSalesOrder, ProjectSalesOrder.id == OrderInquiry.project_sales_order_id)
        .outerjoin(Project, Project.id == ProjectSalesOrder.project_id)
        .filter(
            ProjectSalesOrderLine.product_id == product_id,
            OrderInquiryRow.verb == IV_ORDER_BACK,
            OrderInquiryRow.state != INQUIRY_CANCELLED,
        )
        .order_by(
            OrderInquiryRow.delivery_date.asc().nulls_last(),
            OrderInquiryRow.id.asc(),
        )
        .all()
    )
    if not rows:
        return []

    row_ids = [str(r.OrderInquiryRow.id) for r in rows]

    linked: dict[str, float] = {}
    for row_id, total in (
        db.query(OrderInquiryLink.row_id, func.sum(OrderInquiryLink.qty))
        .filter(OrderInquiryLink.row_id.in_(row_ids))
        .group_by(OrderInquiryLink.row_id)
        .all()
    ):
        linked[str(row_id)] = float(total or 0)

    # `taken_by` (S5/F3, review round): the document(s) currently backing each row's links.
    # `taken_qty` above sums EVERY link regardless of target - a PO placement counts exactly
    # the same as an SPO one - so `taken_by` has to name EVERY link's document too, or a row
    # fully placed on a plain PO read `taken_qty > 0, taken_by: []`, "Another SPO" for a row
    # no SPO ever touched. SPO numbers first (an inner join to `spo_allocations`, so a link
    # whose allocation was removed - which now takes the link with it, `unwind`'s own F9 fix
    # below - never names one that no longer exists), then PO numbers, each group sorted -
    # the same shape `project_order_inquiry_service.links_for_rows` states for a project
    # row's own "Linked to" column, read locally rather than imported (that module is the
    # large stateful class the module docstring already gives the reason not to import).
    spo_names: dict[str, list[str]] = {}
    for row_id, spo_number in (
        db.query(OrderInquiryLink.row_id, SPOAllocation.spo_number)
        .join(SPOAllocation, SPOAllocation.id == OrderInquiryLink.spo_allocation_id)
        .filter(
            OrderInquiryLink.row_id.in_(row_ids),
            SPOAllocation.spo_number.isnot(None),
        )
        .all()
    ):
        spo_names.setdefault(str(row_id), []).append(spo_number)
    po_names: dict[str, list[str]] = {}
    for row_id, po_number in (
        db.query(OrderInquiryLink.row_id, PurchaseOrder.po_number)
        .join(PurchaseOrderLine, PurchaseOrderLine.id == OrderInquiryLink.po_line_id)
        .join(PurchaseOrder, PurchaseOrder.id == PurchaseOrderLine.purchase_order_id)
        .filter(
            OrderInquiryLink.row_id.in_(row_ids),
            PurchaseOrder.po_number.isnot(None),
        )
        .all()
    ):
        po_names.setdefault(str(row_id), []).append(po_number)
    taken_by: dict[str, list[str]] = {
        row_id: sorted(spo_names.get(row_id, [])) + sorted(po_names.get(row_id, []))
        for row_id in set(spo_names) | set(po_names)
    }

    codes = _warehouse_ids_by_code(db)
    out: list[dict] = []
    for row, doc_no, provisional, project_title in rows:
        taken_qty = linked.get(str(row.id), 0.0)
        need = max(float(row.qty or 0) - taken_qty, 0.0)
        location = (row.stock_location or "").strip().upper() or None
        warehouse_id = codes.get(location) if location else None
        out.append({
            "key": f"{_COVERAGE_PROJECT}:{row.id}",
            "kind": _COVERAGE_PROJECT,
            "document": doc_no or provisional,
            "customer_name": project_title,
            "required_date": row.delivery_date.isoformat() if row.delivery_date else None,
            "qty": need,
            "taken_qty": taken_qty,
            "taken_by": taken_by.get(str(row.id), []),
            "warehouse_id": warehouse_id,
            "warehouse_code": location,
            # An order-inquiry row IS project demand by definition (R3, AC-J1) - there is no
            # `sales_orders.demand_class` to read here, so this is a constant, not a lookup.
            "demand_class": _DEMAND_CLASS_PROJECT,
        })
    return out


def _warehouse_ids_by_code(db: Session) -> dict[str, str]:
    """`{WAREHOUSE CODE: id}`. An inquiry row names its location as a CODE, and the split
    needs an id; a code naming no warehouse we hold simply steers nothing."""
    return {
        (code or "").strip().upper(): str(wid)
        for wid, code in db.query(Warehouse.id, Warehouse.warehouse_code).all()
        if code
    }


def _retail_cover_for(
    db: Session, product_id: str, ticked_keys: list[str], placing: float
) -> list[dict]:
    """What this SPO line is being asked to cover, per ticked RETAIL sales-order line.

    Cascaded in the coverage list's own order and capped at what the line actually places,
    so a tick list asking for more than the container holds records what the container can
    honestly answer for rather than the whole wish.
    """
    wanted = [k.split(":", 1)[1] for k in ticked_keys if k.startswith(f"{_COVERAGE_RETAIL}:")]
    if not wanted or placing <= 0:
        return []
    by_id = {
        c["key"].split(":", 1)[1]: c
        for c in _retail_coverage(db, product_id)
    }
    out: list[dict] = []
    left = placing
    for so_line_id in wanted:
        if left <= 0:
            break
        entry = by_id.get(so_line_id)
        if entry is None:
            continue
        take = min(float(entry["qty"]), left)
        if take <= 0:
            continue
        out.append({"so_line_id": so_line_id, "qty": take})
        left -= take
    return out


def _retail_coverage(db: Session, product_id: str) -> list[dict]:
    """Open sales-order book demand for this product - the retail half of the tick list.

    The same predicate `_demand_lines_by_warehouse` uses, so the two answers cannot
    disagree; the difference is that this one carries the LINE, because a tick has to name
    the thing it ticked.

    A line already partly or fully covered by an earlier container is still RETURNED (S5): a
    fully covered one at `qty: 0`, carrying `taken_qty` / `taken_by` so the plan can grey it
    out rather than have it vanish with no word about where it went.
    """
    from app.services.scm.demand import demand_qty, is_open_demand

    owed = demand_qty()
    rows = (
        db.query(
            SalesOrderLine.id,
            SalesOrderLine.warehouse_id,
            Warehouse.warehouse_code,
            SalesOrder.so_number,
            SalesOrder.demand_class,
            Customer.customer_name,
            SalesOrderLine.required_date,
            owed.label("qty"),
        )
        .join(SalesOrder, SalesOrder.id == SalesOrderLine.sales_order_id)
        .outerjoin(Customer, Customer.id == SalesOrder.customer_id)
        .outerjoin(Warehouse, Warehouse.id == SalesOrderLine.warehouse_id)
        .filter(
            SalesOrderLine.product_id == product_id,
            SalesOrder.status == "open",
            is_open_demand(),
        )
        .order_by(
            SalesOrderLine.required_date.asc().nulls_last(),
            SalesOrder.so_number.asc(),
            SalesOrderLine.id.asc(),
        )
        .all()
    )
    # Net of what earlier containers were already pointed at (review finding 10); the SAME
    # read also names WHICH SPO(s) (S5).
    cover = _spo_cover_by_so_line(db, product_id)
    out: list[dict] = []
    for r in rows:
        entries = cover.get(str(r.id), [])
        taken_qty = sum(e["qty"] for e in entries)
        left = max(float(r.qty or 0) - taken_qty, 0.0)
        out.append({
            "key": f"{_COVERAGE_RETAIL}:{r.id}",
            "kind": _COVERAGE_RETAIL,
            "document": r.so_number,
            "customer_name": r.customer_name,
            "required_date": r.required_date.isoformat() if r.required_date else None,
            "qty": left,
            "taken_qty": taken_qty,
            "taken_by": [e["spo_number"] for e in entries],
            "warehouse_id": str(r.warehouse_id) if r.warehouse_id else None,
            "warehouse_code": r.warehouse_code,
            # `sales_orders.demand_class` as stored - 'project' / 'retail' / None (R3, AC-J1).
            # Distinct from `kind` above: `kind` says WHERE the row came from (an inquiry row
            # vs a book line), `demand_class` says what the SO ITSELF is classified as.
            "demand_class": r.demand_class,
        })
    return out


def _so_coverage(db: Session, product_id: str, packed: float) -> list[dict]:
    """What this SPO could be for, and what the default ticks claim (Q4, AC-G3).

    TWO groups, each by need-by date, oldest first (R3, AC-J2, captain's course correction 3
    Sep): PROJECT DEMAND - order-inquiry rows AND book lines whose own sales order is
    project-class, merged and sorted together by date rather than the inquiry rows jumping
    ahead of every book line regardless of date - then everything else (retail / unclassified
    book lines). The old two-way split (inquiry rows, then EVERY book line) conflated "where
    the row comes from" with "what the order is classified as": a retail-book-line SO stamped
    `demand_class = project` used to walk behind every retail line regardless of its own
    priority. A single stable sort over the combined list does the merge; a same-date tie
    keeps each source query's own relative order (Python's sort is stable).

    Dedupe: a book line whose project SO line already carries an ORDER BACK inquiry row is
    the SAME piece of demand as that row, not a second one - `_project_coverage` and
    `_retail_coverage` read two different tables for it (the inquiry row, and the AutoCount
    book line it mirrors) and would otherwise both offer it. Dropped here, once, rather than
    in each reader, because only this function holds both lists at once to compare them.
    """
    from app.models.project_so import OrderInquiryRow, ProjectSalesOrderLine

    project_rows = _project_coverage(db, product_id)
    retail_rows = _retail_coverage(db, product_id)

    inquiry_row_ids = [row["key"].split(":", 1)[1] for row in project_rows]
    covered_core_line_ids: set[str] = set()
    if inquiry_row_ids:
        # The mirror chain `sales_order_service._line_inquiries` reads the other way:
        # `OrderInquiryRow.so_line_id` -> `ProjectSalesOrderLine.id` ->
        # `ProjectSalesOrderLine.core_sales_order_line_id` - the book line's own id, which is
        # the `so_line_id` half of a `retail:<id>` coverage key.
        covered_core_line_ids = {
            str(r[0])
            for r in (
                db.query(ProjectSalesOrderLine.core_sales_order_line_id)
                .join(OrderInquiryRow, OrderInquiryRow.so_line_id == ProjectSalesOrderLine.id)
                .filter(OrderInquiryRow.id.in_(inquiry_row_ids))
                .all()
            )
            if r[0] is not None
        }
    if covered_core_line_ids:
        retail_rows = [
            row
            for row in retail_rows
            if row["key"].split(":", 1)[1] not in covered_core_line_ids
        ]

    coverage = project_rows + retail_rows
    coverage.sort(key=_so_coverage_sort_key)
    left = packed
    for entry in coverage:
        take = min(entry["qty"], left) if left > 0 else 0.0
        entry["default_ticked"] = take > 0
        left -= take
    return coverage


def _so_coverage_sort_key(entry: dict) -> tuple[int, bool, str]:
    """(0, ...) = project demand - an inquiry row OR a book line whose own SO is
    project-class; (1, ...) = everything else. Within a group, ascending by need-by date,
    no-date last - the same `nulls_last` reading the source queries' own `ORDER BY` uses,
    restated here because the merge happens in Python, across two already-sorted lists."""
    is_project_demand = (
        entry["kind"] == _COVERAGE_PROJECT or entry.get("demand_class") == _DEMAND_CLASS_PROJECT
    )
    date = entry.get("required_date")
    return (0 if is_project_demand else 1, date is None, date or "")


def _location_options(db: Session, product_id: str) -> dict:
    """Candidate destination warehouses for a new SPO on this product, ranked - the module
    docstring's "second amendment" section. Every site pool, plus every project bin carrying
    a live figure for the product (`location_stock_service.location_stock_for_product` - the
    SAME per-location reader the reorder Buy-row popup and the fulfilment board already use;
    R16 made the pool rows unconditional, so a site holding none of this item is still a
    destination the buyer can pick), ranked by the
    active Fulfilment Priority policy (`priority.factors_for_demand_rows` - project earlier
    delivery first, then retail, never a second sort); a location with no open demand behind
    it sorts after every ranked one, by warehouse code, mirroring `container_request_service.
    build`'s own "ranked rows, then no-demand rows" order. A product with nothing anywhere
    (never sold or stocked) falls back to the one "counts as available" warehouse
    (`allocation_suggestion_service._default_warehouse`, reused rather than reinvented)."""
    from app.services.scm import priority
    from app.services.scm.allocation_suggestion_service import _default_warehouse
    from app.services.scm.location_stock_service import location_stock_for_product

    locations = location_stock_for_product(db, product_id).get("locations", [])
    if not locations:
        fallback = _default_warehouse(db)
        if not fallback:
            return {"options": [], "suggested_warehouse_id": None}
        return {
            "options": [
                {
                    "warehouse_id": str(fallback["id"]),
                    "warehouse_code": fallback["warehouse_code"],
                    "outstanding_so": 0.0,
                    "on_hand": 0.0,
                    "incoming_spo": 0.0,
                    "available": 0.0,
                    "rank_score": None,
                    "demand_lines": [],
                }
            ],
            "suggested_warehouse_id": str(fallback["id"]),
        }

    demand = _demand_by_warehouse(db, product_id)
    demand_lines = _demand_lines_by_warehouse(db, product_id)
    demand_rows = [
        {
            "row_key": loc["warehouse_id"],
            "required_date": demand[loc["warehouse_id"]]["required_date"],
            "order_date": demand[loc["warehouse_id"]]["order_date"],
            "payment_terms_days": demand[loc["warehouse_id"]]["payment_terms_days"],
            "demand_class": demand[loc["warehouse_id"]]["demand_class"],
        }
        for loc in locations
        if loc["warehouse_id"] in demand
    ]
    scores: dict[str, float] = {}
    if demand_rows:
        factors_by_row = priority.factors_for_demand_rows(db, demand_rows)
        scores = priority.scores_for(factors_by_row)

    options = [
        {
            "warehouse_id": loc["warehouse_id"],
            "warehouse_code": loc["warehouse_code"],
            "outstanding_so": loc["so_qty"],
            "on_hand": loc["on_hand"],
            "incoming_spo": loc["spo_qty"],
            # Signed, never clamped - same rule `location_stock_for_product`'s own
            # `available` already carries. The caller adds whatever qty it proposes landing
            # here to get the "after figure" the plan asks for - computed against the LIVE,
            # edited quantity on screen, never a value this endpoint would have to go stale.
            "available": loc["available"],
            "rank_score": scores.get(loc["warehouse_id"]),
            # "What SO am I covering" - the individual demand lines behind `outstanding_so`,
            # earliest need date first. The FE cascades these against its own live SPO qty
            # (and its own per-location split) to answer which of them this SPO actually
            # serves, and to bucket the SO-coverage schedule matrix.
            "demand_lines": demand_lines.get(loc["warehouse_id"], []),
        }
        for loc in locations
    ]
    options.sort(
        key=lambda o: (
            0 if o["rank_score"] is not None else 1,
            -(o["rank_score"] or 0.0),
            o["warehouse_code"] or "",
        )
    )
    suggested = options[0]["warehouse_id"] if options else None
    return {"options": options, "suggested_warehouse_id": suggested}


def suggest(db: Session, shipment_id: str) -> dict:
    """What "Create SPO" would ask for, per shipment line - the DOCTRINE-CORRECTED arithmetic
    (module docstring, fifth amendment): `suggested_qty` is what an open PO PULLS this SPO up
    to, not what is left after a PO is subtracted. It is a DEFAULT, not a cap (sixth amendment,
    captain's ruling 3 Sep) - the input the buyer sees starts here, but she can type past it,
    up to the line's own REMAINDER (seventh amendment, R1 below); a line with no PO to pull
    from at all is still convertible, only a line with no supplier is not (`cannot_convert`).
    Almost a pure read - the one exception is `_heal_stale_links`, which deletes any link row
    a prior `create` wrote that no longer points at a live PO/PO line (see that function's
    docstring). The caller must `db.commit()` after this, same as every other
    lazily-self-healing GET in this codebase (`explainer.py`'s cached-explanation routes).

    **Seventh amendment (R1, `PLAN-scm-spo-planner-feedback-3sep.md`, captain's ruling 3
    Sep): a container can carry MANY SPOs, not one.** `already_converted` NEVER flips to
    `True` any more - it is kept in the response, always `False`, only so an older FE build
    reading it does not break. `existing_spos` is always populated (every SPO this shipment
    has ever produced, oldest first) and `lines` is always the REMAINDER planner: per line,
    `packed_qty` stays the untouched physical fact (`quantity_shipped`) while the new
    `remaining_qty` is what is left to convert - `packed_qty` minus every prior `create`
    run's own take on this line (`_already_spo_qty_by_line`, summed off
    `ShipmentLineSpoLink` rows a matched line now writes ONE OF PER RUN rather than once
    ever). `suggested_qty` / `po_covered_qty` / `no_po_qty` are the SAME cascade as before,
    simply run against `remaining_qty` instead of `packed_qty` as the ask. A line with
    nothing left to convert (`remaining_qty <= 0`) is `cannot_convert: True`, reason "Already
    on <SPO number(s)>" - the same shape the no-supplier case already used, so the FE's one
    disabled-row rendering covers both."""
    shipment = _shipment_or_404(db, shipment_id)
    lines = (
        db.query(InboundShipmentLine)
        .filter(InboundShipmentLine.shipment_id == shipment.id)
        .all()
    )

    _, healed_count = _heal_stale_links(db, shipment.id, _existing_links(db, shipment.id))
    self_heal_note = (
        f"{healed_count} SPO{'s' if healed_count != 1 else ''} previously linked to this "
        "shipment no longer exist and have been cleared - Create SPO can be run again for them."
        if healed_count
        else None
    )
    existing_spos = _existing_spos(db, shipment.id)
    already_spo = _already_spo_qty_by_line(db, [str(ln.id) for ln in lines])

    product_ids = [str(ln.product_id) for ln in lines if ln.product_id]
    supplier_ids = {str(ln.supplier_id) for ln in lines if ln.supplier_id}
    products = {
        str(pid): (code, name)
        for pid, code, name in db.query(*_product_cols()).filter(_product_id_in(product_ids)).all()
    } if product_ids else {}
    suppliers = (
        {str(s.id): s.supplier_name for s in db.query(Supplier).filter(Supplier.id.in_(supplier_ids)).all()}
        if supplier_ids
        else {}
    )
    stock = _stock_context(db, product_ids)
    # One `_location_options` call per DISTINCT product on the shipment, not per line - two
    # shipment lines of the same product (different suppliers, say) ask the same "where does
    # this land" question and must get the same ranked answer.
    location_cache: dict[str, dict] = {}
    # ONE scan of the whole `crm_spo` book for the whole shipment (F7, review round) - every
    # line below reads out of this same dict rather than re-querying it per line.
    pulls_by_line = _spo_pulls_by_po_line(db)

    out_lines: list[dict] = []
    for ln in sorted(lines, key=lambda l: (products.get(str(l.product_id), (None, None))[0] or "", str(l.id))):
        packed = float(ln.quantity_shipped or 0)
        product = products.get(str(ln.product_id))
        item_code, product_name = product if product else (None, None)
        supplier_id = str(ln.supplier_id) if ln.supplier_id else None
        already = already_spo.get(str(ln.id), {"qty": 0.0, "spo_numbers": []})
        remaining = max(packed - already["qty"], 0.0)

        if supplier_id is None:
            out_lines.append({
                "shipment_line_id": str(ln.id),
                "product_id": str(ln.product_id),
                "item_code": item_code,
                "product_name": product_name,
                "supplier_id": None,
                "supplier_name": None,
                "packed_qty": packed,
                "remaining_qty": remaining,
                "po_covered_qty": 0.0,
                "matched_po_number": None,
                "matched_by": None,
                "po_takes": [],
                "on_hand": 0.0,
                "incoming_spo": 0.0,
                "suggested_qty": 0.0,
                "no_po_qty": remaining,
                "cannot_convert": True,
                "reason": _REASON_NO_SUPPLIER,
                "unit_cost": _f(ln.unit_cost),
                "currency": ln.currency,
                "location_options": [],
                "suggested_warehouse_id": None,
                "so_coverage": [],
            })
            continue

        ctx = stock.get(str(ln.product_id), {"on_hand": 0.0, "incoming_spo": 0.0})

        if remaining <= 0:
            # R1: nothing left on this line for THIS shipment to convert - already fully
            # accounted for by a prior `create` run (or several). Same shape the
            # no-supplier case already used (`cannot_convert`, a `reason`), so the FE's one
            # disabled-row rendering covers both.
            names = ", ".join(already["spo_numbers"]) or "an SPO"
            out_lines.append({
                "shipment_line_id": str(ln.id),
                "product_id": str(ln.product_id),
                "item_code": item_code,
                "product_name": product_name,
                "supplier_id": supplier_id,
                "supplier_name": suppliers.get(supplier_id),
                "packed_qty": packed,
                "remaining_qty": 0.0,
                "po_covered_qty": 0.0,
                "matched_po_number": None,
                "matched_by": None,
                "po_takes": [],
                "on_hand": ctx["on_hand"],
                "incoming_spo": ctx["incoming_spo"],
                "suggested_qty": 0.0,
                "no_po_qty": 0.0,
                "cannot_convert": True,
                "reason": f"Already on {names}.",
                "unit_cost": _f(ln.unit_cost),
                "currency": ln.currency,
                "location_options": [],
                "suggested_warehouse_id": None,
                "so_coverage": [],
            })
            continue

        matched_by, takes = _match_takes_for_line(db, ln, remaining)
        po_covered = sum(t[2] for t in takes)
        matched_po_number = takes[0][1].po_number if takes else None

        # The SAME candidate set the cascade above just walked (S5) - a taken-only row
        # names the SPO(s) that already pulled from a line the cascade did NOT take from,
        # so a line fully (or partly) occupied elsewhere is visible rather than silently
        # absent. `_candidate_lines_for_line` is re-resolved rather than threaded out of
        # `_match_takes_for_line` because `suggest` never ticks (`only_po_lines=None`), so
        # the two calls can never name a different candidate set for this line.
        _, candidates = _candidate_lines_for_line(db, ln)
        taken_ids = {str(line.id) for line, _po, _qty in takes}

        def _po_take_row(line, po, qty: float) -> dict:
            entries = pulls_by_line.get(str(line.id), [])
            return {
                "po_line_id": str(line.id),
                "po_number": po.po_number,
                "qty": qty,
                "expected_date": line.expected_date.isoformat() if line.expected_date else None,
                # The captain's own ask (doctrine correction, per-take drill): the PO's OWN
                # date and supplier, not just its number - a pinned match can resolve to a
                # PO booked under a differently-spelled supplier (fourth amendment), so this
                # is the PO's supplier, not necessarily the shipment line's.
                "po_date": po.issue_date.isoformat() if po.issue_date else None,
                "supplier_name": po.supplier.supplier_name if po.supplier else None,
                # What the LINE has open, not what this cascade took from it. Unticking a
                # take re-runs the walk over the lines still ticked, and the screen cannot
                # mirror that from the slice alone (review finding 9).
                "open_qty": max(
                    float(line.qty_ordered or 0) - float(line.qty_received or 0), 0.0
                ),
                # S5: what an earlier SPO already pulled off this SAME line, and its
                # number(s) - context on a take, the whole row on a taken-only entry below.
                "taken_qty": sum(e["qty"] for e in entries),
                "taken_by": [e["spo_number"] for e in entries],
            }

        po_takes = [_po_take_row(line, po, qty) for line, po, qty in takes]
        # Candidate lines the cascade did NOT take from, but that another SPO already has
        # (S5) - `qty: 0` (nothing left for THIS cascade to claim), kept in the candidates'
        # own order (earliest issue date first) so they read as a continuation of the takes
        # above rather than a second, differently-ordered list.
        for line, po, available in candidates:
            if str(line.id) in taken_ids:
                continue
            # F2 (review round): a taken-only row is for a line THIS cascade left NOTHING
            # open on - one with real open balance the cascade simply did not walk to (a
            # later PO in the queue, say) is a normal, tickable candidate, and showing it
            # here greyed it out and made it untickable for no reason.
            if available > 0:
                continue
            row = _po_take_row(line, po, 0.0)
            if row["taken_qty"] > 0:
                po_takes.append(row)

        # The doctrine correction's own arithmetic (module docstring, fifth amendment),
        # amended again (sixth amendment, captain's ruling 3 Sep; seventh, R1): `suggested_qty`
        # IS `po_covered_qty` - what an open PO pulls this SPO up to - but only as the
        # DEFAULT the input starts at, never a cap; the buyer can type past it up to the
        # line's own REMAINDER, not the raw packed figure. `on_hand` / `incoming_spo` stay
        # on the response as context only.
        suggested = po_covered
        no_po_qty = max(remaining - po_covered, 0.0)
        # Only a missing SUPPLIER blocks conversion (checked earlier, above); this branch
        # already has one, so a line with no open PO at all is still convertible - `reason`
        # stays as information about the shortfall, never a block.
        cannot_convert = False
        if po_covered <= 0:
            reason = _REASON_NO_PO
        elif no_po_qty > 0:
            reason = (
                f"{_g(no_po_qty)} of {_g(remaining)} has no PO to pull from - raise it in "
                "AutoCount first."
            )
        else:
            reason = None

        pid = str(ln.product_id)
        if pid not in location_cache:
            location_cache[pid] = _location_options(db, pid)
        location = location_cache[pid]

        out_lines.append({
            "shipment_line_id": str(ln.id),
            "product_id": str(ln.product_id),
            "item_code": item_code,
            "product_name": product_name,
            "supplier_id": supplier_id,
            "supplier_name": suppliers.get(supplier_id),
            "packed_qty": packed,
            "remaining_qty": remaining,
            "po_covered_qty": po_covered,
            "matched_po_number": matched_po_number,
            "matched_by": matched_by,
            "po_takes": po_takes,
            "on_hand": ctx["on_hand"],
            "incoming_spo": ctx["incoming_spo"],
            "suggested_qty": suggested,
            "no_po_qty": no_po_qty,
            "cannot_convert": cannot_convert,
            "reason": reason,
            "unit_cost": _f(ln.unit_cost),
            "currency": ln.currency,
            "location_options": location["options"],
            "suggested_warehouse_id": location["suggested_warehouse_id"],
            # What this SPO could be FOR, with the default ticks already walked (AC-G3),
            # capped at what is actually LEFT to place (R1) rather than the raw packed qty.
            "so_coverage": _so_coverage(db, str(ln.product_id), remaining),
        })

    return {
        "shipment_id": str(shipment.id),
        "shipment_number": shipment.shipment_number,
        "shipment_status": shipment.shipment_status,
        # R1: never flips any more - kept only so an older FE build does not break.
        "already_converted": False,
        "existing_spos": existing_spos,
        "lines": out_lines,
        "self_heal_note": self_heal_note,
    }


def _product_cols():
    from app.models.product import Product

    return (Product.id, Product.product_code, Product.product_name)


def _product_id_in(ids: list[str]):
    from app.models.product import Product

    return Product.id.in_(ids)


def _spo_number(db: Session) -> str:
    number = NumberingService(db).get_next_number(_NUMBER_DOC_TYPE, _date.today(), commit_rule=False)
    return number or f"{_NUMBER_PREFIX}-{uuid.uuid4().hex[:8]}"


def create(
    db: Session,
    shipment_id: str,
    lines: list[dict],
    *,
    actor: Optional[str] = None,
    actor_user_id: Optional[str] = None,
) -> dict:
    """Confirm the screen `suggest` drew. One `purchase_orders` header per supplier, each new
    line carrying `need = min(requested, remaining_qty)` as its quantity outright (seventh
    amendment, R1 below, generalising the sixth's `min(requested, packed)`) - PULLING what it
    can from open PO line(s), re-derived HERE against live data rather than trusted from
    `suggest`'s earlier read (the same "recompute at write time" rule every write in this
    module already follows), and writing whatever `need` is left over onto the SAME line with
    no pull behind it.

    `lines` is `[{shipment_line_id, qty, include, location_splits}]` - every shipment line, per
    the "same structure on read and write" screen shape: a line the buyer left unticked still
    needs a row here so the whole shipment is accounted for in ONE action, not a part of it
    silently left for later. An untouched or zero-qty line is skipped `_REASON_NOT_SELECTED` -
    "not selectable" holds at write time too, not only on the read that drew the screen. A line
    with a supplier but no open PO at all is NO LONGER skipped (sixth amendment) - it converts
    at `need`, unbacked. A line with NOTHING left to convert (`remaining_qty <= 0`, seventh
    amendment) is skipped too, "Already on <SPO number(s)>", however it was ticked.

    `location_splits` (fourth ask, multi-location) is OPTIONAL and, when present, a list of
    `{warehouse_id, qty}` that must sum to EXACTLY what this line actually pulls (recomputed,
    not the client's stated `qty`) - the same "a split must add up" hard rule `allocation_
    suggestion_service.approve` already enforces for its own splits, applied here because this
    screen now makes the identical kind of decision. Each split writes its own `spo_allocations`
    row in the SAME action that mints the SPO line - one confirm, still one write per DECISION,
    now zero-or-many decisions per line rather than zero-or-one. Absent/empty means no
    allocation is written for that line, byte-identical to a line with no location chosen -
    `scm.on_order_v` stays silent about it until someone allocates it.
    `actor_user_id` is the real user id `spo_allocations.created_by` needs (a UUID column);
    `actor`, kept as-is, is the human-readable name other provenance on this shipment stamps.

    **Seventh amendment (R1, captain's ruling 3 Sep): no blanket refusal any more.** A
    shipment can carry MANY SPOs - this used to refuse (409) the moment ANY link existed for
    the shipment at all; now every line is judged on its OWN remainder (`_already_spo_qty_
    by_line`, re-summed here at write time, never trusted from `suggest`'s earlier read).
    Refused (422, `nothing_left`) only when NO line with a supplier has anything left to
    convert at all - the "everything is already on an SPO" state; `nothing_selected` stays
    the answer for "something is left, but nothing was ticked".
    """
    shipment = _shipment_or_404(db, shipment_id)

    shipment_lines = {
        str(ln.id): ln
        for ln in db.query(InboundShipmentLine)
        .filter(InboundShipmentLine.shipment_id == shipment.id)
        .all()
    }
    if not shipment_lines:
        raise AppException(422, "This shipment has no lines to create an SPO from.")

    by_line = {str(item.get("shipment_line_id") or ""): item for item in (lines or [])}
    missing = [lid for lid in shipment_lines if lid not in by_line]
    if missing:
        raise AppException(
            422,
            "Every line on this shipment must be accounted for.",
            detail=f"{len(missing)} line(s) missing from the request",
        )

    groups: dict[str, dict[str, Any]] = {}
    skipped: list[tuple[str, str]] = []
    ticked_demand: dict[str, list[str]] = {}
    retail_cover: dict[str, list[dict]] = {}

    already_spo = _already_spo_qty_by_line(db, list(shipment_lines.keys()))
    any_supplier_line = False
    any_remaining = False

    for line_id, ln in shipment_lines.items():
        item = by_line[line_id]
        include = bool(item.get("include"))
        requested = float(item.get("qty") or 0)
        splits = [
            {"warehouse_id": str(s.get("warehouse_id") or ""), "qty": float(s.get("qty") or 0)}
            for s in (item.get("location_splits") or [])
            if s.get("warehouse_id")
        ]
        supplier_id = str(ln.supplier_id) if ln.supplier_id else None

        if supplier_id is None:
            skipped.append((line_id, _REASON_NO_SUPPLIER))
            continue
        any_supplier_line = True

        # R1: judged against this LINE's own remainder, re-summed here at write time (never
        # trusted from `suggest`'s earlier read) - a line another confirm already spent
        # since the screen was drawn must not be spent twice.
        already = already_spo.get(line_id, {"qty": 0.0, "spo_numbers": []})
        remaining = max(float(ln.quantity_shipped or 0) - already["qty"], 0.0)
        if remaining <= 0:
            names = ", ".join(already["spo_numbers"]) or "an SPO"
            skipped.append((line_id, f"Already on {names}."))
            continue
        any_remaining = True

        if not include or requested <= 0:
            skipped.append((line_id, _REASON_NOT_SELECTED))
            continue

        # `need` is the SPO line's quantity outright (sixth amendment, captain's ruling 3
        # Sep; seventh, R1) - never the qty the client sent as-is, capped at what is left ON
        # THIS LINE (never the raw packed figure once a prior run has already taken some of
        # it). It is NOT capped a second time at what a PO can back.
        need = min(requested, remaining)
        # Only the takes the buyer left ticked (AC-G1), applied to the CANDIDATES so the
        # cascade runs again over what is left - never `suggest`'s earlier read, so a PO
        # another confirm consumed in between must not be double-spent. ABSENT means "every
        # take you re-derive", which is what every caller before this ask sent; an empty LIST
        # means "draw from none of them" - the line still becomes an SPO line at `need`, simply
        # with no pull behind it.
        take_ids = item.get("po_take_ids")
        only = None if take_ids is None else {str(t) for t in take_ids}
        matched_by, takes = _match_takes_for_line(db, ln, need, only)
        covered_now = sum(t[2] for t in takes)

        if splits:
            split_total = sum(s["qty"] for s in splits)
            if abs(split_total - need) > 1e-6:
                raise AppException(
                    422,
                    "A line's location split has to add up to its SPO qty.",
                    detail=f"{_g(split_total)} split against {_g(need)} on the line.",
                )

        group = groups.setdefault(
            supplier_id,
            {"lines": [], "currency": ln.currency},
        )
        group["lines"].append((ln, need, takes, splits))
        # Which demand this line's quantity was ticked against (AC-G3). Held per shipment
        # line so the links can be written AFTER the allocations exist to point at.
        ticked_demand[line_id] = [str(k) for k in (item.get("so_line_ids") or [])]
        # The RETAIL half of the ticks, with what each one is being covered for. Cascaded
        # in the order the coverage list walks, capped by the FULL quantity this line is
        # actually placing (`need`, sixth amendment) - not only its PO-covered part - the
        # same walk the screen ticked with.
        retail_cover[line_id] = _retail_cover_for(
            db, str(ln.product_id), ticked_demand[line_id], need
        )

    if not groups:
        if any_supplier_line and not any_remaining:
            raise AppException(
                422,
                "Every line of this packing list is already on an SPO.",
                detail="nothing_left",
            )
        raise AppException(
            422,
            "Nothing was selected to create an SPO from.",
            detail="nothing_selected",
        )

    supplier_names = {
        str(s.id): s.supplier_name
        for s in db.query(Supplier).filter(Supplier.id.in_(list(groups.keys()))).all()
    }

    created_spos: list[dict] = []
    line_links: dict[str, tuple[str, str]] = {}  # shipment_line_id -> (po_id, po_line_id)
    # One entry per location split, across every line - read after every SPO/line/link row
    # above is written, so an allocation is never attempted against a purchase-order line
    # that does not exist yet.
    allocation_targets: list[dict[str, Any]] = []

    for supplier_id, group in groups.items():
        po = PurchaseOrder(
            id=_uuid(),
            po_number=_spo_number(db),
            supplier_id=supplier_id,
            issue_date=_date.today(),
            expected_date=shipment.eta_delay_date or shipment.estimated_arrival_date,
            status=_STATUS,
            currency=group["currency"],
            source_system=SOURCE_SYSTEM,
            source_ref=str(shipment.id),
        )
        db.add(po)
        db.flush()

        total_qty = 0.0
        for ln, need, takes, splits in group["lines"]:
            # What the takes above actually pulled - `need` itself is the line's quantity
            # (sixth amendment); this is only the PO-covered PART of it, for `no_po_qty`.
            covered_now = sum(qty for _src_line, _src_po, qty in takes)
            po_line = PurchaseOrderLine(
                id=_uuid(),
                purchase_order_id=po.id,
                product_id=ln.product_id,
                qty_ordered=Decimal(str(need)),
                qty_received=0,
                unit_cost=ln.unit_cost,
                currency=ln.currency,
                expected_date=po.expected_date,
                line_status=_LINE_STATUS,
                source_system=SOURCE_SYSTEM,
                # WHICH open PO line(s) this pull drew from, and how much of each - see the
                # module docstring's fifth amendment for why this is JSON on `source_ref`
                # rather than a new link table. Replaces the shipment-line-id this column
                # used to carry (redundant - `ShipmentLineSpoLink.inbound_shipment_line_id`
                # already names it).
                # WHERE this pull came from, and WHAT it is for. The retail half of the
                # tick list has no link row to hang on (`order_inquiry_links.row_id` is NOT
                # NULL and a retail sales-order line has none), so without this the same 30
                # pieces were promised to SO-A on every container of the month, each one
                # default-ticked at full quantity. Recorded here, beside the pull, because
                # this line IS the thing that serves them.
                # `no_po_qty` - the sixth amendment's own addition: the part of `need` no PO
                # backs, `parse_source_ref` ignores unknown keys so this never breaks an
                # older reader.
                source_ref=json.dumps({
                    "pulls": [
                        {"po_line_id": str(src_line.id), "qty": qty}
                        for src_line, _src_po, qty in takes
                    ],
                    "so_coverage": retail_cover.get(str(ln.id), []),
                    "no_po_qty": max(need - covered_now, 0.0),
                }),
            )
            db.add(po_line)
            db.flush()
            line_links[str(ln.id)] = (po.id, po_line.id)

            # The pull ADVANCES the source PO line's own accounting - the honest choice the
            # module docstring works through (fifth amendment): the same write `allocation_
            # suggestion_service.approve` makes when a shipment draws down a PO line, so the
            # source line's open balance falls by exactly what this SPO line just claimed and
            # the company-wide "ordered" total nets to zero movement, never a double count.
            for src_line, _src_po, qty in takes:
                src_line.qty_received = float(src_line.qty_received or 0) + qty

            for split in splits:
                allocation_targets.append({
                    "shipment_line_id": str(ln.id),
                    "product_id": str(ln.product_id),
                    "warehouse_id": split["warehouse_id"],
                    "qty": split["qty"],
                    "po_number": po.po_number,
                    "po_line_id": po_line.id,
                })
            total_qty += need

        created_spos.append({
            "purchase_order_id": str(po.id),
            "po_number": po.po_number,
            "supplier_id": supplier_id,
            "supplier_name": supplier_names.get(supplier_id),
            "currency": po.currency,
            "lines": len(group["lines"]),
            "qty": total_qty,
        })

    for line_id, (po_id, po_line_id) in line_links.items():
        db.add(ShipmentLineSpoLink(
            id=_uuid(),
            inbound_shipment_id=shipment.id,
            inbound_shipment_line_id=line_id,
            purchase_order_id=po_id,
            purchase_order_line_id=po_line_id,
        ))
    for line_id, reason in skipped:
        db.add(ShipmentLineSpoLink(
            id=_uuid(),
            inbound_shipment_id=shipment.id,
            inbound_shipment_line_id=line_id,
            purchase_order_id=None,
            purchase_order_line_id=None,
            unmatched_reason=reason,
        ))
    db.flush()

    allocations = _write_allocations(db, shipment, allocation_targets, actor_user_id=actor_user_id)
    demand_links = _link_ticked_demand(
        db, allocations, ticked_demand, actor_user_id=actor_user_id
    )

    return {
        "shipment_id": str(shipment.id),
        "shipment_number": shipment.shipment_number,
        "created_spos": created_spos,
        "demand_links": demand_links,
        "skipped": [
            {
                "shipment_line_id": lid,
                "item_code": None,
                "reason": reason,
            }
            for lid, reason in skipped
        ],
        "allocations": allocations,
    }


def _write_allocations(
    db: Session,
    shipment: InboundShipment,
    targets: list[dict[str, Any]],
    *,
    actor_user_id: Optional[str],
) -> list[dict]:
    """The second half of "one confirm, two writes" - `SPOAllocationService.create_allocation`,
    the IDENTICAL write `allocation_suggestion_service.approve` uses, so this action and that
    one can never disagree about what an allocation is. `forward_match=False` per row and one
    sweep per distinct SPO number at the end, for the exact reason `approve` already gives:
    several warehouses can share one SPO number here (a container is routinely several
    factories, several destinations - and, since the fourth ask, several destinations for ONE
    factory's line too), and firing the GRN-forward-match hook per row would place a waiting
    GRN line before every allocation under that number exists.

    `targets` is a FLAT list, one entry per location split - a line with two splits writes two
    allocations here, both against the SAME `po_line_id` (its one new SPO line)."""
    if not targets:
        return []
    from app.schemas.procurement import SPOAllocationCreate
    from app.services.grn_spo_matching import forward_match_grn_lines_for_spo_best_effort
    from app.services.procurement_service import SPOAllocationService

    service = SPOAllocationService(db)
    written: list[dict] = []
    forward_match_targets: set[tuple[str, Optional[str]]] = set()
    for target in targets:
        allocation = service.create_allocation(
            SPOAllocationCreate(
                spo_number=target["po_number"],
                inbound_shipment_id=str(shipment.id),
                product_id=target["product_id"],
                warehouse_id=target["warehouse_id"],
                allocated_quantity=int(round(target["qty"])),
                po_line_id=target["po_line_id"],
            ),
            created_by=actor_user_id,
            forward_match=False,
        )
        written.append({
            "shipment_line_id": target["shipment_line_id"],
            "warehouse_id": target["warehouse_id"],
            "allocation_id": str(allocation.id),
            "qty": target["qty"],
        })
        forward_match_targets.add((
            target["po_number"],
            str(allocation.company_id) if allocation.company_id is not None else None,
        ))

    for spo_number, company_id in forward_match_targets:
        forward_match_grn_lines_for_spo_best_effort(db, spo_number, company_id=company_id)

    return written


def _link_ticked_demand(
    db: Session,
    allocations: list[dict],
    ticked: dict[str, list[str]],
    *,
    actor_user_id: Optional[str],
) -> list[dict]:
    """Tie the ticked PROJECT demand to the SPO allocations that will serve it (AC-G6).

    Written through `ProjectOrderInquiryService.place_on_po_allocations`, the ONE writer of
    `projects.order_inquiry_links` (part 2 I), rather than inserting the row here: that
    service holds the rules a link has to obey - the row is never split, the target must
    still be open for that product, an SPO answers only an ORDER BACK row - and a second
    writer would be a second, quietly different set of them.

    **A RETAIL tick writes no link, by design.** `order_inquiry_links.row_id` is NOT NULL:
    the table hangs off an order-inquiry row, and a retail sales-order line has none. Making
    the column nullable to hang one there would weaken a constraint four readers join on
    (`_allocations_for`, `links_for_rows`, `_allocated_by_po`, the sales-order detail's own
    "Linked to"), for a link no screen reads - retail has never carried one, for a purchase
    order either. The retail tick still counts: it is what put the quantity at that
    warehouse.

    Best-effort per row: a refusal names one row's own problem (its verb changed, the
    allocation moved), and failing the whole confirm - after the SPO, its lines and its
    allocations are already written - would leave the operator with a container they cannot
    re-create and no link either.
    """
    if not allocations or not any(ticked.values()):
        return []

    from app.services.project_order_inquiry_service import ProjectOrderInquiryService

    service = ProjectOrderInquiryService(db)
    by_line: dict[str, list[dict]] = {}
    for allocation in allocations:
        by_line.setdefault(str(allocation["shipment_line_id"]), []).append(allocation)

    out: list[dict] = []
    for line_id, keys in ticked.items():
        pool = [dict(a) for a in by_line.get(line_id, [])]
        if not pool:
            continue
        for key in keys:
            if not key.startswith(f"{_COVERAGE_PROJECT}:"):
                continue
            row_id = key.split(":", 1)[1]
            row = (
                db.query(_inquiry_row_model())
                .filter(_inquiry_row_model().id == row_id)
                .first()
            )
            if row is None:
                continue
            need = float(row.qty or 0) - _linked_qty(db, row_id)
            if need <= 0:
                continue
            # The allocation at this row's OWN location first - that is what the tick put
            # there - then whatever else this line landed, because a row served from the
            # other half of the same container is still served.
            wanted_code = (row.stock_location or "").strip().upper() or None
            codes = _warehouse_ids_by_code(db) if wanted_code else {}
            preferred = codes.get(wanted_code) if wanted_code else None
            pool.sort(key=lambda a: 0 if str(a["warehouse_id"]) == str(preferred) else 1)
            for allocation in pool:
                if need <= 0:
                    break
                left = float(allocation.get("qty") or 0)
                if left <= 0:
                    continue
                take = min(need, left)
                try:
                    service.place_on_po_allocations(
                        row_id,
                        [{"spo_allocation_id": allocation["allocation_id"], "qty": take}],
                        actor_user_id=actor_user_id,
                    )
                except AppException as exc:  # noqa: PERF203 - one row's problem, not the batch's
                    logger.warning(
                        "SPO link refused for order inquiry row %s: %s", row_id, exc.detail
                    )
                    break
                allocation["qty"] = left - take
                need -= take
                out.append({
                    "key": key,
                    "document": row.item_code,
                    "spo_number": _spo_number_of(db, allocation["allocation_id"]),
                    "qty": take,
                })
    return out


def _inquiry_row_model():
    from app.models.project_so import OrderInquiryRow

    return OrderInquiryRow


def _linked_qty(db: Session, row_id: str) -> float:
    from app.models.project_so import OrderInquiryLink

    total = (
        db.query(func.coalesce(func.sum(OrderInquiryLink.qty), 0))
        .filter(OrderInquiryLink.row_id == str(row_id))
        .scalar()
    )
    return float(total or 0)


def _spo_number_of(db: Session, allocation_id: str) -> Optional[str]:
    return (
        db.query(SPOAllocation.spo_number)
        .filter(SPOAllocation.id == str(allocation_id))
        .scalar()
    )


def unwind(db: Session, shipment_id: str, purchase_order_id: Optional[str] = None) -> dict:
    """Undo `create` for one shipment - the Delete action on the planner's "Created SPOs"
    grid (captain live case, 21 Aug: he created SPOs, then deleted their `spo_allocations` on
    the SPO Allocations screen, and had no way back to a clean "suggest" state because the SPO
    itself, and the link naming it, were both still there). Deletes the `purchase_order_lines`
    + `purchase_orders` headers this shipment's `create` minted, every MATCHED
    `shipment_line_spo_link` row pointing at them, and any `spo_allocations` still hanging off
    those PO lines. That last one is not automatic: the FK from `spo_allocations.po_line_id`
    is `ON DELETE SET NULL`, not `CASCADE` (a stock arrival can legitimately have no PO behind
    it), so deleting the PO line alone would leave a now-untraceable allocation still counting
    as incoming supply - the exact half-undone state this action exists to avoid.

    **Seventh amendment (R1, `PLAN-scm-spo-planner-feedback-3sep.md`): scoped to ONE SPO.**
    A shipment can carry several SPOs now (its own docstring, above `suggest`); `purchase_
    order_id` (absent = every SPO this shipment ever produced, the legacy shape every earlier
    caller/test sends) narrows every delete below - PO, lines, links, allocations, pull
    reversal - to that ONE header, so deleting one SPO never touches another's. A `purchase_
    order_id` that does not name one of THIS shipment's own SPOs is 404, same as no SPO at
    all - the caller cannot use this to reach a header belonging elsewhere. SKIP rows
    (`unmatched_reason` set, no `purchase_order_id` to scope by) are left alone on a scoped
    unwind - they simply explain a line's history and do not affect `remaining_qty`, which
    only sums MATCHED rows - and are still swept on a full (unscoped) unwind, as before.

    **F9 (review round): deletes the `OrderInquiryLink` rows this SPO's own allocations
    carry BEFORE deleting the allocations.** `spo_allocation_id`'s FK is `ON DELETE SET
    NULL`, not `CASCADE` - but `ck_order_inquiry_links_one_target` requires exactly ONE of
    `po_line_id` / `spo_allocation_id` to be set, so a SET NULL that left a link with
    NEITHER was never a legal end state for one of these rows: the bulk allocation delete
    below raised a CHECK violation instead of completing. These links are this SPO's OWN
    placements (S7's `place_on_po_allocations`), so deleting them first simply returns the
    row to unlinked, which `_project_coverage` then offers in full - exactly what "delete
    the SPO" should mean for the demand it was serving.

    Since the doctrine correction (module docstring, fifth amendment), also REVERSES the
    `qty_received` `create` advanced on every SOURCE PO line these SPO lines pulled from -
    parsed back out of each deleted line's own `source_ref` JSON. Without this, deleting a CRM
    SPO would leave the source PO permanently short by whatever it lent, with no SPO left to
    explain where it went - a real loss, not merely an undo. Guarded the same defensive way
    `_heal_stale_links` is: a source line already gone by some other path is simply skipped,
    left to that function's own self-heal story rather than failing this one.

    Two hard guards, both refused with a 409 rather than a partial delete:
      * every header touched must carry `source_system == crm_spo` (`SOURCE_SYSTEM`) - an
        AutoCount import must never be deleted from this screen, however it ended up linked;
      * only headers actually LINKED TO THIS SHIPMENT are touched at all - the query never
        reaches a PO belonging to another shipment's conversion.

    After this, `suggest(db, shipment_id)` returns to its normal (non-`already_converted`)
    state - nothing is left behind to make the next `suggest` call answer any differently than
    a shipment that never ran "Create SPO" at all. Same permission as `create`
    (`scm.reorder.run`, enforced by the route) - unwinding a PO-book write is the same class of
    write as making one.
    """
    shipment = _shipment_or_404(db, shipment_id)
    links = _existing_links(db, shipment.id)
    if not links:
        raise AppException(404, "This shipment has no SPO to delete.")

    po_ids = {str(l.purchase_order_id) for l in links if l.purchase_order_id}
    if not po_ids:
        raise AppException(404, "This shipment has no SPO to delete.")

    if purchase_order_id is not None:
        if str(purchase_order_id) not in po_ids:
            raise AppException(404, "This shipment has no SPO with that id to delete.")
        po_ids = {str(purchase_order_id)}
        links = [
            l for l in links
            if l.purchase_order_id and str(l.purchase_order_id) in po_ids
        ]

    pos = db.query(PurchaseOrder).filter(PurchaseOrder.id.in_(po_ids)).all()
    foreign = [po for po in pos if po.source_system != SOURCE_SYSTEM]
    if foreign:
        names = ", ".join(sorted({po.po_number or "?" for po in foreign}))
        raise AppException(
            409,
            f"{names} was not created by Create SPO and cannot be deleted from this screen.",
            detail="not_crm_spo",
        )

    po_numbers = sorted({po.po_number or "?" for po in pos})

    po_lines = (
        db.query(PurchaseOrderLine)
        .filter(PurchaseOrderLine.purchase_order_id.in_(po_ids))
        .all()
    )
    po_line_ids = [str(pl.id) for pl in po_lines]

    restored_source_lines = _reverse_advances(db, po_lines)

    deleted_allocations = 0
    if po_line_ids:
        allocation_ids = [
            str(row[0])
            for row in db.query(SPOAllocation.id)
            .filter(SPOAllocation.po_line_id.in_(po_line_ids))
            .all()
        ]
        if allocation_ids:
            # F9 (review round): this SPO's own order-inquiry placements, deleted BEFORE
            # the allocations they point at - see the docstring above for why the bulk
            # allocation delete raises without this.
            from app.models.project_so import OrderInquiryLink

            db.query(OrderInquiryLink).filter(
                OrderInquiryLink.spo_allocation_id.in_(allocation_ids)
            ).delete(synchronize_session=False)
        deleted_allocations = (
            db.query(SPOAllocation)
            .filter(SPOAllocation.po_line_id.in_(po_line_ids))
            .delete(synchronize_session=False)
        )

    for pl in po_lines:
        db.delete(pl)
    for po in pos:
        db.delete(po)
    for link in links:
        db.delete(link)
    db.flush()

    return {
        "shipment_id": str(shipment.id),
        "shipment_number": shipment.shipment_number,
        "deleted_po_numbers": po_numbers,
        "deleted_spo_count": len(pos),
        "deleted_allocation_count": deleted_allocations,
        "restored_po_line_count": restored_source_lines,
    }


def plan_of(db: Session, purchase_order_id: str) -> dict:
    """The PO detail's "Plan" card (R1, AC-H5): what a CRM SPO's own lines PULLED from, and
    which retail sales-order lines they COVER - read purely off each line's own `source_ref`,
    the fact `create` already wrote (module docstring, fifth/seventh amendments), never a
    second write.

    `{"pulls": [{"purchase_order_id", "po_number", "po_line_label", "qty"}], "covers":
    [{"so_number", "customer", "qty", "warehouse"}]}`. `pulls` names the SOURCE PO line each
    take drew from (its own id, number and product, since the pin can resolve to a PO booked
    under a differently-spelled supplier - fourth amendment) so the FE can link to it; `covers`
    is the RETAIL half of `source_ref.so_coverage`
    only - the project half is written as `OrderInquiryLink` rows instead (a retail line has
    no link row to hang on, module docstring's `_link_ticked_demand`), so a project tick does
    not appear here; it is already visible on the order-inquiry worklist's own "Linked to".

    Empty for a PO this module did not create (`source_system != crm_spo`) - an AutoCount
    import has no plan to read."""
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == purchase_order_id).one_or_none()
    if po is None or po.source_system != SOURCE_SYSTEM:
        return {"pulls": [], "covers": []}

    spo_lines = (
        db.query(PurchaseOrderLine)
        .filter(PurchaseOrderLine.purchase_order_id == po.id)
        .all()
    )
    if not spo_lines:
        return {"pulls": [], "covers": []}

    parsed = [(ln, parse_source_ref(ln.source_ref)) for ln in spo_lines]

    pull_line_ids = {pl_id for _ln, p in parsed for pl_id, _qty in p["pulls"]}
    source_lines = (
        {
            str(row.id): row
            for row in db.query(PurchaseOrderLine)
            .filter(PurchaseOrderLine.id.in_(pull_line_ids))
            .all()
        }
        if pull_line_ids
        else {}
    )
    source_po_ids = {str(sl.purchase_order_id) for sl in source_lines.values()}
    source_pos = (
        {
            str(row.id): row
            for row in db.query(PurchaseOrder).filter(PurchaseOrder.id.in_(source_po_ids)).all()
        }
        if source_po_ids
        else {}
    )

    pulls: list[dict] = []
    for _ln, p in parsed:
        for pl_id, qty in p["pulls"]:
            src_line = source_lines.get(pl_id)
            src_po = source_pos.get(str(src_line.purchase_order_id)) if src_line else None
            pulls.append({
                "purchase_order_id": str(src_po.id) if src_po else None,
                "po_number": src_po.po_number if src_po else None,
                "po_line_label": (
                    src_line.product.product_code
                    if (src_line and src_line.product) else None
                ),
                "qty": qty,
            })

    so_line_ids = {so_id for _ln, p in parsed for so_id, _qty in p["so_coverage"]}
    so_rows: dict[str, tuple] = {}
    if so_line_ids:
        rows = (
            db.query(
                SalesOrderLine.id, SalesOrder.so_number, Customer.customer_name,
                Warehouse.warehouse_code,
            )
            .join(SalesOrder, SalesOrder.id == SalesOrderLine.sales_order_id)
            .outerjoin(Customer, Customer.id == SalesOrder.customer_id)
            .outerjoin(Warehouse, Warehouse.id == SalesOrderLine.warehouse_id)
            .filter(SalesOrderLine.id.in_(so_line_ids))
            .all()
        )
        so_rows = {str(r[0]): r for r in rows}

    covers: list[dict] = []
    for _ln, p in parsed:
        for so_id, qty in p["so_coverage"]:
            row = so_rows.get(so_id)
            covers.append({
                "so_number": row[1] if row else None,
                "customer": row[2] if row else None,
                "qty": qty,
                "warehouse": row[3] if row else None,
            })

    return {"pulls": pulls, "covers": covers}


def parse_source_ref(source_ref: Optional[str]) -> dict[str, list]:
    """What a `crm_spo` line records about where it came from and what it is for.

    `{"pulls": [{"po_line_id", "qty"}], "so_coverage": [{"so_line_id", "qty"}]}`. A BARE
    LIST is the older encoding (pulls only) and still reads, because rows written before the
    retail ticks existed are on file and a delete path must not start raising on them.

    Read defensively throughout - absent, unparsable, or another writer's unrelated
    `source_ref` all read as "nothing recorded" rather than raising.
    """
    empty: dict[str, list] = {"pulls": [], "so_coverage": []}
    if not source_ref:
        return empty
    try:
        data = json.loads(source_ref)
    except (ValueError, TypeError):
        return empty
    if isinstance(data, list):
        raw_pulls, raw_coverage = data, []
    elif isinstance(data, dict):
        raw_pulls = data.get("pulls") or []
        raw_coverage = data.get("so_coverage") or []
    else:
        return empty

    out: dict[str, list] = {"pulls": [], "so_coverage": []}
    for item in raw_pulls if isinstance(raw_pulls, list) else []:
        try:
            out["pulls"].append((str(item["po_line_id"]), float(item["qty"])))
        except (KeyError, TypeError, ValueError):
            continue
    for item in raw_coverage if isinstance(raw_coverage, list) else []:
        try:
            out["so_coverage"].append((str(item["so_line_id"]), float(item["qty"])))
        except (KeyError, TypeError, ValueError):
            continue
    return out


def _parse_pulls(source_ref: Optional[str]) -> list[tuple[str, float]]:
    """Which open PO lines this SPO line drew from, and how much of each."""
    return parse_source_ref(source_ref)["pulls"]


def _reverse_advances(db: Session, po_lines: list[PurchaseOrderLine]) -> int:
    """`unwind`'s own half of the fifth amendment: undo the `qty_received` bump `create` made
    on every SOURCE PO line these (about-to-be-deleted) `crm_spo` lines pulled from. Floored at
    zero rather than allowed to go negative - a source line touched again by something else
    since `create` ran must not be dragged below what IT now legitimately holds."""
    reversals: dict[str, float] = {}
    for pl in po_lines:
        for src_id, qty in _parse_pulls(pl.source_ref):
            reversals[src_id] = reversals.get(src_id, 0.0) + qty
    if not reversals:
        return 0
    sources = db.query(PurchaseOrderLine).filter(PurchaseOrderLine.id.in_(reversals.keys())).all()
    for src in sources:
        taken = reversals.get(str(src.id), 0.0)
        src.qty_received = max(float(src.qty_received or 0) - taken, 0.0)
    return len(sources)


# --------------------------------------------------------------------------- #
# the AutoCount handoff worksheet
# --------------------------------------------------------------------------- #

_COLUMNS = ["SUPPLIER", "MODEL", "DESCRIPTION", "QTY", "UNIT COST", "CURRENCY", "CRM SPO NO"]


def worksheet_payload(db: Session, shipment_id: str) -> dict:
    """What the office keys into AutoCount, read back from what `create` wrote. 404 when
    this shipment has never had "Create SPO" run - there is nothing to hand off yet."""
    shipment = _shipment_or_404(db, shipment_id)
    rows = (
        db.query(ShipmentLineSpoLink, InboundShipmentLine, PurchaseOrder)
        .join(InboundShipmentLine, InboundShipmentLine.id == ShipmentLineSpoLink.inbound_shipment_line_id)
        .outerjoin(PurchaseOrder, PurchaseOrder.id == ShipmentLineSpoLink.purchase_order_id)
        .filter(
            ShipmentLineSpoLink.inbound_shipment_id == shipment.id,
            ShipmentLineSpoLink.purchase_order_id.isnot(None),
        )
        .all()
    )
    if not rows:
        raise AppException(
            404,
            "No SPO has been created from this shipment yet.",
            detail="not_converted",
        )

    product_ids = [str(l.product_id) for _link, l, _po in rows]
    products = {
        str(pid): (code, name)
        for pid, code, name in db.query(*_product_cols()).filter(_product_id_in(product_ids)).all()
    }
    supplier_ids = {str(po.supplier_id) for _l, _sl, po in rows if po and po.supplier_id}
    suppliers = {
        str(s.id): s.supplier_name
        for s in db.query(Supplier).filter(Supplier.id.in_(supplier_ids)).all()
    } if supplier_ids else {}

    lines = []
    for link, ship_line, po in rows:
        code, name = products.get(str(ship_line.product_id), (None, None))
        po_line = (
            db.query(PurchaseOrderLine)
            .filter(PurchaseOrderLine.id == link.purchase_order_line_id)
            .one_or_none()
        )
        lines.append({
            "supplier_name": suppliers.get(str(po.supplier_id)) if po else None,
            "product_code": code,
            "product_name": name,
            "qty": _f(po_line.qty_ordered) if po_line else None,
            "unit_cost": _f(po_line.unit_cost) if po_line else None,
            "currency": po_line.currency if po_line else None,
            "po_number": po.po_number if po else None,
        })
    lines.sort(key=lambda l: (l["supplier_name"] or "", l["product_code"] or ""))

    return {
        "shipment_id": str(shipment.id),
        "shipment_number": shipment.shipment_number,
        "container_no": shipment.shipping_container_number,
        "lines": lines,
    }


def to_xlsx(payload: dict) -> bytes:
    """The worksheet as a workbook - same pattern as `consolidated_packing_list.to_xlsx`."""
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active or wb.create_sheet()
    ws.title = "SPO WORKSHEET"
    bold = Font(bold=True)

    for label, value in (
        ("CONTAINER", payload.get("container_no") or payload.get("shipment_number")),
        ("SHIPMENT", payload.get("shipment_number")),
    ):
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = bold
    ws.append([])

    ws.append(list(_COLUMNS))
    for col in range(1, len(_COLUMNS) + 1):
        ws.cell(row=ws.max_row, column=col).font = bold

    for line in payload["lines"]:
        ws.append([
            line["supplier_name"],
            line["product_code"],
            line["product_name"],
            line["qty"],
            line["unit_cost"],
            line["currency"],
            line["po_number"],
        ])

    for i, width in enumerate([28, 18, 34, 10, 12, 10, 20], start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_filename(payload: dict) -> str:
    import re

    stem = payload.get("container_no") or payload.get("shipment_number") or payload["shipment_id"]
    stem = re.sub(r"[^A-Za-z0-9._-]", "", str(stem)) or str(payload["shipment_id"])
    return f"{stem}-spo-worksheet.xlsx"
