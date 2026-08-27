"""The container request as an `.xlsx`, in the supplier's own layout - R13.

Ms Tee's ask, verbatim: send them the SAME sheet back with the quantity to load filled in. Not
a sheet that carries the same data - the same sheet. So when their stock list was retained,
this loads THEIR workbook and writes ONE column into it: their title row, their header
spellings, their row order, their merged families, their yellow fields, their red figures,
their column widths and their `合计` formulas are all still theirs, because nothing here
rebuilds them.

That is the change from the first version, which replayed their VALUES into a workbook of ours
and lost every merge and every fill on the way (a family's volume printed nine times reads as
nine times the volume), and which fell back to five columns of our own naming when there was no
file to answer in - so a supplier's document changed shape because of a file WE failed to keep.
The five-column sheet is gone (AC-D6): with no retained file the SAME eleven columns are drawn
fresh, in their styling, with no merges (we hold no family information and inventing one would
be wrong on the first product with two sizes).

One model, three renderers (R12): `supplier_document_model.SheetModel` is built once by the
send path and drawn by this module, by the PDF and by the public page. This module never reads
the database.

THE ROUND TRIP IS THE CONTRACT (AC-C2 of part 3). The supplier's next stock list is very often
this file with the numbers changed, so whatever goes out has to come back in through
`supplier_inventory_reader.read_workbook` - asserted on both layouts in
`tests/scm/test_container_request_xlsx.py`.
"""
from __future__ import annotations

import re
from copy import copy
from datetime import datetime
from io import BytesIO

from sqlalchemy.orm import Session

from app.services.scm import supplier_document_model as sheet_model
from app.services.scm.supplier_document_model import (  # re-exported: one spelling, one place
    NO_FILE_TITLE,
    QTY_TO_LOAD_HEADER,
    SheetModel,
)

#: Their own measurements, off the July file, for the document we draw ourselves. Column K
#: repeats column J's width, which is also the rule the in-place path applies.
NO_FILE_WIDTHS = [
    6.0,
    28.7109375,
    10.140625,
    12.7109375,
    13.28515625,
    13.42578125,
    19.0,
    13.7109375,
    17.140625,
    13.85546875,
    13.85546875,
]
DEFAULT_WIDTH = 13.85546875

TITLE_HEIGHT = 39.75
HEADER_HEIGHT = 28.5
DATA_HEIGHT = 18.75
TOTALS_HEIGHT = 22.5

DATA_FONT = "宋体"
HEADER_FONT = "Calibri"
YELLOW_RGB = "FFFFFF00"
RED_RGB = "FFFF0000"


def filename(supplier: dict) -> str:
    """`container-request-{code}-{stamp}.xlsx` - the PDF's own stem, other extension (AC-C1).

    Same sanitising as `supplier_notice_service._request_filename`, so the two files a
    supplier receives in one email are named alike rather than nearly alike.
    """
    code = (supplier.get("supplier_code") or "supplier").replace("/", "-").replace(" ", "-")
    stamp = datetime.utcnow().strftime("%Y%m%d")
    return f"container-request-{code}-{stamp}.xlsx"


def build(db: Session, *, supplier_id: str, lines: list[dict]) -> bytes:
    """Model then render, for a caller that holds no model yet (the download route, tests).

    The send path builds the model once and hands the SAME object here and to the PDF
    (AC-D7), so this convenience must stay a two-line wrapper rather than a second path.
    """
    return render(sheet_model.build(db, supplier_id=supplier_id, lines=lines))


def render(model: SheetModel) -> bytes:
    """The document as bytes: their workbook when we have one, ours when we do not."""
    if model.source is not None:
        return _with_qty_to_load(model)
    return _fresh(model)


# --------------------------------------------------------------------------- their workbook


def _with_qty_to_load(model: SheetModel) -> bytes:
    """Their file, with column K written into it. Nothing else in the sheet is touched."""
    import openpyxl
    from openpyxl.utils import get_column_letter

    src = model.source
    assert src is not None  # `render` guards it; this keeps the type checker honest
    wb = openpyxl.load_workbook(BytesIO(src.data))
    ws = wb[src.sheet_title] if src.sheet_title in wb.sheetnames else wb.active

    qty_col = src.qty_col
    like_col = qty_col - 1  # column J: the last column they styled themselves
    qty_letter = get_column_letter(qty_col)
    like_letter = get_column_letter(like_col)

    appended = [row for row in model.rows if row.appended]
    shift = len(appended)
    new_last_data_row = src.last_data_row + shift

    # The 合计 row moves down BEFORE the appended rows are written, because they are written
    # over the rows it used to occupy.
    if src.totals_row is not None and shift:
        _move_row(ws, src.totals_row, src.totals_row + shift, last_col=qty_col)

    header = ws.cell(row=src.header_row, column=qty_col)
    _copy_style(ws.cell(row=src.header_row, column=like_col), header)
    header.value = QTY_TO_LOAD_HEADER
    ws.column_dimensions[qty_letter].width = (
        ws.column_dimensions[like_letter].width or DEFAULT_WIDTH
    )

    for row in model.rows:
        if row.source_row is None:
            continue
        cell = ws.cell(row=row.source_row, column=qty_col)
        _copy_style(ws.cell(row=row.source_row, column=like_col), cell)
        cell.value = row.cells[-1].value

    for offset, row in enumerate(appended):
        target = src.last_data_row + 1 + offset
        for pos in range(1, qty_col + 1):
            style_from = ws.cell(row=src.last_data_row, column=min(pos, like_col))
            cell = ws.cell(row=target, column=pos)
            _copy_style(style_from, cell)
            cell.value = row.cells[pos - 1].value
        ws.row_dimensions[target].height = ws.row_dimensions[src.last_data_row].height

    if src.totals_row is not None:
        totals_at = src.totals_row + shift
        for pos in range(1, qty_col):
            cell = ws.cell(row=totals_at, column=pos)
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = _widen(cell.value, src.last_data_row, new_last_data_row)
        total = ws.cell(row=totals_at, column=qty_col)
        _copy_style(ws.cell(row=totals_at, column=like_col), total)
        total.value = (
            f"=SUM({qty_letter}{src.first_data_row}:{qty_letter}{new_last_data_row})"
        )

    return _bytes(wb)


def _move_row(ws, source: int, target: int, *, last_col: int) -> None:
    """Move one whole row down, merges included, and leave nothing behind.

    `insert_rows` would be the obvious call and is the wrong one: openpyxl does not carry
    merged ranges or row heights with it, and this row is the one whose merge (`A120:E120`)
    and formulas are the point.
    """
    moved = []
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row == source and rng.max_row == source:
            moved.append((rng.min_col, rng.max_col))
            ws.unmerge_cells(str(rng))

    for pos in range(1, last_col + 1):
        old = ws.cell(row=source, column=pos)
        new = ws.cell(row=target, column=pos)
        _copy_style(old, new)
        new.value = old.value
        old.value = None
    ws.row_dimensions[target].height = ws.row_dimensions[source].height

    for min_col, max_col in moved:
        ws.merge_cells(
            start_row=target, start_column=min_col, end_row=target, end_column=max_col
        )


def _copy_style(source, target) -> None:
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)
    target.number_format = source.number_format


def _widen(formula: str, old_last: int, new_last: int) -> str:
    """`=SUM(F3:F119)` covering four appended rows becomes `=SUM(F3:F123)`.

    Only the END of a range that stopped at their last data row moves: a total that stopped
    short of the rows we added would understate the ask, and a total we re-anchored anywhere
    else would stop being their formula.
    """

    def bump(match: re.Match) -> str:
        return (
            f"{match.group(1)}{new_last}"
            if int(match.group(2)) == old_last
            else match.group(0)
        )

    return re.sub(r":(\$?[A-Z]{1,3}\$?)(\d+)", lambda m: ":" + bump(m), formula)


# --------------------------------------------------------------------------- our workbook


def _fresh(model: SheetModel) -> bytes:
    """The same eleven columns, drawn in their styling, when there is no file to answer in."""
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active or wb.create_sheet()
    ws.title = "Container request"

    ncols = len(model.columns)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    centre = Alignment(horizontal="center", vertical="center")
    yellow = PatternFill(fill_type="solid", fgColor=YELLOW_RGB)

    for pos in range(1, ncols + 1):
        width = NO_FILE_WIDTHS[pos - 1] if pos <= len(NO_FILE_WIDTHS) else DEFAULT_WIDTH
        ws.column_dimensions[get_column_letter(pos)].width = width

    title = ws.cell(row=1, column=1)
    title.value = model.title or NO_FILE_TITLE
    title.font = Font(name=HEADER_FONT, sz=22, b=True)
    title.alignment = centre
    title.border = border
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.row_dimensions[1].height = TITLE_HEIGHT

    for pos, column in enumerate(model.columns, start=1):
        cell = ws.cell(row=2, column=pos)
        cell.value = QTY_TO_LOAD_HEADER if pos == ncols else column.label
        cell.font = Font(name=HEADER_FONT, sz=14, b=True)
        cell.alignment = centre
        cell.border = border
    ws.row_dimensions[2].height = HEADER_HEIGHT

    first_data_row = 3
    for offset, row in enumerate(model.rows):
        r = first_data_row + offset
        for pos, source in enumerate(row.cells[:ncols], start=1):
            cell = ws.cell(row=r, column=pos)
            cell.value = source.value
            cell.font = Font(
                name=DATA_FONT, sz=14, color=RED_RGB if source.red else None
            )
            cell.alignment = centre
            cell.border = border
            if source.fill:
                cell.fill = yellow
        ws.row_dimensions[r].height = DATA_HEIGHT

    last_data_row = first_data_row + len(model.rows) - 1
    _fresh_totals(
        ws,
        model,
        row=max(last_data_row + 1, first_data_row + 1),
        first_data_row=first_data_row,
        last_data_row=last_data_row,
        border=border,
        centre=centre,
    )
    return _bytes(wb)


def _fresh_totals(
    ws, model: SheetModel, *, row: int, first_data_row: int, last_data_row: int, border, centre
) -> None:
    """Their `合计：` row: the label merged across the first columns, sums in red."""
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    totals = model.totals
    if totals is None:
        return

    ncols = len(model.columns)
    for pos in range(1, ncols + 1):
        cell = ws.cell(row=row, column=pos)
        cell.alignment = centre
        cell.border = border
        cell.font = Font(name=DATA_FONT, sz=18 if pos == 1 else 14, color=RED_RGB)

    label = totals.cells[0]
    ws.cell(row=row, column=1).value = label.value
    if label.colspan > 1:
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=min(label.colspan, ncols)
        )

    for pos, source in enumerate(totals.cells[:ncols], start=1):
        if pos == 1 or source.covered or source.value is None:
            continue
        letter = get_column_letter(pos)
        ws.cell(row=row, column=pos).value = (
            f"=SUM({letter}{first_data_row}:{letter}{max(last_data_row, first_data_row)})"
        )
    ws.row_dimensions[row].height = TOTALS_HEIGHT


# --------------------------------------------------------------------------- shared


def _bytes(wb) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


__all__ = ["NO_FILE_TITLE", "QTY_TO_LOAD_HEADER", "build", "filename", "render"]
