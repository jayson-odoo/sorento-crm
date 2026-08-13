"""L3 - reading the Order Inquiry sheet.

This is the file that carries the two things the SO and PO books do not: the **stock
location** a sales-order line ships from, and the **purchase order** that line is waiting on.
Both are maintained by the people who own them, so nothing here has to be curated by hand.

The customer keeps it in two shapes and the reader takes both:

  * `Order Inquiry Form.xlsx` - one sheet, with `STOCK LOCATION` and a `REMARK` column that
    holds either the literal `ORDER` (nothing placed yet) or a purchase-order number.
  * `JAN - DEC 2026 ORDERabc.xlsx` - 35 sheets, one per month plus dated and summary tabs,
    with `SUPPLIER` and `PO NO` as their own columns and no location.

So the columns are found by HEADER NAME, never by position: the two shapes disagree about
both the column order and which columns exist at all, and a positional reader would be right
on one file and silently wrong on the other. Every sheet with a recognisable header row is
read - a workbook of monthly tabs is one book, and reading only the first would drop eleven
months without saying so.

`ORDER` in the remark column is not a parse failure. It is the sheet stating that this line
has nothing on order yet, which is a fact worth keeping: it is the difference between "we
have not bought it" and "we do not know".
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Optional

from app.services.scm.outstanding_reader import sheet_rows

#: Header spellings seen in the customer's own files, per canonical field. Matched after
#: normalisation (upper, collapsed whitespace, punctuation stripped), so `S/O NO` and
#: `PO NO ` with its trailing space both land.
_HEADERS: dict[str, tuple[str, ...]] = {
    "so_number": ("SO NO", "S O NO", "SO NUMBER", "SALES ORDER NO", "SO"),
    "item_code": ("ITEM CODE", "ITEM", "PRODUCT CODE", "STOCK CODE"),
    "qty": ("QTY", "QUANTITY", "ORDER QTY"),
    "so_date": ("SO DATE", "ORDER DATE"),
    "delivery_date": ("DELIVERY DATE", "REQUIRED DATE", "ETA"),
    "project": ("PROJECT CUSTOMER", "PROJECT", "CUSTOMER", "PROJECT NAME"),
    "location": ("STOCK LOCATION", "LOCATION", "WAREHOUSE"),
    "supplier": ("SUPPLIER", "SUPPLIER NAME"),
    "po_number": ("PO NO", "PO NUMBER", "PURCHASE ORDER NO"),
    "remark": ("REMARK", "REMARKS", "NOTE", "NOTES"),
}

#: The remark that means "nothing is on order for this line yet". Matched as a WHOLE WORD:
#: `BACKORDER` and `REORDERED` are not this, and a substring test would read both as
#: "nothing placed".
_NOT_ORDERED = re.compile(r"\bORDER\b", re.I)

#: A purchase-order number in any family the customer uses: `202510-S0025`, `SPO-2020/01-0001`,
#: `PO-2020/01-0001`. Used with `findall`, not `match`, because ONE sales-order line can wait
#: on more than one purchase order - the sheet writes `202606-S0024 & 202607-S0043` - and
#: matching the whole cell drops both.
_PO_NUMBER = re.compile(r"\b(?:[A-Z]{2,4}-)?\d{4}[-/]?\d{0,2}[-/]?[A-Z]?\d{3,4}\b", re.I)


def _normalise(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(re.sub(r"[^A-Z0-9 ]+", " ", str(value).upper()).split())


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _number(value: Any) -> Optional[float]:
    if value is None or isinstance(value, (date, datetime)):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip().replace(",", "")
    try:
        return float(raw) if raw else None
    except ValueError:
        return None


def _as_date(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


@dataclass
class OrderInquiryRow:
    so_number: str
    item_code: str
    qty: Optional[float]
    so_date: Optional[date]
    delivery_date: Optional[date]
    project: str
    #: Warehouse code, when the sheet carries one. The monthly book does not.
    location: str
    supplier: str
    #: Every purchase order this line waits on. Usually one, sometimes two - a line split
    #: across two orders is written `202606-S0024 & 202607-S0043`, and taking only the first
    #: would silently lose half the supply the line is waiting for.
    po_numbers: tuple[str, ...]
    #: True when the sheet said `ORDER` - explicitly, nothing on order for this line, or for
    #: the remainder of it in `202605-S0042 & ORDER`. Distinct from a sheet that simply has
    #: no PO column: "we have not bought it" and "we do not know" are different answers.
    not_ordered: bool
    sheet: str
    source_row: int


@dataclass
class OrderInquiryResult:
    rows: list[OrderInquiryRow] = field(default_factory=list)
    problems: list[str] = field(default_factory=list)
    sheets_read: list[str] = field(default_factory=list)
    sheets_skipped: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return bool(self.rows)

    @property
    def with_location(self) -> int:
        return sum(1 for r in self.rows if r.location)

    @property
    def with_po(self) -> int:
        return sum(1 for r in self.rows if r.po_numbers)

    @property
    def po_claims(self) -> int:
        """One claim per (line, purchase order), so a split line counts twice."""
        return sum(len(r.po_numbers) for r in self.rows)


def _header_map(row: tuple) -> Optional[dict[str, int]]:
    """Column index per canonical field, if this row is a header row."""
    found: dict[str, int] = {}
    for idx, cell in enumerate(row):
        text = _normalise(cell)
        if not text:
            continue
        for fieldname, spellings in _HEADERS.items():
            if fieldname in found:
                continue
            if text in spellings:
                found[fieldname] = idx
    # A sheet is only a data sheet if it names the two things every row must have.
    return found if {"so_number", "item_code"} <= set(found) else None


def _read_po(values: dict[str, Any]) -> tuple[tuple[str, ...], bool]:
    """Every purchase order this line waits on, and whether the sheet said part of it is not
    ordered yet.

    Two columns can carry it. `PO NO` is explicit; `REMARK` holds either the literal `ORDER`
    or the number, which is how the single-sheet variant records it. Both can hold a LIST:
    `202606-S0024 & 202607-S0043` is one line split across two orders, and
    `202605-S0042 & ORDER` is a line partly ordered and partly not - both facts worth keeping,
    and both lost by matching the cell as a whole.
    """
    numbers: list[str] = []
    not_ordered = False
    for key in ("po_number", "remark"):
        raw = _text(values.get(key))
        if not raw:
            continue
        if _NOT_ORDERED.search(raw):
            not_ordered = True
        for found in _PO_NUMBER.findall(raw):
            if found not in numbers:
                numbers.append(found)
    return tuple(numbers), not_ordered


def read_order_inquiry(file_data: bytes) -> OrderInquiryResult:
    """Read every sheet that has a recognisable header row."""
    result = OrderInquiryResult()
    try:
        import openpyxl

        from io import BytesIO

        wb = openpyxl.load_workbook(BytesIO(file_data), data_only=True, read_only=True)
        sheets = [(name, list(wb[name].iter_rows(values_only=True))) for name in wb.sheetnames]
        wb.close()
    except Exception:
        # Not a zip workbook. The legacy path reads one sheet, which is what a `.xls`
        # Order Inquiry would be.
        try:
            sheets = [("Sheet", list(sheet_rows(file_data)))]
        except Exception as exc:  # noqa: BLE001
            result.problems.append(f"this file could not be read: {exc}")
            return result

    for name, rows in sheets:
        header: Optional[dict[str, int]] = None
        header_at = 0
        for row_number, row in enumerate(rows, start=1):
            if header is None:
                header = _header_map(row)
                if header is not None:
                    header_at = row_number
                continue

            values = {f: row[i] for f, i in header.items() if i < len(row)}
            so_number = _text(values.get("so_number"))
            item_code = _text(values.get("item_code"))
            if not so_number or not item_code:
                continue

            po_numbers, not_ordered = _read_po(values)
            result.rows.append(
                OrderInquiryRow(
                    so_number=so_number,
                    item_code=item_code,
                    qty=_number(values.get("qty")),
                    so_date=_as_date(values.get("so_date")),
                    delivery_date=_as_date(values.get("delivery_date")),
                    project=_text(values.get("project")),
                    location=_text(values.get("location")).upper(),
                    supplier=_text(values.get("supplier")),
                    po_numbers=po_numbers,
                    not_ordered=not_ordered,
                    sheet=name,
                    source_row=row_number,
                )
            )

        if header is None:
            # A summary or working tab. Named rather than ignored: a workbook of monthly
            # sheets where one silently fails to parse looks exactly like a quiet month.
            result.sheets_skipped.append(name)
        else:
            result.sheets_read.append(name)
        del header_at

    if not result.rows:
        result.problems.append(
            "this file does not look like an Order Inquiry sheet: no header row naming a "
            "sales order and an item code was found."
        )
    return result
