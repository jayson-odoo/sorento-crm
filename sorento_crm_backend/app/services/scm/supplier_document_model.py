"""ONE model of the supplier document, drawn by three renderers - R12.

Ms Tee's ask, verbatim: the link view, the PDF and the xlsx tally 100% with the supplier's own
`库存明细.xlsx`, cbm merging included, formatting included. Before this module there were three
renderers with three shapes - the xlsx replayed their values, the PDF carried six columns of
our own naming and the public page carried another six - so "tally" was impossible to hold: a
change to one was a change to one.

So the document is built ONCE, here, and the renderers only draw it:

* `container_request_xlsx.render(model)` - their own workbook with column K written into it,
  or a fresh workbook in their styling when there is no file of theirs to answer in;
* `supplier_notice_service._document_html(..., sheet=model)` - the PDF, as an HTML table with
  the merges as `rowspan`;
* `supplier_notice_service.public_request_page` - the same model as JSON for the link page.

**Their ten columns, plus ours** (captain's Q1): 序号 / 型号 / 商标 / 规格 / 品名 / 包装好库存 /
空瓷 / 体积(cbm) / 总体积(cbm) / 备注, then `需装数量 / Qty to load`. Their spellings, their
order, their row order - we append, never rewrite.

**A family is a merge, not a repetition.** Their sheet gives one 序号 and one volume to a whole
family of models (`A3:A11`, `H3:H11`, `I3:I11`), and the merged ranges are NOT aligned across
columns (`A40:A43` sits inside `H40:H50`). So every cell carries its own `rowspan` rather than
the row carrying one family number, and `family_span` is simply the 序号 column's span - the
alternative invents families the supplier did not write.

**Without a retained file we build the same eleven columns from what we know** (AC-D6): no
merges, because we hold no family information and inventing one would be wrong on the first
product with two sizes; 商标 = the company letter the product belongs to; their holdings off
the stock snapshot, which is their own latest statement about their own warehouse.
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Optional

from sqlalchemy import text
from sqlalchemy.orm import Session

from app.services.import_alias_service import AliasResolver
from app.services.scm.outstanding_reader import all_sheet_rows
from app.services.scm.supplier_inventory_reader import DOC_TYPE

logger = logging.getLogger(__name__)

#: Column K, in both languages for the same reason the rest of the document is: their staff
#: read the Chinese, ours have to be able to check what went out.
QTY_TO_LOAD_LABEL_ZH = "需装数量"
QTY_TO_LOAD_LABEL_EN = "Qty to load"
QTY_TO_LOAD_HEADER = f"{QTY_TO_LOAD_LABEL_ZH} / {QTY_TO_LOAD_LABEL_EN}"

#: The title line of a document we draw ourselves. Theirs names their factory and the date
#: their list was taken; ours can only name what the document IS.
NO_FILE_TITLE = "配柜要求 / Container request"

#: What an appended row says about itself, so the supplier can see at a glance that the line
#: is ours and not a row of theirs they have forgotten.
NOT_ON_LIST_REMARK = "不在库存表 / Not on your list"

#: Their own word for the last row. Matched on a prefix because the file writes it with a
#: full-width colon and sometimes trailing spaces.
TOTALS_PREFIX = "合计"

YELLOW = "yellow"

#: Their ten columns as the July file writes them, with the field each resolves to. Used for
#: the no-file document (AC-D6) and as the English second line on every document, so the page
#: can label a column the supplier's own sheet named only in Chinese.
DEFAULT_COLUMNS: list[tuple[str, str, Optional[str]]] = [
    ("序号", "No.", None),
    ("型号", "Model", "item_code"),
    ("商标", "Brand", "brand"),
    ("规格", "Spec", "spec"),
    ("品名", "Description", "product_name"),
    ("包装好库存", "Packed", "qty_packed"),
    ("空瓷", "Unfinished", "qty_unfinished"),
    ("体积(cbm)", "CBM/unit", "cbm_per_unit"),
    ("总体积(cbm)", "Total CBM", "cbm_total"),
    ("备注", "Remarks", "remark"),
]

_LABELS_EN = {field: en for _zh, en, field in DEFAULT_COLUMNS if field}

#: Which of their columns the 合计 row adds up on the no-file document - the same three theirs
#: does, plus ours.
_TOTALLED_FIELDS = ("qty_packed", "qty_unfinished", "cbm_total")


@dataclass
class Column:
    """One column of the document. `label` is theirs; `label_en` is our second line."""

    label: str
    label_en: Optional[str] = None
    field: Optional[str] = None


@dataclass
class Cell:
    value: Any = None
    #: How many rows/columns this cell spans. 1 unless their sheet merged it.
    rowspan: int = 1
    colspan: int = 1
    #: True when a merge that starts above or to the left covers this position, so the
    #: renderer skips it rather than drawing an empty box inside the merge.
    covered: bool = False
    #: The only two marks their sheet makes on a value, both of which mean something to the
    #: person reading it: a yellow field is one they maintain, a red figure is one to notice.
    fill: Optional[str] = None
    red: bool = False

    def to_dict(self) -> dict:
        return {
            "value": _plain(self.value),
            "rowspan": self.rowspan,
            "colspan": self.colspan,
            "covered": self.covered,
            "fill": self.fill,
            "red": self.red,
        }


@dataclass
class Row:
    cells: list[Cell]
    #: The 序号 column's span: how many of their rows this product family covers. 0 on a row
    #: that continues a family started above.
    family_span: int = 1
    #: True for a row we added because they never listed the product (AC-D2).
    appended: bool = False
    #: Where this row sits in their workbook (1-based), so the xlsx renderer can write column
    #: K into their file rather than rebuilding it. None for an appended row.
    source_row: Optional[int] = None

    def to_dict(self) -> dict:
        return {
            "cells": [c.to_dict() for c in self.cells],
            "family_span": self.family_span,
            "appended": self.appended,
        }


@dataclass
class SheetSource:
    """Their workbook, and where in it the model's rows came from.

    Present only when the retained file could be opened for writing. It is what lets the xlsx
    renderer hand back THEIR file with one column added rather than a copy of it (R13).
    """

    data: bytes
    sheet_title: str
    header_row: int
    first_data_row: int
    last_data_row: int
    totals_row: Optional[int]
    qty_col: int


@dataclass
class SheetModel:
    columns: list[Column]
    rows: list[Row] = field(default_factory=list)
    totals: Optional[Row] = None
    title: Optional[str] = None
    source: Optional[SheetSource] = None

    def column_index(self, field_name: str) -> int:
        """The position of a canonical field, or -1 when their sheet has no such column."""
        for pos, col in enumerate(self.columns):
            if col.field == field_name:
                return pos
        return -1

    @property
    def qty_index(self) -> int:
        return len(self.columns) - 1

    def to_dict(self) -> dict:
        return {
            "title": self.title,
            "columns": [{"label": c.label, "label_en": c.label_en} for c in self.columns],
            "rows": [r.to_dict() for r in self.rows],
            "totals": self.totals.to_dict() if self.totals else None,
        }


# --------------------------------------------------------------------------- build


def build(
    db: Session, *, supplier: dict, supplier_id: str, lines: list[dict]
) -> SheetModel:
    """The document for one container request. Never raises: a bad stored file falls back.

    ``lines`` are the reviewed lines as `supplier_notice_service._request_pack` holds them:
    `{product_id, item_code, product_name, qty}`.
    """
    raw = _retained_stock_list(db, supplier_id)
    if raw:
        try:
            return _from_their_sheet(raw, db=db, supplier_id=supplier_id, lines=lines)
        except Exception as exc:  # noqa: BLE001 - a stored file is not the caller's fault
            logger.warning(
                "supplier document: supplier %s has a retained stock list that could not be "
                "answered in (%s); building the document from our own data",
                supplier_id,
                exc,
            )
    return _from_our_data(db, supplier_id=supplier_id, lines=lines)


def _retained_stock_list(db: Session, supplier_id: str) -> Optional[bytes]:
    """The bytes of the stock list they last sent us, if it was retained.

    Best-effort by construction: retention is itself best-effort
    (`store_stock_list_attachment` logs and continues), so an absent or unreachable object is
    an ordinary state here rather than an error.
    """
    try:
        from app.services.resources_service import AttachmentService
        from app.services.scm import supplier_inventory_service

        held = supplier_inventory_service.latest_stock_list_attachment(
            db, supplier_id=supplier_id
        )
        if not held:
            return None
        return AttachmentService(db).get_file_content(held["attachment_id"])
    except Exception as exc:  # noqa: BLE001
        logger.warning(
            "supplier document: could not read supplier %s's retained stock list (%s)",
            supplier_id,
            exc,
        )
        return None


# --------------------------------------------------------------------------- their sheet


def _from_their_sheet(
    data: bytes, *, db: Session, supplier_id: str, lines: list[dict]
) -> SheetModel:
    """Their workbook as a model: values, merges, fills, red figures, plus our column.

    An old `.xls` (OLE2) cannot be opened by openpyxl at all, and those are exactly the
    suppliers who send the oldest files, so that container is read for its VALUES through the
    same reader the import uses and rendered fresh. Their row order and their headers survive
    either way; only their cell styling cannot, because the file never carried it to us.
    """
    workbook = _openable(data)
    if workbook is None:
        return _from_their_values(data, db=db, supplier_id=supplier_id, lines=lines)

    ws = workbook.active
    if ws is None:
        raise ValueError("the retained workbook has no sheet")

    resolver = AliasResolver.for_doc_type(db, DOC_TYPE)
    header_row, fields = _header_row(ws, resolver)
    if header_row is None:
        raise ValueError("no item_code column in the retained stock list")

    ncols = _column_count(ws, header_row)
    columns = [
        Column(
            label=_text(ws.cell(row=header_row, column=c).value) or "",
            label_en=_LABELS_EN.get(fields.get(c) or ""),
            field=fields.get(c),
        )
        for c in range(1, ncols + 1)
    ]
    columns.append(
        Column(label=QTY_TO_LOAD_LABEL_ZH, label_en=QTY_TO_LOAD_LABEL_EN, field="qty_to_load")
    )

    spans, covered = _merge_map(ws)
    totals_row = _totals_row(ws, header_row)
    last_data_row = (totals_row - 1) if totals_row else ws.max_row

    model = SheetModel(
        columns=columns,
        title=_title(ws, header_row),
        source=SheetSource(
            data=data,
            sheet_title=ws.title,
            header_row=header_row,
            first_data_row=header_row + 1,
            last_data_row=last_data_row,
            totals_row=totals_row,
            qty_col=ncols + 1,
        ),
    )

    code_col = next((c for c, f in fields.items() if f == "item_code"), None)
    if code_col is None:
        raise ValueError("no item_code column in the retained stock list")
    asks = _Asks(db, supplier_id=supplier_id, lines=lines)

    for r in range(header_row + 1, last_data_row + 1):
        cells = [
            _cell_of(ws, r, c, spans=spans, covered=covered) for c in range(1, ncols + 1)
        ]
        if all(c.value is None for c in cells) and not any(c.covered for c in cells):
            continue
        code = _text(ws.cell(row=r, column=code_col).value)
        cells.append(Cell(value=_qty(asks.take(code))))
        first = cells[0]
        model.rows.append(
            Row(
                cells=cells,
                family_span=0 if first.covered else first.rowspan,
                source_row=r,
            )
        )

    _append_unlisted(model, asks.left_over(), serial=_next_serial(model))
    model.totals = _totals_of(model, ws=ws, totals_row=totals_row, ncols=ncols)
    return model


def _from_their_values(
    data: bytes, *, db: Session, supplier_id: str, lines: list[dict]
) -> SheetModel:
    """Their rows out of a container openpyxl cannot open for writing (an old `.xls`)."""
    rows = [list(r) for r in all_sheet_rows(data)]
    resolver = AliasResolver.for_doc_type(db, DOC_TYPE)
    header_idx = None
    fields: dict[int, str] = {}
    for idx, raw in enumerate(rows):
        mapped = {
            pos: resolver.field_for_header(cell)
            for pos, cell in enumerate(raw)
            if resolver.field_for_header(cell)
        }
        if "item_code" in mapped.values():
            header_idx, fields = idx, mapped
            break
    if header_idx is None:
        raise ValueError("no item_code column in the retained stock list")

    ncols = max(len(r) for r in rows)
    columns = [
        Column(
            label=_text(rows[header_idx][c] if c < len(rows[header_idx]) else None) or "",
            label_en=_LABELS_EN.get(fields.get(c) or ""),
            field=fields.get(c),
        )
        for c in range(ncols)
    ]
    columns.append(
        Column(label=QTY_TO_LOAD_LABEL_ZH, label_en=QTY_TO_LOAD_LABEL_EN, field="qty_to_load")
    )

    code_col = next(c for c, f in fields.items() if f == "item_code")
    asks = _Asks(db, supplier_id=supplier_id, lines=lines)
    model = SheetModel(
        columns=columns,
        title=_text(rows[0][0]) if header_idx > 0 and rows[0] else None,
    )
    for raw in rows[header_idx + 1 :]:
        padded = list(raw) + [None] * (ncols - len(raw))
        if all(v is None for v in padded):
            continue
        code = _text(padded[code_col])
        cells = [Cell(value=_plain(v)) for v in padded]
        cells.append(Cell(value=_qty(asks.take(code))))
        model.rows.append(Row(cells=cells))

    _append_unlisted(model, asks.left_over(), serial=_next_serial(model))
    model.totals = _computed_totals(model, [c.field for c in model.columns])
    return model


def _openable(data: bytes):
    """Their workbook, opened for WRITING, or None when this container cannot be."""
    try:
        import openpyxl

        return openpyxl.load_workbook(BytesIO(data))
    except Exception:  # noqa: BLE001 - an .xls, or bytes that are not a workbook at all
        return None


def _header_row(ws, resolver: AliasResolver) -> tuple[Optional[int], dict[int, str]]:
    """The header row and its column map, by `supplier_inventory_reader`'s own rule.

    Not row 1 by decree: these files carry a title line and sometimes a blank one above it.
    The header is the first row that resolves an item code - the same test the reader applies,
    so a file it can read is a file this can answer in.
    """
    for r in range(1, min(ws.max_row, 20) + 1):
        mapped: dict[int, str] = {}
        for c in range(1, ws.max_column + 1):
            field_name = resolver.field_for_header(ws.cell(row=r, column=c).value)
            if field_name:
                mapped[c] = field_name
        if "item_code" in mapped.values():
            return r, mapped
    return None, {}


def _column_count(ws, header_row: int) -> int:
    """Their last real column: the last header cell they wrote, or the last data cell."""
    last = 0
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=c).value is not None:
            last = c
    for r in range(header_row + 1, ws.max_row + 1):
        for c in range(last + 1, ws.max_column + 1):
            if ws.cell(row=r, column=c).value is not None:
                last = c
    return last or ws.max_column


def _merge_map(ws) -> tuple[dict[tuple[int, int], tuple[int, int]], set[tuple[int, int]]]:
    spans: dict[tuple[int, int], tuple[int, int]] = {}
    covered: set[tuple[int, int]] = set()
    for rng in ws.merged_cells.ranges:
        rows = rng.max_row - rng.min_row + 1
        cols = rng.max_col - rng.min_col + 1
        spans[(rng.min_row, rng.min_col)] = (rows, cols)
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                if (r, c) != (rng.min_row, rng.min_col):
                    covered.add((r, c))
    return spans, covered


def _cell_of(ws, r: int, c: int, *, spans, covered) -> Cell:
    cell = ws.cell(row=r, column=c)
    rows, cols = spans.get((r, c), (1, 1))
    return Cell(
        value=_plain(cell.value),
        rowspan=rows,
        colspan=cols,
        covered=(r, c) in covered,
        fill=YELLOW if _is_yellow(cell) else None,
        red=_is_red(cell),
    )


def _is_yellow(cell) -> bool:
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return False
    rgb = getattr(fill.fgColor, "rgb", None)
    return isinstance(rgb, str) and rgb.upper().endswith("FFFF00")


def _is_red(cell) -> bool:
    color = cell.font.color if cell.font else None
    rgb = getattr(color, "rgb", None) if color else None
    return isinstance(rgb, str) and rgb.upper().endswith("FF0000")


def _title(ws, header_row: int) -> Optional[str]:
    for r in range(1, header_row):
        for c in range(1, ws.max_column + 1):
            value = _text(ws.cell(row=r, column=c).value)
            if value:
                return value
    return None


def _totals_row(ws, header_row: int) -> Optional[int]:
    for r in range(ws.max_row, header_row, -1):
        first = _text(ws.cell(row=r, column=1).value)
        if first and first.startswith(TOTALS_PREFIX):
            return r
    return None


def _totals_of(model: SheetModel, *, ws, totals_row: Optional[int], ncols: int) -> Optional[Row]:
    """Their 合计 row, with the figures computed rather than left as `=SUM(...)`.

    The xlsx renderer keeps their formulas (it moves the row rather than rewriting it); the
    PDF and the page need a number, and computing it from the rows above is the same answer
    Excel shows - and the only one available before Excel has opened the file.
    """
    if totals_row is None:
        return None
    spans, covered = _merge_map(ws)
    label_cells = [
        _cell_of(ws, totals_row, c, spans=spans, covered=covered) for c in range(1, ncols + 1)
    ]
    totalled = [c for c, cell in enumerate(label_cells) if c > 0 and cell.value is not None]
    for pos in totalled:
        label_cells[pos].value = _sum_of(model, pos)
    label_cells.append(Cell(value=_sum_of(model, model.qty_index)))
    return Row(cells=label_cells, family_span=0, source_row=totals_row)


def _computed_totals(model: SheetModel, fields: list[Optional[str]]) -> Row:
    """The 合计 row for a document we are building ourselves: their three sums, plus ours."""
    cells: list[Cell] = []
    for pos, field_name in enumerate(fields):
        if pos == 0:
            cells.append(Cell(value=f"{TOTALS_PREFIX}：", colspan=5))
        elif pos < 5:
            cells.append(Cell(covered=True))
        elif field_name in _TOTALLED_FIELDS or field_name == "qty_to_load":
            cells.append(Cell(value=_sum_of(model, pos)))
        else:
            cells.append(Cell())
    return Row(cells=cells, family_span=0)


def _sum_of(model: SheetModel, pos: int) -> Optional[float]:
    total = 0.0
    seen = False
    for row in model.rows:
        if pos >= len(row.cells):
            continue
        value = row.cells[pos].value
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            continue
        total += float(value)
        seen = True
    if not seen:
        return None
    return _round(total)


# --------------------------------------------------------------------------- our data


def _from_our_data(db: Session, *, supplier_id: str, lines: list[dict]) -> SheetModel:
    """The same eleven columns, built from what we hold (AC-D6).

    Their holdings come off the snapshot rather than off a file, because on this branch there
    is no file - and a row with no snapshot behind it (a product they have never listed) still
    belongs on the sheet, with the holdings blank rather than zero.
    """
    columns = [Column(label=zh, label_en=en, field=f) for zh, en, f in DEFAULT_COLUMNS]
    columns.append(
        Column(label=QTY_TO_LOAD_LABEL_ZH, label_en=QTY_TO_LOAD_LABEL_EN, field="qty_to_load")
    )
    model = SheetModel(columns=columns, title=NO_FILE_TITLE)

    held = _snapshot(db, supplier_id)
    letters = _company_letters(db, [ln.get("product_id") for ln in lines])

    for serial, line in enumerate(lines, start=1):
        code = _text(line.get("item_code"))
        stock = held.get(code or "", {})
        packed = stock.get("qty_packed")
        per_unit = stock.get("cbm_per_unit")
        cells = [
            Cell(value=serial),
            Cell(value=code, fill=YELLOW),
            Cell(
                value=stock.get("brand") or letters.get(str(line.get("product_id"))),
                fill=YELLOW,
            ),
            Cell(value=stock.get("spec"), fill=YELLOW),
            Cell(value=line.get("product_name") or stock.get("product_name"), fill=YELLOW),
            Cell(value=packed, fill=YELLOW, red=packed == 0),
            Cell(value=stock.get("qty_unfinished"), fill=YELLOW),
            Cell(value=per_unit),
            Cell(
                value=_round(per_unit * packed)
                if per_unit is not None and packed is not None
                else None
            ),
            Cell(value=stock.get("remark"), fill=YELLOW),
            Cell(value=_qty(line.get("qty"))),
        ]
        model.rows.append(Row(cells=cells))

    model.totals = _computed_totals(model, [c.field for c in model.columns])
    return model


def _snapshot(db: Session, supplier_id: str) -> dict[str, dict]:
    from app.models.scm import SupplierInventory

    rows = (
        db.query(
            SupplierInventory.item_code,
            SupplierInventory.qty_packed,
            SupplierInventory.qty_unfinished,
            SupplierInventory.cbm_per_unit,
            SupplierInventory.product_name,
            SupplierInventory.brand,
            SupplierInventory.spec,
            SupplierInventory.remark,
        )
        .filter(SupplierInventory.supplier_id == supplier_id)
        .all()
    )
    return {
        str(r.item_code): {
            "qty_packed": _number(r.qty_packed),
            "qty_unfinished": _number(r.qty_unfinished),
            "cbm_per_unit": _number(r.cbm_per_unit),
            "product_name": r.product_name,
            "brand": r.brand,
            "spec": r.spec,
            "remark": r.remark,
        }
        for r in rows
    }


def _company_letters(db: Session, product_ids: list) -> dict[str, str]:
    """`S` / `C` / `M` - the first letter of the company code the product belongs to.

    That is what their 商标 column carries on their own sheet, and it is the only brand mark
    a supplier needs: which of our companies the model is for.
    """
    ids = [str(p) for p in product_ids if p]
    if not ids:
        return {}
    rows = db.execute(
        text(
            "SELECT p.id, c.code FROM products p JOIN companies c ON c.id = p.company_id "
            "WHERE p.id = ANY(CAST(:ids AS uuid[]))"
        ),
        {"ids": ids},
    ).mappings().all()
    return {str(r["id"]): str(r["code"])[:1].upper() for r in rows if r["code"]}


# --------------------------------------------------------------------------- the ask


class _Asks:
    """Which requested quantity belongs on which of their rows (AC-D3, AC-F12.6).

    Three ways a row of theirs can be the row we are asking about, in order:

    1. their code IS the code we sent (a set line goes out under the supplier's own code, and
       that is the code their sheet already carries);
    2. their code is bound to the product we are asking for by the stock snapshot - their
       model number and ours are different strings, and the snapshot row is what joins them;
    3. neither, in which case the line is appended below their last row rather than dropped:
       it is still part of the ask.
    """

    def __init__(self, db: Session, *, supplier_id: str, lines: list[dict]):
        self._lines = [dict(ln) for ln in lines]
        self._by_code: dict[str, int] = {}
        self._by_product: dict[str, int] = {}
        for pos, line in enumerate(self._lines):
            code = _key(line.get("item_code"))
            if code and code not in self._by_code:
                self._by_code[code] = pos
            product_id = line.get("product_id")
            if product_id and str(product_id) not in self._by_product:
                self._by_product[str(product_id)] = pos
        self._bindings = _snapshot_bindings(db, supplier_id) if self._by_product else {}
        self._taken: set[int] = set()

    def take(self, their_code: Optional[str]) -> Optional[float]:
        pos = self._match(their_code)
        if pos is None:
            return None
        self._taken.add(pos)
        return self._lines[pos].get("qty")

    def _match(self, their_code: Optional[str]) -> Optional[int]:
        key = _key(their_code)
        if not key:
            return None
        pos = self._by_code.get(key)
        if pos is not None and pos not in self._taken:
            return pos
        product_id = self._bindings.get(key)
        if product_id:
            pos = self._by_product.get(product_id)
            if pos is not None and pos not in self._taken:
                return pos
        return None

    def left_over(self) -> list[dict]:
        return [ln for pos, ln in enumerate(self._lines) if pos not in self._taken]


def _snapshot_bindings(db: Session, supplier_id: str) -> dict[str, str]:
    from app.models.scm import SupplierInventory

    rows = (
        db.query(SupplierInventory.item_code, SupplierInventory.product_id)
        .filter(
            SupplierInventory.supplier_id == supplier_id,
            SupplierInventory.product_id.isnot(None),
        )
        .all()
    )
    return {_key(r.item_code): str(r.product_id) for r in rows if _key(r.item_code)}


def _append_unlisted(model: SheetModel, lines: list[dict], *, serial: int) -> None:
    """AC-D2: below their last family, own 序号 continuing the count, and a remark saying so."""
    code_at = model.column_index("item_code")
    name_at = model.column_index("product_name")
    remark_at = model.column_index("remark")
    for offset, line in enumerate(lines):
        cells = [Cell() for _ in model.columns]
        cells[0] = Cell(value=serial + offset)
        if code_at >= 0:
            cells[code_at] = Cell(value=_text(line.get("item_code")), fill=YELLOW)
        if name_at >= 0:
            cells[name_at] = Cell(value=line.get("product_name"), fill=YELLOW)
        if remark_at >= 0:
            cells[remark_at] = Cell(value=NOT_ON_LIST_REMARK, fill=YELLOW)
        cells[-1] = Cell(value=_qty(line.get("qty")))
        model.rows.append(Row(cells=cells, appended=True))


def _next_serial(model: SheetModel) -> int:
    """Their last 序号, plus one. Their count continues; it does not restart under their rows."""
    last = 0
    for row in model.rows:
        value = row.cells[0].value if row.cells else None
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            last = max(last, int(value))
    return last + 1


# --------------------------------------------------------------------------- values


def _qty(value) -> Optional[float]:
    """What to write in the Qty to load cell, and `None` is a real answer (AC-D3).

    A zero would read as "pack none of these", which is a different instruction from "we did
    not ask about these" - and the supplier acts on the difference.
    """
    if value is None:
        return None
    number = float(value)
    if number <= 0:
        return None
    return int(number) if number == int(number) else number


def _number(value) -> Optional[float]:
    if value is None:
        return None
    number = float(value)
    return int(number) if number == int(number) else number


def _round(value: float) -> float:
    rounded = round(float(value), 4)
    return int(rounded) if rounded == int(rounded) else rounded


def _text(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    stripped = str(value).strip()
    return stripped or None


def _key(value) -> Optional[str]:
    text_value = _text(value)
    return text_value.upper() if text_value else None


def _plain(value):
    """A cell value the page can read: no Decimal, no date, no Excel-only object."""
    if isinstance(value, Decimal):
        return _number(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value
