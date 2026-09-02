"""Hold the supplier's proforma invoices, with the prices they actually state.

Two-step like every other upload channel here: `preview` and `validate` describe, `apply`
writes, and `?validate_only=true` returns the same `{valid, errors, warnings, summary}`
verdict a Test means everywhere else in this system.

What is particular to this channel:

  * **One file is several invoices.** Jinbaichuan's pre-loading list is five proforma
    invoices stacked down one sheet and states no document number for any of them, so a
    positional one is derived (`PI-<file stem>-<block>`) - stable across a re-upload, which
    is what makes the second upload update five invoices instead of creating five more.
  * **A price is never stored without its currency.** The document usually says (`RMB`,
    `单价(元)`), the supplier's price list sometimes says, and where neither does the upload
    is refused with the gap named rather than stored in a house default (AC-P3.2).
  * **Products are RESOLVED, never created**, on an exact case-insensitive code match. An
    unmatched line is still stored - it is a real charge on a real invoice - and named in
    the preview so somebody can go and fix the catalogue.
"""
from __future__ import annotations

import re
import uuid
from datetime import date as _date, datetime
from decimal import ROUND_HALF_UP, Decimal
from typing import Any, Optional

from sqlalchemy import func, text
from sqlalchemy.orm import Session, joinedload

from app.models.procurement import InboundShipment, InboundShipmentLine, Supplier
from app.models.product import Product
from app.models.product_set import ProductSet, ProductSetMember
from app.models.scm import (
    ContainerSize,
    ProformaInvoice,
    ProformaInvoiceLine,
    ProformaInvoiceShipmentLink,
    SupplierProductCodeAlias,
)
from app.services.audit_service import log_audit
from app.services.company_scope import get_company_scope, resolve_write_company_id
from app.services.company_scope_sql import company_sql_predicate
from app.services.error_handler import AppException
from app.services.numbering_defaults import (
    INBOUND_SHIPMENT_DRAFT_DOC_TYPE,
    seed_inbound_shipment_draft_rule,
)
from app.services.numbering_service import NumberingService
from app.services.scm.currency_resolution import resolve_currency
from app.services.scm.proforma_invoice_reader import (
    ProformaDocument,
    ProformaReadResult,
    read_workbook,
)
from app.services.scm.supplier_scope import (  # noqa: F401  (assert_supplier is re-exported)
    assert_supplier,
    is_uuid as _is_uuid,
    supplier_check as _supplier_check,
    supplier_label as _supplier_label,
    supplier_mismatch_warning as _supplier_mismatch_warning,
)
from app.services.scm.upload_validation import envelope, named

SOURCE_SYSTEM = "scm_proforma_invoice"

#: What an invoice is called when the document states no number of its own. Positional, so
#: the same file re-uploaded produces the same names and lands on the same invoices.
_DERIVED_PREFIX = "PI"

#: A stated total and the sum of the lines are money; a cent of rounding is not a discrepancy.
_TOTAL_TOLERANCE = 0.01

#: The sentence the operator has to be able to act on. Quoted in the UAC (AC-P3.2), so it is
#: one constant rather than two spellings that drift between Test and Apply.
_NO_CURRENCY = (
    "Nothing says which money these prices are in - state the currency this invoice is in."
)

# --- PI -> draft inbound shipment convert (packing-list amendment, 20 Aug evening) --------

#: Not-yet-a-real-packing-list. Distinct from every OTHER value in the column's vocabulary,
#: all of which describe a container the AGENT has actually loaded or shipped - this one
#: exists only because CRM lines were pre-filled from proforma invoices (AC per
#: PLAN-scm-proforma-to-spo.md's Amendment). The real packing list, when it arrives, is
#: uploaded through the existing `packing_list_service.apply` path same as any other.
_DRAFT_SHIPMENT_STATUS = "draft"

#: `NumberingService` doc_type for a draft packing list's own series - kept distinct from a
#: real container's number (usually the container number itself, or `PRELOAD-...`) so the two
#: can never collide. The rule behind it is `PL-{YYMM}-{NNN}`, defined once in
#: `app.services.numbering_defaults` and seeded from there by migration 440, by
#: `bootstrap_env`, and by `_draft_shipment_number` for a company that has no series yet.
#:
#: There is NO random fallback. It used to invent `SHIP-DRAFT-<hex8>` when the rule was
#: missing, and because the rule had never been seeded that random suffix was what every
#: converted packing list was actually called: two of them sort in no order, neither can be
#: read down a phone, and nobody can tell which came first.
_DRAFT_NUMBER_DOC_TYPE = INBOUND_SHIPMENT_DRAFT_DOC_TYPE

#: Where an invoice's goods have got to. `split` is a real state, not a rounding of
#: `converted`: one invoice legitimately sits in two containers (Q9), and until every line
#: is placed there is still something to convert.
_PLACEMENTS = ("not_converted", "converted", "split")


def _uuid() -> str:
    return str(uuid.uuid4())


def _f(value: Any) -> Optional[float]:
    return None if value is None else float(value)


def _parse(db: Session, data: bytes) -> ProformaReadResult:
    return read_workbook(data, db=db)


def pi_number_for(doc: ProformaDocument, *, source_ref: Optional[str]) -> str:
    """What this invoice is called, so two blocks in one file are two invoices.

    The document's own number when it states one, verbatim. Otherwise the file's own name
    plus the block's position: the pre-loading list numbers none of its five invoices, and
    deriving rather than generating is what makes a re-upload update them in place (AC-P2.5).
    """
    if doc.pi_number:
        return doc.pi_number.strip()[:100]
    # The STEM is truncated, not the composed name: a long filename would otherwise push the
    # block index off the end of a `String(100)` column and turn five distinct invoices into
    # one name that each block in turn overwrites.
    stem = (source_ref or "proforma").rsplit("/", 1)[-1].rsplit(".", 1)[0][:80]
    return f"{_DERIVED_PREFIX}-{stem}-{doc.index}"


def _products_by_code(db: Session, codes: set[str]) -> dict[str, dict]:
    """Catalogue lookup, scoped to the caller's company.

    Raw SQL bypasses the ORM's company filter, and product codes are NOT unique across
    companies - so unscoped, an invoice line could resolve to another company's product.
    Exact and case-insensitive on this rung. Since R16 the CALLER hands the supplier in and
    the ladder (`supplier_code_matcher`) answers the codes this one cannot - the supplier's
    own spelling, remembered - so what stays unmatched here is genuinely unmatched.
    """
    if not codes:
        return {}
    predicate, params = company_sql_predicate(db, "p.company_id", param_prefix="c")
    rows = db.execute(
        text(
            "SELECT p.id, p.product_code FROM products p "
            " WHERE upper(p.product_code) = ANY(:codes) "
            f"   AND {predicate or 'true'}"
        ),
        {"codes": [c.upper() for c in codes], **params},
    ).mappings().all()
    return {str(r["product_code"]).upper(): dict(r) for r in rows}


def _with_supplier_codes(
    db: Session,
    known: dict[str, dict],
    *,
    supplier_id: Optional[str],
    codes: set[str],
    actor: Optional[str] = None,
    remember: bool = True,
) -> dict[str, dict]:
    """Fill the gaps in an exact-match lookup with the supplier-code ladder (R16).

    The exact answers are kept as they are: a rung that agrees with the catalogue outranks
    every derivation, and re-deriving a code that already matched would be a chance to
    disagree with it. Only the codes nothing answered are asked about.
    """
    if not supplier_id:
        return known
    missing = {c for c in codes if c.upper() not in known}
    if not missing:
        return known

    from app.services.scm.supplier_code_matcher import resolve

    found = resolve(db, supplier_id, missing, remember=remember, actor=actor)
    if not found:
        return known
    rows = (
        db.query(Product.id, Product.product_code)
        .filter(Product.id.in_([m.product_id for m in found.values() if m.product_id]))
        .all()
    )
    by_id = {
        str(pid): {"id": str(pid), "product_code": code, "product_set_id": None}
        for pid, code in rows
    }
    # A code can name one of our SETS (R19). It carries no product id at all, so the entry
    # holds the set instead - `product_code` is the set's own code, because that is what
    # every screen showing "what this line matched" has to print.
    set_rows = (
        db.query(ProductSet.id, ProductSet.set_code)
        .filter(
            ProductSet.id.in_([m.product_set_id for m in found.values() if m.product_set_id])
        )
        .all()
    )
    by_set = {
        str(sid): {"id": None, "product_code": code, "product_set_id": str(sid)}
        for sid, code in set_rows
    }
    out = dict(known)
    for code, match in found.items():
        target = (
            by_set.get(str(match.product_set_id))
            if match.product_set_id
            else by_id.get(str(match.product_id))
        )
        if target:
            out[code.upper()] = target
    return out


def _currencies(
    db: Session,
    parsed: ProformaReadResult,
    *,
    supplier_id: Optional[str],
    requested: Optional[str],
) -> dict[int, tuple[Optional[str], str]]:
    """Per document, `(code, source)`. Per document because a file can hold several, and a
    currency shown against the wrong one is worse than no currency at all."""
    return {
        doc.index: resolve_currency(
            db, supplier_id=supplier_id, requested=requested, stated=doc.currency_hint
        )
        for doc in parsed.documents
    }


def _summarise(
    db: Session,
    parsed: ProformaReadResult,
    *,
    supplier_id: Optional[str],
    source_ref: Optional[str],
    requested_currency: Optional[str] = None,
    known: Optional[dict[str, dict]] = None,
    resolved: Optional[dict[int, tuple[Optional[str], str]]] = None,
    revision_candidates: Optional[dict[int, dict]] = None,
) -> dict[str, Any]:
    """What the file holds, described. `known` and `resolved` are injectable so `apply`,
    which needs both to do the writing, does not pay for them a second time to describe
    what it wrote.

    `supplier_check` (AC-G3) rides along here rather than living only in `validate`, so
    `preview` and `apply`'s own summary agree with it too: `{letterhead,
    chosen_supplier_name, other_supplier_name | null}`, or `None` when the file states no
    letterhead. `document_count` is what the verdict card calls "N invoice blocks" and
    `unmatched_items` is its "U codes unknown" (AC-G4) - both already computed below, named
    for what they were before this slice, not renamed for it.
    """
    codes = {ln.item_code for d in parsed.documents for ln in d.lines}
    if known is None:
        known = _products_by_code(db, codes)
    unknown = sorted({c for c in codes if c.upper() not in known})
    if resolved is None:
        resolved = _currencies(
            db, parsed, supplier_id=supplier_id, requested=requested_currency
        )

    documents = []
    priced_without_currency = 0
    for doc in parsed.documents:
        currency, source = resolved.get(doc.index, (None, "none"))
        if doc.priced_lines and not currency:
            priced_without_currency += doc.priced_lines
        documents.append(
            {
                "index": doc.index,
                "pi_number": pi_number_for(doc, source_ref=source_ref),
                "pi_number_stated": bool(doc.pi_number),
                "invoice_date": doc.invoice_date.isoformat() if doc.invoice_date else None,
                "container_no": doc.container_no,
                "bl_no": doc.bl_no,
                "lines": len(doc.lines),
                "qty": doc.total_qty,
                "total": doc.line_total,
                "stated_total": doc.stated_total,
                "unmatched_items": sorted(
                    {ln.item_code for ln in doc.lines if ln.item_code.upper() not in known}
                )[:50],
                "currency": currency,
                "currency_source": source,
                # The invoice on file this block looks like a new revision of, if any. A
                # PROPOSAL for the dialog to pre-select, never applied here (AC-E6).
                "revision_candidate": (revision_candidates or {}).get(doc.index),
            }
        )

    supplier_code, supplier_name = _supplier_label(db, supplier_id)
    file_currency, file_source = next(
        (resolved[d.index] for d in parsed.documents if resolved.get(d.index, (None,))[0]),
        (None, "none"),
    )
    return {
        "supplier_id": supplier_id,
        "supplier_code": supplier_code,
        "supplier_name": supplier_name,
        "documents": documents,
        "document_count": len(documents),
        "line_count": parsed.line_count,
        "priced_lines": parsed.priced_line_count,
        "rows_read": parsed.total_rows,
        "unmatched_item_codes": unknown[:200],
        "unmatched_items": len(unknown),
        "unmapped_headers": parsed.unmapped_headers,
        "currency": file_currency,
        "currency_source": file_source,
        "priced_lines_without_currency": priced_without_currency,
        "supplier_check": _supplier_check(db, parsed.letterhead, supplier_id=supplier_id),
    }


#: How much of a file's item codes must already sit on an un-converted invoice before the
#: upload dialog PROPOSES it as a revision. Half is deliberate and loose: the pre-loading
#: list carries no invoice number at all, so nothing identifies a resend except the goods it
#: names, and a supplier who drops two lines from a five-line container is still resending
#: the same container. It is a proposal, never a rule - the operator confirms (AC-E6).
_REVISION_OVERLAP = 0.5


def _revision_candidates(
    db: Session, parsed: ProformaReadResult, *, supplier_id: Optional[str]
) -> dict[int, dict]:
    """Per parsed document, the invoice on file it most looks like a new revision of.

    Only CURRENT, un-converted invoices of the same supplier are eligible: a superseded one
    has already been replaced, and a converted one's goods are on a shipment, so revising it
    would silently move the ground under a container that is already being loaded.
    """
    if not supplier_id or not parsed.documents:
        return {}

    rows = (
        db.query(ProformaInvoice.id, ProformaInvoice.pi_number, ProformaInvoice.invoice_date,
                 ProformaInvoice.created_at, ProformaInvoiceLine.item_code)
        .join(ProformaInvoiceLine, ProformaInvoiceLine.invoice_id == ProformaInvoice.id)
        .filter(
            ProformaInvoice.supplier_id == str(supplier_id),
            func.coalesce(ProformaInvoice.status, "current") == "current",
            # PLACED, not "has a link row". A SKIP row records a line no product matched
            # when some other invoice's convert ran; the JINBAICHUAN documents on the dev
            # database all carry one, so this excluded exactly the invoices a re-upload was
            # a revision of, and every second send landed as an independent PI.
            ~ProformaInvoice.id.in_(
                db.query(ProformaInvoiceShipmentLink.proforma_invoice_id).filter(
                    ProformaInvoiceShipmentLink.inbound_shipment_line_id.isnot(None)
                )
            ),
        )
        .all()
    )
    if not rows:
        return {}

    codes_by_invoice: dict[str, set[str]] = {}
    meta: dict[str, tuple[str, Any, Any]] = {}
    for inv_id, number, invoice_date, created_at, item_code in rows:
        codes_by_invoice.setdefault(str(inv_id), set()).add((item_code or "").upper())
        meta.setdefault(str(inv_id), (number, invoice_date, created_at))

    out: dict[int, dict] = {}
    for doc in parsed.documents:
        wanted = {ln.item_code.upper() for ln in doc.lines}
        if not wanted:
            continue
        # Ranked, never "whichever the dict yielded first": two invoices carrying the
        # same codes are a real state (the tester left a duplicate), and the offer has to
        # land on the same one every time. Best overlap, then the supplier's LATEST word -
        # its own document date, then when it was read, then the id as the final tiebreak.
        ranked = [
            (
                len(wanted & held) / len(wanted),
                meta[inv_id][1] or _date.min,
                meta[inv_id][2] or datetime.min,
                inv_id,
            )
            for inv_id, held in codes_by_invoice.items()
            if len(wanted & held) / len(wanted) >= _REVISION_OVERLAP
        ]
        if not ranked:
            continue
        ranked.sort(reverse=True)
        best = (ranked[0][0], ranked[0][3])
        number, invoice_date, _created_at = meta[best[1]]
        out[doc.index] = {
            "invoice_id": best[1],
            "pi_number": number,
            "invoice_date": invoice_date.isoformat() if invoice_date else None,
            "overlap_pct": round(best[0] * 100, 2),
            "matched_items": len(wanted & codes_by_invoice[best[1]]),
            "lines": len(wanted),
        }
    return out


def _revision_targets(
    db: Session, revision_of: Optional[dict], *, supplier_id: str
) -> dict[str, ProformaInvoice]:
    """The invoices this upload says it revises, checked before a single row is written.

    Checked up front rather than per document so a five-block file naming one bad target
    fails whole, instead of writing four revisions and then refusing.
    """
    if not revision_of:
        return {}
    ids = {str(v).strip() for v in revision_of.values() if str(v or "").strip()}
    if not ids:
        return {}
    if any(not _is_uuid(i) for i in ids):
        raise AppException(422, "That proforma invoice does not exist.", detail="revision_of")

    found = {
        str(inv.id): inv
        for inv in db.query(ProformaInvoice).filter(ProformaInvoice.id.in_(ids)).all()
    }
    missing = sorted(ids - set(found))
    if missing:
        raise AppException(
            422, "That proforma invoice does not exist.", detail="revision_of"
        )
    for inv in found.values():
        if str(inv.supplier_id) != str(supplier_id):
            raise AppException(
                422,
                f"'{inv.pi_number}' belongs to a different supplier, so this file cannot be "
                "a revision of it.",
                detail="revision_of",
            )
        if (inv.status or "current") == "superseded":
            raise AppException(
                409,
                f"'{inv.pi_number}' has already been superseded by a newer revision.",
                code="superseded",
            )
    return found


def _available_number(
    db: Session,
    supplier_id: str,
    base: str,
    attempt: int,
    marker: str = "R",
    always_suffix: bool = False,
) -> str:
    """A document number free for THIS supplier, starting from what the file derived.

    Identity is (company, supplier, pi_number) and the pre-loading list derives its number
    positionally from the file, so a revision taken from the same file lands on the number it
    is revising. `-R2` is appended rather than a random suffix, because the number is read by
    people and "PI-预装清单-1-R2" says what it is. `marker` is empty for a document filed as
    NEW rather than as a revision: it is not revision 2 of anything, it is a second
    document, and "-2" is what says so. `always_suffix` starts the search AT the suffix
    rather than trying the bare base first: a document filed as new is never the stem on its
    own, it is the next ordinal after it.
    """
    number = f"{base[:90]}-{marker}{attempt}" if always_suffix else base
    if always_suffix:
        attempt += 1
    while (
        db.query(ProformaInvoice)
        .filter(
            ProformaInvoice.supplier_id == str(supplier_id),
            ProformaInvoice.pi_number == number,
        )
        .first()
        is not None
    ):
        number = f"{base[:90]}-{marker}{attempt}"
        attempt += 1
    return number[:100]


def _chain(db: Session, invoice: ProformaInvoice) -> list[ProformaInvoice]:
    """Every revision of this document, oldest first.

    Walked with two bounded loops rather than a recursive CTE: a chain is two or three
    documents in practice, and the loop is readable by the next person. The bound is what
    stops a `revision_of_id` cycle - which `mark_as_revision_of` refuses to create, but a
    hand-edited row could still hold - from hanging the detail page.
    """
    seen: list[ProformaInvoice] = [invoice]
    guard = 0
    node = invoice
    while node.revision_of_id and guard < 50:
        prior = (
            db.query(ProformaInvoice)
            .filter(ProformaInvoice.id == str(node.revision_of_id))
            .first()
        )
        if prior is None or any(str(prior.id) == str(x.id) for x in seen):
            break
        seen.insert(0, prior)
        node = prior
        guard += 1

    node = invoice
    guard = 0
    while guard < 50:
        nxt = (
            db.query(ProformaInvoice)
            .filter(ProformaInvoice.revision_of_id == str(node.id))
            .first()
        )
        if nxt is None or any(str(nxt.id) == str(x.id) for x in seen):
            break
        seen.append(nxt)
        node = nxt
        guard += 1
    return seen


def _supplier_line_figures(db: Session, invoice_id: str) -> dict[tuple[str, int], dict]:
    """Per line, what the SUPPLIER stated on that invoice.

    The diff compares two of their statements, never our adjustment of one against their
    other: `qty` and `unit_price` are ours to trim, and diffing those would report a price
    change on a line whose price never moved.

    Keyed by item code AND its occurrence on the document, which is the only line identity
    carried across two sends: a revision's row numbers are its own, and Kailu's proforma
    names the same model on two lines at two prices, so keying on the code alone would
    compare one of them against neither.
    """
    lines = (
        db.query(ProformaInvoiceLine)
        .filter(ProformaInvoiceLine.invoice_id == str(invoice_id))
        .order_by(ProformaInvoiceLine.line_no)
        .all()
    )
    out: dict[tuple[str, int], dict] = {}
    seen: dict[str, int] = {}
    for ln in lines:
        occurrence = seen.get(ln.item_code, 0) + 1
        seen[ln.item_code] = occurrence
        qty = _f(ln.supplier_qty if ln.supplier_qty is not None else ln.qty)
        price = _f(
            ln.supplier_unit_price if ln.supplier_unit_price is not None else ln.unit_price
        )
        amount = round(qty * price, 4) if qty is not None and price is not None else _f(ln.amount)
        out[(ln.item_code, occurrence)] = {
            "item_code": ln.item_code,
            "occurrence": occurrence,
            "description": ln.description,
            "qty": qty,
            "unit_price": price,
            "amount": amount,
        }
    return out


def _moved(was: Optional[float], now: Optional[float]) -> bool:
    """Whether two stated figures actually differ. Money to the cent, quantities exactly."""
    if was is None and now is None:
        return False
    if was is None or now is None:
        return True
    return abs(was - now) > 0.005


def _diff(db: Session, invoice: ProformaInvoice) -> Optional[dict]:
    """What the supplier changed between the previous revision and this one (AC-E8)."""
    if not invoice.revision_of_id:
        return None
    previous = (
        db.query(ProformaInvoice)
        .filter(ProformaInvoice.id == str(invoice.revision_of_id))
        .first()
    )
    if previous is None:
        return None

    before = _supplier_line_figures(db, str(previous.id))
    after = _supplier_line_figures(db, str(invoice.id))

    changes: list[dict] = []
    price_changed = qty_changed = 0
    for key, now in after.items():
        code = now["item_code"]
        was = before.get(key)
        if was is None:
            changes.append({
                "item_code": code, "occurrence": now["occurrence"],
                "description": now["description"], "status": "added",
                "qty_was": None, "qty_now": now["qty"], "qty_changed": True,
                "unit_price_was": None, "unit_price_now": now["unit_price"],
                "unit_price_changed": True,
                "amount_was": None, "amount_now": now["amount"],
            })
            continue
        price_moved = _moved(was["unit_price"], now["unit_price"])
        qty_moved = _moved(was["qty"], now["qty"])
        price_changed += 1 if price_moved else 0
        qty_changed += 1 if qty_moved else 0
        if not price_moved and not qty_moved:
            continue
        changes.append({
            "item_code": code, "occurrence": now["occurrence"],
            "description": now["description"], "status": "changed",
            "qty_was": was["qty"], "qty_now": now["qty"], "qty_changed": qty_moved,
            "unit_price_was": was["unit_price"], "unit_price_now": now["unit_price"],
            "unit_price_changed": price_moved,
            "amount_was": was["amount"], "amount_now": now["amount"],
        })
    for key, was in before.items():
        if key in after:
            continue
        changes.append({
            "item_code": was["item_code"], "occurrence": was["occurrence"],
            "description": was["description"], "status": "removed",
            "qty_was": was["qty"], "qty_now": None, "qty_changed": True,
            "unit_price_was": was["unit_price"], "unit_price_now": None,
            "unit_price_changed": True,
            "amount_was": was["amount"], "amount_now": None,
        })

    return {
        "compared_to_id": str(previous.id),
        "compared_to_pi_number": previous.pi_number,
        "price_changed_lines": price_changed,
        "qty_changed_lines": qty_changed,
        "added_lines": sum(1 for c in changes if c["status"] == "added"),
        "removed_lines": sum(1 for c in changes if c["status"] == "removed"),
        "changes": changes,
    }


def mark_as_revision_of(db: Session, invoice_id: str, previous_id: str) -> dict:
    """Link a PI that was uploaded as new to the document it actually revises (AC-E11).

    The matching in the upload dialog is a proposal, and a wrong "New PI" has to be fixable
    without deleting and re-uploading - the pre-loading list carries no invoice number, so
    the mistake is an easy one to make and an expensive one to be stuck with.
    """
    invoice = get_or_404(db, invoice_id)
    if str(invoice_id) == str(previous_id):
        raise AppException(
            422, "A proforma invoice cannot be a revision of itself.", detail="previous_id"
        )
    previous = get_or_404(db, previous_id)
    if str(previous.supplier_id) != str(invoice.supplier_id):
        raise AppException(
            422,
            f"'{previous.pi_number}' belongs to a different supplier.",
            detail="previous_id",
        )
    if (previous.status or "current") == "superseded":
        raise AppException(
            409,
            f"'{previous.pi_number}' has already been superseded by a newer revision.",
            code="superseded",
        )
    if invoice.revision_of_id:
        raise AppException(
            409,
            f"'{invoice.pi_number}' is already a revision of another document.",
            code="already_a_revision",
        )
    # A -> B, then B -> A closes the loop: the superseded original comes back as current,
    # both documents answer "what is this container costing", and the chain walk terminates
    # only because it counts its own steps. Walk BACK from the proposed predecessor - if it
    # reaches this invoice, this invoice is its ancestor and the link would be a cycle.
    ancestor = previous
    guard = 0
    while ancestor is not None and guard < 50:
        if str(ancestor.id) == str(invoice.id):
            raise AppException(
                409,
                f"'{previous.pi_number}' is already a revision of '{invoice.pi_number}', "
                "so it cannot also be the document it revises.",
                code="revision_cycle",
            )
        if not ancestor.revision_of_id:
            break
        ancestor = (
            db.query(ProformaInvoice)
            .filter(ProformaInvoice.id == str(ancestor.revision_of_id))
            .first()
        )
        guard += 1

    invoice.revision_of_id = str(previous.id)
    invoice.revision_no = int(previous.revision_no or 1) + 1
    invoice.status = "current"
    previous.status = "superseded"
    db.flush()
    return serialize(db, invoice)


def preview(
    db: Session,
    data: bytes,
    *,
    supplier_id: str,
    currency: Optional[str] = None,
    source_ref: Optional[str] = None,
) -> dict:
    """What this file holds, and what it would create, before anything is written."""
    parsed = _parse(db, data)
    out = _summarise(
        db, parsed, supplier_id=supplier_id, source_ref=source_ref,
        requested_currency=currency,
        revision_candidates=_revision_candidates(db, parsed, supplier_id=supplier_id),
    )
    out["ok"] = parsed.ok
    out["missing_columns"] = parsed.missing_columns
    out["problems"] = [p.reason for p in parsed.problems][:50]
    return out


def validate(
    db: Session,
    data: bytes,
    *,
    supplier_id: str,
    currency: Optional[str] = None,
    source_ref: Optional[str] = None,
) -> dict:
    """The `{valid, errors, warnings, summary}` verdict a Test means everywhere here."""
    parsed = _parse(db, data)
    summary = _summarise(
        db, parsed, supplier_id=supplier_id, source_ref=source_ref,
        requested_currency=currency,
    )

    # NOT the row problems: `apply` refuses only an unreadable file or an unresolved
    # currency, so anything else counted as an error here would make Test say no to a file
    # Apply would happily take - and the operator has no way to tell which one is lying.
    problems: list[str] = []
    if parsed.missing_columns:
        problems.append(
            "The file does not name "
            + named(len(parsed.missing_columns), parsed.missing_columns,
                    one="the column", many="the columns")
        )
    elif not parsed.documents:
        problems.append("No proforma invoice was found in this file.")
    elif summary["priced_lines_without_currency"]:
        problems.append(_NO_CURRENCY)

    warnings: list[str] = [p.reason for p in parsed.problems][:50]
    if summary["unmatched_items"]:
        warnings.append(
            "No product matches "
            + named(summary["unmatched_items"], summary["unmatched_item_codes"],
                    one="the code", many="the codes")
            + ". Those lines are still loaded, with no product against them."
        )
    if parsed.unmapped_headers:
        warnings.append(
            "Ignored " + named(len(parsed.unmapped_headers), parsed.unmapped_headers,
                               one="the column", many="the columns")
        )
    disagreeing = [
        d["pi_number"]
        for d in summary["documents"]
        if d["stated_total"] is not None
        and abs(float(d["stated_total"]) - float(d["total"])) > _TOTAL_TOLERANCE
    ]
    if disagreeing:
        warnings.append(
            "The stated total does not match the sum of the lines on "
            + named(len(disagreeing), disagreeing, one="invoice", many="invoices")
        )
    mismatch = _supplier_mismatch_warning(summary.get("supplier_check"))
    if mismatch:
        warnings.append(mismatch)

    return envelope(ok=not problems, problems=problems, warnings=warnings, summary=summary)


def _plan_of_this_supplier(
    db: Session, loading_plan_id: Optional[str], *, supplier_id: str
) -> None:
    """Refuse a plan that belongs to somebody else (AC-F3).

    Checked BEFORE anything is written, like the revision targets: stamping one supplier's
    invoices onto another supplier's plan would make every read on that plan wrong - its
    holdings, its unknown codes and its own subtitle - and nothing on the record would say so.
    """
    if not loading_plan_id:
        return
    from app.models.scm import LoadingPlan

    try:
        uuid.UUID(str(loading_plan_id))
    except (ValueError, AttributeError, TypeError):
        raise AppException(422, "That loading plan does not exist.", detail="loading_plan_id")
    plan = db.query(LoadingPlan).filter(LoadingPlan.id == loading_plan_id).first()
    if plan is None:
        raise AppException(422, "That loading plan does not exist.", detail="loading_plan_id")
    if str(plan.supplier_id) != str(supplier_id):
        raise AppException(
            422,
            "That loading plan belongs to a different supplier.",
            detail="invoice_supplier_mismatch",
        )


def apply(
    db: Session,
    data: bytes,
    *,
    supplier_id: str,
    currency: Optional[str] = None,
    source_ref: Optional[str] = None,
    actor: Optional[str] = None,
    revision_of: Optional[dict] = None,
    file_as_new: Optional[list] = None,
    loading_plan_id: Optional[str] = None,
) -> dict:
    """Write one proforma invoice per document in the file. Idempotent by identity.

    Idempotent because the invoice NUMBER is the document's own or derived from the file
    rather than generated: the same file uploaded twice resolves to the same invoices and
    replaces their lines, which is AC-P1.4 and is also what stops a nervous second click
    doubling an invoice. Does not commit.

    `loading_plan_id` (S6, AC-F3) stamps EVERY invoice this call creates or revises with the
    plan it was uploaded into. That stamp is what lets the plan sum its own five blocks
    instead of reading whichever single invoice sorted first for the supplier - which, over
    five blocks sharing an invoice date and a transaction timestamp, was decided by the UUID.
    A superseded prior keeps the plan that took IT, so an older plan goes on reading its own.
    """
    _plan_of_this_supplier(db, loading_plan_id, supplier_id=supplier_id)
    parsed = _parse(db, data)
    if not parsed.ok:
        raise AppException(
            422,
            "This file could not be read as a proforma invoice.",
            detail=", ".join(parsed.missing_columns) or "no invoice was found in it",
        )

    resolved = _currencies(db, parsed, supplier_id=supplier_id, requested=currency)
    if any(d.priced_lines and not resolved.get(d.index, (None,))[0] for d in parsed.documents):
        raise AppException(422, _NO_CURRENCY, detail="currency")

    # Checked before anything is written: a five-block file naming one bad revision target
    # fails whole rather than writing four revisions and then refusing.
    targets = _revision_targets(db, revision_of, supplier_id=supplier_id)
    # Documents the operator explicitly UNTICKED the revision offer on. An untick is an
    # instruction - file this as a new document - and without it the same file's derived
    # number fell into the in-place replace below, so the upload reported "updated in
    # place" and produced no new row at all.
    filed_as_new = {str(i) for i in (file_as_new or [])}

    known = _products_by_code(
        db, {ln.item_code for d in parsed.documents for ln in d.lines}
    )
    # The supplier's own spelling, worked out and remembered (R16). Keyed like `known` -
    # upper-cased - so the line binding below reads one map and cannot prefer the looser
    # answer where the exact one exists.
    known = _with_supplier_codes(
        db, known, supplier_id=supplier_id,
        codes={ln.item_code for d in parsed.documents for ln in d.lines},
        actor=actor,
    )
    # Built here, from the catalogue lookup and the currencies this apply is about to use,
    # rather than re-read afterwards: the summary then describes exactly what was written.
    summary = _summarise(
        db, parsed, supplier_id=supplier_id, source_ref=source_ref,
        requested_currency=currency, known=known, resolved=resolved,
    )

    created = updated = 0
    results: list[dict] = []

    for doc in parsed.documents:
        number = pi_number_for(doc, source_ref=source_ref)
        code, source = resolved.get(doc.index, (None, "none"))
        prior = targets.get(str((revision_of or {}).get(str(doc.index)) or ""))

        if prior is not None:
            # A revision is always a NEW row: the prior one is what the supplier sent on the
            # day, it is what the diff is read against, and overwriting it would delete the
            # only evidence that anything changed (AC-E7).
            revision_no = int(prior.revision_no or 1) + 1
            invoice = ProformaInvoice(
                id=_uuid(),
                supplier_id=supplier_id,
                pi_number=_available_number(db, supplier_id, number, revision_no),
                revision_of_id=str(prior.id),
                revision_no=revision_no,
                status="current",
            )
            db.add(invoice)
            prior.status = "superseded"
            existed = False
        elif str(doc.index) in filed_as_new:
            # The STEM is the base, not the number the file derived. A derived number is
            # `PI-<file stem>-<block>`, so appending to it produced `<stem>-1-2` and then
            # `<stem>-1-3`; the ordinal after the stem is which document this is, and the
            # search skips the ones already taken. A STATED number is not an ordinal and is
            # never chopped - `202605-S0060` would become `202605`.
            base = number if doc.pi_number else re.sub(r"-\d+$", "", number)
            invoice = ProformaInvoice(
                id=_uuid(),
                supplier_id=supplier_id,
                # The next free number for this supplier: `-2`, `-3`. Not `-R2` - it is a
                # second document rather than a second version of one.
                pi_number=_available_number(
                    db, supplier_id, base, 2, marker="", always_suffix=True
                ),
            )
            db.add(invoice)
            existed = False
        else:
            invoice = (
                db.query(ProformaInvoice)
                .filter(
                    ProformaInvoice.supplier_id == supplier_id,
                    ProformaInvoice.pi_number == number,
                )
                .first()
            )
            existed = invoice is not None
            if invoice is None:
                invoice = ProformaInvoice(
                    id=_uuid(), supplier_id=supplier_id, pi_number=number
                )
                db.add(invoice)
            else:
                _replaceable_or_409(db, invoice)

        invoice.invoice_date = doc.invoice_date
        # A PRICED document without a currency never reaches here - it was refused above. An
        # unpriced one has nothing to denominate and stays NULL. Never a house default (AC-P3.3).
        invoice.currency = code or invoice.currency
        invoice.container_ref = doc.container_no
        invoice.bl_ref = doc.bl_no
        # The document's own total when it states one - it is the number on the paper the
        # supplier sent, and the line sum is what we make of it. Where it states none, the
        # sum is the honest stand-in.
        invoice.total_amount = doc.stated_total if doc.stated_total is not None else doc.line_total
        invoice.line_count = len(doc.lines)
        invoice.source_ref = source_ref
        invoice.block_index = doc.index
        invoice.uploaded_by = actor
        # Only ever SET, never cleared: a standalone re-upload of a file that a plan is
        # reading must not quietly unbind it from that plan.
        if loading_plan_id:
            invoice.loading_plan_id = loading_plan_id
        db.flush()

        # Replace rather than merge: the file is the document of record, and a line the new
        # version no longer carries is a line the supplier withdrew.
        db.query(ProformaInvoiceLine).filter(
            ProformaInvoiceLine.invoice_id == invoice.id
        ).delete(synchronize_session=False)
        db.flush()

        unmatched: list[str] = []
        for n, ln in enumerate(doc.lines, start=1):
            product = known.get(ln.item_code.upper())
            if product is None:
                unmatched.append(ln.item_code)
            db.add(
                ProformaInvoiceLine(
                    id=_uuid(),
                    invoice_id=invoice.id,
                    line_no=n,
                    row_number=ln.row_number,
                    item_code=ln.item_code[:100],
                    description=ln.description,
                    qty=ln.qty,
                    uom=ln.uom[:20] if ln.uom else None,
                    unit_price=ln.unit_price,
                    amount=ln.amount,
                    po_ref=ln.po_ref[:100] if ln.po_ref else None,
                    remark=ln.remark,
                    cartons=ln.cartons,
                    cbm_per_unit=ln.cbm_per_unit,
                    cbm_total=ln.cbm_total,
                    net_weight=ln.net_weight,
                    gross_weight=ln.gross_weight,
                    # What it is made of and how it is boxed. The container workbook derives
                    # its carton count and its volume from these, and the conversion carries
                    # them onto the packing-list line so nobody re-types the supplier's own
                    # measurements (AC-F3.5).
                    material=ln.material[:255] if ln.material else None,
                    pcs_per_carton=ln.pcs_per_carton,
                    carton_length_cm=ln.carton_length_cm,
                    carton_width_cm=ln.carton_width_cm,
                    carton_height_cm=ln.carton_height_cm,
                    # The supplier's own figures, frozen here and never written again: `qty`
                    # and `unit_price` above are ours to trim to fit the container, and the
                    # whole journey rests on the two never being confused (AC-E2).
                    supplier_qty=ln.qty,
                    supplier_unit_price=ln.unit_price,
                    product_id=product["id"] if product else None,
                    product_set_id=product.get("product_set_id") if product else None,
                )
            )
        db.flush()

        created += 0 if existed else 1
        updated += 1 if existed else 0
        results.append(
            {
                "index": doc.index,
                "invoice_id": str(invoice.id),
                "pi_number": invoice.pi_number,
                "invoice_date": invoice.invoice_date.isoformat() if invoice.invoice_date else None,
                "currency": invoice.currency or None,
                "currency_source": source,
                "lines": len(doc.lines),
                "revision_no": int(invoice.revision_no or 1),
                "revision_of_id": str(invoice.revision_of_id) if invoice.revision_of_id else None,
                "total_amount": _f(invoice.total_amount),
                "unmatched_items": sorted(set(unmatched))[:50],
                "created": not existed,
            }
        )

    summary.update(
        {
            "documents_created": created,
            "documents_updated": updated,
            "results": results,
        }
    )
    return {
        "documents_created": created,
        "documents_updated": updated,
        "results": results,
        "summary": summary,
    }


def _draft_shipment_number(db: Session) -> str:
    """The next `PL-YYMM-NNN` from the numbering rule, or a refusal naming what is missing.

    Scoped to the writing company, the same one the row itself will be stamped with: a running
    number printed on a document two companies share is not a series (migration 327). An
    ambiguous scope resolves to `None`, which draws from whatever rule exists rather than
    refusing a conversion over a scope question.

    A company with no series of its own gets one HERE, in the caller's transaction, from the
    same definition migration 440 seeds: 440 ran once, over the companies that existed at that
    instant, so a company created afterwards had no rule and its first convert was refused with
    `numbering_rule_missing` - a 500 on a screen where the operator had done nothing wrong. The
    refusal is kept for the case it actually describes: a rule that EXISTS and still cannot
    number (disabled in Setup), which is a decision somebody made and not an absence to fill in.
    """
    company_id = resolve_write_company_id(get_company_scope(db), ambiguous=None)

    def _next() -> Optional[str]:
        return NumberingService(db).get_next_number(
            _DRAFT_NUMBER_DOC_TYPE, _date.today(), company_id=company_id, commit_rule=False
        )

    number = _next()
    if not number:
        seed_inbound_shipment_draft_rule(db, company_id=company_id)
        number = _next()
    if not number:
        raise AppException(
            500,
            "No numbering rule is configured for a packing list, so one cannot be numbered. "
            "Add the 'inbound_shipment_draft' rule under System > Numbering.",
            code="numbering_rule_missing",
        )
    return number


def _product_base_uoms(db: Session, product_ids: set[str]) -> dict[str, Optional[str]]:
    if not product_ids:
        return {}
    from app.models.product import Product

    rows = (
        db.query(Product.id, Product.base_uom_id)
        .filter(Product.id.in_(product_ids))
        .all()
    )
    return {str(pid): (str(uom_id) if uom_id else None) for pid, uom_id in rows}


def _placements(db: Session, invoice_ids: list[str]) -> dict[str, dict]:
    """Per invoice, how much of it has reached a packing list and which ones (AC-F6).

    Only rows that actually became a shipment line count. A SKIP row records why a line went
    nowhere, and counting it as placed would report goods on a container that never held
    them.
    """
    ids = [i for i in set(invoice_ids) if i]
    if not ids:
        return {}
    rows = (
        db.query(
            ProformaInvoiceShipmentLink.proforma_invoice_id,
            ProformaInvoiceShipmentLink.inbound_shipment_id,
            InboundShipment.shipment_number,
            InboundShipment.shipment_status,
            func.sum(func.coalesce(ProformaInvoiceShipmentLink.qty, 0)),
            func.count(func.distinct(ProformaInvoiceShipmentLink.proforma_invoice_line_id)),
        )
        .join(
            InboundShipment,
            InboundShipment.id == ProformaInvoiceShipmentLink.inbound_shipment_id,
        )
        .filter(
            ProformaInvoiceShipmentLink.proforma_invoice_id.in_(ids),
            ProformaInvoiceShipmentLink.inbound_shipment_line_id.isnot(None),
        )
        .group_by(
            ProformaInvoiceShipmentLink.proforma_invoice_id,
            ProformaInvoiceShipmentLink.inbound_shipment_id,
            InboundShipment.shipment_number,
            InboundShipment.shipment_status,
        )
        .all()
    )
    out: dict[str, dict] = {}
    for invoice_id, shipment_id, number, status, qty, lines in rows:
        entry = out.setdefault(str(invoice_id), {"placed_qty": 0.0, "packing_lists": []})
        entry["placed_qty"] += float(qty or 0)
        entry["packing_lists"].append(
            {
                "shipment_id": str(shipment_id),
                "shipment_number": number,
                "shipment_status": status,
                "qty": float(qty or 0),
                "lines": int(lines or 0),
            }
        )
    for entry in out.values():
        entry["packing_lists"].sort(key=lambda p: (p["shipment_number"] or "", p["shipment_id"]))
    return out


def _line_placements(db: Session, invoice_id: str) -> dict[str, list[dict]]:
    """Per LINE of one invoice, where its goods went and how much of them."""
    rows = (
        db.query(
            ProformaInvoiceShipmentLink.proforma_invoice_line_id,
            ProformaInvoiceShipmentLink.inbound_shipment_id,
            InboundShipment.shipment_number,
            func.sum(func.coalesce(ProformaInvoiceShipmentLink.qty, 0)),
        )
        .join(
            InboundShipment,
            InboundShipment.id == ProformaInvoiceShipmentLink.inbound_shipment_id,
        )
        .filter(
            ProformaInvoiceShipmentLink.proforma_invoice_id == str(invoice_id),
            ProformaInvoiceShipmentLink.inbound_shipment_line_id.isnot(None),
        )
        .group_by(
            ProformaInvoiceShipmentLink.proforma_invoice_line_id,
            ProformaInvoiceShipmentLink.inbound_shipment_id,
            InboundShipment.shipment_number,
        )
        .all()
    )
    out: dict[str, list[dict]] = {}
    for line_id, shipment_id, number, qty in rows:
        out.setdefault(str(line_id), []).append(
            {
                "shipment_id": str(shipment_id),
                "shipment_number": number,
                "qty": float(qty or 0),
            }
        )
    return out


def _set_members(db: Session, set_ids) -> dict[str, list]:
    """`{set_id: its members, in the author's own order}` (R21).

    Ordered by `sort_order` then product code, ending on the member id so the order is
    TOTAL: the FIRST member is where a set line's price and its volume land on the
    container, and a rule that could pick a different member on a second run would put the
    same invoice's cost on a different line each time it was converted.
    """
    wanted = [str(i) for i in dict.fromkeys(set_ids) if i]
    if not wanted:
        return {}
    rows = (
        db.query(ProductSetMember)
        .options(joinedload(ProductSetMember.product))
        .filter(ProductSetMember.product_set_id.in_(wanted))
        .all()
    )
    out: dict[str, list] = {}
    for member in sorted(
        rows,
        key=lambda m: (
            m.sort_order or 0,
            (getattr(m.product, "product_code", None) or ""),
            str(m.id),
        ),
    ):
        out.setdefault(str(member.product_set_id), []).append(member)
    return out


def _placeable(line: ProformaInvoiceLine) -> float:
    """What of this line COULD go on a container.

    A line bound to neither a product nor a SET cannot: `inbound_shipment_lines.product_id`
    is NOT NULL, so there is nowhere to write it. Counting it as outstanding would leave the
    invoice reading Split for ever, with a remainder nobody can ever place.

    A set line CAN (R21): it explodes into one container line per member, and its placed
    quantity is counted in the invoice's own units - sets - so what is placeable is the
    line's own quantity, exactly as it is for a product.
    """
    if line.product_id is None and line.product_set_id is None:
        return 0.0
    qty = float(line.qty or 0)
    return qty if qty > 0 else 0.0


def _quantities(db: Session, invoice_ids: list[str]) -> dict[str, dict]:
    """Per invoice: what the document totals, what could be placed, and what still can be."""
    lines = (
        db.query(ProformaInvoiceLine)
        .filter(ProformaInvoiceLine.invoice_id.in_([str(i) for i in invoice_ids]))
        .all()
    )
    placements = _placements(db, [str(i) for i in invoice_ids])
    out: dict[str, dict] = {
        str(i): {"total_qty": 0.0, "placeable_qty": 0.0} for i in invoice_ids
    }
    for line in lines:
        entry = out.setdefault(str(line.invoice_id), {"total_qty": 0.0, "placeable_qty": 0.0})
        entry["total_qty"] += float(line.qty or 0)
        entry["placeable_qty"] += _placeable(line)
    for invoice_id, entry in out.items():
        placed = placements.get(invoice_id, {}).get("placed_qty", 0.0)
        entry["placed_qty"] = placed
        entry["remaining_qty"] = max(entry["placeable_qty"] - placed, 0.0)
        entry["packing_lists"] = placements.get(invoice_id, {}).get("packing_lists", [])
        entry["placement"] = (
            "not_converted" if placed <= 0
            else ("converted" if entry["remaining_qty"] <= 0 else "split")
        )
    return out


def _shipment_actor_name(db: Session, shipment_id: str) -> str:
    """Who created this container, by NAME.

    `inbound_shipments.created_by` holds a user id, and the origin line on the packing list
    printed it: "Created from PI-x by 9993276c-a55e-4a54-9686-7138f5fa1306". A UUID on
    screen is forbidden here and tells its reader nothing anyway.

    "System" covers both silences - a container nobody is recorded against (an import, an
    n8n write) and an actor whose user row has since gone. Neither can be resolved to a
    person, and the id is not an acceptable stand-in for one.
    """
    actor = (
        db.query(InboundShipment.created_by)
        .filter(InboundShipment.id == str(shipment_id))
        .scalar()
    )
    if not actor:
        return "System"
    row = db.execute(
        text(
            "SELECT COALESCE(NULLIF(TRIM(name), ''), email) AS label "
            "FROM users WHERE id = :id"
        ),
        {"id": str(actor)},
    ).mappings().first()
    return (row["label"] if row and row["label"] else None) or "System"


def source_proforma_invoices(db: Session, shipment_id: str) -> dict:
    """Which proforma invoices this container's lines were charged on (AC-F9).

    One read for the four places that show it - the Details card, the Lines column, the
    Timeline entry and the Documents list - because they are four readings of the same link
    rows, and four fetches would be four chances for them to disagree.

    Empty rather than 404 on a container that came off a real packing-list upload: having no
    proforma invoice behind it is a normal state, not a missing record.
    """
    if not _is_uuid(shipment_id):
        raise AppException(404, "Packing list not found.", detail="shipment_id")

    rows = (
        db.query(ProformaInvoiceShipmentLink, ProformaInvoice)
        .join(ProformaInvoice, ProformaInvoice.id == ProformaInvoiceShipmentLink.proforma_invoice_id)
        .filter(
            ProformaInvoiceShipmentLink.inbound_shipment_id == str(shipment_id),
            ProformaInvoiceShipmentLink.inbound_shipment_line_id.isnot(None),
        )
        .all()
    )
    created_by = _shipment_actor_name(db, shipment_id)
    if not rows:
        return {"invoices": [], "by_shipment_line": {}, "created_by": created_by}

    prices: dict[str, Optional[float]] = {}
    line_ids = {str(link.proforma_invoice_line_id) for link, _ in rows}
    for pi_line in (
        db.query(ProformaInvoiceLine)
        .filter(ProformaInvoiceLine.id.in_(list(line_ids)))
        .all()
    ):
        prices[str(pi_line.id)] = _f(pi_line.unit_price)

    totals = _quantities(db, list({str(inv.id) for _, inv in rows}))
    line_counts: dict[str, int] = {}
    for invoice_id, count in (
        db.query(ProformaInvoiceLine.invoice_id, func.count(1))
        .filter(ProformaInvoiceLine.invoice_id.in_(list({str(inv.id) for _, inv in rows})))
        .group_by(ProformaInvoiceLine.invoice_id)
        .all()
    ):
        line_counts[str(invoice_id)] = int(count)

    suppliers = _supplier_labels(db, [str(inv.supplier_id) for _, inv in rows])
    invoices: dict[str, dict] = {}
    by_line: dict[str, list[dict]] = {}
    for link, invoice in rows:
        qty = float(link.qty or 0)
        entry = invoices.get(str(invoice.id))
        if entry is None:
            entry = {
                "id": str(invoice.id),
                "pi_number": invoice.pi_number,
                "supplier_id": str(invoice.supplier_id) if invoice.supplier_id else None,
                "supplier_name": suppliers.get(str(invoice.supplier_id), (None, None))[1],
                "invoice_date": (
                    invoice.invoice_date.isoformat() if invoice.invoice_date else None
                ),
                "revision_no": int(invoice.revision_no or 1),
                # How many revisions the chain holds, so the card can read "Revision 2 of 2"
                # (AC-F9): which version the container was loaded from is the whole point of
                # keeping the chain, and the number alone does not say it.
                "revision_count": len(_chain(db, invoice)),
                "status": invoice.status or "current",
                "source_ref": invoice.source_ref,
                "currency": invoice.currency or None,
                "lines": 0,
                "total_lines": line_counts.get(str(invoice.id), 0),
                "qty": 0.0,
                "total_qty": totals.get(str(invoice.id), {}).get("total_qty", 0.0),
                "amount": None,
            }
            invoices[str(invoice.id)] = entry
        # DISTINCT PI lines, not link rows: a set line becomes one container line per
        # member (R21), and counting the rows would report one invoice line as two.
        entry.setdefault("_line_ids", set()).add(str(link.proforma_invoice_line_id))
        entry["lines"] = len(entry["_line_ids"])
        entry["qty"] += qty
        price = prices.get(str(link.proforma_invoice_line_id))
        if price is not None:
            entry["amount"] = round((entry["amount"] or 0.0) + price * qty, 2)

        if link.inbound_shipment_line_id:
            by_line.setdefault(str(link.inbound_shipment_line_id), []).append(
                {
                    "proforma_invoice_id": str(invoice.id),
                    "pi_number": invoice.pi_number,
                    # NULL, never 0, on the sibling line of an exploded SET (R21): the
                    # quantity is the PI line's own and is recorded once, on the first
                    # member, so a 0 beside the cistern would read as goods that never came.
                    "qty": qty if link.qty is not None else None,
                }
            )

    for entry in invoices.values():
        entry.pop("_line_ids", None)
    return {
        "invoices": sorted(invoices.values(), key=lambda i: i["pi_number"]),
        "by_shipment_line": by_line,
        "created_by": created_by,
    }


def _over_capacity(
    db: Session,
    invoices: list[ProformaInvoice],
    placing: dict[str, float],
) -> list[str]:
    """One sentence per invoice that will not fit, naming both figures (AC-E5).

    Judged on what is ACTUALLY being loaded, not on the whole document: `placing` is the
    volume of the quantities this convert is placing, so half of a 69 cbm invoice is 35 cbm
    and goes in the box - which is the split Q9 exists for, and was unreachable while the
    gate read the whole invoice however little of it was moving.

    Judged against an EMPTY box, because every convert opens a new one (Q6): there is
    never anything already in it to load on top of.

    An invoice whose lines state NO volume is not over capacity - it is unmeasured, and
    refusing it would break the Kailu shape that has converted since G3b (AC-H3). Silence
    about a number nobody has is the honest answer; a guess is not.
    """
    sizes_by_id, default_size = _container_sizes(db)
    out: list[str] = []
    for invoice in invoices:
        placed_cbm = placing.get(str(invoice.id))
        if placed_cbm is None:
            continue
        fit = _fit(invoice, placed_cbm, sizes_by_id, default_size)
        if fit["over_by_cbm"]:
            out.append(
                f"{invoice.pi_number} is {_num(placed_cbm)} cbm and the "
                f"{fit['container_size_code'] or 'container'} holds "
                f"{_num(fit['container_cbm'])} - over by {_num(fit['over_by_cbm'])} cbm."
            )
    return out


def _dec(value: Any) -> Decimal:
    """A quantity as a Decimal. `proforma_invoice_line.qty` is Numeric, and float arithmetic
    on it is how 12.5 placed twice stops adding up to 25."""
    if value is None:
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value))


def _whole(value: Decimal) -> int:
    """A Decimal quantity as the whole pieces an integer column can hold, half up."""
    return int(_dec(value).to_integral_value(rounding=ROUND_HALF_UP))


def _placing_volume(line: ProformaInvoiceLine, qty: float) -> Optional[float]:
    """How much room `qty` of this line takes. None when the document measured nothing."""
    if line.cbm_per_unit is not None:
        return float(line.cbm_per_unit) * qty
    if line.cbm_total is not None and line.qty:
        return float(line.cbm_total) * (qty / float(line.qty))
    return None


def _record_over_capacity(
    db: Session,
    shipment_id: str,
    over: list[str],
    override_reason: Optional[str],
    created_by: Optional[str],
) -> None:
    """Why a container was loaded past its planned volume, on the container's Timeline.

    It used to be appended to `notes`, which is the operator's OWN field: the next person to
    edit the container found a sentence there that nobody had typed, and a note they wrote
    themselves sat beside it with nothing to tell the two apart. The decision belongs to the
    record of what happened to this container, which is its audit trail (R17).
    """
    reason = (override_reason or "").strip()
    if not over or not reason:
        return
    log_audit(
        db,
        "inbound_shipments",
        str(shipment_id),
        "UPDATE",
        user_id=created_by if _is_uuid(created_by) else None,
        description=f"Converted over capacity: {' '.join(over)} Reason: {reason}",
    )


def convert_to_draft_shipment(
    db: Session,
    invoice_ids: list[str],
    *,
    created_by: Optional[str] = None,
    override_capacity: bool = False,
    override_reason: Optional[str] = None,
    line_quantities: Optional[dict] = None,
) -> dict:
    """One or more proforma invoices become ONE NEW draft inbound shipment (the packing-list
    amendment, `PLAN-scm-proforma-to-spo.md`): "pick one or more PIs -> the system creates a
    DRAFT inbound shipment pre-filled with their lines". Several PIs are allowed to come from
    DIFFERENT suppliers on purpose - a container is routinely one factory's PI joining three
    others' inside the same box (the real KAILU + FSCU8103365 documents this was built
    against) - so every shipment line carries its OWN `supplier_id`, never the header's.

    Not built here: the real packing list replacing/reconciling this draft (that is the
    EXISTING upload path, `packing_list_service.apply`, unchanged by this function - a draft's
    `shipment_number` is its own series so it never collides with what a real upload derives,
    and nothing here teaches that upload to find this row; reconciling onto the exact draft
    is follow-up work) and the "Create SPO" action off the shipment (the next slice).

    Idempotent by refusal, not by silently updating: a PI that has ANY conversion outcome
    recorded already (`ProformaInvoiceShipmentLink`, matched or skipped) is refused with a 409
    naming the shipment it already went to - converting it again would double what counts as
    incoming stock until somebody notices two drafts.

    A line with no catalogue product match, or a non-positive quantity, is reported and
    SKIPPED rather than dropped silently - `inbound_shipment_lines.product_id` is NOT NULL,
    so there is nowhere to write it, and a skip is recorded (`unmatched_reason`) so the PI
    detail page can still say what happened to it. Refuses only when NOTHING in the whole
    selection can convert; a partial match still creates the shipment.
    """
    ids = list(dict.fromkeys(str(i).strip() for i in (invoice_ids or []) if str(i).strip()))
    if not ids:
        raise AppException(
            422, "Select at least one proforma invoice to convert.", detail="proforma_invoice_ids"
        )

    invoices = db.query(ProformaInvoice).filter(ProformaInvoice.id.in_(ids)).all()
    found_by_id = {str(inv.id): inv for inv in invoices}
    missing = [i for i in ids if i not in found_by_id]
    if missing:
        raise AppException(
            404,
            "One or more proforma invoices could not be found.",
            detail=", ".join(missing),
        )

    # Take the row lock BEFORE reading what is placed. Migration 429 dropped the unique
    # index that made a double convert impossible, and what replaced it - "what is placed
    # against what can be placed" - is a READ: two overlapping converts both saw an unplaced
    # invoice and both placed it, doubling what the office is asked to key in. The second
    # one now waits here and then sees the first one's work.
    db.query(ProformaInvoice.id).filter(ProformaInvoice.id.in_(ids)).with_for_update().all()

    # What is left to place. Since Q9 an invoice may go into two containers, so "already
    # converted" is no longer "has a link row" - it is "has nothing left", which is
    # arithmetic rather than existence. An invoice with nothing left is SKIPPED and named
    # (AC-F7); only a selection where every invoice is finished is refused outright.
    quantities = _quantities(db, ids)
    skipped_invoices: list[dict] = []
    live_ids: list[str] = []
    for invoice_id in ids:
        entry = quantities.get(invoice_id, {})
        if entry.get("remaining_qty", 0) > 0:
            live_ids.append(invoice_id)
            continue
        if entry.get("placed_qty", 0) <= 0:
            # Nothing was ever placed and nothing ever can be - every line names a code
            # this catalogue does not hold. "Already converted" is the wrong sentence: it
            # sends the reader looking for a container that does not exist. Let it through
            # to the grouping below, which refuses it by NAME.
            live_ids.append(invoice_id)
            continue
        names = ", ".join(
            p["shipment_number"] or "a draft" for p in entry.get("packing_lists", [])
        )
        skipped_invoices.append(
            {
                "id": invoice_id,
                "pi_number": found_by_id[invoice_id].pi_number,
                "reason": f"Already in {names}." if names else "Nothing left to place.",
            }
        )
    if not live_ids:
        named_parts = [f"{i['pi_number']} -> {i['reason'].rstrip('.')}" for i in skipped_invoices]
        raise AppException(
            409,
            "Already converted: " + "; ".join(sorted(named_parts)) + ".",
            code="already_converted",
        )
    ids = live_ids
    invoices = [found_by_id[i] for i in ids]

    # A superseded revision is what the supplier sent on a day that has passed; converting
    # it would load the container from a document already replaced (AC-E10).
    superseded = [inv.pi_number for inv in invoices if (inv.status or "current") == "superseded"]
    if superseded:
        raise AppException(
            409,
            "Superseded by a newer revision: " + ", ".join(sorted(superseded)) + ".",
            code="superseded",
        )

    lines: list[ProformaInvoiceLine] = (
        db.query(ProformaInvoiceLine)
        .filter(ProformaInvoiceLine.invoice_id.in_(ids))
        .order_by(ProformaInvoiceLine.invoice_id, ProformaInvoiceLine.line_no)
        .all()
    )

    # Group by (product, supplier) - the same grain a real packing list writes on, and the
    # one that lets two PIs from the same factory naming the same model become one shipment
    # line rather than two. `source_lines` is what makes the provenance link total: every PI
    # line that fed a group links to whatever ONE shipment line the group becomes.
    groups: dict[tuple[str, Optional[str]], dict[str, Any]] = {}
    skipped: list[tuple[ProformaInvoiceLine, str]] = []
    product_ids: set[str] = set()
    #: Per invoice, the volume of the quantities THIS convert is placing - what the capacity
    #: gate judges, rather than the whole document however little of it is moving.
    placing: dict[str, float] = {}

    # How much of each line is still to place, and how much of it this convert is taking.
    # The default is the whole remainder, because "all of what is left" is the normal case
    # and a split is the exception somebody types (AC-F10).
    placed_per_line = {
        line_id: sum(p["qty"] for p in entries)
        for invoice_id in ids
        for line_id, entries in _line_placements(db, invoice_id).items()
    }
    requested = {str(k): _dec(v) for k, v in (line_quantities or {}).items()}

    # A line bound to a SET explodes into its members (R21): stock lives on the members and
    # `inbound_shipment_lines.product_id` is NOT NULL, so a set has nowhere to be written on
    # a container. Read once for the whole convert, in the author's own order - the first
    # member is where the set's price and its volume land, so which one it is has to be
    # settled and repeatable.
    members_by_set = _set_members(db, {ln.product_set_id for ln in lines if ln.product_set_id})

    for ln in lines:
        invoice = found_by_id[str(ln.invoice_id)]
        if ln.product_id is None and ln.product_set_id is None:
            skipped.append((ln, "No catalogue product matches this line's item code."))
            continue
        if ln.product_id is None and not members_by_set.get(str(ln.product_set_id)):
            skipped.append((ln, "This set has no members to ship."))
            continue
        remaining = _dec(ln.qty) - _dec(placed_per_line.get(str(ln.id), 0))
        if remaining <= 0:
            # Already on a container. Not a skip WITH A REASON - there is nothing wrong
            # with it, it is simply finished, and a second skip row would claim otherwise.
            continue
        asked = requested.get(str(ln.id), remaining)
        if asked < 0 or asked > remaining:
            raise AppException(
                422,
                f"{ln.item_code} has {_num(remaining)} left to place, not {_num(asked)}.",
                detail="line_quantities",
            )
        # DECIMAL, never int(): `proforma_invoice_line.qty` is Numeric, and truncating
        # 12.5 to 12 left half a piece outstanding on a line nothing would ever place
        # again - so the invoice read Split for ever and could not be converted.
        qty = asked
        if qty <= 0:
            skipped.append((ln, "This line has no positive quantity to ship."))
            continue
        # `(which product, how many of it)`. One pair for an ordinary line; one per member
        # for a set line, at `qty x member.quantity`. The FIRST pair is the PRIMARY: the
        # set's price and its volume land there and nowhere else, because a set priced once
        # must not become N priced lines and a set that takes one carton must not become N
        # cartons - a container reading several times its real size is refused by the
        # capacity gate for goods that fit.
        if ln.product_id is not None:
            targets = [(str(ln.product_id), qty)]
        else:
            targets = [
                (str(m.product_id), qty * _dec(m.quantity if m.quantity is not None else 1))
                for m in members_by_set[str(ln.product_set_id)]
            ]

        for target_index, (target_product_id, target_qty) in enumerate(targets):
            primary = target_index == 0
            product_ids.add(target_product_id)
            key = (
                target_product_id,
                str(invoice.supplier_id) if invoice.supplier_id else None,
            )
            group = groups.get(key)
            if group is None:
                group = {
                    "product_id": target_product_id,
                    "supplier_id": str(invoice.supplier_id) if invoice.supplier_id else None,
                    "quantity_shipped": Decimal("0"),
                    "unit_cost": None,
                    "currency": None,
                    # Volume and cartons ADD across the lines a group merges, and stay None
                    # until something states one: 0 cbm would be summed on the packing list
                    # as a container that takes no room (AC-F1).
                    "cbm": None,
                    "cartons": None,
                    # How the goods are made and boxed. These do NOT add across merged
                    # lines - a carton is the size it is however many lines name it - so the
                    # first line that states one wins and the rest are ignored, the same
                    # rule the unit price follows two blocks below (AC-F3.5).
                    "measurements": {},
                    "remarks": [],
                    "source_lines": [],
                }
                groups[key] = group
            prev_qty = group["quantity_shipped"]
            group["quantity_shipped"] = prev_qty + target_qty
            # A part-placed line brings its SHARE of the volume and the cartons, not the
            # whole line's: half a line in this container takes half the room in it.
            share = (qty / _dec(ln.qty)) if ln.qty else Decimal("1")
            if primary and ln.cbm_per_unit is not None:
                group["cbm"] = (group["cbm"] or Decimal("0")) + _dec(ln.cbm_per_unit) * qty
            elif primary and ln.cbm_total is not None:
                group["cbm"] = (group["cbm"] or Decimal("0")) + _dec(ln.cbm_total) * share
            if primary and ln.cartons is not None:
                group["cartons"] = (group["cartons"] or 0) + int(
                    (_dec(ln.cartons) * share).to_integral_value(rounding=ROUND_HALF_UP)
                )
            # The supplier's own measurements, carried across so the container workbook can
            # be printed without anyone re-typing them. `net_weight` / `gross_weight` are
            # PER CARTON on both documents, which is why they land on the per-carton
            # columns. On the primary alone, for the same reason the carton count is: they
            # describe the box the whole set travels in.
            if primary:
                for source, target in (
                    ("material", "material"),
                    ("pcs_per_carton", "pcs_per_carton"),
                    ("carton_length_cm", "carton_length_cm"),
                    ("carton_width_cm", "carton_width_cm"),
                    ("carton_height_cm", "carton_height_cm"),
                    ("net_weight", "net_weight_per_carton"),
                    ("gross_weight", "gross_weight_per_carton"),
                ):
                    value = getattr(ln, source, None)
                    if value is not None and group["measurements"].get(target) is None:
                        group["measurements"][target] = value
            group["placed"] = group.get("placed", {})
            if primary:
                # HOW MUCH of the PI line came here, in the PI line's OWN units, recorded
                # ONCE per line per container. A set line placed as two member lines is
                # still ten sets placed, and a quantity on each of them would sum to twenty
                # and finish the invoice twice over. The sibling's link row therefore
                # carries no quantity at all, which every reader already coalesces to zero.
                group["placed"][str(ln.id)] = group["placed"].get(str(ln.id), 0) + qty
                placing_cbm = _placing_volume(ln, float(qty))
                if placing_cbm is not None:
                    placing[str(ln.invoice_id)] = (
                        placing.get(str(ln.invoice_id), 0.0) + placing_cbm
                    )
            if primary and ln.unit_price is not None:
                # The price is the SET's, so it lands on the first member and nowhere else -
                # which is Sorento's own catalogue convention, where the whole set's list
                # price parks on the pedestal and the cistern reads 0.00.
                if group["unit_cost"] is None:
                    group["unit_cost"] = ln.unit_price
                    group["currency"] = invoice.currency
                elif (group["currency"] or None) == (invoice.currency or None):
                    total_qty = Decimal(str(prev_qty)) + Decimal(str(target_qty))
                    if total_qty > 0:
                        group["unit_cost"] = (
                            Decimal(str(group["unit_cost"])) * Decimal(str(prev_qty))
                            + Decimal(str(ln.unit_price)) * Decimal(str(target_qty))
                        ) / total_qty
                # A currency mismatch across the merged lines is not arithmetic - the first
                # price stands and the disagreement is visible on the PI lines themselves.
            if primary and ln.remark:
                group["remarks"].append(ln.remark)
            # EVERY member line links back to the PI line it came from, so the provenance is
            # total: a cistern on a container has to be able to say which invoice charged it.
            group["source_lines"].append(ln)

    if not groups:
        raise AppException(
            422,
            "None of the selected invoices' lines match a product we hold, so there is "
            "nothing to draft a shipment from.",
            detail="unmatched",
        )

    # Checked HERE, after the placed quantities are known and before the first write: the
    # question is whether what is being loaded fits in the box it is going into (AC-E5).
    over = _over_capacity(db, invoices, placing)
    if over and not override_capacity:
        raise AppException(
            409,
            " ".join(over) + " Convert anyway to load it regardless.",
            code="over_capacity",
        )
    if over and override_capacity and not (override_reason or "").strip():
        raise AppException(
            422,
            "Say why this container is being loaded over its planned volume.",
            detail="override_reason",
        )

    uoms = _product_base_uoms(db, product_ids)

    invoice_dates = [inv.invoice_date for inv in invoices if inv.invoice_date]

    # A NEW packing list, every time (Q6). "Add to an existing draft" is gone: a convert
    # that could land in somebody else's box needed the box picking, and the pick was the
    # dialog this screen no longer has.
    shipment = InboundShipment(
        id=_uuid(),
        shipment_number=_draft_shipment_number(db),
        shipment_date=min(invoice_dates) if invoice_dates else _date.today(),
        shipment_status=_DRAFT_SHIPMENT_STATUS,
        created_by=created_by,
    )
    db.add(shipment)
    db.flush()
    # `notes` is left alone either way (R17): which invoices a container was drafted from is
    # the Proforma invoices tab's answer, and it is a table's worth rather than a sentence.
    _record_over_capacity(db, shipment.id, over, override_reason, created_by)

    # The box is new, so nothing is on it yet: two invoices naming the same model were
    # already merged into ONE group above, which is what the unique index on
    # (shipment, product, supplier) needs.
    shipment_lines_by_key: dict[tuple[str, Optional[str]], InboundShipmentLine] = {}
    for key, group in groups.items():
        line = InboundShipmentLine(
            id=_uuid(),
            shipment_id=shipment.id,
            product_id=group["product_id"],
            supplier_id=group["supplier_id"],
            # `inbound_shipment_lines.quantity_shipped` is an INTEGER column: a
            # container ships whole pieces. The fraction is not lost - it is on the
            # link, which is what the placement arithmetic reads.
            quantity_shipped=_whole(group["quantity_shipped"]),
            uom_id=uoms.get(group["product_id"]),
            unit_cost=group["unit_cost"],
            currency=group["currency"],
            cbm=group["cbm"],
            # `cartons_count` is NOT NULL with a default of 1, so an unstated carton
            # count keeps that default rather than being written as 0 - which would read
            # as a line that shipped in no box at all.
            **({"cartons_count": group["cartons"]} if group["cartons"] is not None else {}),
            **group["measurements"],
            remarks="; ".join(group["remarks"]) if group["remarks"] else None,
        )
        db.add(line)
        shipment_lines_by_key[key] = line
    db.flush()

    # Header supplier: whatever the written lines agree on, else none - the same rule
    # `InboundShipmentService._derive_header_supplier` applies to a real packing list.
    all_lines = (
        db.query(InboundShipmentLine)
        .filter(InboundShipmentLine.shipment_id == shipment.id)
        .all()
    )
    line_suppliers = {str(l.supplier_id) for l in all_lines if l.supplier_id}
    shipment.supplier_id = next(iter(line_suppliers)) if len(line_suppliers) == 1 else None
    shipment.total_items_shipped = sum(l.quantity_shipped or 0 for l in all_lines)

    for key, group in groups.items():
        shipment_line = shipment_lines_by_key[key]
        for source_line in group["source_lines"]:
            db.add(
                ProformaInvoiceShipmentLink(
                    id=_uuid(),
                    proforma_invoice_id=source_line.invoice_id,
                    proforma_invoice_line_id=source_line.id,
                    inbound_shipment_id=shipment.id,
                    inbound_shipment_line_id=shipment_line.id,
                    # HOW MUCH of the line came here. The line's own quantity is no longer
                    # the answer: since Q9 it may be split across two containers.
                    qty=group["placed"].get(str(source_line.id)),
                )
            )
    # A skip is recorded ONCE per line. A repeat convert of the same invoice would otherwise
    # write a second identical "no catalogue match" row every time, and the PI detail would
    # report the same line going nowhere twice.
    already_skipped: set[str] = set()
    if skipped:
        # Guarded rather than passed an empty list: `IN ('')` against a uuid column is a
        # DataError, not an empty result.
        already_skipped = {
            str(row[0])
            for row in db.query(ProformaInvoiceShipmentLink.proforma_invoice_line_id)
            .filter(
                ProformaInvoiceShipmentLink.proforma_invoice_line_id.in_(
                    [str(l.id) for l, _ in skipped]
                ),
                ProformaInvoiceShipmentLink.inbound_shipment_line_id.is_(None),
            )
            .all()
        }
    for source_line, reason in skipped:
        if str(source_line.id) in already_skipped:
            continue
        db.add(
            ProformaInvoiceShipmentLink(
                id=_uuid(),
                proforma_invoice_id=source_line.invoice_id,
                proforma_invoice_line_id=source_line.id,
                inbound_shipment_id=shipment.id,
                inbound_shipment_line_id=None,
                unmatched_reason=reason,
            )
        )
    db.flush()

    from app.services.procurement_service import InboundShipmentService

    # Flushes its own changes; does not commit - same "the caller commits" contract as
    # every other write in this module (`apply` above).
    InboundShipmentService(db).refresh_shipment_line_statuses(shipment.id)

    return {
        "shipment_id": str(shipment.id),
        "shipment_number": shipment.shipment_number,
        "shipment_status": shipment.shipment_status,
        "supplier_id": str(shipment.supplier_id) if shipment.supplier_id else None,
        "lines_created": len(groups),
        "lines_skipped": len(skipped),
        # Invoices in the selection that had nothing left to place. Named rather than
        # silently dropped, so the caller can say which ones did not move (AC-F7).
        "skipped_invoices": skipped_invoices,
        "invoices": [
            {
                "id": str(inv.id),
                "pi_number": inv.pi_number,
                "supplier_id": str(inv.supplier_id) if inv.supplier_id else None,
                "supplier_name": _supplier_label(db, str(inv.supplier_id))[1],
            }
            for inv in invoices
        ],
        "unmatched": [
            {
                "proforma_invoice_id": str(source_line.invoice_id),
                "pi_number": found_by_id[str(source_line.invoice_id)].pi_number,
                "line_no": source_line.line_no,
                "item_code": source_line.item_code,
                "reason": reason,
            }
            for source_line, reason in skipped
        ],
    }


#: The pre-loading list's own column spellings, in its own order, for the sheet that goes
#: BACK to the supplier (AC-E4). Their header, so the document they receive is recognisably
#: the one they sent - a Sorento-worded sheet would have to be read before it can be acted
#: on. Only the columns we actually hold: 规格 / 商标 are read from their file and never
#: stored, so inventing empty ones here would suggest we lost them.
_EXPORT_COLUMNS = (
    "序号",
    "产品型号",
    "品名",
    "数量",
    "箱数",
    "体积(cbm)",
    "总体积(cbm)",
    "单价",
    "金额",
    "备注",
)
_EXPORT_WIDTHS = (6, 24, 18, 10, 10, 12, 14, 12, 14, 34)


def _num(value: Optional[float]) -> str:
    """A number as a person writes it: `65`, `69.36`, never `65.0` or `4.359999999`."""
    if value is None:
        return "-"
    text_value = f"{float(value):.4f}".rstrip("0").rstrip(".")
    return text_value or "0"


def to_xlsx(payload: dict) -> bytes:
    """The adjusted invoice in the supplier's own block layout (AC-E4).

    Quantities, volumes and amounts are OURS as adjusted; where a line's quantity differs
    from what they sent, their own figure travels in 备注 - the remarks column their sheet
    already has - so the difference is visible on the page rather than only in our database.
    """
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active or wb.create_sheet()
    ws.title = "PROFORMA INVOICE"
    bold = Font(bold=True)

    for label, value in (
        ("货单号 / PI No.", payload.get("pi_number")),
        ("供应商 / Supplier", payload.get("supplier_name")),
        ("日期 / Date", payload.get("invoice_date")),
        ("货柜号 / Container No.", payload.get("container_no")),
        ("提单号 / B/L No.", payload.get("bl_no")),
    ):
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = bold
    ws.append([])

    ws.append(list(_EXPORT_COLUMNS))
    for col in range(1, len(_EXPORT_COLUMNS) + 1):
        ws.cell(row=ws.max_row, column=col).font = bold

    qty_total = 0.0
    carton_total = 0.0
    cbm_total = 0.0
    amount_total = 0.0
    for i, line in enumerate(payload.get("lines") or [], start=1):
        qty = line.get("qty")
        supplier_qty = line.get("supplier_qty")
        remarks = [line.get("remark")] if line.get("remark") else []
        if supplier_qty is not None and qty is not None and float(supplier_qty) != float(qty):
            remarks.append(f"原数量 / Supplier qty: {_num(supplier_qty)}")
        ws.append(
            [
                i,
                line.get("item_code"),
                line.get("description"),
                qty,
                line.get("cartons"),
                line.get("cbm_per_unit"),
                line.get("cbm_total"),
                line.get("unit_price"),
                line.get("amount"),
                "; ".join(remarks) or None,
            ]
        )
        qty_total += float(qty or 0)
        carton_total += float(line.get("cartons") or 0)
        cbm_total += float(line.get("cbm_total") or 0)
        amount_total += float(line.get("amount") or 0)

    ws.append(
        [
            "总计 / Total", None, None, qty_total, carton_total, None,
            round(cbm_total, 4), None, round(amount_total, 2), None,
        ]
    )
    for col in range(1, len(_EXPORT_COLUMNS) + 1):
        ws.cell(row=ws.max_row, column=col).font = bold

    capacity = payload.get("container_cbm")
    if capacity:
        over = payload.get("over_by_cbm")
        ws.append(
            [
                f"货柜 / Container {payload.get('container_size_code') or ''}".strip(),
                None, None, None, None, None, _num(capacity), None,
                f"超出 / Over by {_num(over)} cbm" if over else None, None,
            ]
        )

    for i, width in enumerate(_EXPORT_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    from io import BytesIO

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_filename(payload: dict) -> str:
    """`<PI number>-proforma.xlsx`, with anything a filesystem would argue about removed.

    Named after the invoice, never after its id: a downloaded file called
    `9f3a1c...xlsx` tells the supplier who opens it nothing at all.
    """
    stem = re.sub(r"[^A-Za-z0-9._-]", "", str(payload.get("pi_number") or "")) or "proforma"
    return f"{stem}-proforma.xlsx"


def bulk_delete(db: Session, invoice_ids: list[str]) -> dict:
    """Hard delete several proforma invoices at once, same shape as the PO book's bulk
    delete (`PurchaseOrderService.bulk_delete`): ids not found (already deleted, or another
    company's) are skipped rather than failing the whole batch.

    A PI that has already been converted (a `ProformaInvoiceShipmentLink` row that reached a
    shipment LINE) is REFUSED rather than deleted - the safer of the two choices the plan
    named. A SKIP row (`inbound_shipment_line_id` NULL: the convert looked at the line and
    put nothing on the container) does not count, the same reading `_placements` gives the
    list; before this the list said "Not converted" and the delete said "already converted"
    about the same invoice (captain, 27 Aug, three all-skipped PIs on the dev copy). Cascading
    the link instead would silently sever a draft shipment's line from the document that
    justified it, and a shipment already visible on `/scm/incoming` losing its "why" with no
    trace is worse than a delete the operator has to go and untangle by hand (delete the
    shipment first, or accept the PI stays on file). Named per invoice so the caller knows
    exactly which ones were blocked and why, rather than the batch failing outright.
    """
    ids = list(dict.fromkeys(str(i).strip() for i in (invoice_ids or []) if str(i).strip()))
    if not ids:
        raise AppException(422, "Select at least one proforma invoice to delete.")

    invoices = db.query(ProformaInvoice).filter(ProformaInvoice.id.in_(ids)).all()
    if not invoices:
        return {"deleted": 0, "blocked": []}

    converted_invoice_ids = {
        str(row[0])
        for row in (
            db.query(ProformaInvoiceShipmentLink.proforma_invoice_id)
            .filter(
                ProformaInvoiceShipmentLink.proforma_invoice_id.in_([str(i.id) for i in invoices]),
                ProformaInvoiceShipmentLink.inbound_shipment_line_id.isnot(None),
            )
            .distinct()
            .all()
        )
    }
    shipment_numbers: dict[str, str] = {}
    if converted_invoice_ids:
        rows = (
            db.query(ProformaInvoiceShipmentLink.proforma_invoice_id, InboundShipment.shipment_number)
            .join(InboundShipment, InboundShipment.id == ProformaInvoiceShipmentLink.inbound_shipment_id)
            .filter(
                ProformaInvoiceShipmentLink.proforma_invoice_id.in_(converted_invoice_ids),
                ProformaInvoiceShipmentLink.inbound_shipment_line_id.isnot(None),
            )
            .all()
        )
        for inv_id, number in rows:
            shipment_numbers.setdefault(str(inv_id), number or "?")

    blocked: list[dict] = []
    deleted = 0
    for invoice in invoices:
        inv_id = str(invoice.id)
        if inv_id in converted_invoice_ids:
            blocked.append(
                {
                    "id": inv_id,
                    "pi_number": invoice.pi_number,
                    "shipment_number": shipment_numbers.get(inv_id),
                }
            )
            continue
        db.delete(invoice)
        deleted += 1
    db.flush()
    return {"deleted": deleted, "blocked": blocked}


def _container_sizes(db: Session) -> tuple[dict[str, Any], Optional[Any]]:
    """Every active container size by id, and whichever one is the tenant's default.

    Read once per serialization rather than per invoice: a page of 25 invoices asks the
    same three-row question 25 times otherwise.
    """
    rows = db.query(ContainerSize).filter(ContainerSize.is_active.is_(True)).all()
    return {str(r.id): r for r in rows}, next((r for r in rows if r.is_default), None)


def _volumes(db: Session, invoice_ids: list[str]) -> dict[str, tuple[Optional[float], int]]:
    """Per invoice, `(total cbm, lines with no volume)` in ONE query.

    The total is NULL rather than 0 when no line states a volume: 0 cbm would read as an
    empty container on a document that simply never measured itself (AC-D2).
    """
    ids = [i for i in set(invoice_ids) if i]
    if not ids:
        return {}
    rows = (
        db.query(
            ProformaInvoiceLine.invoice_id,
            func.sum(ProformaInvoiceLine.cbm_total),
            func.count(1).filter(ProformaInvoiceLine.cbm_total.is_(None)),
        )
        .filter(ProformaInvoiceLine.invoice_id.in_(ids))
        .group_by(ProformaInvoiceLine.invoice_id)
        .all()
    )
    return {str(inv): (_f(total), int(unmeasured or 0)) for inv, total, unmeasured in rows}


def _fit(
    invoice: ProformaInvoice,
    total_cbm: Optional[float],
    sizes_by_id: dict[str, Any],
    default_size: Optional[Any],
) -> dict[str, Any]:
    """How full this invoice's container is, and which container that is.

    The size is RESOLVED at read time from the tenant's default when the invoice names none,
    rather than copied onto the row at import: a PI uploaded before anybody thought about
    capacity should follow the default, not freeze whatever it happened to be that day
    (AC-D4). An invoice that DOES name one keeps it, which is what makes it changeable.
    """
    size = sizes_by_id.get(str(invoice.container_size_id)) if invoice.container_size_id else None
    if size is None:
        size = default_size
    capacity = _f(size.cbm) if size is not None else None
    fill = (
        (total_cbm / capacity) * 100
        if total_cbm is not None and capacity
        else None
    )
    over = (
        round(total_cbm - capacity, 4)
        if total_cbm is not None and capacity and total_cbm > capacity
        else None
    )
    return {
        "container_size_id": str(size.id) if size is not None else None,
        "container_size_code": size.code if size is not None else None,
        "container_cbm": capacity,
        "total_cbm": total_cbm,
        "fill_pct": round(fill, 2) if fill is not None else None,
        "over_by_cbm": over,
    }


def _placed_invoice_ids(db: Session, invoice_ids: list[str]) -> dict[str, str]:
    """Of these invoices, the ones whose goods are actually ON a container, and where.

    The ONE definition of "placed", shared by every reader and every guard: a link row that
    names a real shipment LINE. A row with a NULL `inbound_shipment_line_id` is not a
    placement - it is either a SKIP (a line that matched no product, recorded so the detail
    page can say what happened to it) or a link whose shipment line has since been deleted
    and the FK set null. The placement readers have always counted it that way; the guards
    counted any row at all, so an invoice whose only link was a skip refused every write for
    ever and read as converted on a container that never held it.
    """
    ids = [str(i) for i in invoice_ids if i]
    if not ids:
        return {}
    rows = (
        db.query(
            ProformaInvoiceShipmentLink.proforma_invoice_id, InboundShipment.shipment_number
        )
        .join(
            InboundShipment,
            InboundShipment.id == ProformaInvoiceShipmentLink.inbound_shipment_id,
        )
        .filter(
            ProformaInvoiceShipmentLink.proforma_invoice_id.in_(ids),
            ProformaInvoiceShipmentLink.inbound_shipment_line_id.isnot(None),
        )
        .all()
    )
    out: dict[str, str] = {}
    for invoice_id, number in rows:
        out.setdefault(str(invoice_id), number or "?")
    return out


def _replaceable_or_409(db: Session, invoice: ProformaInvoice) -> None:
    """May a plain re-upload REPLACE this invoice's lines?

    `apply` is idempotent by identity: the same file uploaded twice lands on the same
    invoices and replaces their lines, which is what stops a nervous second Confirm doubling
    a document (AC-P1.4). That is only harmless while nobody has done anything to the
    invoice since. Three states make it harmful, and all three were reachable:

      * ADJUSTED - the replacement throws away the quantities Ms Tee trimmed to fit the box,
        with nothing on screen saying so;
      * PLACED - the line delete CASCADES `proforma_invoice_shipment_link` away, so the
        invoice reads not-converted again while its goods sit on a container;
      * SUPERSEDED - it is what the supplier sent on a day that has passed, and rewriting it
        rewrites the diff the current revision is read against.

    Each refusal names the invoice and points at the revision path, because that IS the
    supported way to bring a supplier's second send in (AC-E6).
    """
    number = invoice.pi_number
    if (invoice.status or "current") == "superseded":
        raise AppException(
            409,
            f"'{number}' has been superseded by a newer revision and cannot be replaced. "
            "Upload the new file as a revision of the current one instead.",
            code="superseded",
        )
    placed = _placed_invoice_ids(db, [str(invoice.id)])
    if placed:
        raise AppException(
            409,
            f"'{number}' is already in packing list '{placed[str(invoice.id)]}', so this "
            "file cannot replace it. Upload it as a revision instead.",
            code="already_converted",
        )
    if invoice.adjusted_at is not None:
        raise AppException(
            409,
            f"'{number}' has been adjusted since it was uploaded, so this file cannot "
            "replace it. Upload it as a revision instead.",
            code="already_adjusted",
        )


def _editable_or_409(db: Session, invoice: ProformaInvoice) -> None:
    """Refuse an adjustment that would make the document disagree with reality.

    Two states are closed. A SUPERSEDED revision is what the supplier sent on a day that has
    passed, and editing it would rewrite history nobody can see (AC-E7). A CONVERTED invoice
    has already had its goods drafted onto a shipment, so trimming it afterwards leaves the
    two documents disagreeing with nothing on screen saying which one the container was
    loaded from.
    """
    if (invoice.status or "current") == "superseded":
        raise AppException(
            409,
            f"'{invoice.pi_number}' has been superseded by a newer revision and is read-only.",
            code="superseded",
        )
    # The SAME "live link" predicate the placement readers use (`_placed_invoice_ids`): a
    # skip row, or a link whose shipment line has been deleted, is not a container the goods
    # are on, and counting it here froze an invoice nothing was holding.
    placed = _placed_invoice_ids(db, [str(invoice.id)])
    if placed:
        raise AppException(
            409,
            f"'{invoice.pi_number}' is already in packing list "
            f"'{placed[str(invoice.id)]}', so its quantities can no longer be changed.",
            code="already_converted",
        )


def _line_or_404(db: Session, invoice: ProformaInvoice, line_id: str) -> ProformaInvoiceLine:
    """One line OF THIS INVOICE. A line id belonging to another invoice is a 404, never a
    silent write against a document the caller was not looking at."""
    if not _is_uuid(line_id):
        raise AppException(404, "That line is not on this proforma invoice.", detail="line_id")
    line = (
        db.query(ProformaInvoiceLine)
        .filter(
            ProformaInvoiceLine.id == str(line_id),
            ProformaInvoiceLine.invoice_id == invoice.id,
        )
        .first()
    )
    if line is None:
        raise AppException(404, "That line is not on this proforma invoice.", detail="line_id")
    return line


def _restate(db: Session, invoice: ProformaInvoice, *, actor: Optional[str]) -> None:
    """Re-derive what the invoice says about itself after a line changed, and stamp who did it.

    `total_amount` follows the lines here even when the document stated its own total on
    import: once Sorento has trimmed a line, the supplier's printed total describes an
    invoice that no longer exists, and leaving it would put a figure on screen that agrees
    with nothing under it.
    """
    lines = (
        db.query(ProformaInvoiceLine)
        .filter(ProformaInvoiceLine.invoice_id == invoice.id)
        .order_by(ProformaInvoiceLine.line_no)
        .all()
    )
    total = Decimal("0")
    for ln in lines:
        if ln.amount is not None:
            total += Decimal(str(ln.amount))
        elif ln.unit_price is not None and ln.qty is not None:
            total += Decimal(str(ln.unit_price)) * Decimal(str(ln.qty))
    invoice.total_amount = total
    invoice.line_count = len(lines)
    invoice.adjusted_by = actor
    invoice.adjusted_at = datetime.utcnow()
    db.flush()


def adjust_line(
    db: Session, invoice_id: str, line_id: str, *, qty: float, actor: Optional[str] = None
) -> dict:
    """Sorento's own quantity for one line (AC-E1). Does not commit.

    `supplier_qty` and `supplier_unit_price` are NOT touched - they are the supplier's
    statement and this is ours (AC-E2). The volume and the money follow the new quantity,
    because both are per-unit facts multiplied by it, and a fill bar computed from a stale
    total is exactly the number that gets a container booked twice.
    """
    invoice = get_or_404(db, invoice_id)
    _editable_or_409(db, invoice)
    line = _line_or_404(db, invoice, line_id)

    try:
        new_qty = Decimal(str(qty))
    except Exception:  # noqa: BLE001 - the message is for whoever typed it
        raise AppException(422, "Enter a quantity of zero or more.", detail="qty")
    if new_qty < 0:
        raise AppException(422, "Enter a quantity of zero or more.", detail="qty")

    line.qty = new_qty
    per_unit = line.cbm_per_unit
    if per_unit is not None:
        line.cbm_total = Decimal(str(per_unit)) * new_qty
    if line.unit_price is not None:
        line.amount = Decimal(str(line.unit_price)) * new_qty
    db.flush()

    _restate(db, invoice, actor=actor)
    return serialize(db, invoice)


def remove_line(
    db: Session, invoice_id: str, line_id: str, *, actor: Optional[str] = None
) -> dict:
    """Take one line off the invoice entirely - it is not going in this container.

    A hard delete, per the CRUD standard, behind the caller's confirmation dialog. The
    remaining lines are NOT renumbered: `line_no` is where the line sat on the document the
    supplier sent, and re-flowing it would make our copy disagree with their paper.
    """
    invoice = get_or_404(db, invoice_id)
    _editable_or_409(db, invoice)
    line = _line_or_404(db, invoice, line_id)
    db.delete(line)
    db.flush()
    _restate(db, invoice, actor=actor)
    return serialize(db, invoice)


def set_container_size(
    db: Session, invoice_id: str, container_size_id: Optional[str]
) -> dict:
    """Which box this invoice is being fitted into. `None` means the tenant's default."""
    invoice = get_or_404(db, invoice_id)
    _set_container_size(db, invoice, container_size_id)
    db.flush()
    return serialize(db, invoice)


#: "The caller did not mention this field", which is a different instruction from "set it to
#: null" - `container_size_id: null` means the tenant default, and an absent one means leave
#: the invoice's own choice alone.
UNSET: Any = object()


def _set_container_size(db: Session, invoice: ProformaInvoice, container_size_id: Optional[str]) -> None:
    """Resolve and set the box. Shared by the PATCH and the whole-document PUT."""
    if container_size_id:
        if not _is_uuid(container_size_id):
            raise AppException(422, "That container size does not exist.", detail="container_size_id")
        size = (
            db.query(ContainerSize)
            .filter(ContainerSize.id == str(container_size_id))
            .first()
        )
        if size is None:
            raise AppException(404, "That container size does not exist.", detail="container_size_id")
        invoice.container_size_id = str(size.id)
    else:
        invoice.container_size_id = None


def _rename(db: Session, invoice: ProformaInvoice, pi_number: Optional[str]) -> None:
    """The document's own number, corrected by hand.

    A derived number (`PI-<file stem>-<block>`) is a guess the file forced on us, and the
    operator reading the paper is the one who knows what the supplier actually called it.
    Identity is `(company, supplier, pi_number)`, so the rename is refused by NAME rather
    than left to the unique index, which would surface as a 500 with a constraint in it.
    """
    number = (pi_number or "").strip()
    if not number:
        raise AppException(422, "The invoice needs a number.", detail="pi_number")
    if len(number) > 100:
        raise AppException(422, "That invoice number is too long.", detail="pi_number")
    if number == invoice.pi_number:
        return
    clash = (
        db.query(ProformaInvoice.id)
        .filter(
            ProformaInvoice.supplier_id == invoice.supplier_id,
            ProformaInvoice.pi_number == number,
            ProformaInvoice.id != invoice.id,
        )
        .first()
    )
    if clash:
        raise AppException(
            409,
            f"This supplier already has a proforma invoice numbered '{number}'.",
            code="duplicate_pi_number",
            detail="pi_number",
        )
    invoice.pi_number = number


def _write_lines(db: Session, invoice: ProformaInvoice, rows: list[dict]) -> None:
    """The whole line set in one write: rows with an id update, rows without create, and a
    line the payload no longer names is deleted.

    The same upsert shape `PurchaseOrderService` uses for a purchase order's lines, and for
    the same reason: the screen edits a DRAFT of the document and presses Save once, so the
    server is told what the invoice IS rather than replayed the keystrokes that got there.

    `line_no` is NOT re-flowed. It is where the line sat on the paper the supplier sent, so a
    removed line leaves a gap and a new one is appended after the highest number in use.
    """
    existing = {
        str(ln.id): ln
        for ln in db.query(ProformaInvoiceLine)
        .filter(ProformaInvoiceLine.invoice_id == invoice.id)
        .all()
    }
    next_line_no = max((int(ln.line_no or 0) for ln in existing.values()), default=0) + 1
    kept: set[str] = set()

    for row in rows:
        code = (row.get("item_code") or "").strip()
        if not code:
            raise AppException(422, "Every line needs an item code.", detail="item_code")
        try:
            qty = Decimal(str(row.get("qty")))
        except Exception:  # noqa: BLE001 - the message is for whoever typed it
            raise AppException(422, "Enter a quantity of zero or more.", detail="qty")
        if qty < 0:
            raise AppException(422, "Enter a quantity of zero or more.", detail="qty")

        product_id = row.get("product_id") or None
        if product_id:
            if not _is_uuid(str(product_id)):
                raise AppException(422, "That product does not exist.", detail="product_id")
            found = db.query(Product.id).filter(Product.id == str(product_id)).first()
            if found is None:
                raise AppException(404, "That product does not exist.", detail="product_id")

        line_id = row.get("id")
        if line_id:
            line = existing.get(str(line_id))
            if line is None:
                raise AppException(
                    404, "That line is not on this proforma invoice.", detail="line_id"
                )
            kept.add(str(line_id))
        else:
            line = ProformaInvoiceLine(
                id=_uuid(), invoice_id=invoice.id, line_no=next_line_no, item_code=code, qty=qty
            )
            next_line_no += 1
            db.add(line)

        unit_price = row.get("unit_price")
        cbm_per_unit = row.get("cbm_per_unit")
        uom = (row.get("uom") or "").strip()
        line.item_code = code[:100]
        line.description = row.get("description")
        line.qty = qty
        line.uom = uom[:20] or None
        line.cartons = None if row.get("cartons") is None else Decimal(str(row["cartons"]))
        line.cbm_per_unit = None if cbm_per_unit is None else Decimal(str(cbm_per_unit))
        # Derived, never asked for: the total volume is the per-unit figure times what is
        # being shipped, and a stored total that disagreed with the two beside it is the
        # number that gets a container booked twice.
        line.cbm_total = None if cbm_per_unit is None else Decimal(str(cbm_per_unit)) * qty
        line.unit_price = None if unit_price is None else Decimal(str(unit_price))
        line.amount = None if unit_price is None else Decimal(str(unit_price)) * qty
        line.net_weight = (
            None if row.get("net_weight") is None else Decimal(str(row["net_weight"]))
        )
        line.gross_weight = (
            None if row.get("gross_weight") is None else Decimal(str(row["gross_weight"]))
        )
        line.product_id = str(product_id) if product_id else None

    for line_id, line in existing.items():
        if line_id not in kept:
            db.delete(line)
    db.flush()


def update_invoice(
    db: Session,
    invoice_id: str,
    *,
    pi_number: Any = UNSET,
    container_size_id: Any = UNSET,
    lines: Any = UNSET,
    actor: Optional[str] = None,
) -> dict:
    """The whole document as the edit screen holds it, written in ONE call. Does not commit.

    The detail page edits a local draft and nothing is written until Save, so this is the
    only write that save makes: rename, container size and the whole line set travel
    together and either all land or none do. A superseded revision and an invoice whose
    goods are already on a container are refused here, by the same rules that refuse a
    single-line adjustment.
    """
    invoice = get_or_404(db, invoice_id)
    _editable_or_409(db, invoice)

    if pi_number is not UNSET:
        _rename(db, invoice, pi_number)
    if container_size_id is not UNSET:
        _set_container_size(db, invoice, container_size_id)
    if lines is not UNSET:
        _write_lines(db, invoice, list(lines or []))
        # Stamped only when the LINES moved: renaming a document or choosing a different box
        # is not trimming it to fit, and "adjusted by" is what tells the screen the
        # supplier's figures are no longer what we are shipping.
        _restate(db, invoice, actor=actor)
    else:
        db.flush()
    return serialize(db, invoice)


def get_or_404(db: Session, invoice_id: str) -> ProformaInvoice:
    """One invoice, or a refusal. Never another company's.

    A 404 rather than an empty document: the id came from somewhere, and rendering an invoice
    the caller's company does not own would be worse than saying no. The ORM query carries
    the company filter, so an id belonging to another company reads as "not found" here -
    which is the honest answer to a caller who cannot see it.
    """
    if not _is_uuid(invoice_id):
        raise AppException(404, "Proforma invoice not found.", detail="invoice_id")
    invoice = db.query(ProformaInvoice).filter(ProformaInvoice.id == str(invoice_id)).first()
    if invoice is None:
        raise AppException(404, "Proforma invoice not found.", detail="invoice_id")
    return invoice


def list_for_supplier(
    db: Session,
    *,
    supplier_id: Optional[str] = None,
    placement: Optional[str] = None,
    query: Optional[str] = None,
    limit: int = 25,
    offset: int = 0,
) -> dict:
    """Invoices we have read, newest first. `supplier_id` narrows it to one supplier.

    `query` is the list screen's search box: PI number, supplier (name or code), container
    number or bill of lading, case-insensitively. One box over the four identifiers a person
    actually arrives with, rather than four filters they have to choose between.

    `total` is counted separately from the page, so it is how many invoices are held rather
    than how many came back: `len(rows)` after a `LIMIT` can never exceed the page size, and a
    supplier with more invoices than one page would have read as having exactly a page of them.
    `offset` is what makes the rest of them reachable.

    The id breaks ties in the sort, and the ties are guaranteed rather than unlikely:
    `created_at` defaults to `now()`, which is the TRANSACTION timestamp, so all five invoices
    of one stacked pre-loading list share it to the microsecond. Postgres promises no order
    among equal keys, so without the tiebreaker a caller walking offset=0 then offset=25 can
    be handed one invoice twice and never shown another.
    """
    if supplier_id and not _is_uuid(supplier_id):
        raise AppException(422, "That supplier does not exist.", detail="supplier_id")

    if placement and placement not in _PLACEMENTS:
        raise AppException(
            422, "That is not a packing-list state.", detail="placement"
        )

    q = db.query(ProformaInvoice)
    if supplier_id:
        q = q.filter(ProformaInvoice.supplier_id == str(supplier_id))

    needle = (query or "").strip()
    if needle:
        # The four things somebody has in their hand when they come looking for an invoice:
        # its number, whose it is, which box it went in and which bill of lading covers it.
        # The supplier is matched through a scoped subquery rather than a join, so the
        # company filter the ORM puts on `Supplier` still applies and the page count below
        # stays one query.
        like = f"%{needle}%"
        supplier_ids = (
            db.query(Supplier.id)
            .filter(
                func.lower(Supplier.supplier_name).like(func.lower(like))
                | func.lower(Supplier.supplier_code).like(func.lower(like))
            )
            .subquery()
        )
        q = q.filter(
            func.lower(ProformaInvoice.pi_number).like(func.lower(like))
            | func.lower(ProformaInvoice.container_ref).like(func.lower(like))
            | func.lower(ProformaInvoice.bl_ref).like(func.lower(like))
            | ProformaInvoice.supplier_id.in_(db.query(supplier_ids.c.id))
        )

    if placement:
        # Filtered in Python rather than in SQL, deliberately: "split" is the comparison of
        # what is placed against what CAN be placed, and what can be placed excludes lines
        # with no catalogue product - a predicate that would take three correlated
        # subqueries to write and one line to get wrong. The candidate set here is one
        # supplier's invoices, not a table scan.
        candidates = q.order_by(
            ProformaInvoice.created_at.desc(), ProformaInvoice.id.desc()
        ).all()
        states = _quantities(db, [str(r.id) for r in candidates])
        matching = [
            r for r in candidates
            if states.get(str(r.id), {}).get("placement") == placement
        ]
        total = len(matching)
        rows = matching[max(offset, 0): max(offset, 0) + limit]
    else:
        total = q.count()
        rows = (
            q.order_by(ProformaInvoice.created_at.desc(), ProformaInvoice.id.desc())
            .offset(max(offset, 0))
            .limit(limit)
            .all()
        )

    labels = _supplier_labels(db, [str(r.supplier_id) for r in rows])
    volumes = _volumes(db, [str(r.id) for r in rows])
    sizes = _container_sizes(db)
    placements = _quantities(db, [str(r.id) for r in rows])
    return {
        "data": [
            serialize(
                db,
                r,
                with_lines=False,
                supplier_labels=labels,
                volumes=volumes,
                container_sizes=sizes,
                placements=placements,
            )
            for r in rows
        ],
        "total": total,
        "limit": limit,
        "offset": max(offset, 0),
    }


def delete(db: Session, invoice_id: str) -> None:
    """Hard delete, with its lines (the FK cascades), per the CRUD standard. Does not commit.

    Refused when this invoice has already been converted to a draft shipment (any
    `ProformaInvoiceShipmentLink` row against it) - `bulk_delete`'s docstring has the reasoning
    for refusing rather than cascading the link away silently.
    """
    invoice = get_or_404(db, invoice_id)
    link = (
        db.query(InboundShipment.shipment_number)
        .join(
            ProformaInvoiceShipmentLink,
            ProformaInvoiceShipmentLink.inbound_shipment_id == InboundShipment.id,
        )
        .filter(ProformaInvoiceShipmentLink.proforma_invoice_id == invoice.id)
        .first()
    )
    if link:
        raise AppException(
            409,
            f"'{invoice.pi_number}' has already been converted to shipment "
            f"'{link[0] or '?'}' and cannot be deleted.",
            detail="already_converted",
        )
    db.delete(invoice)
    db.flush()


def serialize(
    db: Session,
    invoice: ProformaInvoice,
    *,
    with_lines: bool = True,
    supplier_labels: Optional[dict[str, tuple[Optional[str], Optional[str]]]] = None,
    volumes: Optional[dict[str, tuple[Optional[float], int]]] = None,
    container_sizes: Optional[tuple[dict[str, Any], Optional[Any]]] = None,
    placements: Optional[dict[str, dict]] = None,
) -> dict:
    """One invoice as the API returns it: codes and names, never a bare identifier.

    `supplier_labels`, `volumes` and `container_sizes` are the page's own lookups, resolved
    once by a caller listing several invoices; a single serialization resolves its own. All
    three are per-page rather than per-row because each is one query that would otherwise be
    asked twenty-five times for the same answer.
    """
    if supplier_labels is not None:
        supplier_code, supplier_name = supplier_labels.get(
            str(invoice.supplier_id), (None, None)
        )
    else:
        supplier_code, supplier_name = _supplier_label(db, str(invoice.supplier_id))
    out: dict[str, Any] = {
        "id": str(invoice.id),
        "supplier_id": str(invoice.supplier_id),
        "supplier_code": supplier_code,
        "supplier_name": supplier_name,
        "pi_number": invoice.pi_number,
        "invoice_date": invoice.invoice_date.isoformat() if invoice.invoice_date else None,
        "currency": invoice.currency or None,
        "container_no": invoice.container_ref,
        "bl_no": invoice.bl_ref,
        "total_amount": _f(invoice.total_amount),
        "line_count": invoice.line_count,
        "source_ref": invoice.source_ref,
        "block_index": invoice.block_index,
        "uploaded_by": invoice.uploaded_by,
        "created_at": invoice.created_at.isoformat() if invoice.created_at else None,
        "updated_at": invoice.updated_at.isoformat() if invoice.updated_at else None,
        # Who trimmed this document to fit, and when. Null on one nobody has touched, which
        # is what tells the screen to show the supplier's figures unqualified.
        "adjusted_by": invoice.adjusted_by,
        "adjusted_at": invoice.adjusted_at.isoformat() if invoice.adjusted_at else None,
        "is_adjusted": invoice.adjusted_at is not None,
        "status": invoice.status or "current",
        "revision_no": int(invoice.revision_no or 1),
    }

    chain = _chain(db, invoice)
    out["revision_count"] = len(chain)
    if with_lines:
        out["revisions"] = [
            {
                "id": str(r.id),
                "pi_number": r.pi_number,
                "revision_no": int(r.revision_no or 1),
                "status": r.status or "current",
                "invoice_date": r.invoice_date.isoformat() if r.invoice_date else None,
                "total_amount": _f(r.total_amount),
                "line_count": r.line_count,
            }
            for r in chain
        ]
        previous = next(
            (r for r in chain if invoice.revision_of_id and str(r.id) == str(invoice.revision_of_id)),
            None,
        )
        out["revision_of_pi_number"] = previous.pi_number if previous else None
        out["diff"] = _diff(db, invoice)

    quantities = (
        placements if placements is not None else _quantities(db, [str(invoice.id)])
    ).get(str(invoice.id), {})
    out.update(
        {
            "placement": quantities.get("placement", "not_converted"),
            "placed_qty": quantities.get("placed_qty", 0.0),
            "total_qty": quantities.get("total_qty", 0.0),
            "placeable_qty": quantities.get("placeable_qty", 0.0),
            "remaining_qty": quantities.get("remaining_qty", 0.0),
            "packing_lists": quantities.get("packing_lists", []),
        }
    )

    sizes_by_id, default_size = (
        container_sizes if container_sizes is not None else _container_sizes(db)
    )
    if volumes is None:
        volumes = _volumes(db, [str(invoice.id)])
    total_cbm, unmeasured = volumes.get(str(invoice.id), (None, 0))
    out.update(_fit(invoice, total_cbm, sizes_by_id, default_size))
    out["unmeasured_lines"] = unmeasured

    if not with_lines:
        return out

    lines = (
        db.query(ProformaInvoiceLine)
        .filter(ProformaInvoiceLine.invoice_id == invoice.id)
        .order_by(ProformaInvoiceLine.line_no)
        .all()
    )
    codes = _product_codes(db, [str(ln.product_id) for ln in lines if ln.product_id])
    # A line can be bound to one of our SETS instead (R19): no product carries the code the
    # supplier wrote, so the set's own code is what "this line matched" means for it.
    set_codes = {
        str(row.id): row.set_code
        for row in db.query(ProductSet)
        .filter(
            ProductSet.id.in_(
                [str(ln.product_set_id) for ln in lines if ln.product_set_id] or [""]
            )
        )
        .all()
    } if any(ln.product_set_id for ln in lines) else {}
    line_placements = _line_placements(db, str(invoice.id))
    # How each line's code came to mean a product, where it was not an exact agreement
    # (R16): an automatic bind has to be visible AS one, or nobody can be asked to check it.
    aliases = {
        str(row.supplier_code).strip().upper(): row
        for row in db.query(SupplierProductCodeAlias)
        .filter(SupplierProductCodeAlias.supplier_id == str(invoice.supplier_id))
        .all()
    }

    # Where each line went, if it has been converted (the PI -> draft-shipment convert):
    # `(shipment_id, shipment_number)` when it became a real shipment line, or an
    # `unmatched_reason` when the convert ran but this line was skipped. A line absent from
    # this map has simply never been converted.
    link_rows = (
        db.query(ProformaInvoiceShipmentLink, InboundShipment.shipment_number)
        .join(
            InboundShipment,
            InboundShipment.id == ProformaInvoiceShipmentLink.inbound_shipment_id,
        )
        .filter(ProformaInvoiceShipmentLink.proforma_invoice_id == invoice.id)
        .all()
    )
    links_by_line: dict[str, tuple[Any, str]] = {
        str(link.proforma_invoice_line_id): (link, shipment_number)
        for link, shipment_number in link_rows
    }
    # Header summary: every DISTINCT shipment this invoice's lines actually went to.
    #
    # A SKIP row is NOT one. It names the shipment whose convert ran, and records why
    # this line could not go on it - so counting it here put a container in the header
    # of an invoice that had never been on one, and the screen reading it locked the
    # lines panel ("already in a packing list, so the quantities are fixed") on an
    # invoice nothing was holding. Same rule as `_placed_invoice_ids`: a placement
    # names a shipment LINE.
    seen_shipments: dict[str, str] = {}
    for link, shipment_number in link_rows:
        if link.inbound_shipment_line_id is None:
            continue
        seen_shipments[str(link.inbound_shipment_id)] = shipment_number
    out["converted_shipments"] = [
        {"shipment_id": sid, "shipment_number": num} for sid, num in seen_shipments.items()
    ]

    out["lines"] = [
        {
            "id": str(ln.id),
            "line_no": ln.line_no,
            "row_number": ln.row_number,
            "item_code": ln.item_code,
            "description": ln.description,
            "qty": _f(ln.qty),
            "uom": ln.uom,
            "unit_price": _f(ln.unit_price),
            "amount": _f(ln.amount),
            "po_ref": ln.po_ref,
            "remark": ln.remark,
            "cartons": _f(ln.cartons),
            "cbm_per_unit": _f(ln.cbm_per_unit),
            "cbm_total": _f(ln.cbm_total),
            # What the supplier printed the line weighs. Null on a document that states
            # neither, never 0 - see the column's own note on the model.
            "net_weight": _f(ln.net_weight),
            "gross_weight": _f(ln.gross_weight),
            # What the SUPPLIER stated, frozen at import - the "was" beside our "now".
            "supplier_qty": _f(ln.supplier_qty),
            "supplier_unit_price": _f(ln.supplier_unit_price),
            # The product we hold, by CODE. A line that matched nothing says so rather than
            # carrying an id nobody can read.
            "product_code": (
                codes.get(str(ln.product_id))
                if ln.product_id
                else set_codes.get(str(ln.product_set_id))
                if ln.product_set_id
                else None
            ),
            # The SET this line names, when it names one - so the cell can badge it rather
            # than reading as an ordinary product code the catalogue does not hold.
            "set_code": set_codes.get(str(ln.product_set_id)) if ln.product_set_id else None,
            "matched": ln.product_id is not None or ln.product_set_id is not None,
            # Null on an exact match: the codes agreed, and nothing was worked out.
            "matched_by": (
                aliases[ln.item_code.strip().upper()].matched_by
                if (ln.product_id or ln.product_set_id)
                and ln.item_code.strip().upper() in aliases
                else None
            ),
            "match_source": (
                aliases[ln.item_code.strip().upper()].source
                if (ln.product_id or ln.product_set_id)
                and ln.item_code.strip().upper() in aliases
                else None
            ),
            "match_id": (
                str(aliases[ln.item_code.strip().upper()].id)
                if (ln.product_id or ln.product_set_id)
                and ln.item_code.strip().upper() in aliases
                else None
            ),
            # Where this line went, once converted - null on every line until the first
            # convert. `shipment_number`/`shipment_id` are set only when it became a real
            # shipment line; `unmatched_reason` is set only when the convert skipped it.
            "shipment_id": (
                str(links_by_line[str(ln.id)][0].inbound_shipment_id)
                if str(ln.id) in links_by_line and links_by_line[str(ln.id)][0].inbound_shipment_line_id
                else None
            ),
            "shipment_number": (
                links_by_line[str(ln.id)][1]
                if str(ln.id) in links_by_line and links_by_line[str(ln.id)][0].inbound_shipment_line_id
                else None
            ),
            "unmatched_reason": (
                links_by_line[str(ln.id)][0].unmatched_reason
                if str(ln.id) in links_by_line
                else None
            ),
            # Where this LINE went, and what is left of it - a line is what gets split
            # across two containers, and the remainder is what the next convert pre-fills.
            "placed_qty": sum(p["qty"] for p in line_placements.get(str(ln.id), [])),
            "remaining_qty": max(
                _placeable(ln) - sum(p["qty"] for p in line_placements.get(str(ln.id), [])),
                0.0,
            ),
            "packing_lists": line_placements.get(str(ln.id), []),
        }
        for ln in lines
    ]
    return out


def _supplier_labels(
    db: Session, supplier_ids: list[str]
) -> dict[str, tuple[Optional[str], Optional[str]]]:
    """`(supplier_code, supplier_name)` per id in one query, company-scoped by the ORM."""
    ids = [sid for sid in set(supplier_ids) if _is_uuid(sid)]
    if not ids:
        return {}
    rows = (
        db.query(Supplier.id, Supplier.supplier_code, Supplier.supplier_name)
        .filter(Supplier.id.in_(ids))
        .all()
    )
    return {str(sid): (code, name) for sid, code, name in rows}


def _product_codes(db: Session, product_ids: list[str]) -> dict[str, str]:
    """Product code per id, for the lines that matched one. Through the ORM, so the company
    filter applies without restating it - and because `id = ANY(:ids)` on a uuid column with
    string parameters is `operator does not exist: uuid = text`."""
    if not product_ids:
        return {}
    from app.models.product import Product

    rows = (
        db.query(Product.id, Product.product_code)
        .filter(Product.id.in_(list(set(product_ids))))
        .all()
    )
    return {str(pid): code for pid, code in rows}
