"""The Sorento packing list: one container, every factory that loaded it.

A container is routinely filled by two or three factories, each of which sends its own list.
Ms Tee used to retype all of them into one sheet, work out what each factory owed against the
loading plan we sent it, and split the whole thing between SORENTO and MOCHA by hand. This is
that sheet, derived.

Nothing here is typed in. The factory is the supplier the line was uploaded as, the company is
the product's brand, the volume is what the supplier stated or what our own catalogue measures,
and the remarks are the supplier's own.

It prints the SHIPMENT and nothing else (R20). It used to compare the container against the
loading plan each factory was sent and add its own remarks ("Loading plan asked 500, packed
490") plus a row per model the plan asked for and nobody loaded. Both are gone: this is the
document the forwarder, the factories and the clearance agent read, and a line on it for goods
that never went into the container is read as goods that shipped.

Reads only: `build` never writes, and `to_xlsx` is a pure function of what `build` returned.
"""
from __future__ import annotations

import re
from datetime import date, datetime
from io import BytesIO
from typing import Optional

from sqlalchemy.orm import Session

from app.models.procurement import InboundShipment, InboundShipmentLine, Supplier
from app.models.product import Brand, Product
from app.services.error_handler import AppException
from app.services.scm.supplier_scope import is_uuid

#: The captain's own file counts SANDEL / CABANA / blank under SORENTO, so this is a single
#: named brand rather than a lookup table: MOCHA is the one that is invoiced separately.
MOCHA = "MOCHA"
SORENTO = "SORENTO"
COMPANIES = (SORENTO, MOCHA)

#: What a factory with no name is called on the sheet. It sorts last.
UNASSIGNED = "Unassigned"

#: Catalogue dimensions are millimetres; volume is cubic metres.
_MM3_PER_M3 = 1_000_000_000.0


def _f(v) -> Optional[float]:
    return None if v is None else float(v)


def _qty(v: float):
    """A quantity as a person writes it: 500, not 500.0, unless it really has a fraction."""
    return int(v) if float(v).is_integer() else float(v)


def _num(v: float) -> str:
    return str(_qty(v))


def _catalogue_cbm(product: Product, qty: float) -> Optional[float]:
    """Volume from the catalogue when the packing list did not state one.

    The formula is COPIED from `loading_plan_service._catalogue_cbm` rather than imported:
    that module is being changed by another lane, and two lines of arithmetic are a smaller
    cost than a merge conflict across the two features. Same mm^3 basis, times the quantity
    on the line (the loading plan wants a per-unit figure; a packing list wants the line's).
    """
    l, w, h = product.dimensions_length, product.dimensions_width, product.dimensions_height
    if l is None or w is None or h is None:
        return None
    return round(float(l) * float(w) * float(h) / _MM3_PER_M3 * float(qty), 6)


def _totals(lines: list[dict]) -> dict:
    """Sum a set of lines, and say how much of the volume it actually knows.

    `cbm_known_lines` is not decoration. A container planned against a volume that only two
    thirds of its lines contributed to is planned against a number nobody measured, and a
    partial figure printed as a full one is exactly how that happens.
    """
    known = [l["cbm"] for l in lines if l["cbm"] is not None]
    return {
        "lines": len(lines),
        "qty": sum(int(l["qty"]) for l in lines),
        "cartons": sum(int(l["cartons"] or 0) for l in lines),
        "cbm": round(sum(known), 6),
        "cbm_known_lines": len(known),
    }


def _shipment_or_404(db: Session, shipment_id: str) -> InboundShipment:
    if not is_uuid(shipment_id):
        raise AppException(404, "Inbound shipment not found")
    row = db.query(InboundShipment).filter(InboundShipment.id == shipment_id).one_or_none()
    if row is None:
        raise AppException(404, "Inbound shipment not found")
    return row


def _as_date(value) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    return value if isinstance(value, date) else None


def _iso(value) -> Optional[str]:
    return value.isoformat() if value is not None else None


def build(db: Session, shipment_id: str) -> dict:
    """The consolidated list for one container. 404 only when the container is unknown.

    A container with no lines is a container that has been read but not unpacked, which the
    screen has to be able to draw - so it is an empty list of factories, not an error.
    """
    shipment = _shipment_or_404(db, shipment_id)

    # One query per table. The product and its brand travel with the line because every line
    # needs both: the model number to print and the brand to split the company by.
    rows = (
        db.query(InboundShipmentLine, Product, Brand)
        .join(Product, Product.id == InboundShipmentLine.product_id)
        .outerjoin(Brand, Brand.id == Product.brand_id)
        .filter(InboundShipmentLine.shipment_id == str(shipment.id))
        .all()
    )

    supplier_ids = {str(line.supplier_id) for line, _p, _b in rows if line.supplier_id}
    suppliers = (
        {
            str(s.id): s
            for s in db.query(Supplier).filter(Supplier.id.in_(supplier_ids)).all()
        }
        if supplier_ids
        else {}
    )
    grouped: dict[Optional[str], list[dict]] = {}
    for line, product, brand in rows:
        brand_code = (brand.brand_code or "").strip() if brand else None
        qty = int(line.quantity_shipped or 0)
        cbm = _f(line.cbm)
        if cbm is None:
            cbm = _catalogue_cbm(product, qty)
        grouped.setdefault(str(line.supplier_id) if line.supplier_id else None, []).append(
            {
                "line_id": str(line.id),
                "product_id": str(product.id),
                "product_code": product.product_code,
                "product_name": product.product_name,
                "brand": brand_code or None,
                "company": MOCHA if (brand_code or "").upper() == MOCHA else SORENTO,
                "qty": qty,
                "cartons": int(line.cartons_count or 0),
                "cbm": cbm,
                "remarks": line.remarks,
                # What the container workbook measures the line by. All optional: a line
                # nobody measured prints an empty cell, and the sheet's own formula for
                # that row is left off rather than dividing by a figure nobody stated.
                "material": line.material,
                "pcs_per_carton": _f(line.pcs_per_carton),
                "carton_length_cm": _f(line.carton_length_cm),
                "carton_width_cm": _f(line.carton_width_cm),
                "carton_height_cm": _f(line.carton_height_cm),
                "net_weight": _f(line.net_weight_per_carton),
                # `weight_per_carton` is the column the line has always had, and every
                # packing list that states one weight states the GROSS. Read as the fallback
                # rather than migrated, so the containers already carrying it keep it.
                "gross_weight": _f(
                    line.gross_weight_per_carton
                    if line.gross_weight_per_carton is not None
                    else line.weight_per_carton
                ),
                "unit_cost": _f(line.unit_cost),
                "currency": line.currency,
            }
        )

    factories: list[dict] = []
    for supplier_id, lines in grouped.items():
        supplier = suppliers.get(supplier_id) if supplier_id else None
        lines.sort(key=lambda l: (l["product_code"] or "", l["line_id"]))
        factories.append(
            {
                "supplier_id": supplier_id,
                "supplier_code": supplier.supplier_code if supplier else None,
                "supplier_name": supplier.supplier_name if supplier else UNASSIGNED,
                "lines": lines,
                "subtotal": _totals(lines),
            }
        )

    # By factory name, with the one we were never told last: an unnamed group at the top
    # would be the first thing read and the least useful.
    factories.sort(key=lambda f: (f["supplier_id"] is None, (f["supplier_name"] or "").upper()))

    all_lines = [line for factory in factories for line in factory["lines"]]
    return {
        "shipment_id": str(shipment.id),
        "shipment_number": shipment.shipment_number,
        "container_no": shipment.shipping_container_number,
        "bl_no": shipment.bill_of_lading_number,
        "status": shipment.shipment_status,
        # The twelve lines the workbook prints above the goods. Every one of them is a
        # column on the container already - nothing here is worked out, and the factory
        # list is the one derived value, read off the lines that were actually loaded.
        "header": {
            "loading_date": _iso(shipment.loading_date),
            "etd": _iso(shipment.etd_date),
            # The REVISED eta where there is one: `eta_delay_date` is the accurate figure
            # and `estimated_arrival_date` is what was first published (migration 314).
            "eta": _iso(shipment.eta_delay_date or shipment.estimated_arrival_date),
            "container_no": shipment.shipping_container_number,
            "seal_number": shipment.seal_number,
            "forwarder_order_ref": shipment.forwarder_order_ref,
            "consignee": shipment.consignee,
            "shipper": shipment.shipper,
            # The sheet calls the China-side forwarder the china agent.
            "china_agent": shipment.china_forwarder,
            # Only the factories we were actually told about: the group of lines whose
            # factory nobody named is called `Unassigned` on the blocks below, and printing
            # that word on the header line would read as a company on the bill of lading.
            "factories": ", ".join(
                f["supplier_name"] for f in factories if f.get("supplier_id")
            ),
            "free_days": shipment.free_days_available,
            "delivery_warehouse": shipment.delivery_warehouse,
        },
        # Typed per container. The sheet's footer apportions the first two by each
        # company's share of the volume and the third by its share of the amount.
        "costs": {
            "clearance_cost": _f(shipment.clearance_cost),
            "china_freight_cost": _f(shipment.china_freight_cost),
            "insurance_rate": _f(shipment.insurance_rate),
        },
        "factories": factories,
        "total": _totals(all_lines),
        # Both rows always, zeros included: an absent company reads as a missing figure
        # rather than as nothing having shipped under it.
        "split": [
            {"company": company, **_totals([l for l in all_lines if l["company"] == company])}
            for company in COMPANIES
        ],
    }


# --------------------------------------------------------------------------- #
# the workbook
# --------------------------------------------------------------------------- #

#: The sheet, its columns and its arithmetic are `FSCU8103365.xlsx` tab `RMB` - the file Ms
#: Tee builds by hand and the one the forwarder, the factories and our own clearance agent
#: all read. It is copied exactly on purpose: a workbook that is nearly the familiar one is
#: worse than none, because every reader has to work out which column moved.
#:
#: Two headers, because `SIZE (CM)` spans three columns and `NW` / `GW` state their unit on
#: the second line. Row 15 is the wide header, row 16 the sub-header, lines start on 18.
_SHEET_TITLE = "RMB"
_HEADER_ROW = 15
_SUBHEADER_ROW = 16
_FIRST_LINE_ROW = 18

#: (column letter, row-15 label, row-16 label). A blank row-16 label means the row-15 cell
#: is merged down over both.
_COLUMNS_SPEC = [
    ("A", "FACTORY", None),
    ("B", "NO", None),
    ("C", "MODEL", None),
    ("D", "DESCRIPTION", None),
    ("E", "MATERIAL", None),
    ("F", "QTY", None),
    ("G", "PCS / CTN", None),
    ("H", "CTN QTY", None),
    ("I", "SIZE (CM)", "L"),
    ("J", None, "W"),
    ("K", None, "H"),
    ("L", "CBM\n/ CTN", None),
    ("M", "TOTAL CBM", None),
    ("N", "NW", "KG"),
    ("O", "GW", "KG"),
    ("P", "TOTAL NW", "KG"),
    ("Q", "TOTAL GW", "KG"),
    ("R", "LOGO", None),
    ("S", "REMARKS", None),
    ("T", "RMB", None),
    # Their spelling, not a typo of ours: the reference workbook heads the column `TOTAL RM`,
    # and a heading that differs from the file everybody already has is the one thing a
    # fidelity copy must not do.
    ("U", "TOTAL RM", None),
    ("V", None, None),
]

#: As measured off the source file, so a printed sheet lines up with the ones already filed.
#: `K` is absent on purpose - the reference sizes every column but that one, and giving it a
#: width here makes the three SIZE (CM) cells print unevenly.
_WIDTHS = {
    "A": 18.86, "B": 7.29, "C": 27.86, "D": 45.86, "E": 10.71, "F": 9.0, "G": 8.0,
    "H": 7.71, "I": 8.57, "J": 7.71, "L": 9.57, "M": 10.14, "N": 9.14,
    "O": 8.29, "P": 11.57, "Q": 11.43, "R": 15.14, "S": 29.29, "T": 15.57, "V": 15.57,
}

#: Type, off the same file. Everything on the sheet is Calibri 12; the subtotals are bold red
#: and the grand total bold black, which is how a reader tells a block's figures from the
#: container's at a glance.
_FONT_NAME = "Calibri"
_FONT_SIZE = 12
_RED = "FFFF0000"

#: Row heights, off the same file: a line wraps its description over two or three lines, and a
#: subtotal is a thin rule between blocks. Row 12 (the delivery warehouse) wraps too.
_LINE_HEIGHT = 35.1
_SUBTOTAL_HEIGHT = 15.95
_HEADER_BLOCK_ROW_HEIGHT = {12: 31.5}

#: Number formats, off the same file. The `[Red]` sections are the reference's own: a negative
#: volume or weight is a data fault, and it is meant to be visible as one.
_FMT_2DP = "0.00"
_FMT_2DP_RED = "0.00;[Red]0.00"
_FMT_MONEY = "#,##0.00"
_FMT_MONEY_RED = "#,##0.00;[Red]#,##0.00"
_FMT_SUBTOTAL_INT = "0_);[Red]\\(0\\)"
_FMT_SUBTOTAL_2DP = "0.00_);[Red]\\(0.00\\)"
_FMT_DATE = "[$-14409]dd/mm/yyyy;@"

#: The format each line column prints in. Anything not named here stays General, which is what
#: the reference does for the text columns.
_LINE_FORMATS = {
    "I": _FMT_2DP, "J": _FMT_2DP, "K": _FMT_2DP,
    "L": _FMT_2DP_RED, "M": _FMT_2DP_RED, "N": _FMT_2DP_RED,
    "O": _FMT_2DP_RED, "P": _FMT_2DP_RED, "Q": _FMT_2DP_RED,
    "T": _FMT_2DP, "U": _FMT_MONEY, "V": _FMT_MONEY,
}

#: What the CONTAINER's own total prints in. The money column differs from a block subtotal's
#: on purpose - that is what the reference does, and the two totals are meant to look
#: different from each other.
_GRAND_TOTAL_FORMATS = {
    "F": _FMT_SUBTOTAL_INT, "G": _FMT_SUBTOTAL_INT, "H": _FMT_SUBTOTAL_INT,
    "M": _FMT_SUBTOTAL_2DP, "P": _FMT_SUBTOTAL_2DP, "Q": _FMT_SUBTOTAL_2DP,
    "U": _FMT_SUBTOTAL_2DP,
}

#: What a subtotal prints in: whole pieces and cartons, two decimals for volume and weight,
#: money with its thousands separator.
_SUBTOTAL_FORMATS = {
    "F": _FMT_SUBTOTAL_INT, "G": _FMT_SUBTOTAL_INT, "H": _FMT_SUBTOTAL_INT,
    "M": _FMT_SUBTOTAL_2DP, "P": _FMT_SUBTOTAL_2DP, "Q": _FMT_SUBTOTAL_2DP,
    "U": _FMT_MONEY, "V": _FMT_MONEY,
}

#: The sheet's columns, A to V, in order - so a column letter and its 1-based index are one
#: lookup rather than two spellings that drift.
_LETTERS = [spec[0] for spec in _COLUMNS_SPEC]

#: The header-block keys that carry a DATE. Written as dates, not as the ISO strings the
#: payload holds, or Excel treats them as text.
_HEADER_DATE_KEYS = {"loading_date", "etd", "eta"}

#: Round 2 (captain, 3 Sep): the header block used to be forced centre/centre/wrap-text, and
#: the captain read that as a defect - "headers should be left aligned", "very ugly on first
#: open". The reference's own hand-typed file leaves every label at Excel's default and left
#: -aligns only these value rows (dates, container no, SO no); the rest are default too, off
#: `FSCU8103365.xlsx`. Read off the reference rather than guessed at, same as everything else
#: in this module - a rule that invented "always left" would drift the moment a fifth row in
#: the reference turned out to disagree.
_HEADER_LEFT_B_ROWS = {1, 2, 3, 4, 6}

#: The header block above the lines: (row, label, payload key on `header`).
_HEADER_BLOCK = [
    (1, "LOADING : ", "loading_date"),
    (2, "ETD :", "etd"),
    (3, "ETA : ", "eta"),
    (4, "CONTAINER :", "container_no"),
    (5, "SEAL NO : ", "seal_number"),
    (6, "SO :", "forwarder_order_ref"),
    (7, "CONSIGNEE :", "consignee"),
    (8, "SHIPPER :", "shipper"),
    (9, "CHINA AGENT : ", "china_agent"),
    (10, "FACTORY :", "factories"),
    (11, "FREE DAYS : ", "free_days"),
    (12, "DELIVERY WAREHOUSE : ", "delivery_warehouse"),
]


def _container_label(payload: dict) -> str:
    return payload.get("container_no") or payload.get("shipment_number") or ""


def _remarks(line: dict) -> Optional[str]:
    """The supplier's own note on the line, and only that (R20).

    Our own comparison against the loading plan used to be appended here in the same cell.
    On the document the factory reads back, that is us writing in their column.
    """
    return line.get("remarks") or None


def _blocks(payload: dict) -> list[dict]:
    """The sheet's blocks: one per factory, and one more for its MOCHA goods.

    MOCHA is invoiced separately, so its lines are listed under their own heading even
    though the same factory packed them - the footer's per-company apportionment has to be
    readable off the blocks above it, and a mixed block cannot be read that way.

    Split HERE rather than in `build`, because a factory is one supplier however its goods
    are invoiced: the screen that reads `build` shows one card per factory, and splitting it
    upstream would put the same factory on the page twice.
    """
    out: list[dict] = []
    for factory in payload.get("factories") or []:
        name = factory.get("supplier_name") or UNASSIGNED
        for company in COMPANIES:
            lines = [ln for ln in factory["lines"] if ln.get("company") == company]
            if not lines:
                continue
            out.append(
                {
                    "name": name if company == SORENTO else f"{name} ({company})",
                    "company": company,
                    "lines": lines,
                }
            )
    return out


def _f_or_none(value) -> Optional[float]:
    """A number Excel can add up, or nothing. Decimals arrive as strings on some paths."""
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def to_xlsx(payload: dict) -> bytes:
    """The container workbook in the FSCU layout, arithmetic included as FORMULAS.

    The derived cells (carton count, volume, line weights, line amount, every subtotal) are
    written as formulas rather than as computed numbers on purpose. The recipient corrects a
    quantity or a carton size in Excel - that is what the sheet is for - and a workbook of
    frozen numbers would keep printing the old totals underneath the corrected line.

    The type, the row heights and the number formats are the reference file's, asserted cell
    by cell in `tests/test_consolidated_packing_list_fidelity.py`. Where the reference
    disagrees with itself (`/10^6` against `/1000000`, `=PRODUCT(T,F)` against `=T*F`) one
    form is used throughout and the choice is recorded in that test.
    """
    import openpyxl
    from openpyxl.styles import Alignment, Font

    wb = openpyxl.Workbook()
    ws = wb.active or wb.create_sheet()
    ws.title = _SHEET_TITLE
    centred = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # The header block (rows 1-12) is NOT centred - see `_HEADER_LEFT_B_ROWS`.
    header_label = Alignment(vertical="center")
    header_label_wrap = Alignment(vertical="center", wrap_text=True)
    header_value = Alignment(vertical="center")
    header_value_left = Alignment(horizontal="left", vertical="center")

    def style(
        row: int, column: int, *, bold: bool = False, red: bool = False, fmt=None, alignment=None
    ):
        """Every cell the sheet writes goes through here, so nothing is left at Excel's
        default 11pt in the middle of a 12pt document."""
        cell = ws.cell(row=row, column=column)
        cell.font = Font(
            name=_FONT_NAME, size=_FONT_SIZE, bold=bold, color=_RED if red else None
        )
        cell.alignment = alignment if alignment is not None else centred
        if fmt:
            cell.number_format = fmt
        return cell

    def style_row(row: int, *, bold: bool = False, red: bool = False, formats=None):
        for index, letter in enumerate(_LETTERS, start=1):
            style(row, index, bold=bold, red=red, fmt=(formats or {}).get(letter))

    header = payload.get("header") or {}
    costs = payload.get("costs") or {}

    # ---- the header block ------------------------------------------------- #
    for row, label, key in _HEADER_BLOCK:
        label_alignment = header_label_wrap if row == 12 else header_label
        style(row, 1, alignment=label_alignment).value = label
        value = header.get(key)
        if key == "free_days" and value is not None:
            value = f"{_qty(value)} FREEDAYS"
        value_alignment = header_value_left if row in _HEADER_LEFT_B_ROWS else header_value
        cell = style(row, 2, alignment=value_alignment)
        # A date is written AS a date, in the reference's own format. Written as the ISO
        # string it arrives as, Excel left-aligns it as text: it cannot be sorted, cannot be
        # added to, and prints 2026-07-17 on a document everybody else reads dd/mm/yyyy.
        if key in _HEADER_DATE_KEYS and value:
            cell.value = _as_date(datetime.fromisoformat(str(value)))
            cell.number_format = _FMT_DATE
        else:
            cell.value = value
    for row, height in _HEADER_BLOCK_ROW_HEIGHT.items():
        ws.row_dimensions[row].height = height

    # ---- the two-row column header ---------------------------------------- #
    for letter, top, sub in _COLUMNS_SPEC:
        column = _LETTERS.index(letter) + 1
        if top is not None:
            style(_HEADER_ROW, column, bold=True).value = top
        if sub is not None:
            style(_SUBHEADER_ROW, column, bold=True).value = sub
        # A column with a row-16 label of its own keeps two cells; everything else is one
        # header merged down over both rows.
        if sub is None and top is not None:
            ws.merge_cells(f"{letter}{_HEADER_ROW}:{letter}{_SUBHEADER_ROW}")
    ws.merge_cells(f"I{_HEADER_ROW}:K{_HEADER_ROW}")
    ws.merge_cells(f"V{_HEADER_ROW}:V{_SUBHEADER_ROW}")

    # ---- one block per factory -------------------------------------------- #
    row = _FIRST_LINE_ROW
    number = 1
    subtotal_rows: list[tuple[str, int]] = []  # (company, subtotal row)

    for block in _blocks(payload):
        first_row = row
        for line in block["lines"]:
            qty = line.get("qty") or 0
            pcs = _f_or_none(line.get("pcs_per_carton"))
            length = _f_or_none(line.get("carton_length_cm"))
            width = _f_or_none(line.get("carton_width_cm"))
            height = _f_or_none(line.get("carton_height_cm"))
            net = _f_or_none(line.get("net_weight"))
            gross = _f_or_none(line.get("gross_weight"))
            price = _f_or_none(line.get("unit_cost"))
            cartons = line.get("cartons")

            style_row(row, formats=_LINE_FORMATS)
            ws.row_dimensions[row].height = _LINE_HEIGHT

            ws.cell(row=row, column=1, value=block["name"])
            ws.cell(row=row, column=2, value=number)
            ws.cell(row=row, column=3, value=line.get("product_code"))
            ws.cell(row=row, column=4, value=line.get("product_name"))
            ws.cell(row=row, column=5, value=line.get("material"))
            ws.cell(row=row, column=6, value=_qty(qty))
            ws.cell(row=row, column=7, value=pcs)
            # The carton count is DERIVED where the factory stated how many pieces go in a
            # box, and stated where it did not - the source file does exactly this, because
            # some factories give the carton count and no pack size.
            ws.cell(
                row=row,
                column=8,
                value=f"=F{row}/G{row}" if pcs else (cartons if cartons else None),
            )
            ws.cell(row=row, column=9, value=length)
            ws.cell(row=row, column=10, value=width)
            ws.cell(row=row, column=11, value=height)
            has_size = None not in (length, width, height)
            if has_size:
                ws.cell(row=row, column=12, value=f"=I{row}*J{row}*K{row}/10^6")
                ws.cell(row=row, column=13, value=f"=H{row}*L{row}")
            else:
                # Nothing measured the carton, so the volume is whatever the packing list
                # itself stated (or our catalogue worked out). Written flat: a formula with
                # no inputs would print a zero the container is then planned against.
                ws.cell(row=row, column=13, value=line.get("cbm"))
            ws.cell(row=row, column=14, value=net)
            ws.cell(row=row, column=15, value=gross)
            if net is not None:
                ws.cell(row=row, column=16, value=f"=N{row}*H{row}")
            if gross is not None:
                ws.cell(row=row, column=17, value=f"=O{row}*H{row}")
            ws.cell(row=row, column=18, value=line.get("brand"))
            ws.cell(row=row, column=19, value=_remarks(line))
            ws.cell(row=row, column=20, value=price)
            if price is not None:
                ws.cell(row=row, column=21, value=f"=T{row}*F{row}")

            number += 1
            row += 1

        last_row = row - 1
        # The block's own amount, merged down its rows: it is the figure the factory is paid,
        # and it sits beside its lines rather than under them. Red, like the reference, so it
        # reads as a total rather than as one more line figure.
        style(first_row, 22, red=True, fmt=_FMT_MONEY).value = (
            f"=SUM(U{first_row}:U{last_row})"
        )
        if last_row > first_row:
            ws.merge_cells(f"V{first_row}:V{last_row}")

        style_row(row, bold=True, red=True, formats=_SUBTOTAL_FORMATS)
        ws.row_dimensions[row].height = _SUBTOTAL_HEIGHT
        for column in ("F", "G", "H", "M", "P", "Q", "U"):
            ws[f"{column}{row}"].value = f"=SUM({column}{first_row}:{column}{last_row})"
        ws[f"V{row}"].value = f"=SUM(V{first_row})"
        subtotal_rows.append((block["company"], row))
        row += 1

    # ---- the rule, then the container's own totals ------------------------- #
    style_row(row)
    for column in range(1, len(_LETTERS) + 1):
        ws.cell(row=row, column=column, value="-")
    row += 1

    total_row = row
    style_row(total_row, bold=True, formats=_GRAND_TOTAL_FORMATS)
    for column in ("F", "G", "H", "M", "P", "Q", "U"):
        cell = ws[f"{column}{total_row}"]
        # Summed off the SUBTOTALS, not off the line range: a plain range would swallow the
        # subtotal rows sitting inside it and count every quantity twice.
        refs = ",".join(f"{column}{r}" for _company, r in subtotal_rows)
        cell.value = f"=SUM({refs})" if refs else 0
    row += 2

    # ---- the split, and what each company owes on it ----------------------- #
    # Clearance and China freight follow the VOLUME, insurance follows the AMOUNT: that is
    # how the forwarder bills them and how the source file apportions them.
    clearance = _f_or_none(costs.get("clearance_cost"))
    freight = _f_or_none(costs.get("china_freight_cost"))
    insurance_rate = _f_or_none(costs.get("insurance_rate"))

    company_rows: dict[str, int] = {}
    for company in COMPANIES:
        rows_for = [r for c, r in subtotal_rows if c == company]
        cbm_ref = ",".join(f"M{r}" for r in rows_for)
        amount_ref = ",".join(f"U{r}" for r in rows_for)
        style_row(row)
        style(row, 12).value = company
        style(row, 13, bold=True, fmt=_FMT_MONEY_RED).value = (
            f"=SUM({cbm_ref})" if cbm_ref else 0
        )
        if clearance is not None:
            style(row, 14, bold=True, fmt=_FMT_MONEY_RED).value = (
                f"=M{row}/M{total_row}*{clearance}"
            )
        if insurance_rate is not None:
            style(row, 15, bold=True, fmt=_FMT_MONEY_RED).value = (
                f"=U{row}/U{total_row}*{insurance_rate}"
            )
        if freight is not None:
            style(row, 16, bold=True, fmt=_FMT_MONEY_RED).value = (
                f"=M{row}/M{total_row}*{freight}"
            )
        style(row, 20).value = company
        style(row, 21, bold=True, fmt=_FMT_MONEY_RED).value = (
            f"=SUM({amount_ref})" if amount_ref else 0
        )
        company_rows[company] = row
        row += 1

    split_rows = [company_rows[c] for c in COMPANIES]
    # The volume and the amount always total; a cost only totals when it was typed. Summing
    # two blank cells prints 0, and a zero under CLEARANCE reads as a container that cost
    # nothing to clear rather than as one nobody has priced yet.
    totalled = ["M", "U"]
    if clearance is not None:
        totalled.append("N")
    if insurance_rate is not None:
        totalled.append("O")
    if freight is not None:
        totalled.append("P")
    style_row(row)
    for column in totalled:
        cell = style(row, _LETTERS.index(column) + 1, bold=True, fmt=_FMT_MONEY_RED)
        cell.value = "=" + "+".join(f"{column}{r}" for r in split_rows)
    row += 1

    style_row(row, bold=True)
    for column, label in (
        (13, "CBM"),
        (14, "CLEARANCE"),
        (15, "INSURANCE"),
        (16, "CHINA FREIGHT"),
        (21, "TOTAL AMOUNT"),
    ):
        ws.cell(row=row, column=column, value=label)
    # The three identifiers a forwarder quotes back at us, in the wording their own
    # paperwork uses.
    ws.cell(row=row, column=3, value=f"订单号:{header.get('forwarder_order_ref') or ''}")
    ws.cell(row=row + 1, column=3, value=f"柜号:{_container_label(payload)}")
    ws.cell(row=row + 2, column=3, value=f"封号:{header.get('seal_number') or ''}")

    for letter, width in _WIDTHS.items():
        ws.column_dimensions[letter].width = width
    # Round 2 (captain, 3 Sep): a frozen band read as "freezing at the bottom" on first open.
    # The reference freezes nothing, and neither does this sheet now.

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_filename(payload: dict) -> str:
    """`<container>-packing-list.xlsx`, with anything a filesystem would argue about removed."""
    stem = _container_label(payload) or payload["shipment_id"]
    stem = re.sub(r"[^A-Za-z0-9._-]", "", str(stem)) or str(payload["shipment_id"])
    return f"{stem}-packing-list.xlsx"
