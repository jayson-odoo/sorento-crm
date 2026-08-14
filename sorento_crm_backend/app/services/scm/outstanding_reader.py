"""Read an outstanding SO / PO extract into `outstanding_diff.Line` objects.

Column headers are resolved through `import_field_alias` rather than a hardcoded candidate
tuple, so onboarding a client whose export says `S/O NO` instead of `SO NO`, or `型号`
instead of `ITEM CODE`, is an INSERT rather than a code change. Header order does not
matter and unmapped columns are reported rather than ignored, because a silently dropped
column is how an import produces confidently wrong numbers.

What this module deliberately does NOT do: resolve products, warehouses or customers, or
touch the database. It turns a sheet into rows plus a list of complaints. Resolution and
the diff happen above it, which keeps the parsing testable against a file alone.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any, Optional

from app.services.import_alias_service import AliasResolver, normalize_header
from app.services.scm.outstanding_diff import Line

SO = "outstanding_so"
PO = "outstanding_po"

# Without these a row cannot be placed on the plan at all, so a row missing one is rejected
# rather than guessed at. Everything else is optional and merely reduces what we can show.
_REQUIRED = {
    SO: ("so_number", "item_code", "qty_outstanding"),
    PO: ("po_number", "item_code", "qty_ordered"),
}

# Header-level requirements: a file with no delivery-date column at all is not an
# outstanding-orders extract, it is a different report, and importing it would wipe every
# date in scope. Checked once against the header row, not per line.
_REQUIRED_COLUMNS = {
    SO: ("so_number", "item_code", "qty_outstanding", "required_date"),
    # `expected_date` is required for exactly the reason above. It was harmless while the PO
    # path was read-only; now that it WRITES, a file with no ETA column blanks every arrival
    # date in scope and the plan loses its time axis (coverage timelines, the arrival
    # calendar, every "when does this land" answer) while on-order still looks right.
    PO: ("po_number", "item_code", "qty_ordered", "expected_date"),
}


@dataclass
class RowProblem:
    row_number: int
    reason: str
    value: str = ""


@dataclass
class ReadResult:
    doc_type: str
    lines: list[Line] = field(default_factory=list)
    problems: list[RowProblem] = field(default_factory=list)
    unmapped_headers: list[str] = field(default_factory=list)
    missing_columns: list[str] = field(default_factory=list)
    total_rows: int = 0
    #: Rows carrying neither an item code nor a quantity. Captions and spacers rather than
    #: failed lines, counted so a big number is visible instead of being silently dropped.
    layout_rows: int = 0
    # Per-row fields the DIFF does not care about but the WRITE does: counterparty code,
    # unit cost, currency. Kept in a side map keyed by `Line.row_ref` rather than added to
    # `Line`, because `outstanding_diff` is deliberately document-agnostic and adding
    # purchase-only columns to its value type would leak one document's shape into both.
    extras: dict[str, dict] = field(default_factory=dict)

    @property
    def ok(self) -> bool:
        """A file is readable when its header carries every column the plan needs.

        Individual bad rows do NOT make a file unusable: an extract of 4,000 lines with 3
        unresolvable rows should import 3,997 and report 3, not refuse everything.
        """
        return not self.missing_columns


def _to_float(value: Any) -> Optional[float]:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    # Extracts arrive with thousands separators and trailing units.
    cleaned = str(value).strip().replace(",", "").replace(" ", "")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _to_date(value: Any) -> Optional[date]:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    # Day-first before month-first: these are Malaysian exports, where 03/08/2026 is 3 Aug.
    # Guessing the other way silently moves a delivery five months.
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d %b %Y", "%d-%b-%Y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


# The two spreadsheet containers, told apart by their MAGIC BYTES rather than by the
# filename. A legacy `.xls` is an OLE2 compound document; `.xlsx`/`.xlsm` are zips. Sniffing
# the content is the honest check: an AutoCount export saved with the wrong extension, or a
# file renamed on the way through email, still reads as what it actually is - and the
# alternative is refusing a workbook we can plainly open.
_ZIP_MAGIC = b"PK\x03\x04"
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


def _xls_rows(file_data: bytes):
    """Rows of a legacy BIFF `.xls`, shaped like openpyxl's `values_only=True`.

    `xlrd` 2.x reads `.xls` and nothing else, which is exactly the half we need: openpyxl
    covers the zip formats and refuses BIFF outright.

    Dates arrive as floats (an Excel serial) with a cell TYPE that says so, and the type is
    the only way to tell 43832 the date from 43832 the quantity - so the conversion happens
    here, where the type is still in scope, rather than downstream where it is guesswork.
    """
    import xlrd

    book = xlrd.open_workbook(file_contents=file_data)
    sheet = book.sheet_by_index(0)
    for r in range(sheet.nrows):
        out = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            if cell.ctype == xlrd.XL_CELL_DATE:
                y, mo, d, hh, mm, ss = xlrd.xldate_as_tuple(cell.value, book.datemode)
                out.append(date(y, mo, d) if not (hh or mm or ss) else datetime(y, mo, d, hh, mm, ss))
            elif cell.ctype == xlrd.XL_CELL_EMPTY:
                out.append(None)
            else:
                out.append(cell.value)
        yield tuple(out)


def sheet_rows(file_data: bytes):
    """The first sheet's rows, whatever container the bytes are in.

    One entry point so every caller reads the same formats: the route that accepts an upload
    and the reader that parses it cannot disagree about what is supported.
    """
    if file_data[:8] == _OLE2_MAGIC:
        return _xls_rows(file_data)

    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(file_data), data_only=True, read_only=True)
    sheet = wb.active
    if sheet is None:
        raise ValueError("the workbook has no sheet")
    return sheet.iter_rows(values_only=True)


def read_workbook(file_data: bytes, doc_type: str, resolver: AliasResolver) -> ReadResult:
    """Parse the first sheet of an uploaded workbook into lines plus complaints."""
    result = ReadResult(doc_type=doc_type)
    try:
        rows = sheet_rows(file_data)
    except Exception as exc:  # noqa: BLE001 - surfaced to the user, never swallowed
        result.problems.append(RowProblem(0, f"could not read the workbook: {exc}"))
        result.missing_columns = list(_REQUIRED_COLUMNS[doc_type])
        return result
    try:
        header = next(rows)
    except StopIteration:
        result.missing_columns = list(_REQUIRED_COLUMNS[doc_type])
        return result

    # header index -> canonical field
    col_field: dict[int, str] = {}
    for idx, cell in enumerate(header):
        f = resolver.field_for_header(cell)
        if f:
            col_field[idx] = f
        elif normalize_header(cell):
            result.unmapped_headers.append(_clean(cell))

    present = set(col_field.values())
    result.missing_columns = [c for c in _REQUIRED_COLUMNS[doc_type] if c not in present]
    if result.missing_columns:
        # No point reading 4,000 rows against a header that cannot answer the question.
        return result

    qty_field = "qty_outstanding" if doc_type == SO else "qty_ordered"
    doc_field = "so_number" if doc_type == SO else "po_number"
    date_field = "required_date" if doc_type == SO else "expected_date"
    label_field = "customer_name" if doc_type == SO else "supplier_name"
    order_date_field = "so_date" if doc_type == SO else "po_date"

    # (document, item, location, date) -> the row that first stated it. The same line stated
    # twice in one file is a spreadsheet accident, and it is invisible downstream: the diff
    # pairs by content, so the first copy reads as `unchanged` and the second as `added`,
    # doubling the supply and then sitting perfectly stable so no later upload surfaces it.
    # Both rows are still carried - what the write should do with them is arguable, that a
    # human is told is not.
    first_seen: dict[tuple, int] = {}

    for row_number, row in enumerate(rows, start=2):
        if not any(c is not None and str(c).strip() for c in row):
            continue
        result.total_rows += 1
        rec = {col_field[i]: v for i, v in enumerate(row) if i in col_field}

        doc = _clean(rec.get(doc_field))
        item = _clean(rec.get(item_key := "item_code"))
        qty = _to_float(rec.get(qty_field))

        # A row with no item AND no quantity states nothing, so it is layout, not an error.
        # The AutoCount detail listing puts an "ITEM PACKAGE :" caption above each package
        # and leaves the item and quantity cells empty: 9,144 of them in the client's own
        # 81,361-row export. Reporting each as a failed row buries the handful of rows that
        # really did fail under nine thousand that were never lines at all. Counted, so the
        # skip is visible rather than silent. A row that names a quantity but no item is
        # still a problem - that is the totals row, or a genuine gap.
        if not item and (qty is None or qty == 0):
            result.layout_rows += 1
            continue

        missing = [f for f, v in ((doc_field, doc), (item_key, item)) if not v]
        if qty is None:
            missing.append(qty_field)
        if missing:
            result.problems.append(RowProblem(
                row_number, f"missing {', '.join(missing)}", value=doc or item))
            continue

        # Three real file shapes, resolved in a fixed order of preference rather than by
        # column position. Getting this wrong is not cosmetic: reading ORDERED as
        # OUTSTANDING inflates committed demand on every partly-delivered line, and inflates
        # every buy computed from it, on a file that imports perfectly cleanly.
        #
        #   1. the remaining quantity stated outright  -> use it
        #   2. quantity plus what has already gone out -> the difference
        #   3. a single netted quantity column         -> use it as-is
        #
        # The purchase-order path has always done (2) with `qty_received`; this is the same
        # rule, named, and extended to the sales-order side which needed it for the
        # AutoCount detail listing (Qty + Transfered Qty + Remaining Qty).
        remaining = _to_float(rec.get("qty_remaining"))
        fulfilled_field = "qty_received" if doc_type == PO else "qty_delivered"
        fulfilled = _to_float(rec.get(fulfilled_field))
        if remaining is not None:
            qty = remaining
        elif fulfilled is not None:
            qty = qty - fulfilled

        if qty <= 0:
            # Nothing outstanding is not an error - it is a line that belongs in the
            # "closed" side of the diff, reached by its absence rather than by a zero.
            continue

        raw_date = rec.get(date_field)
        when = _to_date(raw_date)
        if when is None and _clean(raw_date):
            result.problems.append(RowProblem(
                row_number, f"could not read the date in {date_field}",
                value=_clean(raw_date)))

        location = _clean(rec.get("stock_location"))
        # Codes are compared uppercase because every consumer normalises with `upper()`.
        dup_key = (doc.upper(), item.upper(), location.upper(), when)
        if dup_key in first_seen:
            result.problems.append(RowProblem(
                row_number,
                f"the same line is stated twice: {doc} / {item} at the same location and "
                f"date already appears on row {first_seen[dup_key]}",
                value=doc))
        else:
            first_seen[dup_key] = row_number

        result.lines.append(Line(
            doc_number=doc,
            item_code=item,
            location=location,
            qty=qty,
            required_date=when,
            row_ref=str(row_number),
            label=_clean(rec.get(label_field)),
        ))
        # An absent cost stays absent. A zero would read as free goods, and the cash
        # co-pilot ranks on this column. `order_date` is the DOCUMENT's own date (PO DATE /
        # SO DATE): carried per row because that is the shape of the file, folded onto the
        # header by the write path. Without it `scm.receipt_lead_v` has no `po.issue_date` to
        # measure lead days against, so an imported order contributes no lead-time evidence.
        # `order_type` is the same shape and OPTIONAL: no export carries it today, and when
        # one does it is what the sales book's demand class is stamped from, so an order
        # this upload creates can be classified instead of reported as unclassifiable.
        result.extras[str(row_number)] = {
            "party_code": _clean(rec.get("debtor_code" if doc_type == SO
                                         else "creditor_code")),
            "unit_cost": _to_float(rec.get("unit_cost")),
            "currency": _clean(rec.get("currency")) or None,
            "order_date": _to_date(rec.get(order_date_field)),
            "order_type": _clean(rec.get("order_type")) or None,
        }

    return result
