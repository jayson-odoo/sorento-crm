"""Order service for business logic."""
import logging
import uuid
import time
from typing import Any, Optional
from io import BytesIO
from datetime import datetime

from sqlalchemy.orm import Session, joinedload
from sqlalchemy import or_, and_, func, exists, false
from decimal import Decimal
from app.models.order import Order, OrderStatus, Customer, OrderLine, Transporter
from app.models.product import Product
from app.models.inventory import Warehouse
from app.schemas.order import (
    OrderCreate, OrderUpdate, CustomerCreate, CustomerUpdate,
    OrderStatusCreate, OrderStatusUpdate,
    OrderLineCreate, OrderLineUpdate,
)
from app.services.error_handler import handle_not_found, handle_conflict
from app.services.import_log_service import ImportLogService
from app.services.calendar_service import CalendarService
from app.services.identifier_resolver import resolve_identifier
from app.services.company_scope import (
    build_company_predicate,
    get_company_scope,
    stamp_lookup_companies,
)
from app.services.embedding_change_listener import (
    suppress_embedding_events,
    bulk_enqueue_embedding_events,
)
from app.services.fuzzy_resolver import resolve_via_embedding_then_ilike
from app.services.entity_resolver import (
    EntityFilterBuckets,
    resolve_entities_to_filters,
)

logger = logging.getLogger(__name__)


# Date-axis relaxation cap (§3.4). When a customer-scoped delivery query returns
# zero DOs on the asked date, surface at most this many nearest DISTINCT delivery
# dates (either side of the asked date) that DO have a DO for the same scope.
DATE_RELAX_LIMIT = 3

# Canonical "handed to the customer" status codes. Mirrors
# ComplaintFulfilmentService.DELIVERED_STATUS_CODES - an order counts as
# delivered ONLY when its status is one of these AND actual_delivery_date is set
# (a Status alone without a date, or a date under a non-delivered status like
# "New Order", does NOT count as delivered). Drives the `order_status`
# outstanding/delivered bucket filter.
DELIVERED_STATUS_CODES = ("delivered", "completed")


def _delivered_status_ids(db) -> list:
    """Ids of the order statuses that count as delivered (see DELIVERED_STATUS_CODES)."""
    return [
        s.id
        for s in db.query(OrderStatus.id)
        .filter(func.lower(OrderStatus.status_code).in_(DELIVERED_STATUS_CODES))
        .all()
    ]


def _delivered_clause(delivered_status_ids: list):
    """The canonical delivered predicate: delivered/completed status AND a set date.

    With no delivered statuses configured nothing is delivered (``false()``), so a
    "delivered" bucket returns no rows and a summary counts zero delivered.
    """
    if not delivered_status_ids:
        return false()
    return and_(
        Order.order_status_id.in_(delivered_status_ids),
        Order.actual_delivery_date.is_not(None),
    )


def _outstanding_clause(delivered_status_ids: list):
    """Null-safe negation of ``_delivered_clause``: NULL status, non-delivered
    status, or no date. ``~_delivered_clause`` is NOT this - for a NULL
    ``order_status_id`` it evaluates to SQL NULL and drops the row from a
    FILTER (WHERE ...) aggregate while ``total - delivered`` still counts it
    (cross-model review, 2026-08-25)."""
    if not delivered_status_ids:
        return None  # everything is outstanding; no filter
    return or_(
        Order.order_status_id.is_(None),
        Order.order_status_id.not_in(delivered_status_ids),
        Order.actual_delivery_date.is_(None),
    )


def _order_status_bucket_filter(db, order_status: Optional[str]):
    """Translate the ``order_status`` bucket ('outstanding' | 'delivered') into a
    filter clause, or None when the bucket is absent / unrecognised / a no-op.

    Shared by ``list_orders`` and ``list_orders_by_product`` so the two order
    tools cannot disagree on what "delivered" means (4,832 orders carried a
    delivery date under a non-delivered status on 2026-08-24 - a date-only
    predicate and this one are NOT the same filter).
    """
    bucket = (order_status or "").strip().lower()
    if bucket not in ("outstanding", "delivered"):
        return None
    delivered_status_ids = _delivered_status_ids(db)
    if bucket == "delivered":
        return _delivered_clause(delivered_status_ids)
    return _outstanding_clause(delivered_status_ids)


def _plain_number(v):
    """Decimal/float -> int when integral, else float. None stays None."""
    if v is None:
        return None
    try:
        d = Decimal(str(v))
    except Exception:
        return v
    if d == d.to_integral_value():
        return int(d)
    return float(d)


GROUPS_CEILING = 500


def stamp_order_summary(db, payload: dict, filtered_q, *, product_ids=None) -> None:
    """Stamp ``payload["summary"]`` - measures over the WHOLE filter, never the page.

    ``filtered_q`` is the list query AFTER every filter and BEFORE offset/limit
    (ordering is stripped here). External/agent callers see at most 20 rows of
    ``data``; a consumer that summed those rows would state a wrong total for
    exactly the accounts that ask (81 customer x product pairs exceed 20 DOs).
    So the numbers are computed in SQL over the same predicate the rows came from.

    Shape (absent-not-null everywhere, like ``lookup_companies``)::

        {"scope": "filter", "row_count": 35,
         "order_count": 35, "delivered_count": 33, "pending_count": 2,
         "customers": ["ECO WORLD SDN BHD"], "customer_count": 1,
         "delivered_from": "2026-03-02", "delivered_to": "2026-07-15",
         "products": [{"product_code": "SRTWC8605",
                       "delivered_quantity": 48, "pending_quantity": 12}],
         "groups": [{"customer": "ECO WORLD SDN BHD", "product_code": "SRTWC8605",
                     "order_count": 3, "delivered_quantity": 48, "pending_quantity": 12}]}

    ``groups`` is the unit the question is really about - WHICH customer took HOW
    MUCH of WHICH product (captain, 2026-08-25). One row per (customer, product)
    over the same filter, sorted customer then product, every row (no top-N; a
    safety ceiling of GROUPS_CEILING rows with ``groups_truncated`` = how many
    were cut, so a runaway filter cannot blow the payload). ``products`` stays
    as the per-product totals - the header line above the groups; the consumer
    never sums rows itself.

    Computed ONLY when the caller asked for it (``include_summary=true`` on the
    endpoint - n8n sends it when the parser saw a quantity question: "how many",
    "have take", "have sales"). "Any delivery to Hanlim for SRTWC286" wants the
    DOs, not a headline, and pays no extra query. ``products`` (per product code) and
    ``groups`` (customer x product) are ALWAYS emitted with an asked summary
    (amendment 4); a product narrower restricts them to that product. Nothing is
    ever summed across products. ``delivered_from/to`` are present only when something was
    delivered. Delivered = the canonical predicate (``_delivered_clause``),
    pending = its negation.

    Cost when asked: one aggregate over the filtered order ids (same class as
    the ``COUNT(*)`` the list already runs for ``pagination``), one distinct-name
    pass, and one grouped SUM over the lines when product-scoped.

    Best-effort: any failure warns and leaves the payload untouched, so a list
    answer never dies because its headline could not be computed. An empty
    result gets no summary at all.
    """
    pids = [str(p) for p in (product_ids or []) if p]
    try:
        # COMPANY SCOPE, EXPLICITLY. The session's do_orm_execute listener scopes ORM
        # ENTITIES via with_loader_criteria; every query below is column-only
        # (aggregates over `select_from(Order)` / `select_from(OrderLine)`), and a
        # `.subquery()` of an entity query loses the criteria too. Measured
        # 2026-08-25 (test_qs_c9): under a Sorento-only scope the ids subquery held a
        # Mocha order and the customer list named its debtor. So the same predicate
        # the listener would inject is ANDed in here by hand, per entity. Harmless
        # where the listener also fires (duplicate AND); load-bearing where it does not.
        _scope = get_company_scope(db)
        _p_order = build_company_predicate(Order, _scope)
        _p_line = build_company_predicate(OrderLine, _scope)
        _p_cust = build_company_predicate(Customer, _scope)
        _scoped = lambda q, pred: q.filter(pred) if pred is not None else q  # noqa: E731

        ids_sq = (
            _scoped(filtered_q.order_by(None), _p_order)
            .with_entities(Order.id)
            .distinct()
            .subquery()
        )
        in_ids = Order.id.in_(db.query(ids_sq.c.id))
        _dsids = _delivered_status_ids(db)
        delivered = _delivered_clause(_dsids)
        pending = _outstanding_clause(_dsids)
        if pending is None:
            from sqlalchemy import true as _true
            pending = _true()

        total, n_delivered, d_from, d_to = _scoped(
            db.query(
                func.count(Order.id),
                func.count(Order.id).filter(delivered),
                func.min(Order.actual_delivery_date).filter(delivered),
                func.max(Order.actual_delivery_date).filter(delivered),
            ).select_from(Order).filter(in_ids),
            _p_order,
        ).one()
        total = int(total or 0)
        if total == 0:
            return
        n_delivered = int(n_delivered or 0)

        # NULLIF: a blank/whitespace debtor_name must fall back to the customer name,
        # not win the coalesce and then be filtered out (cross-model review, 2026-08-25).
        name_col = func.coalesce(
            func.nullif(func.btrim(Order.debtor_name), ""),
            func.btrim(Customer.customer_name),
        )
        _cust_on = Customer.id == Order.customer_id
        if _p_cust is not None:
            _cust_on = and_(_cust_on, _p_cust)  # scope the joined customer in the ON clause
        names_q = _scoped(
            db.query(name_col)
            .select_from(Order)
            .outerjoin(Customer, _cust_on)
            .filter(in_ids)
            .filter(name_col.is_not(None), name_col != ""),
            _p_order,
        )
        # The COUNT is its own query over every distinct name; the LIST is capped.
        # Counting the capped list would say "50 customers" for 200 (reviewer F-1).
        customer_count = int(names_q.with_entities(func.count(func.distinct(name_col))).scalar() or 0)
        names = [r[0] for r in names_q.distinct().order_by(name_col.asc()).limit(3).all()]

        summary: dict[str, Any] = {
            "scope": "filter",
            "row_count": total,
            "order_count": total,
            "delivered_count": n_delivered,
            "pending_count": total - n_delivered,
            "customers": names,
            "customer_count": customer_count,
        }
        if n_delivered and d_from is not None and d_to is not None:
            summary["delivered_from"] = d_from.isoformat()
            summary["delivered_to"] = d_to.isoformat()

        # Amendment 4 (captain, 2026-08-25): per-product totals and customer x product
        # groups are ALWAYS emitted with an asked summary - a breakdown by product is
        # meaningful under any filter (it was the cross-product grand total that was
        # not). When the caller narrowed by product, the lines are restricted to it.
        _line_scope = OrderLine.product_id.in_(pids) if pids else None
        if True:
            prod_q = (
                db.query(
                    Product.product_code,
                    func.count(func.distinct(Order.id)),
                    func.sum(OrderLine.quantity).filter(delivered),
                    func.sum(OrderLine.quantity).filter(pending),
                    func.min(Order.actual_delivery_date).filter(delivered),
                    func.max(Order.actual_delivery_date).filter(delivered),
                    # exact per-product customer count - the presenter must not derive
                    # it from a groups[] slice that the ceiling may have cut
                    func.count(func.distinct(name_col)),
                )
                .select_from(OrderLine)
                .join(Order, Order.id == OrderLine.order_id)
                .join(Product, Product.id == OrderLine.product_id)
                .outerjoin(Customer, _cust_on)
                .filter(in_ids)
            )
            if _line_scope is not None:
                prod_q = prod_q.filter(_line_scope)
            prod_q = _scoped(_scoped(prod_q, _p_order), _p_line)
            prod_rows = (
                prod_q.group_by(Product.product_code)
                .order_by(Product.product_code.asc())
                .limit(GROUPS_CEILING + 1)
                .all()
            )
            summary["products"] = [
                {
                    "product_code": code,
                    "order_count": int(oc or 0),
                    "customer_count": int(cc or 0),
                    "delivered_quantity": _plain_number(dq) or 0,
                    "pending_quantity": _plain_number(pq) or 0,
                    **({"delivered_from": pf.isoformat(), "delivered_to": pt.isoformat()}
                       if (pf is not None and pt is not None) else {}),
                }
                for code, oc, dq, pq, pf, pt, cc in prod_rows[:GROUPS_CEILING]
            ]
            if len(prod_rows) > GROUPS_CEILING:
                summary["products_truncated"] = True

            # customer x product: the row the question is about. Same joins, same
            # scope predicates, same delivered/pending clauses; grouped one level
            # finer. Customer name resolved the same way as `customers` above.
            grp_q = (
                db.query(
                    name_col,
                    Product.product_code,
                    func.count(func.distinct(Order.id)),
                    func.sum(OrderLine.quantity).filter(delivered),
                    func.sum(OrderLine.quantity).filter(pending),
                    func.min(Order.actual_delivery_date).filter(delivered),
                    func.max(Order.actual_delivery_date).filter(delivered),
                )
                .select_from(OrderLine)
                .join(Order, Order.id == OrderLine.order_id)
                .join(Product, Product.id == OrderLine.product_id)
                .outerjoin(Customer, _cust_on)
                .filter(in_ids)
            )
            if _line_scope is not None:
                grp_q = grp_q.filter(_line_scope)
            grp_q = _scoped(_scoped(grp_q, _p_order), _p_line)
            grp_rows = (
                grp_q.group_by(name_col, Product.product_code)
                .order_by(name_col.asc().nulls_last(), Product.product_code.asc())
                .limit(GROUPS_CEILING + 1)
                .all()
            )
            summary["groups"] = [
                {
                    "customer": (cname or "").strip() or None,
                    "product_code": code,
                    "order_count": int(oc or 0),
                    "delivered_quantity": _plain_number(dq) or 0,
                    "pending_quantity": _plain_number(pq) or 0,
                    **({"delivered_from": gf.isoformat(), "delivered_to": gt.isoformat()}
                       if (gf is not None and gt is not None) else {}),
                }
                for cname, code, oc, dq, pq, gf, gt in grp_rows[:GROUPS_CEILING]
            ]
            if len(grp_rows) > GROUPS_CEILING:
                # Only the ceiling tells us "more" - the exact remainder would be
                # another COUNT over the groups; say that some were cut, honestly.
                summary["groups_truncated"] = True
        payload["summary"] = summary
    except Exception as exc:  # pragma: no cover - best-effort by contract
        logger.warning("stamp_order_summary skipped: %s", exc)


class OrderService:
    """Service for order operations."""
    
    def __init__(self, db: Session):
        self.db = db
    
    def list_orders(
        self,
        page: int = 1,
        limit: int = 50,
        query: Optional[str] = None,
        customer_query: Optional[str] = None,
        product_query: Optional[str] = None,
        transporter_query: Optional[str] = None,
        customer_id: Optional[str] = None,
        order_status_id: Optional[str] = None,
        order_status: Optional[str] = None,
        has_order_lines: Optional[str] = None,
        has_actual_delivery_date: Optional[str] = None,
        order_date_from: Optional[datetime] = None,
        order_date_to: Optional[datetime] = None,
        actual_delivery_date_from: Optional[datetime] = None,
        actual_delivery_date_to: Optional[datetime] = None,
        sort_field: str = "created_at",
        sort_dir: str = "asc",
        advanced_filter_clause: Optional[Any] = None,
        entities: Optional[list[str]] = None,
        order_ids: Optional[list[str]] = None,
        customer_ids: Optional[list[str]] = None,
        product_ids: Optional[list[str]] = None,
        transporter_ids: Optional[list[str]] = None,
        ids_only: bool = False,
        include_summary: bool = False,
    ):
        """List orders with filtering and pagination.

        ``include_summary`` stamps ``payload["summary"]`` (filter-wide measures,
        see ``stamp_order_summary``); off by default so a plain list pays nothing.

        When ``ids_only`` is True the method skips counting, pagination, and
        eager-loading and returns the full ordered list of order ids
        (``list[str]``) for the active filter+sort. This is the exact same query
        object the paginated list builds, so the prev/next ``neighbours`` pager
        and the grid can never drift. Any early-return shortcut (an identifier
        that resolves to nothing) yields ``[]`` in this mode.

        `entities` is the single-bag entity filter for AI/MCP callers: pass a list of
        free-text strings (customer names, product codes, transporter labels, order
        numbers). The server resolves each via the entity_resolver and applies the
        matched filters internally. The resolution is echoed back as
        `_resolved_entities` on the response so the agent can show the user what
        was actually matched. Typed params (`customer_query`, `product_query`, ...)
        remain for direct callers and are AND'd with `entities`.
        """
        # Capture typed UUID params into locals up-front so later code that
        # reuses the names (e.g. `customer_ids = resolve_identifier(...)`)
        # cannot shadow them.
        _order_uuid_filter = list(order_ids) if order_ids else None
        _customer_uuid_filter = list(customer_ids) if customer_ids else None
        _product_uuid_filter = list(product_ids) if product_ids else None
        _transporter_uuid_filter = list(transporter_ids) if transporter_ids else None

        # Date-axis relaxation (§3.4) bookkeeping. `_customer_scoped` gates the
        # nearest-date probe so we never dump every customer's dates on an empty
        # result; `_asked_delivery_date` is the date we measure distance from.
        _customer_scoped = bool(_customer_uuid_filter) or bool(
            (customer_query or "").strip()
        )
        _asked_delivery_date = actual_delivery_date_from or actual_delivery_date_to

        q = self.db.query(Order).filter(Order.deleted_at.is_(None))

        entity_buckets: Optional[EntityFilterBuckets] = None
        if entities:
            entity_buckets = resolve_entities_to_filters(
                self.db,
                entities,
                allowed_entity_types=(
                    "product", "customer", "customer_order", "transporter",
                    "inbound_shipment", "spo_allocation", "grn", "promotion",
                    "attachment", "form",
                ),
            )
            # Caller passed `entities` but nothing resolved above the RAG
            # similarity threshold → return empty rather than fall through and
            # unintentionally list every order. Echo the unresolved inputs so the
            # agent surfaces "no match" to the user.
            if not entity_buckets.has_resolved_filter:
                if ids_only:
                    return []
                payload = {
                    "data": [],
                    "pagination": {"total": 0, "page": page, "limit": limit},
                    "empty": True,
                    "resolved_entities": entity_buckets.as_echo(),
                }
                stamp_lookup_companies(
                    self.db, payload, [], product_ids=_product_uuid_filter
                )
                return payload

        filters = []

        if _order_uuid_filter:
            filters.append(Order.id.in_(_order_uuid_filter))

        if _customer_uuid_filter:
            # Match Order.customer_id IN (...) OR legacy debtor_name match.
            # Backfill 207 should make the OR branch redundant for new data, but
            # historical rows with NULL customer_id stay reachable.
            from sqlalchemy import func as _func
            customer_names_subq = (
                self.db.query(Customer.customer_name)
                .filter(Customer.id.in_(_customer_uuid_filter))
                .subquery()
            )
            filters.append(
                or_(
                    Order.customer_id.in_(_customer_uuid_filter),
                    _func.lower(_func.btrim(Order.debtor_name)).in_(
                        self.db.query(_func.lower(_func.btrim(customer_names_subq.c.customer_name)))
                    ),
                )
            )

        if _product_uuid_filter:
            filters.append(
                Order.lines.any(OrderLine.product_id.in_(_product_uuid_filter))
            )

        if _transporter_uuid_filter:
            from sqlalchemy import func as _func
            transporter_names_subq = (
                self.db.query(Transporter.name)
                .filter(Transporter.id.in_(_transporter_uuid_filter))
                .subquery()
            )
            filters.append(
                or_(
                    Order.transporter_id.in_(_transporter_uuid_filter),
                    _func.lower(_func.btrim(Order.transporter)).in_(
                        self.db.query(_func.lower(_func.btrim(transporter_names_subq.c.name)))
                    ),
                )
            )

        customer_ids = resolve_identifier(
            self.db,
            customer_id,
            Customer,
            code_fields=("customer_code", "customer_name"),
        )
        if customer_ids is not None:
            if not customer_ids:
                if ids_only:
                    return []
                payload = {
                    "data": [],
                    "pagination": {"total": 0, "page": page, "limit": limit},
                    "empty": True,
                }
                stamp_lookup_companies(
                    self.db, payload, [], product_ids=_product_uuid_filter
                )
                return payload
            filters.append(Order.customer_id.in_(customer_ids))
            _customer_scoped = True

        status_ids = resolve_identifier(
            self.db,
            order_status_id,
            OrderStatus,
            code_fields=("status_code", "status_name"),
        )
        if status_ids is not None:
            if not status_ids:
                if ids_only:
                    return []
                payload = {
                    "data": [],
                    "pagination": {"total": 0, "page": page, "limit": limit},
                    "empty": True,
                }
                stamp_lookup_companies(
                    self.db, payload, [], product_ids=_product_uuid_filter
                )
                return payload
            filters.append(Order.order_status_id.in_(status_ids))

        # Semantic outstanding/delivered bucket (CRM-003). Distinct from the exact
        # `order_status_id` match above and AND'd with it. Delivered predicate is the
        # canonical one: a delivered/completed status code AND a set actual_delivery_date
        # (so a "New Order" with a stray delivery date reads as outstanding, not
        # delivered). Outstanding = the null-safe negation.
        bucket_clause = _order_status_bucket_filter(self.db, order_status)
        if bucket_clause is not None:
            filters.append(bucket_clause)

        hol = (has_order_lines or "").strip().lower()
        if hol == "yes":
            filters.append(exists().where(OrderLine.order_id == Order.id))
        elif hol == "no":
            filters.append(~exists().where(OrderLine.order_id == Order.id))

        hadd = (has_actual_delivery_date or "").strip().lower()
        if hadd == "yes":
            filters.append(Order.actual_delivery_date.is_not(None))
        elif hadd == "no":
            filters.append(Order.actual_delivery_date.is_(None))

        # Date-range clauses are tracked apart from entity/scope clauses so the
        # date-axis relaxation (§3.4) can rebuild a customer-scoped query with the
        # date constraint dropped when the asked date yields zero DOs.
        date_range_filters: list[Any] = []
        if order_date_from is not None:
            date_range_filters.append(Order.order_date >= order_date_from)

        if order_date_to is not None:
            date_range_filters.append(Order.order_date <= order_date_to)

        if actual_delivery_date_from is not None:
            date_range_filters.append(Order.actual_delivery_date >= actual_delivery_date_from)

        if actual_delivery_date_to is not None:
            date_range_filters.append(Order.actual_delivery_date <= actual_delivery_date_to)

        query_filter = None
        if query and (query := (query or "").strip()):
            term = f"%{query}%"
            # Split into UNION of matching id sets. The original OR-with-EXISTS-on-customers
            # forces a Seq Scan on orders even with pg_trgm indexes; pulling the cross-table
            # match into its own subquery lets each branch use its trigram index independently.
            direct_ids = self.db.query(Order.id).filter(
                or_(
                    Order.order_number.ilike(term),
                    Order.debtor_name.ilike(term),
                    Order.debtor_code.ilike(term),
                    # Legacy free-text transporter column kept for back-compat
                    # rows that pre-date the transporters FK.
                    Order.transporter.ilike(term),
                )
            )
            customer_ids = (
                self.db.query(Order.id)
                .join(Customer, Order.customer_id == Customer.id)
                .filter(
                    or_(
                        Customer.customer_name.ilike(term),
                        Customer.customer_code.ilike(term),
                    )
                )
            )
            # Canonical transporter join: matches Transporter.code / name so the
            # DO grid free-text search resolves transporter searches the same
            # way the Filters panel does (transporter_id FK).
            transporter_ids = (
                self.db.query(Order.id)
                .join(Transporter, Order.transporter_id == Transporter.id)
                .filter(
                    or_(
                        Transporter.code.ilike(term),
                        Transporter.name.ilike(term),
                    )
                )
            )
            query_filter = Order.id.in_(
                direct_ids.union(customer_ids).union(transporter_ids)
            )
        if customer_query and (customer_query := (customer_query or "").strip()):
            customer_clause, _ = resolve_via_embedding_then_ilike(
                self.db,
                customer_query,
                source_type="customer",
                ilike_columns=[Order.debtor_name, Order.debtor_code],
                canonical_model=Customer,
                canonical_fields=("customer_name", "customer_code"),
                extra_filter_builders=[
                    lambda like: Order.customer.has(Customer.customer_name.ilike(like)),
                    lambda like: Order.customer.has(Customer.customer_code.ilike(like)),
                ],
            )
            if customer_clause is not None:
                filters.append(customer_clause)
        if product_query and (product_query := (product_query or "").strip()):
            product_clause, _ = resolve_via_embedding_then_ilike(
                self.db,
                product_query,
                source_type="product",
                ilike_columns=[],
                canonical_model=Product,
                canonical_fields=("product_code", "product_name"),
                extra_filter_builders=[
                    lambda like: Order.lines.any(
                        OrderLine.product.has(
                            or_(
                                Product.product_code.ilike(like),
                                Product.product_name.ilike(like),
                                Product.description.ilike(like),
                            )
                        )
                    ),
                ],
            )
            if product_clause is not None:
                filters.append(product_clause)
        if transporter_query and (transporter_query := (transporter_query or "").strip()):
            transporter_clause, _ = resolve_via_embedding_then_ilike(
                self.db,
                transporter_query,
                source_type="transporter",
                ilike_columns=[Order.transporter],
            )
            if transporter_clause is not None:
                filters.append(transporter_clause)

        if entity_buckets is not None and entity_buckets.has_resolved_filter:
            entity_clauses = self._build_entity_filter_clauses(entity_buckets)
            filters.extend(entity_clauses)
            if (
                entity_buckets.debtor_names
                or entity_buckets.customer_names
                or entity_buckets.customer_codes
            ):
                _customer_scoped = True

        if advanced_filter_clause is not None:
            filters.append(advanced_filter_clause)

        all_filters = filters + date_range_filters
        base_q = q.filter(and_(*all_filters)) if all_filters else q
        q = base_q.filter(query_filter) if query_filter is not None else base_q

        total = q.count()
        if query_filter is not None and total == 0 and all_filters:
            # General fallback: if text query yields nothing but structured filters exist,
            # prioritize structured filters instead of returning a false empty result.
            q = base_q
            total = q.count()
        summary_q = q  # the rows' own filter, before sort / offset / limit

        # Nullable columns: use nulls_last so NULLs don't break sort order
        nullable_sort_fields = {
            "order_date",
            "estimated_delivery_date",
            "actual_delivery_date",
            "delivery_days",
            "debtor_name",
            "debtor_code",
            "agent",
            "remarks_cs",
            "order_type",
        }
        sort_map = {
            "order_number": Order.order_number,
            "order_date": Order.order_date,
            "estimated_delivery_date": Order.estimated_delivery_date,
            "actual_delivery_date": Order.actual_delivery_date,
            "delivery_days": Order.delivery_days,
            "debtor_name": Order.debtor_name,
            "debtor_code": Order.debtor_code,
            "agent": Order.agent,
            "is_cancelled": Order.is_cancelled,
            "remarks_cs": Order.remarks_cs,
            "order_type": Order.order_type,
            "total_amount": Order.total_amount,
            "created_at": Order.created_at,
            "updated_at": Order.updated_at,
        }
        # Order.id is always appended as a deterministic tie-breaker so offset
        # position and prev/next neighbours are unambiguous when the primary
        # sort column has equal values (e.g. many rows share a created_at).
        if sort_field == "order_status.status_name":
            q = q.outerjoin(Order.order_status)
            status_col = OrderStatus.status_name
            if sort_dir == "desc":
                q = q.order_by(status_col.desc().nulls_last(), Order.id.asc())
            else:
                q = q.order_by(status_col.asc().nulls_last(), Order.id.asc())
        else:
            sort_column = sort_map.get(sort_field, Order.created_at)
            nullable_sort = sort_field in nullable_sort_fields
            if sort_dir == "desc":
                if nullable_sort:
                    q = q.order_by(sort_column.desc().nulls_last(), Order.id.asc())
                else:
                    q = q.order_by(sort_column.desc(), Order.id.asc())
            else:
                if nullable_sort:
                    q = q.order_by(sort_column.asc().nulls_last(), Order.id.asc())
                else:
                    q = q.order_by(sort_column.asc(), Order.id.asc())

        if ids_only:
            # Same filter+sort query object as the paginated list, but select only
            # the ordered ids (no count/offset/eager-load) for the neighbours pager.
            return [str(row[0]) for row in q.with_entities(Order.id).all()]

        offset = (page - 1) * limit
        # Eager-load relations referenced by OrderResponse to avoid N+1 lazy loads
        # during Pydantic serialization. joinedload for to-one (customer, order_status),
        # selectinload for the lines collection (1 batched IN query, then per-line
        # product is also batched).
        from sqlalchemy.orm import selectinload
        orders = (
            q.options(
                joinedload(Order.customer),
                joinedload(Order.order_status),
                selectinload(Order.lines).joinedload(OrderLine.product),
            )
            .offset(offset)
            .limit(limit)
            .all()
        )

        payload = {
            "data": orders,
            "pagination": {
                "total": total,
                "page": page,
                "limit": limit
            },
            "empty": total == 0
        }
        if include_summary:
            stamp_order_summary(self.db, payload, summary_q, product_ids=_product_uuid_filter)
        # Per-company labelling when the lookup spans more than one company - on the
        # empty path too, so an empty answer can name the companies searched.
        stamp_lookup_companies(
            self.db, payload, orders, product_ids=_product_uuid_filter
        )
        if entity_buckets is not None:
            payload["resolved_entities"] = entity_buckets.as_echo()
        # Date-axis relaxation (§3.4): the customer (and any product/status scope)
        # resolved but had zero DOs on the asked delivery date. Offer the nearest
        # delivery dates that DO carry a DO for the same scope. Only on the empty
        # path - a non-empty result stays byte-identical (AC-R1).
        if total == 0 and _asked_delivery_date is not None and _customer_scoped:
            alternatives = self._nearest_delivery_date_alternatives(
                scope_filters=filters,
                asked_date=_asked_delivery_date,
            )
            if alternatives:
                payload["alternatives"] = alternatives
                payload["relaxed_axis"] = "date"
        return payload

    def _nearest_delivery_date_alternatives(
        self,
        *,
        scope_filters: list[Any],
        asked_date: datetime,
        limit: int = DATE_RELAX_LIMIT,
    ) -> list[dict[str, Any]]:
        """Nearest-date substitutes for an empty customer-scoped delivery query (§3.4).

        `scope_filters` are the non-date entity/scope clauses already built for the
        main query (customer / product / status / transporter), so the substitute
        dates stay within the SAME scope the user asked about - only the date axis
        is relaxed. Rows are ranked by absolute distance from ``asked_date`` (either
        side) and deduped to at most ``limit`` DISTINCT delivery dates. Best-effort:
        a probe failure never turns the already-computed empty result into a 500.
        """
        try:
            from datetime import date as _date
            from sqlalchemy import Date

            # `asked_date` arrives as a query-param string ("2026-04-05") or a
            # date/datetime. Normalize to a python date; Postgres cannot resolve
            # `timestamp - '<string>'` (it collapses to date-date=integer and then
            # `extract(epoch FROM integer)` errors). Cast the column to DATE and
            # subtract dates -> integer day-distance, which is all we need to rank.
            if isinstance(asked_date, str):
                asked = _date.fromisoformat(asked_date[:10])
            elif isinstance(asked_date, datetime):
                asked = asked_date.date()
            else:
                asked = asked_date
            distance = func.abs(func.cast(Order.actual_delivery_date, Date) - asked)
            rows = (
                self.db.query(Order.actual_delivery_date, Order.order_number)
                .filter(
                    Order.deleted_at.is_(None),
                    Order.actual_delivery_date.isnot(None),
                    *scope_filters,
                )
                .order_by(distance.asc(), Order.actual_delivery_date.asc())
                .limit(limit * 5)
                .all()
            )
            alternatives: list[dict[str, Any]] = []
            seen: set[str] = set()
            for adate, order_number in rows:
                if adate is None:
                    continue
                # Column is DateTime but some rows deserialize as date; normalize.
                key = (adate.date() if isinstance(adate, datetime) else adate).isoformat()
                if key in seen:
                    continue
                seen.add(key)
                alternatives.append(
                    {
                        "value": key,
                        "display": f"{key} (DO {order_number})" if order_number else key,
                        "order_number": order_number,
                    }
                )
                if len(alternatives) >= limit:
                    break
            return alternatives
        except Exception:
            logger.exception("nearest-delivery-date probe failed")
            return []

    def _build_entity_filter_clauses(
        self,
        buckets: "EntityFilterBuckets",
        *,
        product_join_already_present: bool = False,
    ) -> list[Any]:
        """Translate resolved entity buckets into SQL filter clauses for Order queries.

        Same-type entities are OR'd (e.g. two product codes → IN). Different types are
        returned as separate clauses so the caller AND's them across types - the user
        said "orders for X and product Y" should intersect, not union. When the caller
        has already joined OrderLine + Product (e.g. list_orders_by_product), pass
        `product_join_already_present=True` so the product clause filters on the
        existing join instead of adding a redundant EXISTS subquery.
        """
        clauses: list[Any] = []
        if buckets.product_codes:
            lowered_codes = [c.lower() for c in buckets.product_codes]
            if product_join_already_present:
                clauses.append(func.lower(Product.product_code).in_(lowered_codes))
            else:
                clauses.append(
                    Order.lines.any(
                        OrderLine.product.has(
                            func.lower(Product.product_code).in_(lowered_codes)
                        )
                    )
                )
        if buckets.order_numbers:
            clauses.append(
                func.lower(Order.order_number).in_([o.lower() for o in buckets.order_numbers])
            )
        customer_subclauses: list[Any] = []
        if buckets.debtor_names:
            lowered = [d.lower() for d in buckets.debtor_names]
            customer_subclauses.append(func.lower(Order.debtor_name).in_(lowered))
        if buckets.customer_codes:
            lowered = [c.lower() for c in buckets.customer_codes]
            customer_subclauses.append(func.lower(Order.debtor_code).in_(lowered))
            customer_subclauses.append(
                Order.customer.has(func.lower(Customer.customer_code).in_(lowered))
            )
        if buckets.customer_names:
            lowered = [n.lower() for n in buckets.customer_names]
            customer_subclauses.append(
                Order.customer.has(func.lower(Customer.customer_name).in_(lowered))
            )
            customer_subclauses.append(func.lower(Order.debtor_name).in_(lowered))
        if customer_subclauses:
            clauses.append(or_(*customer_subclauses))
        if buckets.transporter_codes or buckets.transporter_names:
            from app.models.order import Transporter

            transporter_subclauses: list[Any] = []
            transporter_ids: list[str] = []
            if buckets.transporter_codes:
                lowered = [c.lower() for c in buckets.transporter_codes]
                rows = (
                    self.db.query(Transporter.id)
                    .filter(func.lower(Transporter.code).in_(lowered))
                    .all()
                )
                transporter_ids = [str(r[0]) for r in rows if r[0]]
                if transporter_ids:
                    transporter_subclauses.append(Order.transporter_id.in_(transporter_ids))
                for c in buckets.transporter_codes:
                    transporter_subclauses.append(Order.transporter.ilike(f"%{c}%"))
            if buckets.transporter_names:
                for n in buckets.transporter_names:
                    transporter_subclauses.append(Order.transporter.ilike(f"%{n}%"))
                rows = (
                    self.db.query(Transporter.id)
                    .filter(
                        or_(
                            func.lower(Transporter.name).in_(
                                [n.lower() for n in buckets.transporter_names]
                            ),
                            func.lower(Transporter.normalized_name).in_(
                                [n.lower() for n in buckets.transporter_names]
                            ),
                        )
                    )
                    .all()
                )
                extra_ids = [str(r[0]) for r in rows if r[0] and str(r[0]) not in transporter_ids]
                if extra_ids:
                    transporter_subclauses.append(Order.transporter_id.in_(extra_ids))
            if transporter_subclauses:
                clauses.append(or_(*transporter_subclauses))
        return clauses

    def get_order(self, order_id: str):
        """Get a single order by UUID or order_number (accepts either form)."""
        resolved_ids = resolve_identifier(
            self.db,
            order_id,
            Order,
            code_fields=("order_number",),
        )
        if not resolved_ids:
            raise handle_not_found("Order", order_id)
        order = (
            self.db.query(Order)
            .filter(Order.id.in_(resolved_ids), Order.deleted_at.is_(None))
            .options(
                joinedload(Order.lines).joinedload(OrderLine.product),
                joinedload(Order.lines).joinedload(OrderLine.warehouse),
            )
            .first()
        )
        if not order:
            raise handle_not_found("Order", order_id)
        self._attach_remarks_cs_locked(order)
        return order

    def _attach_remarks_cs_locked(self, order) -> None:
        """Stamp the transient ``remarks_cs_locked`` flag the OrderResponse exposes.

        Delivered DO linked to >=1 complaint => its Remarks CS is frozen. Cheap
        single-row EXISTS; only called on the single-order GET / PUT path.
        """
        try:
            from app.services.complaint_fulfilment_service import (
                ComplaintFulfilmentService,
            )

            order.remarks_cs_locked = ComplaintFulfilmentService(
                self.db
            ).is_remarks_cs_locked(order)
        except Exception:
            order.remarks_cs_locked = False

    def order_neighbours(
        self,
        order_id: str,
        query: Optional[str] = None,
        customer_id: Optional[str] = None,
        order_status_id: Optional[str] = None,
        order_status: Optional[str] = None,
        has_order_lines: Optional[str] = None,
        has_actual_delivery_date: Optional[str] = None,
        order_date_from: Optional[datetime] = None,
        order_date_to: Optional[datetime] = None,
        actual_delivery_date_from: Optional[datetime] = None,
        actual_delivery_date_to: Optional[datetime] = None,
        customer_query: Optional[str] = None,
        product_query: Optional[str] = None,
        transporter_query: Optional[str] = None,
        sort_field: str = "created_at",
        sort_dir: str = "asc",
    ) -> dict:
        """Resolve prev/next neighbours for ``order_id`` within the active list query.

        Reuses ``list_orders(ids_only=True)`` - the exact same filter+sort query the
        paginated grid builds - so the pager and list can never drift. Selects only
        the ordered ids, then defers the position/wrap math to the pure
        ``compute_neighbours`` helper. If the record is not in the filtered set (deep
        link, or filtered out after an edit), falls back to the unfiltered,
        default-sorted set so the pager is never dead (D2).

        ``order_id`` may be a UUID or an order_number; it is resolved to the canonical
        UUID first so it matches the ids produced by the list query.
        """
        from app.services.record_navigation import compute_neighbours

        resolved_ids = resolve_identifier(
            self.db,
            order_id,
            Order,
            code_fields=("order_number",),
        )
        current_id = resolved_ids[0] if resolved_ids else str(order_id)

        filtered_ids = self.list_orders(
            query=query,
            customer_id=customer_id,
            order_status_id=order_status_id,
            order_status=order_status,
            has_order_lines=has_order_lines,
            has_actual_delivery_date=has_actual_delivery_date,
            order_date_from=order_date_from,
            order_date_to=order_date_to,
            actual_delivery_date_from=actual_delivery_date_from,
            actual_delivery_date_to=actual_delivery_date_to,
            customer_query=customer_query,
            product_query=product_query,
            transporter_query=transporter_query,
            sort_field=sort_field,
            sort_dir=sort_dir,
            ids_only=True,
        )
        result = compute_neighbours(filtered_ids, current_id)
        if result["index"] is not None:
            return result

        # D2: current record not in the filtered set -> fall back to the unfiltered,
        # default-sorted set so prev/next still works and total reflects all orders.
        unfiltered_ids = self.list_orders(ids_only=True)
        return compute_neighbours(unfiltered_ids, current_id)

    # ------------------------------------------------------------------ #
    # Aggregation / analytics
    # ------------------------------------------------------------------ #
    ANALYTICS_METRICS = ("count", "total_value", "avg_delivery_days")
    ANALYTICS_GROUP_BY = ("customer", "product", "month", "none")

    def order_analytics(
        self,
        *,
        metric: str,
        group_by: str = "none",
        customer_ids: Optional[list[str]] = None,
        product_ids: Optional[list[str]] = None,
        product_code: Optional[str] = None,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        limit: int = 50,
    ) -> dict:
        """Compute an aggregate metric over customer sales orders.

        metric:
        - count             → number of matching orders
        - total_value       → SUM(orders.total_amount) (order-level revenue);
                                 when group_by=product it is the per-product line
                                 revenue (SUM of the matching order lines' totals)
        - avg_delivery_days → AVG(actual_delivery_date - order_date) in days,
                                 over orders that have BOTH dates set.

        group_by buckets the result by customer (debtor_name), product
        (product_code), calendar month (order_date YYYY-MM), or none (single
        overall figure). Returns ranked group rows + an order-level overall total.
        Only aggregates are exposed - never per-order pricing / cost fields.
        """
        from app.services.error_handler import AppException
        from fastapi import status as _status

        metric = (metric or "").strip().lower()
        if metric not in self.ANALYTICS_METRICS:
            raise AppException(
                status_code=_status.HTTP_422_UNPROCESSABLE_ENTITY,
                message=(
                    f"Unknown metric {metric!r}. Use one of: "
                    + ", ".join(self.ANALYTICS_METRICS)
                ),
                code="INVALID_METRIC",
            )
        group_by = (group_by or "none").strip().lower() or "none"
        if group_by not in self.ANALYTICS_GROUP_BY:
            raise AppException(
                status_code=_status.HTTP_422_UNPROCESSABLE_ENTITY,
                message=(
                    f"Unknown group_by {group_by!r}. Use one of: "
                    + ", ".join(self.ANALYTICS_GROUP_BY)
                ),
                code="INVALID_GROUP_BY",
            )
        try:
            limit = max(1, min(int(limit), 500))
        except (TypeError, ValueError):
            limit = 50

        customer_ids = list(customer_ids) if customer_ids else None
        product_ids = list(product_ids) if product_ids else None
        product_code = (product_code or "").strip() or None

        def _apply_order_filters(q):
            q = q.filter(Order.deleted_at.is_(None))
            if customer_ids:
                # Mirror list_orders: Order.customer_id IN (...) OR legacy
                # debtor_name match for historical rows with NULL customer_id.
                names_subq = (
                    self.db.query(Customer.customer_name)
                    .filter(Customer.id.in_(customer_ids))
                    .subquery()
                )
                q = q.filter(
                    or_(
                        Order.customer_id.in_(customer_ids),
                        func.lower(func.btrim(Order.debtor_name)).in_(
                            self.db.query(
                                func.lower(func.btrim(names_subq.c.customer_name))
                            )
                        ),
                    )
                )
            if product_ids or product_code:
                line_conds = []
                if product_ids:
                    line_conds.append(OrderLine.product_id.in_(product_ids))
                if product_code:
                    like = f"%{product_code.lower()}%"
                    line_conds.append(
                        OrderLine.product.has(func.lower(Product.product_code).like(like))
                    )
                q = q.filter(Order.lines.any(or_(*line_conds)))
            if date_from is not None:
                q = q.filter(Order.order_date >= date_from)
            if date_to is not None:
                q = q.filter(Order.order_date <= date_to)
            return q

        # Order value = header total_amount when set, else the sum of the order's
        # line totals. In this dataset orders.total_amount is 0 across the board - 
        # the real monetary value lives on the order lines - so we coalesce to the
        # line sum to return a truthful figure (spec said SUM(total_amount); the
        # header is unpopulated, so this is the honest equivalent).
        line_sum_subq = (
            self.db.query(
                OrderLine.order_id.label("oid"),
                func.sum(
                    func.coalesce(OrderLine.total_including_tax, OrderLine.total, 0)
                ).label("lsum"),
            )
            .group_by(OrderLine.order_id)
            .subquery()
        )

        # ---- order-level overall totals (grouping-independent, distinct orders)
        overall_rows = _apply_order_filters(
            self.db.query(
                Order.id,
                Order.debtor_name,
                Order.order_date,
                Order.total_amount,
                Order.actual_delivery_date,
                func.coalesce(line_sum_subq.c.lsum, 0).label("line_sum"),
            ).outerjoin(line_sum_subq, line_sum_subq.c.oid == Order.id)
        ).all()

        def _order_value(row) -> float:
            header = float(row.total_amount or 0)
            return header if header else float(row.line_sum or 0)

        def _day_diff(order_date, delivery_date) -> Optional[float]:
            if order_date is None or delivery_date is None:
                return None
            try:
                return (delivery_date - order_date).total_seconds() / 86400.0
            except (TypeError, AttributeError):
                return None

        def _metric_value(order_total_sum, order_count, day_diffs, line_total_sum):
            if metric == "count":
                return order_count
            if metric == "total_value":
                base = line_total_sum if group_by == "product" else order_total_sum
                return round(float(base or 0), 2)
            # avg_delivery_days
            vals = [d for d in day_diffs if d is not None]
            if not vals:
                return None
            return round(sum(vals) / len(vals), 2)

        overall_total_amount = sum((_order_value(r) for r in overall_rows), 0.0)
        overall_day_diffs = [
            _day_diff(r.order_date, r.actual_delivery_date) for r in overall_rows
        ]
        overall_value = _metric_value(
            overall_total_amount, len(overall_rows), overall_day_diffs, overall_total_amount
        )
        overall = {
            "group_key": "__all__",
            "group_label": "All orders",
            "metric": metric,
            "value": overall_value,
            "order_count": len(overall_rows),
        }

        # ---- grouped rows
        groups: list[dict] = []
        if group_by == "product":
            like = f"%{product_code.lower()}%" if product_code else None
            pq = (
                self.db.query(
                    Product.product_code.label("product_code"),
                    Product.product_name.label("product_name"),
                    Order.id.label("order_id"),
                    Order.total_amount.label("total_amount"),
                    Order.order_date.label("order_date"),
                    Order.actual_delivery_date.label("actual_delivery_date"),
                    func.coalesce(
                        OrderLine.total_including_tax, OrderLine.total, 0
                    ).label("line_total"),
                )
                .join(OrderLine, OrderLine.order_id == Order.id)
                .join(Product, Product.id == OrderLine.product_id)
                .filter(Order.deleted_at.is_(None))
            )
            if customer_ids:
                names_subq = (
                    self.db.query(Customer.customer_name)
                    .filter(Customer.id.in_(customer_ids))
                    .subquery()
                )
                pq = pq.filter(
                    or_(
                        Order.customer_id.in_(customer_ids),
                        func.lower(func.btrim(Order.debtor_name)).in_(
                            self.db.query(
                                func.lower(func.btrim(names_subq.c.customer_name))
                            )
                        ),
                    )
                )
            # Product filter applied to the joined line directly so only matching
            # product lines are grouped (not every product on a matching order).
            if product_ids:
                pq = pq.filter(OrderLine.product_id.in_(product_ids))
            if like:
                pq = pq.filter(func.lower(Product.product_code).like(like))
            if date_from is not None:
                pq = pq.filter(Order.order_date >= date_from)
            if date_to is not None:
                pq = pq.filter(Order.order_date <= date_to)

            buckets: dict[str, dict] = {}
            for r in pq.all():
                key = r.product_code or "(unknown)"
                b = buckets.setdefault(
                    key,
                    {
                        "label": (
                            f"{r.product_code} - {r.product_name}"
                            if r.product_name
                            else (r.product_code or "(unknown)")
                        ),
                        "orders": {},
                        "line_total": 0.0,
                    },
                )
                b["orders"][r.order_id] = _day_diff(
                    r.order_date, r.actual_delivery_date
                )
                b["line_total"] += float(r.line_total or 0)
            for key, b in buckets.items():
                groups.append(
                    {
                        "group_key": key,
                        "group_label": b["label"],
                        "metric": metric,
                        "value": _metric_value(
                            0.0, len(b["orders"]), list(b["orders"].values()), b["line_total"]
                        ),
                        "order_count": len(b["orders"]),
                    }
                )
        elif group_by == "none":
            groups.append({**overall, "group_key": "all", "group_label": "All orders"})
        else:
            buckets: dict[str, dict] = {}
            for r in overall_rows:
                if group_by == "customer":
                    label = (r.debtor_name or "(no customer)").strip() or "(no customer)"
                    key = label.lower()
                else:  # month
                    key = r.order_date.strftime("%Y-%m") if r.order_date else "(no date)"
                    label = key
                b = buckets.setdefault(
                    key, {"label": label, "total_amount": 0.0, "day_diffs": [], "count": 0}
                )
                b["total_amount"] += _order_value(r)
                b["day_diffs"].append(_day_diff(r.order_date, r.actual_delivery_date))
                b["count"] += 1
            for key, b in buckets.items():
                groups.append(
                    {
                        "group_key": key,
                        "group_label": b["label"],
                        "metric": metric,
                        "value": _metric_value(
                            b["total_amount"], b["count"], b["day_diffs"], b["total_amount"]
                        ),
                        "order_count": b["count"],
                    }
                )

        # Rank: highest value first (None sinks to the bottom); cap to `limit`.
        groups.sort(
            key=lambda g: (g["value"] is not None, g["value"] if g["value"] is not None else 0),
            reverse=True,
        )
        groups = groups[:limit]

        return {
            "metric": metric,
            "group_by": group_by,
            "filters": {
                "customer_ids": customer_ids,
                "product_ids": product_ids,
                "product_code": product_code,
                "date_from": date_from.isoformat() if date_from else None,
                "date_to": date_to.isoformat() if date_to else None,
            },
            "groups": groups,
            "total": overall,
            "empty": len(overall_rows) == 0,
        }

    def list_orders_by_product(
        self,
        page: int = 1,
        limit: int = 50,
        query: Optional[str] = None,
        customer_query: Optional[str] = None,
        product_query: Optional[str] = None,
        product_id: Optional[str] = None,
        has_actual_delivery_date: Optional[str] = None,
        order_date_from: Optional[datetime] = None,
        order_date_to: Optional[datetime] = None,
        actual_delivery_date_from: Optional[datetime] = None,
        actual_delivery_date_to: Optional[datetime] = None,
        sort_field: str = "order_date",
        sort_dir: str = "desc",
        entities: Optional[list[str]] = None,
        product_ids: Optional[list[str]] = None,
        customer_ids: Optional[list[str]] = None,
        transporter_ids: Optional[list[str]] = None,
        order_status: Optional[str] = None,
        include_summary: bool = False,
    ):
        """List distinct orders matched by product search.

        ``order_status`` is the same 'outstanding' | 'delivered' bucket as
        ``list_orders`` (canonical delivered predicate), so the two order tools
        agree on what delivered means.

        `entities` is the single-bag entity filter (see list_orders). At least one
        entity must resolve to a product - the endpoint is product-centric and falls
        back to empty otherwise. Customer / transporter / order_number entities are
        applied as additional AND filters.
        """
        # Capture typed UUID kwargs before any local variable shadows them.
        _product_uuid_filter = list(product_ids) if product_ids else None
        _customer_uuid_filter = list(customer_ids) if customer_ids else None
        _transporter_uuid_filter = list(transporter_ids) if transporter_ids else None

        q = (
            self.db.query(Order)
            .join(Order.lines)
            .join(OrderLine.product)
            .filter(Order.deleted_at.is_(None))
            .distinct()
        )

        if _product_uuid_filter:
            q = q.filter(OrderLine.product_id.in_(_product_uuid_filter))

        if _customer_uuid_filter:
            from sqlalchemy import func as _func
            customer_names_subq = (
                self.db.query(Customer.customer_name)
                .filter(Customer.id.in_(_customer_uuid_filter))
                .subquery()
            )
            q = q.filter(
                or_(
                    Order.customer_id.in_(_customer_uuid_filter),
                    _func.lower(_func.btrim(Order.debtor_name)).in_(
                        self.db.query(_func.lower(_func.btrim(customer_names_subq.c.customer_name)))
                    ),
                )
            )

        if _transporter_uuid_filter:
            from sqlalchemy import func as _func
            transporter_names_subq = (
                self.db.query(Transporter.name)
                .filter(Transporter.id.in_(_transporter_uuid_filter))
                .subquery()
            )
            q = q.filter(
                or_(
                    Order.transporter_id.in_(_transporter_uuid_filter),
                    _func.lower(_func.btrim(Order.transporter)).in_(
                        self.db.query(_func.lower(_func.btrim(transporter_names_subq.c.name)))
                    ),
                )
            )

        entity_buckets: Optional[EntityFilterBuckets] = None
        if entities:
            entity_buckets = resolve_entities_to_filters(
                self.db,
                entities,
                allowed_entity_types=(
                    "product", "customer", "customer_order", "transporter",
                    "inbound_shipment", "spo_allocation", "grn", "promotion",
                    "attachment", "form",
                ),
            )
            # Caller passed `entities` but nothing resolved above the RAG
            # similarity threshold → return empty rather than fall through and
            # unintentionally list every order. Echo the unresolved inputs so the
            # agent surfaces "no match" to the user.
            if not entity_buckets.has_resolved_filter:
                payload = {
                    "data": [],
                    "pagination": {"total": 0, "page": page, "limit": limit},
                    "empty": True,
                    "resolved_entities": entity_buckets.as_echo(),
                }
                stamp_lookup_companies(
                    self.db, payload, [], product_ids=_product_uuid_filter
                )
                return payload

        filters = []
        product_match_filters: list = []

        # Drive matched_products enrichment off the typed UUID filter too - 
        # otherwise rows come back with matched_products=[] when the caller
        # narrows by `product_ids` (the MCP/typed path) instead of the legacy
        # `product_id` token resolver below. The order-level filter is applied
        # separately above via _product_uuid_filter on `q`.
        if _product_uuid_filter:
            product_match_filters.append(
                OrderLine.product_id.in_(_product_uuid_filter)
            )

        # Normalize `product_id` into a list of raw tokens. Accepts a single
        # UUID/SKU, a comma-separated string ("A,B"), or a Python list (n8n
        # repeated ?product_id=A&product_id=B). Each token is resolved through
        # the standard exact path first; tokens that still don't match a
        # product_code fall through to an embedding fuzzy resolve so the LLM
        # can hand us partials like "SRTWC8608" or "SRTWC8608-SC-UF" without
        # n8n having to canonicalise the SKU upstream.
        raw_tokens: list[str] = []
        if isinstance(product_id, list):
            for item in product_id:
                if item is None:
                    continue
                for piece in str(item).split(","):
                    piece = piece.strip()
                    if piece:
                        raw_tokens.append(piece)
        elif product_id is not None:
            for piece in str(product_id).split(","):
                piece = piece.strip()
                if piece:
                    raw_tokens.append(piece)
        raw_tokens = list(dict.fromkeys(raw_tokens))

        product_ids: list[str] = []
        if raw_tokens:
            for token in raw_tokens:
                ids = resolve_identifier(
                    self.db,
                    token,
                    Product,
                    code_fields=("product_code",),
                )
                if ids:
                    product_ids.extend(ids)
                    continue
                # Prefix / substring ILIKE on product_code - deterministic
                # match for partial SKUs like "SRTWC8608" → "SRTWC8608-SC",
                # "SRTWC8608-SC-UF". Capped to avoid runaway broad tokens.
                like_rows = (
                    self.db.query(Product.id, Product.product_code)
                    .filter(Product.product_code.ilike(f"%{token}%"))
                    .limit(20)
                    .all()
                )
                if like_rows:
                    product_ids.extend(str(r[0]) for r in like_rows)
                    continue
                # Embedding fuzzy fallback only when prefix/substring code
                # match yields nothing. Useful for descriptive tokens (e.g.
                # "kitchen sink") rather than partial codes.
                _fuzzy_clause, canonical_values = resolve_via_embedding_then_ilike(
                    self.db,
                    token,
                    source_type="product",
                    ilike_columns=[Product.product_code],
                    canonical_model=Product,
                    canonical_fields=("product_code",),
                )
                if canonical_values:
                    fuzzy_rows = (
                        self.db.query(Product.id)
                        .filter(Product.product_code.in_(canonical_values))
                        .all()
                    )
                    product_ids.extend(str(r[0]) for r in fuzzy_rows)
            product_ids = list(dict.fromkeys(product_ids))
            if not product_ids:
                payload = {
                    "data": [],
                    "pagination": {"total": 0, "page": page, "limit": limit},
                    "empty": True,
                }
                stamp_lookup_companies(
                    self.db, payload, [], product_ids=_product_uuid_filter
                )
                return payload
            filters.append(OrderLine.product_id.in_(product_ids))
            product_match_filters.append(OrderLine.product_id.in_(product_ids))

        hadd = (has_actual_delivery_date or "").strip().lower()
        if hadd == "yes":
            filters.append(Order.actual_delivery_date.is_not(None))
        elif hadd == "no":
            filters.append(Order.actual_delivery_date.is_(None))

        bucket_clause = _order_status_bucket_filter(self.db, order_status)
        if bucket_clause is not None:
            filters.append(bucket_clause)

        if order_date_from is not None:
            filters.append(Order.order_date >= order_date_from)

        if order_date_to is not None:
            filters.append(Order.order_date <= order_date_to)

        if actual_delivery_date_from is not None:
            filters.append(Order.actual_delivery_date >= actual_delivery_date_from)

        if actual_delivery_date_to is not None:
            filters.append(Order.actual_delivery_date <= actual_delivery_date_to)

        query_filter = None
        if query and (query := (query or "").strip()):
            term = f"%{query}%"
            # Transporter sub-id set joined to the canonical transporters table
            # so by-product search also finds DOs by transporter code / name
            # (matches `list_orders` behavior).
            transporter_id_subq = (
                self.db.query(Transporter.id).filter(
                    or_(
                        Transporter.code.ilike(term),
                        Transporter.name.ilike(term),
                    )
                )
            )
            query_filter = or_(
                Order.order_number.ilike(term),
                Order.debtor_name.ilike(term),
                Order.transporter.ilike(term),  # legacy free-text column
                Order.transporter_id.in_(transporter_id_subq),
                Product.product_code.ilike(term),
                Product.product_name.ilike(term),
                Product.description.ilike(term),
            )
            product_match_filters.append(
                or_(
                    Product.product_code.ilike(term),
                    Product.product_name.ilike(term),
                    Product.description.ilike(term),
                )
            )
        if customer_query and (customer_query := (customer_query or "").strip()):
            customer_clause, _ = resolve_via_embedding_then_ilike(
                self.db,
                customer_query,
                source_type="customer",
                ilike_columns=[Order.debtor_name, Order.debtor_code],
                canonical_model=Customer,
                canonical_fields=("customer_name", "customer_code"),
                extra_filter_builders=[
                    lambda like: Order.customer.has(Customer.customer_name.ilike(like)),
                    lambda like: Order.customer.has(Customer.customer_code.ilike(like)),
                ],
            )
            if customer_clause is not None:
                filters.append(customer_clause)
        if product_query and (product_query := (product_query or "").strip()):
            product_clause, _ = resolve_via_embedding_then_ilike(
                self.db,
                product_query,
                source_type="product",
                ilike_columns=[
                    Product.product_code,
                    Product.product_name,
                    Product.description,
                ],
                canonical_model=Product,
                canonical_fields=("product_code", "product_name"),
            )
            if product_clause is not None:
                filters.append(product_clause)
                product_match_filters.append(product_clause)

        if entity_buckets is not None and entity_buckets.has_resolved_filter:
            entity_clauses = self._build_entity_filter_clauses(
                entity_buckets, product_join_already_present=True
            )
            filters.extend(entity_clauses)
            if entity_buckets.product_codes:
                product_match_filters.append(
                    func.lower(Product.product_code).in_(
                        [c.lower() for c in entity_buckets.product_codes]
                    )
                )

        # By-product endpoint is product-centric: if the caller passed `entities` but
        # zero of them resolved to a product (and no other product filter is in
        # play), there's nothing to look up. Return an explicit empty rather than
        # silently fanning across every order.
        if (
            entity_buckets is not None
            and entities
            and not entity_buckets.product_codes
            and not product_id
            and not product_query
            and not (query and query.strip())
        ):
            payload = {
                "data": [],
                "pagination": {"total": 0, "page": page, "limit": limit},
                "empty": True,
                "resolved_entities": entity_buckets.as_echo(),
            }
            stamp_lookup_companies(
                self.db,
                payload,
                [],
                product_ids=[*(_product_uuid_filter or []), *(product_ids or [])],
            )
            return payload

        base_q = q.filter(and_(*filters)) if filters else q
        q = base_q.filter(query_filter) if query_filter is not None else base_q

        sort_map = {
            "order_number": Order.order_number,
            "order_date": Order.order_date,
            "actual_delivery_date": Order.actual_delivery_date,
            "debtor_name": Order.debtor_name,
            "created_at": Order.created_at,
            "updated_at": Order.updated_at,
        }
        sort_column = sort_map.get(sort_field, Order.order_date)
        if (sort_dir or "").lower() == "asc":
            q = q.order_by(sort_column.asc().nulls_last())
        else:
            q = q.order_by(sort_column.desc().nulls_last())

        total = q.count()
        if query_filter is not None and total == 0 and filters:
            # General fallback for noisy intent-like query text.
            q = base_q
            total = q.count()
        summary_q = q  # the rows' own filter, before offset / limit
        offset = (page - 1) * limit
        orders = q.offset(offset).limit(limit).all()

        # Attach matched_products per order so callers know which order line
        # caused the match (esp. important when product_id was an ilike /
        # fuzzy hit rather than an exact UUID).
        order_ids = [str(o.id) for o in orders]
        matched_by_order: dict[str, list[dict]] = {}
        if order_ids and product_match_filters:
            mq = (
                self.db.query(
                    OrderLine.order_id,
                    Product.id,
                    Product.product_code,
                    Product.product_name,
                    OrderLine.quantity,
                    Warehouse.warehouse_code,
                    Warehouse.warehouse_name,
                )
                .join(Product, Product.id == OrderLine.product_id)
                .outerjoin(Warehouse, Warehouse.id == OrderLine.warehouse_id)
                .filter(OrderLine.order_id.in_(order_ids))
                .filter(and_(*product_match_filters))
            )
            seen: set[tuple[str, str, str]] = set()
            for oid, pid, pcode, pname, qty, wcode, wname in mq.distinct().all():
                key = (str(oid), str(pid), str(wcode))
                if key in seen:
                    continue
                seen.add(key)
                matched_by_order.setdefault(str(oid), []).append(
                    {
                        "id": str(pid),
                        "product_code": pcode,
                        "product_name": pname,
                        "quantity": float(qty) if qty is not None else None,
                        "warehouse_code": wcode,
                        "warehouse_name": wname,
                    }
                )
        for o in orders:
            setattr(o, "matched_products", matched_by_order.get(str(o.id), []))

        payload = {
            "data": orders,
            "pagination": {
                "total": total,
                "page": page,
                "limit": limit,
            },
            "empty": total == 0,
        }
        if include_summary:
            stamp_order_summary(
                self.db,
                payload,
                summary_q,
                product_ids=[*(_product_uuid_filter or []), *(product_ids or [])],
            )
        # Per-company labelling when the lookup spans more than one company - on the
        # empty path too, so an empty answer can name the companies searched. Both
        # the typed uuid filter and the ids resolved from free-text product tokens
        # count as "what was asked about".
        stamp_lookup_companies(
            self.db,
            payload,
            orders,
            product_ids=[*(_product_uuid_filter or []), *(product_ids or [])],
        )
        if entity_buckets is not None:
            payload["resolved_entities"] = entity_buckets.as_echo()
        return payload

    @staticmethod
    def _normalize_uuid_fields(data: dict, keys: tuple = ("customer_id", "order_status_id", "billing_address_id", "shipping_address_id")) -> dict:
        """Set empty-string UUID fields to None so PostgreSQL accepts them."""
        out = dict(data)
        for key in keys:
            if key in out and out[key] == "":
                out[key] = None
        return out

    def _upsert_customer_from_debtor(
        self,
        debtor_name: Optional[str],
        debtor_code: Optional[str],
    ) -> Optional[str]:
        """Find-or-create a Customer row from order debtor fields, return its id.

        Match key is the (customer_code, customer_name) PAIR - case + whitespace
        insensitive - because one Sage code can carry multiple distinct debtor
        names (e.g. "300-D093" maps to both "Deluxe Home Center (KTN)" and
        "Deluxe Home Center AC (I)"). Each unique pair gets its own customers
        row. Matches the composite unique index created in migration 220.

        When debtor_code is missing, fall back to a deterministic `DBR-<hash>`
        slug derived from the name so the same blank-code debtor doesn't
        explode into duplicate rows.
        """
        name = (debtor_name or "").strip()
        if not name:
            return None
        code = (debtor_code or "").strip()
        if not code:
            import hashlib
            code = "DBR-" + hashlib.md5(name.lower().encode("utf-8")).hexdigest()[:10]
        # Pair match (case + whitespace normalized on both columns).
        existing = (
            self.db.query(Customer)
            .filter(
                func.lower(func.btrim(Customer.customer_code)) == code.lower(),
                func.lower(func.btrim(Customer.customer_name)) == name.lower(),
            )
            .first()
        )
        if existing is not None:
            return str(existing.id)
        new_customer = Customer(
            customer_code=code,
            customer_name=name,
            customer_type="company",
            is_active=True,
        )
        self.db.add(new_customer)
        # Flush so the new id is available before the order persists; commit
        # remains with the order so a single transaction wraps both writes.
        self.db.flush()
        return str(new_customer.id)

    def _upsert_transporter_from_text(self, transporter_text: Optional[str]) -> Optional[str]:
        """Find-or-create a Transporter row keyed by normalized_name, return id.

        Mirrors migration 200's seeding rule. `normalized_name = lower(btrim(...))`
        is the unique key - code + name default to the raw stripped text on
        first sight.
        """
        raw = (transporter_text or "").strip()
        if not raw:
            return None
        normalized = raw.lower()
        existing = (
            self.db.query(Transporter)
            .filter(Transporter.normalized_name == normalized)
            .first()
        )
        if existing is not None:
            return str(existing.id)
        new_row = Transporter(
            code=raw,
            name=raw,
            normalized_name=normalized,
        )
        self.db.add(new_row)
        self.db.flush()
        return str(new_row.id)

    def _sync_order_master_refs(self, order_dict: dict) -> dict:
        """Populate `customer_id` and `transporter_id` from text fields, in place.

        Called by create_order / update_order before persistence. Only writes
        the FK when (a) the caller did not supply one explicitly, OR (b) the
        existing value does not match the text field (text edit re-points the
        FK). Idempotent: master rows are deduped via the unique
        `customer_code` / `transporters.normalized_name` indexes.
        """
        out = dict(order_dict)

        debtor_name = out.get("debtor_name")
        if debtor_name:
            cid = self._upsert_customer_from_debtor(debtor_name, out.get("debtor_code"))
            if cid:
                out["customer_id"] = cid  # always re-point; debtor_name edit can change customer

        transporter_text = out.get("transporter")
        if transporter_text:
            tid = self._upsert_transporter_from_text(transporter_text)
            if tid:
                out["transporter_id"] = tid

        return out

    def create_order(self, order_data: OrderCreate, created_by: str):
        """Create a new order."""
        # Check if order_number already exists
        existing = self.db.query(Order).filter(Order.order_number == order_data.order_number).first()
        if existing:
            raise handle_conflict("Order number already exists. Please use a different number.")
        
        # Calculate total if not provided
        subtotal = order_data.subtotal_amount or Decimal("0")
        discount = order_data.discount_amount or Decimal("0")
        tax = order_data.tax_amount or Decimal("0")
        total = subtotal - discount + tax
        
        order_dict = order_data.model_dump()
        order_dict = self._normalize_uuid_fields(order_dict)
        # Auto-upsert customer + transporter master rows from text fields so the
        # FKs stay populated for every new order (mirrors migrations 200/206/207).
        order_dict = self._sync_order_master_refs(order_dict)
        order_dict["total_amount"] = total
        order_dict["created_by"] = created_by

        order = Order(**order_dict)
        self.db.add(order)
        self.db.commit()
        self.db.refresh(order)

        # Complaint <-> DO auto-fulfilment on CREATE: a DO born with a Remarks CS
        # reference (optionally already delivered/cancelled) must link + (re)fulfil
        # immediately, mirroring update_order. Only run when there's something to
        # link - a blank Remarks CS on a fresh order can't affect any complaint.
        # Best-effort - never fail the create.
        if (order.remarks_cs or "").strip():
            try:
                from app.services.complaint_fulfilment_service import (
                    ComplaintFulfilmentService,
                )

                ComplaintFulfilmentService(self.db).apply_for_orders(
                    [{"order": order, "old_remarks": None}]
                )
                self._attach_remarks_cs_locked(order)
            except Exception:
                logger.exception(
                    "Order %s: complaint fulfilment recompute failed after create",
                    order.id,
                )
        return order

    def update_order(self, order_id: str, order_data: OrderUpdate, updated_by: str):
        """Update an order."""
        order = self.get_order(order_id)

        update_data = order_data.model_dump(exclude_unset=True)
        if update_data:
            update_data = self._normalize_uuid_fields(update_data)

            # Freeze guard: a delivered DO linked to a complaint has historical,
            # already-notified fulfilment - its Remarks CS is read-only. Reject any
            # change rather than silently keep the old value (FE renders it readonly).
            old_remarks_cs = order.remarks_cs
            remarks_cs_changing = (
                "remarks_cs" in update_data
                and (update_data.get("remarks_cs") or "").strip()
                != (old_remarks_cs or "").strip()
            )
            old_actual_delivery = order.actual_delivery_date
            old_is_cancelled = bool(order.is_cancelled)
            old_status_id = order.order_status_id
            if remarks_cs_changing:
                # Evaluate the freeze against the INCOMING (status, delivery) state, so
                # moving the DO off the delivered status in this SAME save unfreezes
                # Remarks CS (the unfreeze-via-status flow). Frozen = delivered (date
                # set AND delivered/completed status) AND linked to a complaint.
                from app.services.complaint_fulfilment_service import (
                    ComplaintFulfilmentService,
                )

                incoming_status_id = update_data.get(
                    "order_status_id", order.order_status_id
                )
                incoming_delivery = (
                    update_data["actual_delivery_date"]
                    if "actual_delivery_date" in update_data
                    else order.actual_delivery_date
                )
                if ComplaintFulfilmentService(self.db).is_locked_for_state(
                    order, incoming_status_id, incoming_delivery
                ):
                    from app.services.error_handler import AppException
                    from fastapi import status as _status

                    raise AppException(
                        status_code=_status.HTTP_422_UNPROCESSABLE_ENTITY,
                        message=(
                            "Remarks CS is locked: this delivery order is delivered and "
                            "linked to a complaint, so its complaint reference cannot be changed."
                        ),
                        code="REMARKS_CS_LOCKED",
                    )
            # Recalculate total if amounts changed
            if any(k in update_data for k in ["subtotal_amount", "discount_amount", "tax_amount"]):
                subtotal = update_data.get("subtotal_amount", order.subtotal_amount) or Decimal("0")
                discount = update_data.get("discount_amount", order.discount_amount) or Decimal("0")
                tax = update_data.get("tax_amount", order.tax_amount) or Decimal("0")
                update_data["total_amount"] = subtotal - discount + tax

            # Re-sync customer_id / transporter_id whenever debtor_name /
            # debtor_code / transporter text changes on update. Carry the
            # pre-existing values forward so the upsert sees full context even
            # if the caller only patched one field.
            text_view = {
                "debtor_name": update_data.get("debtor_name", order.debtor_name),
                "debtor_code": update_data.get("debtor_code", order.debtor_code),
                "transporter": update_data.get("transporter", order.transporter),
            }
            synced = self._sync_order_master_refs(text_view)
            if "customer_id" in synced:
                update_data["customer_id"] = synced["customer_id"]
            if "transporter_id" in synced:
                update_data["transporter_id"] = synced["transporter_id"]

            update_data["updated_by"] = updated_by
            for key, value in update_data.items():
                setattr(order, key, value)

            self.db.commit()
            self.db.refresh(order)

            # Complaint <-> DO auto-fulfilment: a Remarks CS, delivery-date, or
            # is_cancelled change can (un)link complaints, (re)fulfil them, and
            # fire a per-DO delivery notice. Best-effort - never fail the update.
            try:
                fulfilment_relevant = (
                    remarks_cs_changing
                    or (order.actual_delivery_date is not None) != (old_actual_delivery is not None)
                    or bool(order.is_cancelled) != old_is_cancelled
                    # Status change flips the delivered-state (delivered = date AND
                    # delivered/completed status), so recompute on a status change too.
                    or str(order.order_status_id or "") != str(old_status_id or "")
                )
                if fulfilment_relevant:
                    from app.services.complaint_fulfilment_service import (
                        ComplaintFulfilmentService,
                    )

                    ComplaintFulfilmentService(self.db).apply_for_orders(
                        [{"order": order, "old_remarks": old_remarks_cs}]
                    )
                    self._attach_remarks_cs_locked(order)
            except Exception:
                logger.exception(
                    "Order %s: complaint fulfilment recompute failed after update", order_id
                )

        return order

    def create_order_line(self, order_id: str, data: OrderLineCreate):
        """Add a line to an order. Lines are ordered by line_sequence (multiple lines may share product+warehouse)."""
        self.get_order(order_id)  # ensure order exists
        next_seq = (
            self.db.query(func.coalesce(func.max(OrderLine.line_sequence), 0))
            .filter(OrderLine.order_id == order_id)
            .scalar()
        )
        next_seq = int(next_seq or 0) + 1
        line = OrderLine(order_id=order_id, line_sequence=next_seq, **data.model_dump())
        self.db.add(line)
        self.db.commit()
        self.db.refresh(line)
        return line

    def update_order_line(self, order_id: str, line_id: str, data: OrderLineUpdate):
        """Update an order line."""
        line = (
            self.db.query(OrderLine)
            .filter(OrderLine.id == line_id, OrderLine.order_id == order_id)
            .first()
        )
        if not line:
            raise handle_not_found("Order line", line_id)
        for key, value in data.model_dump(exclude_unset=True).items():
            setattr(line, key, value)
        self.db.commit()
        self.db.refresh(line)
        return line

    def delete_order_line(self, order_id: str, line_id: str):
        """Remove an order line."""
        line = (
            self.db.query(OrderLine)
            .filter(OrderLine.id == line_id, OrderLine.order_id == order_id)
            .first()
        )
        if not line:
            raise handle_not_found("Order line", line_id)
        self.db.delete(line)
        self.db.commit()
        return {"message": "Order line deleted"}

    def bulk_delete_order_lines(self, order_id: str, line_ids: list[str]):
        """Delete multiple lines from an order by line IDs."""
        if not line_ids:
            return {"message": "No order lines to delete", "deleted_count": 0}
        deleted = (
            self.db.query(OrderLine)
            .filter(OrderLine.order_id == order_id, OrderLine.id.in_(line_ids))
            .delete(synchronize_session=False)
        )
        self.db.commit()
        return {"message": f"Deleted {deleted} order line(s)", "deleted_count": deleted}

    def get_order_by_order_number(self, order_number: str):
        """Get order by order_number (doc no). Returns None if not found."""
        if not (order_number or "").strip():
            return None
        return (
            self.db.query(Order)
            .filter(Order.deleted_at.is_(None), Order.order_number == (order_number or "").strip())
            .first()
        )

    def _get_order_any(self, order_id: str):
        """Get order by ID including archived."""
        order = self.db.query(Order).filter(Order.id == order_id).first()
        if not order:
            raise handle_not_found("Order", order_id)
        return order

    def archive_order(self, order_id: str):
        """Archive an order (soft delete). Data remains for retention."""
        from datetime import datetime, timezone
        order = self.get_order(order_id)
        order.deleted_at = datetime.now(timezone.utc)
        self.db.commit()
        return {"message": "Order archived successfully"}

    def restore_order(self, order_id: str):
        """Restore an archived order."""
        order = self._get_order_any(order_id)
        if order.deleted_at is None:
            raise handle_conflict("Order is not archived")
        order.deleted_at = None
        self.db.commit()
        return {"message": "Order restored successfully"}

    def delete_order(self, order_id: str):
        """Hard delete an order (permanent). Use archive for retention."""
        order = self._get_order_any(order_id)
        self.db.delete(order)
        self.db.commit()
        return {"message": "Order deleted successfully"}

    def bulk_delete_orders(self, order_ids: list[str]):
        """Delete multiple orders by ID. Returns count of deleted."""
        if not order_ids:
            return {"message": "No orders to delete", "deleted_count": 0}
        deleted = self.db.query(Order).filter(Order.id.in_(order_ids)).delete(synchronize_session=False)
        self.db.commit()
        return {"message": f"Deleted {deleted} order(s)", "deleted_count": deleted}

    def bulk_import_orders(self, orders_data: list[dict], user_id: str):
        """Bulk import orders from Excel data.
        
        Args:
            orders_data: List of dictionaries containing order data from Excel
            user_id: ID of the user performing the import
            
        Returns:
            dict with created, updated counts and errors
        """
        from datetime import datetime
        import re
        
        created = 0
        updated = 0
        errors = []
        warnings = []
        
        # Column mapping from Excel headers to database fields
        column_mapping = {
            'id': 'id',
            'ID': 'id',
            'order_number': 'order_number',
            'Order Number': 'order_number',
            'order_date': 'order_date',
            'Order Date': 'order_date',
            'estimated_delivery_date': 'estimated_delivery_date',
            'Estimated Delivery Date': 'estimated_delivery_date',
            'Promised Delivery Date': 'estimated_delivery_date',  # legacy Excel headers
            'promised_delivery_date': 'estimated_delivery_date',  # legacy column / export key
            'actual_delivery_date': 'actual_delivery_date',
            'Actual Delivery Date': 'actual_delivery_date',
            'customer_id': 'customer_id',
            'Customer ID': 'customer_id',
            'order_status_id': 'order_status_id',
            'Order Status ID': 'order_status_id',
            'subtotal_amount': 'subtotal_amount',
            'Subtotal Amount': 'subtotal_amount',
            'discount_amount': 'discount_amount',
            'Discount Amount': 'discount_amount',
            'tax_amount': 'tax_amount',
            'Tax Amount': 'tax_amount',
            'total_amount': 'total_amount',
            'Total Amount': 'total_amount',
            'remarks': 'remarks',
            'Remarks': 'remarks',
            'billing_address_id': 'billing_address_id',
            'Billing Address ID': 'billing_address_id',
            'shipping_address_id': 'shipping_address_id',
            'Shipping Address ID': 'shipping_address_id',
        }
        
        def parse_date(value):
            """Parse date from various formats."""
            if not value or value == '':
                return None
            if isinstance(value, datetime):
                return value
            if isinstance(value, str):
                # Try ISO format first
                try:
                    return datetime.fromisoformat(value.replace('Z', '+00:00'))
                except:
                    pass
                # Try other common formats
                for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y']:
                    try:
                        return datetime.strptime(value, fmt)
                    except:
                        continue
            return None
        
        def parse_decimal(value):
            """Parse decimal value."""
            if value is None or value == '':
                return Decimal("0")
            try:
                return Decimal(str(value))
            except:
                return Decimal("0")
        
        for idx, row_data in enumerate(orders_data, start=1):
            try:
                # Map Excel columns to database fields
                mapped_data = {}
                for excel_key, value in row_data.items():
                    db_key = column_mapping.get(excel_key, excel_key.lower())
                    if db_key in ['order_date', 'estimated_delivery_date', 'actual_delivery_date']:
                        mapped_data[db_key] = parse_date(value)
                    elif db_key in ['subtotal_amount', 'discount_amount', 'tax_amount', 'total_amount']:
                        mapped_data[db_key] = parse_decimal(value)
                    elif db_key in ['customer_id', 'order_status_id', 'billing_address_id', 'shipping_address_id']:
                        # Convert to string, handle empty values
                        mapped_data[db_key] = str(value).strip() if value and str(value).strip() else None
                    elif db_key in ['order_number', 'remarks']:
                        mapped_data[db_key] = str(value).strip() if value else None
                    elif db_key == 'id':
                        # ID is optional - if provided, will update; if not, will create
                        mapped_data[db_key] = str(value).strip() if value and str(value).strip() else None
                
                order_id = mapped_data.pop('id', None)
                
                # Validate required fields
                if not mapped_data.get('order_number'):
                    errors.append(f"Row {idx}: Order Number is required")
                    continue
                
                # Check if order exists (by ID first, then by order_number)
                existing_order = None
                
                # First, try to find by ID if provided
                if order_id:
                    existing_order = self.db.query(Order).filter(
                        Order.id == order_id,
                        Order.deleted_at.is_(None)
                    ).first()
                
                # If not found by ID, try to find by order_number
                if not existing_order:
                    existing_order = self.db.query(Order).filter(
                        Order.order_number == mapped_data['order_number'],
                        Order.deleted_at.is_(None)
                    ).first()
                
                # If creating new order, check for duplicate order_number
                if not existing_order:
                    duplicate = self.db.query(Order).filter(
                        Order.order_number == mapped_data['order_number'],
                        Order.deleted_at.is_(None)
                    ).first()
                    if duplicate:
                        errors.append(f"Row {idx}: Order Number '{mapped_data['order_number']}' already exists")
                        continue
                
                # Calculate total if not provided
                if 'total_amount' not in mapped_data or not mapped_data['total_amount']:
                    subtotal = mapped_data.get('subtotal_amount', Decimal("0")) or Decimal("0")
                    discount = mapped_data.get('discount_amount', Decimal("0")) or Decimal("0")
                    tax = mapped_data.get('tax_amount', Decimal("0")) or Decimal("0")
                    mapped_data['total_amount'] = subtotal - discount + tax
                
                if existing_order:
                    # Update existing order
                    for key, value in mapped_data.items():
                        if key != 'order_number':  # Don't update order_number
                            setattr(existing_order, key, value)
                    existing_order.updated_by = user_id
                    updated += 1
                else:
                    # Create new order
                    mapped_data['created_by'] = user_id
                    order = Order(**mapped_data)
                    self.db.add(order)
                    created += 1
                
            except Exception as e:
                error_msg = f"Row {idx}: {str(e)}"
                errors.append(error_msg)
                continue
        
        try:
            self.db.commit()
        except Exception as e:
            self.db.rollback()
            errors.append(f"Database error: {str(e)}")
        
        return {
            "created": created,
            "updated": updated,
            "errors": errors
        }

    def import_excel_tracking(self, file_data: bytes, user_id: str, validate_only: bool = False, outcome=None):
        """Import orders from Excel file with Master and Overall Tracking sheets. If validate_only=True, run validation and return errors/warnings without persisting (rollback)."""
        from io import BytesIO
        from datetime import datetime, date, time as dt_time, timedelta
        import openpyxl
        from decimal import Decimal

        from app.services import import_outcome_codes as _oc
        from app.services.import_outcome import ImportOutcome as _ImportOutcome

        # Per-row attribution for the job detail page. Row numbers repeat across the
        # two sheets, so identity carries the sheet name to disambiguate.
        if outcome is None:
            outcome = _ImportOutcome(None, persist=False)

        created = 0
        updated = 0
        errors = []
        warnings = []
        kpi_warnings = []
        import_session_id = str(uuid.uuid4())
        start_time = time.monotonic()
        master_rows = 0
        tracking_rows = 0
        calendar_service = CalendarService(self.db)

        # Resolve "new" and "delivered" status ids (case-insensitive match to status_code)
        status_rows = (
            self.db.query(OrderStatus)
            .filter(func.lower(OrderStatus.status_code).in_(["new", "delivered", "completed"]))
            .all()
        )
        status_by_code_lower = {str(s.status_code).lower(): s.id for s in status_rows}
        new_status_id = status_by_code_lower.get("new")
        delivered_status_id = status_by_code_lower.get("delivered")
        # Status ids that count as "delivered" for the complaint-DO freeze (date set
        # AND a delivered/completed status). Mirrors ComplaintFulfilmentService.
        delivered_status_ids = {
            str(sid)
            for code, sid in status_by_code_lower.items()
            if code in ("delivered", "completed") and sid
        }

        try:
            workbook = openpyxl.load_workbook(BytesIO(file_data), data_only=True)
        except Exception as exc:
            raise ValueError(f"Failed to read Excel file: {exc}") from exc

        logger.info("Order tracking import: opened workbook, sheets=%s", workbook.sheetnames)

        required_sheets = {"Master", "Overall Tracking"}
        if not required_sheets.issubset(set(workbook.sheetnames)):
            missing = required_sheets.difference(set(workbook.sheetnames))
            raise ValueError(f"Missing required sheet(s): {', '.join(sorted(missing))}")

        master_sheet = workbook["Master"]
        tracking_sheet = workbook["Overall Tracking"]

        master_mapping = {
            "Doc. No.": "order_number",
            "Date": "order_date",
            "Created Time": "created_time",
            "Debtor Code": "debtor_code",
            "Debtor Name": "debtor_name",
            "Agent": "agent",
            "Cancel": "is_cancelled",
            "Remarks CS": "remarks_cs",
            "Type": "order_type",
        }

        tracking_mapping = {
            "Doc Number": "order_number",
            "Doc No.": "order_number",
            "Date": "actual_delivery_date",
            "Delivery Date": "actual_delivery_date",
            "Time": "pickup_time",
            "Checker": "checker",
            "Transporter": "transporter",
            "Driver Name": "driver_name",
            "Lorry Plate": "lorry_plate",
            "Customer": "customer_ref",
            "Remarks CS": "delivery_remarks_cs",
            "Remarks": "delivery_remarks",
            "Salesman": "salesman",
            "Trips": "trips",
            "W/H": "warehouse",
            "Doc Date": "doc_date",
        }

        def normalize_header(value):
            return str(value).strip() if value is not None else ""

        def iter_sheet_rows(sheet, sheet_name_for_log=""):
            headers = [normalize_header(cell.value) for cell in sheet[1]]
            if sheet_name_for_log:
                logger.info("Order tracking import: sheet %r headers (row 1)=%s", sheet_name_for_log, headers)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(cell is not None and str(cell).strip() for cell in row):
                    continue
                row_data = {}
                for idx, value in enumerate(row):
                    header = headers[idx] if idx < len(headers) else ""
                    if header:
                        row_data[header] = value
                yield row_idx, row_data

        def chunked(values, size):
            for i in range(0, len(values), size):
                yield values[i:i + size]

        def normalize_order_number(value):
            if value is None:
                return ""
            return str(value).strip()

        master_rows_list = list(iter_sheet_rows(master_sheet, "Master"))
        tracking_rows_list = list(iter_sheet_rows(tracking_sheet, "Overall Tracking"))
        master_rows = len(master_rows_list)
        tracking_rows = len(tracking_rows_list)

        master_order_numbers = {
            normalize_order_number(row_data.get("Doc. No.")).lower()
            for _, row_data in master_rows_list
            if normalize_order_number(row_data.get("Doc. No."))
        }
        tracking_order_numbers = {
            normalize_order_number(row_data.get("Doc Number") or row_data.get("Doc No.")).lower()
            for _, row_data in tracking_rows_list
            if normalize_order_number(row_data.get("Doc Number") or row_data.get("Doc No."))
        }
        all_order_numbers = master_order_numbers.union(tracking_order_numbers)

        existing_orders = {}
        if all_order_numbers:
            for number_chunk in chunked(sorted(all_order_numbers), 500):
                orders = self.db.query(Order).filter(
                    func.lower(Order.order_number).in_(number_chunk),
                    Order.deleted_at.is_(None)
                ).all()
                for order in orders:
                    if order.order_number:
                        existing_orders[order.order_number.lower()] = order

        # Complaint <-> DO freeze: preload (once, batched) which existing orders are
        # already linked to a complaint and the complaint numbers, so the Master loop
        # can freeze Remarks CS on delivered+linked DOs without per-row queries.
        from app.models.complaints import ComplaintFulfilmentOrder, Complaint
        linked_order_ids: set[str] = set()
        linked_complaints_by_order: dict[str, list[str]] = {}
        _existing_ids = [str(o.id) for o in existing_orders.values() if getattr(o, "id", None)]
        for id_chunk in chunked(_existing_ids, 500):
            for oid, cnum in (
                self.db.query(ComplaintFulfilmentOrder.order_id, Complaint.complaint_number)
                .join(Complaint, Complaint.id == ComplaintFulfilmentOrder.complaint_id)
                .filter(ComplaintFulfilmentOrder.order_id.in_(id_chunk))
                .all()
            ):
                linked_order_ids.add(str(oid))
                linked_complaints_by_order.setdefault(str(oid), []).append(cnum)
        # Pre-change snapshot per order (remarks / delivered / cancelled), keyed by
        # lower order_number - captured BEFORE any setattr so the reconcile pass can
        # diff old vs new and run delta-only.
        pre_state: dict[str, dict] = {}

        def _snapshot_order(order_obj, number_lower: str) -> None:
            if number_lower in pre_state:
                return
            pre_state[number_lower] = {
                "remarks": getattr(order_obj, "remarks_cs", None),
                "delivered": getattr(order_obj, "actual_delivery_date", None) is not None,
                "cancelled": bool(getattr(order_obj, "is_cancelled", False)),
            }

        def parse_date_value(value, doc_date_value=None):
            if value is None or value == "":
                return None
            if isinstance(value, datetime):
                return value
            if isinstance(value, date):
                return datetime.combine(value, dt_time.min)
            if isinstance(value, str):
                cleaned = value.strip()
                for fmt in ["%d/%m/%Y", "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"]:
                    try:
                        return datetime.strptime(cleaned, fmt)
                    except ValueError:
                        continue
                try:
                    parsed = datetime.strptime(cleaned, "%d-%b")
                    year = doc_date_value.year if isinstance(doc_date_value, (datetime, date)) else datetime.now().year
                    return parsed.replace(year=year)
                except ValueError:
                    return None
            return None

        def parse_time_value(value):
            if value is None or value == "":
                return None
            if isinstance(value, dt_time):
                return value
            if isinstance(value, datetime):
                return value.time()
            if isinstance(value, str):
                cleaned = value.strip()
                for fmt in ["%I:%M %p", "%I:%M%p", "%H:%M", "%H:%M:%S"]:
                    try:
                        return datetime.strptime(cleaned, fmt).time()
                    except ValueError:
                        continue
            return None

        def parse_bool(value):
            if value is None or value == "":
                return False
            if isinstance(value, bool):
                return value
            return str(value).strip().upper() not in {"F", "FALSE", "0", "NO", "N"}

        def json_safe(value):
            if isinstance(value, (datetime, date, dt_time)):
                return value.isoformat()
            if isinstance(value, Decimal):
                return float(value)
            if isinstance(value, dict):
                return {str(k): json_safe(v) for k, v in value.items()}
            if isinstance(value, (list, tuple)):
                return [json_safe(v) for v in value]
            return value

        working_weekdays = calendar_service.get_working_weekdays()

        mapped_master_rows = []
        min_master_date = None
        max_master_date = None
        touched_orders: list[Order] = []

        # Process Master sheet
        for row_idx, row_data in master_rows_list:
            try:
                mapped = {}
                for header, value in row_data.items():
                    if header not in master_mapping:
                        continue
                    db_key = master_mapping[header]
                    if db_key in {"order_date", "created_time"}:
                        mapped[db_key] = parse_date_value(value)
                    elif db_key == "is_cancelled":
                        mapped[db_key] = parse_bool(value)
                    else:
                        mapped[db_key] = str(value).strip() if value is not None else None

                order_number = (mapped.get("order_number") or "").strip()
                if not order_number:
                    errors.append({"row": row_idx, "error": "Doc. No. is required", "data": row_data})
                    outcome.skip(
                        row=row_idx, code=_oc.MISSING_DOC_NO,
                        message="Doc. No. is required",
                        identity={"sheet": "Master"},
                    )
                    continue
                mapped["order_number"] = order_number
                order_date_value = mapped.get("order_date")
                if isinstance(order_date_value, (datetime, date)):
                    date_value = order_date_value.date() if isinstance(order_date_value, datetime) else order_date_value
                    min_master_date = date_value if min_master_date is None else min(min_master_date, date_value)
                    max_master_date = date_value if max_master_date is None else max(max_master_date, date_value)
                mapped_master_rows.append((row_idx, row_data, mapped))
            except Exception as exc:
                errors.append({"row": row_idx, "error": str(exc), "data": row_data})

        holidays_master = set()
        if min_master_date and max_master_date:
            holidays_master = calendar_service.get_public_holidays_between(
                min_master_date,
                max_master_date + timedelta(days=30),
            )

        for row_idx, row_data, mapped in mapped_master_rows:
            try:
                if mapped.get("order_date"):
                    mapped["estimated_delivery_date"] = calendar_service.add_business_days(
                        mapped["order_date"],
                        2,
                        working_weekdays=working_weekdays,
                        holidays=holidays_master,
                    )

                order_number = mapped["order_number"]
                number_lower = order_number.lower()
                existing_order = existing_orders.get(number_lower)

                if existing_order:
                    # Snapshot BEFORE mutation for the delta-only reconcile pass.
                    _snapshot_order(existing_order, number_lower)
                    # Freeze guard: a delivered DO linked to a complaint has historical,
                    # already-notified fulfilment - keep the DB Remarks CS, skip that
                    # field only, and warn (surfaces in test + real import alike).
                    if (
                        "remarks_cs" in mapped
                        and existing_order.actual_delivery_date is not None
                        and str(existing_order.order_status_id or "") in delivered_status_ids
                        and str(existing_order.id) in linked_order_ids
                        and (mapped.get("remarks_cs") or "").strip()
                        != (existing_order.remarks_cs or "").strip()
                    ):
                        _nums = ", ".join(
                            linked_complaints_by_order.get(str(existing_order.id), [])
                        )
                        warnings.append({
                            "row": row_idx,
                            "warning": (
                                f"Order {order_number}: Remarks CS change ignored - "
                                f"delivery order delivered and linked to complaint {_nums}"
                            ),
                            "data": row_data,
                        })
                        mapped.pop("remarks_cs", None)
                    # Upsert customer/transporter master rows from the debtor /
                    # transporter text so the FKs stay populated on import too
                    # (parity with update_order). Idempotent.
                    mapped = self._sync_order_master_refs(mapped)
                    for key, value in mapped.items():
                        if key != "order_number":
                            setattr(existing_order, key, value)
                    # Orders without actual delivery date should be "new"; Tracking will set
                    # "delivered" where applicable. But the Master sheet re-imports the FULL
                    # order history on every run while the Overall Tracking tab is only a
                    # partial/rolling window (real import_logs: master_rows routinely far
                    # exceeds tracking_rows) - resetting unconditionally wiped out a delivery
                    # already recorded by an earlier run just because this run's Tracking tab
                    # doesn't happen to include the order again. Only reset when the order
                    # doesn't already carry a real delivery.
                    already_delivered = (
                        existing_order.actual_delivery_date is not None
                        and str(existing_order.order_status_id or "") in delivered_status_ids
                    )
                    if new_status_id and not already_delivered:
                        existing_order.order_status_id = new_status_id
                    existing_order.updated_by = user_id
                    updated += 1
                    outcome.updated(
                        row=row_idx,
                        message=f"Order updated: {order_number}",
                        value=order_number,
                        identity={"sheet": "Master", "doc_no": order_number},
                        entity_type="order",
                        entity_id=getattr(existing_order, "id", None),
                    )
                    touched_orders.append(existing_order)
                else:
                    # Brand-new order: pre-change state is empty (no prior remarks /
                    # delivery / cancel), so any non-empty Remarks CS enters the reconcile.
                    pre_state.setdefault(
                        order_number.lower(),
                        {"remarks": None, "delivered": False, "cancelled": False},
                    )
                    mapped["created_by"] = user_id
                    if new_status_id:
                        mapped["order_status_id"] = new_status_id
                    # Upsert customer/transporter master rows from the debtor /
                    # transporter text so the FKs stay populated on import too
                    # (parity with create_order). Idempotent.
                    mapped = self._sync_order_master_refs(mapped)
                    order = Order(**mapped)
                    self.db.add(order)
                    created += 1
                    outcome.success(
                        row=row_idx,
                        message=f"Order created: {order_number}",
                        value=order_number,
                        identity={"sheet": "Master", "doc_no": order_number},
                        entity_type="order",
                    )
                    existing_orders[order_number.lower()] = order
                    touched_orders.append(order)
            except Exception as exc:
                errors.append({"row": row_idx, "error": str(exc), "data": row_data})
                outcome.fail(
                    row=row_idx, code=_oc.ROW_ERROR, message=str(exc),
                    identity={"sheet": "Master"},
                )

        tracking_updates = []
        min_tracking_date = None
        max_tracking_date = None

        # Process Overall Tracking sheet
        for row_idx, row_data in tracking_rows_list:
            try:
                mapped = {}
                doc_date_value = None
                for header, value in row_data.items():
                    if header not in tracking_mapping:
                        continue
                    db_key = tracking_mapping[header]
                    if db_key == "doc_date":
                        doc_date_value = parse_date_value(value)
                        continue
                    if db_key == "actual_delivery_date":
                        mapped[db_key] = parse_date_value(value, doc_date_value=doc_date_value)
                    elif db_key == "trips":
                        mapped[db_key] = int(value) if value not in (None, "") else None
                    else:
                        mapped[db_key] = str(value).strip() if value is not None else None

                order_number = (mapped.get("order_number") or "").strip()
                if not order_number:
                    errors.append({"row": row_idx, "error": "Doc Number is required", "data": row_data})
                    outcome.skip(
                        row=row_idx, code=_oc.MISSING_DOC_NO,
                        message="Doc Number is required",
                        identity={"sheet": "Overall Tracking"},
                    )
                    continue

                order = existing_orders.get(order_number.lower())
                if not order:
                    warnings.append({"row": row_idx, "warning": f"Order '{order_number}' not found in Master sheet", "data": row_data})
                    outcome.skip(
                        row=row_idx, code=_oc.ORDER_NOT_IN_MASTER,
                        message=f"Order '{order_number}' not found in Master sheet",
                        value=order_number,
                        identity={"sheet": "Overall Tracking", "doc_no": order_number},
                    )
                    continue

                # Snapshot pre-tracking state for tracking-only orders (Master-touched
                # ones were already snapshotted before their delivery date is applied).
                _snapshot_order(order, order_number.lower())

                delivery_date = mapped.get("actual_delivery_date")
                pickup_time = parse_time_value(row_data.get("Time"))
                if delivery_date:
                    if pickup_time:
                        mapped["actual_delivery_date"] = datetime.combine(
                            delivery_date.date() if isinstance(delivery_date, datetime) else delivery_date,
                            pickup_time,
                        )
                    else:
                        # Date only: store as midnight for DateTime column
                        d = delivery_date.date() if isinstance(delivery_date, datetime) else delivery_date
                        mapped["actual_delivery_date"] = datetime.combine(d, dt_time.min)
                    delivery_date = mapped["actual_delivery_date"]

                # Overall Tracking sheet carries the transporter text - upsert the
                # transporter master row + FK here (parity with update_order). Idempotent.
                mapped = self._sync_order_master_refs(mapped)
                for key, value in mapped.items():
                    if key != "order_number":
                        setattr(order, key, value)

                if order.order_date and delivery_date:
                    start_date_value = order.order_date.date() if isinstance(order.order_date, datetime) else order.order_date
                    end_date_value = delivery_date.date() if isinstance(delivery_date, datetime) else delivery_date
                    min_tracking_date = start_date_value if min_tracking_date is None else min(min_tracking_date, start_date_value)
                    max_tracking_date = end_date_value if max_tracking_date is None else max(max_tracking_date, end_date_value)
                    tracking_updates.append((order, delivery_date))
                    if delivered_status_id:
                        order.order_status_id = delivered_status_id
                else:
                    order.delivery_days = None
                    order.kpi_warning = False
                    if new_status_id:
                        order.order_status_id = new_status_id

                order.updated_by = user_id
                updated += 1
                outcome.updated(
                    row=row_idx,
                    message=f"Delivery tracking applied: {order_number}",
                    value=order_number,
                    identity={"sheet": "Overall Tracking", "doc_no": order_number},
                    entity_type="order",
                    entity_id=getattr(order, "id", None),
                )
                touched_orders.append(order)
            except Exception as exc:
                errors.append({"row": row_idx, "error": str(exc), "data": row_data})
                outcome.fail(
                    row=row_idx, code=_oc.ROW_ERROR, message=str(exc),
                    identity={"sheet": "Overall Tracking"},
                )

        holidays_tracking = set()
        if min_tracking_date and max_tracking_date:
            holidays_tracking = calendar_service.get_public_holidays_between(min_tracking_date, max_tracking_date)

        for order, delivery_date in tracking_updates:
            delivery_days = calendar_service.business_days_between(
                order.order_date,
                delivery_date,
                working_weekdays=working_weekdays,
                holidays=holidays_tracking,
            )
            order.delivery_days = delivery_days
            order.kpi_warning = delivery_days > 2
            if order.kpi_warning:
                kpi_warnings.append({"order_number": order.order_number, "days": delivery_days})

        logger.info(
            "Order tracking import: Master rows=%s, Tracking rows=%s, created=%s, updated=%s, errors=%s",
            master_rows, tracking_rows, created, updated, len(errors),
        )
        for i, err in enumerate(errors[:5]):
            if isinstance(err, dict):
                data = err.get("data")
                if isinstance(data, dict):
                    data_keys = list(data.keys())
                elif isinstance(data, list):
                    data_keys = f"list({len(data)})"
                else:
                    data_keys = type(data).__name__ if data is not None else None
                logger.warning(
                    "Order tracking import error [%s]: row=%s %s | data keys=%s",
                    i + 1,
                    err.get("row"),
                    err.get("error"),
                    data_keys,
                )
            else:
                logger.warning("Order tracking import error [%s]: %s", i + 1, err)
        for i, warn in enumerate(warnings[:5]):
            if isinstance(warn, dict):
                data = warn.get("data")
                if isinstance(data, dict):
                    data_keys = list(data.keys())
                elif isinstance(data, list):
                    data_keys = f"list({len(data)})"
                else:
                    data_keys = type(data).__name__ if data is not None else None
                logger.warning(
                    "Order tracking import warning [%s]: row=%s %s | data keys=%s",
                    i + 1,
                    warn.get("row"),
                    warn.get("warning"),
                    data_keys,
                )
            else:
                logger.warning("Order tracking import warning [%s]: %s", i + 1, warn)
        if len(errors) > 5:
            logger.warning("Order tracking import: ... and %s more errors", len(errors) - 5)

        # Complaint <-> DO auto-fulfilment: reconcile links + recompute fulfilment
        # for orders whose Remarks CS / delivery / cancel actually changed (delta-only;
        # one batched complaint lookup). On validate_only it computes warnings only and
        # writes nothing. Real-path link/status writes ride the commit below; the per-DO
        # delivery notifications are dispatched AFTER that commit (best-effort).
        fulfilment_notify_payloads: list[dict] = []
        changes: list[dict] = []
        _seen_change: set[str] = set()
        for o in touched_orders:
            num_lower = (getattr(o, "order_number", None) or "").lower()
            if not num_lower or num_lower in _seen_change:
                continue
            _seen_change.add(num_lower)
            pre = pre_state.get(num_lower) or {"remarks": None, "delivered": False, "cancelled": False}
            new_remarks = getattr(o, "remarks_cs", None)
            new_delivered = getattr(o, "actual_delivery_date", None) is not None
            new_cancelled = bool(getattr(o, "is_cancelled", False))
            if (
                (new_remarks or "").strip() != (pre["remarks"] or "").strip()
                or new_delivered != pre["delivered"]
                or new_cancelled != pre["cancelled"]
            ):
                changes.append({"order": o, "old_remarks": pre["remarks"]})
        if changes:
            try:
                from app.services.complaint_fulfilment_service import (
                    ComplaintFulfilmentService,
                )

                _f_warnings, fulfilment_notify_payloads = ComplaintFulfilmentService(
                    self.db
                ).reconcile_links_and_status(changes, dry_run=validate_only)
                for _w in _f_warnings:
                    warnings.append({"row": None, "warning": _w, "data": None})
            except Exception:
                logger.exception("Order import: complaint fulfilment reconcile failed")

        if validate_only:
            self.db.rollback()
            err_list = [e.get("error", str(e)) if isinstance(e, dict) else str(e) for e in errors]
            warn_list = [w.get("warning", str(w)) if isinstance(w, dict) else str(w) for w in warnings]
            return {
                "valid": len(errors) == 0,
                "errors": err_list,
                "warnings": warn_list,
                "summary": {
                    "master_rows": master_rows,
                    "tracking_rows": tracking_rows,
                    "would_create": created,
                    "would_update": updated,
                    "error_count": len(errors),
                    "warning_count": len(warnings),
                    "kpi_warnings": kpi_warnings,
                },
            }

        try:
            with suppress_embedding_events("order"):
                self.db.flush()
                seen: set[str] = set()
                touched_ids: list[str] = []
                for o in touched_orders:
                    oid = getattr(o, "id", None)
                    if oid is None:
                        continue
                    sid = str(oid)
                    if sid in seen:
                        continue
                    seen.add(sid)
                    touched_ids.append(sid)
                enqueued = bulk_enqueue_embedding_events(
                    self.db, "order", touched_ids, event_type="order.updated"
                )
                logger.info(
                    "Order tracking import: bulk-enqueued %s embedding events (touched=%s)",
                    enqueued, len(touched_orders),
                )
                self.db.commit()
            _import_commit_ok = True
        except Exception as exc:
            self.db.rollback()
            _import_commit_ok = False
            errors.append({"row": None, "error": f"Database error: {exc}", "data": None})
            logger.exception("Order tracking import: database commit failed")

        # Per-DO delivery notifications (customer + Complaint team) - only after the
        # link/status writes committed. Best-effort; never fail the import.
        if _import_commit_ok and fulfilment_notify_payloads:
            try:
                from app.services.complaint_fulfilment_service import (
                    ComplaintFulfilmentService,
                )

                ComplaintFulfilmentService(self.db).dispatch_delivery_notifications(
                    fulfilment_notify_payloads
                )
            except Exception:
                logger.exception("Order import: fulfilment delivery notify dispatch failed")

        duration_ms = int((time.monotonic() - start_time) * 1000)
        if not validate_only:
            try:
                import_log_service = ImportLogService(self.db)
                import_log_service.create_import_log(
                    entity_type="order",
                    entity_table="orders",
                    import_session_id=import_session_id,
                    filename=None,
                    import_type="EXCEL_IMPORT",
                    total_rows=master_rows + tracking_rows,
                    successful_rows=created + updated,
                    created_rows=created,
                    updated_rows=updated,
                    failed_rows=len(errors),
                    skipped_rows=len(warnings),
                    warnings=json_safe(warnings),
                    errors=json_safe(errors),
                    summary=json_safe({"kpi_warnings": kpi_warnings, "master_rows": master_rows, "tracking_rows": tracking_rows, "warnings": len(warnings)}),
                    imported_by=user_id,
                    duration_ms=duration_ms,
                )
            except Exception:
                pass

        return {
            "import_session_id": import_session_id,
            "created": created,
            "updated": updated,
            "errors": errors,
            "warnings": warnings,
            "kpi_warnings": kpi_warnings,
            "master_rows": master_rows,
            "tracking_rows": tracking_rows,
        }

    def validate_delivery_order_detail_excel(self, file_data: bytes) -> dict:
        """Validate delivery order detail import file (order lines) without writing to DB."""
        import openpyxl
        from app.api.v1.external.utils import normalize_code

        def _normalize_header(value) -> str:
            if value is None:
                return ""
            return str(value).strip().lower()

        def _find(row_d: dict, *candidates: str):
            for c in candidates:
                key = c.lower().strip()
                if key in row_d:
                    return row_d.get(key)
            return None

        try:
            workbook = openpyxl.load_workbook(BytesIO(file_data), data_only=True)
        except Exception as exc:
            raise ValueError(f"Failed to read Excel file: {exc}") from exc

        sheet = workbook.active
        if not sheet:
            raise ValueError("Workbook has no active sheet")

        headers = [_normalize_header(cell.value) for cell in sheet[1]]
        parsed_rows = []
        all_doc_nos = set()
        all_product_codes = set()
        all_locations = set()
        errors: list[str] = []

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue

            row_data: dict = {}
            for idx, value in enumerate(row):
                if idx < len(headers) and headers[idx]:
                    row_data[headers[idx]] = value

            doc_no = (_find(row_data, "doc no", "doc number", "order number") and str(_find(row_data, "doc no", "doc number", "order number")).strip()) or None
            item_code = (_find(row_data, "item code", "product code") and str(_find(row_data, "item code", "product code")).strip()) or None
            location = (_find(row_data, "location", "warehouse", "warehouse code") and str(_find(row_data, "location", "warehouse", "warehouse code")).strip()) or None

            # Ignore trailing / formatted "empty" rows (e.g. stray 0, F, or styled blanks) with no line identity
            if not doc_no and not item_code and not location:
                continue

            if not doc_no:
                errors.append(f"Row {row_idx}: Missing doc no")
            if not item_code:
                errors.append(f"Row {row_idx}: Missing item code")
            if not location:
                errors.append(f"Row {row_idx}: Missing location")

            if doc_no:
                all_doc_nos.add(doc_no)
            if item_code:
                all_product_codes.add(item_code)
            if location:
                all_locations.add(location)

            parsed_rows.append((row_idx, doc_no, item_code, location))

        orders_by_number = {}
        if all_doc_nos:
            for order in self.db.query(Order).filter(Order.order_number.in_(all_doc_nos), Order.deleted_at.is_(None)).all():
                orders_by_number[order.order_number] = order

        products_by_code = {}
        if all_product_codes:
            for product in self.db.query(Product).filter(Product.product_code.in_(all_product_codes)).all():
                products_by_code[(product.product_code or "").strip()] = product

        warehouses_map = {}
        if all_locations:
            all_warehouses = self.db.query(Warehouse).all()
            for wh in all_warehouses:
                if wh.warehouse_code:
                    warehouses_map[normalize_code(wh.warehouse_code)] = wh
                if wh.warehouse_name:
                    warehouses_map[normalize_code(wh.warehouse_name)] = wh

        for row_idx, doc_no, item_code, location in parsed_rows:
            if not doc_no or not item_code or not location:
                continue
            if doc_no not in orders_by_number:
                errors.append(f"Row {row_idx}: Order not found: {doc_no}")
            if item_code not in products_by_code:
                errors.append(f"Row {row_idx}: Product not found: {item_code}")
            if normalize_code(location) not in warehouses_map:
                errors.append(f"Row {row_idx}: Warehouse not found: {location}")

        total_rows = len(parsed_rows)
        return {
            "valid": len(errors) == 0,
            "errors": errors,
            "warnings": [],
            "summary": {
                "rows": total_rows,
                "error_count": len(errors),
            },
        }


class CustomerService:
    """Service for customer operations."""
    
    def __init__(self, db: Session):
        self.db = db
    
    # Whitelisted sort columns for the customers list. Anything else falls back
    # to created_at so an unknown sort param can't crash the query.
    _CUSTOMER_SORT_MAP = {
        "customer_code": Customer.customer_code,
        "customer_name": Customer.customer_name,
        "email": Customer.email,
        "phone_number": Customer.phone_number,
        "is_active": Customer.is_active,
        "created_at": Customer.created_at,
        "updated_at": Customer.updated_at,
    }

    def _build_customer_list_query(
        self,
        query: Optional[str] = None,
        sort_field: str = "created_at",
        sort_dir: str = "desc",
    ):
        """Build the filtered + sorted customers query shared by ``list_customers``
        and ``neighbours`` so the two can never drift.

        The ORDER BY always appends ``Customer.id`` as a deterministic tie-breaker
        so offset position and prev/next neighbours are unambiguous when the
        primary sort column has equal values.
        """
        q = self.db.query(Customer)

        if query:
            q = q.filter(
                or_(
                    Customer.customer_code.ilike(f"%{query}%"),
                    Customer.customer_name.ilike(f"%{query}%"),
                )
            )

        sort_column = self._CUSTOMER_SORT_MAP.get(sort_field, Customer.created_at)
        if sort_dir == "desc":
            q = q.order_by(sort_column.desc(), Customer.id.asc())
        else:
            q = q.order_by(sort_column.asc(), Customer.id.asc())
        return q

    def list_customers(
        self,
        page: int = 1,
        limit: int = 50,
        query: Optional[str] = None,
        sort_field: str = "created_at",
        sort_dir: str = "desc",
    ):
        """List customers with pagination, search, and sorting."""
        q = self._build_customer_list_query(
            query=query,
            sort_field=sort_field,
            sort_dir=sort_dir,
        )

        total = q.count()
        offset = (page - 1) * limit
        customers = q.offset(offset).limit(limit).all()

        return {
            "data": customers,
            "pagination": {"total": total, "page": page, "limit": limit},
            "empty": total == 0
        }

    def neighbours(
        self,
        customer_id: str,
        query: Optional[str] = None,
        sort_field: str = "created_at",
        sort_dir: str = "desc",
    ) -> dict:
        """Resolve prev/next neighbours for ``customer_id`` within the active list
        query.

        Selects only the ordered ids (not full rows), then defers the position/wrap
        math to the pure ``compute_neighbours`` helper. If the record is not in the
        filtered set (deep link, or filtered out after an edit), falls back to the
        unfiltered, default-sorted set so the pager is never dead (D2).
        """
        from app.services.record_navigation import compute_neighbours

        def _ordered_ids(q) -> list[str]:
            return [str(row[0]) for row in q.with_entities(Customer.id).all()]

        filtered_q = self._build_customer_list_query(
            query=query,
            sort_field=sort_field,
            sort_dir=sort_dir,
        )
        result = compute_neighbours(_ordered_ids(filtered_q), customer_id)
        if result["index"] is not None:
            return result

        # D2: current record not in the filtered set -> fall back to the unfiltered,
        # default-sorted set so prev/next still works and total reflects all customers.
        unfiltered_q = self._build_customer_list_query()
        return compute_neighbours(_ordered_ids(unfiltered_q), customer_id)

    def get_customer(self, customer_id: str):
        """Get a customer by ID."""
        customer = self.db.query(Customer).filter(Customer.id == customer_id).first()
        if not customer:
            raise handle_not_found("Customer", customer_id)
        return customer
    
    def create_customer(self, customer_data: CustomerCreate):
        """Create a new customer.

        Uniqueness is on the (customer_code, customer_name) pair - case +
        whitespace insensitive - so the same Sage code can legitimately host
        multiple debtor names (e.g. "300-D093" for "Deluxe Home Center (KTN)"
        and "Deluxe Home Center AC (I)").
        """
        code = (customer_data.customer_code or "").strip()
        name = (customer_data.customer_name or "").strip()
        existing = self.db.query(Customer).filter(
            func.lower(func.btrim(Customer.customer_code)) == code.lower(),
            func.lower(func.btrim(Customer.customer_name)) == name.lower(),
        ).first()
        if existing:
            raise handle_conflict("Customer with this code + name already exists.")

        customer = Customer(**customer_data.model_dump())
        self.db.add(customer)
        self.db.commit()
        self.db.refresh(customer)
        return customer
    
    def update_customer(self, customer_id: str, customer_data: CustomerUpdate):
        """Update a customer."""
        customer = self.get_customer(customer_id)
        
        update_data = customer_data.model_dump(exclude_unset=True)
        for key, value in update_data.items():
            setattr(customer, key, value)
        
        self.db.commit()
        self.db.refresh(customer)
        return customer


class OrderStatusService:
    """Service for order status operations."""
    
    def __init__(self, db: Session):
        self.db = db
    
    def list_order_statuses(self, page: int = 1, limit: int = 50):
        """List order statuses."""
        q = self.db.query(OrderStatus).order_by(OrderStatus.sequence)
        
        total = q.count()
        offset = (page - 1) * limit
        statuses = q.offset(offset).limit(limit).all()
        
        return {
            "data": statuses,
            "pagination": {"total": total, "page": page, "limit": limit},
            "empty": total == 0
        }
    
    def get_order_status(self, status_id: str):
        """Get an order status by ID."""
        status = self.db.query(OrderStatus).filter(OrderStatus.id == status_id).first()
        if not status:
            raise handle_not_found("Order Status", status_id)
        return status
    
    def create_order_status(self, status_data: OrderStatusCreate):
        """Create a new order status."""
        existing = self.db.query(OrderStatus).filter(
            OrderStatus.status_code == status_data.status_code
        ).first()
        if existing:
            raise handle_conflict("Status code already exists.")
        
        status = OrderStatus(**status_data.model_dump())
        self.db.add(status)
        self.db.commit()
        self.db.refresh(status)
        return status
    
    def update_order_status(self, status_id: str, status_data: OrderStatusUpdate):
        """Update an order status."""
        status = self.get_order_status(status_id)
        
        update_data = status_data.model_dump(exclude_unset=True)
        for key, value in update_data.items():
            setattr(status, key, value)
        
        self.db.commit()
        self.db.refresh(status)
        return status
