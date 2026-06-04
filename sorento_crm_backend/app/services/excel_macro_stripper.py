"""Strip VBA macros from `.xlsm` workbooks and re-emit as clean `.xlsx`.

Used at every Excel upload boundary so `.xlsm` bytes never reach the import
pipeline or persistent storage. `openpyxl` re-reads the workbook with
`keep_vba=False` (the default), which drops `xl/vbaProject.bin` and the
`bin`/`xlsm` content-types on save.
"""
from __future__ import annotations

from io import BytesIO
from typing import Iterable, Optional, Tuple

from openpyxl import load_workbook


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

TEMPLATE_SHEET_NAME = "Template"


class MacroWorkbookError(ValueError):
    """Raised when a `.xlsm` upload can't be reduced to a single data sheet.

    Either the workbook is unreadable, or it has multiple sheets and none is
    named `Template`. Routes translate this to a 422 so the uploader gets an
    actionable message instead of a 500.
    """


def is_xlsm_filename(filename: str | None) -> bool:
    return bool(filename) and filename.lower().endswith(".xlsm")


def _prune_sheets(wb, keep: Iterable[str]) -> None:
    """Remove all sheets except those whose names are in `keep`."""
    keep_set = {str(name) for name in keep}
    for name in list(wb.sheetnames):
        if name not in keep_set:
            del wb[name]


def strip_macros_to_xlsx(
    file_bytes: bytes,
    filename: str,
    *,
    keep_sheets: Optional[Iterable[str]] = None,
) -> Tuple[bytes, str]:
    """Return (cleaned_bytes, new_filename). No-op when filename is not `.xlsm` and no `keep_sheets`."""
    is_xlsm = is_xlsm_filename(filename)
    if not is_xlsm and not keep_sheets:
        return file_bytes, filename
    wb = load_workbook(BytesIO(file_bytes), data_only=False, keep_vba=False)
    if keep_sheets is not None:
        _prune_sheets(wb, keep_sheets)
    out = BytesIO()
    wb.save(out)
    new_name = filename[:-5] + ".xlsx" if is_xlsm else filename
    return out.getvalue(), new_name


def list_sheet_names(file_bytes: bytes) -> list[str]:
    """Cheap read-only peek for workbook sheet names (empty list when unreadable)."""
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=False, keep_vba=False, read_only=True)
    except Exception:
        return []
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def first_sheet_name(file_bytes: bytes) -> Optional[str]:
    names = list_sheet_names(file_bytes)
    return names[0] if names else None


def _resolve_data_sheet(
    file_bytes: bytes, preferred: Optional[Iterable[str]]
) -> Optional[str]:
    """Return first sheet name from `preferred` that exists in the workbook.

    Match is case-insensitive. Falls back to the workbook's first sheet when
    no preferred name matches. Returns None for unreadable workbooks.
    """
    names = list_sheet_names(file_bytes)
    if not names:
        return None
    if preferred:
        lookup = {n.lower(): n for n in names}
        for candidate in preferred:
            hit = lookup.get(str(candidate).lower())
            if hit:
                return hit
    return names[0]


def extract_macro_template_xlsx(
    file_bytes: bytes,
    filename: str | None,
    content_type: str | None,
    *,
    template_sheet: str = TEMPLATE_SHEET_NAME,
    require_data: bool = False,
) -> Tuple[bytes, str | None, str | None]:
    """Stock-list macro pipeline: `.xlsm` → values-only `.xlsx` with just the Template sheet.

    Non-`.xlsm` files pass through untouched (the `.xls`/`.xlsx` flow must not
    change). For `.xlsm`:
    - VBA is dropped (`keep_vba=False`) and formulas are baked to their cached
      values (`data_only=True`) so pruning sibling sheets can't leave `#REF!`
      in the file n8n/the chatbot reads.
    - Sheet selection: `template_sheet` (case-insensitive) if present; a
      single-sheet workbook keeps its only sheet; multi-sheet without the
      template raises `MacroWorkbookError`.
    - `require_data=True` additionally raises `MacroWorkbookError` when the
      kept sheet has no data rows below the header — macro workbooks whose
      Template the user forgot to populate (run the macro) before uploading.
    Returns (bytes, filename, mime) with the `.xlsm` suffix and MIME rewritten
    to `.xlsx`.
    """
    if not is_xlsm_filename(filename):
        return file_bytes, filename, content_type
    assert filename is not None

    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True, keep_vba=False)
    except Exception as exc:
        raise MacroWorkbookError(
            "Uploaded .xlsm file could not be read as an Excel workbook."
        ) from exc

    names = list(wb.sheetnames)
    lookup = {n.lower(): n for n in names}
    target = lookup.get(template_sheet.lower())
    if target is None:
        if len(names) == 1:
            target = names[0]
        else:
            raise MacroWorkbookError(
                f"Macro workbook must contain a '{template_sheet}' sheet "
                f"(found sheets: {', '.join(names)})."
            )

    if require_data:
        ws = wb[target]
        has_data = any(
            any(cell is not None and str(cell).strip() for cell in row)
            for row in ws.iter_rows(min_row=2, values_only=True)
        )
        if not has_data:
            raise MacroWorkbookError(
                f"The '{target}' sheet has no data rows. Run the macro to "
                "populate it, save, and upload again."
            )

    _prune_sheets(wb, [target])
    out = BytesIO()
    wb.save(out)
    return out.getvalue(), filename[:-5] + ".xlsx", XLSX_MIME


def maybe_strip(file_bytes: bytes, filename: str | None) -> Tuple[bytes, str]:
    """Convenience: pass through when filename is None/empty."""
    if not filename:
        return file_bytes, filename or ""
    return strip_macros_to_xlsx(file_bytes, filename)


def maybe_strip_upload(
    file_bytes: bytes,
    filename: str | None,
    content_type: str | None,
    *,
    keep_data_sheet: bool = False,
    data_sheet_candidates: Optional[Iterable[str]] = None,
) -> Tuple[bytes, str, str | None]:
    """Strip macros if `.xlsm`; optionally retain only the data sheet importers read.

    When `keep_data_sheet=True` the workbook is always re-saved through openpyxl
    so extra sheets are removed even from clean `.xlsx` uploads. `data_sheet_candidates`
    is a priority list of sheet names to look for (case-insensitive); the first hit
    wins, and the workbook's first sheet is used as a fallback. MIME normalizes to
    xlsx whenever a rewrite happens.
    """
    is_xlsm = is_xlsm_filename(filename or "")
    if not is_xlsm and not keep_data_sheet:
        return file_bytes, filename or "", content_type

    keep_arg: Optional[Iterable[str]] = None
    if keep_data_sheet:
        target = _resolve_data_sheet(file_bytes, data_sheet_candidates)
        keep_arg = [target] if target else None

    new_bytes, new_name = strip_macros_to_xlsx(file_bytes, filename or "", keep_sheets=keep_arg)
    new_mime = XLSX_MIME if (is_xlsm or keep_data_sheet) else content_type
    return new_bytes, new_name, new_mime
