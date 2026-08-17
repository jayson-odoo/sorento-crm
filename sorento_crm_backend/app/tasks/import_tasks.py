"""Background tasks for imports."""
import json
import os
import re
import time
import zipfile
import hashlib
import mimetypes
import logging
from collections import defaultdict
from datetime import date, datetime, time as dt_time
from decimal import Decimal
from io import BytesIO
from typing import Callable, Optional, List, Any, Dict, cast

from sqlalchemy import func

from app.database import SessionLocal
from app.services.inventory_service import StockService, WarehouseService
from app.services.order_service import OrderService
from app.services.product_service import ProductService
from app.services.import_log_service import ImportLogService
from app.services.job_service import JobService
from app.services.audit_service import log_import_audit
from app.services.procurement_service import (
    SPOAllocationService,
    PickingHeaderService,
    AllocationReceivedGuardError,
)
from app.services.grn_spo_matching import (
    build_allocation_pool,
    draw_fifo,
    forward_match_grn_lines_for_spo_best_effort,
)
from app.services.resources_service import (
    AttachmentDirectoryService,
    AttachmentService,
    AttachmentTypeService,
)
from app.services.attachment_webhook_helper import create_and_send_webhook
from app.api.v1.external.utils import (
    get_inbound_shipment_by_container_number,
    get_products_by_code_exact,
    get_warehouses_by_code_or_name,
    parse_date_value,
)
from app.models.job import ImportJob, JobStatus
from app.services import import_outcome_codes as oc
from app.services.import_outcome import ImportOutcome
from app.services.company_scope import set_company_scope
from app.models.order import Order, OrderLine
from app.schemas.resources import AttachmentCreate
from app.schemas.procurement import SPOAllocationCreate

logger = logging.getLogger(__name__)

# Sentinel for "the caller did not tell us its company scope". Distinct from
# ``None``, which is a REAL scope meaning all companies.
_SCOPE_NOT_GIVEN = object()


def _apply_preview_scope(db, company_scope: Any) -> None:
    """Give a validation preview the SAME company scope the real import will run at.

    A preview that reads across all companies answers a different question than the
    import: it resolves products, warehouses and existing headers the scoped import
    cannot see, so it reports "would succeed" on rows the import then skips as
    not-found. Preview and import must disagree about nothing.

    The caller is an HTTP route, so it already holds the resolved scope on its own
    session (``get_company_scope(db)``) — it passes that through rather than us
    re-deriving it. ``_SCOPE_NOT_GIVEN`` keeps non-HTTP callers (scripts, tests)
    working system-scoped, with a warning so a route that forgets is visible.
    """
    if company_scope is _SCOPE_NOT_GIVEN:
        logger.warning(
            "Import validation preview ran with no company scope; reading all companies"
        )
        set_company_scope(db, None)
        return
    set_company_scope(db, company_scope)


def _apply_import_job_scope(db, db_job_id: Optional[str]) -> None:
    """Re-establish the request's company scope on the worker session (multi-company
    isolation, AC-D2/D3/K4).

    The worker registers the scope enforcement but every ``SessionLocal()`` starts
    at UNSET (fail-closed), which would (a) make owned reads return 0 rows and
    (b) make owned inserts raise (nothing to auto-stamp). We read the ImportJob's
    ``company_id`` snapshot (captured at enqueue) and set a single-company scope so
    owned inserts auto-stamp + reads isolate. A NULL snapshot (system / None-scope
    import, or a pre-isolation job) runs system-scoped (``None`` = all companies)
    with a warning — back-compat for the pre-multi-company behaviour.
    """
    company_id = None
    if db_job_id:
        try:
            job = db.query(ImportJob).filter(ImportJob.id == db_job_id).first()
            company_id = getattr(job, "company_id", None) if job else None
        except Exception:
            company_id = None
    if company_id:
        set_company_scope(db, frozenset({str(company_id)}))
    else:
        logger.warning(
            "Import job %s has no company snapshot; running system-scoped (all companies)",
            db_job_id,
        )
        set_company_scope(db, None)


# Batch size for attachment bulk import (files per batch); sleep between batches to avoid overload
ATTACHMENT_BULK_IMPORT_BATCH_SIZE = 5
ATTACHMENT_BULK_IMPORT_BATCH_DELAY_SECONDS = 0.5


def _normalize_zip_path(name: str) -> str:
    return name.replace("\\", "/").strip("/")


def _is_macos_metadata_path(normalized_path: str) -> bool:
    """True if path is macOS metadata (._* files or under __MACOSX)."""
    if not normalized_path:
        return False
    parts = [p for p in normalized_path.split("/") if p.strip()]
    if any(p == "__MACOSX" for p in parts):
        return True
    if parts and parts[-1].strip().startswith("._"):
        return True
    return False


def _write_import_audit(
    db,
    *,
    entity_type: str,
    label: str,
    row_count,
    user_id: Optional[str],
    entity_id: Optional[str],
    status: str = "success",
    details_builder: Optional[Callable[[], Dict[str, Any]]] = None,
) -> None:
    """Coarse per-job import audit at the job boundary. Best-effort (post-commit
    side effect): a failure here must NEVER break the import, so we swallow and
    warn. Bulk imports bypass the ORM audit listener, so this is the only audit
    row an import job produces. Commits the audit row on the same session AFTER
    the import data (and the ImportJob status) have already committed.

    ``row_count`` may be an int OR a zero-argument callable that computes one. The
    callable form exists because computing the figure is itself part of the side
    effect: evaluated at the call site it runs OUTSIDE this guard, so a caller whose
    result dict is shaped unexpectedly raises AFTER ``complete_job`` has committed -
    and the enclosing handler then FAILS a job that actually succeeded. Inside the
    try it costs a warning, which is what a best-effort side effect is supposed to
    cost.

    ``details_builder`` produces the row's ``new_values``: what the run did, plus
    the import job id that reaches the per-row outcomes in ``import_job_rows``.
    A CALLABLE, not a dict, for the same reason as the callable ``row_count``:
    built at the call site it would run after the import has already committed
    and outside this try, where a stale attribute read (the job row is expired by
    that commit) would escape into the task's own except clause and mark a
    finished import FAILED."""
    try:
        log_import_audit(
            db,
            entity_type=entity_type,
            label=label,
            row_count=row_count() if callable(row_count) else row_count,
            user_id=user_id,
            entity_id=entity_id,
            status=status,
            details=details_builder() if details_builder is not None else None,
        )
        db.commit()
    except Exception:
        logger.warning(
            "Import audit write failed (entity_type=%s, entity_id=%s)",
            entity_type,
            entity_id,
            exc_info=True,
        )
        try:
            db.rollback()
        except Exception:
            pass


def _json_safe(value):
    if isinstance(value, (datetime, date, dt_time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(v) for v in value]
    return value


def process_stock_import(db_job_id: str, stock_data: list, user_id: str):
    """Process stock import in background."""
    from rq import get_current_job
    
    db = SessionLocal()
    _apply_import_job_scope(db, db_job_id)
    job_service = JobService(db)
    
    # Get RQ job ID
    rq_job = get_current_job()
    rq_job_id = rq_job.id if rq_job else None
    
    # Find job by DB ID or RQ job ID
    job = job_service.get_job_by_db_id(db_job_id) if db_job_id else None
    if not job and rq_job_id:
        job = job_service.get_job(rq_job_id)
    
    if not job:
        logger.error(f"Job not found: db_job_id={db_job_id}, rq_job_id={rq_job_id}")
        db.close()
        return

    job_id_str: str = str(job.job_id)
    outcome = ImportOutcome(getattr(job, "id", None), session_factory=SessionLocal)
    try:
        # Mark job as started
        job_service.start_job(job_id_str)
        
        # Process import
        stock_service = StockService(db)
        result = stock_service.bulk_import_stock(stock_data, user_id, outcome=outcome)
        
        # Mark job as completed
        job_service.complete_job(
            job_id=job_id_str,
            result=outcome.finalize(
                "Stock import completed",
                total_rows=len(stock_data),
                # legacy keys, kept one release
                created=result['created'],
                updated=result['updated'],
                skipped=result['skipped'],
                errors=result['errors'],
                warnings=result['warnings'],
                import_session_id=result['import_session_id'],
            ),
            **outcome.completion_counts(total_rows=len(stock_data)),
        )

        logger.info(f"Stock import job {job.job_id} completed successfully")
        _write_import_audit(
            db,
            entity_type="inbound_shipment",
            label=f"Stock import {job.filename or ''}".strip(),
            row_count=result['created'] + result['updated'],
            user_id=user_id,
            entity_id=job_id_str,
            status="partial" if result['errors'] else "success",
        )

    except Exception as e:
        logger.error(f"Stock import job {job.job_id} failed: {str(e)}", exc_info=True)
        # A crashed import still owes the operator the rows it did classify. The
        # recorder writes on its own session, so flushing here survives the rollback.
        outcome.flush()
        job_service.fail_job(job_id_str, str(e))
        _write_import_audit(
            db,
            entity_type="inbound_shipment",
            label=f"Stock import {job.filename or ''}".strip(),
            row_count=0,
            user_id=user_id,
            entity_id=job_id_str,
            status="failed",
        )
    finally:
        db.close()


def process_warehouse_import(db_job_id: str, warehouses_data: list, user_id: str):
    """Process warehouse import in background. Upserts by warehouse_code."""
    from rq import get_current_job

    db = SessionLocal()
    _apply_import_job_scope(db, db_job_id)
    job_service = JobService(db)

    rq_job = get_current_job()
    rq_job_id = rq_job.id if rq_job else None

    job = job_service.get_job_by_db_id(db_job_id) if db_job_id else None
    if not job and rq_job_id:
        job = job_service.get_job(rq_job_id)

    if not job:
        logger.error(f"Job not found: db_job_id={db_job_id}, rq_job_id={rq_job_id}")
        db.close()
        return

    job_id_str: str = str(job.job_id)
    outcome = ImportOutcome(getattr(job, "id", None), session_factory=SessionLocal)
    try:
        job_service.start_job(job_id_str)

        warehouse_service = WarehouseService(db)
        result = warehouse_service.bulk_import_warehouses(
            warehouses_data, user_id, outcome=outcome
        )

        job_service.complete_job(
            job_id=job_id_str,
            result=outcome.finalize(
                "Warehouse import completed",
                total_rows=len(warehouses_data),
                # legacy keys, kept one release
                created=result["created"],
                updated=result["updated"],
                skipped=result["skipped"],
                errors=result["errors"],
                warnings=result["warnings"],
                import_session_id=result["import_session_id"],
            ),
            **outcome.completion_counts(total_rows=len(warehouses_data)),
        )

        logger.info(f"Warehouse import job {job.job_id} completed successfully")
        _write_import_audit(
            db,
            entity_type="warehouse",
            label=f"Warehouse import {job.filename or ''}".strip(),
            row_count=result["created"] + result["updated"],
            user_id=user_id,
            entity_id=job_id_str,
            status="partial" if result["errors"] else "success",
        )

    except Exception as e:
        logger.error(f"Warehouse import job {job.job_id} failed: {str(e)}", exc_info=True)
        # A crashed import still owes the operator the rows it did classify. The
        # recorder writes on its own session, so flushing here survives the rollback.
        outcome.flush()
        job_service.fail_job(job_id_str, str(e))
        _write_import_audit(
            db,
            entity_type="warehouse",
            label=f"Warehouse import {job.filename or ''}".strip(),
            row_count=0,
            user_id=user_id,
            entity_id=job_id_str,
            status="failed",
        )
    finally:
        db.close()


def process_product_import(db_job_id: str, products_data: list, user_id: str):
    """Process product import in background."""
    from rq import get_current_job

    db = SessionLocal()
    _apply_import_job_scope(db, db_job_id)
    job_service = JobService(db)

    rq_job = get_current_job()
    rq_job_id = rq_job.id if rq_job else None

    job = job_service.get_job_by_db_id(db_job_id) if db_job_id else None
    if not job and rq_job_id:
        job = job_service.get_job(rq_job_id)

    if not job:
        logger.error(f"Job not found: db_job_id={db_job_id}, rq_job_id={rq_job_id}")
        db.close()
        return

    job_id_str: str = str(job.job_id)
    outcome = ImportOutcome(getattr(job, "id", None), session_factory=SessionLocal)
    try:
        job_service.start_job(job_id_str)
        job_service.update_job_progress(job_id_str, total_rows=len(products_data))

        def on_progress(processed: int, successful: int, failed: int, skipped: int) -> None:
            counts = outcome.completion_counts()
            counts["processed_rows"] = processed
            job_service.update_job_progress(job_id_str, **counts)

        product_service = ProductService(db)
        result = product_service.bulk_import_products(
            products_data, user_id, on_progress=on_progress, outcome=outcome
        )

        job_service.complete_job(
            job_id=job_id_str,
            result=outcome.finalize(
                "Product import completed",
                total_rows=len(products_data),
                # Master data the import had to create for itself (unknown Item
                # Group / Item Brand / UOM), so the operator can see what appeared.
                created_categories=result.get('created_categories', 0),
                created_brands=result.get('created_brands', 0),
                created_uoms=result.get('created_uoms', 0),
                # AutoCount's reorder columns, when the file carried them.
                levels_applied=result.get('levels_applied', 0),
                levels_cleared=result.get('levels_cleared', 0),
                level_conflicts=result.get('level_conflicts', 0),
                # Rendered by the job detail page's existing Warnings list: a planning
                # level a person set that this file disagreed with, named so it can be
                # settled rather than only counted.
                warnings=result.get('level_conflict_warnings', []),
                # legacy keys, kept one release
                created=result['created'],
                updated=result['updated'],
                errors=result['errors'][:200],
            ),
            **outcome.completion_counts(total_rows=len(products_data)),
        )

        logger.info(
            "Product import job %s completed: created=%s, updated=%s, errors=%s, "
            "new_categories=%s, new_brands=%s, new_uoms=%s",
            job_id_str, result['created'], result['updated'], len(result['errors']),
            result.get('created_categories', 0), result.get('created_brands', 0),
            result.get('created_uoms', 0),
        )
        _write_import_audit(
            db,
            entity_type="product",
            label=f"Product import {job.filename or ''}".strip(),
            row_count=result['created'] + result['updated'],
            user_id=user_id,
            entity_id=job_id_str,
            status="partial" if result['errors'] else "success",
        )
    except Exception as e:
        logger.error("Product import job %s failed: %s", job.job_id, str(e), exc_info=True)
        # A crashed import still owes the operator the rows it did classify. The
        # recorder writes on its own session, so flushing here survives the rollback.
        outcome.flush()
        job_service.fail_job(job_id_str, str(e))
        _write_import_audit(
            db,
            entity_type="product",
            label=f"Product import {job.filename or ''}".strip(),
            row_count=0,
            user_id=user_id,
            entity_id=job_id_str,
            status="failed",
        )
    finally:
        db.close()


def process_order_tracking_import(db_job_id: str, file_data: bytes, user_id: str):
    """Process order tracking import in background."""
    from rq import get_current_job
    
    db = SessionLocal()
    _apply_import_job_scope(db, db_job_id)
    job_service = JobService(db)
    
    # Get RQ job ID
    rq_job = get_current_job()
    rq_job_id = rq_job.id if rq_job else None
    
    # Find job by DB ID or RQ job ID
    job = job_service.get_job_by_db_id(db_job_id) if db_job_id else None
    if not job and rq_job_id:
        job = job_service.get_job(rq_job_id)
    
    if not job:
        logger.error(f"Job not found: db_job_id={db_job_id}, rq_job_id={rq_job_id}")
        db.close()
        return
    
    job_id_str: str = str(job.job_id)
    outcome = ImportOutcome(getattr(job, "id", None), session_factory=SessionLocal)
    try:
        # Mark job as started
        job_service.start_job(job_id_str)
        
        # Process import
        order_service = OrderService(db)
        result = order_service.import_excel_tracking(file_data, user_id, outcome=outcome)

        created = result.get('created', 0)
        updated = result.get('updated', 0)
        errors = result.get('errors', [])
        warnings = result.get('warnings', [])
        master_rows = result.get('master_rows', 0)
        tracking_rows = result.get('tracking_rows', 0)
        failed_count = len(errors)
        successful_rows = created + updated
        processed_rows = master_rows + tracking_rows
        total_rows = processed_rows

        # Mark job as completed with correct counts
        job_service.complete_job(
            job_id=job_id_str,
            result=outcome.finalize(
                "Order tracking import completed",
                total_rows=total_rows,
                # legacy keys, kept one release
                created=created,
                updated=updated,
                failed=failed_count,
                master_rows=master_rows,
                tracking_rows=tracking_rows,
                kpi_warnings=result.get('kpi_warnings', []),
                import_session_id=result.get('import_session_id'),
                errors=_json_safe(errors[:50]),
                warnings=_json_safe(warnings[:50]),
            ),
            **outcome.completion_counts(total_rows=total_rows),
        )

        logger.info(
            "Order tracking import job %s completed: processed=%s (Master=%s, Tracking=%s), created=%s, updated=%s, failed=%s",
            job_id_str, processed_rows, master_rows, tracking_rows, created, updated, failed_count,
        )
        if errors:
            for i, err in enumerate(errors[:3]):
                logger.warning("Order tracking import job %s error [%s]: row=%s %s", job_id_str, i + 1, err.get('row'), err.get('error'))
        if warnings:
            for i, warn in enumerate(warnings[:3]):
                logger.warning("Order tracking import job %s warning [%s]: row=%s %s", job_id_str, i + 1, warn.get('row'), warn.get('warning'))

        _write_import_audit(
            db,
            entity_type="order",
            label=f"Order tracking import {job.filename or ''}".strip(),
            row_count=successful_rows,
            user_id=user_id,
            entity_id=job_id_str,
            status="partial" if failed_count else "success",
        )

    except Exception as e:
        logger.error("Order tracking import job %s failed: %s", job_id_str, str(e), exc_info=True)
        db.rollback()
        # A crashed import still owes the operator the rows it did classify. The
        # recorder writes on its own session, so flushing here survives the rollback.
        outcome.flush()
        job_service.fail_job(job_id_str, str(e))
        try:
            import_log_service = ImportLogService(db)
            import_log_service.create_import_log(
                entity_type="order",
                entity_table="orders",
                import_session_id=job_id_str,
                filename=str(job.filename) if (job and job.filename is not None) else None,
                import_type="EXCEL_IMPORT",
                total_rows=0,
                successful_rows=0,
                created_rows=0,
                updated_rows=0,
                failed_rows=0,
                skipped_rows=0,
                warnings=None,
                errors=_json_safe([{"row": None, "error": str(e), "data": None}]),
                summary=None,
                imported_by=user_id,
                duration_ms=None,
            )
        except Exception:
            pass
        _write_import_audit(
            db,
            entity_type="order",
            label=f"Order tracking import {job.filename or ''}".strip(),
            row_count=0,
            user_id=user_id,
            entity_id=job_id_str,
            status="failed",
        )
    finally:
        db.close()


def process_attachment_bulk_import(
    db_job_id: str,
    storage_key: str,
    attachment_type_id: str,
    access_levels_json: str,
    parent_directory_id: Optional[str],
    user_id: str,
    on_conflict: str = "skip",
    storage_provider: Optional[str] = None,
):
    """Process attachment bulk import (ZIP) in background with batch processing.

    TCK-2026-000020 ZIP collision behaviour (`on_conflict`):
      * `skip` (default): drop the second occurrence (intra-zip dupe) or keep
        the existing system row (zip-vs-system dupe). Counted as `skipped`.
      * `copy`: rename the incoming file with the Google-Drive style suffix
        loop until free in the resolved directory.
      * `replace`: update the colliding live attachment row in place
        (preserves attachment_id + every linkage) and re-fires the
        `attachment_replaced` webhook.

    Detection key: `(resolved directory_id, lower(filename))` — same as the
    interactive single-upload path. Empty `directory_id` (root) is also
    scoped correctly because the partial unique index on attachments only
    applies when directory_id IS NOT NULL.
    """
    on_conflict = (on_conflict or "skip").strip().lower()
    if on_conflict not in ("skip", "copy", "replace"):
        on_conflict = "skip"
    from rq import get_current_job

    db = SessionLocal()
    _apply_import_job_scope(db, db_job_id)
    job_service = JobService(db)

    rq_job = get_current_job()
    rq_job_id = rq_job.id if rq_job else None

    job = job_service.get_job_by_db_id(db_job_id) if db_job_id else None
    if not job and rq_job_id:
        job = job_service.get_job(rq_job_id)

    # Storage routing for the staged ZIP. `storage_provider` arg is the value
    # captured at enqueue time; falls back to current default for older queued
    # jobs that pre-date the cross-pod fix.
    from app.services.storage_router import (
        cdn_base_url,
        default_provider,
        extract_key,
        get_backend,
        sanitize_storage_filename,
    )
    import uuid as _uuid

    zip_storage_provider = storage_provider or default_provider()
    zip_storage_backend = get_backend(zip_storage_provider)

    def _cleanup_staged_zip() -> None:
        try:
            zip_storage_backend.delete_file(storage_key)
        except Exception:
            logger.warning("bulk-import zip cleanup failed for key=%s", storage_key)

    if not job:
        logger.error("Attachment bulk import job not found: db_job_id=%s, rq_job_id=%s", db_job_id, rq_job_id)
        db.close()
        _cleanup_staged_zip()
        return

    job_id_str: str = str(job.job_id)
    outcome = ImportOutcome(getattr(job, "id", None), session_factory=SessionLocal)
    # Attachment is __audit_track__; suppress the per-row ORM audit for this bulk
    # job (no request actor here → would log N rows as "System"). One coarse,
    # correctly-attributed job row is written at completion instead.
    # setdefault-union, not assignment: two suppressions on one session must coexist.
    db.info.setdefault("skip_audit_entity_types", set()).add("attachment")
    try:
        dir_service = AttachmentDirectoryService(db)
        attachment_service = AttachmentService(db)
        type_service = AttachmentTypeService(db)
        # Attachment-write provider — separate variable from the staged-zip
        # provider so a future split (e.g. zips on R2, writes on S3) is just a
        # config change.
        storage_provider = default_provider()
        storage_backend = get_backend(storage_provider)

        # Pull the staged zip into memory ONCE. zipfile.ZipFile is re-opened
        # multiple times below (namelist + per-file read); BytesIO is cheap to
        # rewind, avoids re-downloading on every batch entry.
        try:
            zip_bytes = zip_storage_backend.download_file(storage_key)
        except Exception as exc:
            job_service.fail_job(job_id_str, f"Failed to fetch staged zip from storage: {exc}")
            db.close()
            _cleanup_staged_zip()
            return
        zip_buffer = BytesIO(zip_bytes)

        access_levels_payload = None
        try:
            parsed = json.loads(access_levels_json or "[]")
            if isinstance(parsed, list):
                access_levels_payload = parsed
        except Exception:
            pass
        from app.services.contact_access_type_service import ContactAccessTypeService
        access_svc = ContactAccessTypeService(db)
        if access_levels_payload:
            try:
                access_levels_payload = access_svc.validate_access_levels(access_levels_payload, field_name="access_levels")
            except Exception:
                access_levels_payload = access_svc.get_default_access_levels()
        else:
            access_levels_payload = access_svc.get_default_access_levels()

        job_service.start_job(job_id_str)

        try:
            attachment_type = type_service.get_type(attachment_type_id)
        except Exception:
            job_service.fail_job(job_id_str, "Invalid attachment type ID")
            db.close()
            return

        _allowed_raw = attachment_type.allowed_extensions
        _allowed_str = str(_allowed_raw) if _allowed_raw is not None else ""
        allowed_extensions: set[str] = set(
            ext.strip().lower().replace(".", "")
            for ext in _allowed_str.split(",")
            if ext.strip()
        )
        _mb = attachment_type.max_file_size_mb
        max_bytes = (int(cast(int, _mb)) if _mb is not None else 10) * 1024 * 1024

        zip_buffer.seek(0)
        with zipfile.ZipFile(zip_buffer, "r") as zf:
            all_names = [_normalize_zip_path(n) for n in zf.namelist()]

        dir_paths = set()
        file_paths: List[str] = []
        for name in all_names:
            if not name:
                continue
            if _is_macos_metadata_path(name):
                continue
            if name.endswith("/"):
                dir_path = name.rstrip("/")
                if dir_path and not _is_macos_metadata_path(dir_path):
                    dir_paths.add(dir_path)
            else:
                file_paths.append(name)

        # Create all directories first
        for dir_path in sorted(dir_paths):
            parts = [p for p in dir_path.split("/") if p.strip()]
            if parts:
                dir_service.get_or_create_path(parent_directory_id, parts)

        total_files = len(file_paths)
        job_service.update_job_progress(job_id_str, total_rows=total_files, result={"directories_created": len(dir_paths)})

        created_attachments: List[dict] = []
        errors: List[str] = []
        successful = 0
        failed = 0
        skipped = 0
        processed = 0
        # TCK-2026-000020 collision tracking.
        collisions_skipped = 0
        collisions_renamed_copy = 0
        collisions_replaced = 0
        # Intra-zip dedup key: (directory_id_or_None, lower(filename)).
        seen_in_run: set[tuple[Optional[str], str]] = set()
        # Helpers from the single-upload path. Local import avoids module-level
        # circular dependency (attachments route imports services; this task
        # module also imports services).
        from app.api.v1.resources.attachments import (
            _suffix_copy_name,
            _find_filename_collision,
        )

        def _name_taken_in_dir(_dir_id: Optional[str], _name: str) -> bool:
            key = (_dir_id, _name.lower())
            if key in seen_in_run:
                return True
            if _dir_id and _find_filename_collision(db, _dir_id, _name) is not None:
                return True
            return False

        # Zip entries have no spreadsheet row, so the entry's 1-based position in the
        # archive is what the operator can match against.
        file_row_index = {path: idx for idx, path in enumerate(file_paths, start=1)}

        for i in range(0, len(file_paths), ATTACHMENT_BULK_IMPORT_BATCH_SIZE):
            batch = file_paths[i : i + ATTACHMENT_BULK_IMPORT_BATCH_SIZE]
            for file_path in batch:
                entry_row = file_row_index.get(file_path)
                entry_identity = {"file": file_path}
                try:
                    zip_buffer.seek(0)
                    with zipfile.ZipFile(zip_buffer, "r") as zf:
                        raw_name = next((n for n in zf.namelist() if _normalize_zip_path(n) == file_path), None)
                    if not raw_name:
                        errors.append(f"Not found in zip: {file_path}")
                        outcome.fail(
                            row=entry_row, code=oc.NOT_FOUND_IN_ZIP,
                            message=f"Not found in zip: {file_path}",
                            value=file_path, identity=entry_identity,
                        )
                        failed += 1
                        processed += 1
                        continue
                    zip_buffer.seek(0)
                    with zipfile.ZipFile(zip_buffer, "r") as zf:
                        with zf.open(raw_name, "r") as entry:
                            file_content = entry.read()

                    original_filename = file_path.split("/")[-1]
                    ext = (original_filename.split(".")[-1] or "").lower()
                    if allowed_extensions and ext not in allowed_extensions:
                        errors.append(f"Skipped (extension .{ext} not allowed): {file_path}")
                        outcome.skip(
                            row=entry_row, code=oc.EXTENSION_NOT_ALLOWED,
                            message=f"Extension .{ext} is not allowed for this attachment type",
                            value=f".{ext}", identity=entry_identity,
                        )
                        skipped += 1
                        processed += 1
                        continue
                    if len(file_content) > max_bytes:
                        errors.append(f"Skipped (file too large): {file_path}")
                        outcome.skip(
                            row=entry_row, code=oc.FILE_TOO_LARGE,
                            message=(
                                f"File is {len(file_content)} bytes, over the "
                                f"{max_bytes} byte limit for this attachment type"
                            ),
                            value=file_path, identity=entry_identity,
                        )
                        skipped += 1
                        processed += 1
                        continue

                    dir_parts = [p for p in file_path.split("/")[:-1] if p.strip()]
                    directory_id = dir_service.get_or_create_path(parent_directory_id, dir_parts)

                    # ------------------------------------------------------------------
                    # Collision detection (intra-zip + zip-vs-system).
                    # ------------------------------------------------------------------
                    intra_key = (directory_id, original_filename.lower())
                    intra_collision = intra_key in seen_in_run
                    sys_collision = (
                        _find_filename_collision(db, directory_id, original_filename)
                        if directory_id
                        else None
                    )

                    existing_to_replace = None
                    if intra_collision or sys_collision is not None:
                        if on_conflict == "skip":
                            errors.append(
                                f"Skipped (filename already exists in target folder): {file_path}"
                            )
                            outcome.skip(
                                row=entry_row, code=oc.FILENAME_COLLISION,
                                message="A file with this name already exists in the target folder",
                                value=original_filename, identity=entry_identity,
                            )
                            collisions_skipped += 1
                            skipped += 1
                            processed += 1
                            continue
                        if on_conflict == "copy":
                            candidate = _suffix_copy_name(original_filename)
                            while _name_taken_in_dir(directory_id, candidate):
                                candidate = _suffix_copy_name(candidate)
                            original_filename = candidate
                            collisions_renamed_copy += 1
                        else:  # replace
                            if sys_collision is None:
                                # Intra-zip-only collision: cannot "replace" a row that
                                # doesn't exist yet in the system, so rename the second
                                # occurrence to keep both.
                                candidate = _suffix_copy_name(original_filename)
                                while _name_taken_in_dir(directory_id, candidate):
                                    candidate = _suffix_copy_name(candidate)
                                original_filename = candidate
                                collisions_renamed_copy += 1
                            else:
                                existing_to_replace = sys_collision

                    # Mirror the single-upload contract (PLAN-attachment-key-uuid-segregation):
                    #   stored_filename   = raw display name (editable, what the folder dup-check matches)
                    #   original_filename = sanitized, immutable → the object-key basename
                    #   key = {entity_type}/{attachment_id}/{basename} — the uuid dir makes every
                    #   key unique, so two same-named files across folders can NEVER clobber.
                    bulk_attachment_id = str(_uuid.uuid4())
                    key_basename = sanitize_storage_filename(original_filename) or "file"
                    stored_filename = original_filename
                    entity_type = (attachment_type.type_name or "general").lower().replace(" ", "_")
                    if existing_to_replace is not None:
                        # Replace overwrites the existing object's own key (no orphan).
                        s3_file_path = extract_key(existing_to_replace.file_path) or (
                            f"{entity_type}/{existing_to_replace.id}/{key_basename}"
                        )
                    else:
                        s3_file_path = f"{entity_type}/{bulk_attachment_id}/{key_basename}"
                    guessed_type, _ = mimetypes.guess_type(original_filename)
                    s3_key, _ = storage_backend.upload_file(
                        file_content=file_content,
                        file_path=s3_file_path,
                        content_type=guessed_type,
                    )
                    # Grid thumbnail (images only) — same small-variant path as the
                    # single-file upload route so bulk-imported photos also render
                    # at ~320px in the Files grid. Best-effort; never fails import.
                    from app.services.image_thumbnailer import store_thumbnail
                    bulk_thumbnail_path = store_thumbnail(
                        storage_backend, storage_provider, s3_key, file_content, guessed_type
                    )

                    if existing_to_replace is not None:
                        # Replace-in-place: preserve attachment_id + every linkage.
                        from datetime import datetime as _dt

                        existing_to_replace.attachment_type_id = attachment_type_id  # type: ignore[assignment]
                        existing_to_replace.stored_filename = stored_filename  # type: ignore[assignment]
                        existing_to_replace.file_path = cdn_base_url(storage_provider, s3_key)  # type: ignore[assignment]
                        existing_to_replace.thumbnail_path = bulk_thumbnail_path  # type: ignore[assignment]
                        existing_to_replace.file_size_bytes = len(file_content)  # type: ignore[assignment]
                        existing_to_replace.mime_type = guessed_type or "application/octet-stream"  # type: ignore[assignment]
                        existing_to_replace.file_hash = hashlib.sha256(file_content).hexdigest()  # type: ignore[assignment]
                        existing_to_replace.uploaded_by = user_id  # type: ignore[assignment]
                        existing_to_replace.uploaded_at = _dt.utcnow()  # type: ignore[assignment]
                        existing_to_replace.access_levels = access_levels_payload  # type: ignore[assignment]
                        existing_to_replace.storage_provider = storage_provider  # type: ignore[assignment]
                        # Tag with the import_job.job_id so the upload-activity
                        # endpoint can group every file in this ZIP into one
                        # session row. See documentation/plans/PLAN-upload-activity-drawer.md §4.2.
                        existing_to_replace.upload_batch_id = job_id_str  # type: ignore[assignment]
                        db.commit()
                        db.refresh(existing_to_replace)
                        attachment = existing_to_replace
                        try:
                            create_and_send_webhook(
                                db,
                                attachment,
                                attachment_type,
                                access_levels_payload,
                                user_id,
                                event_type="attachment_replaced",
                            )
                        except Exception as e:
                            logger.warning(
                                "Webhook (replace) creation failed for %s: %s",
                                attachment.id,
                                e,
                            )
                        collisions_replaced += 1
                        created_attachments.append({"id": attachment.id, "path": file_path, "replaced": True})
                        outcome.updated(
                            row=entry_row, code=oc.REPLACED,
                            message=f"Replaced existing attachment: {stored_filename}",
                            value=stored_filename, identity=entry_identity,
                            entity_type="attachment", entity_id=attachment.id,
                        )
                        successful += 1
                        seen_in_run.add((directory_id, original_filename.lower()))
                        processed += 1
                        continue

                    attachment_data = AttachmentCreate(
                        id=bulk_attachment_id,
                        attachment_type_id=attachment_type_id,
                        original_filename=key_basename,
                        stored_filename=stored_filename,
                        file_path=cdn_base_url(storage_provider, s3_key),
                        thumbnail_path=bulk_thumbnail_path,
                        file_size_bytes=len(file_content),
                        mime_type=guessed_type or "application/octet-stream",
                        file_hash=hashlib.sha256(file_content).hexdigest(),
                        entity_type=None,
                        entity_id=None,
                        directory_id=directory_id,
                        access_levels=access_levels_payload,
                        storage_provider=storage_provider,
                        # Tag with the import_job.job_id so the upload-activity
                        # endpoint can group every file in this ZIP into one
                        # session row. See documentation/plans/PLAN-upload-activity-drawer.md §4.2.
                        upload_batch_id=job_id_str,
                    )
                    attachment = attachment_service.create_attachment(attachment_data, user_id)
                    if attachment is not None:
                        try:
                            create_and_send_webhook(
                                db, attachment, attachment_type, access_levels_payload, user_id
                            )
                        except Exception as e:
                            logger.warning("Webhook creation failed for %s: %s", attachment.id, e)
                        created_attachments.append({"id": attachment.id, "path": file_path})
                    outcome.success(
                        row=entry_row,
                        code=oc.RENAMED_COPY if original_filename != file_path.split("/")[-1] else oc.CREATED,
                        message=f"Uploaded: {original_filename}",
                        value=original_filename, identity=entry_identity,
                        entity_type="attachment",
                        entity_id=getattr(attachment, "id", None) if attachment is not None else None,
                    )
                    successful += 1
                    seen_in_run.add((directory_id, original_filename.lower()))
                except Exception as e:
                    errors.append(f"{file_path}: {e}")
                    logger.exception("Bulk import file failed: %s", file_path)
                    outcome.fail(
                        row=entry_row, code=oc.ROW_ERROR, message=str(e),
                        value=file_path, identity=entry_identity,
                    )
                    failed += 1
                processed += 1

            job_service.update_job_progress(
                job_id_str,
                processed_rows=processed,
                successful_rows=successful,
                failed_rows=failed,
                skipped_rows=skipped,
                result={
                    "directories_created": len(dir_paths),
                    "attachments_created": len(created_attachments),
                    "attachments": created_attachments[-100:],  # Last 100 for response size
                    "errors": errors[-50:],
                    "on_conflict": on_conflict,
                    "collisions_skipped": collisions_skipped,
                    "collisions_renamed_copy": collisions_renamed_copy,
                    "collisions_replaced": collisions_replaced,
                },
            )
            if i + len(batch) < len(file_paths):
                time.sleep(ATTACHMENT_BULK_IMPORT_BATCH_DELAY_SECONDS)

        outcome_result = outcome.finalize(
            "Bulk import completed",
            total_rows=total_files,
            # legacy keys, kept one release
            directories_created=len(dir_paths),
            attachments_created=len(created_attachments),
            attachments=created_attachments,
            errors=errors,
            on_conflict=on_conflict,
            collisions_skipped=collisions_skipped,
            collisions_renamed_copy=collisions_renamed_copy,
            collisions_replaced=collisions_replaced,
        )
        job_service.complete_job(
            job_id=job_id_str,
            result=outcome_result,
            **outcome.completion_counts(total_rows=total_files),
        )
        # Coarse job-level audit: one correctly-attributed row for the whole ZIP
        # (per-row attachment audit is suppressed above for this bulk job).
        _write_import_audit(
            db,
            entity_type="attachment",
            label="Attachment bulk import (ZIP)",
            row_count=successful,
            user_id=user_id,
            entity_id=job_id_str,
            status="partial" if failed else "success",
        )
        logger.info(
            "Attachment bulk import job %s completed: %s files, %s created, %s failed, %s skipped",
            job_id_str, processed, successful, failed, skipped,
        )
        return outcome_result
    except Exception as e:
        logger.exception("Attachment bulk import job %s failed", job_id_str)
        # A crashed import still owes the operator the rows it did classify. The
        # recorder writes on its own session, so flushing here survives the rollback.
        outcome.flush()
        job_service.fail_job(job_id_str, str(e))
    finally:
        db.close()
        _cleanup_staged_zip()


def _spo_import_normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


_SPO_NO_PACKING_LIST_REASON = "Packing list not found for container"


def partition_spo_skip_reasons(
    skipped_rows_detail: List[Dict[str, Any]],
) -> tuple[List[str], List[str]]:
    """Split SPO import skips into blocking errors vs. "no packing list" warnings.

    A row whose only problem is a missing inbound shipment (packing list) is a
    warning, not an error: that line is skipped but the rest of the file still
    imports, so the uploader is warned rather than blocked. Every other reason
    (bad qty, unknown product/warehouse, missing container) stays an error.
    Used by BOTH validate_spo_import (test) and process_spo_import so the two
    surface identical messages.
    """
    errors: List[str] = []
    warnings: List[str] = []
    for s in skipped_rows_detail:
        msg = f"Row {s['row']}: {s['reason']}"
        if str(s.get("reason", "")).startswith(_SPO_NO_PACKING_LIST_REASON):
            warnings.append(msg)
        else:
            errors.append(msg)
    return errors, warnings


def _spo_import_find_column(row_data: dict, *candidates: str) -> Any:
    """Return first matching column value (case-insensitive key match)."""
    keys_lower = {k.lower(): k for k in row_data if k}
    for c in candidates:
        cl = c.lower().strip()
        if cl in keys_lower:
            return row_data.get(keys_lower[cl])
    return None


def _spo_import_extract_container(loading_date_value: Any) -> Optional[str]:
    """Extract shipping container number: text after first space in Loading Date cell."""
    if loading_date_value is None:
        return None
    s = str(loading_date_value).strip()
    if not s:
        return None
    parts = s.split(None, 1)  # max 2 parts
    if len(parts) < 2:
        return None
    return parts[1].strip() or None


def process_spo_import(db_job_id: str, file_data: bytes, filename: str, user_id: str):
    """Process SPO allocation import from Excel in background.

    Parsing rules:
    - Filename (without extension) = SPO number.
    - Item Code = product code.
    - Loading Date = cell text; text after first space = shipping container number (link to inbound shipment).
    - Location = warehouse code.
    - Qty = allocated quantity.

    Rows are grouped by (spo_number, product_id, warehouse_id); quantities are summed.
    One allocation is created per group, linked to the first valid inbound shipment for that group.
    """
    from rq import get_current_job
    import openpyxl

    db = SessionLocal()
    _apply_import_job_scope(db, db_job_id)
    job_service = JobService(db)
    rq_job = get_current_job()
    rq_job_id = rq_job.id if rq_job else None

    job = job_service.get_job_by_db_id(db_job_id) if db_job_id else None
    if not job and rq_job_id:
        job = job_service.get_job(rq_job_id)

    if not job:
        logger.error("SPO import job not found: db_job_id=%s, rq_job_id=%s", db_job_id, rq_job_id)
        db.close()
        return

    job_id_str: str = str(job.job_id)
    outcome = ImportOutcome(getattr(job, "id", None), session_factory=SessionLocal)

    try:
        job_service.start_job(job_id_str)

        spo_number = re.sub(r"\.xlsx?$", "", filename or "", flags=re.IGNORECASE).strip()
        if not spo_number:
            job_service.fail_job(job_id_str, "Filename must provide SPO number (e.g. SPO-2025.10-0050.xlsx)")
            db.close()
            return

        try:
            workbook = openpyxl.load_workbook(BytesIO(file_data), data_only=True)
        except Exception as exc:
            job_service.fail_job(job_id_str, f"Failed to read Excel file: {exc}")
            db.close()
            return

        sheet = workbook.active
        if not sheet:
            job_service.fail_job(job_id_str, "Workbook has no active sheet")
            db.close()
            return

        from app.api.v1.external.utils import normalize_code

        headers = [_spo_import_normalize_header(cell.value) for cell in sheet[1]]
        # Collect all data rows (non-empty) and extract codes/locations for lookup
        data_rows: List[tuple[int, dict]] = []  # (row_idx, row_data dict)
        all_product_codes = set()
        all_locations = set()

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            row_data = {}
            for idx, value in enumerate(row):
                if idx < len(headers) and headers[idx]:
                    row_data[headers[idx]] = value
            item_code_raw = _spo_import_find_column(
                row_data, "Item Code", "Item code", "Product Code", "Product code"
            )
            location_raw = _spo_import_find_column(
                row_data, "Location", "Warehouse", "Warehouse Code"
            )
            qty_raw = _spo_import_find_column(
                row_data, "Qty", "Quantity", "Allocated", "Allocated Quantity"
            )
            loading_date_raw = _spo_import_find_column(
                row_data, "Loading Date", "Loading date"
            )
            item_code = (item_code_raw and str(item_code_raw).strip()) or None
            location = (location_raw and str(location_raw).strip()) or None
            try:
                qty = int(float(qty_raw)) if qty_raw is not None else 0
            except (TypeError, ValueError):
                qty = 0
            container = _spo_import_extract_container(loading_date_raw)
            if item_code:
                all_product_codes.add(item_code)
            if location:
                all_locations.add(location)
            data_rows.append((row_idx, row_data))

        total_data_rows = len(data_rows)
        if not total_data_rows:
            job_service.complete_job(
                job_id=job_id_str,
                result=outcome.finalize("No valid data rows found", total_rows=0),
                **outcome.completion_counts(total_rows=0),
            )
            db.close()
            return

        products_by_code = get_products_by_code_exact(db, all_product_codes)
        warehouses_map = get_warehouses_by_code_or_name(db, all_locations)

        resolved_rows: List[tuple[str, str, str, int, int]] = []  # product_id, warehouse_id, shipment_id, qty, row_idx
        skipped_rows_detail: List[dict] = []  # legacy key, kept one release
        # Allocations are upserted per (product, warehouse) group, but the operator
        # asked which spreadsheet ROW landed - keep the mapping to fan the verdict back.
        spo_group_rows: Dict[tuple[str, str], List[int]] = defaultdict(list)
        spo_row_identity: Dict[tuple[str, str], Dict[str, Any]] = {}

        for row_idx, row_data in data_rows:
            item_code_raw = _spo_import_find_column(
                row_data, "Item Code", "Item code", "Product Code", "Product code"
            )
            loading_date_raw = _spo_import_find_column(
                row_data, "Loading Date", "Loading date"
            )
            location_raw = _spo_import_find_column(
                row_data, "Location", "Warehouse", "Warehouse Code"
            )
            qty_raw = _spo_import_find_column(
                row_data, "Qty", "Quantity", "Allocated", "Allocated Quantity"
            )

            item_code = (item_code_raw and str(item_code_raw).strip()) or None
            location = (location_raw and str(location_raw).strip()) or None
            try:
                qty = int(float(qty_raw)) if qty_raw is not None else 0
            except (TypeError, ValueError):
                qty = 0

            container = _spo_import_extract_container(loading_date_raw)
            spo_identity = {
                "spo_number": spo_number,
                "item_code": item_code,
                "location": location,
                "qty": qty,
                "container": container,
            }

            def _spo_skip(code: str, reason: str, value=None) -> None:
                outcome.skip(
                    row=row_idx, code=code, message=reason, value=value, identity=spo_identity
                )
                skipped_rows_detail.append({"row": row_idx, "reason": reason})

            if not item_code:
                _spo_skip(oc.MISSING_ITEM_CODE, "Missing Item Code / Product Code")
                continue
            if not location:
                _spo_skip(oc.MISSING_LOCATION, "Missing Location / Warehouse")
                continue
            if qty <= 0:
                _spo_skip(oc.INVALID_QUANTITY, "Invalid or zero Qty")
                continue
            if not container:
                _spo_skip(
                    oc.MISSING_CONTAINER,
                    "Missing or invalid Loading Date (no container number)",
                )
                continue

            product = products_by_code.get(item_code)
            warehouse = warehouses_map.get(normalize_code(location)) if location else None
            shipment = get_inbound_shipment_by_container_number(db, container)

            if not product:
                _spo_skip(oc.PRODUCT_NOT_FOUND, f"Product not found: {item_code}", item_code)
                continue
            if not warehouse:
                _spo_skip(oc.WAREHOUSE_NOT_FOUND, f"Warehouse not found: {location}", location)
                continue
            if not shipment:
                _spo_skip(
                    oc.PACKING_LIST_NOT_FOUND,
                    f"Packing list not found for container: {container}",
                    container,
                )
                continue
            resolved_rows.append((str(product.id), str(warehouse.id), str(shipment.id), qty, row_idx))
            spo_row_identity[(str(product.id), str(warehouse.id))] = spo_identity
            spo_group_rows[(str(product.id), str(warehouse.id))].append(row_idx)

        # Group by (product_id, warehouse_id): sum qty, keep first shipment_id
        groups: Dict[tuple[str, str], tuple[int, str]] = defaultdict(lambda: (0, ""))
        for product_id, warehouse_id, shipment_id, qty, row_idx in resolved_rows:
            key = (product_id, warehouse_id)
            total, first_shipment = groups[key]
            if not first_shipment:
                first_shipment = shipment_id
            groups[key] = (total + qty, first_shipment)

        num_groups = len(groups)
        row_level_skipped = len(skipped_rows_detail)
        job_service.update_job_progress(job_id_str, total_rows=total_data_rows)

        successful = 0
        updated = 0
        unchanged = 0
        failed = 0
        guarded_skipped = 0
        processed = 0
        errors: List[str] = []
        proc_service = SPOAllocationService(db)
        # (spo_number, company_id) pairs to forward match once the WHOLE file has
        # landed. The file is upserted one (product, warehouse) group at a time, so
        # a hook fired per allocation row runs while the rest of the file does not
        # exist yet and places a waiting GRN line against whichever allocation
        # happened to be written first rather than the one covering its warehouse -
        # which is upload-order dependence, the exact thing this feature removes.
        forward_match_targets: set[tuple[str, Optional[str]]] = set()

        for (product_id, warehouse_id), (total_qty, shipment_id) in groups.items():
            processed += 1
            try:
                allocation_data = SPOAllocationCreate(
                    spo_number=spo_number,
                    inbound_shipment_id=shipment_id,
                    warehouse_id=warehouse_id,
                    product_id=product_id,
                    allocated_quantity=total_qty,
                    receipt_status="pending",
                    quantity_received=0,
                    quantity_rejected=0,
                )
                action, _allocation = proc_service.upsert_allocation(
                    allocation_data, user_id, forward_match=False
                )
                if action in ("created", "updated") and _allocation.spo_number:
                    forward_match_targets.add(
                        (_allocation.spo_number, _allocation.company_id)
                    )
                group_key = (product_id, warehouse_id)
                identity = spo_row_identity.get(group_key) or {}
                if action == "created":
                    successful += 1
                elif action == "updated":
                    updated += 1
                else:  # "unchanged"
                    unchanged += 1
                code = {"created": oc.CREATED, "updated": oc.UPDATED}.get(action, oc.UNCHANGED)
                row_outcome = {
                    "created": oc.OUTCOME_CREATED,
                    "updated": oc.OUTCOME_UPDATED,
                }.get(action, oc.OUTCOME_UNCHANGED)
                for source_row in spo_group_rows.get(group_key, []):
                    outcome.success(
                        row=source_row,
                        outcome=row_outcome,
                        code=code,
                        message=f"Allocation {action}: {spo_number}",
                        value=spo_number,
                        identity=identity,
                        entity_type="spo_allocation",
                    )
            except AllocationReceivedGuardError as e:
                guarded_skipped += 1
                errors.append(str(e))
                group_key = (product_id, warehouse_id)
                for source_row in spo_group_rows.get(group_key, []):
                    outcome.skip(
                        row=source_row,
                        code=oc.ALREADY_RECEIVED_GUARD,
                        message=str(e),
                        value=spo_number,
                        identity=spo_row_identity.get(group_key),
                    )
            except Exception as e:
                failed += 1
                errors.append(f"Upsert allocation: {e}")
                group_key = (product_id, warehouse_id)
                for source_row in spo_group_rows.get(group_key, []):
                    outcome.fail(
                        row=source_row,
                        code=oc.UPSERT_ERROR,
                        message=f"Upsert allocation: {e}",
                        value=spo_number,
                        identity=spo_row_identity.get(group_key),
                    )

            progress = outcome.completion_counts()
            progress["processed_rows"] = processed
            job_service.update_job_progress(
                job_id_str,
                result={"errors": errors[-50:], "skipped_rows_detail": skipped_rows_detail[-200:]},
                **progress,
            )

        # The whole file has landed, so the pool is now complete: any GRN line that
        # stated one of these SPOs and could not be placed when it was imported is
        # placeable, against the allocation that actually covers its warehouse.
        # Post-commit and best-effort - a side effect must not fail an import whose
        # allocations are already written.
        for target_spo, target_company in forward_match_targets:
            forward_match_grn_lines_for_spo_best_effort(
                db, target_spo, company_id=target_company
            )

        total_skipped = row_level_skipped + guarded_skipped
        _skip_errors, skip_warnings = partition_spo_skip_reasons(skipped_rows_detail)
        job_service.complete_job(
            job_id=job_id_str,
            result=outcome.finalize(
                "SPO import completed",
                total_rows=total_data_rows,
                # legacy keys, kept one release
                data_rows=total_data_rows,
                allocations_created=successful,
                allocations_updated=updated,
                allocations_unchanged=unchanged,
                skipped_rows_detail=_json_safe(skipped_rows_detail[-200:]),
                skipped_rows_count=total_skipped,
                warnings=_json_safe(skip_warnings[-100:]),
                errors=_json_safe(errors[-100:]),
            ),
            **outcome.completion_counts(total_rows=total_data_rows),
        )
        logger.info(
            "SPO import job %s completed: %s data rows, %s ok, %s failed, %s skipped",
            job_id_str, total_data_rows, successful, failed, total_skipped,
        )
        _write_import_audit(
            db,
            entity_type="spo",
            label=f"SPO import {filename or ''}".strip(),
            row_count=successful,
            user_id=user_id,
            entity_id=job_id_str,
            status="partial" if failed else "success",
        )
    except Exception as e:
        logger.exception("SPO import job %s failed", job_id_str)
        # A crashed import still owes the operator the rows it did classify. The
        # recorder writes on its own session, so flushing here survives the rollback.
        outcome.flush()
        job_service.fail_job(job_id_str, str(e))
        _write_import_audit(
            db,
            entity_type="spo",
            label=f"SPO import {filename or ''}".strip(),
            row_count=0,
            user_id=user_id,
            entity_id=job_id_str,
            status="failed",
        )
    finally:
        db.close()


def validate_spo_import(
    file_data: bytes, filename: str, *, company_scope: Any = _SCOPE_NOT_GIVEN
) -> Dict[str, Any]:
    """Run SPO import validation (same parsing and row validation as process_spo_import). No allocations created.

    ``company_scope`` is the caller's resolved scope, so the preview reads exactly
    what the import will read (see ``_apply_preview_scope``).
    """
    import openpyxl

    spo_number = re.sub(r"\.xlsx?$", "", filename or "", flags=re.IGNORECASE).strip()
    if not spo_number:
        return {"valid": False, "errors": ["Filename must provide SPO number (e.g. SPO-2025.10-0050.xlsx)"], "warnings": [], "summary": {}}

    db = SessionLocal()
    _apply_preview_scope(db, company_scope)
    try:
        workbook = openpyxl.load_workbook(BytesIO(file_data), data_only=True)
    except Exception as exc:
        db.close()
        return {"valid": False, "errors": [f"Failed to read Excel: {exc}"], "warnings": [], "summary": {}}
    sheet = workbook.active
    if not sheet:
        db.close()
        return {"valid": False, "errors": ["Workbook has no active sheet"], "warnings": [], "summary": {}}

    headers = [_spo_import_normalize_header(cell.value) for cell in sheet[1]]
    data_rows: List[tuple[int, dict]] = []
    all_product_codes = set()
    all_locations = set()
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        row_data = {}
        for idx, value in enumerate(row):
            if idx < len(headers) and headers[idx]:
                row_data[headers[idx]] = value
        item_code_raw = _spo_import_find_column(row_data, "Item Code", "Item code", "Product Code", "Product code")
        location_raw = _spo_import_find_column(row_data, "Location", "Warehouse", "Warehouse Code")
        qty_raw = _spo_import_find_column(row_data, "Qty", "Quantity", "Allocated", "Allocated Quantity")
        loading_date_raw = _spo_import_find_column(row_data, "Loading Date", "Loading date")
        item_code = (item_code_raw and str(item_code_raw).strip()) or None
        location = (location_raw and str(location_raw).strip()) or None
        try:
            qty = int(float(qty_raw)) if qty_raw is not None else 0
        except (TypeError, ValueError):
            qty = 0
        container = _spo_import_extract_container(loading_date_raw)
        if item_code:
            all_product_codes.add(item_code)
        if location:
            all_locations.add(location)
        data_rows.append((row_idx, row_data))

    if not data_rows:
        db.close()
        return {"valid": True, "errors": [], "warnings": [], "summary": {"total_data_rows": 0}}

    from app.api.v1.external.utils import normalize_code
    products_by_code = get_products_by_code_exact(db, all_product_codes)
    warehouses_map = get_warehouses_by_code_or_name(db, all_locations)
    skipped_rows_detail: List[dict] = []
    would_succeed = 0
    for row_idx, row_data in data_rows:
        item_code_raw = _spo_import_find_column(row_data, "Item Code", "Item code", "Product Code", "Product code")
        loading_date_raw = _spo_import_find_column(row_data, "Loading Date", "Loading date")
        location_raw = _spo_import_find_column(row_data, "Location", "Warehouse", "Warehouse Code")
        qty_raw = _spo_import_find_column(row_data, "Qty", "Quantity", "Allocated", "Allocated Quantity")
        item_code = (item_code_raw and str(item_code_raw).strip()) or None
        location = (location_raw and str(location_raw).strip()) or None
        try:
            qty = int(float(qty_raw)) if qty_raw is not None else 0
        except (TypeError, ValueError):
            qty = 0
        container = _spo_import_extract_container(loading_date_raw)
        if not item_code:
            skipped_rows_detail.append({"row": row_idx, "reason": "Missing Item Code / Product Code"})
            continue
        if not location:
            skipped_rows_detail.append({"row": row_idx, "reason": "Missing Location / Warehouse"})
            continue
        if qty <= 0:
            skipped_rows_detail.append({"row": row_idx, "reason": "Invalid or zero Qty"})
            continue
        if not container:
            skipped_rows_detail.append({"row": row_idx, "reason": "Missing or invalid Loading Date (no container number)"})
            continue
        product = products_by_code.get(item_code)
        warehouse = warehouses_map.get(normalize_code(location)) if location else None
        shipment = get_inbound_shipment_by_container_number(db, container)
        if not product:
            skipped_rows_detail.append({"row": row_idx, "reason": f"Product not found: {item_code}"})
            continue
        if not warehouse:
            skipped_rows_detail.append({"row": row_idx, "reason": f"Warehouse not found: {location}"})
            continue
        if not shipment:
            skipped_rows_detail.append({"row": row_idx, "reason": f"Packing list not found for container: {container}"})
            continue
        would_succeed += 1

    errors, warnings = partition_spo_skip_reasons(skipped_rows_detail)
    db.close()
    return {
        "valid": len(errors) == 0,
        "errors": errors,
        "warnings": warnings,
        "summary": {
            "spo_number": spo_number,
            "total_data_rows": len(data_rows),
            "would_succeed": would_succeed,
            "would_skip": len(skipped_rows_detail),
            "skipped_rows_detail": skipped_rows_detail[-100:],
        },
    }


def _grn_import_normalize_header(value: Any) -> str:
    """Normalize Excel header for GRN import (lowercase, trim, punctuation-insensitive)."""
    if value is None:
        return ""
    s = str(value).strip().lower()
    if not s:
        return ""
    # Treat variants like "Doc. No.", "Doc No", "DOC-NO" as the same key.
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


# GRN listing "Transfer from" and line sheet "From Doc. No." both carry the SPO number.
_GRN_SPO_COLUMN_CANDIDATES = (
    "transfer from",
    "transfer from ",
    "spo number",
    "from doc no",
    "from doc number",
    "from document no",
    "from document number",
    "our po no",
    "our po number",
    "our p o no",
    "our p o number",
)

# The GRN DETAIL (line) export carries the SPO on the LINE itself - AutoCount
# calls it "Our PO No.". That is the stronger linkage: one GRN can be received
# against several SPOs, and only the line knows which one it belongs to. The
# header's "Transfer from" is the FALLBACK, used when the line says nothing.
#
# Order matters, and so does emptiness: `_first_filled` walks these in order and
# takes the first candidate that is actually POPULATED on the row, so a sheet
# carrying both an empty "Our PO No." and a repeated "Transfer From" still
# resolves to the header value instead of stopping at the blank line column.
_GRN_LINE_SPO_COLUMN_CANDIDATES = (
    "our po no",
    "our po number",
    # "Our P.O. No." normalizes to separate letters, so it needs its own entry.
    "our p o no",
    "our p o number",
    "our po",
    "from doc no",
    "from doc number",
    "from document no",
    "from document number",
    "spo number",
    "transfer from",
    "transfer from ",
)


def _resolve_line_spo(row_data: dict) -> Optional[str]:
    """The SPO a GRN LINE was received against, from the line's own columns.

    None when the line names no single SPO (blank, or a multi-SPO cell) - the
    caller then falls back to the GRN header's `spo_number`, which is the older
    "Transfer from" linkage.
    """
    return _single_spo_or_none(_first_filled(row_data, *_GRN_LINE_SPO_COLUMN_CANDIDATES))


def _first_filled(row: dict, *candidates: str) -> Any:
    """First candidate column that is PRESENT AND NON-BLANK on this row.

    Distinct from a present-key lookup: an export that ships the line-level SPO
    column but leaves it empty on a given row must fall through to the next
    candidate, not resolve to None and lose the header fallback.
    """
    for c in candidates:
        cl = _grn_import_normalize_header(c)
        if cl not in row:
            continue
        value = row.get(cl)
        if value is None:
            continue
        if str(value).strip():
            return value
    return None

# AutoCount puts every SPO a GRN was received against into the ONE "Transfer
# from" cell ("SPO-2026/06-0020, SPO-2026/06-0021, ..."), which overflowed the
# old varchar(50) and aborted the import. Migration 317 widened
# `picking_headers.spo_number` to 255 so the header can SAY which SPOs it covers.
_GRN_SPO_SEPARATORS = re.compile(r"[,;\n\r]+|\s{2,}")
# Storage width of `picking_headers.spo_number` after migration 317.
_SPO_NUMBER_MAX_LEN = 255


def _split_spo_cell(raw: Any) -> list[str]:
    """Split a GRN SPO cell into its individual SPO numbers.

    One value in, one value out for the normal case; a multi-SPO GRN yields the
    full list (see `_header_spo_display` and `_single_spo_or_none`).
    """
    if raw is None:
        return []
    return [part.strip() for part in _GRN_SPO_SEPARATORS.split(str(raw)) if part.strip()]


def _header_spo_display(raw: Any) -> Optional[str]:
    """What `picking_headers.spo_number` stores: every SPO the GRN covers.

    Normalized to a single ", "-joined form so the same cell always stores the
    same string regardless of which separator AutoCount exported.

    This column is DISPLAY for the multi-SPO case. Matching is scalar and stays
    scalar - `_spo_match_key`, `procurement_service._normalize_spo_number` and the
    packing-list grouping all compare ONE normalized SPO - so a joined value
    equals no single SPO and can never false-link. The per-line SPO
    (`_single_spo_or_none`) is what drives allocation matching.
    """
    parts = _split_spo_cell(raw)
    if not parts:
        return None
    return ", ".join(parts)


def _single_spo_or_none(raw: Any) -> Optional[str]:
    """The one SPO number in a GRN SPO cell, or None when there is not exactly one.

    Used where the value FEEDS MATCHING (the per-line SPO): a GRN line received
    against several SPOs names no single allocation, so it stays unlinked rather
    than carrying a joined string that `_spo_match_key` can never match.
    """
    parts = _split_spo_cell(raw)
    if len(parts) != 1:
        return None
    return parts[0]


def _run_grn_listing_import_core(
    db,
    file_data: bytes,
    job_service: Optional[JobService] = None,
    job_id_str: Optional[str] = None,
    outcome: Optional[ImportOutcome] = None,
    created_by: Optional[str] = None,
    import_job_db_id: Optional[str] = None,
) -> Dict[str, Any]:
    """Parse GRN listing Excel and run upsert loop. Returns counts and errors. Caller must commit or rollback.

    ``outcome`` records per-row attribution. The validation preview passes a
    non-persisting recorder so preview and import speak the same reason codes.

    ``created_by`` / ``import_job_db_id`` are stamped onto headers this run
    CREATES, so a GRN can say who imported it and from which file. The preview
    passes neither (it writes nothing).
    """
    import openpyxl

    if outcome is None:
        outcome = ImportOutcome(None, persist=False)

    try:
        workbook = openpyxl.load_workbook(BytesIO(file_data), data_only=True)
    except Exception as exc:
        return {"error": f"Failed to read Excel: {exc}", "successful": 0, "failed": 0, "skipped": 0, "errors": [], "skipped_rows_detail": [], "total_data_rows": 0}

    sheet = workbook.active
    if not sheet:
        return {"error": "Workbook has no active sheet", "successful": 0, "failed": 0, "skipped": 0, "errors": [], "skipped_rows_detail": [], "total_data_rows": 0}

    headers = [_grn_import_normalize_header(cell.value) for cell in sheet[1]]
    data_rows: List[dict] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        row_data = {}
        for idx, value in enumerate(row):
            if idx < len(headers) and headers[idx]:
                row_data[headers[idx]] = value
        data_rows.append(row_data)

    def _find(row: dict, *candidates: str) -> Any:
        for c in candidates:
            cl = _grn_import_normalize_header(c)
            if cl in row:
                return row.get(cl)
        return None

    total_data_rows = len(data_rows)
    if job_service and job_id_str:
        job_service.update_job_progress(job_id_str, total_rows=total_data_rows)

    proc = PickingHeaderService(db)
    errors: List[str] = []
    skipped_rows_detail: List[dict] = []  # legacy key, kept one release

    def _progress(row_idx: int) -> None:
        if job_service and job_id_str:
            counts = outcome.completion_counts()
            counts["processed_rows"] = row_idx - 1
            job_service.update_job_progress(job_id_str, **counts)

    for row_idx, row in enumerate(data_rows, start=2):
        doc_num = _find(row, "doc number", "doc. no.", "doc. number", "doc number ", "grn number")
        grn_number = (doc_num and str(doc_num).strip()) or None
        transfer_from = _first_filled(row, *_GRN_SPO_COLUMN_CANDIDATES)
        # Every SPO the GRN covers, normalized. Display for the multi-SPO case;
        # allocation matching runs off the per-line SPO.
        spo_number = _header_spo_display(transfer_from)
        identity = {"grn_number": grn_number, "spo_number": spo_number}
        if not grn_number:
            reason = "Missing doc number / GRN number"
            outcome.skip(row=row_idx, code=oc.MISSING_DOC_NO, message=reason, identity=identity)
            skipped_rows_detail.append({"row": row_idx, "reason": reason})
            _progress(row_idx)
            continue
        if spo_number and len(spo_number) > _SPO_NUMBER_MAX_LEN:
            # Past 255 even the widened column cannot hold it: bad source data.
            # Skip the row with the reason instead of letting Postgres abort the
            # transaction (which used to take every later row down with it).
            reason = (
                f"SPO number longer than {_SPO_NUMBER_MAX_LEN} characters: {spo_number[:60]}"
            )
            outcome.skip(row=row_idx, code=oc.UPSERT_ERROR, message=reason, identity=identity)
            skipped_rows_detail.append({"row": row_idx, "reason": reason})
            _progress(row_idx)
            continue
        date_val = _find(row, "date", "picking date", "picking date ")
        try:
            _pd = parse_date_value(date_val) if date_val is not None else date.today()
            picking_date = _pd if _pd is not None else date.today()
        except Exception:
            picking_date = date.today()
        try:
            header, was_created = proc.upsert_grn_header_for_import(
                grn_number,
                spo_number,
                picking_date,
                created_by=created_by,
                import_job_id=import_job_db_id,
            )
            # Report which of the two actually happened. Reporting every success as
            # `created` made the last person to re-run a file look like the author
            # of every GRN in it, which is what made "who created this GRN"
            # unanswerable from import_job_rows.
            outcome.success(
                row=row_idx,
                code=oc.CREATED if was_created else oc.UPDATED,
                # `outcome` is the column a query filters on and it defaults to
                # "created", so it has to be set alongside `code` or the two
                # disagree and the filter still lies.
                outcome=oc.OUTCOME_CREATED if was_created else oc.OUTCOME_UPDATED,
                message=(
                    f"GRN header created: {grn_number}"
                    if was_created
                    else f"GRN header updated: {grn_number}"
                ),
                value=grn_number,
                identity=identity,
                entity_type="picking_header",
                entity_id=str(header.id),
            )
        except Exception as e:
            # `upsert_grn_header_for_import` commits per row, so a failed flush
            # leaves the session needing a rollback. Without this, every LATER row
            # dies with "transaction has been rolled back due to a previous
            # exception" — one bad cell fails the whole file and the job report
            # blames rows that were fine.
            try:
                db.rollback()
            except Exception:
                logger.exception("rollback after GRN header upsert failure also failed")
            outcome.fail(
                row=row_idx,
                code=oc.UPSERT_ERROR,
                message=f"{grn_number}: {e}",
                value=grn_number,
                identity=identity,
            )
            errors.append(f"{grn_number}: {e}")
        _progress(row_idx)

    return {
        "successful": outcome.successful,
        "failed": outcome.failed,
        "skipped": outcome.skipped,
        "errors": errors,
        "skipped_rows_detail": skipped_rows_detail,
        "total_data_rows": total_data_rows,
        "outcome": outcome,
    }


def validate_grn_listing_import(
    file_data: bytes, *, company_scope: Any = _SCOPE_NOT_GIVEN
) -> Dict[str, Any]:
    """Run GRN listing validation (same logic as import, then rollback). No DB writes.

    Scoped to the caller's company so the preview cannot claim a row would succeed
    against a header the scoped import will never see (``_apply_preview_scope``).
    """
    db = SessionLocal()
    _apply_preview_scope(db, company_scope)
    try:
        result = _run_grn_listing_import_core(db, file_data)
        db.rollback()
        if "error" in result:
            return {"valid": False, "errors": [result["error"]], "warnings": [], "summary": result}
        return {
            "valid": result["failed"] == 0 and len(result["errors"]) == 0,
            "errors": result["errors"],
            "warnings": [],
            "summary": {
                "total_data_rows": result["total_data_rows"],
                "would_succeed": result["successful"],
                "would_fail": result["failed"],
                "would_skip": result["skipped"],
                "skipped_rows_detail": result["skipped_rows_detail"][-100:],
            },
        }
    finally:
        db.close()


def validate_grn_lines_import(
    file_data: bytes, *, company_scope: Any = _SCOPE_NOT_GIVEN
) -> Dict[str, Any]:
    """Run GRN lines validation: parse file and run grouping-phase checks (same as import). No line creation.

    Scoped to the caller's company: product / warehouse / GRN-header lookups here
    must resolve exactly what the scoped import resolves (``_apply_preview_scope``).
    """
    import openpyxl
    from app.api.v1.external.utils import normalize_code

    db = SessionLocal()
    _apply_preview_scope(db, company_scope)
    try:
        workbook = openpyxl.load_workbook(BytesIO(file_data), data_only=True)
    except Exception as exc:
        db.close()
        return {"valid": False, "errors": [f"Failed to read Excel: {exc}"], "warnings": [], "summary": {}}
    sheet = workbook.active
    if not sheet:
        db.close()
        return {"valid": False, "errors": ["Workbook has no active sheet"], "warnings": [], "summary": {}}

    headers = [_grn_import_normalize_header(cell.value) for cell in sheet[1]]
    data_rows: List[tuple] = []
    all_doc_nos = set()
    all_product_codes = set()
    all_locations = set()

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        row_data = {}
        for idx, value in enumerate(row):
            if idx < len(headers) and headers[idx]:
                row_data[headers[idx]] = value

        def _find(row_d: dict, *candidates: str) -> Any:
            for c in candidates:
                cl = _grn_import_normalize_header(c)
                if cl in row_d:
                    return row_d.get(cl)
            return None

        doc_no = (_find(row_data, "doc no", "doc number", "grn number") and str(_find(row_data, "doc no", "doc number", "grn number")).strip()) or None
        item_code = (_find(row_data, "item code", "item code ", "product code", "product code ") and str(_find(row_data, "item code", "item code ", "product code", "product code ")).strip()) or None
        location = (_find(row_data, "location", "warehouse", "warehouse code") and str(_find(row_data, "location", "warehouse", "warehouse code")).strip()) or None
        qty_raw = _find(row_data, "qty", "quantity", "qty ")
        # Line-level SPO first ("Our PO No."), header column only as fallback -
        # the preview must resolve exactly what the import resolves.
        line_spo = _resolve_line_spo(row_data)
        try:
            qty = int(float(qty_raw)) if qty_raw is not None else 0
        except (TypeError, ValueError):
            qty = 0
        if doc_no:
            all_doc_nos.add(doc_no)
        if item_code:
            all_product_codes.add(item_code)
        if location:
            all_locations.add(location)
        data_rows.append((doc_no, item_code, location, qty, line_spo))

    if not data_rows:
        db.close()
        return {"valid": True, "errors": [], "warnings": [], "summary": {"total_data_rows": 0}}

    proc = PickingHeaderService(db)
    products_by_code = get_products_by_code_exact(db, all_product_codes)
    warehouses_map = get_warehouses_by_code_or_name(db, all_locations)
    headers_by_number: Dict[str, Any] = {}
    for pn in all_doc_nos:
        h = proc.get_grn_by_picking_number(pn)
        if h:
            headers_by_number[pn] = h

    errors = []
    skipped_detail: List[dict] = []
    would_succeed = 0
    for row_idx, row_tuple in enumerate(data_rows, start=2):
        doc_no, item_code, location, qty, _line_spo = row_tuple
        if not doc_no:
            skipped_detail.append({"row": row_idx, "reason": "Missing doc no"})
            continue
        if not item_code:
            skipped_detail.append({"row": row_idx, "reason": "Missing item code"})
            continue
        if not location:
            skipped_detail.append({"row": row_idx, "reason": "Missing location"})
            continue
        if qty <= 0:
            skipped_detail.append({"row": row_idx, "reason": "Invalid quantity"})
            continue
        if doc_no not in headers_by_number:
            skipped_detail.append({"row": row_idx, "reason": f"GRN header not found: {doc_no}"})
            continue
        product = products_by_code.get((item_code or "").strip())
        warehouse = warehouses_map.get(normalize_code(location)) if location else None
        if not product:
            skipped_detail.append({"row": row_idx, "reason": f"Product not found: {item_code}"})
            continue
        if not warehouse:
            skipped_detail.append({"row": row_idx, "reason": f"Warehouse not found: {location}"})
            continue
        would_succeed += 1

    errors = [f"Row {s['row']}: {s['reason']}" for s in skipped_detail]
    db.close()
    return {
        "valid": len(errors) == 0,
        "errors": errors,
        "warnings": [],
        "summary": {
            "total_data_rows": len(data_rows),
            "would_succeed": would_succeed,
            "would_skip": len(skipped_detail),
            "skipped_rows_detail": skipped_detail[-100:],
        },
    }


def process_grn_listing_import(db_job_id: str, file_data: bytes, filename: str, user_id: str):
    """Process GRN listing Excel: create/update picking headers. Idempotent."""
    from rq import get_current_job

    db = SessionLocal()
    _apply_import_job_scope(db, db_job_id)
    job_service = JobService(db)
    rq_job = get_current_job()
    rq_job_id = rq_job.id if rq_job else None
    job = job_service.get_job_by_db_id(db_job_id) if db_job_id else None
    if not job and rq_job_id:
        job = job_service.get_job(rq_job_id)
    if not job:
        logger.error("GRN listing import job not found: db_job_id=%s", db_job_id)
        db.close()
        return

    job_id_str: str = str(job.job_id)
    outcome = ImportOutcome(getattr(job, "id", None), session_factory=SessionLocal)
    try:
        job_service.start_job(job_id_str)
        result = _run_grn_listing_import_core(
            db,
            file_data,
            job_service=job_service,
            job_id_str=job_id_str,
            outcome=outcome,
            # Provenance for the headers this run creates: the uploader, and the
            # job (which carries the file name and the company snapshot).
            created_by=user_id,
            import_job_db_id=str(getattr(job, "id", "") or "") or None,
        )
        if "error" in result:
            job_service.fail_job(job_id_str, result["error"])
            db.close()
            return
        db.commit()
        job_service.complete_job(
            job_id=job_id_str,
            result=outcome.finalize(
                "GRN listing import completed",
                total_rows=result["total_data_rows"],
                # legacy keys, kept one release
                errors=result["errors"][-100:],
                skipped_rows_detail=_json_safe(result["skipped_rows_detail"][-500:]),
            ),
            **outcome.completion_counts(total_rows=result["total_data_rows"]),
        )
        logger.info(
            "GRN listing import job %s completed: %s ok, %s failed, %s skipped",
            job_id_str, result["successful"], result["failed"], result["skipped"],
        )
        _write_import_audit(
            db,
            entity_type="grn",
            label=f"GRN listing import {filename or ''}".strip(),
            row_count=result["successful"],
            user_id=user_id,
            entity_id=job_id_str,
            status="partial" if result["failed"] else "success",
        )
    except Exception as e:
        logger.exception("GRN listing import job %s failed", job_id_str)
        # A crashed import still owes the operator the rows it did classify. The
        # recorder writes on its own session, so flushing here survives the rollback.
        outcome.flush()
        job_service.fail_job(job_id_str, str(e))
        _write_import_audit(
            db,
            entity_type="grn",
            label=f"GRN listing import {filename or ''}".strip(),
            row_count=0,
            user_id=user_id,
            entity_id=job_id_str,
            status="failed",
        )
    finally:
        db.close()


def process_grn_lines_import(db_job_id: str, file_data: bytes, filename: str, user_id: str):
    """Process GRN lines Excel: create/update picking lines. Idempotent.
    Columns: doc no -> GRN picking_number; item code; location -> warehouse; qty.
    SPO source, in order: the LINE's own column ("Our PO No." / "From Doc. No."),
    then the sheet's repeated header column ("Transfer From"), then the GRN
    header's stored spo_number. Groups split by effective SPO so mixed SPOs on
    one GRN do not merge.
    """
    from rq import get_current_job
    import openpyxl

    db = SessionLocal()
    _apply_import_job_scope(db, db_job_id)
    job_service = JobService(db)
    rq_job = get_current_job()
    rq_job_id = rq_job.id if rq_job else None
    job = job_service.get_job_by_db_id(db_job_id) if db_job_id else None
    if not job and rq_job_id:
        job = job_service.get_job(rq_job_id)
    if not job:
        logger.error("GRN lines import job not found: db_job_id=%s", db_job_id)
        db.close()
        return

    job_id_str: str = str(job.job_id)
    outcome = ImportOutcome(getattr(job, "id", None), session_factory=SessionLocal)
    try:
        job_service.start_job(job_id_str)
        try:
            workbook = openpyxl.load_workbook(BytesIO(file_data), data_only=True)
        except Exception as exc:
            job_service.fail_job(job_id_str, f"Failed to read Excel: {exc}")
            db.close()
            return

        sheet = workbook.active
        if not sheet:
            job_service.fail_job(job_id_str, "Workbook has no active sheet")
            db.close()
            return

        headers = [_grn_import_normalize_header(cell.value) for cell in sheet[1]]
        data_rows: List[tuple] = []  # (doc_no, item_code, location, qty, line_spo)
        all_doc_nos = set()
        all_product_codes = set()
        all_locations = set()

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            row_data = {}
            for idx, value in enumerate(row):
                if idx < len(headers) and headers[idx]:
                    row_data[headers[idx]] = value

            def _find(row_d: dict, *candidates: str) -> Any:
                for c in candidates:
                    cl = _grn_import_normalize_header(c)
                    if cl in row_d:
                        return row_d.get(cl)
                return None

            doc_no = (_find(row_data, "doc no", "doc number", "grn number") and str(_find(row_data, "doc no", "doc number", "grn number")).strip()) or None
            item_code = (_find(row_data, "item code", "item code ", "product code", "product code ") and str(_find(row_data, "item code", "item code ", "product code", "product code ")).strip()) or None
            location = (_find(row_data, "location", "warehouse", "warehouse code") and str(_find(row_data, "location", "warehouse", "warehouse code")).strip()) or None
            qty_raw = _find(row_data, "qty", "quantity", "qty ")
            # "Our PO No." is the LINE's own SPO and wins; the repeated header
            # column ("Transfer From") is only consulted when the line is blank,
            # and the DB header's spo_number is the last fallback below.
            # Same scalar rule as the header: a multi-SPO cell names no single
            # allocation, so the line stays unlinked rather than carrying a blob
            # that `_spo_match_key` can never match.
            line_spo = _resolve_line_spo(row_data)
            try:
                qty = int(float(qty_raw)) if qty_raw is not None else 0
            except (TypeError, ValueError):
                qty = 0
            if doc_no:
                all_doc_nos.add(doc_no)
            if item_code:
                all_product_codes.add(item_code)
            if location:
                all_locations.add(location)
            data_rows.append((doc_no, item_code, location, qty, line_spo))

        total_data_rows = len(data_rows)
        # Set total_rows immediately after reading all rows, before any processing
        job_service.update_job_progress(job_id_str, total_rows=total_data_rows, processed_rows=0)

        if not data_rows:
            job_service.complete_job(
                job_id=job_id_str,
                result=outcome.finalize("No valid data rows", total_rows=total_data_rows),
                **outcome.completion_counts(total_rows=total_data_rows),
            )
            db.close()
            return

        proc = PickingHeaderService(db)
        products_by_code = get_products_by_code_exact(db, all_product_codes)
        warehouses_map = get_warehouses_by_code_or_name(db, all_locations)
        from app.api.v1.external.utils import normalize_code

        product_id_to_code = {str(p.id): code for code, p in products_by_code.items()}
        warehouse_id_to_display = {
            str(w.id): (getattr(w, "code", None) or getattr(w, "name", None) or str(w.id))
            for w in warehouses_map.values()
        }

        # Resolve headers by picking_number
        headers_by_number: Dict[str, Any] = {}
        for pn in all_doc_nos:
            h = proc.get_grn_by_picking_number(pn)
            if h:
                headers_by_number[pn] = h

        # Group by (doc_no, product_id, warehouse_id) and sum quantity; track skipped with Excel row number
        groups: Dict[tuple[str, str, str, Optional[str]], Dict[str, Any]] = {}
        skipped_detail: List[dict] = []  # legacy key, kept one release
        # Rows that made it into a group, so a group-level failure can be attributed
        # back to the source rows that fed it.
        group_source_rows: Dict[tuple[str, str, str, Optional[str]], List[int]] = defaultdict(list)
        group_row_identity: Dict[tuple[str, str, str, Optional[str]], Dict[str, Any]] = {}

        def _line_identity(doc_no, item_code, location, qty, spo=None) -> Dict[str, Any]:
            return {
                "doc_no": doc_no,
                "item_code": item_code,
                "location": location,
                "qty": qty,
                "spo_number": spo,
            }

        def _grn_skip(row_idx: int, code: str, reason: str, identity: Dict[str, Any], value=None) -> None:
            outcome.skip(row=row_idx, code=code, message=reason, value=value, identity=identity)
            skipped_detail.append({"row": row_idx, "reason": reason})
            counts = outcome.completion_counts()
            counts["processed_rows"] = row_idx - 1
            job_service.update_job_progress(job_id_str, total_rows=total_data_rows, **counts)

        for row_idx, row_tuple in enumerate(data_rows, start=2):  # Excel row 2 = first data row
            doc_no, item_code, location, qty, line_spo = row_tuple
            identity = _line_identity(doc_no, item_code, location, qty, line_spo)
            if not doc_no:
                _grn_skip(row_idx, oc.MISSING_DOC_NO, "Missing doc no", identity)
                continue
            if not item_code:
                _grn_skip(row_idx, oc.MISSING_ITEM_CODE, "Missing item code", identity)
                continue
            if not location:
                _grn_skip(row_idx, oc.MISSING_LOCATION, "Missing location", identity)
                continue
            if qty <= 0:
                _grn_skip(row_idx, oc.INVALID_QUANTITY, "Invalid quantity", identity)
                continue
            header = headers_by_number.get(doc_no)
            if not header:
                _grn_skip(
                    row_idx,
                    oc.GRN_HEADER_NOT_FOUND,
                    f"GRN header not found: {doc_no}",
                    identity,
                    doc_no,
                )
                continue
            # The header column may hold EVERY SPO the GRN covers (display form).
            # Only a single-SPO header can name the allocation for a line, so a
            # joined header falls back to no SPO rather than to a string that
            # `_spo_match_key` can never match.
            hdr_spo = _single_spo_or_none(getattr(header, "spo_number", None))
            effective_spo: Optional[str] = line_spo if line_spo else hdr_spo
            product = products_by_code.get((item_code or "").strip())
            warehouse = warehouses_map.get(normalize_code(location)) if location else None
            if not product:
                _grn_skip(
                    row_idx,
                    oc.PRODUCT_NOT_FOUND,
                    f"Product not found: {item_code}",
                    identity,
                    item_code,
                )
                continue
            if not warehouse:
                _grn_skip(
                    row_idx,
                    oc.WAREHOUSE_NOT_FOUND,
                    f"Warehouse not found: {location}",
                    identity,
                    location,
                )
                continue
            # Group by (doc_no, product_id, warehouse_id, effective_spo) so line-level SPO (e.g. From Doc. No.) is not merged across SPOs.
            key = (doc_no, str(product.id), str(warehouse.id), effective_spo)
            if key not in groups:
                groups[key] = {"qty": 0, "warehouse_id": str(warehouse.id)}
            groups[key]["qty"] += qty
            group_source_rows[key].append(row_idx)
            group_row_identity[key] = _line_identity(doc_no, item_code, location, groups[key]["qty"], effective_spo)
            counts = outcome.completion_counts()
            counts["processed_rows"] = row_idx - 1
            job_service.update_job_progress(job_id_str, total_rows=total_data_rows, **counts)

        # After grouping phase: all data rows are "processed"; skipped count is final for this phase
        job_service.update_job_progress(
            job_id_str,
            total_rows=total_data_rows,
            processed_rows=total_data_rows,
            successful_rows=0,
            failed_rows=0,
            skipped_rows=len(skipped_detail),
        )

        # FIFO matching: match SPO allocations by spo_number + product (ignore warehouse)
        # Process GR lines grouped by (doc_no, product_id) to share SPO pool FIFO
        #
        # NOTE: `successful` / `failed` below count PICKING LINES, not source rows -
        # FIFO splits one grouped quantity across several allocations. Per-source-row
        # attribution is done once at the end from `group_line_failed`, so the outcome
        # counters stay reconcilable against the number of rows in the file.
        successful = 0
        failed = 0
        errors: List[str] = []
        successful_detail: List[Dict[str, Any]] = []
        first_line_error: Optional[str] = None
        group_line_failed: set = set()
        group_line_error: Dict[tuple, str] = {}

        def _record_success(grn_number: str, product_id: str, warehouse_id: str, qty: int) -> None:
            successful_detail.append({
                "grn_number": grn_number,
                "product_code": product_id_to_code.get(product_id, product_id),
                "warehouse": warehouse_id_to_display.get(warehouse_id, warehouse_id),
                "quantity": qty,
            })

        def _safe_upsert_line(
            *,
            picking_header_id: Any,
            product_id: str,
            source_warehouse_id: str,
            quantity: int,
            spo_allocation_id: Optional[str],
            spo_number_raw: Optional[str] = None,
            company_id: Optional[str] = None,
            group_key: Optional[tuple] = None,
        ) -> bool:
            """Upsert one GRN line inside a SAVEPOINT.

            A single bad row (constraint / type / FK error) must not abort the
            whole job's transaction: without the savepoint, the swallowed error
            below would leave the Postgres tx in a failed state and every later
            statement — including ``update_job_progress`` and ``fail_job`` —
            would blow up with StaleDataError / PendingRollbackError.
            """
            nonlocal first_line_error
            try:
                with db.begin_nested():
                    proc.upsert_grn_line_for_import(
                        picking_header_id=picking_header_id,
                        product_id=product_id,
                        source_warehouse_id=source_warehouse_id,
                        quantity=quantity,
                        spo_allocation_id=spo_allocation_id,
                        spo_number_raw=spo_number_raw,
                        company_id=company_id,
                    )
                return True
            except Exception as e:
                if first_line_error is None:
                    first_line_error = str(e)
                    logger.warning(
                        "GRN lines import job %s: first line upsert error (savepoint rolled back): %s",
                        job_id_str,
                        e,
                        exc_info=True,
                    )
                errors.append(str(e))
                if group_key is not None:
                    group_line_failed.add(group_key)
                    group_line_error[group_key] = str(e)
                return False

        # Group GR lines by (doc_no, product_id, effective_spo) for FIFO SPO matching
        # Keep warehouse info for creating picking lines
        gr_lines_by_product: Dict[tuple, List[tuple]] = defaultdict(list)
        for (doc_no, product_id, warehouse_id, effective_spo), group_data in groups.items():
            header = headers_by_number.get(doc_no)
            if not header:
                # Defensive: the per-row loop already skips rows with no header, so
                # this is unreachable in practice - but it used to bump `failed`
                # silently, which is exactly the class of bug this work removes.
                failed += 1
                group_line_failed.add((doc_no, product_id, warehouse_id, effective_spo))
                group_line_error[(doc_no, product_id, warehouse_id, effective_spo)] = (
                    f"GRN header not found: {doc_no}"
                )
                continue
            gr_lines_by_product[(doc_no, product_id, effective_spo)].append((warehouse_id, group_data["qty"], header))

        # Process each (doc_no, product, effective_spo) group with shared FIFO SPO pool
        for (doc_no, product_id, effective_spo), gr_line_list in gr_lines_by_product.items():
            header = headers_by_number.get(doc_no)
            if not header:
                failed += 1
                for _wh_id, _qty, _hdr in gr_line_list:
                    gk = (doc_no, product_id, _wh_id, effective_spo)
                    group_line_failed.add(gk)
                    group_line_error[gk] = f"GRN header not found: {doc_no}"
                continue

            # Every row this group writes carries the GRN's OWN company, and every
            # pool it builds is confined to it. Both halves or neither (AC-FM-27):
            # a job with no company snapshot runs system-scoped, where the insert
            # hook would stamp the incumbent instead.
            header_company_id = (
                str(header.company_id) if header.company_id is not None else None
            )

            spo_number: Optional[str] = effective_spo
            if not spo_number:
                # No SPO number, create all lines without spo_allocation_id
                for warehouse_id, qty, hdr in gr_line_list:
                    if _safe_upsert_line(
                        picking_header_id=hdr.id,
                        product_id=product_id,
                        source_warehouse_id=warehouse_id,
                        quantity=qty,
                        spo_allocation_id=None,
                        spo_number_raw=None,
                        company_id=header_company_id,
                        group_key=(doc_no, product_id, warehouse_id, effective_spo),
                    ):
                        successful += 1
                        _record_success(doc_no, product_id, warehouse_id, qty)
                    else:
                        failed += 1
                continue

            # The pool and the draw are `app.services.grn_spo_matching`'s, not this
            # loop's. Forward matching runs the SAME two functions over the lines a
            # previous import stated but could not place, and a second copy of the
            # two-pass rule here is exactly how the two directions would come to
            # disagree about which allocation a line belongs to.
            #
            # The pool is built ONCE per (doc_no, product, SPO) group so its lines
            # share it, and this GRN is excluded from its own consumption - a
            # re-import must not see the rows it is about to rewrite as capacity
            # somebody else already took.
            pool = build_allocation_pool(
                db,
                product_id=product_id,
                spo_number=spo_number,
                exclude_header_ids={str(hdr.id) for _wh, _q, hdr in gr_line_list},
                # An import job with no company snapshot runs system-scoped (all
                # companies), where the scope layer constrains nothing - so the
                # GRN's own company is stated rather than assumed.
                company_id=header_company_id,
            )

            for warehouse_id, gr_qty, hdr in gr_line_list:
                for draw in draw_fifo(pool, warehouse_id=warehouse_id, quantity=gr_qty):
                    if _safe_upsert_line(
                        picking_header_id=hdr.id,
                        product_id=product_id,
                        source_warehouse_id=warehouse_id,
                        quantity=draw.quantity,
                        spo_allocation_id=draw.allocation_id,
                        spo_number_raw=spo_number,
                        # The pool above is already confined to this header's
                        # company, so the row it produces has to carry that company
                        # too - otherwise the GRN draws correctly and shows none of
                        # what it drew, and the mis-stamped row never counts as
                        # consumption either, so a re-import over-draws.
                        company_id=header_company_id,
                        group_key=(doc_no, product_id, warehouse_id, effective_spo),
                    ):
                        successful += 1
                        _record_success(doc_no, product_id, warehouse_id, draw.quantity)
                    else:
                        failed += 1

            job_service.update_job_progress(
                job_id_str,
                total_rows=total_data_rows,
                processed_rows=total_data_rows,
                successful_rows=successful,
                failed_rows=failed,
                skipped_rows=len(skipped_detail),
            )

        db.commit()

        # Per-SOURCE-ROW attribution. The loops above counted picking lines (FIFO
        # splits one grouped quantity across allocations); the operator asked which
        # spreadsheet ROW succeeded, so fan the group verdict back onto its rows.
        for group_key, source_rows in group_source_rows.items():
            identity = group_row_identity.get(group_key) or {}
            if group_key in group_line_failed:
                message = group_line_error.get(group_key, "GRN line could not be written")
                for source_row in source_rows:
                    outcome.fail(
                        row=source_row,
                        code=oc.UPSERT_ERROR,
                        message=message,
                        value=identity.get("doc_no"),
                        identity=identity,
                    )
            else:
                for source_row in source_rows:
                    outcome.success(
                        row=source_row,
                        code=oc.CREATED,
                        message=f"GRN line saved: {identity.get('doc_no') or ''}".strip(),
                        value=identity.get("doc_no"),
                        identity=identity,
                        entity_type="picking_line",
                    )

        # After GRN lines import: reflect received quantities to SPO allocations (confirmed GRN)
        for header in headers_by_number.values():
            try:
                proc.sync_grn_received_to_spo(header.id)
            except Exception as e:
                logger.warning("Sync GRN received to SPO failed for header %s: %s", header.id, e)

        job_service.complete_job(
            job_id=job_id_str,
            result=outcome.finalize(
                "GRN lines import completed",
                total_rows=total_data_rows,
                # legacy keys, kept one release
                first_line_error=first_line_error,
                errors=_json_safe(errors[-100:]),
                skipped_rows_detail=_json_safe(skipped_detail[-500:]),
                successful_rows_detail=_json_safe(successful_detail[-500:]),
                picking_lines_written=successful,
                picking_lines_failed=failed,
            ),
            **outcome.completion_counts(total_rows=total_data_rows),
        )
        logger.info("GRN lines import job %s completed: %s ok, %s failed", job_id_str, successful, failed)
        _write_import_audit(
            db,
            entity_type="grn",
            label=f"GRN lines import {filename or ''}".strip(),
            row_count=successful,
            user_id=user_id,
            entity_id=job_id_str,
            status="partial" if failed else "success",
        )
    except Exception as e:
        logger.exception("GRN lines import job %s failed", job_id_str)
        # A crashed import still owes the operator the rows it did classify. The
        # recorder writes on its own session, so flushing here survives the rollback.
        outcome.flush()
        job_service.fail_job(job_id_str, str(e))
        _write_import_audit(
            db,
            entity_type="grn",
            label=f"GRN lines import {filename or ''}".strip(),
            row_count=0,
            user_id=user_id,
            entity_id=job_id_str,
            status="failed",
        )
    finally:
        db.close()


def _delivery_order_detail_normalize_header(value: Any) -> str:
    """Normalize column header for delivery order detail import."""
    if value is None:
        return ""
    s = str(value).strip().lower()
    if not s:
        return ""
    return s


def _decimal_or_none(value: Any) -> Optional[Decimal]:
    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    try:
        return Decimal(str(value))
    except Exception:
        return None


def _decimal_or_none_pct(value: Any) -> Optional[Decimal]:
    """Decimal parser that also accepts trailing '%'. Returns the percent value as a Decimal
    (e.g. '45%' -> Decimal('45')). Empty / unparseable -> None."""
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    s = str(value).strip()
    if not s:
        return None
    if s.endswith("%"):
        s = s[:-1].strip()
        if not s:
            return None
    try:
        return Decimal(s)
    except Exception:
        return None


def _norm_decimal_for_key(v: Any) -> Decimal:
    """Normalize a Decimal-or-None for composite-key equality. None -> 0."""
    if v is None:
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    try:
        return Decimal(str(v))
    except Exception:
        return Decimal("0")


def process_delivery_order_detail_import(db_job_id: str, file_data: bytes, filename: str, user_id: str):
    """Process delivery order detail Excel: one new order line per spreadsheet row.
    Columns: doc no -> order (order_number), item code -> product, location -> warehouse,
    qty, unit price, discount, total, tax, total excluding tax, total including tax.
    Duplicate product+warehouse on the same order produce separate lines (sequenced).
    """
    from rq import get_current_job
    import openpyxl
    from app.api.v1.external.utils import normalize_code

    db = SessionLocal()
    _apply_import_job_scope(db, db_job_id)
    job_service = JobService(db)
    rq_job = get_current_job()
    rq_job_id = rq_job.id if rq_job else None
    job = job_service.get_job_by_db_id(db_job_id) if db_job_id else None
    if not job and rq_job_id:
        job = job_service.get_job(rq_job_id)
    if not job:
        logger.error("Delivery order detail import job not found: db_job_id=%s", db_job_id)
        db.close()
        return

    job_id_str = str(job.job_id)
    # Bind the recorder to THIS module's SessionLocal (not app.database's) so it
    # follows the same engine the task is using - keeps tests hermetic and keeps
    # row capture on its own session.
    outcome = ImportOutcome(getattr(job, "id", None), session_factory=SessionLocal)
    try:
        job_service.start_job(job_id_str)
        try:
            workbook = openpyxl.load_workbook(BytesIO(file_data), data_only=True)
        except Exception as exc:
            job_service.fail_job(job_id_str, f"Failed to read Excel: {exc}")
            db.close()
            return

        sheet = workbook.active
        if not sheet:
            job_service.fail_job(job_id_str, "Workbook has no active sheet")
            db.close()
            return

        headers = [_delivery_order_detail_normalize_header(cell.value) for cell in sheet[1]]
        data_rows: List[Dict[str, Any]] = []
        all_doc_nos = set()
        all_product_codes = set()
        all_locations = set()

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            row_data = {}
            for idx, value in enumerate(row):
                if idx < len(headers) and headers[idx]:
                    row_data[headers[idx]] = value

            def _find(row_d: dict, *candidates: str) -> Any:
                for c in candidates:
                    cl = c.lower().strip()
                    if cl in row_d:
                        return row_d.get(cl)
                return None

            doc_no = (_find(row_data, "doc no", "doc number", "order number") and str(_find(row_data, "doc no", "doc number", "order number")).strip()) or None
            item_code = (_find(row_data, "item code", "product code") and str(_find(row_data, "item code", "product code")).strip()) or None
            location = (_find(row_data, "location", "warehouse", "warehouse code") and str(_find(row_data, "location", "warehouse", "warehouse code")).strip()) or None

            # Skip rows with no order-line identity (trailing formatted blanks, stray numeric defaults, etc.)
            if not doc_no and not item_code and not location:
                continue

            qty = _decimal_or_none(_find(row_data, "qty", "quantity"))
            unit_price = _decimal_or_none(_find(row_data, "unit price", "unit price "))
            discount = _decimal_or_none_pct(_find(row_data, "discount"))
            total = _decimal_or_none(_find(row_data, "total"))
            tax = _decimal_or_none(_find(row_data, "tax"))
            total_excl = _decimal_or_none(_find(row_data, "total excluding tax", "total excluding tax "))
            total_incl = _decimal_or_none(_find(row_data, "total including tax", "total including tax "))

            if doc_no:
                all_doc_nos.add(doc_no)
            if item_code:
                all_product_codes.add(item_code)
            if location:
                all_locations.add(location)
            data_rows.append({
                "doc_no": doc_no,
                "item_code": item_code,
                "location": location,
                "quantity": qty,
                "unit_price": unit_price,
                "discount": discount,
                "total": total,
                "tax": tax,
                "total_excluding_tax": total_excl,
                "total_including_tax": total_incl,
            })

        total_data_rows = len(data_rows)
        job_service.update_job_progress(job_id_str, total_rows=total_data_rows, processed_rows=0)

        if not data_rows:
            job_service.complete_job(
                job_id=job_id_str,
                result=outcome.finalize("No valid data rows", total_rows=total_data_rows),
                **outcome.completion_counts(total_rows=total_data_rows),
            )
            db.close()
            return

        doc_strips = {(d or "").strip() for d in all_doc_nos if d}
        orders_by_number: Dict[str, Any] = {}
        if doc_strips:
            batch_orders = (
                db.query(Order)
                .filter(Order.deleted_at.is_(None), Order.order_number.in_(list(doc_strips)))
                .all()
            )
            by_order_number = {(o.order_number or "").strip(): o for o in batch_orders}
            for doc_no in all_doc_nos:
                o = by_order_number.get((doc_no or "").strip())
                if o:
                    orders_by_number[doc_no] = o

        products_by_code = get_products_by_code_exact(db, all_product_codes)
        warehouses_map = get_warehouses_by_code_or_name(db, all_locations)

        # Next line_sequence per order (append after existing lines if key does not exist yet)
        seq_next: Dict[str, int] = {o.id: 0 for o in orders_by_number.values()}
        if seq_next:
            oids = list(seq_next.keys())
            for oid, mx in (
                db.query(OrderLine.order_id, func.max(OrderLine.line_sequence))
                .filter(OrderLine.order_id.in_(oids))
                .group_by(OrderLine.order_id)
                .all()
            ):
                seq_next[oid] = int(mx or 0)

        def _line_key(
            order_id: str,
            product_id: str,
            warehouse_id: str,
            quantity: Any,
            unit_price: Any,
            discount: Any,
            total: Any,
        ) -> tuple[str, str, str, Decimal, Decimal, Decimal, Decimal]:
            """Composite identity for an order line within an order. Same product+warehouse with
            different quantity/unit_price/discount/total are SEPARATE lines (no aggregation).
            Used both to deduplicate against existing rows on re-upload and to keep import
            rows distinct when source data has same product/warehouse repeated."""
            return (
                str(order_id),
                str(product_id),
                str(warehouse_id),
                _norm_decimal_for_key(quantity),
                _norm_decimal_for_key(unit_price),
                _norm_decimal_for_key(discount),
                _norm_decimal_for_key(total),
            )

        # Resolve each import row to a per-row entry. NO pre-aggregation — every row stays distinct
        # so two source rows with same product/warehouse but different price/discount produce two
        # separate order lines, matching the source spreadsheet.
        resolved_rows: List[Dict[str, Any]] = []
        errors: List[Dict[str, Any]] = []
        progress_every = max(100, min(500, total_data_rows // 50 or 100))

        def _identity(row_d: Dict[str, Any]) -> Dict[str, Any]:
            """The row's mapped business columns — what the operator needs to find it."""
            return {
                "doc_no": row_d.get("doc_no"),
                "item_code": row_d.get("item_code"),
                "location": row_d.get("location"),
                "qty": row_d.get("quantity"),
                "unit_price": row_d.get("unit_price"),
            }

        def _skip(row_idx: int, code: str, message: str, row_d: Dict[str, Any], value=None) -> None:
            outcome.skip(
                row=row_idx, code=code, message=message, value=value, identity=_identity(row_d)
            )
            errors.append({"row": row_idx, "error": message})  # legacy key, one release

        for row_idx, row_data in enumerate(data_rows, start=2):
            try:
                doc_no = row_data.get("doc_no")
                item_code = row_data.get("item_code")
                location = row_data.get("location")
                if not doc_no:
                    _skip(row_idx, oc.MISSING_DOC_NO, "Missing doc no", row_data)
                    continue
                if not item_code:
                    _skip(row_idx, oc.MISSING_ITEM_CODE, "Missing item code", row_data)
                    continue
                if not location:
                    _skip(row_idx, oc.MISSING_LOCATION, "Missing location", row_data)
                    continue

                order = orders_by_number.get(doc_no)
                product = products_by_code.get((item_code or "").strip())
                warehouse = warehouses_map.get(normalize_code(location)) if location else None
                if not order:
                    _skip(row_idx, oc.ORDER_NOT_FOUND, f"Order not found: {doc_no}", row_data, doc_no)
                    continue
                if not product:
                    _skip(
                        row_idx,
                        oc.PRODUCT_NOT_FOUND,
                        f"Product not found: {item_code}",
                        row_data,
                        item_code,
                    )
                    continue
                if not warehouse:
                    _skip(
                        row_idx,
                        oc.WAREHOUSE_NOT_FOUND,
                        f"Warehouse not found: {location}",
                        row_data,
                        location,
                    )
                    continue

                resolved_rows.append({
                    "row_idx": row_idx,
                    "order_id": str(order.id),
                    "product_id": str(product.id),
                    "warehouse_id": str(warehouse.id),
                    "identity": _identity(row_data),
                    "quantity": row_data.get("quantity"),
                    "unit_price": row_data.get("unit_price"),
                    "discount": row_data.get("discount"),
                    "total": row_data.get("total"),
                    "tax": row_data.get("tax"),
                    "total_excluding_tax": row_data.get("total_excluding_tax"),
                    "total_including_tax": row_data.get("total_including_tax"),
                })
            finally:
                cur = row_idx - 1
                if cur % progress_every == 0 or cur == total_data_rows:
                    progress = outcome.completion_counts()
                    progress["processed_rows"] = cur  # rows read, not rows decided
                    job_service.update_job_progress(job_id_str, **progress)

        # Load existing lines for the involved orders and count multiplicity per composite key.
        # Re-uploading an incremental file should NOT create duplicates: for each incoming row,
        # if an unmatched existing line with the identical composite key remains, we consume it
        # (skip insert); only the surplus (new rows in the file) becomes new OrderLine inserts.
        order_ids = list({r["order_id"] for r in resolved_rows})
        existing_remaining: Dict[
            tuple[str, str, str, Decimal, Decimal, Decimal, Decimal],
            int,
        ] = {}
        if order_ids:
            existing_lines = (
                db.query(OrderLine)
                .filter(OrderLine.order_id.in_(order_ids))
                .all()
            )
            for line in existing_lines:
                ex_key = _line_key(
                    str(line.order_id),
                    str(line.product_id),
                    str(line.warehouse_id),
                    getattr(line, "quantity", None),
                    getattr(line, "unit_price", None),
                    getattr(line, "discount", None),
                    getattr(line, "total", None),
                )
                existing_remaining[ex_key] = existing_remaining.get(ex_key, 0) + 1

        def _build_line(entry: Dict[str, Any], sequence: int) -> OrderLine:
            return OrderLine(
                order_id=entry["order_id"],
                product_id=entry["product_id"],
                warehouse_id=entry["warehouse_id"],
                line_sequence=sequence,
                quantity=entry.get("quantity") or Decimal("0"),
                unit_price=entry.get("unit_price"),
                discount=entry.get("discount"),
                total=entry.get("total"),
                tax=entry.get("tax"),
                total_excluding_tax=entry.get("total_excluding_tax"),
                total_including_tax=entry.get("total_including_tax"),
            )

        pending: List[Dict[str, Any]] = []  # entries actually queued for insert
        for entry in resolved_rows:
            try:
                key = _line_key(
                    entry["order_id"],
                    entry["product_id"],
                    entry["warehouse_id"],
                    entry["quantity"],
                    entry["unit_price"],
                    entry["discount"],
                    entry["total"],
                )
                if existing_remaining.get(key, 0) > 0:
                    # An identical line is already on this order, so this row is a
                    # re-upload of something already imported. This used to bump the
                    # counter silently, which is how 4,018 rows vanished from a green
                    # job; it is now a named, drillable outcome.
                    existing_remaining[key] -= 1
                    outcome.skip(
                        row=entry.get("row_idx"),
                        code=oc.DUPLICATE_LINE,
                        message=(
                            "Identical line already exists on this order "
                            "(same product, warehouse, qty, unit price, discount, total)"
                        ),
                        value=(entry.get("identity") or {}).get("doc_no"),
                        identity=entry.get("identity"),
                    )
                    continue

                seq_next[entry["order_id"]] += 1
                entry["_sequence"] = seq_next[entry["order_id"]]
                db.add(_build_line(entry, entry["_sequence"]))
                pending.append(entry)
            except Exception as e:
                outcome.fail(
                    row=entry.get("row_idx"),
                    code=oc.ROW_ERROR,
                    message=str(e),
                    identity=entry.get("identity"),
                )
                errors.append({"row": entry.get("row_idx"), "error": str(e)})

        # Single bulk commit on the happy path. If it blows up we still owe the
        # operator the row that caused it, so replay row-by-row inside savepoints
        # purely to attribute the failure (AC-A10). Costs nothing unless it fails.
        try:
            db.commit()
            for entry in pending:
                outcome.success(
                    row=entry.get("row_idx"),
                    code=oc.CREATED,
                    message="Order line created",
                    value=(entry.get("identity") or {}).get("doc_no"),
                    identity=entry.get("identity"),
                    entity_type="order_line",
                )
        except Exception as commit_error:
            db.rollback()
            logger.warning(
                "Delivery order detail import job %s: bulk commit failed (%s); "
                "replaying row-by-row to attribute the failure",
                job_id_str,
                commit_error,
            )
            for entry in pending:
                try:
                    with db.begin_nested():
                        db.add(_build_line(entry, entry["_sequence"]))
                    outcome.success(
                        row=entry.get("row_idx"),
                        code=oc.CREATED,
                        message="Order line created",
                        value=(entry.get("identity") or {}).get("doc_no"),
                        identity=entry.get("identity"),
                        entity_type="order_line",
                    )
                except Exception as row_error:
                    outcome.fail(
                        row=entry.get("row_idx"),
                        code=oc.ROW_ERROR,
                        message=str(row_error),
                        value=(entry.get("identity") or {}).get("doc_no"),
                        identity=entry.get("identity"),
                    )
                    errors.append({"row": entry.get("row_idx"), "error": str(row_error)})
            db.commit()

        successful = outcome.successful
        failed = outcome.failed
        skipped = outcome.skipped
        job_service.complete_job(
            job_id=job_id_str,
            result=outcome.finalize(
                "Delivery order detail import completed",
                total_rows=total_data_rows,
                errors=_json_safe(errors[:100]),  # legacy key, kept one release
            ),
            **outcome.completion_counts(total_rows=total_data_rows),
        )
        logger.info("Delivery order detail import job %s completed: %s ok, %s failed, %s skipped", job_id_str, successful, failed, skipped)
        _write_import_audit(
            db,
            entity_type="picking",
            label=f"DO detail import {filename or ''}".strip(),
            row_count=successful,
            user_id=user_id,
            entity_id=job_id_str,
            status="partial" if failed else "success",
        )
    except Exception as e:
        logger.exception("Delivery order detail import job %s failed", job_id_str)
        # A crashed import still owes the operator the rows it did classify. The
        # recorder writes on its own session, so flushing here survives the rollback.
        outcome.flush()
        job_service.fail_job(job_id_str, str(e))
        _write_import_audit(
            db,
            entity_type="picking",
            label=f"DO detail import {filename or ''}".strip(),
            row_count=0,
            user_id=user_id,
            entity_id=job_id_str,
            status="failed",
        )
    finally:
        db.close()


def validate_container_status_import(file_data: bytes, filename: str) -> Dict[str, Any]:
    """Dry run for the container status workbook. Reads only, never writes.

    Returns the shape the shared frontend upload dialog renders:
    ``{valid, errors[], warnings[], summary{total_rows, would_update, would_create,
    error_count}}``. Those summary keys are not free choice - the dialog renders
    exactly those and silently drops anything else.
    """
    from app.services.container_status_service import ContainerStatusImportService

    db = SessionLocal()
    # A preview reads across every company: there is no job snapshot yet, and the
    # operator needs the true would-update count, not a company-filtered one.
    set_company_scope(db, None)
    try:
        return ContainerStatusImportService(db).validate(file_data)
    except Exception as exc:  # noqa: BLE001 - a preview must never 500
        logger.exception("Container status validation failed for %s", filename)
        return {
            "valid": False,
            "errors": [f"Could not validate this workbook: {exc}"],
            "warnings": [],
            "summary": {
                "total_rows": 0,
                "would_update": 0,
                "would_create": 0,
                "error_count": 1,
            },
        }
    finally:
        db.close()


def process_container_status_import(
    db_job_id: str, file_data: bytes, filename: str, user_id: str
):
    """Import the container status workbook onto `inbound_shipments`.

    Update-only: it adds clearance dates to packing lists that ALREADY exist and
    never creates one, because the sheet carries no lines, supplier or quantities
    (D32). Rows for containers the system does not have are skipped and counted.

    Matching is on the normalized container number across EVERY shipment status -
    318 of the 407 rows are archived containers, and their clearance history still
    belongs on their row. A blank cell never clears, and a row whose values already
    agree is not touched at all, so a daily re-upload is a genuine no-op rather than
    407 phantom edits.
    """
    from rq import get_current_job

    from app.services.container_status_import import (
        ContainerStatusParseError,
        parse_container_status_workbook,
    )
    from app.services.container_status_service import ContainerStatusImportService

    db = SessionLocal()
    _apply_import_job_scope(db, db_job_id)
    job_service = JobService(db)

    rq_job = get_current_job()
    rq_job_id = rq_job.id if rq_job else None
    job = job_service.get_job_by_db_id(db_job_id) if db_job_id else None
    if not job and rq_job_id:
        job = job_service.get_job(rq_job_id)
    if not job:
        logger.error("Container status import job not found: db_job_id=%s", db_job_id)
        db.close()
        return

    job_id_str = str(job.job_id)
    outcome = ImportOutcome(getattr(job, "id", None), session_factory=SessionLocal)
    try:
        job_service.start_job(job_id_str)

        try:
            parsed = parse_container_status_workbook(file_data)
        except ContainerStatusParseError as exc:
            # Not a container status sheet at all. Fail the job with the reason
            # rather than importing zero rows and reporting success.
            outcome.flush()
            job_service.fail_job(job_id_str, str(exc))
            _write_import_audit(
                db,
                entity_type="inbound_shipment",
                label=f"Container status import {filename or ''}".strip(),
                row_count=0,
                user_id=user_id,
                entity_id=job_id_str,
                status="failed",
            )
            return

        service = ContainerStatusImportService(db)
        counts = service.apply(parsed, user_id=user_id, outcome=outcome)
        db.commit()

        errors = service._errors_from(parsed)
        total = len(parsed.rows) + len(parsed.rejected)

        job_service.complete_job(
            job_id=job_id_str,
            # The recorder owns the counters, so every counted row is attributed
            # to a reason (a tally that moves without one is the bug the guard in
            # tests/test_import_outcome_guard.py exists to prevent). The local
            # `counts` survive only as the legacy result keys the UI still reads:
            # "unchanged" and "skipped_no_packing_list" are both skips to the job,
            # but the operator needs to see WHICH, or a sheet that matched nothing
            # looks the same as one that changed nothing.
            result=outcome.finalize(
                "Container status import completed",
                total_rows=total,
                updated=counts["updated"],
                unchanged=counts["unchanged"],
                skipped_no_packing_list=counts["skipped"],
                rejected=counts["rejected"],
                blocks=len(parsed.blocks),
                blank_rows=parsed.blank_row_count,
                errors=errors[:50],
                warnings=parsed.warnings,
            ),
            **outcome.completion_counts(total_rows=total),
        )

        # Publish the retained workbook as a Container Status attachment so
        # "send me the container status" has an answer. Best-effort by design:
        # the import has already succeeded, and failing to catalogue a document
        # must not report that success as a failure.
        from app.services.container_status_document import publish_import_source

        publish_import_source(db, job)

        _write_import_audit(
            db,
            entity_type="inbound_shipment",
            label=f"Container status import {filename or ''}".strip(),
            row_count=len(parsed.rows),
            user_id=user_id,
            entity_id=job_id_str,
            status="partial" if counts["rejected"] else "success",
        )
    except Exception as exc:  # noqa: BLE001 - the job must record why it died
        logger.exception("Container status import job %s failed", job_id_str)
        db.rollback()
        # The recorder writes on its own session, so the rows it already
        # classified survive this rollback.
        outcome.flush()
        job_service.fail_job(job_id_str, str(exc))
        _write_import_audit(
            db,
            entity_type="inbound_shipment",
            label=f"Container status import {filename or ''}".strip(),
            row_count=0,
            user_id=user_id,
            entity_id=job_id_str,
            status="failed",
        )
    finally:
        db.close()


# ---------------------------------------------------------------------------
# Customers (debtor listing)
# ---------------------------------------------------------------------------

#: The audit entity type for a customer. MUST equal what the audit listener derives
#: from the model (`Customer.__audit_entity_type__`), because the per-row suppression
#: below matches on this exact string and a mismatch fails silently - the import just
#: writes one audit row per line again, 4,000 of them, all reading "System". Pinned by
#: tests/test_customer_audit.py.
_CUSTOMER_ENTITY_TYPE = "customer"


def _customer_import_details(
    job: Any, filename: str, result: Optional[Dict[str, Any]] = None
) -> Dict[str, Any]:
    """What the coarse audit row carries: the file, the job, and what changed.

    `import_job_id` is the FK `import_job_rows.import_job_id`, so a reader a year later
    gets from this one row to every row's outcome and identity.
    """
    result = result or {}
    return {
        "filename": filename or None,
        "import_job_id": str(getattr(job, "id", "") or "") or None,
        "job_id": str(getattr(job, "job_id", "") or "") or None,
        "total_rows": int(result.get("total_rows", 0)),
        "created": int(result.get("created", 0)),
        "updated": int(result.get("updated", 0)),
        "unchanged": int(result.get("unchanged", 0)),
        "skipped": int(result.get("skipped", 0)),
        "failed": int(result.get("failed", 0)),
    }


def _customer_import_shape(result: Dict[str, Any]) -> Dict[str, Any]:
    """Turn the service's outcome into the shape the import dialog renders.

    `valid` answers "would anything import at all", NOT "is every row perfect": a file
    with three bad rows out of 900 imports 897, so row problems are WARNINGS the user
    acknowledges rather than errors that block (AC-5.6). Only an unreadable file, or one
    with no customer code / name column, is invalid.
    """
    problems = result.get("problems", [])
    unmapped = result.get("unmapped_headers", [])
    unknown_segments = result.get("unknown_market_segments", [])
    unknown_segment_rows = int(result.get("unknown_market_segment_rows", 0))
    needs_review = int(result.get("needs_review", 0))

    errors: List[str] = []
    if not result.get("readable"):
        missing = ", ".join(
            c.replace("_", " ") for c in result.get("missing_columns", [])
        )
        errors.append(
            f"The file has no {missing} column."
            if missing
            else "The file could not be read."
        )
        for problem in problems:
            errors.append(f"Row {problem['row']}: {problem['reason']}")

    warnings: List[str] = []
    if result.get("readable"):
        warnings.extend(f"Row {p['row']}: {p['reason']}" for p in problems)
        warnings.extend(f"Column not recognised: {header}" for header in unmapped)
        warnings.extend(
            f"Market segment not recognised, left unset: {code}"
            for code in unknown_segments
        )
        if unknown_segment_rows:
            # The segment decides SCM demand class and fulfilment priority, so how MANY
            # customers land without one is the part that matters. Each affected row also
            # carries its own outcome code on the job detail.
            warnings.append(
                f"{unknown_segment_rows} row(s) import with no market segment. "
                "Each is listed on the job."
            )
        if needs_review:
            warnings.append(
                f"{needs_review} row(s) carry a name close to one already on the same "
                "customer code. They import; each is listed on the job."
            )

    return {
        "valid": bool(result.get("readable")),
        "errors": errors,
        "warnings": warnings,
        "summary": {
            "total_rows": int(result.get("total_rows", 0)),
            "would_create": int(result.get("created", 0)),
            "would_update": int(result.get("updated", 0)),
            "would_unchanged": int(result.get("unchanged", 0)),
            "would_skip": int(result.get("skipped", 0)) + int(result.get("failed", 0)),
            "needs_review": needs_review,
            "unmapped_headers": list(unmapped),
            "missing_columns": list(result.get("missing_columns", [])),
            "problems": problems,
        },
    }


def validate_customer_import(
    file_data: bytes, *, company_scope: Any = _SCOPE_NOT_GIVEN
) -> Dict[str, Any]:
    """Dry run for a customer listing. Reads only, never writes.

    Runs at the SAME company scope the real import will run at: the existing-customer
    lookup is what decides create vs update, so a preview reading every company would
    report "would update" on rows the scoped import inserts.
    """
    from app.services import customer_import_service

    db = SessionLocal()
    _apply_preview_scope(db, company_scope)
    try:
        return _customer_import_shape(customer_import_service.preview(db, file_data))
    except Exception as exc:  # noqa: BLE001 - a preview must never 500
        logger.exception("Customer import validation failed")
        return {
            "valid": False,
            "errors": [f"Could not read this file: {exc}"],
            "warnings": [],
            "summary": {},
        }
    finally:
        db.rollback()
        db.close()


def process_customer_import(db_job_id: str, file_data: bytes, filename: str, user_id: str):
    """Import a customer listing into `customers` for the job's company.

    Create-or-update on (company, code, name), all three; never a rename, never a delete,
    and never a touch of the fields a person curates (account owner, notes, active flag,
    a market segment already set). One bad row never fails the file.
    """
    from rq import get_current_job

    from app.services import customer_import_service

    db = SessionLocal()
    _apply_import_job_scope(db, db_job_id)
    job_service = JobService(db)

    rq_job = get_current_job()
    rq_job_id = rq_job.id if rq_job else None
    job = job_service.get_job_by_db_id(db_job_id) if db_job_id else None
    if not job and rq_job_id:
        job = job_service.get_job(rq_job_id)
    if not job:
        logger.error("Customer import job not found: db_job_id=%s", db_job_id)
        db.close()
        return

    job_id_str = str(job.job_id)
    outcome = ImportOutcome(getattr(job, "id", None), session_factory=SessionLocal)
    # Customer is __audit_track__; suppress the per-row ORM audit for this bulk job.
    # In production this is defence-in-depth rather than the thing doing the work:
    # worker.py registers the company-scope listeners only, never
    # register_audit_listeners, so no per-row audit fires in an RQ process at all
    # today. It bites for every OTHER caller of this task - the in-process test suite,
    # and the worker itself the day it starts registering the audit listeners - where
    # a 4,000-line debtor listing would otherwise become 4,000 audit rows, every one
    # of them reading "System" because a worker has no request actor. One coarse,
    # correctly-attributed job row is written at completion instead, and every row's
    # own outcome is already in import_job_rows.
    # setdefault-union, not assignment: a second suppression in the same session
    # (another audited model written by the same job) must not erase this one.
    db.info.setdefault("skip_audit_entity_types", set()).add(_CUSTOMER_ENTITY_TYPE)
    try:
        job_service.start_job(job_id_str)
        result = customer_import_service.apply(
            db,
            file_data,
            outcome,
            actor=user_id,
            # Publish the row total the moment the sheet is read. Without it `total_rows`
            # first appears in `complete_job` and the upload drawer shows 0/0 for the
            # whole run, which reads as stuck. Same as process_grn_lines_import.
            on_total_rows=lambda total: job_service.update_job_progress(
                job_id_str, total_rows=total, processed_rows=0
            ),
        )

        if not result.get("readable"):
            # Not a customer listing at all. Fail the job with the reason rather than
            # importing zero rows and reporting success.
            missing = ", ".join(
                c.replace("_", " ") for c in result.get("missing_columns", [])
            )
            db.rollback()
            outcome.flush()
            job_service.fail_job(
                job_id_str,
                f"The file has no {missing} column." if missing
                else "The file could not be read.",
            )
            _write_import_audit(
                db,
                entity_type=_CUSTOMER_ENTITY_TYPE,
                label=f"Customer import {filename or ''}".strip(),
                row_count=0,
                user_id=user_id,
                entity_id=job_id_str,
                status="failed",
                details_builder=lambda: _customer_import_details(job, filename, result),
            )
            return

        db.commit()
        total = int(result.get("total_rows", 0))
        job_service.complete_job(
            job_id=job_id_str,
            result=outcome.finalize(
                "Customer import completed",
                total_rows=total,
                created=result["created"],
                updated=result["updated"],
                unchanged=result["unchanged"],
                needs_review=result["needs_review"],
                review_rows=result["review_rows"][:50],
                unmapped_headers=result["unmapped_headers"],
                unknown_market_segments=result.get("unknown_market_segments", []),
            ),
            **outcome.completion_counts(total_rows=total),
        )
        _write_import_audit(
            db,
            entity_type=_CUSTOMER_ENTITY_TYPE,
            label=f"Customer import {filename or ''}".strip(),
            row_count=result["created"] + result["updated"],
            user_id=user_id,
            entity_id=job_id_str,
            status="partial" if (result["failed"] or result["skipped"]) else "success",
            details_builder=lambda: _customer_import_details(job, filename, result),
        )
    except Exception as exc:  # noqa: BLE001 - the job must record why it died
        logger.exception("Customer import job %s failed", job_id_str)
        db.rollback()
        # The recorder writes on its own session, so the rows it already classified
        # survive this rollback.
        outcome.flush()
        job_service.fail_job(job_id_str, str(exc))
        _write_import_audit(
            db,
            entity_type=_CUSTOMER_ENTITY_TYPE,
            label=f"Customer import {filename or ''}".strip(),
            row_count=0,
            user_id=user_id,
            entity_id=job_id_str,
            status="failed",
            details_builder=lambda: _customer_import_details(job, filename),
        )
    finally:
        db.close()


# ---------------------------------------------------------------------------
# SCM upload channels
# ---------------------------------------------------------------------------
#
# Five files feed the reorder plan: the outstanding sales-order and purchase-order books,
# purchase and sales history, and the Order Inquiry sheet. They ran INLINE in the request
# until the sales book (72,000 lines) timed the gateway out mid-write, which is the failure
# mode this queue exists to prevent. They now go through exactly the machinery every other
# importer uses - job row, company snapshot, retained source file, per-row outcomes - so the
# operator watches them in the upload drawer and reads what happened on the job page.
#
# One runner rather than five copies of the same seventy lines. The five differ ONLY in which
# service they call and how that service words "I could not read this file"; everything else
# (scope, the row recorder, the three exit paths, the audit row) is identical, and five copies
# of it is how one of them quietly stops flushing its rows on failure.


def _run_scm_upload_job(
    db_job_id: str,
    filename: str,
    user_id: str,
    *,
    job_label: str,
    entity_type: str,
    apply_fn,
    unreadable_message,
    written_rows,
    total_rows_of,
):
    """Run one queued SCM upload: apply the file, then terminate the job honestly.

    ``apply_fn(db, outcome, on_total_rows) -> dict`` does the write. ``unreadable_message``
    turns that dict into the reason the FILE could not be read, or None when it could - this
    is where the old routes' `400 unreadable` semantics land now: the JOB fails, carrying the
    problems, rather than the HTTP request the operator has already walked away from.
    """
    from rq import get_current_job

    db = SessionLocal()
    _apply_import_job_scope(db, db_job_id)
    job_service = JobService(db)

    rq_job = get_current_job()
    rq_job_id = rq_job.id if rq_job else None
    job = job_service.get_job_by_db_id(db_job_id) if db_job_id else None
    if not job and rq_job_id:
        job = job_service.get_job(rq_job_id)
    if not job:
        logger.error("%s job not found: db_job_id=%s", job_label, db_job_id)
        db.close()
        return

    job_id_str = str(job.job_id)
    label = f"{job_label} {filename or ''}".strip()
    outcome = ImportOutcome(getattr(job, "id", None), session_factory=SessionLocal)
    try:
        job_service.start_job(job_id_str)
        result = apply_fn(
            db,
            outcome,
            # Publish the row total the moment the file is read. Without it `total_rows`
            # first appears in `complete_job` and the drawer shows 0/0 for the whole run,
            # which reads as stuck.
            lambda total: job_service.update_job_progress(
                job_id_str, total_rows=total, processed_rows=0
            ),
        )

        problem = unreadable_message(result)
        if problem:
            # Not this kind of file at all. The job FAILS with the reason rather than
            # importing zero rows and reporting success.
            db.rollback()
            outcome.flush()
            job_service.fail_job(job_id_str, problem)
            _write_import_audit(db, entity_type=entity_type, label=label, row_count=0,
                                user_id=user_id, entity_id=job_id_str, status="failed")
            return

        db.commit()
        total = int(total_rows_of(result) or 0)
        job_service.complete_job(
            job_id=job_id_str,
            # The channel's own answer goes under ONE key rather than spread across the
            # envelope. Its shape is unchanged - `counts`, `applied`, `links`,
            # `unmapped_agents` all read exactly as the synchronous response did - and
            # nesting is what keeps its `counts` (the diff: added / closed / unchanged) from
            # overwriting the envelope's `counts` (the row totals every job page reads).
            result=outcome.finalize(f"{job_label} completed", total_rows=total,
                                    upload=result),
            **outcome.completion_counts(total_rows=total),
        )
        _write_import_audit(
            # Passed as a builder, not a value: the job is COMPLETE and committed by now, so
            # a `written_rows` that raises here must cost a warning rather than flip a
            # finished job to failed through the handler below.
            db, entity_type=entity_type, label=label,
            row_count=lambda: written_rows(result),
            user_id=user_id, entity_id=job_id_str,
            status="partial" if (outcome.failed or outcome.skipped) else "success",
        )
    except Exception as exc:  # noqa: BLE001 - the job must record why it died
        logger.exception("%s job %s failed", job_label, job_id_str)
        db.rollback()
        # The recorder writes on its own session, so the rows it already classified survive
        # this rollback.
        outcome.flush()
        job_service.fail_job(job_id_str, str(exc))
        _write_import_audit(db, entity_type=entity_type, label=label, row_count=0,
                            user_id=user_id, entity_id=job_id_str, status="failed")
    finally:
        db.close()


def _missing_columns_message(result: dict) -> Optional[str]:
    """The outstanding books' way of saying "this is not the file you think it is"."""
    if result.get("ok"):
        return None
    missing = ", ".join(c.replace("_", " ") for c in result.get("missing_columns") or [])
    return (f"The file is missing required columns: {missing}." if missing
            else "The file could not be read.")


def _problems_message(result: dict) -> Optional[str]:
    """The history and inquiry channels' way of saying the same thing."""
    if result.get("ok"):
        return None
    problems = result.get("problems") or ["This file could not be read."]
    return "; ".join(str(p) for p in problems)


def process_outstanding_import(db_job_id: str, file_data: bytes, filename: str,
                               user_id: str, doc_type: str):
    """Import an outstanding order book (sales or purchase) for the job's company.

    The diff is computed against what we already hold and applied line by line: added,
    quantity or date changed, unchanged, and CLOSED - a line we hold that the file no longer
    carries. Every one of those is recorded per source row, closures included, because
    closing is the destructive half and the job detail is the only place it can be found
    afterwards.
    """
    from app.services.scm import outstanding_import_service

    _run_scm_upload_job(
        db_job_id, filename, user_id,
        job_label=("Outstanding sales order import" if doc_type == "outstanding_so"
                   else "Outstanding purchase order import"),
        entity_type="sales_order" if doc_type == "outstanding_so" else "purchase_order",
        apply_fn=lambda db, outcome, on_total: outstanding_import_service.apply(
            db, file_data, doc_type, actor=user_id, outcome=outcome,
            on_total_rows=on_total,
        ),
        unreadable_message=_missing_columns_message,
        written_rows=lambda r: sum(
            int(r.get("applied", {}).get(k, 0)) for k in ("added", "updated", "closed")
        ),
        # The file's rows PLUS the lines it closed by absence: a closure carries an outcome
        # and no source row, and the file's own count alone would put processed past the
        # total. `file_rows` is still on the result as the operator's own number.
        total_rows_of=lambda r: r.get("total_rows", 0),
    )


def process_po_history_import(db_job_id: str, file_data: bytes, filename: str, user_id: str):
    """Import a purchase book as HISTORY, then resolve the SO<->PO claims.

    Either export shape (the banded listing, or the flat PO + SPO extract) and both document
    families; the service decides from the file itself. History is written closed and fully
    received, so it can never read as incoming supply, whichever family it belongs to.

    The link resolve runs inside the job because these files name sales orders - per document
    in the banded report, per LINE in the structured one - so an upload can complete a pairing
    the other side claimed months ago.
    """
    from app.services.scm import order_link_service, po_history_service

    def _apply(db, outcome, on_total):
        result = po_history_service.apply(db, file_data, actor=user_id, outcome=outcome,
                                          on_total_rows=on_total)
        if result.get("ok"):
            result["links"] = order_link_service.resolve(db)
        return result

    _run_scm_upload_job(
        db_job_id, filename, user_id,
        job_label="Purchase history import",
        entity_type="purchase_order",
        apply_fn=_apply,
        unreadable_message=_problems_message,
        written_rows=lambda r: int(r.get("lines_created", 0)),
        # The rows of the file, not its purchase lines: this is a banded report and most of
        # it - headers, SO notes, spacers - was never a line. Each of those carries its own
        # `not_a_line` outcome, so the total is the source rows and processed reaches it.
        total_rows_of=lambda r: r.get("total_rows", 0),
    )


def process_sales_history_import(db_job_id: str, file_data: bytes, filename: str,
                                 user_id: str):
    """Absorb the sales-order listing as HISTORY for the job's company.

    The channel that timed out: 11,275 documents and 81,361 lines in the client's own export.
    Every line lands closed and fully delivered, so absorbed history contributes nothing to
    committed demand.
    """
    from app.services.scm import so_history_service

    _run_scm_upload_job(
        db_job_id, filename, user_id,
        job_label="Sales history import",
        entity_type="sales_order",
        apply_fn=lambda db, outcome, on_total: so_history_service.apply(
            db, file_data, actor=user_id, outcome=outcome, on_total_rows=on_total,
        ),
        unreadable_message=_problems_message,
        written_rows=lambda r: int(r.get("lines_created", 0)) + int(r.get("lines_updated", 0)),
        # Every non-blank row, the 9,144 package captions included: each carries its own
        # `not_a_line` outcome, so a total that counts them is still reachable.
        total_rows_of=lambda r: r.get("total_rows", 0),
    )


def process_order_inquiry_import(db_job_id: str, file_data: bytes, filename: str,
                                 user_id: str):
    """Import the Order Inquiry sheet: project demand, stock locations, and PO claims.

    The importer is owned by Project Sales (ADR 0010), so it is read from
    `app/services/project_order_inquiry_import_service.py` rather than from `app.services.scm`.
    The job type, the queue and the route that enqueues it are unchanged.
    """
    from app.services import project_order_inquiry_import_service as order_inquiry_service
    from app.services.scm import order_link_service

    def _apply(db, outcome, on_total):
        result = order_inquiry_service.apply(db, file_data, actor=user_id, outcome=outcome,
                                             on_total_rows=on_total)
        if result.get("ok"):
            result["links"] = order_link_service.resolve(db)
        return result

    _run_scm_upload_job(
        db_job_id, filename, user_id,
        job_label="Order inquiry import",
        entity_type="sales_order",
        apply_fn=_apply,
        unreadable_message=_problems_message,
        written_rows=lambda r: int(r.get("lines_created", 0))
        + int(r.get("lines_refreshed", 0)),
        # The sheet's rows PLUS the instalments it withdrew: a withdrawal is reached by the
        # sheet's silence, so it carries an outcome and no row, and `rows` alone would put
        # processed past the total. `rows` is still on the result as the file's own count.
        total_rows_of=lambda r: r.get("total_rows", r.get("rows", 0)),
    )
